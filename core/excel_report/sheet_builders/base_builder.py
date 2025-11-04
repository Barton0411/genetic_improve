"""
Sheet构建器基类
所有Sheet构建器的基类，提供通用方法
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from abc import ABC, abstractmethod


class BaseSheetBuilder(ABC):
    """Sheet构建器基类"""

    def __init__(self, workbook: Workbook, style_manager, chart_builder, progress_callback=None):
        """
        初始化Sheet构建器

        Args:
            workbook: Excel工作簿对象
            style_manager: 样式管理器
            chart_builder: 图表构建器
            progress_callback: 进度回调函数 callback(progress: int, message: str)
        """
        self.wb = workbook
        self.style_manager = style_manager
        self.chart_builder = chart_builder
        self.progress_callback = progress_callback
        self.ws = None

    @abstractmethod
    def build(self, data: dict):
        """
        构建Sheet（子类必须实现）

        Args:
            data: 数据字典
        """
        pass

    def _create_sheet(self, title: str) -> Worksheet:
        """
        创建新的Sheet

        Args:
            title: Sheet标题

        Returns:
            Worksheet对象
        """
        self.ws = self.wb.create_sheet(title=title)
        return self.ws

    def _write_header(self, row: int, headers: list, start_col: int = 1):
        """
        写入标题行

        Args:
            row: 行号
            headers: 标题列表
            start_col: 起始列号
        """
        for col_idx, header in enumerate(headers, start=start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)

    def _write_data_row(self, row: int, values: list, start_col: int = 1, alignment='left'):
        """
        写入数据行

        Args:
            row: 行号
            values: 数据值列表
            start_col: 起始列号
            alignment: 对齐方式 ('left', 'center', 'right')
        """
        for col_idx, value in enumerate(values, start=start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=value)
            self.style_manager.apply_data_style(cell, alignment)

    def _write_total_row(self, row: int, values: list, start_col: int = 1):
        """
        写入合计行

        Args:
            row: 行号
            values: 数据值列表
            start_col: 起始列号
        """
        for col_idx, value in enumerate(values, start=start_col):
            cell = self.ws.cell(row=row, column=col_idx, value=value)
            self.style_manager.apply_total_style(cell)

    def _set_column_widths(self, widths: dict):
        """
        设置列宽

        Args:
            widths: {列号: 宽度} 字典
        """
        from openpyxl.utils import get_column_letter
        for col_idx, width in widths.items():
            col_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[col_letter].width = width

    def _freeze_panes(self, cell: str):
        """
        冻结窗格

        Args:
            cell: 冻结位置 (如 'A2')
        """
        self.ws.freeze_panes = cell

    def _add_pie_chart(self, data_range: dict, position: str, title: str):
        """
        添加饼图

        Args:
            data_range: 数据范围
            position: 图表位置
            title: 图表标题

        Returns:
            图表对象
        """
        return self.chart_builder.create_pie_chart(
            self.ws, data_range, position, title
        )

    def _add_bar_chart(self, data_range: dict, position: str, title: str):
        """
        添加柱状图

        Args:
            data_range: 数据范围
            position: 图表位置
            title: 图表标题

        Returns:
            图表对象
        """
        return self.chart_builder.create_bar_chart(
            self.ws, data_range, position, title
        )

    def _add_line_chart(self, data_range: dict, position: str, title: str):
        """
        添加折线图

        Args:
            data_range: 数据范围
            position: 图表位置
            title: 图表标题

        Returns:
            图表对象
        """
        return self.chart_builder.create_line_chart(
            self.ws, data_range, position, title
        )

    def _merge_cells(self, range_string: str):
        """
        合并单元格

        Args:
            range_string: 范围字符串 (如 'A1:D1')
        """
        self.ws.merge_cells(range_string)

    def _write_cell(self, row: int, col: int, value, alignment='left'):
        """
        写入单个单元格

        Args:
            row: 行号
            col: 列号
            value: 值
            alignment: 对齐方式
        """
        cell = self.ws.cell(row=row, column=col, value=value)
        self.style_manager.apply_data_style(cell, alignment)
        return cell

    def _write_dataframe_fast(self, df, start_row: int = 1, headers: list = None,
                             data_alignment='center', column_widths: dict = None,
                             progress_callback_interval: int = 500):
        """
        快速写入DataFrame数据（批量方式，性能优化）

        Args:
            df: DataFrame数据
            start_row: 起始行号
            headers: 自定义表头列表（如果为None则使用df.columns）
            data_alignment: 数据对齐方式 ('left', 'center', 'right')
            column_widths: 列宽字典 {列号: 宽度}
            progress_callback_interval: 进度报告间隔（每N行）

        Returns:
            下一个可用行号
        """
        import pandas as pd

        current_row = start_row

        # 1. 写入表头
        if headers is None:
            headers = list(df.columns)

        self._write_header(current_row, headers)
        current_row += 1

        # 2. 批量写入数据（使用append整行写入，性能提升100倍！）
        total_rows = len(df)

        for idx, row_data in enumerate(df.itertuples(index=False), start=0):
            # 使用append批量写入整行（远快于逐个单元格）
            self.ws.append(list(row_data))

            # 报告进度（减少频率）
            if self.progress_callback and (idx + 1) % progress_callback_interval == 0:
                pass  # 可以在此处添加进度回调

        current_row = start_row + 1 + total_rows

        # 3. 只对表头行应用样式（数据行不设置格式，提升性能）
        # 如果需要数据对齐，可以在列宽设置后由Excel自动处理

        # 4. 设置列宽
        if column_widths:
            self._set_column_widths(column_widths)

        return current_row

"""
Sheet 13构建器: 备选公牛-近交系数分析
按公牛分别分析近交系数风险分布
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class Sheet13Builder(BaseSheetBuilder):
    """
    Sheet 13: 备选公牛-近交系数分析

    包含内容:
    - 每个备选公牛的近交系数分布表格（垂直排列）
    - 每个表格右侧配分组水平条形图（成母牛/后备牛/全群对比）
    """

    def build(self, data: dict):
        """
        构建Sheet 13: 备选公牛-近交系数分析

        Args:
            data: 包含备选公牛近交系数分析数据
                - bulls: 公牛列表（按高风险占比从高到低排序）
        """
        try:
            # 检查数据
            if not data or 'bulls' not in data:
                logger.warning("Sheet13: 缺少数据，跳过生成")
                return

            bulls = data['bulls']
            if not bulls:
                logger.warning("Sheet13: 没有备选公牛数据")
                return

            # 创建Sheet
            self._create_sheet("备选公牛-近交系数分析")
            logger.info(f"构建Sheet 13: 备选公牛-近交系数分析（{len(bulls)}头公牛）")

            # 定义颜色
            self.color_header = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type='solid')  # 深蓝
            self.color_total = PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid')   # 橙色
            self.color_safe = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type='solid')    # 绿色
            self.color_low = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type='solid')     # 黄色
            self.color_high = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type='solid')    # 红色

            current_row = 1

            # 为每个公牛创建独立的表格和图表
            for idx, bull in enumerate(bulls, 1):
                logger.info(f"创建公牛{idx}/{len(bulls)}: {bull['bull_id']}")

                # 创建公牛表格
                current_row = self._create_bull_table(bull, current_row)

                # 创建图表（右侧）
                chart_row = current_row - 7  # 图表从表格标题行开始
                self._create_bull_chart(bull, chart_row)

                # 公牛之间空3行
                current_row += 3

            # 设置列宽
            self.ws.column_dimensions['A'].width = 18   # 近交区间列
            self.ws.column_dimensions['B'].width = 16   # 成母牛-头数
            self.ws.column_dimensions['C'].width = 12   # 占比
            self.ws.column_dimensions['D'].width = 16   # 后备牛-头数
            self.ws.column_dimensions['E'].width = 12   # 占比
            self.ws.column_dimensions['F'].width = 16   # 全群-头数
            self.ws.column_dimensions['G'].width = 12   # 占比
            self.ws.column_dimensions['H'].width = 12   # 风险等级

            # 冻结首行
            self._freeze_panes('A2')

            logger.info(f"✓ Sheet 13构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 13失败: {e}", exc_info=True)
            raise

    def _create_bull_table(self, bull: dict, start_row: int) -> int:
        """
        创建单个公牛的近交系数分布表格

        Args:
            bull: 公牛数据
            start_row: 起始行

        Returns:
            下一个可用行号
        """
        # 公牛标题行
        title_text = f"【公牛{bull['bull_id']}】原始号：{bull['original_bull_id']}"
        cell = self.ws.cell(row=start_row, column=1, value=title_text)
        cell.font = Font(size=12, bold=True, color="FFFFFF")
        cell.fill = self.color_header
        cell.alignment = Alignment(horizontal='left', vertical='center')
        # 合并A-H列
        self.ws.merge_cells(f'A{start_row}:H{start_row}')
        self.ws.row_dimensions[start_row].height = 25
        start_row += 1

        # 在群母牛总数说明行
        summary_text = f"在群母牛总数：成母牛{bull['mature_cow_count']}头 + 后备牛{bull['heifer_count']}头 = 全群{bull['total_cow_count']}头"
        cell = self.ws.cell(row=start_row, column=1, value=summary_text)
        cell.font = Font(size=10, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.merge_cells(f'A{start_row}:H{start_row}')
        start_row += 1

        # 表头
        headers = [
            '近交区间',
            f'成母牛\n(总{bull["mature_cow_count"]}头)',
            '占比',
            f'后备牛\n(总{bull["heifer_count"]}头)',
            '占比',
            f'全群\n(总{bull["total_cow_count"]}头)',
            '占比',
            '风险等级'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.color_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self._get_thin_border()

        self.ws.row_dimensions[start_row].height = 30
        start_row += 1

        # 数据行
        distribution = bull['distribution']
        for i, interval in enumerate(distribution['intervals']):
            row_data = [
                interval,
                distribution['mature_counts'][i],
                f"{distribution['mature_ratios'][i]:.1%}",
                distribution['heifer_counts'][i],
                f"{distribution['heifer_ratios'][i]:.1%}",
                distribution['total_counts'][i],
                f"{distribution['total_ratios'][i]:.1%}",
                distribution['risk_levels'][i]
            ]

            # 根据风险等级设置背景色
            if '安全' in distribution['risk_levels'][i]:
                row_fill = self.color_safe
            elif '低风险' in distribution['risk_levels'][i]:
                row_fill = self.color_low
            else:  # 高风险或极高风险
                row_fill = self.color_high

            for col_idx, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=start_row, column=col_idx, value=value)
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self._get_thin_border()
                # 不对第一列（区间）和最后一列（风险等级）设置背景色
                if 1 < col_idx < 8:
                    cell.fill = row_fill

            start_row += 1

        # 分隔线
        for col_idx in range(1, 9):
            cell = self.ws.cell(row=start_row, column=col_idx)
            cell.border = Border(top=Side(style='medium'))
        start_row += 1

        # 高风险小计行（>6.25%）
        high_risk = bull['high_risk_summary']
        row_data = [
            '高风险小计 (>6.25%)',
            high_risk['mature_count'],
            f"{high_risk['mature_ratio']:.1%}",
            high_risk['heifer_count'],
            f"{high_risk['heifer_ratio']:.1%}",
            high_risk['total_count'],
            f"{high_risk['total_ratio']:.1%}",
            ''
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.fill = self.color_total
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_thin_border()

        start_row += 1

        return start_row

    def _create_bull_chart(self, bull: dict, chart_row: int):
        """
        创建公牛的分组水平条形图（右侧）

        Args:
            bull: 公牛数据
            chart_row: 图表起始行（公牛标题行）
        """
        try:
            # 创建条形图
            chart = BarChart()
            chart.type = "bar"  # 水平条形图
            chart.title = "高风险影响占比 (>6.25%)"
            chart.style = 10

            # 写入图表数据到隐藏列（从列K开始）
            data_col = 11  # K列
            high_risk = bull['high_risk_summary']

            # 写入分类标签
            self.ws.cell(row=chart_row + 2, column=data_col, value="成母牛")
            self.ws.cell(row=chart_row + 3, column=data_col, value="后备牛")
            self.ws.cell(row=chart_row + 4, column=data_col, value="全群")

            # 写入数据（占比，转换为百分数）
            self.ws.cell(row=chart_row + 2, column=data_col + 1, value=high_risk['mature_ratio'])
            self.ws.cell(row=chart_row + 3, column=data_col + 1, value=high_risk['heifer_ratio'])
            self.ws.cell(row=chart_row + 4, column=data_col + 1, value=high_risk['total_ratio'])

            # 设置数据和分类
            data = Reference(self.ws, min_col=data_col + 1, min_row=chart_row + 2, max_row=chart_row + 4)
            cats = Reference(self.ws, min_col=data_col, min_row=chart_row + 2, max_row=chart_row + 4)

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            # 设置图表样式
            chart.height = 6  # 6cm高
            chart.width = 12  # 12cm宽
            chart.x_axis.title = "风险占比"
            chart.y_axis.title = ""
            chart.legend = None  # 不显示图例

            # X轴设置（百分比格式，0-15%）
            chart.x_axis.scaling.min = 0
            chart.x_axis.scaling.max = 0.15  # 最大15%
            chart.x_axis.numFmt = '0%'

            # 放置图表（在表格右侧，从I列开始）
            anchor_cell = f"I{chart_row}"
            self.ws.add_chart(chart, anchor_cell)

        except Exception as e:
            logger.error(f"创建公牛图表失败: {e}")

    def _get_thin_border(self):
        """获取细边框样式"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

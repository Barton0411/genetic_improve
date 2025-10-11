"""
Sheet 6构建器: 配种记录-近交系数分析
v1.2版本 - 配种记录中的近交系数分析
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class Sheet6Builder(BaseSheetBuilder):
    """
    Sheet 6: 配种记录-近交系数分析

    包含内容:
    1. 近交系数分布汇总表（全部配次）
    2. 近交系数分布汇总表（近12个月）
    3. 分布柱状图 × 2
    4. 风险占比饼图 × 2
    5. 按年份近交系数趋势表
    """

    def build(self, data: dict):
        """
        构建Sheet 6: 配种记录-近交系数分析

        Args:
            data: 包含配种记录中近交系数分析数据
                - all_years_distribution: 全部年份分布
                - recent_12m_distribution: 近12个月分布
                - date_range: 近12个月日期范围
                - yearly_trend: 按年份趋势
        """
        try:
            # 检查数据
            if not data or 'all_years_distribution' not in data:
                logger.warning("Sheet6: 缺少数据，跳过生成")
                return

            # 创建Sheet
            self._create_sheet("配种记录-近交系数分析")
            logger.info("构建Sheet 6: 配种记录-近交系数分析")

            # 定义颜色
            self.color_light_blue = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type='solid')
            self.color_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type='solid')
            self.color_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type='solid')
            self.color_deep_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')

            current_row = 1

            # 1. 构建表1：全部配次的分布
            all_years_dist = data.get('all_years_distribution', {})
            if all_years_dist:
                title = f"近交系数分布汇总表（全部配种记录年份，共{all_years_dist.get('total', 0)}配次）"
                table_start = current_row
                current_row = self._build_distribution_table(
                    title=title,
                    distribution_data=all_years_dist,
                    start_row=current_row
                )

                # 数据行起始位置 = 表起始行 + 标题行 + 表头行
                data_start_row = table_start + 2

                # 图表1：柱状图（与表格第一行平齐）
                chart_row = table_start
                self._create_bar_chart(
                    title="近交系数分布（全部配次）",
                    data_start_row=data_start_row,
                    chart_position=f"F{chart_row}"
                )

                # 图表2：饼图（与表格第一行平齐）
                self._create_pie_chart(
                    title="风险等级占比（全部配次）",
                    data_start_row=data_start_row,
                    chart_position=f"Q{chart_row}"
                )

                current_row += 18  # 图表高度预留

            # 2. 构建表2：近12个月的分布
            recent_12m_dist = data.get('recent_12m_distribution', {})
            date_range = data.get('date_range', {})
            if recent_12m_dist and date_range:
                # 格式化日期范围标题
                start_date = date_range.get('start', '')
                end_date = date_range.get('end', '')
                if start_date and end_date:
                    start_formatted = self._format_date_chinese(start_date)
                    end_formatted = self._format_date_chinese(end_date)
                    title = f"近交系数分布汇总表——{start_formatted}至{end_formatted}（共{recent_12m_dist.get('total', 0)}配次）"
                else:
                    title = f"近交系数分布汇总表（近12个月，共{recent_12m_dist.get('total', 0)}配次）"

                table_start = current_row
                current_row = self._build_distribution_table(
                    title=title,
                    distribution_data=recent_12m_dist,
                    start_row=current_row
                )

                # 数据行起始位置 = 表起始行 + 标题行 + 表头行
                data_start_row = table_start + 2

                # 图表3：柱状图（与表格第一行平齐）
                chart_row = table_start
                self._create_bar_chart(
                    title="近交系数分布（近12个月）",
                    data_start_row=data_start_row,
                    chart_position=f"F{chart_row}"
                )

                # 图表4：饼图（与表格第一行平齐）
                self._create_pie_chart(
                    title="风险等级占比（近12个月）",
                    data_start_row=data_start_row,
                    chart_position=f"Q{chart_row}"
                )

                current_row += 18  # 图表高度预留

            # 3. 构建表3：按年份趋势
            yearly_trend = data.get('yearly_trend', [])
            if yearly_trend:
                current_row = self._build_yearly_trend_table(
                    data=yearly_trend,
                    start_row=current_row
                )

            # 设置列宽
            self._set_column_widths({
                1: 20,  # 近交系数区间/年份
                2: 12,  # 配种头次/总配种次数
                3: 20,  # 占比/高风险配种数（6.25%-12.5%）
                4: 12,  # 风险等级/高风险占比
                5: 20,  # 极高风险配种数（>12.5%）
                6: 12,  # 极高风险占比
            })

            # 冻结首行
            self._freeze_panes('A2')

            logger.info("✓ Sheet 6构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 6失败: {e}", exc_info=True)
            raise

    def _build_distribution_table(self, title: str, distribution_data: dict, start_row: int) -> int:
        """
        构建近交系数分布汇总表

        Args:
            title: 表格标题
            distribution_data: 分布数据
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头行
        headers = ["近交系数区间", "配种头次", "占比", "风险等级"]
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据行
        intervals = distribution_data.get('intervals', [])
        counts = distribution_data.get('counts', [])
        ratios = distribution_data.get('ratios', [])
        risk_levels = distribution_data.get('risk_levels', [])

        # 颜色映射
        colors = [self.color_light_blue, self.color_yellow, self.color_red, self.color_deep_red]

        for i in range(len(intervals)):
            # 近交系数区间
            cell = self.ws.cell(row=current_row, column=1, value=intervals[i])
            cell.fill = colors[i]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border
            cell.font = Font(name='微软雅黑', size=10, bold=(i >= 2))  # 高风险和极高加粗

            # 配种头次
            cell = self.ws.cell(row=current_row, column=2, value=counts[i])
            cell.fill = colors[i]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border
            cell.font = Font(name='微软雅黑', size=10)

            # 占比
            cell = self.ws.cell(row=current_row, column=3, value=f"{ratios[i]*100:.1f}%")
            cell.fill = colors[i]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border
            cell.font = Font(name='微软雅黑', size=10)

            # 风险等级
            cell = self.ws.cell(row=current_row, column=4, value=risk_levels[i])
            cell.fill = colors[i]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border
            cell.font = Font(name='微软雅黑', size=10, bold=(i >= 2))

            current_row += 1

        # 总计行
        total_count = distribution_data.get('total', 0)
        cell = self.ws.cell(row=current_row, column=1, value="总计")
        self.style_manager.apply_total_style(cell)

        cell = self.ws.cell(row=current_row, column=2, value=total_count)
        self.style_manager.apply_total_style(cell)

        cell = self.ws.cell(row=current_row, column=3, value="100.0%")
        self.style_manager.apply_total_style(cell)

        cell = self.ws.cell(row=current_row, column=4, value="-")
        self.style_manager.apply_total_style(cell)

        current_row += 2  # 空1行

        return current_row

    def _create_bar_chart(self, title: str, data_start_row: int, chart_position: str):
        """
        创建柱状图

        Args:
            title: 图表标题
            data_start_row: 数据起始行（表头之后的第一行数据）
            chart_position: 图表位置（如'F10'）
        """
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
        from openpyxl.chart.series import DataPoint

        chart = BarChart()
        chart.title = title
        chart.style = 10
        chart.height = 10
        chart.width = 13  # 10 * 4/3 ≈ 13.33

        # 数据引用（4个区间）
        data = Reference(self.ws, min_col=2, min_row=data_start_row, max_row=data_start_row + 3)
        cats = Reference(self.ws, min_col=1, min_row=data_start_row, max_row=data_start_row + 3)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # 设置轴标签
        chart.x_axis.title = "近交系数区间"
        chart.y_axis.title = "配种头次"

        # 设置柱子颜色（对应风险等级）
        # 浅蓝色、黄色、红色、深红色
        colors = ["9BC2E6", "FFEB9C", "FFC7CE", "FF0000"]

        if chart.series:
            series = chart.series[0]
            # 为每个数据点设置颜色
            for idx, color in enumerate(colors):
                # 创建DataPoint对象
                pt = DataPoint(idx=idx)

                # 设置填充颜色
                pt.spPr = GraphicalProperties()
                pt.spPr.solidFill = ColorChoice(srgbClr=color)

                series.dPt.append(pt)

        # 添加图表到worksheet
        self.ws.add_chart(chart, chart_position)

    def _create_pie_chart(self, title: str, data_start_row: int, chart_position: str):
        """
        创建饼图

        Args:
            title: 图表标题
            data_start_row: 数据起始行
            chart_position: 图表位置
        """
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import ColorChoice
        from openpyxl.chart.series import DataPoint

        chart = PieChart()
        chart.title = title
        chart.style = 10
        chart.height = 10
        chart.width = 10  # 减小宽度，与柱状图保持一致

        # 数据引用
        data = Reference(self.ws, min_col=2, min_row=data_start_row, max_row=data_start_row + 3)
        labels = Reference(self.ws, min_col=4, min_row=data_start_row, max_row=data_start_row + 3)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)

        # 设置扇形颜色（对应风险等级）
        # 浅蓝色、黄色、红色、深红色
        colors = ["9BC2E6", "FFEB9C", "FFC7CE", "FF0000"]

        if chart.series:
            series = chart.series[0]
            # 为每个数据点设置颜色
            for idx, color in enumerate(colors):
                # 创建DataPoint对象
                pt = DataPoint(idx=idx)

                # 设置填充颜色
                pt.spPr = GraphicalProperties()
                pt.spPr.solidFill = ColorChoice(srgbClr=color)

                series.dPt.append(pt)

        # 显示百分比
        chart.dataLabels = None  # 简化处理

        # 添加图表到worksheet
        self.ws.add_chart(chart, chart_position)

    def _build_yearly_trend_table(self, data: list, start_row: int) -> int:
        """
        构建按年份近交系数趋势表

        Args:
            data: 年份趋势数据列表
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        title_cell = self.ws.cell(row=current_row, column=1, value="按年份近交系数趋势")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头行
        headers = ["年份", "总配种次数", "高风险配种数（6.25%-12.5%）", "高风险占比",
                   "极高风险配种数（>12.5%）", "极高风险占比"]
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据行
        for year_data in data:
            # 年份
            cell = self.ws.cell(row=current_row, column=1, value=year_data['year'])
            self.style_manager.apply_data_style(cell, alignment='center')

            # 总配种次数
            cell = self.ws.cell(row=current_row, column=2, value=year_data['total_count'])
            self.style_manager.apply_data_style(cell, alignment='center')

            # 高风险配种数（6.25%-12.5%）
            cell = self.ws.cell(row=current_row, column=3, value=year_data['high_risk_count'])
            self.style_manager.apply_data_style(cell, alignment='center')
            # 如果有高风险，标红
            if year_data['high_risk_count'] > 0:
                cell.font = Font(name='微软雅黑', size=10, color="FF0000", bold=True)

            # 高风险占比
            cell = self.ws.cell(row=current_row, column=4, value=f"{year_data['high_risk_ratio']*100:.1f}%")
            self.style_manager.apply_data_style(cell, alignment='center')
            if year_data['high_risk_ratio'] > 0:
                cell.font = Font(name='微软雅黑', size=10, color="FF0000", bold=True)

            # 极高风险配种数（>12.5%）
            cell = self.ws.cell(row=current_row, column=5, value=year_data['extreme_risk_count'])
            self.style_manager.apply_data_style(cell, alignment='center')
            # 如果有极高风险，标深红
            if year_data['extreme_risk_count'] > 0:
                cell.font = Font(name='微软雅黑', size=10, color="FF0000", bold=True)

            # 极高风险占比
            cell = self.ws.cell(row=current_row, column=6, value=f"{year_data['extreme_risk_ratio']*100:.1f}%")
            self.style_manager.apply_data_style(cell, alignment='center')
            if year_data['extreme_risk_ratio'] > 0:
                cell.font = Font(name='微软雅黑', size=10, color="FF0000", bold=True)

            current_row += 1

        return current_row

    def _format_date_chinese(self, date_str: str) -> str:
        """
        将YYYY-MM-DD格式转换为YYYY年MM月DD日

        Args:
            date_str: 日期字符串 (YYYY-MM-DD)

        Returns:
            中文日期格式 (YYYY年MM月DD日)
        """
        try:
            parts = date_str.split('-')
            if len(parts) == 3:
                year, month, day = parts
                return f"{year}年{int(month)}月{int(day)}日"
            return date_str
        except Exception as e:
            logger.warning(f"日期格式转换失败: {date_str}, 错误: {e}")
            return date_str

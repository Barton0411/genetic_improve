"""
Sheet 4构建器: 育种指数分布分析
"""

from .base_builder import BaseSheetBuilder
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import pandas as pd
import numpy as np
import logging
import tempfile
import os

logger = logging.getLogger(__name__)


class Sheet4IndexDistributionBuilder(BaseSheetBuilder):
    """Sheet 4: 育种指数分布分析"""

    def build(self, data: dict):
        """
        构建Sheet 4

        Args:
            data: {
                'distribution_present': 在群母牛育种指数分布DataFrame,
                'distribution_all': 全部母牛育种指数分布DataFrame,
                'detail_df': 明细数据DataFrame（用于正态分布图）
            }
        """
        try:
            # 创建Sheet
            self._create_sheet("育种指数分布分析")
            logger.info("构建Sheet 4: 育种指数分布分析")

            present_df = data.get('distribution_present')
            all_df = data.get('distribution_all')
            detail_df = data.get('detail_df')

            if present_df is None or present_df.empty:
                logger.warning("Sheet 3-2: 在群母牛育种指数分布数据为空，跳过构建")
                return

            current_row = 1

            # === 1. 上方：分布统计表（在群母牛和全部母牛并排，间隔2列） ===
            # 在群母牛表（从A列开始）
            cell = self.ws.cell(row=current_row, column=1, value="在群母牛育种指数分布")
            self.style_manager.apply_title_style(cell)
            current_row += 1

            # 写入在群母牛表头和数据
            headers_present = list(present_df.columns)
            self._write_header(current_row, headers_present, start_col=1)
            current_row += 1

            for idx, row in present_df.iterrows():
                values = list(row)
                self._write_data_row(current_row, values, start_col=1, alignment='center')
                current_row += 1

            # 全部母牛表（间隔6列）
            if all_df is not None and not all_df.empty:
                right_start_col = len(headers_present) + 1 + 6  # 左表列数 + 1 + 空6列

                current_row_right = 1
                cell = self.ws.cell(row=current_row_right, column=right_start_col, value="全部母牛育种指数分布")
                self.style_manager.apply_title_style(cell)
                current_row_right += 1

                # 写入全部母牛表头和数据
                headers_all = list(all_df.columns)
                self._write_header(current_row_right, headers_all, start_col=right_start_col)
                current_row_right += 1

                for idx, row in all_df.iterrows():
                    values = list(row)
                    self._write_data_row(current_row_right, values, start_col=right_start_col, alignment='center')
                    current_row_right += 1

            # === 2. 下方：图表区域（垂直排列：饼图、柱状图、正态分布图） ===
            chart_start_row = max(current_row, current_row_right if 'current_row_right' in locals() else current_row) + 2

            # 图表大小（保持原始尺寸）
            chart_width = 15  # 15cm宽
            chart_height = 10  # 10cm高
            # 图表占用空间：横7列 × 竖20行（用于定位）
            chart_row_span = 20  # 每个图表占20行
            row_gap = 2  # 图表纵向间隔2行

            # === 在群母牛图表（左侧） ===
            current_chart_row = chart_start_row
            left_data_col = 1  # 在群母牛数据从第1列开始

            # 饼图
            self._create_pie_chart(present_df, "在群母牛育种指数分布占比",
                                  current_chart_row, 1, chart_height, chart_width, left_data_col)
            current_chart_row += chart_row_span + row_gap

            # 柱状图
            self._create_bar_chart(present_df, "在群母牛育种指数分布柱状图",
                                  current_chart_row, 1, chart_height, chart_width, left_data_col)
            current_chart_row += chart_row_span + row_gap

            # === 全部母牛图表（右侧） ===
            # 图表占用：横7列 + 空2列 = 从第10列开始（在条件外定义，避免后续使用时未定义）
            right_chart_col = 1 + 7 + 2  # 左侧图表起始列(1) + 图表占用(7列) + 间隔(2列)
            if all_df is not None and not all_df.empty:
                current_chart_row_right = chart_start_row
                right_data_col = right_start_col  # 全部母牛数据的起始列

                # 饼图
                self._create_pie_chart(all_df, "全部母牛育种指数分布占比",
                                      current_chart_row_right, right_chart_col, chart_height, chart_width, right_data_col)
                current_chart_row_right += chart_row_span + row_gap

                # 柱状图
                self._create_bar_chart(all_df, "全部母牛育种指数分布柱状图",
                                      current_chart_row_right, right_chart_col, chart_height, chart_width, right_data_col)
                current_chart_row_right += chart_row_span + row_gap

            # === 3. 整体正态分布 ===
            overall_start_row = current_chart_row + row_gap + 2

            # 添加分类标题
            from openpyxl.styles import Font
            title_cell = self.ws.cell(row=overall_start_row, column=1, value="整体正态分布")
            title_cell.font = Font(size=14, bold=True)
            overall_start_row += 2

            # 先创建正态分布图（使用明细数据）
            if detail_df is not None and not detail_df.empty:
                # 在群母牛正态分布图（左侧）
                present_detail_df = detail_df[(detail_df['是否在场'] == '是') & (detail_df['sex'] == '母')].copy()
                present_stats = self._create_normal_distribution_chart(present_detail_df, 'index_score', "在群母牛育种指数正态分布",
                                                       overall_start_row, 1, chart_height, chart_width, 52)

                # 全部母牛正态分布图（右侧）
                all_detail_df = detail_df[detail_df['sex'] == '母'].copy()
                all_stats = self._create_normal_distribution_chart(all_detail_df, 'index_score', "全部母牛育种指数正态分布",
                                                       overall_start_row, right_chart_col, chart_height, chart_width, 56)

                # 在图表右侧添加统计信息表
                if present_stats or all_stats:
                    stats_start_col = right_chart_col + 7 + 2  # 右侧图表后面再空2列
                    self._create_single_chart_stats_table(present_stats, all_stats,
                                                          overall_start_row, stats_start_col, '育种指数')

                # 图表占用约20行
                matplotlib_chart_rows = 20
                overall_start_row += matplotlib_chart_rows + 3
            else:
                logger.warning("明细数据不存在，跳过整体正态分布图")

            # 在正态分布图之后，创建五等份分析表
            if detail_df is not None and not detail_df.empty:
                # 为在群母牛和全部母牛生成整体五等份分析表
                present_detail_df = detail_df[(detail_df['是否在场'] == '是') & (detail_df['sex'] == '母')].copy()
                all_detail_df = detail_df[detail_df['sex'] == '母'].copy()

                # 生成在群母牛统计
                present_stats = self._create_single_group_distribution(
                    detail_df, 'index_score', '在群母牛', lambda df: (df['是否在场'] == '是') & (df['sex'] == '母'),
                    filter_present=False  # 已经在lambda中过滤了
                )

                # 生成全部母牛统计
                all_stats = self._create_single_group_distribution(
                    detail_df, 'index_score', '全部母牛', lambda df: df['sex'] == '母',
                    filter_present=False  # 已经在lambda中过滤了
                )

                # 创建横向并排的表格
                if present_stats or all_stats:
                    rows_used = self._create_group_pair_tables(
                        '育种指数五等份分析', present_stats, all_stats, '育种指数', overall_start_row
                    )
                    overall_start_row += rows_used + 3  # 表格高度 + 间隔3行

            # === 4. 其他分组的正态分布分析 ===
            if detail_df is not None and not detail_df.empty:
                # 在整体正态分布表格下方开始（只需要2行间隔）
                next_section_row = overall_start_row + 2

                # 4.1 成母牛/后备牛2组分布图
                next_row = self._create_maturity_group_charts(detail_df, 'index_score', next_section_row)

                # 4.2 胎次-月龄4组分布图
                next_row = self._create_parity_age_group_charts(detail_df, 'index_score', next_row + 3)

                # 4.3 出生年份分组分布图
                next_row = self._create_birth_year_group_charts(detail_df, 'index_score', next_row + 3)

            # 设置列宽
            for col in range(1, 40):  # 设置足够多的列
                self.ws.column_dimensions[get_column_letter(col)].width = 12

            # 冻结首行
            self._freeze_panes('A2')

            logger.info(f"✓ Sheet 3-2构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 4失败: {e}", exc_info=True)
            raise

    def _create_bar_chart(self, df, title, chart_row, chart_col, height_cm, width_cm, data_col):
        """创建柱状图"""
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = title
            chart.style = 10
            chart.y_axis.title = "头数"
            chart.x_axis.title = "分布区间"

            # 数据范围（第2行是表头，第3行开始是数据）
            # data_col是数据起始列，data_col+1是头数列
            data = Reference(self.ws, min_col=data_col + 1, min_row=2, max_row=2 + len(df))
            cats = Reference(self.ws, min_col=data_col, min_row=3, max_row=2 + len(df))

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # 设置图表大小
            chart.width = width_cm
            chart.height = height_cm

            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"
            self.ws.add_chart(chart, anchor_cell)

        except Exception as e:
            logger.error(f"创建柱状图失败: {e}")

    def _create_pie_chart(self, df, title, chart_row, chart_col, height_cm, width_cm, data_col):
        """创建饼图"""
        try:
            chart = PieChart()
            chart.title = title
            chart.style = 10

            # 数据范围（第2行是表头，第3行开始是数据）
            # data_col是数据起始列，data_col+1是头数列
            data = Reference(self.ws, min_col=data_col + 1, min_row=2, max_row=2 + len(df))
            cats = Reference(self.ws, min_col=data_col, min_row=3, max_row=2 + len(df))

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # 设置图表大小
            chart.width = width_cm
            chart.height = height_cm

            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"
            self.ws.add_chart(chart, anchor_cell)

        except Exception as e:
            logger.error(f"创建饼图失败: {e}")

    def _create_normal_distribution_chart(self, detail_df, score_column, title, chart_row, chart_col, height_cm, width_cm, data_start_col=52):
        """
        创建正态分布曲线图（仅使用matplotlib）

        Args:
            detail_df: 明细数据DataFrame
            score_column: 分数列名（如'index_score'）
            title: 图表标题
            chart_row: 图表起始行
            chart_col: 图表起始列
            height_cm: 图表高度（未使用，保留以兼容旧调用）
            width_cm: 图表宽度（未使用，保留以兼容旧调用）
            data_start_col: 数据写入起始列（未使用，保留以兼容旧调用）

        Returns:
            dict: 统计数据字典，包含 sample_size, mean, std, cv, median, q1, q3, min, max
        """
        try:
            import numpy as np

            # 获取有效数据
            valid_data = detail_df[score_column].dropna()

            if len(valid_data) == 0:
                logger.warning(f"没有有效的{score_column}数据，跳过正态分布图")
                return None

            # 计算统计指标
            n = len(valid_data)
            mu = valid_data.mean()
            sigma = valid_data.std()
            cv = (sigma / mu * 100) if mu != 0 else 0
            median = valid_data.median()
            q1 = valid_data.quantile(0.25)
            q3 = valid_data.quantile(0.75)
            min_val = valid_data.min()
            max_val = valid_data.max()

            # 直接创建matplotlib分布图
            try:
                self._create_matplotlib_distribution(valid_data, title, mu, sigma,
                                                    chart_row, chart_col, score_column)
            except Exception as e:
                logger.warning(f"创建matplotlib分布图失败: {e}")

            # 返回统计数据（键名需要与_create_single_chart_stats_table方法期望的一致）
            return {
                'sample_size': n,
                'mean': mu,
                'std': sigma,
                'cv': cv,
                'median': median,
                'q1': q1,
                'q3': q3,
                'min': min_val,
                'max': max_val
            }

        except Exception as e:
            logger.error(f"创建正态分布图失败: {e}")
            return None

    def _create_simple_distribution_chart(self, detail_df, score_column, title, chart_row, chart_col, height_cm, width_cm):
        """创建简化的频率分布图（当scipy不可用时）"""
        try:
            import numpy as np

            # 获取有效数据
            valid_data = detail_df[score_column].dropna()

            if len(valid_data) == 0:
                logger.warning(f"没有有效的{score_column}数据")
                return

            # 计算直方图数据
            counts, bin_edges = np.histogram(valid_data, bins=20)
            bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2

            # 在工作表中写入数据（放在远离可见区域）
            data_start_row = chart_row
            data_start_col = 52  # AZ列（第52列），远离可见区域

            for i, (x_val, y_val) in enumerate(zip(bin_centers, counts)):
                self.ws.cell(row=data_start_row + i + 1, column=data_start_col, value=int(round(x_val)))
                self.ws.cell(row=data_start_row + i + 1, column=data_start_col + 1, value=int(y_val))

            # 创建折线图
            chart = LineChart()
            chart.title = title
            chart.style = 13
            chart.y_axis.title = "频数"
            chart.x_axis.title = "育种指数值"

            data = Reference(self.ws, min_col=data_start_col + 1, min_row=data_start_row + 1,
                           max_row=data_start_row + len(bin_centers))
            cats = Reference(self.ws, min_col=data_start_col, min_row=data_start_row + 1,
                           max_row=data_start_row + len(bin_centers))

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            chart.width = width_cm
            chart.height = height_cm

            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"
            self.ws.add_chart(chart, anchor_cell)

        except Exception as e:
            logger.error(f"创建简化分布图失败: {e}")

    def _create_parity_age_group_charts(self, detail_df, score_column, start_row):
        """
        按胎次-月龄分组创建正态分布图（多曲线）

        新布局（参考示意表）：
        - 分类标题："不同阶段牛群正态分布"
        - 2个多曲线matplotlib图（在群母牛 + 全部母牛），每个图包含4条曲线
        - 图表下方：4个组，每个组包含2个横向并排的表（在群母牛 + 全部母牛）
        - 组之间垂直排列，组内表格横向并排

        Args:
            detail_df: 明细数据DataFrame
            score_column: 分数列名（如'index_score', 'TPI_score'）
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        try:
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            # 提取性状名称
            trait_name = score_column.replace('_score', '')

            # 分类标题
            title_cell = self.ws.cell(row=start_row, column=1, value="不同阶段牛群正态分布")
            title_cell.font = Font(size=14, bold=True)
            start_row += 2

            # 定义分组（修正：age是年龄，单位为年）
            groups = [
                ('2胎及以上组', lambda df: df['lac'] >= 2, '#1f77b4'),
                ('1胎组', lambda df: df['lac'] == 1, '#ff7f0e'),
                ('12月龄以上0胎牛', lambda df: (df['lac'] == 0) & (df['age'] >= 1.0), '#2ca02c'),
                ('12月龄以下0胎牛', lambda df: (df['lac'] == 0) & (df['age'] < 1.0), '#d62728')
            ]

            # === 1. 创建两个多曲线matplotlib图 ===
            current_row = start_row

            # 1.1 在群母牛多曲线图（左侧）
            present_title = f"在群母牛{trait_name}不同阶段牛群正态分布"
            present_stats = self._create_multi_group_distribution(detail_df, score_column, groups,
                                                  present_title, current_row, 1,
                                                  filter_present=True)

            # 1.2 全部母牛多曲线图（右侧）
            all_title = f"全部母牛{trait_name}不同阶段牛群正态分布"
            right_chart_col = 1 + 7 + 2  # 左侧图表(1) + 占用7列 + 间隔2列
            all_stats = self._create_multi_group_distribution(detail_df, score_column, groups,
                                                  all_title, current_row, right_chart_col,
                                                  filter_present=False)

            # === 1.3 在图表右侧添加统计信息表 ===
            if present_stats or all_stats:
                stats_start_col = right_chart_col + 7 + 2  # 右侧图表后面再空2列
                self._create_multi_group_stats_table(present_stats, all_stats,
                                                     current_row, stats_start_col, trait_name)

            # matplotlib图表占用约20行
            matplotlib_chart_rows = 20
            current_row += matplotlib_chart_rows + 3

            # === 2. 在图表下方创建表格 ===
            # 为每个组分别生成统计数据
            for group_name, filter_func, color in groups:
                # 为在群母牛生成统计
                present_stats = self._create_single_group_distribution(
                    detail_df, score_column, group_name, filter_func,
                    filter_present=True
                )

                # 为全部母牛生成统计
                all_stats = self._create_single_group_distribution(
                    detail_df, score_column, group_name, filter_func,
                    filter_present=False
                )

                # 生成横向并排的两个表
                if present_stats or all_stats:
                    rows_used = self._create_group_pair_tables(
                        group_name, present_stats, all_stats, trait_name, current_row
                    )
                    current_row += rows_used + 2  # 表格高度 + 间隔2行

            logger.info(f"✓ 已添加{trait_name}胎次-月龄分组分析（2个多曲线图 + 4组表格）")

            return current_row

        except ImportError as e:
            logger.warning(f"matplotlib或scipy未安装，跳过胎次-月龄分组图: {e}")
            return start_row
        except Exception as e:
            logger.error(f"创建胎次-月龄分组图表失败: {e}", exc_info=True)
            return start_row

    def _create_single_group_distribution(self, detail_df, score_column, group_name, filter_func, filter_present=True):
        """
        为单个组生成统计数据（不绘制图表）

        Args:
            detail_df: 明细数据
            score_column: 分数列名
            group_name: 组名
            filter_func: 过滤函数
            filter_present: 是否只统计在群母牛

        Returns:
            统计数据字典 {'group_name', 'total_count', 'max_val', 'min_val', 'mean', 'quintiles'} 或 None
        """
        try:
            import pandas as pd

            # 筛选数据
            if filter_present:
                group_data = detail_df[(detail_df['是否在场'] == '是') &
                                      (detail_df['sex'] == '母') &
                                      filter_func(detail_df)].copy()
            else:
                group_data = detail_df[(detail_df['sex'] == '母') &
                                      filter_func(detail_df)].copy()

            if len(group_data) == 0 or group_data[score_column].notna().sum() == 0:
                return None

            valid_data = group_data[score_column].dropna()
            n = len(valid_data)
            mu = valid_data.mean()
            max_val = valid_data.max()
            min_val = valid_data.min()

            if n < 2:
                return None

            # 计算20%分位统计数据
            quintile_stats = self._calculate_quintile_stats(valid_data, group_name, score_column)

            return {
                'group_name': group_name,
                'total_count': n,
                'max_val': max_val,
                'min_val': min_val,
                'mean': mu,
                'quintiles': quintile_stats
            }

        except Exception as e:
            logger.error(f"生成单组统计数据失败 ({group_name}): {e}")
            return None

    def _calculate_quintile_stats(self, valid_data, group_name, score_column):
        """
        计算某组数据的20%分位统计

        Args:
            valid_data: 有效数据Series
            group_name: 组名
            score_column: 分数列名

        Returns:
            分位统计列表 [(排名分组, 头数, 最大值, 最小值, 平均得分), ...]
        """
        import pandas as pd
        import numpy as np

        # 按分数降序排列
        sorted_data = valid_data.sort_values(ascending=False).reset_index(drop=True)
        total_count = len(sorted_data)

        quintile_labels = ['最高20%', '2nd 20%', '3rd 20%', '4th 20%', '最差20%']
        quintile_stats = []

        for i in range(5):
            start_idx = int(total_count * i / 5)
            end_idx = int(total_count * (i + 1) / 5)

            if start_idx >= end_idx:
                continue

            quintile_data = sorted_data.iloc[start_idx:end_idx]

            if len(quintile_data) > 0:
                quintile_stats.append((
                    quintile_labels[i],
                    len(quintile_data),
                    quintile_data.max(),
                    quintile_data.min(),
                    quintile_data.mean()
                ))

        return quintile_stats

    def _create_multi_group_distribution(self, detail_df, score_column, groups, title,
                                         chart_row, chart_col, filter_present=True):
        """
        创建包含多个组的正态分布图

        Args:
            detail_df: 明细数据
            score_column: 分数列名
            groups: 分组列表 [(组名, 过滤函数, 颜色), ...]
            title: 图表标题
            chart_row: 图表起始行
            chart_col: 图表起始列
            filter_present: 是否只统计在群母牛

        Returns:
            统计数据列表，每个元素是包含组信息和quintiles的字典
        """
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            matplotlib.use('Agg')
            from matplotlib.ticker import MaxNLocator
            from scipy import stats
            import numpy as np
            import tempfile
            from openpyxl.drawing.image import Image
            from openpyxl.utils import get_column_letter
            import pandas as pd

            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False

            # 创建图表
            fig, ax = plt.subplots(figsize=(10, 6))

            # 为每个组绘制正态分布曲线并收集统计数据
            all_stats = []
            all_data_min = None
            all_data_max = None
            all_curve_heights = []  # 收集所有组的峰值高度（用于排除异常值）

            for group_name, filter_func, color in groups:
                # 筛选数据
                if filter_present:
                    group_data = detail_df[(detail_df['是否在场'] == '是') &
                                          (detail_df['sex'] == '母') &
                                          filter_func(detail_df)].copy()
                else:
                    group_data = detail_df[(detail_df['sex'] == '母') &
                                          filter_func(detail_df)].copy()

                if len(group_data) == 0 or group_data[score_column].notna().sum() == 0:
                    continue

                valid_data = group_data[score_column].dropna()
                n = len(valid_data)
                mu = valid_data.mean()
                sigma = valid_data.std()
                max_val = valid_data.max()
                min_val = valid_data.min()

                # 更新全局最小最大值
                if all_data_min is None or min_val < all_data_min:
                    all_data_min = min_val
                if all_data_max is None or max_val > all_data_max:
                    all_data_max = max_val

                if n < 2 or sigma == 0:
                    continue

                # 绘制直方图（频率密度）
                ax.hist(valid_data, bins=30, density=True, alpha=0.3,
                       color=color, edgecolor='none')

                # 绘制正态分布曲线
                x = np.linspace(valid_data.min(), valid_data.max(), 200)
                y = stats.norm.pdf(x, mu, sigma)
                ax.plot(x, y, linewidth=2, color=color,
                       label=f'{group_name} (n={n}, μ={mu:.1f}, σ={sigma:.1f})')

                # 收集曲线峰值高度
                curve_max = y.max()
                all_curve_heights.append(curve_max)

                # 计算20%分位统计数据
                quintile_stats = self._calculate_quintile_stats(valid_data, group_name, score_column)

                # 收集统计数据（包含分位数详情）
                all_stats.append({
                    'group_name': group_name,
                    'total_count': n,
                    'max_val': max_val,
                    'min_val': min_val,
                    'mean': mu,
                    'quintiles': quintile_stats  # 20%分位详细数据
                })

            # 设置X轴范围（动态，基于所有组的数据范围，并留出10%的边距）
            if all_data_min is not None and all_data_max is not None:
                x_margin = (all_data_max - all_data_min) * 0.1
                ax.set_xlim(all_data_min - x_margin, all_data_max + x_margin)

            # 设置Y轴范围：如果最高峰值是第二高峰值的3倍以上，则使用第二高峰值的2.5倍
            if len(all_curve_heights) > 0:
                # 对峰值排序（从大到小）
                sorted_heights = sorted(all_curve_heights, reverse=True)

                if len(sorted_heights) > 1:
                    highest = sorted_heights[0]
                    second_highest = sorted_heights[1]

                    # 判断最高峰值是否为异常值（是第二高的3倍以上）
                    if highest >= second_highest * 3:
                        # 存在异常值，使用第二高峰值的2.5倍作为Y轴上限
                        y_max = second_highest * 2.5
                        logger.info(f"检测到异常峰值：最高={highest:.4f}，第二高={second_highest:.4f}（最高/第二高={highest/second_highest:.2f}倍），Y轴上限采用第二高×2.5={y_max:.4f}")
                    else:
                        # 无异常值，使用最高峰值的1.3倍
                        y_max = highest * 1.3
                        logger.info(f"未检测到异常峰值：最高={highest:.4f}，第二高={second_highest:.4f}，Y轴上限采用最高×1.3={y_max:.4f}")
                else:
                    # 只有一个分组，使用其峰值的1.3倍
                    y_max = sorted_heights[0] * 1.3

                ax.set_ylim(0, y_max)

            # 设置图表样式（简化版，图例只显示组名）
            ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
            ax.set_xlabel('指数值', fontsize=11)
            ax.set_ylabel('频率密度', fontsize=11)

            # 简化图例，只显示组名
            handles, labels = ax.get_legend_handles_labels()
            simple_labels = [label.split(' (')[0] for label in labels]
            ax.legend(handles, simple_labels, loc='upper right', fontsize=9, framealpha=0.9)

            ax.grid(True, alpha=0.3, linestyle='--')
            ax.xaxis.set_major_locator(MaxNLocator(integer=True, nbins=10))

            plt.tight_layout()

            # 保存为临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            plt.savefig(temp_file.name, dpi=150, bbox_inches='tight',
                       facecolor='white', edgecolor='none')
            plt.close(fig)

            # 插入图片到Excel
            img = Image(temp_file.name)
            img.width = int(15 * 0.65 * 150 / 2.54)  # 9.75cm宽
            img.height = int(10 * 0.65 * 150 / 2.54)  # 6.5cm高

            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"
            self.ws.add_image(img, anchor_cell)

            logger.info(f"✓ 多组正态分布图已嵌入: {title}, 包含{len(all_stats)}个组")

            return all_stats

        except Exception as e:
            logger.error(f"创建多组分布图失败: {e}", exc_info=True)
            return []

    def _create_matplotlib_distribution(self, valid_data, title, mu, sigma, chart_row, chart_col, score_column):
        """
        使用matplotlib创建高质量的正态分布图并嵌入Excel

        Args:
            valid_data: 有效数据Series
            title: 图表标题
            mu: 均值
            sigma: 标准差
            chart_row: 图表起始行
            chart_col: 图表起始列
            score_column: 分数列名
        """
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            from scipy import stats

            # 设置中文字体
            matplotlib.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
            matplotlib.rcParams['axes.unicode_minus'] = False

            # 创建图表
            fig, ax = plt.subplots(figsize=(8, 6), dpi=150)

            # 计算数据范围
            data_min = valid_data.min()
            data_max = valid_data.max()
            data_range = data_max - data_min

            # 绘制直方图（频率密度）
            n, bins, patches = ax.hist(valid_data, bins=30, density=True,
                                       alpha=0.6, color='#4472C4',
                                       edgecolor='white', linewidth=0.5,
                                       label='实际分布')

            # 绘制正态分布曲线
            x_range = np.linspace(data_min, data_max, 200)
            normal_curve = stats.norm.pdf(x_range, mu, sigma)
            ax.plot(x_range, normal_curve, 'r-', linewidth=2.5,
                   label=f'正态分布曲线\n(μ={mu:.2f}, σ={sigma:.2f})')

            # 添加均值线
            ax.axvline(mu, color='red', linestyle='--', linewidth=1.5,
                      alpha=0.7, label=f'均值: {mu:.2f}')

            # 计算正态分布曲线的最高点，用于设置Y轴范围
            curve_max_height = normal_curve.max()

            # 设置X轴范围（动态，基于数据范围，留出10%的边距）
            x_margin = data_range * 0.1
            ax.set_xlim(data_min - x_margin, data_max + x_margin)

            # 设置Y轴范围：从0到正态分布曲线最高点的2倍
            ax.set_ylim(0, curve_max_height * 2)

            # 设置标题和标签
            ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
            trait_name = score_column.replace('_score', '')
            ax.set_xlabel(f'{trait_name}值', fontsize=11)
            ax.set_ylabel('概率密度', fontsize=11)

            # 简化图例
            ax.legend(loc='upper right', fontsize=9, framealpha=0.9)

            # 添加网格
            ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)

            # 美化坐标轴
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_linewidth(1.2)
            ax.spines['bottom'].set_linewidth(1.2)

            # 设置X轴刻度为整数
            from matplotlib.ticker import MaxNLocator
            ax.xaxis.set_major_locator(MaxNLocator(integer=True, nbins=10))

            plt.tight_layout()

            # 计算统计指标（用于返回）
            median = valid_data.median()
            q1 = valid_data.quantile(0.25)
            q3 = valid_data.quantile(0.75)
            cv = (sigma / mu * 100) if mu != 0 else 0

            # 保存为临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            plt.savefig(temp_file.name, dpi=150, bbox_inches='tight',
                       facecolor='white', edgecolor='none')
            plt.close(fig)

            # 插入图片到Excel
            img = Image(temp_file.name)
            # 设置图片大小为原来的0.65倍：9.75cm×6.5cm（厘米转像素，150 DPI）
            img.width = int(15 * 0.65 * 150 / 2.54)  # 9.75cm宽 ≈ 576像素
            img.height = int(10 * 0.65 * 150 / 2.54)  # 6.5cm高 ≈ 384像素

            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"
            self.ws.add_image(img, anchor_cell)

            # 注意：临时文件会在Excel保存后自动清理，这里不删除
            logger.info(f"✓ matplotlib正态分布图已嵌入: {title}")

            # 返回统计数据
            return {
                'sample_size': n,
                'mean': mu,
                'std': sigma,
                'cv': cv,
                'median': median,
                'q1': q1,
                'q3': q3,
                'min': valid_data.min(),
                'max': valid_data.max()
            }

        except ImportError as e:
            logger.warning(f"matplotlib未安装，跳过高质量分布图: {e}")
            return None
        except Exception as e:
            logger.error(f"创建matplotlib分布图失败: {e}", exc_info=True)
            return None

    def _create_group_stats_table(self, stats_data, title, start_row, start_col):
        """
        创建分组统计表格

        Args:
            stats_data: 统计数据列表 [(组名, 头数, 最大值, 最小值, 加权平均分), ...]
            title: 表格标题
            start_row: 起始行
            start_col: 起始列
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # 表头
        header_cell = self.ws.cell(row=start_row, column=start_col, value=title)
        header_cell.font = Font(size=11, bold=True)
        start_row += 1

        # 列标题
        headers = ['排名分组', '头数', '最大值', '最小值', 'NM指数平均得分']
        header_fill = PatternFill(start_color="DAA520", end_color="DAA520", fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header_text in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=start_col + col_idx, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        current_row = start_row + 1

        # 数据行
        for group_name, count, max_val, min_val, avg_val in stats_data:
            # 排名分组
            cell = self.ws.cell(row=current_row, column=start_col, value=group_name)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

            # 头数
            cell = self.ws.cell(row=current_row, column=start_col + 1, value=count)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 最大值
            cell = self.ws.cell(row=current_row, column=start_col + 2, value=int(max_val))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 最小值
            cell = self.ws.cell(row=current_row, column=start_col + 3, value=int(min_val))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 加权平均分
            cell = self.ws.cell(row=current_row, column=start_col + 4, value=int(avg_val))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            current_row += 1

        # 合计行
        total_count = sum(s[1] for s in stats_data)
        cell = self.ws.cell(row=current_row, column=start_col, value='总计')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

        cell = self.ws.cell(row=current_row, column=start_col + 1, value=total_count)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 设置列宽
        self.ws.column_dimensions[get_column_letter(start_col)].width = 18
        for i in range(1, 5):
            self.ws.column_dimensions[get_column_letter(start_col + i)].width = 12

    def _create_group_pair_tables(self, group_name, present_data, all_data, trait_name, start_row):
        """
        为一个组创建横向并排的两个表（在群母牛 + 全部母牛）

        Args:
            group_name: 组名（如 '2胎及以上组', '成母牛组'）
            present_data: 在群母牛数据字典 {'group_name', 'total_count', 'max_val', 'min_val', 'mean', 'quintiles'}
            all_data: 全部母牛数据字典
            trait_name: 性状名称（如 '育种指数'）
            start_row: 起始行

        Returns:
            表格占用的行数
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        current_row = start_row

        # 标题行：组名
        title_cell = self.ws.cell(row=current_row, column=3, value=f"在群母牛-{group_name}")
        title_cell.font = Font(size=11, bold=True)

        title_cell2 = self.ws.cell(row=current_row, column=11, value=f"全部母牛-{group_name}")
        title_cell2.font = Font(size=11, bold=True)
        current_row += 1

        # 转置后的表头（列）
        quintile_headers = ['排名分组', '最高20%', '2nd 20%', '3rd 20%', '4th 20%', '最低20%', '总计']
        header_fill = PatternFill(start_color="DAA520", end_color="DAA520", fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 左表表头（在群母牛，列B-H）
        for col_idx, header_text in enumerate(quintile_headers):
            cell = self.ws.cell(row=current_row, column=2 + col_idx, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 右表表头（全部母牛，列J-P）
        for col_idx, header_text in enumerate(quintile_headers):
            cell = self.ws.cell(row=current_row, column=10 + col_idx, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        current_row += 1

        # 准备数据
        present_quintiles = present_data['quintiles'] if present_data else []
        all_quintiles = all_data['quintiles'] if all_data else []

        # 转置：行是指标，列是分组
        row_labels = ['头数', '最大值', '最小值', f'{trait_name}平均得分']

        for row_idx, row_label in enumerate(row_labels):
            # 左表（在群母牛）
            # 第一列：指标名称
            cell = self.ws.cell(row=current_row + row_idx, column=2, value=row_label)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

            # 计算总计
            present_total_count = sum(q[1] for q in present_quintiles)
            present_total_max = max((q[2] for q in present_quintiles), default=0)
            present_total_min = min((q[3] for q in present_quintiles), default=0)
            present_total_mean = sum(q[1] * q[4] for q in present_quintiles) / present_total_count if present_total_count > 0 else 0

            # 各个分组的数据
            for col_idx in range(5):
                if col_idx < len(present_quintiles):
                    quintile_label, count, qmax, qmin, qmean = present_quintiles[col_idx]
                    if row_idx == 0:  # 头数
                        value = count
                    elif row_idx == 1:  # 最大值
                        value = int(qmax)
                    elif row_idx == 2:  # 最小值
                        value = int(qmin)
                    else:  # 平均得分
                        value = int(qmean)
                else:
                    value = ''

                cell = self.ws.cell(row=current_row + row_idx, column=3 + col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 总计列
            if row_idx == 0:  # 头数
                total_value = present_total_count
            elif row_idx == 1:  # 最大值
                total_value = int(present_total_max)
            elif row_idx == 2:  # 最小值
                total_value = int(present_total_min)
            else:  # 平均得分
                total_value = int(present_total_mean)

            cell = self.ws.cell(row=current_row + row_idx, column=8, value=total_value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 右表（全部母牛）
            # 第一列：指标名称
            cell = self.ws.cell(row=current_row + row_idx, column=10, value=row_label)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

            # 计算总计
            all_total_count = sum(q[1] for q in all_quintiles)
            all_total_max = max((q[2] for q in all_quintiles), default=0)
            all_total_min = min((q[3] for q in all_quintiles), default=0)
            all_total_mean = sum(q[1] * q[4] for q in all_quintiles) / all_total_count if all_total_count > 0 else 0

            # 各个分组的数据
            for col_idx in range(5):
                if col_idx < len(all_quintiles):
                    quintile_label, count, qmax, qmin, qmean = all_quintiles[col_idx]
                    if row_idx == 0:  # 头数
                        value = count
                    elif row_idx == 1:  # 最大值
                        value = int(qmax)
                    elif row_idx == 2:  # 最小值
                        value = int(qmin)
                    else:  # 平均得分
                        value = int(qmean)
                else:
                    value = ''

                cell = self.ws.cell(row=current_row + row_idx, column=11 + col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 总计列
            if row_idx == 0:  # 头数
                total_value = all_total_count
            elif row_idx == 1:  # 最大值
                total_value = int(all_total_max)
            elif row_idx == 2:  # 最小值
                total_value = int(all_total_min)
            else:  # 平均得分
                total_value = int(all_total_mean)

            cell = self.ws.cell(row=current_row + row_idx, column=16, value=total_value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 更新当前行（转置后的表格只有4行数据）
        current_row += 4

        # 返回占用的行数（标题1行 + 表头1行 + 数据4行 = 6行）
        return 6

    def _create_group_quintile_table(self, group_data, trait_name, start_row, start_col):
        """
        为单个组创建20%分位表格（旧方法，保留以备用）

        Args:
            group_data: 组数据字典，包含 'group_name', 'total_count', 'quintiles'等
            trait_name: 性状名称（如 '育种指数'）
            start_row: 起始行
            start_col: 起始列

        Returns:
            表格占用的行数
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        group_name = group_data['group_name']
        quintiles = group_data['quintiles']
        total_count = group_data['total_count']
        max_val = group_data['max_val']
        min_val = group_data['min_val']
        mean = group_data['mean']

        # 标题：组名
        title_cell = self.ws.cell(row=start_row, column=start_col, value=group_name)
        title_cell.font = Font(size=11, bold=True)
        current_row = start_row + 1

        # 表头
        headers = ['排名分组', '头数', '最大值', '最小值', f'{trait_name}平均得分']
        header_fill = PatternFill(start_color="DAA520", end_color="DAA520", fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header_text in enumerate(headers):
            cell = self.ws.cell(row=current_row, column=start_col + col_idx, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        current_row += 1

        # 数据行 - 5个20%分位
        for quintile_label, count, qmax, qmin, qmean in quintiles:
            # 排名分组
            cell = self.ws.cell(row=current_row, column=start_col, value=quintile_label)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

            # 头数
            cell = self.ws.cell(row=current_row, column=start_col + 1, value=count)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 最大值
            cell = self.ws.cell(row=current_row, column=start_col + 2, value=int(qmax))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 最小值
            cell = self.ws.cell(row=current_row, column=start_col + 3, value=int(qmin))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 平均得分
            cell = self.ws.cell(row=current_row, column=start_col + 4, value=int(qmean))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            current_row += 1

        # 总计行
        cell = self.ws.cell(row=current_row, column=start_col, value='总计')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

        cell = self.ws.cell(row=current_row, column=start_col + 1, value=total_count)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 总计行的最大值
        cell = self.ws.cell(row=current_row, column=start_col + 2, value=int(max_val))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 总计行的最小值
        cell = self.ws.cell(row=current_row, column=start_col + 3, value=int(min_val))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 总计行的平均分
        cell = self.ws.cell(row=current_row, column=start_col + 4, value=int(mean))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 设置列宽
        self.ws.column_dimensions[get_column_letter(start_col)].width = 12
        for i in range(1, 5):
            self.ws.column_dimensions[get_column_letter(start_col + i)].width = 11

        # 返回表格占用的行数（标题1行 + 表头1行 + 5个数据行 + 总计行 = 8行）
        return 8

    def _create_maturity_group_charts(self, detail_df, score_column, start_row):
        """
        按成母牛/后备牛分组创建正态分布图

        新布局（参考示意表）：
        - 分类标题："成母牛/后备牛正态分布"
        - 2个多曲线matplotlib图（在群母牛 + 全部母牛），每个图包含2条曲线
        - 图表下方：2个组，每个组包含2个横向并排的表（在群母牛 + 全部母牛）
        - 组之间垂直排列，组内表格横向并排

        Args:
            detail_df: 明细数据DataFrame
            score_column: 分数列名
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        try:
            from openpyxl.styles import Font

            # 提取性状名称
            trait_name = score_column.replace('_score', '')

            # 分类标题
            title_cell = self.ws.cell(row=start_row, column=1, value="成母牛/后备牛正态分布")
            title_cell.font = Font(size=14, bold=True)
            start_row += 2

            # 定义分组
            groups = [
                ('成母牛组', lambda df: df['lac'] > 0, '#1f77b4'),
                ('后备牛组', lambda df: (df['lac'] == 0) | df['lac'].isna(), '#ff7f0e')
            ]

            # === 1. 创建两个多曲线matplotlib图 ===
            current_row = start_row

            # 1.1 在群母牛多曲线图（左侧）
            present_title = f"在群母牛{trait_name}成母牛/后备牛正态分布"
            present_stats = self._create_multi_group_distribution(detail_df, score_column, groups,
                                                  present_title, current_row, 1,
                                                  filter_present=True)

            # 1.2 全部母牛多曲线图（右侧）
            all_title = f"全部母牛{trait_name}成母牛/后备牛正态分布"
            right_chart_col = 1 + 7 + 2  # 左侧图表(1) + 占用7列 + 间隔2列
            all_stats = self._create_multi_group_distribution(detail_df, score_column, groups,
                                                  all_title, current_row, right_chart_col,
                                                  filter_present=False)

            # === 1.3 在图表右侧添加统计信息表 ===
            if present_stats or all_stats:
                stats_start_col = right_chart_col + 7 + 2  # 右侧图表后面再空2列
                self._create_multi_group_stats_table(present_stats, all_stats,
                                                     current_row, stats_start_col, trait_name)

            # matplotlib图表占用约20行
            matplotlib_chart_rows = 20
            current_row += matplotlib_chart_rows + 3

            # === 2. 在图表下方创建表格 ===
            # 为每个组分别生成统计数据
            for group_name, filter_func, color in groups:
                # 为在群母牛生成统计
                present_stats = self._create_single_group_distribution(
                    detail_df, score_column, group_name, filter_func,
                    filter_present=True
                )

                # 为全部母牛生成统计
                all_stats = self._create_single_group_distribution(
                    detail_df, score_column, group_name, filter_func,
                    filter_present=False
                )

                # 生成横向并排的两个表
                if present_stats or all_stats:
                    rows_used = self._create_group_pair_tables(
                        group_name, present_stats, all_stats, trait_name, current_row
                    )
                    current_row += rows_used + 2  # 表格高度 + 间隔2行

            logger.info(f"✓ 已添加{trait_name}成母牛/后备牛分组分析（2个多曲线图 + 2组表格）")

            return current_row

        except Exception as e:
            logger.error(f"创建成母牛/后备牛分组图表失败: {e}", exc_info=True)
            return start_row

    def _create_birth_year_group_charts(self, detail_df, score_column, start_row):
        """
        按出生年份分组创建正态分布图

        新布局（参考示意表）：
        - 分类标题："不同出生年份牛群正态分布"
        - 2个多曲线matplotlib图（在群母牛 + 全部母牛），每个图包含n条曲线
        - 图表下方：n个组（n个年份），每个组包含2个横向并排的表（在群母牛 + 全部母牛）
        - 组之间垂直排列，组内表格横向并排

        Args:
            detail_df: 明细数据DataFrame
            score_column: 分数列名
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        try:
            from openpyxl.styles import Font
            import pandas as pd

            # 提取性状名称
            trait_name = score_column.replace('_score', '')

            # 分类标题
            title_cell = self.ws.cell(row=start_row, column=1, value="不同出生年份牛群正态分布")
            title_cell.font = Font(size=14, bold=True)
            start_row += 2

            # 动态识别出生年份
            if 'birth_year' not in detail_df.columns:
                logger.warning("缺少birth_year列，跳过出生年份分组图")
                return start_row

            years = sorted(detail_df['birth_year'].dropna().unique(), reverse=True)
            if len(years) == 0:
                logger.warning("没有出生年份数据")
                return start_row

            # 取最多6个年份，颜色循环
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
            years = years[:6]

            # 定义分组
            groups = []
            for idx, year in enumerate(years):
                year_int = int(year)
                groups.append((
                    f'{year_int}年出生',
                    lambda df, y=year_int: df['birth_year'] == y,
                    colors[idx % len(colors)]
                ))

            # === 1. 创建两个多曲线matplotlib图 ===
            current_row = start_row

            # 1.1 在群母牛多曲线图（左侧）
            present_title = f"在群母牛{trait_name}不同出生年份牛群正态分布"
            present_stats = self._create_multi_group_distribution(detail_df, score_column, groups,
                                                  present_title, current_row, 1,
                                                  filter_present=True)

            # 1.2 全部母牛多曲线图（右侧）
            all_title = f"全部母牛{trait_name}不同出生年份牛群正态分布"
            right_chart_col = 1 + 7 + 2  # 左侧图表(1) + 占用7列 + 间隔2列
            all_stats = self._create_multi_group_distribution(detail_df, score_column, groups,
                                                  all_title, current_row, right_chart_col,
                                                  filter_present=False)

            # === 1.3 在图表右侧添加统计信息表 ===
            if present_stats or all_stats:
                stats_start_col = right_chart_col + 7 + 2  # 右侧图表后面再空2列
                self._create_multi_group_stats_table(present_stats, all_stats,
                                                     current_row, stats_start_col, trait_name)

            # matplotlib图表占用约20行
            matplotlib_chart_rows = 20
            current_row += matplotlib_chart_rows + 3

            # === 2. 在图表下方创建表格 ===
            # 为每个组分别生成统计数据
            for group_name, filter_func, color in groups:
                # 为在群母牛生成统计
                present_stats = self._create_single_group_distribution(
                    detail_df, score_column, group_name, filter_func,
                    filter_present=True
                )

                # 为全部母牛生成统计
                all_stats = self._create_single_group_distribution(
                    detail_df, score_column, group_name, filter_func,
                    filter_present=False
                )

                # 生成横向并排的两个表
                if present_stats or all_stats:
                    rows_used = self._create_group_pair_tables(
                        group_name, present_stats, all_stats, trait_name, current_row
                    )
                    current_row += rows_used + 2  # 表格高度 + 间隔2行

            logger.info(f"✓ 已添加{trait_name}出生年份分组分析（2个多曲线图 + {len(groups)}组表格）")

            return current_row

        except Exception as e:
            logger.error(f"创建出生年份分组图表失败: {e}", exc_info=True)
            return start_row
    def _create_multi_group_stats_table(self, present_stats, all_stats, start_row, start_col, trait_name):
        """
        为多组正态分布图创建统计信息表格，显示在图的右侧
        
        Args:
            present_stats: 在群母牛统计数据列表
            all_stats: 全部母牛统计数据列表
            start_row: 起始行
            start_col: 起始列
            trait_name: 性状名称
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import pandas as pd

            current_row = start_row
            
            # 为在群母牛和全部母牛分别创建统计表
            for stats_list, title in [(present_stats, f"在群母牛{trait_name}统计"),
                                      (all_stats, f"全部母牛{trait_name}统计")]:
                if not stats_list:
                    continue
                
                # 标题
                title_cell = self.ws.cell(row=current_row, column=start_col, value=title)
                title_cell.font = Font(size=11, bold=True)
                current_row += 1
                
                # 表头
                headers = ['分组', '样本量', '均值', '标准差', '变异系数', '中位数', 'Q1-Q3']
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type='solid')
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                for col_idx, header_text in enumerate(headers):
                    cell = self.ws.cell(row=current_row, column=start_col + col_idx, value=header_text)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                current_row += 1
                
                # 数据行
                for stat in stats_list:
                    group_name = stat['group_name']
                    n = stat['total_count']
                    mu = stat['mean']
                    
                    # 从quintiles中提取统计数据
                    quintiles = stat.get('quintiles', [])
                    if quintiles:
                        # 从quintiles数据计算统计指标
                        df_temp = pd.DataFrame(quintiles)
                        if 'NM指数平均得分' in df_temp.columns:
                            score_col = 'NM指数平均得分'
                        elif 'TPI平均得分' in df_temp.columns:
                            score_col = 'TPI平均得分'
                        else:
                            score_col = df_temp.columns[-1]  # 使用最后一列作为分数列
                        
                        scores = df_temp[score_col].values
                        # 简单计算标准差（基于五个分位的平均分）
                        sigma = scores.std()
                        cv = (sigma / mu * 100) if mu != 0 else 0
                        median = scores[2] if len(scores) >= 3 else mu  # 中间分位作为中位数的近似
                        q1 = scores[0] if len(scores) >= 1 else mu
                        q3 = scores[4] if len(scores) >= 5 else mu
                    else:
                        sigma = 0
                        cv = 0
                        median = mu
                        q1 = mu
                        q3 = mu
                    
                    row_data = [
                        group_name,
                        n,
                        f"{mu:.1f}",
                        f"{sigma:.1f}",
                        f"{cv:.1f}%",
                        f"{median:.1f}",
                        f"{q1:.1f} - {q3:.1f}"
                    ]
                    
                    for col_idx, value in enumerate(row_data):
                        cell = self.ws.cell(row=current_row, column=start_col + col_idx, value=value)
                        if col_idx == 0:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    current_row += 1
                
                current_row += 2  # 两个表之间空2行

            # 添加指标说明备注
            current_row += 1
            note_cell = self.ws.cell(row=current_row, column=start_col, value="指标说明：")
            note_cell.font = Font(size=10, bold=True, color="666666")
            current_row += 1

            notes = [
                "· 样本量：参与统计的牛只数量",
                "· 均值：所有样本的平均值，反映整体水平",
                "· 标准差：数据离散程度，越大表示差异越大",
                "· 变异系数：标准差/均值×100%，用于比较不同指标的离散程度",
                "· 中位数：将数据从小到大排列后的中间值，不受极端值影响",
                "· Q1-Q3：第25%分位数到第75%分位数的范围，包含中间50%的数据",
                "",
                "分析建议：",
                "1. 变异系数<10%表示数据较为集中，>30%表示差异较大",
                "2. 对比均值和中位数：若接近则分布对称，差异大则数据有偏",
                "3. Q1-Q3范围越窄，说明核心群体水平越一致",
                "4. 对比不同分组的均值和标准差，了解各组间的差异"
            ]

            for note in notes:
                note_cell = self.ws.cell(row=current_row, column=start_col, value=note)
                note_cell.font = Font(size=9, color="666666")
                current_row += 1

            logger.info(f"✓ 已添加多组分布图统计表和指标说明")

        except Exception as e:
            logger.error(f"创建多组统计表失败: {e}", exc_info=True)

    def _create_single_chart_stats_table(self, present_stats, all_stats, start_row, start_col, trait_name):
        """
        为单图（整体正态分布）创建统计信息表格（转置格式：指标作为列，分组作为行）

        Args:
            present_stats: 在群母牛统计数据字典
            all_stats: 全部母牛统计数据字典
            start_row: 起始行
            start_col: 起始列
            trait_name: 性状名称
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            current_row = start_row

            # 定义样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # 标题
            title_cell = self.ws.cell(row=current_row, column=start_col, value=f"{trait_name}统计")
            title_cell.font = Font(size=11, bold=True)
            current_row += 1

            # 转置后的表头（列）：分组、样本量、均值、标准差、变异系数、中位数、Q1-Q3
            headers = ['分组', '样本量', '均值', '标准差', '变异系数', '中位数', 'Q1-Q3']

            for col_idx, header_text in enumerate(headers):
                cell = self.ws.cell(row=current_row, column=start_col + col_idx, value=header_text)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            current_row += 1

            # 数据行：在群母牛 和 全部母牛
            for stats_dict, group_name in [(present_stats, "在群母牛"), (all_stats, "全部母牛")]:
                if not stats_dict:
                    continue

                # 分组名称
                cell = self.ws.cell(row=current_row, column=start_col, value=group_name)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.font = Font(bold=True)

                # 样本量
                cell = self.ws.cell(row=current_row, column=start_col + 1,
                                  value=int(stats_dict['sample_size']))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # 均值
                cell = self.ws.cell(row=current_row, column=start_col + 2,
                                  value=round(stats_dict['mean'], 1))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # 标准差
                cell = self.ws.cell(row=current_row, column=start_col + 3,
                                  value=round(stats_dict['std'], 1))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # 变异系数
                cell = self.ws.cell(row=current_row, column=start_col + 4,
                                  value=f"{stats_dict['cv']:.1f}%")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # 中位数
                cell = self.ws.cell(row=current_row, column=start_col + 5,
                                  value=round(stats_dict['median'], 1))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                # Q1-Q3
                cell = self.ws.cell(row=current_row, column=start_col + 6,
                                  value=f"{stats_dict['q1']:.1f} - {stats_dict['q3']:.1f}")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

                current_row += 1

            # 添加指标说明备注
            current_row += 2
            note_cell = self.ws.cell(row=current_row, column=start_col, value="指标说明：")
            note_cell.font = Font(size=10, bold=True, color="666666")
            current_row += 1

            notes = [
                "· 样本量：参与统计的牛只数量",
                "· 均值：所有样本的平均值，反映整体水平",
                "· 标准差：数据离散程度，越大表示差异越大",
                "· 变异系数：标准差/均值×100%，用于比较不同指标的离散程度",
                "· 中位数：将数据从小到大排列后的中间值，不受极端值影响",
                "· Q1-Q3：第25%分位数到第75%分位数的范围，包含中间50%的数据",
                "",
                "分析建议：",
                "1. 变异系数<10%表示数据较为集中，>30%表示差异较大",
                "2. 对比均值和中位数：若接近则分布对称，差异大则数据有偏",
                "3. Q1-Q3范围越窄，说明核心群体水平越一致"
            ]

            for note in notes:
                note_cell = self.ws.cell(row=current_row, column=start_col, value=note)
                note_cell.font = Font(size=9, color="666666")
                current_row += 1

            logger.info(f"✓ 已添加单图统计表（转置格式）和指标说明")

        except Exception as e:
            logger.error(f"创建单图统计表失败: {e}", exc_info=True)

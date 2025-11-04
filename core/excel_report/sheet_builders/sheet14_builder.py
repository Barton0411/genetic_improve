"""
Sheet 14构建器: 备选公牛-隐性基因及近交系数明细表
直接展示完整的分析结果数据
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)


class Sheet14Builder(BaseSheetBuilder):
    """
    Sheet 14: 备选公牛-隐性基因及近交系数明细表

    包含内容:
    - 完整的备选公牛分析结果数据表
    - 包含所有母牛×公牛配对的详细信息
    """

    def build(self, data: dict):
        """
        构建Sheet 14: 备选公牛-隐性基因及近交系数明细表

        Args:
            data: 包含备选公牛明细数据
                - file_path: 源文件路径
                - data: DataFrame数据
        """
        try:
            # 检查数据
            if not data or 'data' not in data:
                logger.warning("Sheet14: 缺少数据，跳过生成")
                return

            df = data['data']
            if df is None or len(df) == 0:
                logger.warning("Sheet14: 数据为空")
                return

            # 创建Sheet
            self._create_sheet("备选公牛-明细表")
            logger.info(f"构建Sheet 14: 备选公牛-明细表（{len(df)}行 × {len(df.columns)}列）")

            # 定义颜色
            self.color_header = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type='solid')  # 深蓝

            # 写入数据（使用批量写入优化性能）
            logger.info("开始写入数据...")

            # 使用append()批量写入整行（比逐个单元格写入快100倍！）
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                self.ws.append(row)

                # 每5000行输出一次进度
                if r_idx % 5000 == 0:
                    logger.info(f"已写入 {r_idx}/{len(df)+1} 行...")

            logger.info(f"数据写入完成，共 {len(df)+1} 行（含表头）")

            # 只对表头行设置格式
            logger.info("设置表头格式...")
            for col_idx in range(1, len(df.columns) + 1):
                cell = self.ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = self.color_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 设置表头行高
            self.ws.row_dimensions[1].height = 35

            # 设置列宽（根据列名自动调整）
            self._auto_adjust_column_widths(df)

            # 冻结首行
            self._freeze_panes('A2')

            # 添加筛选器
            self.ws.auto_filter.ref = self.ws.dimensions

            logger.info(f"✓ Sheet 14构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 14失败: {e}", exc_info=True)
            raise

    def _auto_adjust_column_widths(self, df):
        """
        根据列名和内容自动调整列宽

        Args:
            df: DataFrame数据
        """
        for col_idx, column in enumerate(df.columns, 1):
            column_name = str(column)

            # 设置基础宽度
            if '母牛号' in column_name or '父号' in column_name or '公牛号' in column_name:
                width = 15
            elif '近交' in column_name:
                width = 12
            elif '详情' in column_name:
                width = 30
            elif '原始' in column_name:
                width = 15
            elif column_name in ['是否在场', 'sex', 'lac']:
                width = 10
            elif '(' in column_name:  # 基因携带状态列
                width = 12
            else:
                # 根据列名长度设置
                width = max(len(column_name) * 1.5, 10)

            # 限制最大宽度
            width = min(width, 50)

            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[col_letter].width = width

    def _get_thin_border(self):
        """获取细边框样式"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

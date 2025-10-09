"""
Sheet 1A构建器: 牧场牛群原始数据
直接复制原始Excel文件内容
"""

from .base_builder import BaseSheetBuilder
from openpyxl import load_workbook
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class Sheet1ABuilder(BaseSheetBuilder):
    """Sheet 1A: 牧场牛群原始数据（直接复制原始Excel）"""

    def build(self, data: dict):
        """
        构建Sheet 1A - 直接复制原始Excel文件的内容

        Args:
            data: {
                'raw_file_path': Path  # 原始Excel文件路径
            }
        """
        try:
            logger.info("构建Sheet 1A: 牧场牛群原始数据")

            raw_file_path = data.get('raw_file_path')
            if not raw_file_path or not Path(raw_file_path).exists():
                logger.warning(f"原始母牛数据文件不存在: {raw_file_path}")
                # 创建空Sheet
                self._create_sheet("牧场牛群原始数据")
                self.ws.cell(row=1, column=1, value="暂无原始数据")
                return

            # 加载原始Excel文件
            source_wb = load_workbook(raw_file_path)
            source_ws = source_wb.active  # 获取第一个sheet

            # 在目标workbook中创建新sheet
            target_ws = self.wb.create_sheet("牧场牛群原始数据")

            # 复制所有单元格的值和格式
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws[cell.coordinate]
                    # 复制值
                    target_cell.value = cell.value
                    # 复制格式
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()

            # 复制列宽
            for col_letter in source_ws.column_dimensions:
                if col_letter in source_ws.column_dimensions:
                    target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

            # 复制行高
            for row_num in source_ws.row_dimensions:
                if row_num in source_ws.row_dimensions:
                    target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height

            # 复制合并单元格
            for merged_cell_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_cell_range))

            # 冻结窗格
            if source_ws.freeze_panes:
                target_ws.freeze_panes = source_ws.freeze_panes

            logger.info(f"✓ Sheet 1A构建完成，已复制原始Excel文件")

        except Exception as e:
            logger.error(f"构建Sheet 1A失败: {e}", exc_info=True)
            raise

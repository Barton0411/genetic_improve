"""
Sheet 8构建器: 已用公牛性状汇总分析
v1.2.2版本 - 动态性状+动态年份+折线图+散点图

包含内容:
1. 按年份汇总表（使用公牛数、配种头次、各性状平均值）
2. 性状进展折线图（按配置分组生成多个图表）
3. 性状进展数据表（逐年增长、年均增长、累计增长）
4. 散点图：全部配种记录时间线
5. 散点图:近12个月配种记录时间线
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, ScatterChart, BarChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image as OpenpyxlImage
import pandas as pd
import logging
import matplotlib
matplotlib.use('Agg')  # 使用非GUI后端
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO
from datetime import datetime

logger = logging.getLogger(__name__)


class Sheet8Builder(BaseSheetBuilder):
    """
    Sheet 8: 已用公牛性状汇总分析

    包含内容:
    1. 按年份汇总表（使用公牛数、配种头次、各性状平均值）
    2. 性状进展折线图（按配置分组生成多个图表）
    3. 性状进展数据表（逐年增长、年均增长、累计增长）
    4. 散点图：全部配种记录时间线
    5. 散点图：近12个月配种记录时间线
    """

    def build(self, data: dict):
        """
        构建Sheet 8: 已用公牛性状汇总分析

        Args:
            data: 包含以下键的字典:
                - summary_data: 汇总表数据 (DataFrame)
                - progress_data: 性状进展数据 (DataFrame)
                - trait_columns: 性状列列表
                - year_range: 年份范围列表
                - scatter_data_all: 全部配种记录散点图数据
                - scatter_data_12m: 近12个月配种记录散点图数据
                - chart_groups: 折线图分组配置列表
        """
        try:
            # 检查数据
            if not data or 'summary_data' not in data:
                logger.warning("Sheet8: 缺少数据，跳过生成")
                return

            # 创建Sheet
            self._create_sheet("已用公牛性状汇总")
            logger.info("构建Sheet 8: 已用公牛性状汇总分析")

            current_row = 1

            # 1. 构建汇总表
            summary_data = data.get('summary_data')
            if summary_data is not None and not summary_data.empty:
                logger.info(f"  构建汇总表: {len(summary_data)} 行")
                summary_start_row = current_row
                current_row = self._build_summary_table(summary_data, current_row)
                current_row += 3  # 空3行

            # 2. 构建性状进展折线图（多个图表）
            chart_groups = data.get('chart_groups', [])
            year_range = data.get('year_range', [])
            if chart_groups and year_range and summary_data is not None:
                logger.info(f"  构建折线图: {len(chart_groups)} 个图表组")
                current_row = self._build_trait_progress_charts(
                    summary_data, chart_groups, year_range, summary_start_row, current_row
                )
                current_row += 3  # 空3行

            # 3. 构建性状进展数据表
            progress_data = data.get('progress_data')
            if progress_data is not None and not progress_data.empty:
                logger.info(f"  构建进展数据表: {len(progress_data)} 个性状")
                current_row = self._build_progress_table(progress_data, current_row)
                current_row += 3  # 空3行

            # 4. 构建两个散点图：并排显示在同一行
            scatter_data_all = data.get('scatter_data_all')
            scatter_data_12m = data.get('scatter_data_12m')

            if (scatter_data_all is not None and not scatter_data_all.empty) or \
               (scatter_data_12m is not None and not scatter_data_12m.empty):
                logger.info(f"  构建散点图（并排布局）")
                current_row = self._build_scatter_charts_side_by_side(
                    scatter_data_all,
                    scatter_data_12m,
                    current_row
                )

            # 设置列宽（根据实际列数动态调整）
            self._set_default_column_widths(summary_data)

            # 冻结首行
            self._freeze_panes('A2')

            logger.info("✓ Sheet 8构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 8失败: {e}", exc_info=True)
            raise

    def _build_summary_table(self, summary_df: pd.DataFrame, start_row: int) -> int:
        """
        构建按年份汇总表

        Args:
            summary_df: 汇总数据DataFrame
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        title = "已用公牛性状汇总表（按年份）"
        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头行
        headers = summary_df.columns.tolist()
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据行
        for _, row_data in summary_df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=current_row, column=col_idx, value=value)

                # 年份列和总平均行使用不同样式
                if col_idx == 1:  # 年份列
                    self.style_manager.apply_data_style(cell, alignment='center')
                    if value == '总平均':
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
                else:  # 数值列
                    # 判断是否为数值
                    if isinstance(value, (int, float)) and pd.notna(value):
                        # 根据列名判断格式
                        if '使用公牛数' in headers[col_idx-1] or '配种头次' in headers[col_idx-1]:
                            cell.value = int(value)
                            self.style_manager.apply_data_style(cell, alignment='center')
                        else:
                            # 性状平均值保留2位小数
                            self.style_manager.apply_data_style(cell, alignment='center')
                            cell.number_format = '0.00'
                    else:
                        self.style_manager.apply_data_style(cell, alignment='center')

                    # 总平均行加粗
                    if row_data.iloc[0] == '总平均':
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')

            current_row += 1

        return current_row

    def _build_trait_progress_charts(
        self,
        summary_df: pd.DataFrame,
        chart_groups: list,
        year_range: list,
        summary_start_row: int,
        start_row: int
    ) -> int:
        """
        构建性状进展折线图（多个图表）

        Args:
            summary_df: 汇总数据DataFrame
            chart_groups: 图表分组配置列表
            year_range: 年份范围列表
            summary_start_row: 汇总表开始行号（用于图表数据引用）
            start_row: 折线图区域起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题
        title = "性状进展折线图"
        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 汇总表数据位置
        # 表头在summary_start_row + 1, 数据从summary_start_row + 2开始
        header_row = summary_start_row + 1
        data_start_row = summary_start_row + 2
        data_end_row = data_start_row + len(year_range) - 1  # 不包含总平均行

        logger.info(f"    汇总表位置: header={header_row}, data={data_start_row}-{data_end_row}")

        # 为每个图表组生成一个折线图
        chart_col = 1
        chart_row = current_row
        charts_per_row = 2  # 每行放2个图表
        chart_count = 0

        for group in chart_groups:
            group_name = group['name']
            traits = group['traits']
            y_axis_label = group.get('y_axis_label', '性状值')
            y_range = group.get('range', (None, None))
            invert = group.get('invert', False)

            logger.info(f"    生成折线图: {group_name} ({len(traits)}个性状)")

            # 1. 先计算该组性状的实际数据范围（动态Y轴）
            all_values = []
            for trait in traits:
                col_name = f'平均{trait}'
                if col_name in summary_df.columns:
                    # 获取该性状的所有年份数据（排除总平均行）
                    yearly_data = summary_df[summary_df['年份'] != '总平均'][col_name]
                    valid_values = yearly_data.dropna().tolist()
                    all_values.extend(valid_values)

            # 计算动态Y轴范围
            if all_values:
                data_min = min(all_values)
                data_max = max(all_values)
                # 添加10%的边距，让数据线不贴边
                margin = (data_max - data_min) * 0.1
                y_min = data_min - margin
                y_max = data_max + margin
                logger.info(f"      动态Y轴范围: {y_min:.2f} 到 {y_max:.2f} (数据范围: {data_min:.2f} - {data_max:.2f})")
            else:
                # 如果没有有效数据，使用配置的范围或默认范围
                y_min = y_range[0] if y_range[0] is not None else None
                y_max = y_range[1] if y_range[1] is not None else None
                logger.warning(f"      无有效数据，使用配置Y轴范围: {y_min} - {y_max}")

            # 2. 创建折线图
            chart = LineChart()
            chart.title = group_name
            chart.style = 10
            chart.y_axis.title = y_axis_label
            chart.x_axis.title = "年份"
            chart.width = 16
            chart.height = 10

            # 设置动态Y轴范围
            if y_min is not None:
                chart.y_axis.scaling.min = y_min
            if y_max is not None:
                chart.y_axis.scaling.max = y_max

            # 3. 为每个性状添加一条折线
            series_added = 0
            for trait in traits:
                col_name = f'平均{trait}'
                if col_name not in summary_df.columns:
                    logger.warning(f"      跳过性状 {trait}: 列 '{col_name}' 不存在于汇总表中")
                    continue

                # 找到该列在summary_df中的索引
                col_idx = summary_df.columns.tolist().index(col_name) + 1

                logger.info(f"      添加系列 {trait}: 列索引={col_idx}, 数据范围={data_start_row}-{data_end_row}")

                # 数据引用（Y轴）- 只包含数据行
                yvalues = Reference(
                    self.ws,
                    min_col=col_idx,
                    min_row=data_start_row,  # 只包含数据行
                    max_row=data_end_row
                )

                # 创建数据系列，手动指定标题
                series = Series(yvalues, title=trait)
                chart.series.append(series)
                series_added += 1

            # 设置X轴类别（年份）
            year_col_idx = 1
            xvalues = Reference(
                self.ws,
                min_col=year_col_idx,
                min_row=data_start_row,  # 只包含数据行
                max_row=data_end_row
            )
            chart.set_categories(xvalues)

            if series_added == 0:
                logger.warning(f"      图表 {group_name}: 没有添加任何数据系列！")
            else:
                logger.info(f"      图表 {group_name}: 成功添加 {series_added} 个数据系列，X轴范围={data_start_row}-{data_end_row}")

            # 计算图表位置
            chart_col_letter = get_column_letter(chart_col)
            chart.anchor = f"{chart_col_letter}{chart_row}"
            self.ws.add_chart(chart)

            # 更新图表位置
            chart_count += 1
            if chart_count % charts_per_row == 0:
                # 换行
                chart_row += 22  # 图表高度约22行（增加间距避免重叠）
                chart_col = 1
            else:
                # 横向移动
                chart_col += 9  # 图表宽度约9列

        # 返回图表区域之后的行号
        if chart_count > 0:
            rows_used = ((chart_count - 1) // charts_per_row + 1) * 22
            current_row += rows_used

        return current_row

    def _build_progress_table(self, progress_df: pd.DataFrame, start_row: int) -> int:
        """
        构建性状进展数据表

        Args:
            progress_df: 性状进展数据DataFrame
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        title = "性状进展数据表（逐年增长分析）"
        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头行
        headers = progress_df.columns.tolist()
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据行
        for _, row_data in progress_df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=current_row, column=col_idx, value=value)

                # 性状名称列左对齐
                if col_idx == 1:
                    self.style_manager.apply_data_style(cell, alignment='left')
                else:
                    # 数值列
                    if isinstance(value, (int, float)) and pd.notna(value):
                        self.style_manager.apply_data_style(cell, alignment='center')

                        # 判断列名，增长率列使用百分比格式
                        col_name = headers[col_idx - 1]
                        if col_name == '增长率':
                            # 增长率：显示为百分比（数据已经乘以100，所以用0.00%格式）
                            cell.number_format = '0.00"%"'
                        else:
                            # 其他数值列：保留3位小数
                            cell.number_format = '0.000'
                    else:
                        self.style_manager.apply_data_style(cell, alignment='center')

            current_row += 1

        # 在表格右侧添加计算公式说明
        current_row = self._add_progress_formula_notes(progress_df, start_row, current_row)

        return current_row

    def _add_progress_formula_notes(self, progress_df: pd.DataFrame, table_start_row: int, current_row: int) -> int:
        """
        在性状进展表右侧添加计算公式说明

        Args:
            progress_df: 性状进展数据DataFrame
            table_start_row: 表格起始行（标题行）
            current_row: 当前行号

        Returns:
            下一个可用行号
        """
        # 计算说明放在表格右侧（表格列数+2列的位置）
        note_col = len(progress_df.columns) + 3

        # 说明框标题
        note_start_row = table_start_row + 1  # 从表头行开始
        title_cell = self.ws.cell(row=note_start_row, column=note_col, value="📊 计算公式说明")
        title_cell.font = Font(size=12, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        note_start_row += 1

        # 公式说明内容（注意：不能以"="开头，否则Excel会当作公式）
        formulas = [
            ("逐年增长", "后一年值 - 前一年值", "例：2025年NM$ - 2024年NM$"),
            ("", "", ""),
            ("年均增长", "所有逐年增长的平均值", "例：(2024→2025 + 2025→2026) ÷ 2"),
            ("", "", ""),
            ("N年累计", "最后一年值 - 第一年值", "例：2025年NM$ - 2023年NM$ = 206.347"),
            ("", "", ""),
            ("增长率", "(N年累计 ÷ 第一年绝对值) × 100%", "例：(206.347 ÷ 99.345) × 100% = 207.71%"),
        ]

        for idx, (term, formula, example) in enumerate(formulas):
            row = note_start_row + idx

            # 术语列
            if term:
                term_cell = self.ws.cell(row=row, column=note_col, value=term)
                term_cell.font = Font(size=10, bold=True, color="2E5090")
                term_cell.alignment = Alignment(horizontal='left', vertical='center')

            # 公式列
            if formula:
                formula_cell = self.ws.cell(row=row, column=note_col + 1, value=formula)
                formula_cell.font = Font(size=10, color="404040")
                formula_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 示例列
            if example:
                example_cell = self.ws.cell(row=row, column=note_col + 2, value=example)
                example_cell.font = Font(size=9, italic=True, color="666666")
                example_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 设置说明列的列宽
        self.ws.column_dimensions[get_column_letter(note_col)].width = 12
        self.ws.column_dimensions[get_column_letter(note_col + 1)].width = 28
        self.ws.column_dimensions[get_column_letter(note_col + 2)].width = 35

        # 添加边框（可选，让说明框更明显）
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        for row_idx in range(note_start_row, note_start_row + len(formulas)):
            for col_offset in range(3):
                cell = self.ws.cell(row=row_idx, column=note_col + col_offset)
                cell.border = thin_border

                # 添加浅色背景
                if row_idx == note_start_row:
                    # 标题行背景
                    cell.fill = PatternFill(start_color="E7EFF8", end_color="E7EFF8", fill_type='solid')
                elif formulas[row_idx - note_start_row][0]:  # 非空行
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type='solid')

        return current_row

    def _build_scatter_chart(self, scatter_df: pd.DataFrame, chart_title: str, start_row: int) -> int:
        """
        构建散点图（配种记录时间线）

        Args:
            scatter_df: 散点图数据DataFrame（配种日期、冻精编号、配种类型）
            chart_title: 图表标题
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题
        title_cell = self.ws.cell(row=current_row, column=1, value=chart_title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 将散点图数据写入Sheet（用于图表引用）
        # 表头
        scatter_headers = ['配种日期', '冻精编号', '配种类型']
        for col_idx, header in enumerate(scatter_headers, start=1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
        data_start_row = current_row + 1
        current_row += 1

        # 数据行（最多显示1000条，避免数据过多）
        max_rows = min(1000, len(scatter_df))
        for idx in range(max_rows):
            row_data = scatter_df.iloc[idx]
            date_val = row_data.get('配种日期')
            bull_val = row_data.get('冻精编号')
            type_val = row_data.get('配种类型')

            self.ws.cell(row=current_row, column=1, value=date_val)
            self.ws.cell(row=current_row, column=2, value=bull_val)
            self.ws.cell(row=current_row, column=3, value=type_val)

            # 应用样式
            for col_idx in range(1, 4):
                cell = self.ws.cell(row=current_row, column=col_idx)
                self.style_manager.apply_data_style(cell, alignment='center')

            current_row += 1

        data_end_row = current_row - 1

        # 创建散点图
        # 注意：散点图需要数值型X轴，这里使用行号作为X轴（代表时间顺序）
        chart = ScatterChart()
        chart.title = chart_title
        chart.style = 13
        chart.x_axis.title = "时间顺序"
        chart.y_axis.title = "公牛编号"
        chart.width = 20
        chart.height = 12

        # 由于散点图需要按配种类型分组显示（性控、常规、其他）
        # 这里暂时创建一个简单的散点图，实际可按类型分系列

        # X轴：行号（时间顺序）
        xvalues = Reference(self.ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        # Y轴：公牛编号（这里用数值型列代替，实际需要将公牛编号映射为数值）
        # 暂时不添加系列，只显示数据表

        # 图表位置（放在数据表右侧）
        chart.anchor = f"E{start_row}"
        self.ws.add_chart(chart)

        # 添加说明
        note_row = data_end_row + 2
        note = f"注：由于数据量较大，仅显示前{max_rows}条记录" if len(scatter_df) > max_rows else f"共{len(scatter_df)}条配种记录"
        note_cell = self.ws.cell(row=note_row, column=1, value=note)
        note_cell.font = Font(size=10, italic=True, color="666666")

        return note_row + 1

    def _set_default_column_widths(self, summary_df):
        """设置默认列宽"""
        col_widths = {
            1: 12,  # 年份
            2: 14,  # 使用公牛数
            3: 14,  # 配种头次
        }

        # 性状列默认宽度
        if summary_df is not None:
            for col_idx in range(4, len(summary_df.columns) + 1):
                col_widths[col_idx] = 15

        self._set_column_widths(col_widths)

    def _build_scatter_charts_side_by_side(
        self,
        scatter_data_all: pd.DataFrame,
        scatter_data_12m: pd.DataFrame,
        start_row: int
    ) -> int:
        """
        构建两个散点图，上下排列显示（避免重叠）

        Args:
            scatter_data_all: 全部配种记录数据
            scatter_data_12m: 近12个月配种记录数据
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        title_cell = self.ws.cell(row=current_row, column=1, value="配种记录时间线")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 上方：全部记录散点图（列A开始）
        if scatter_data_all is not None and not scatter_data_all.empty:
            logger.info(f"    上方：全部记录 ({len(scatter_data_all)} 条)")
            current_row = self._build_single_scatter_chart(
                scatter_data_all,
                "全部记录",
                current_row,
                start_col=1  # A列开始
            )
            # 图表间距
            current_row += 3

        # 下方：近1年散点图（列A开始）
        if scatter_data_12m is not None and not scatter_data_12m.empty:
            logger.info(f"    下方：近1年 ({len(scatter_data_12m)} 条)")
            current_row = self._build_single_scatter_chart(
                scatter_data_12m,
                "近1年",
                current_row,
                start_col=1  # A列开始
            )
            # 最后图表后留间距
            current_row += 3

        return current_row

    def _build_single_scatter_chart(
        self,
        scatter_df: pd.DataFrame,
        chart_subtitle: str,
        start_row: int,
        start_col: int
    ) -> int:
        """
        构建单个配种记录分布散点图（单值图）- 使用matplotlib生成

        X轴：配种冻精（公牛编号）
        Y轴：配种日期
        每个点：一次配种记录
        按配种类型分系列显示

        Args:
            scatter_df: 配种记录数据DataFrame
            chart_subtitle: 图表子标题
            start_row: 起始行号
            start_col: 起始列号（1-based）

        Returns:
            该图表使用的最后一行行号
        """
        current_row = start_row

        # 小标题
        subtitle_cell = self.ws.cell(row=current_row, column=start_col, value=f"【{chart_subtitle}】")
        subtitle_cell.font = Font(size=12, bold=True)
        subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        if scatter_df is not None and not scatter_df.empty:
            # 确保配种日期是datetime类型
            scatter_df = scatter_df.copy()
            scatter_df['配种日期'] = pd.to_datetime(scatter_df['配种日期'])

            # 统一冻精类型名称（兼容不同写法）
            type_mapping = {
                '超级性控': '超级性控',
                '性控': '性控冻精',
                '性控冻精': '性控冻精',
                '普通': '普通冻精',
                '普通冻精': '普通冻精',
                '常规': '普通冻精',
                '其他': '其他'
            }
            scatter_df['冻精类型标准'] = scatter_df['配种类型'].map(lambda x: type_mapping.get(x, '其他'))

            # 按优先级排序冻精类型
            type_order = ['超级性控', '性控冻精', '普通冻精', '其他']
            types = [t for t in type_order if t in scatter_df['冻精类型标准'].unique()]

            # 统计每头公牛的使用次数
            bull_usage = scatter_df.groupby('冻精编号').size().to_dict()

            # 按冻精类型分组公牛，并按使用频率排序
            type_bulls = {}
            type_positions = {}
            current_x = 0

            for breeding_type in type_order:
                # 获取该类型的所有公牛
                type_df = scatter_df[scatter_df['冻精类型标准'] == breeding_type]
                if len(type_df) == 0:
                    continue

                # 统计每头公牛的使用次数并排序
                bulls_in_type = type_df['冻精编号'].unique()
                bulls_with_count = [(bull, bull_usage[bull]) for bull in bulls_in_type]
                bulls_with_count.sort(key=lambda x: x[1], reverse=True)  # 按使用次数降序

                # 为每头公牛分配X轴位置（使用次数多的占据更宽的空间）
                type_positions[breeding_type] = {}
                for bull, count in bulls_with_count:
                    # 使用次数越多，宽度越大（但有上限）
                    width = min(5, max(0.5, count / 10))  # 宽度范围：0.5-5
                    type_positions[breeding_type][bull] = (current_x + width / 2, width)
                    current_x += width

                type_bulls[breeding_type] = [bull for bull, _ in bulls_with_count]
                current_x += 2  # 类型之间留间隔

            total_width = current_x
            bull_count = len(scatter_df['冻精编号'].unique())

            # 为每条记录分配X位置（在该公牛的宽度范围内随机分布）
            import numpy as np
            np.random.seed(42)  # 固定随机种子，确保结果可重现

            def get_x_position_with_jitter(row):
                breeding_type = row['冻精类型标准']
                bull = row['冻精编号']
                if breeding_type in type_positions and bull in type_positions[breeding_type]:
                    center_x, width = type_positions[breeding_type][bull]
                    # 在该公牛的宽度范围内随机分布（±width/2）
                    jitter = np.random.uniform(-width/2, width/2)
                    return center_x + jitter
                return 0

            scatter_df['X位置'] = scatter_df.apply(get_x_position_with_jitter, axis=1)

            # 使用matplotlib生成散点图
            try:
                # 设置中文字体
                plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
                plt.rcParams['axes.unicode_minus'] = False

                # 计算图表大小（根据总宽度动态调整）
                fig_width = min(20, max(10, total_width * 0.3))
                fig_height = 6

                fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=150)  # 提高DPI到150

                # 冻精类型颜色映射
                type_colors = {
                    '超级性控': '#2E86AB',  # 蓝色
                    '性控冻精': '#06A77D',  # 绿色
                    '普通冻精': '#F77F00',  # 橙色
                    '其他': '#999999'      # 灰色
                }

                # 冻精类型背景色（更浅的版本）
                type_bg_colors = {
                    '超级性控': '#E3F2FD',  # 浅蓝色
                    '性控冻精': '#E8F5E9',  # 浅绿色
                    '普通冻精': '#FFF3E0',  # 浅橙色
                    '其他': '#F5F5F5'      # 浅灰色
                }

                # 添加背景色分区
                current_x_bg = 0
                type_x_ranges = {}  # 保存每个类型的X轴范围
                for breeding_type in type_order:
                    if breeding_type not in type_bulls or len(type_bulls[breeding_type]) == 0:
                        continue

                    # 计算该类型的X轴范围
                    type_width = sum(type_positions[breeding_type][bull][1] for bull in type_bulls[breeding_type])

                    # 绘制背景色
                    ax.axvspan(
                        current_x_bg,
                        current_x_bg + type_width,
                        alpha=0.2,
                        color=type_colors[breeding_type],
                        zorder=0
                    )

                    # 保存X轴范围（稍后添加标签）
                    type_x_ranges[breeding_type] = (current_x_bg, current_x_bg + type_width)
                    current_x_bg += type_width + 2

                # 绘制每头公牛的"柱子"边界线
                for breeding_type in type_order:
                    if breeding_type not in type_bulls:
                        continue

                    for bull in type_bulls[breeding_type]:
                        center_x, width = type_positions[breeding_type][bull]
                        left_edge = center_x - width / 2
                        right_edge = center_x + width / 2

                        # 绘制左右边界虚线
                        ax.axvline(left_edge, color=type_colors[breeding_type],
                                   linestyle=':', alpha=0.3, linewidth=0.5, zorder=1)
                        ax.axvline(right_edge, color=type_colors[breeding_type],
                                   linestyle=':', alpha=0.3, linewidth=0.5, zorder=1)

                # 按类型绘制散点
                for breeding_type in types:
                    type_df = scatter_df[scatter_df['冻精类型标准'] == breeding_type]
                    color = type_colors.get(breeding_type, '#999999')

                    # 为每条记录计算点的大小（根据该公牛的使用频率）
                    sizes = []
                    for _, row in type_df.iterrows():
                        bull = row['冻精编号']
                        count = bull_usage[bull]
                        # 使用次数多的点稍大
                        size = min(50, max(20, count / 5))
                        sizes.append(size)

                    ax.scatter(
                        type_df['X位置'],
                        type_df['配种日期'],
                        c=color,
                        label=breeding_type,
                        alpha=0.7,  # 稍微提高不透明度
                        s=sizes,  # 动态点大小
                        edgecolors='white',
                        linewidths=0.5,
                        zorder=2
                    )

                # 设置X轴
                ax.set_xlabel('配种冻精（按类型分组，宽度表示使用频率）', fontsize=10, fontweight='bold')

                # 为每个公牛添加X轴刻度和标签
                tick_positions = []
                tick_labels = []

                for breeding_type in type_order:
                    if breeding_type not in type_bulls or len(type_bulls[breeding_type]) == 0:
                        continue

                    # 为该类型的每头公牛添加标签
                    for bull in type_bulls[breeding_type]:
                        x_pos, width = type_positions[breeding_type][bull]
                        tick_positions.append(x_pos)
                        # 显示完整公牛编号
                        tick_labels.append(bull)

                ax.set_xticks(tick_positions)
                ax.set_xticklabels(tick_labels, rotation=90, ha='center', fontsize=7)
                ax.set_xlim(-1, total_width)

                # 设置Y轴：配种日期
                ax.set_ylabel('配种日期', fontsize=10, fontweight='bold')
                ax.yaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
                ax.yaxis.set_major_locator(mdates.AutoDateLocator())

                # 设置标题
                ax.set_title(f'配种日期/冻精 单值图 - {chart_subtitle}', fontsize=12, fontweight='bold', pad=20)

                # 在X轴顶部添加类型标签
                y_max = ax.get_ylim()[1]
                for breeding_type, (x_start, x_end) in type_x_ranges.items():
                    ax.text(
                        (x_start + x_end) / 2,
                        y_max,
                        breeding_type,
                        horizontalalignment='center',
                        verticalalignment='bottom',
                        fontsize=9,
                        fontweight='bold',
                        color=type_colors[breeding_type],
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor=type_colors[breeding_type], alpha=0.8)
                    )

                # 计算统计信息
                total_count = len(scatter_df)
                date_min = scatter_df['配种日期'].min()
                date_max = scatter_df['配种日期'].max()
                date_range_str = f"{date_min.strftime('%Y/%m/%d')} 至 {date_max.strftime('%Y/%m/%d')}"

                # 在图表上添加统计信息文本框（左上角）
                stats_text = f'时间范围: {date_range_str}\n总配种头次: {total_count}\n使用公牛数: {bull_count}'
                ax.text(
                    0.02, 0.98,  # 位置：左上角
                    stats_text,
                    transform=ax.transAxes,
                    fontsize=8,
                    verticalalignment='top',
                    bbox=dict(boxstyle='round', facecolor='white', alpha=0.9, edgecolor='#CCCCCC', linewidth=1)
                )

                # 添加图例
                ax.legend(loc='upper right', fontsize=9, framealpha=0.9)

                # 添加网格
                ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)

                # 调整日期标签
                fig.autofmt_xdate()

                # 调整布局
                plt.tight_layout()

                # 保存到内存
                img_buffer = BytesIO()
                plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')  # 提高DPI到150
                plt.close(fig)
                img_buffer.seek(0)

                # 将图片嵌入Excel
                img = OpenpyxlImage(img_buffer)
                chart_col_letter = get_column_letter(start_col)
                img.anchor = f"{chart_col_letter}{current_row + 1}"
                self.ws.add_image(img)

                logger.info(f"      {chart_subtitle}: {total_count}配次, {bull_count}头公牛, {len(types)}种类型, matplotlib散点图")

                # 返回位置：图片大约占用45行（6英寸高度 + tight bbox增加的边距）
                # Excel默认行高约15像素，图片900像素高度 + 边距 ≈ 1000像素 ≈ 45行
                return current_row + 45

            except Exception as e:
                logger.error(f"生成matplotlib散点图失败: {e}", exc_info=True)
                # 如果matplotlib失败，显示错误信息
                error_cell = self.ws.cell(row=current_row + 1, column=start_col, value=f"图表生成失败: {str(e)}")
                error_cell.font = Font(size=10, italic=True, color="FF0000")
                return current_row + 2

        else:
            # 无数据
            no_data_cell = self.ws.cell(row=current_row + 1, column=start_col, value="暂无数据")
            no_data_cell.font = Font(size=10, italic=True, color="999999")
            return current_row + 2

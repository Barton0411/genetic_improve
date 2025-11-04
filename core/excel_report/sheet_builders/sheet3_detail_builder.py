"""
Sheet 3-4构建器: 育种性状明细表
"""

from .base_builder import BaseSheetBuilder
from openpyxl.utils import get_column_letter
import pandas as pd
import logging

logger = logging.getLogger(__name__)


class Sheet3DetailBuilder(BaseSheetBuilder):
    """Sheet 3-4: 育种性状明细表"""

    # 基础列名中英文映射
    COLUMN_MAPPING = {
        'cow_id': '耳号',
        'breed': '品种',
        'sex': '性别',
        'sire': '父号',
        'dam': '母号',
        'mgs': '外祖父号',
        'mgd': '外祖母号',
        'mmgs': '外外祖父号',
        'lac': '胎次',
        'calving_date': '产犊日期',
        'birth_date': '出生日期',
        'birth_date_dam': '母牛出生日期',
        'birth_date_mgd': '外祖母出生日期',
        'age': '月龄',
        'services_time': '配种次数',
        'DIM': '泌乳天数',
        'peak_milk': '峰值奶量',
        'milk_305': '305天奶量',
        'repro_status': '繁殖状态',
        'group': '组别',
        '是否在场': '是否在场',
        'birth_year': '出生年份',
    }

    @staticmethod
    def _get_trait_name(score_column: str) -> str:
        """
        从_score列名获取性状名称
        例如: 'NM$_score' -> 'NM$'
        """
        if score_column.endswith('_score'):
            return score_column[:-6]  # 去掉'_score'后缀
        return score_column

    def build(self, data: dict):
        """
        构建Sheet 3-4

        Args:
            data: {
                'detail_df': 育种性状明细DataFrame
            }
        """
        try:
            # 创建Sheet
            self._create_sheet("育种性状明细")
            logger.info("构建Sheet 3-4: 育种性状明细")

            df = data.get('detail_df')

            if df is None or df.empty:
                logger.warning("Sheet 3-4: 育种性状明细数据为空，跳过构建")
                return

            # 定义基础列（固定顺序）
            base_columns = [
                'cow_id', 'breed', 'sex', 'sire', 'dam', 'mgs', 'mgd', 'mmgs',
                'lac', 'calving_date', 'birth_date', 'birth_date_dam', 'birth_date_mgd',
                'age', 'services_time', 'DIM', 'peak_milk', 'milk_305',
                'repro_status', 'group', '是否在场', 'birth_year'
            ]

            # 动态获取所有_score列（性状列）
            score_columns = [col for col in df.columns if col.endswith('_score')]
            logger.info(f"检测到 {len(score_columns)} 个性状列")

            # 组合基础列和性状列
            display_columns = base_columns + score_columns

            # 检查哪些列存在
            existing_columns = [col for col in display_columns if col in df.columns]
            df_display = df[existing_columns].copy()

            # 转换列名为中文
            chinese_headers = []
            for col in existing_columns:
                if col in self.COLUMN_MAPPING:
                    chinese_headers.append(self.COLUMN_MAPPING[col])
                elif col.endswith('_score'):
                    # 动态生成性状列的中文名
                    chinese_headers.append(self._get_trait_name(col))
                else:
                    chinese_headers.append(col)

            current_row = 1

            # 写入表头
            self._write_header(current_row, chinese_headers, start_col=1)
            current_row += 1

            # 写入数据（使用append批量写入整行）
            for idx, row in df_display.iterrows():
                values = list(row)
                self.ws.append(values)

            current_row += len(df_display)

            # 设置列宽
            base_col_count = len([col for col in existing_columns if not col.endswith('_score')])
            for col_idx in range(1, len(chinese_headers) + 1):
                col_letter = get_column_letter(col_idx)
                # 根据列内容设置不同宽度
                if col_idx <= 8:  # 耳号、品种等基础信息
                    self.ws.column_dimensions[col_letter].width = 15
                elif col_idx <= base_col_count:  # 日期、月龄等基础列
                    self.ws.column_dimensions[col_letter].width = 12
                else:  # 性状分数（动态列）
                    self.ws.column_dimensions[col_letter].width = 10

            # 冻结首行
            self._freeze_panes('A2')

            logger.info(f"✓ Sheet 3-4构建完成: {len(df_display)}行数据, {len(score_columns)}个性状列")

        except Exception as e:
            logger.error(f"构建Sheet 3-4失败: {e}", exc_info=True)
            raise

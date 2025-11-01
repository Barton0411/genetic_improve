"""
Sheet 7构建器: 配种记录-隐性基因/近交系数明细
v1.2版本 - 直接复制已配公牛分析结果文件
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import Alignment
import logging
import pandas as pd

logger = logging.getLogger(__name__)


class Sheet7Builder(BaseSheetBuilder):
    """
    Sheet 7: 配种记录-隐性基因/近交系数明细

    直接复制"已配公牛_近交系数及隐性基因分析结果"文件内容
    """

    def build(self, data: dict):
        """
        构建Sheet 7: 配种记录明细

        Args:
            data: 包含配种记录明细数据
                - data: DataFrame数据
                - file_path: 源文件路径
        """
        try:
            # 检查数据
            if not data or 'data' not in data:
                logger.warning("Sheet7: 缺少数据，跳过生成")
                return

            df = data['data']
            if df.empty:
                logger.warning("Sheet7: 数据为空，跳过生成")
                return

            # 创建Sheet（Excel sheet名称不能包含斜杠）
            self._create_sheet("配种记录-隐性基因及近交系数明细")
            logger.info("构建Sheet 7: 配种记录-隐性基因及近交系数明细")

            # 准备列宽
            column_widths = {}
            for col_idx, col_name in enumerate(df.columns, start=1):
                # 基础列宽
                if '(' in str(col_name):  # 带括号的基因列，较窄
                    width = 10
                elif col_name in ['母牛号', '配种公牛号', '原始公牛号']:
                    width = 15
                elif col_name in ['配种日期']:
                    width = 12
                elif col_name in ['后代近交详情']:
                    width = 30
                else:
                    width = 12

                column_widths[col_idx] = width

            # 使用快速方法写入数据
            current_row = self._write_dataframe_fast(
                df,
                start_row=1,
                data_alignment='center',
                column_widths=column_widths
            )

            # 冻结首行
            self._freeze_panes('A2')

            logger.info(f"✓ Sheet 7构建完成: {len(df)}行 × {len(df.columns)}列")

        except Exception as e:
            logger.error(f"构建Sheet 7失败: {e}", exc_info=True)
            raise

    def _get_column_letter(self, col_idx: int) -> str:
        """获取列字母（1->A, 2->B, ...）"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_idx)

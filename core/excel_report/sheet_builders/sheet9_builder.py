"""
Sheet 9构建器: 已用公牛性状明细
v1.2.2版本 - 按年份分组的公牛使用明细

每年一个表格，包含该年使用的所有公牛及其性状
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import logging

logger = logging.getLogger(__name__)


class Sheet9Builder(BaseSheetBuilder):
    """
    Sheet 9: 已用公牛性状明细

    按年份展示每年使用的冻精明细:
    - 每年一个表格
    - 包含冻精编号、配种类型、使用次数、所有性状
    - 表格末尾添加统计行
    - 年份从新到旧排列
    """

    def build(self, data: dict):
        """
        构建Sheet 9: 已用公牛性状明细

        Args:
            data: 包含以下键的字典:
                - yearly_details: 字典，键为年份，值为该年的公牛明细DataFrame
                - trait_columns: 性状列列表
                - year_range: 年份范围列表（倒序）
        """
        try:
            # 检查数据
            if not data or 'yearly_details' not in data:
                logger.warning("Sheet9: 缺少数据，跳过生成")
                return

            # 创建Sheet
            self._create_sheet("已用公牛性状明细")
            logger.info("构建Sheet 9: 已用公牛性状明细")

            yearly_details = data.get('yearly_details', {})
            trait_columns = data.get('trait_columns', [])
            year_range = data.get('year_range', [])

            if not yearly_details:
                logger.warning("Sheet9: 没有年度数据")
                return

            current_row = 1

            # 按年份（从新到旧）构建表格
            for idx, year in enumerate(year_range):
                detail_df = yearly_details.get(year)

                if detail_df is None or detail_df.empty:
                    logger.warning(f"  {year}年: 无数据，跳过")
                    continue

                logger.info(f"  构建{year}年表格: {len(detail_df)}头公牛, {detail_df['使用次数'].sum()}配次")

                current_row = self._build_year_table(
                    year, detail_df, trait_columns, current_row
                )
                current_row += 3  # 每个表格之间空3行

            # 设置列宽
            self._set_default_column_widths(trait_columns)

            # 冻结首行
            self._freeze_panes('A2')

            logger.info("✓ Sheet 9构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 9失败: {e}", exc_info=True)
            raise

    def _build_year_table(
        self,
        year: int,
        detail_df: pd.DataFrame,
        trait_columns: list,
        start_row: int
    ) -> int:
        """
        构建单个年份的表格

        Args:
            year: 年份
            detail_df: 该年的公牛明细DataFrame
            trait_columns: 性状列列表
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        total_bulls = len(detail_df)
        total_times = int(detail_df['使用次数'].sum())
        title = f"{year}年已用公牛明细（{total_bulls}头公牛，{total_times}配次）"

        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头行和数据行（使用快速方法）
        headers = ['冻精编号', '配种类型', '使用次数'] + trait_columns

        # 使用快速方法写入数据
        current_row = self._write_dataframe_fast(
            detail_df[headers],
            start_row=current_row,
            headers=headers,
            data_alignment='center',
            column_widths=None  # 后面统一设置
        )

        # 统计行
        stat_row_data = {
            '冻精编号': f'小计({total_bulls}头)',
            '配种类型': '-',
            '使用次数': total_times
        }

        # 计算各性状平均值（加权平均，按使用次数加权）
        for trait in trait_columns:
            if trait == 'Eval Date':
                stat_row_data[trait] = '-'
            else:
                # 计算加权平均（按使用次数加权）
                if trait in detail_df.columns:
                    valid_data = detail_df[[trait, '使用次数']].dropna(subset=[trait])
                    if len(valid_data) > 0:
                        weighted_avg = (
                            valid_data[trait] * valid_data['使用次数']
                        ).sum() / valid_data['使用次数'].sum()
                        stat_row_data[trait] = weighted_avg
                    else:
                        stat_row_data[trait] = None
                else:
                    stat_row_data[trait] = None

        # 写入统计行
        for col_idx, col_name in enumerate(headers, start=1):
            value = stat_row_data.get(col_name)
            cell = self.ws.cell(row=current_row, column=col_idx, value=value)

            # 样式
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')

            # 格式
            if col_name == '冻精编号':
                self.style_manager.apply_data_style(cell, alignment='left')
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type='solid')
            elif col_name == '配种类型':
                self.style_manager.apply_data_style(cell, alignment='center')
            elif col_name == '使用次数':
                self.style_manager.apply_data_style(cell, alignment='center')
            elif col_name == 'Eval Date':
                self.style_manager.apply_data_style(cell, alignment='center')
            else:
                # 性状列
                if isinstance(value, (int, float)) and pd.notna(value):
                    self.style_manager.apply_data_style(cell, alignment='center')
                    cell.number_format = '0.00'
                else:
                    self.style_manager.apply_data_style(cell, alignment='center')

        current_row += 1

        return current_row

    def _set_default_column_widths(self, trait_columns: list):
        """设置默认列宽"""
        col_widths = {
            1: 18,  # 冻精编号
            2: 12,  # 配种类型
            3: 12,  # 使用次数
        }

        # 性状列默认宽度
        for col_idx in range(4, len(trait_columns) + 4):
            col_widths[col_idx] = 15

        self._set_column_widths(col_widths)

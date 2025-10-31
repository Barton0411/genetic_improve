"""
Sheet 11构建器: 个体选配推荐结果
v1.2版本 - 选配推荐明细及统计摘要
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import logging

logger = logging.getLogger(__name__)


class Sheet11Builder(BaseSheetBuilder):
    """
    Sheet 11: 个体选配推荐结果

    包含内容:
    1. 选配统计摘要
    2. 选配推荐明细表
    """

    def build(self, data: dict):
        """
        构建Sheet 11: 个体选配推荐结果

        Args:
            data: 包含选配推荐数据
                - mating_details: 选配推荐明细DataFrame
                - mating_summary: 选配统计摘要字典
        """
        try:
            logger.info("构建Sheet 11: 个体选配推荐结果")

            # 检查数据
            if not data or 'mating_details' not in data:
                logger.warning("Sheet11: 缺少选配数据，跳过生成")
                return

            # 创建Sheet
            self._create_sheet("个体选配推荐结果")

            mating_details = data.get('mating_details')
            mating_summary = data.get('mating_summary', {})

            if mating_details is None or mating_details.empty:
                logger.warning("选配明细数据为空，跳过构建")
                return

            current_row = 1

            # 1. 构建统计摘要
            if mating_summary:
                current_row = self._build_summary_section(current_row, mating_summary)
                current_row += 2  # 空2行

            # 2. 构建选配明细表
            current_row = self._build_mating_details_table(current_row, mating_details)

            logger.info("✓ Sheet 11构建完成")

        except Exception as e:
            logger.error(f"Sheet 11构建失败: {e}", exc_info=True)
            raise

    def _build_summary_section(self, start_row: int, summary: dict) -> int:
        """
        构建统计摘要部分

        Args:
            start_row: 起始行号
            summary: 统计摘要字典

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题
        title_cell = self.ws.cell(row=current_row, column=1, value="选配统计摘要")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        # 总体统计
        total_cows = summary.get('total_cows', 0)
        sexed_count = summary.get('has_sexed_recommendation', 0)
        regular_count = summary.get('has_regular_recommendation', 0)

        stats_data = [
            ('总母牛数', total_cows),
            ('有性控推荐', sexed_count),
            ('有常规推荐', regular_count),
            ('无推荐', total_cows - max(sexed_count, regular_count))
        ]

        for label, value in stats_data:
            self.ws.cell(row=current_row, column=1, value=label)
            self.ws.cell(row=current_row, column=2, value=value)
            self.style_manager.apply_header_style(self.ws.cell(row=current_row, column=1))
            self.style_manager.apply_data_style(self.ws.cell(row=current_row, column=2), alignment='center')
            current_row += 1

        current_row += 1  # 空行

        # 按分组统计
        groups = summary.get('groups', {})
        if groups:
            # 分组统计标题
            group_title = self.ws.cell(row=current_row, column=1, value="按分组统计")
            group_title.font = Font(bold=True)
            current_row += 1

            # 表头
            headers = ['分组', '母牛数', '性控推荐数', '常规推荐数']
            for col_idx, header in enumerate(headers, 1):
                cell = self.ws.cell(row=current_row, column=col_idx, value=header)
                self.style_manager.apply_header_style(cell)
            current_row += 1

            # 数据行
            for group_name, group_stats in sorted(groups.items()):
                self.ws.cell(row=current_row, column=1, value=group_name)
                self.ws.cell(row=current_row, column=2, value=group_stats['count'])
                self.ws.cell(row=current_row, column=3, value=group_stats['sexed_count'])
                self.ws.cell(row=current_row, column=4, value=group_stats['regular_count'])

                for col_idx in range(1, 5):
                    self.style_manager.apply_data_style(
                        self.ws.cell(row=current_row, column=col_idx),
                        alignment='center'
                    )
                current_row += 1

        return current_row

    def _build_mating_details_table(self, start_row: int, df: pd.DataFrame) -> int:
        """
        构建选配明细表

        Args:
            start_row: 起始行号
            df: 选配明细DataFrame

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题
        title_cell = self.ws.cell(row=current_row, column=1, value="选配推荐明细")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        # 合并标题行（根据实际列数）
        num_cols = len(df.columns)
        self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        # 表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=col_name)
            self.style_manager.apply_header_style(cell)

            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if '母牛号' in col_name or '父亲' in col_name or '外祖父' in col_name:
                self.ws.column_dimensions[col_letter].width = 15
            elif '选' in col_name and ('性控' in col_name or '常规' in col_name):
                self.ws.column_dimensions[col_letter].width = 14
            elif '备注' in col_name:
                self.ws.column_dimensions[col_letter].width = 50
            elif '分组' in col_name:
                self.ws.column_dimensions[col_letter].width = 20
            elif col_name in ['胎次', '配次', '月龄', '品种', '繁育状态']:
                self.ws.column_dimensions[col_letter].width = 10
            else:
                self.ws.column_dimensions[col_letter].width = 12

        current_row += 1
        header_row = current_row - 1

        # 数据行
        for _, row_data in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, 1):
                value = row_data[col_name]

                # 处理NaN和NaT
                if pd.isna(value):
                    value = ""
                elif isinstance(value, pd.Timestamp):
                    value = value.strftime('%Y-%m-%d')

                cell = self.ws.cell(row=current_row, column=col_idx, value=value)

                # 应用格式
                if col_name in ['胎次', '配次', '月龄', '泌乳天数']:
                    self.style_manager.apply_data_style(cell, alignment='center')
                elif col_name == '母牛指数得分':
                    self.style_manager.apply_data_style(cell, alignment='center')
                    cell.number_format = '0.00'
                elif '备注' in col_name:
                    self.style_manager.apply_data_style(cell, alignment='left')
                    # 如果包含警告信息，标记为黄色
                    if value and ('风险' in str(value) or '近交系数' in str(value)):
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                else:
                    self.style_manager.apply_data_style(cell, alignment='center')

            current_row += 1

        # 冻结首行和首列
        self.ws.freeze_panes = self.ws.cell(row=header_row + 1, column=2)

        logger.info(f"✓ Sheet 11明细表构建完成: {len(df)}行数据")

        return current_row

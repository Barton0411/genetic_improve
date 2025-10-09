"""
Sheet 1构建器: 牧场基础信息
"""

from .base_builder import BaseSheetBuilder
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class Sheet1Builder(BaseSheetBuilder):
    """Sheet 1: 牧场基础信息"""

    def build(self, data: dict):
        """
        构建Sheet 1 - 横向布局（三个部分并排）

        Args:
            data: {
                'basic_info': {...},
                'herd_structure': {...},
                'upload_summary': {...}
            }
        """
        try:
            # 创建Sheet
            self._create_sheet("牧场基础信息")
            logger.info("构建Sheet 1: 牧场基础信息")

            # 横向布局：四个部分并排
            # 第一列（A-B）：基本信息
            # 第二列（D-F）：成母牛/后备牛分布统计
            # 第三列（H-J）：不同胎次分布统计
            # 第四列（L-N）：上传数据概览

            # 1. 基本信息部分（A-B列，从第1行开始）
            self._build_basic_info_section_horizontal(data['basic_info'], start_row=1, start_col=1)

            # 2. 成母牛/后备牛分布统计（D-F列，从第1行开始）
            self._build_cow_type_section_horizontal(data['herd_structure'], start_row=1, start_col=4)

            # 3. 不同胎次分布统计（H-J列，从第1行开始）
            self._build_lactation_section_horizontal(data['herd_structure'], start_row=1, start_col=8)

            # 4. 上传数据概览部分（L-N列，从第1行开始）
            self._build_upload_summary_section_horizontal(data['upload_summary'], start_row=1, start_col=12)

            # 设置列宽（根据图表宽度调整）
            # 图表宽度15厘米，1厘米 ≈ 4个Excel列宽单位，所以15cm ≈ 60列宽单位
            # 第2、3部分各3列总和需为60列宽单位
            self._set_column_widths({
                1: 18,  # A列 - 基本信息（标签列）
                2: 25,  # B列 - 基本信息（值列，需要显示"2025-10-08 11:02:00"）
                3: 3,   # C列（空列）
                4: 24,  # D列 - 成母牛/后备牛（类型列）
                5: 18,  # E列 - 数量(头)
                6: 18,  # F列 - 占比(%)  (24+18+18=60)
                7: 3,   # G列（空列）
                8: 20,  # H列 - 不同胎次（胎次列）
                9: 20,  # I列 - 数量(头)
                10: 20, # J列 - 占比(%)  (20+20+20=60)
                11: 3,  # K列（空列）
                12: 18, # L列 - 上传数据概览（数据类型列）
                13: 12, # M列 - 记录数
                14: 55, # N列（备注列，需要显示完整备注信息）
            })

            logger.info("✓ Sheet 1构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 1失败: {e}", exc_info=True)
            raise

    def _build_basic_info_section_horizontal(self, basic_info: dict, start_row: int, start_col: int):
        """
        构建基本信息部分（横向布局）

        Args:
            basic_info: 基本信息数据
            start_row: 起始行号
            start_col: 起始列号（A=1, D=4, I=9）
        """
        current_row = start_row

        # 标题
        cell = self.ws.cell(row=current_row, column=start_col, value="一、基本信息")
        self.style_manager.apply_title_style(cell)
        current_row += 1

        # 基本信息表格（深蓝色标签，浅蓝色值）
        info_items = [
            ("牧场名称", basic_info.get('farm_name', '')),
            ("报告生成时间", basic_info.get('report_time', '')),
            ("牧场服务人员", basic_info.get('service_staff', '未指定')),
        ]

        for label, value in info_items:
            # 标签列（深蓝色背景，白字）
            label_cell = self.ws.cell(row=current_row, column=start_col, value=label)
            self.style_manager.apply_basic_info_label_style(label_cell)

            # 值列（浅蓝色背景）
            value_cell = self.ws.cell(row=current_row, column=start_col+1, value=value)
            self.style_manager.apply_basic_info_value_style(value_cell)
            current_row += 1

    def _build_cow_type_section_horizontal(self, herd_structure: dict, start_row: int, start_col: int):
        """
        构建成母牛/后备牛分布统计部分（横向布局）

        Args:
            herd_structure: 牛群结构数据
            start_row: 起始行号
            start_col: 起始列号
        """
        current_row = start_row

        # 标题
        cell = self.ws.cell(row=current_row, column=start_col, value="二、成母牛/后备牛分布统计")
        self.style_manager.apply_title_style(cell)
        current_row += 1

        # 表格标题行
        headers = ["类型", "数量(头)", "占比(%)"]
        for idx, header in enumerate(headers):
            cell = self.ws.cell(row=current_row, column=start_col+idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据
        total_count = herd_structure.get('total_count', 0)
        lactating_count = herd_structure.get('lactating_count', 0)
        heifer_count = herd_structure.get('heifer_count', 0)

        lactating_pct = round(lactating_count / total_count * 100, 1) if total_count > 0 else 0
        heifer_pct = round(heifer_count / total_count * 100, 1) if total_count > 0 else 0

        # 成母牛行
        lactating_row = current_row
        self._write_data_row(current_row, ["成母牛(胎次>0)", lactating_count, lactating_pct], start_col=start_col, alignment='left')
        current_row += 1

        # 后备牛行
        heifer_row = current_row
        self._write_data_row(current_row, ["后备牛(胎次=0)", heifer_count, heifer_pct], start_col=start_col, alignment='left')
        current_row += 1

        # 合计行
        self._write_total_row(current_row, ["合计", total_count, 100], start_col=start_col)
        current_row += 2

        # 平均胎次和泌乳天数（添加边框）
        avg_lactation = herd_structure.get('avg_lactation', 0)
        avg_dim = herd_structure.get('avg_dim', 0)

        label_cell = self.ws.cell(row=current_row, column=start_col, value="在群母牛平均胎次")
        self.style_manager.apply_data_style(label_cell, alignment='left')
        value_cell = self.ws.cell(row=current_row, column=start_col+1, value=avg_lactation)
        self.style_manager.apply_data_style(value_cell, alignment='center')
        current_row += 1

        label_cell = self.ws.cell(row=current_row, column=start_col, value="在群母牛平均泌乳天数")
        self.style_manager.apply_data_style(label_cell, alignment='left')
        value_cell = self.ws.cell(row=current_row, column=start_col+1, value=f"{avg_dim}天")
        self.style_manager.apply_data_style(value_cell, alignment='center')
        current_row += 2

        # 饼图：成母牛/后备牛占比（10cm高）
        if lactating_count > 0 or heifer_count > 0:
            pie_row = current_row
            chart_data_range = {
                'labels': (lactating_row, heifer_row, start_col),
                'values': (lactating_row, heifer_row, start_col+1)
            }
            chart_col_letter = get_column_letter(start_col)
            self._add_pie_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{pie_row}',
                title="成母牛/后备牛占比"
            )
            current_row += 20  # 饼图与柱状图间距

        # 柱状图：成母牛/后备牛数量（7.5cm高）
        if lactating_count > 0 or heifer_count > 0:
            bar_row = current_row
            chart_col_letter = get_column_letter(start_col)
            self._add_bar_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{bar_row}',
                title="成母牛/后备牛数量"
            )

    def _build_lactation_section_horizontal(self, herd_structure: dict, start_row: int, start_col: int):
        """
        构建不同胎次分布统计部分（横向布局）

        Args:
            herd_structure: 牛群结构数据
            start_row: 起始行号
            start_col: 起始列号
        """
        current_row = start_row

        # 标题
        cell = self.ws.cell(row=current_row, column=start_col, value="三、不同胎次分布统计")
        self.style_manager.apply_title_style(cell)
        current_row += 1

        # 表格标题行
        headers = ["胎次", "数量(头)", "占比(%)"]
        for idx, header in enumerate(headers):
            cell = self.ws.cell(row=current_row, column=start_col+idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 按胎次分布数据
        total_count = herd_structure.get('total_count', 0)
        lactation_distribution = herd_structure.get('lactation_distribution', {})
        lac_data_rows = []

        for lac_label in ['0胎', '1胎', '2胎', '3胎及以上']:
            count = lactation_distribution.get(lac_label, 0)
            pct = round(count / total_count * 100, 1) if total_count > 0 else 0
            self._write_data_row(current_row, [lac_label, count, pct], start_col=start_col, alignment='left')
            lac_data_rows.append(current_row)
            current_row += 1

        # 合计行
        self._write_total_row(current_row, ["合计", total_count, 100], start_col=start_col)
        current_row += 2

        # 饼图：按胎次分布占比（10cm高）
        if lac_data_rows:
            pie_row = current_row
            chart_data_range = {
                'labels': (lac_data_rows[0], lac_data_rows[-1], start_col),
                'values': (lac_data_rows[0], lac_data_rows[-1], start_col+1)
            }
            chart_col_letter = get_column_letter(start_col)
            self._add_pie_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{pie_row}',
                title="按胎次分布占比"
            )
            current_row += 20  # 饼图与柱状图间距

        # 柱状图：按胎次分布数量（7.5cm高）
        if lac_data_rows:
            bar_row = current_row
            chart_col_letter = get_column_letter(start_col)
            self._add_bar_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{bar_row}',
                title="按胎次分布数量"
            )

    def _build_herd_structure_section_horizontal_old(self, herd_structure: dict, start_row: int, start_col: int):
        """
        构建牛群结构统计部分（横向布局）

        Args:
            herd_structure: 牛群结构数据
            start_row: 起始行号
            start_col: 起始列号
        """
        current_row = start_row

        # 标题
        cell = self.ws.cell(row=current_row, column=start_col, value="二、牛群结构统计")
        self.style_manager.apply_title_style(cell)
        current_row += 1

        # 2.1 成母牛/后备牛统计表格
        headers = ["类型", "数量(头)", "占比(%)", ""]
        for idx, header in enumerate(headers[:3]):  # 只写前3列
            cell = self.ws.cell(row=current_row, column=start_col+idx, value=header)
            self.style_manager.apply_header_style(cell)
        table1_start_row = current_row
        current_row += 1

        # 数据
        total_count = herd_structure.get('total_count', 0)
        lactating_count = herd_structure.get('lactating_count', 0)
        heifer_count = herd_structure.get('heifer_count', 0)

        lactating_pct = round(lactating_count / total_count * 100, 1) if total_count > 0 else 0
        heifer_pct = round(heifer_count / total_count * 100, 1) if total_count > 0 else 0

        # 成母牛行
        lactating_row = current_row
        self.ws.cell(row=current_row, column=start_col, value="成母牛(胎次>0)")
        self.ws.cell(row=current_row, column=start_col+1, value=lactating_count)
        self.ws.cell(row=current_row, column=start_col+2, value=lactating_pct)
        current_row += 1

        # 后备牛行
        heifer_row = current_row
        self.ws.cell(row=current_row, column=start_col, value="后备牛(胎次=0)")
        self.ws.cell(row=current_row, column=start_col+1, value=heifer_count)
        self.ws.cell(row=current_row, column=start_col+2, value=heifer_pct)
        current_row += 1

        # 合计行
        self._write_total_row(current_row, ["合计", total_count, 100.0], start_col=start_col)
        current_row += 2

        # 平均胎次和泌乳天数
        avg_lactation = herd_structure.get('avg_lactation', 0)
        avg_dim = herd_structure.get('avg_dim', 0)
        self.ws.cell(row=current_row, column=start_col, value="在群母牛平均胎次")
        self.ws.cell(row=current_row, column=start_col+1, value=avg_lactation)
        current_row += 1
        self.ws.cell(row=current_row, column=start_col, value="在群母牛平均泌乳天数")
        self.ws.cell(row=current_row, column=start_col+1, value=f"{avg_dim}天")
        current_row += 2

        # 饼图：成母牛/后备牛占比（在表格下方）
        if lactating_count > 0 or heifer_count > 0:
            pie_row = current_row
            chart_data_range = {
                'labels': (lactating_row, heifer_row, start_col),
                'values': (lactating_row, heifer_row, start_col+1)
            }
            chart_col_letter = get_column_letter(start_col)
            self._add_pie_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{pie_row}',
                title="成母牛/后备牛占比"
            )
            current_row += 16  # 饼图高度

        # 柱状图：成母牛/后备牛数量（在饼图下方）
        if lactating_count > 0 or heifer_count > 0:
            bar_row = current_row
            chart_col_letter = get_column_letter(start_col)
            self._add_bar_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{bar_row}',
                title="成母牛/后备牛数量"
            )
            current_row += 16  # 柱状图高度

        # 2.2 按胎次分布统计表格（在柱状图下方）
        current_row += 2
        table2_start_row = current_row
        headers = ["胎次", "数量(头)", "占比(%)", ""]
        for idx, header in enumerate(headers[:3]):
            cell = self.ws.cell(row=current_row, column=start_col+idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 按胎次分布数据
        lactation_distribution = herd_structure.get('lactation_distribution', {})
        lac_data_rows = []

        for lac_label in ['0胎', '1胎', '2胎', '3胎及以上']:
            count = lactation_distribution.get(lac_label, 0)
            pct = round(count / total_count * 100, 1) if total_count > 0 else 0
            self.ws.cell(row=current_row, column=start_col, value=lac_label)
            self.ws.cell(row=current_row, column=start_col+1, value=count)
            self.ws.cell(row=current_row, column=start_col+2, value=pct)
            lac_data_rows.append(current_row)
            current_row += 1

        # 合计行
        self._write_total_row(current_row, ["合计", total_count, 100.0], start_col=start_col)
        current_row += 2

        # 饼图：按胎次分布占比
        if lac_data_rows:
            pie_row = current_row
            chart_data_range = {
                'labels': (lac_data_rows[0], lac_data_rows[-1], start_col),
                'values': (lac_data_rows[0], lac_data_rows[-1], start_col+1)
            }
            chart_col_letter = get_column_letter(start_col)
            self._add_pie_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{pie_row}',
                title="按胎次分布占比"
            )
            current_row += 16

        # 柱状图：按胎次分布数量
        if lac_data_rows:
            bar_row = current_row
            chart_col_letter = get_column_letter(start_col)
            self._add_bar_chart(
                data_range=chart_data_range,
                position=f'{chart_col_letter}{bar_row}',
                title="按胎次分布数量"
            )

    def _build_upload_summary_section_horizontal(self, upload_summary: dict, start_row: int, start_col: int):
        """
        构建上传数据概览部分（横向布局）

        Args:
            upload_summary: 上传数据概览数据
            start_row: 起始行号
            start_col: 起始列号
        """
        current_row = start_row

        # 标题
        cell = self.ws.cell(row=current_row, column=start_col, value="四、上传数据概览")
        self.style_manager.apply_title_style(cell)
        current_row += 1

        # 表格标题行
        headers = ["数据类型", "记录数", "备注"]
        for idx, header in enumerate(headers):
            cell = self.ws.cell(row=current_row, column=start_col+idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据行
        summary_items = [
            ("母牛信息", upload_summary.get('cow_data_count', 0),
             upload_summary.get('cow_data_detail', f"在场: {upload_summary.get('cow_data_active', 0)}头, 离场: {upload_summary.get('cow_data_inactive', 0)}头")),
            ("配种记录", upload_summary.get('breeding_records', 0), ""),
            ("备选公牛", upload_summary.get('bull_count', 0), ""),
            ("体型外貌数据", upload_summary.get('body_conformation_count', 0), ""),
            ("基因组数据", upload_summary.get('genomic_count', 0), ""),
        ]

        for data_type, count, note in summary_items:
            self._write_data_row(current_row, [data_type, count, note], start_col=start_col, alignment='left')
            current_row += 1

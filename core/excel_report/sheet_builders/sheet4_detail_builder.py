"""
Sheet 4-2构建器: 母牛指数排名明细表
"""

from .base_builder import BaseSheetBuilder
from openpyxl.utils import get_column_letter
import pandas as pd
import logging

logger = logging.getLogger(__name__)


class Sheet4DetailBuilder(BaseSheetBuilder):
    """Sheet 4-2: 母牛指数排名明细表"""

    # 基础列名中英文映射
    COLUMN_MAPPING = {
        'cow_id': '耳号',
        'breed': '品种',
        'sex': '性别',
        'sire': '父号',
        'dam': '母号',
        'mgs': '外祖父号',
        'mgd': '外祖母号',
        'mmgs': '外外祖父号',
        'lac': '胎次',
        'calving_date': '产犊日期',
        'birth_date': '出生日期',
        'birth_date_dam': '母牛出生日期',
        'birth_date_mgd': '外祖母出生日期',
        'age': '月龄',
        'services_time': '配种次数',
        'DIM': '泌乳天数',
        'peak_milk': '峰值奶量',
        'milk_305': '305天奶量',
        'repro_status': '繁殖状态',
        'group': '组别',
        '是否在场': '是否在场',
        'birth_year': '出生年份',
        'ranking': '排名',
    }

    @staticmethod
    def _get_trait_name(score_column: str) -> str:
        """
        从_score列名获取性状名称
        例如: 'NM$_score' -> 'NM$'
        """
        if score_column.endswith('_score'):
            return score_column[:-6]  # 去掉'_score'后缀
        return score_column

    @staticmethod
    def _get_index_name(index_column: str) -> str:
        """
        从_index列名获取指数名称
        例如: '测试_index' -> '测试指数'
        """
        if index_column.endswith('_index'):
            return index_column[:-6] + '指数'  # 去掉'_index'后缀，加'指数'
        return index_column

    def build(self, data: dict):
        """
        构建Sheet 4-2

        Args:
            data: {
                'detail_df': 母牛指数明细DataFrame
            }
        """
        try:
            # 创建Sheet
            self._create_sheet("母牛指数排名明细")
            logger.info("构建Sheet 4-2: 母牛指数排名明细")

            df = data.get('detail_df')

            if df is None or df.empty:
                logger.warning("Sheet 4-2: 母牛指数明细数据为空，跳过构建")
                return

            # 定义基础列（固定顺序）
            base_columns = [
                'cow_id', 'breed', 'sex', 'sire', 'dam', 'mgs', 'mgd', 'mmgs',
                'lac', 'calving_date', 'birth_date', 'birth_date_dam', 'birth_date_mgd',
                'age', 'services_time', 'DIM', 'peak_milk', 'milk_305',
                'repro_status', '是否在场', 'birth_year'
            ]

            # 动态获取所有_score列（性状列）
            score_columns = [col for col in df.columns if col.endswith('_score')]
            logger.info(f"检测到 {len(score_columns)} 个性状列")

            # 动态获取所有_index列（指数列）
            index_columns = [col for col in df.columns if col.endswith('_index')]
            logger.info(f"检测到 {len(index_columns)} 个指数列")

            # 获取ranking列
            ranking_columns = ['ranking'] if 'ranking' in df.columns else []

            # 组合列顺序: cow_id, *_index, ranking, 基础列（去掉cow_id）, 性状列
            display_columns = ['cow_id'] + index_columns + ranking_columns

            # 添加基础列（排除已有的cow_id）
            for col in base_columns:
                if col != 'cow_id' and col in df.columns and col not in display_columns:
                    display_columns.append(col)

            # 添加性状列
            display_columns.extend(score_columns)

            # 检查哪些列存在
            existing_columns = [col for col in display_columns if col in df.columns]
            df_display = df[existing_columns].copy()

            # 按ranking排序（如果存在），否则按第一个_index列排序
            if 'ranking' in df_display.columns:
                df_display = df_display.sort_values('ranking', ascending=True)
                logger.info("按ranking列排序")
            elif index_columns:
                df_display = df_display.sort_values(index_columns[0], ascending=False)
                logger.info(f"按{index_columns[0]}列排序")

            # 转换列名为中文
            chinese_headers = []
            for col in existing_columns:
                if col in self.COLUMN_MAPPING:
                    chinese_headers.append(self.COLUMN_MAPPING[col])
                elif col.endswith('_score'):
                    # 动态生成性状列的中文名
                    chinese_headers.append(self._get_trait_name(col))
                elif col.endswith('_index'):
                    # 动态生成指数列的中文名
                    chinese_headers.append(self._get_index_name(col))
                else:
                    chinese_headers.append(col)

            # 直接写入表头（不需要标题行）
            current_row = 1
            self._write_header(current_row, chinese_headers)
            current_row += 1

            # 写入数据
            for idx, row in df_display.iterrows():
                values = list(row[existing_columns])

                # 判断是否在场
                if '是否在场' in df_display.columns:
                    is_present = row['是否在场'] == '是'
                else:
                    is_present = True

                # 在场牛只正常显示，离场牛只灰色背景
                if is_present:
                    self._write_data_row(current_row, values, alignment='center')
                else:
                    # 离场牛只用灰色背景
                    for col_idx, value in enumerate(values):
                        cell = self.ws.cell(row=current_row, column=col_idx + 1, value=value)
                        self.style_manager.apply_data_style(cell, alignment='center')
                        # 应用灰色背景
                        from openpyxl.styles import PatternFill
                        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                current_row += 1

            # 设置列宽
            for col_idx in range(1, len(chinese_headers) + 1):
                col_letter = get_column_letter(col_idx)
                col_name = chinese_headers[col_idx - 1]

                # 根据列内容设置宽度
                if '排名' in col_name:
                    self.ws.column_dimensions[col_letter].width = 8
                elif '耳号' in col_name or 'ID' in col_name:
                    self.ws.column_dimensions[col_letter].width = 15
                elif '日期' in col_name:
                    self.ws.column_dimensions[col_letter].width = 12
                elif '指数' in col_name:
                    self.ws.column_dimensions[col_letter].width = 12
                else:
                    self.ws.column_dimensions[col_letter].width = 12

            # 冻结首行
            self._freeze_panes('A2')

            # 添加筛选器
            self.ws.auto_filter.ref = f'A1:{get_column_letter(len(chinese_headers))}{current_row - 1}'

            logger.info(f"✓ Sheet 4-2构建完成: {len(df_display)}行数据, {len(score_columns)}个性状列, {len(index_columns)}个指数列")

        except Exception as e:
            logger.error(f"构建Sheet 4-2失败: {e}", exc_info=True)
            raise

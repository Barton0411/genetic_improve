"""
Sheet 2构建器: 系谱识别分析
"""

from .base_builder import BaseSheetBuilder
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class Sheet2Builder(BaseSheetBuilder):
    """Sheet 2: 系谱识别分析"""

    def build(self, data: dict):
        """
        构建Sheet 2

        Args:
            data: {
                'summary': 汇总数据DataFrame,
                'detail': 明细数据DataFrame,
                'total_stats': 总体统计dict
            }
        """
        try:
            # 创建Sheet
            self._create_sheet("系谱识别分析")
            logger.info("构建Sheet 2: 系谱识别分析")

            if data.get('summary') is None or data['summary'].empty:
                logger.warning("Sheet 2数据为空，跳过构建")
                return

            current_row = 1

            # 1. 汇总表标题
            title_cell = self.ws.cell(row=current_row, column=1, value="一、系谱识别情况汇总（按出生年份）（在群母牛）")
            self.style_manager.apply_title_style(title_cell)
            current_row += 2

            # 2. 汇总表
            summary_df = data['summary']
            summary_start_row = current_row

            # 表头
            headers = list(summary_df.columns)
            self._write_header(current_row, headers)
            current_row += 1

            # 数据行
            for idx, row in summary_df.iterrows():
                values = list(row)

                # 判断是否为合计行
                if row['出生年份'] == '合计':
                    self._write_total_row(current_row, values)
                else:
                    self._write_data_row(current_row, values, alignment='center')

                    # 检查识别率并标红（<80%）
                    for col_idx, header in enumerate(headers):
                        if '占比' in header:
                            rate_str = str(row[header])
                            if rate_str.endswith('%'):
                                rate_value = float(rate_str.rstrip('%'))
                                if rate_value < 80:
                                    cell = self.ws.cell(row=current_row, column=col_idx+1)
                                    self.style_manager.apply_risk_color(cell, 'high')

                current_row += 1

            current_row += 2

            # 3. 三个饼图（并排）
            total_stats = data.get('total_stats', {})

            if total_stats:
                # 饼图标题
                chart_title_cell = self.ws.cell(row=current_row, column=1, value="二、系谱识别率可视化分析（在群母牛）")
                self.style_manager.apply_title_style(chart_title_cell)
                current_row += 2

                # 为饼图准备数据（在隐藏区域）
                chart_data_row = current_row

                # 父号识别数据
                self.ws.cell(row=chart_data_row, column=1, value="已识别")
                self.ws.cell(row=chart_data_row + 1, column=1, value="未识别")
                self.ws.cell(row=chart_data_row, column=2, value=total_stats.get('sire_identified', 0))
                self.ws.cell(row=chart_data_row + 1, column=2, value=total_stats.get('sire_unidentified', 0))

                # 外祖父识别数据
                self.ws.cell(row=chart_data_row, column=4, value="已识别")
                self.ws.cell(row=chart_data_row + 1, column=4, value="未识别")
                self.ws.cell(row=chart_data_row, column=5, value=total_stats.get('mgs_identified', 0))
                self.ws.cell(row=chart_data_row + 1, column=5, value=total_stats.get('mgs_unidentified', 0))

                # 外曾外祖父识别数据
                self.ws.cell(row=chart_data_row, column=7, value="已识别")
                self.ws.cell(row=chart_data_row + 1, column=7, value="未识别")
                self.ws.cell(row=chart_data_row, column=8, value=total_stats.get('mggs_identified', 0))
                self.ws.cell(row=chart_data_row + 1, column=8, value=total_stats.get('mggs_unidentified', 0))

                current_row += 4

                # 创建3个饼图（并排，每个10cm宽）
                # 父号识别饼图（A列开始）
                self._add_pie_chart(
                    data_range={
                        'labels': (chart_data_row, chart_data_row + 1, 1),
                        'values': (chart_data_row, chart_data_row + 1, 2)
                    },
                    position=f'A{current_row}',
                    title='父号识别情况'
                )

                # 外祖父识别饼图（E列开始，留出足够空间）
                self._add_pie_chart(
                    data_range={
                        'labels': (chart_data_row, chart_data_row + 1, 4),
                        'values': (chart_data_row, chart_data_row + 1, 5)
                    },
                    position=f'E{current_row}',
                    title='外祖父识别情况'
                )

                # 外曾外祖父识别饼图（I列开始，留出足够空间）
                self._add_pie_chart(
                    data_range={
                        'labels': (chart_data_row, chart_data_row + 1, 7),
                        'values': (chart_data_row, chart_data_row + 1, 8)
                    },
                    position=f'I{current_row}',
                    title='外曾外祖父识别情况'
                )

                current_row += 18  # 饼图高度

            # 4. 设置列宽（确保饼图不重叠）
            self._set_column_widths({
                1: 20,  # 出生年份/识别状态
                2: 15,  # 总头数/数量
                3: 18,  # 可识别父号牛数
                4: 18,  # 可识别父号占比
                5: 20,  # 可识别外祖父牛数
                6: 18,  # 可识别外祖父占比
                7: 22,  # 可识别外曾外祖父牛数
                8: 20,  # 可识别外曾外祖父占比
            })

            # 5. 冻结首行
            self._freeze_panes('A3')

            logger.info("✓ Sheet 2构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 2失败: {e}", exc_info=True)
            raise

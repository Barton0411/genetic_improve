"""
Sheet 5构建器: 配种记录-隐性基因分析
v1.2版本 - 聚焦配种记录中的隐性基因纯合风险分析
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class Sheet5Builder(BaseSheetBuilder):
    """
    Sheet 5: 配种记录-隐性基因分析

    包含内容:
    1. 隐性基因纯合汇总表（全部配种记录年份）
    2. 隐性基因纯合汇总表（近12个月配种记录）
    """

    def build(self, data: dict):
        """
        构建Sheet 5: 配种记录-隐性基因分析

        Args:
            data: 包含配种记录中隐性基因分析数据
                - all_years_summary: 全部年份隐性基因纯合汇总（包含所有16个基因）
                - recent_12m_summary: 近12个月隐性基因纯合汇总
                - date_range: 近12个月的日期范围
        """
        try:
            # 检查数据
            if not data or 'all_years_summary' not in data:
                logger.warning("Sheet5: 缺少数据，跳过生成")
                return

            # 创建Sheet
            self._create_sheet("配种记录-隐性基因分析")
            logger.info("构建Sheet 5: 配种记录-隐性基因分析")

            # 定义颜色
            self.color_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type='solid')      # RGB(255, 199, 206)
            self.color_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type='solid')   # RGB(255, 235, 156)
            self.color_orange = PatternFill(start_color="FFD966", end_color="FFD966", fill_type='solid')   # RGB(255, 217, 102)
            self.color_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type='solid')     # RGB(217, 217, 217)

            current_row = 1

            # 1. 构建表1：全部年份的汇总
            all_years_summary = data.get('all_years_summary', [])
            all_years_total = data.get('all_years_total', 0)
            if all_years_summary:
                title = f"隐性基因纯合汇总表（全部配种记录年份，共{all_years_total}配次）"
                current_row = self._build_gene_summary_table(
                    title=title,
                    summary_data=all_years_summary,
                    start_row=current_row
                )
                current_row += 3  # 空3行

            # 2. 构建表2：近12个月的汇总
            recent_12m_summary = data.get('recent_12m_summary', [])
            recent_12m_total = data.get('recent_12m_total', 0)
            date_range = data.get('date_range', {})
            if recent_12m_summary and date_range:
                # 格式化日期范围标题
                start_date = date_range.get('start', '')
                end_date = date_range.get('end', '')
                if start_date and end_date:
                    # 将YYYY-MM-DD转换为YYYY年MM月DD日
                    start_formatted = self._format_date_chinese(start_date)
                    end_formatted = self._format_date_chinese(end_date)
                    title = f"隐性基因纯合汇总表——{start_formatted}至{end_formatted}（共{recent_12m_total}配次）"
                else:
                    title = f"隐性基因纯合汇总表（近12个月，共{recent_12m_total}配次）"

                current_row = self._build_gene_summary_table(
                    title=title,
                    summary_data=recent_12m_summary,
                    start_row=current_row
                )

            # 设置列宽
            self._set_column_widths({
                1: 20,  # 基因名称
                2: 20,  # 翻译
                3: 12,  # 纯合配次
                4: 14,  # 占总配种比例
                5: 16,  # 仅公牛携带配次
                6: 12,  # 占比
                7: 20,  # 仅母牛父亲携带配次
                8: 12,  # 占比
                9: 14,  # 数据缺少配次
                10: 12,  # 占比
            })

            # 冻结首行
            self._freeze_panes('A2')

            logger.info("✓ Sheet 5构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 5失败: {e}", exc_info=True)
            raise

    def _build_gene_summary_table(self, title: str, summary_data: list, start_row: int) -> int:
        """
        构建隐性基因纯合汇总表

        Args:
            title: 表格标题
            summary_data: 汇总数据列表
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题行
        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头行
        headers = [
            "基因名称",
            "翻译",
            "纯合配次",
            "占总配种比例",
            "仅公牛携带配次",
            "占比",
            "仅母牛父亲携带配次",
            "占比",
            "数据缺少配次",
            "占比"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
        current_row += 1

        # 数据行
        for gene_data in summary_data:
            gene_name = gene_data.get('gene_name', '')
            gene_translation = gene_data.get('gene_translation', '')
            homozygous_count = gene_data.get('homozygous_count', 0)
            homozygous_ratio = gene_data.get('homozygous_ratio', 0)
            bull_only_count = gene_data.get('bull_only_count', 0)
            bull_only_ratio = gene_data.get('bull_only_ratio', 0)
            dam_sire_only_count = gene_data.get('dam_sire_only_count', 0)
            dam_sire_only_ratio = gene_data.get('dam_sire_only_ratio', 0)
            missing_data_count = gene_data.get('missing_data_count', 0)
            missing_data_ratio = gene_data.get('missing_data_ratio', 0)

            # 基因名称（左对齐，无背景色）
            cell = self.ws.cell(row=current_row, column=1, value=gene_name)
            self.style_manager.apply_data_style(cell, alignment='left')

            # 翻译（左对齐，无背景色）
            cell = self.ws.cell(row=current_row, column=2, value=gene_translation)
            self.style_manager.apply_data_style(cell, alignment='left')

            # 纯合配次（红色背景）
            cell = self.ws.cell(row=current_row, column=3, value=homozygous_count)
            cell.fill = self.color_red
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 纯合占比（红色背景，百分比格式）
            cell = self.ws.cell(row=current_row, column=4, value=f"{homozygous_ratio*100:.1f}%")
            cell.fill = self.color_red
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 仅公牛携带配次（黄色背景）
            cell = self.ws.cell(row=current_row, column=5, value=bull_only_count)
            cell.fill = self.color_yellow
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 仅公牛携带占比（黄色背景）
            cell = self.ws.cell(row=current_row, column=6, value=f"{bull_only_ratio*100:.1f}%")
            cell.fill = self.color_yellow
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 仅母牛父亲携带配次（橙色背景）
            cell = self.ws.cell(row=current_row, column=7, value=dam_sire_only_count)
            cell.fill = self.color_orange
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 仅母牛父亲携带占比（橙色背景）
            cell = self.ws.cell(row=current_row, column=8, value=f"{dam_sire_only_ratio*100:.1f}%")
            cell.fill = self.color_orange
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 数据缺少配次（灰色背景）
            cell = self.ws.cell(row=current_row, column=9, value=missing_data_count)
            cell.fill = self.color_gray
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            # 数据缺少占比（灰色背景）
            cell = self.ws.cell(row=current_row, column=10, value=f"{missing_data_ratio*100:.1f}%")
            cell.fill = self.color_gray
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.style_manager.border

            current_row += 1

        return current_row

    def _format_date_chinese(self, date_str: str) -> str:
        """
        将YYYY-MM-DD格式转换为YYYY年MM月DD日

        Args:
            date_str: 日期字符串 (YYYY-MM-DD)

        Returns:
            中文日期格式 (YYYY年MM月DD日)
        """
        try:
            parts = date_str.split('-')
            if len(parts) == 3:
                year, month, day = parts
                return f"{year}年{int(month)}月{int(day)}日"
            return date_str
        except Exception as e:
            logger.warning(f"日期格式转换失败: {date_str}, 错误: {e}")
            return date_str

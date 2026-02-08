"""
Sheet 10构建器: 备选公牛排名
v1.3版本 - 按育种指数排名的备选公牛列表，带技术标准对比和标红
"""

from .base_builder import BaseSheetBuilder
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import logging
import platform
from pathlib import Path
from datetime import datetime
from ..config.bull_quality_standards import (
    US_PROGENY_STANDARDS, US_GENOMIC_STANDARDS,
    DEFECT_GENES, QUALITY_STANDARD_TABLE
)

# 导入系统标准翻译字典
from core.breeding_calc.key_traits_page import TRAITS_TRANSLATION

logger = logging.getLogger(__name__)


# 基因翻译字典
GENE_TRANSLATIONS = {
    'HH1': '荷斯坦繁殖缺陷1型',
    'HH2': '荷斯坦繁殖缺陷2型',
    'HH3': '荷斯坦繁殖缺陷3型',
    'HH4': '荷斯坦繁殖缺陷4型',
    'HH5': '荷斯坦繁殖缺陷5型',
    'HH6': '荷斯坦繁殖缺陷6型',
    'MW': '早发肌无力',
    'HMW': '早发肌无力-单倍型',
    'BLAD': '牛白细胞粘附缺陷病',
    'Chondrodysplasia': '软骨发育异常',
    'Citrullinemia': '瓜氨酸血症',
    'DUMPS': '尿苷单磷酸合成酶缺乏',
    'Factor XI': '凝血因子XI缺乏',
    'CVM': '犊牛脊椎畸形综合征',
    'Brachyspina': '短脊椎综合症征',
    'Mulefoot': '单趾畸形',
    'Cholesterol deficiency': '胆固醇缺乏症'
}


def _build_header_map(columns: list) -> dict:
    """
    动态构建表头映射，使用系统标准翻译

    Args:
        columns: DataFrame的列名列表

    Returns:
        表头映射字典 {英文列名: 中英文表头}
    """
    header_map = {}

    # 固定列的翻译
    fixed_translations = {
        'ranking': '排名\nRanking',
        'bull_id': '公牛号\nBull ID',
        'semen_type': '精液类型\nSemen Type',
        '测试_index': '育种指数\nBreeding Index',
        '支数': '支数\nDoses'
    }

    for col in columns:
        if col in fixed_translations:
            # 固定列：使用预定义翻译
            header_map[col] = fixed_translations[col]
        elif col in GENE_TRANSLATIONS:
            # 基因列：使用基因翻译
            header_map[col] = f'{col}\n{GENE_TRANSLATIONS[col]}'
        elif col in TRAITS_TRANSLATION:
            # 性状列：使用系统标准翻译
            header_map[col] = f'{col}\n{TRAITS_TRANSLATION[col]}'
        else:
            # 未知列：保持原样
            header_map[col] = col

    return header_map


class Sheet10Builder(BaseSheetBuilder):
    """
    Sheet 10: 备选公牛排名

    包含内容:
    1. 按育种指数排名表
    2. 包含主要育种性状值
    """

    def __init__(self, workbook, style_manager, chart_builder, progress_callback=None, output_dir=None):
        """
        初始化Sheet10构建器

        Args:
            workbook: Excel工作簿对象
            style_manager: 样式管理器
            chart_builder: 图表构建器
            progress_callback: 进度回调函数
            output_dir: 图片输出目录（用于导出排名表图片）
        """
        super().__init__(workbook, style_manager, chart_builder, progress_callback)
        self.output_dir = output_dir

    def build(self, data: dict):
        """
        构建Sheet 10: 备选公牛排名

        Args:
            data: 包含备选公牛排名数据
                - bull_rankings: 公牛排名DataFrame（按ranking排序）
                - total_bulls: 公牛总数
                - sexed_bulls: 性控公牛数
                - regular_bulls: 常规公牛数
        """
        try:
            logger.info("构建Sheet 10: 备选公牛排名")

            # 检查数据
            if not data or 'bull_rankings' not in data:
                logger.warning("Sheet10: 缺少排名数据，跳过生成")
                return

            # 创建Sheet
            self._create_sheet("备选公牛排名")

            bull_rankings = data.get('bull_rankings')

            if bull_rankings is None or bull_rankings.empty:
                logger.warning("备选公牛排名数据为空，跳过构建")
                return

            current_row = 1

            # 1. 构建技术标准表
            current_row = self._build_quality_standards_table(current_row)
            current_row += 2  # 空2行

            # 2. 构建排名表
            current_row = self._build_ranking_table(current_row, bull_rankings, data)

            logger.info("✓ Sheet 10构建完成")

        except Exception as e:
            logger.error(f"Sheet 10构建失败: {e}", exc_info=True)
            raise

    def _build_quality_standards_table(self, start_row: int) -> int:
        """
        构建优质冻精技术标准表

        Args:
            start_row: 起始行号

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 标题
        title = "优质冻精技术标准"
        title_cell = self.ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # 表头 - 第一行
        headers = ['育种\n指标\n标识', '种\n标类型', 'TPI\n(GTPI)\n综合生产\n指数',
                   'NM$\n净效益\n指数', 'Milk\n奶量育\n种值\n(磅)', 'Prot蛋\n白量育\n种值\n(磅)',
                   'UDC\n乳房\n综合\n指数', 'FE\n饲料\n转化\n指数', 'PL\n生产\n寿命',
                   'DPR女\n儿怀孕\n率', '缺陷基因']

        header_start_row = current_row
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=header)
            self.style_manager.apply_header_style(cell)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:  # 育种指标标识
                self.ws.column_dimensions[col_letter].width = 8
            elif col_idx == 2:  # 种标类型
                self.ws.column_dimensions[col_letter].width = 8
            elif col_idx == 11:  # 缺陷基因
                self.ws.column_dimensions[col_letter].width = 15
            else:
                self.ws.column_dimensions[col_letter].width = 10

        current_row += 1

        # 数据行1: 美国 - 后裔
        row_data = ['美国', '后裔', '≥2850', '≥300', '≥200', '≥20', '≥0', '≥100', '≥1.5', '≥-2', 'HH1-HH6\nHMW']
        for col_idx, value in enumerate(row_data, 1):
            cell = self.ws.cell(row=current_row, column=col_idx, value=value)
            self.style_manager.apply_data_style(cell, alignment='center')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 第一行的"美国"单元格需要合并（跨2行）
        self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
        country_cell = self.ws.cell(row=current_row, column=1)
        country_cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # 数据行2: 基因组
        row_data = ['', '基因组', '≥2900', '≥400', '≥300', '≥30', '≥0', '≥100', '≥1.5', '≥-2', 'HH1-HH6\nHMW']
        for col_idx, value in enumerate(row_data[1:], 2):  # 跳过第一列（已合并）
            cell = self.ws.cell(row=current_row, column=col_idx, value=value)
            self.style_manager.apply_data_style(cell, alignment='center')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_row += 1

        logger.info("✓ 技术标准表构建完成")
        return current_row

    def _check_below_standard(self, trait_name: str, value, use_genomic: bool = False) -> tuple:
        """
        检查性状值是否低于标准

        Args:
            trait_name: 性状名称
            value: 性状值
            use_genomic: 是否使用基因组标准

        Returns:
            (below_progeny, below_genomic): 低于后裔标准, 低于基因组标准
        """
        # 处理NaN值
        if pd.isna(value):
            return (False, False)

        try:
            value = float(value)
        except (ValueError, TypeError):
            return (False, False)

        below_progeny = False
        below_genomic = False

        # 检查后裔标准
        if trait_name in US_PROGENY_STANDARDS:
            standard_value = US_PROGENY_STANDARDS[trait_name]
            below_progeny = value < standard_value

        # 检查基因组标准
        if trait_name in US_GENOMIC_STANDARDS:
            standard_value = US_GENOMIC_STANDARDS[trait_name]
            below_genomic = value < standard_value

        return (below_progeny, below_genomic)

    def _build_ranking_table(self, start_row: int, df: pd.DataFrame, summary: dict) -> int:
        """
        构建备选公牛排名表

        Args:
            start_row: 起始行号
            df: 排名DataFrame
            summary: 统计摘要

        Returns:
            下一个可用行号
        """
        current_row = start_row

        # 动态生成表头映射（使用系统标准翻译）
        header_map = _build_header_map(df.columns.tolist())

        # 标题
        title_cell = self.ws.cell(row=current_row, column=1, value="备选公牛排名")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        # 合并标题行
        num_cols = len(df.columns)
        self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        # 统计信息行
        stats_text = f"总计: {summary.get('total_bulls', 0)}头公牛 (性控: {summary.get('sexed_bulls', 0)}, 常规: {summary.get('regular_bulls', 0)})"
        stats_cell = self.ws.cell(row=current_row, column=1, value=stats_text)
        stats_cell.font = Font(size=11, italic=True)
        stats_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

        # 表头 - 使用动态生成的中英文映射
        for col_idx, col_name in enumerate(df.columns, 1):
            # 获取中英文表头（动态翻译）
            header_text = header_map.get(col_name, col_name)
            cell = self.ws.cell(row=current_row, column=col_idx, value=header_text)
            self.style_manager.apply_header_style(cell)
            # 设置自动换行以显示多行表头
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if 'bull_id' in col_name:
                self.ws.column_dimensions[col_letter].width = 16
            elif 'semen_type' in col_name:
                self.ws.column_dimensions[col_letter].width = 14
            elif col_name == 'ranking':
                self.ws.column_dimensions[col_letter].width = 12
            elif col_name == '测试_index':
                self.ws.column_dimensions[col_letter].width = 16
            elif col_name == '支数':
                self.ws.column_dimensions[col_letter].width = 10
            elif col_name in DEFECT_GENES:
                self.ws.column_dimensions[col_letter].width = 10
            else:
                self.ws.column_dimensions[col_letter].width = 14

        current_row += 1
        header_row = current_row - 1

        # 数据行
        for _, row_data in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, 1):
                value = row_data[col_name]

                # 处理NaN
                if pd.isna(value):
                    value = ""

                cell = self.ws.cell(row=current_row, column=col_idx, value=value)

                # 默认应用格式
                self.style_manager.apply_data_style(cell, alignment='center')

                # 特殊格式
                if col_name == 'ranking':
                    # 排名列加粗
                    cell.font = Font(bold=True, size=11)
                elif col_name == '测试_index':
                    # 育种指数保留2位小数
                    cell.number_format = '0.00'
                elif col_name in ['DPR', 'PROT%', 'FAT %']:
                    # 浮点数保留2位小数
                    cell.number_format = '0.00'

                # 标色逻辑
                should_red = False
                should_pink = False

                # 1. 检查性状值是否低于标准
                if col_name in US_PROGENY_STANDARDS:
                    below_progeny, below_genomic = self._check_below_standard(col_name, value)
                    if below_progeny:
                        # 低于后裔标准，标红
                        should_red = True
                    elif below_genomic:
                        # 高于后裔但低于基因组，标粉色
                        should_pink = True

                # 2. 检查基因是否携带（C表示携带）
                if col_name in DEFECT_GENES:
                    if value == 'C':
                        should_red = True

                # 应用颜色填充
                if should_red:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)  # 白色文字，加粗
                elif should_pink:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    cell.font = Font(color="000000", bold=True)  # 黑色文字，加粗

            current_row += 1

        # 冻结首行
        self.ws.freeze_panes = self.ws.cell(row=header_row + 1, column=1)

        logger.info(f"✓ Sheet 10排名表构建完成: {len(df)}头公牛")

        # 导出排名表为图片（供PPT使用）
        if self.output_dir:
            self._export_ranking_table_image(df, summary, header_map)

        return current_row

    def _export_ranking_table_image(self, df: pd.DataFrame, summary: dict, header_map: dict):
        """
        使用matplotlib将排名表导出为图片

        Args:
            df: 排名DataFrame
            summary: 统计摘要
            header_map: 表头映射
        """
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            matplotlib.use('Agg')  # 使用非交互式后端

            # 设置中文字体（Mac和Windows兼容）
            system = platform.system()
            if system == 'Darwin':  # macOS
                font_candidates = ['PingFang SC', 'Heiti SC', 'STHeiti', 'Arial Unicode MS']
            else:  # Windows
                font_candidates = ['Microsoft YaHei', 'SimHei', 'SimSun', 'Arial Unicode MS']

            # 查找可用字体
            from matplotlib import font_manager
            available_fonts = [f.name for f in font_manager.fontManager.ttflist]
            selected_font = None
            for font in font_candidates:
                if font in available_fonts:
                    selected_font = font
                    break

            if selected_font:
                plt.rcParams['font.sans-serif'] = [selected_font]
            plt.rcParams['axes.unicode_minus'] = False

            # 准备表格数据
            # 表头（使用中英文映射）
            headers = [header_map.get(col, col) for col in df.columns]

            # 数据行
            table_data = []
            cell_colors = []

            for _, row in df.iterrows():
                row_data = []
                row_colors = []
                for col in df.columns:
                    value = row[col]
                    if pd.isna(value):
                        value = ""
                    row_data.append(str(value) if not isinstance(value, str) else value)

                    # 确定单元格颜色
                    color = 'white'
                    if col in US_PROGENY_STANDARDS:
                        below_progeny, below_genomic = self._check_below_standard(col, row[col])
                        if below_progeny:
                            color = '#FF0000'  # 红色
                        elif below_genomic:
                            color = '#FFB6C1'  # 粉色
                    if col in DEFECT_GENES and row[col] == 'C':
                        color = '#FF0000'  # 红色

                    row_colors.append(color)

                table_data.append(row_data)
                cell_colors.append(row_colors)

            # 计算图片尺寸
            num_cols = len(df.columns)
            num_rows = len(df) + 1  # +1 for header
            col_width = 1.2  # 每列宽度（英寸）
            row_height = 0.4  # 每行高度（英寸）

            fig_width = max(num_cols * col_width, 12)
            fig_height = max(num_rows * row_height + 1, 4)

            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            ax.axis('off')

            # 添加标题
            title_text = f"备选公牛排名\n总计: {summary.get('total_bulls', 0)}头公牛 (性控: {summary.get('sexed_bulls', 0)}, 常规: {summary.get('regular_bulls', 0)})"
            ax.set_title(title_text, fontsize=14, fontweight='bold', pad=20)

            # 创建表格
            table = ax.table(
                cellText=table_data,
                colLabels=headers,
                cellColours=cell_colors,
                colColours=['#4472C4'] * num_cols,  # 表头蓝色
                cellLoc='center',
                loc='center'
            )

            # 设置表格样式
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.2, 1.5)

            # 设置表头文字颜色为白色
            for j in range(num_cols):
                cell = table[(0, j)]
                cell.set_text_props(color='white', fontweight='bold')

            # 设置红色单元格的文字为白色
            for i, row_colors in enumerate(cell_colors):
                for j, color in enumerate(row_colors):
                    if color == '#FF0000':
                        cell = table[(i + 1, j)]
                        cell.set_text_props(color='white', fontweight='bold')

            plt.tight_layout()

            # 保存图片
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(self.output_dir) / f"备选公牛排名表_{timestamp}.png"
            plt.savefig(output_path, dpi=100, bbox_inches='tight',
                       facecolor='white', edgecolor='none')
            plt.close(fig)

            logger.info(f"✓ 排名表图片已导出: {output_path.name}")

        except Exception as e:
            logger.warning(f"导出排名表图片失败: {e}")
            logger.debug("错误详情:", exc_info=True)

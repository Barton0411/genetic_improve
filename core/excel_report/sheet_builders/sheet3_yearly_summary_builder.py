"""
Sheet 3-1构建器: 年份汇总与性状进展
"""

from .base_builder import BaseSheetBuilder
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter
import pandas as pd
import logging
import re

logger = logging.getLogger(__name__)


class Sheet3YearlySummaryBuilder(BaseSheetBuilder):
    """Sheet 3-1: 年份汇总与性状进展"""

    def build(self, data: dict):
        """
        构建Sheet 3-1

        Args:
            data: {
                'present_summary': 在群母牛年份汇总DataFrame,
                'all_summary': 全部母牛年份汇总DataFrame,
                'comparison_data': {
                    'farms': [对比牧场数据列表],
                    'references': [外部参考数据列表，仅用于折线图]
                }
            }
        """
        try:
            logger.info("开始构建Sheet 3-1: 年份汇总与性状进展")

            # 创建Sheet
            try:
                self._create_sheet("年份汇总与性状进展")
                logger.info("✓ Sheet创建成功")
            except Exception as e:
                logger.error(f"✗ 创建Sheet失败: {e}", exc_info=True)
                raise

            present_df = data.get('present_summary')
            all_df = data.get('all_summary')
            comparison_data = data.get('comparison_data', {'farms': [], 'references': []})
            comparison_farms = comparison_data.get('farms', [])

            if present_df is None or present_df.empty:
                logger.warning("Sheet 3-1: 在群母牛数据为空，跳过构建")
                return

            current_row = 1

            # === 1. 左侧：在群母牛年份汇总表 ===
            cell = self.ws.cell(row=current_row, column=1, value="在群母牛年份汇总")
            self.style_manager.apply_title_style(cell)
            current_row += 1

            # 写入表头
            headers_present = list(present_df.columns)
            self._write_header(current_row, headers_present, start_col=1)
            current_row += 1

            # 写入数据
            for idx, row in present_df.iterrows():
                values = list(row)
                # 如果是总计行，使用合计行样式
                if '总计' in str(values[0]):
                    self._write_total_row(current_row, values, start_col=1)
                else:
                    self._write_data_row(current_row, values, start_col=1, alignment='center')
                current_row += 1

            # 添加对比牧场总计行（如果有）
            if comparison_farms:
                for farm in comparison_farms:
                    # 构建对比牧场行：智能匹配性状
                    farm_row = self._build_comparison_farm_row(
                        farm, headers_present, data_type='present'
                    )
                    self._write_data_row(current_row, farm_row, start_col=1, alignment='center')
                    current_row += 1

            # === 2. 右侧：全部母牛年份汇总表（间隔6列） ===
            if all_df is not None and not all_df.empty:
                right_start_col = len(headers_present) + 1 + 6  # 左表列数 + 1 + 空6列

                current_row_right = 1
                cell = self.ws.cell(row=current_row_right, column=right_start_col, value="全部母牛年份汇总")
                self.style_manager.apply_title_style(cell)
                current_row_right += 1

                # 写入表头
                headers_all = list(all_df.columns)
                self._write_header(current_row_right, headers_all, start_col=right_start_col)
                current_row_right += 1

                # 写入数据
                for idx, row in all_df.iterrows():
                    values = list(row)
                    if '总计' in str(values[0]):
                        self._write_total_row(current_row_right, values, start_col=right_start_col)
                    else:
                        self._write_data_row(current_row_right, values, start_col=right_start_col, alignment='center')
                    current_row_right += 1

                # 添加对比牧场总计行（如果有）
                if comparison_farms:
                    for farm in comparison_farms:
                        # 构建对比牧场行：智能匹配性状
                        farm_row = self._build_comparison_farm_row(
                            farm, headers_all, data_type='all'
                        )
                        self._write_data_row(current_row_right, farm_row, start_col=right_start_col, alignment='center')
                        current_row_right += 1

            # === 3. 在最右侧写入对比数据（用于折线图） ===
            logger.info(f"开始写入对比数据，farms: {len(comparison_data.get('farms', []))}, references: {len(comparison_data.get('references', []))}")
            chart_data_start_col = self._write_comparison_chart_data(
                present_df, headers_present, comparison_data
            )
            logger.info(f"对比数据写入完成，起始列: {chart_data_start_col}")

            # === 4. 下方：性状进展折线图（3个一排） ===
            logger.info("开始创建性状进展折线图...")
            chart_start_row = max(current_row, current_row_right if 'current_row_right' in locals() else current_row) + 2

            # 获取需要绘制折线图的性状（所有"平均"开头的列）
            trait_columns = [col for col in headers_present if col.startswith('平均')]

            # 每排3个图表
            charts_per_row = 3
            # 图表间隔设置（不影响图表实际显示尺寸）
            col_gap = 2  # 图表横向间隔2列
            row_gap = 2  # 图表纵向间隔2行

            # 绘制所有性状的折线图
            # 每个图表实际占用空间：横10列 × 竖20行（用于定位）
            chart_col_span = 10
            chart_row_span = 20

            logger.info(f"准备创建 {len(trait_columns)} 个性状折线图")
            for idx, trait_col in enumerate(trait_columns):
                row_idx = idx // charts_per_row
                col_idx = idx % charts_per_row

                chart_row = chart_start_row + row_idx * (chart_row_span + row_gap)
                chart_col = 1 + col_idx * (chart_col_span + col_gap)

                # 创建折线图（传入对比数据和数据起始列）
                trait_name = trait_col.replace('平均', '')
                logger.info(f"创建第 {idx+1}/{len(trait_columns)} 个图表: {trait_name}")
                self._create_trait_trend_chart(
                    present_df, trait_col, trait_name,
                    chart_row, chart_col, chart_row_span, chart_col_span,
                    comparison_data, chart_data_start_col  # 传入对比数据和数据列位置
                )

            # 设置列宽
            for col in range(1, right_start_col + len(headers_all) if 'right_start_col' in locals() else len(headers_present) + 1):
                self.ws.column_dimensions[get_column_letter(col)].width = 12

            # 冻结首行
            self._freeze_panes('A2')

            logger.info(f"✓ Sheet 3-1构建完成")

        except Exception as e:
            logger.error(f"构建Sheet 3-1失败: {e}", exc_info=True)
            raise

    def _create_trait_trend_chart(self, df, trait_col, trait_name, chart_row, chart_col,
                                  row_span, col_span, comparison_data: dict, chart_data_start_col: int = None):
        """
        创建性状进展折线图（支持多条对比线）

        Args:
            df: 当前牧场数据DataFrame
            trait_col: 性状列名 (e.g., '平均NM$')
            trait_name: 性状显示名 (e.g., 'NM$')
            chart_row, chart_col: 图表位置
            row_span, col_span: 图表占用空间
            comparison_data: 对比数据 {'farms': [...], 'references': [...]}
        """
        try:
            # 过滤当前牧场有效数据（排除总计行和对比牧场行）
            valid_df = df[~df['出生年份'].str.contains('总计|对比', na=False)]
            if len(valid_df) == 0:
                logger.warning(f"性状 {trait_name} 没有有效数据，跳过图表创建")
                return

            # 创建折线图
            chart = LineChart()
            chart.title = f"{trait_name}性状进展"
            chart.style = 13
            chart.y_axis.title = trait_name
            chart.x_axis.title = "出生年份"

            # 收集所有数据值（用于Y轴范围计算）
            all_values = []

            # === 1. 添加当前牧场数据系列（实线） ===
            trait_values = valid_df[trait_col].dropna()
            if len(trait_values) == 0:
                logger.warning(f"性状 {trait_name} 所有值都为空，跳过图表创建")
                return

            all_values.extend(trait_values.tolist())

            # 数据范围（假设数据从第3行开始）
            data_col = df.columns.get_loc(trait_col) + 1
            data = Reference(self.ws, min_col=data_col,
                           min_row=3, max_row=3 + len(valid_df) - 1)
            cats = Reference(self.ws, min_col=1, min_row=3, max_row=3 + len(valid_df) - 1)

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            # 设置主系列名称为"当前牧场"（不设置名称，使用默认）
            # openpyxl的series.title需要SeriesLabel对象，这里不设置了

            # 获取当前牧场的年份列表（用于对齐）
            current_years = valid_df['出生年份'].tolist()

            # === 2. 添加对比系列（从已写入的数据中读取） ===
            if chart_data_start_col:
                try:
                    # 查找当前性状的对比数据列
                    # 列标题格式："{牧场/参考名}_{性状列名}"
                    for col_idx in range(chart_data_start_col + 1, self.ws.max_column + 1):
                        col_header = self.ws.cell(row=2, column=col_idx).value
                        if not col_header:
                            break

                        # 检查是否是当前性状的列
                        if not col_header.endswith(f"_{trait_col}"):
                            continue

                        # 提取系列名称（去掉性状后缀）
                        series_name = col_header.replace(f"_{trait_col}", "")

                        # 检查该列是否有数据
                        has_data = False
                        for row_idx in range(3, 3 + len(current_years)):
                            if self.ws.cell(row=row_idx, column=col_idx).value is not None:
                                has_data = True
                                # 添加到all_values用于Y轴范围计算
                                all_values.append(self.ws.cell(row=row_idx, column=col_idx).value)

                        if not has_data:
                            continue

                        # 创建数据引用
                        data_ref = Reference(self.ws, min_col=col_idx,
                                           min_row=2, max_row=2 + len(current_years))

                        # 添加系列
                        series = Series(data_ref, title_from_data=True)

                        # 设置虚线样式
                        line_props = LineProperties()
                        line_props.dashStyle = "dash"

                        # 查找对应的颜色
                        series_color = None
                        for farm in comparison_data.get('farms', []):
                            if farm.get('name') == series_name:
                                series_color = farm.get('color', '#95a5a6')
                                break
                        if not series_color:
                            for ref in comparison_data.get('references', []):
                                if ref.get('name') == series_name:
                                    series_color = ref.get('color', '#e67e22')
                                    break

                        if series_color:
                            rgb_color = series_color.lstrip('#').upper()
                            line_props.solidFill = rgb_color

                        series.graphicalProperties.line = line_props

                        # 设置标记点
                        marker = Marker('circle')
                        marker.size = 5
                        if series_color:
                            marker.graphicalProperties.solidFill = rgb_color
                        series.marker = marker

                        chart.series.append(series)
                        logger.debug(f"  添加对比系列: {series_name}")

                except Exception as e:
                    logger.error(f"添加对比系列失败: {e}", exc_info=True)

            # === 4. 自动调整Y轴范围 ===
            if all_values:
                min_val = min(all_values)
                max_val = max(all_values)
                value_range = max_val - min_val
                if value_range > 0:
                    chart.y_axis.scaling.min = min_val - value_range * 0.1
                    chart.y_axis.scaling.max = max_val + value_range * 0.1

            # === 5. 设置图表大小和显示 ===
            chart.width = 20  # 20cm宽
            chart.height = 10  # 10cm高

            # 显示图例
            chart.legend.position = 'r'  # 图例在右侧

            # 添加图表到sheet
            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"
            self.ws.add_chart(chart, anchor_cell)

            logger.info(f"✓ 创建折线图成功: {trait_name}，共{len(chart.series)}个系列")

        except Exception as e:
            logger.error(f"✗ 创建折线图失败 {trait_name}: {e}", exc_info=True)
            import traceback
            logger.error(traceback.format_exc())

    def _write_comparison_chart_data(self, present_df, headers: list, comparison_data: dict) -> int:
        """
        在sheet最右侧写入对比数据（用于折线图）

        Args:
            present_df: 在群母牛DataFrame
            headers: 表头列表
            comparison_data: 对比数据字典

        Returns:
            数据起始列号（如果没有对比数据返回None）
        """
        try:
            # 检查是否有对比数据
            if not comparison_data:
                logger.info("没有对比数据，跳过写入")
                return None

            comparison_farms = comparison_data.get('farms', [])
            references = comparison_data.get('references', [])

            if not comparison_farms and not references:
                logger.info("没有勾选任何对比牧场或外部参考数据")
                return None

            # 过滤出有效年份数据（排除总计行）
            valid_df = present_df[~present_df['出生年份'].str.contains('总计|对比', na=False)]
            if len(valid_df) == 0:
                logger.warning("没有有效年份数据")
                return None

            current_years = valid_df['出生年份'].tolist()

            # 确定数据写入位置（表格最右侧 + 5列空白）
            data_start_col = len(headers) * 2 + 10

            # 写入标题行（第1行）
            self.ws.cell(row=1, column=data_start_col, value="对比数据（用于图表）")

            # 写入年份列（第2行标题，第3行开始数据）
            # 去掉"及以前"，只保留年份
            self.ws.cell(row=2, column=data_start_col, value="年份")
            clean_years = []  # 清理后的年份（用于对比数据）
            for i, year in enumerate(current_years):
                # 提取纯年份（去掉"及以前"）
                clean_year = year
                if '及以前' in year:
                    match = re.search(r'(\d{4})年', year)
                    if match:
                        clean_year = f"{match.group(1)}年"

                self.ws.cell(row=3 + i, column=data_start_col, value=clean_year)
                clean_years.append(clean_year)

            current_col = data_start_col + 1

            # 写入对比牧场数据（每个牧场一列，所有性状）
            for farm in comparison_farms:
                farm_name = farm.get('name', '对比牧场')
                farm_year_data = farm.get('year_data', {})  # 年份数据字典
                farm_year_rows = farm.get('year_rows', [])  # 年份列表

                # 为每个性状列创建一列数据
                for trait_col in headers:
                    if not trait_col.startswith('平均'):
                        continue

                    # 写入列标题（包含牧场名和性状）
                    self.ws.cell(row=2, column=current_col, value=f"{farm_name}_{trait_col}")

                    # 写入数据（按年份对齐，智能匹配）
                    for i, clean_year in enumerate(clean_years):
                        value = None

                        # 1. 尝试精确匹配纯年份
                        year_values = farm_year_data.get(clean_year, {})
                        value = year_values.get(trait_col)

                        # 2. 如果没匹配到，尝试匹配"xxxx年及以前"
                        if value is None:
                            year_with_before = f"{clean_year.replace('年', '')}年及以前"
                            year_values = farm_year_data.get(year_with_before, {})
                            value = year_values.get(trait_col)

                        if value is not None:
                            self.ws.cell(row=3 + i, column=current_col, value=value)

                    current_col += 1

            # 写入外部参考数据（每个参考源一列，所有性状）
            for ref in references:
                ref_name = ref.get('name', '外部参考')
                year_data = ref.get('year_data', {})
                ref_traits = ref.get('traits', [])

                # 为每个性状列创建一列数据
                for trait_col in headers:
                    if not trait_col.startswith('平均'):
                        continue

                    if trait_col not in ref_traits:
                        continue

                    # 写入列标题
                    self.ws.cell(row=2, column=current_col, value=f"{ref_name}_{trait_col}")

                    # 写入数据（按年份对齐，智能匹配）
                    for i, clean_year in enumerate(clean_years):
                        value = None

                        # 1. 尝试精确匹配纯年份
                        year_values = year_data.get(clean_year, {})
                        value = year_values.get(trait_col)

                        # 2. 如果没匹配到，尝试匹配"xxxx年及以前"
                        if value is None:
                            year_with_before = f"{clean_year.replace('年', '')}年及以前"
                            year_values = year_data.get(year_with_before, {})
                            value = year_values.get(trait_col)

                        if value is not None:
                            self.ws.cell(row=3 + i, column=current_col, value=value)

                    current_col += 1

            logger.info(f"✓ 对比数据已写入sheet，起始列: {data_start_col}")
            return data_start_col

        except Exception as e:
            logger.error(f"写入对比数据失败: {e}", exc_info=True)
            return None

    def _build_comparison_farm_row(self, farm: dict, headers: list, data_type: str) -> list:
        """
        构建对比牧场行，智能匹配性状

        Args:
            farm: 对比牧场数据字典
                {
                    'name': '测试1',
                    'latest_year': 2024,
                    'present_cow_count': 665,
                    'all_cow_count': 800,
                    'present_data': {'平均NM$': 900, '平均TPI': 2600, ...},
                    'all_data': {'平均NM$': 875, '平均TPI': 2550, ...}
                }
            headers: 表头列表 ['出生年份', '头数', '平均NM$', ...]
            data_type: 'present' 或 'all'

        Returns:
            行数据列表，与headers对齐
        """
        # 选择对应的数据源
        farm_data = farm.get('present_data', {}) if data_type == 'present' else farm.get('all_data', {})
        # 构建行数据
        row = []
        for col_name in headers:
            if col_name == '出生年份':
                # 第一列：显示"牧场名 (年份范围)"
                name = farm.get('name', '对比牧场')

                # 优先根据年份列表计算年份范围（最小-最大）
                year_range_str = None
                year_rows = farm.get('year_rows') or []
                if year_rows:
                    # 过滤掉“总计”等非年份行
                    actual_years = [y for y in year_rows if '总计' not in str(y)]
                    if actual_years:
                        first_year = actual_years[0]
                        last_year = actual_years[-1]

                        # 提取阿拉伯数字年份，例如“2021年及以前” -> 2021
                        match_first = re.search(r'(\d{4})', str(first_year))
                        match_last = re.search(r'(\d{4})', str(last_year))

                        if match_first and match_last:
                            # 形如“2020年 - 2024年”
                            year_range_str = f"{match_first.group(1)}年 - {match_last.group(1)}年"

                # 如果无法解析年份范围，则退回到最后一年
                if not year_range_str:
                    year = farm.get('latest_year')
                    if year:
                        year_range_str = f"{year}年"

                if year_range_str:
                    row.append(f"{name} ({year_range_str})")
                else:
                    row.append(name)
            elif col_name == '头数':
                # 头数列：根据data_type选择对应的头数
                if data_type == 'present':
                    cow_count = farm.get('present_cow_count', 0)
                else:  # 'all'
                    cow_count = farm.get('all_cow_count', 0)
                row.append(cow_count if cow_count > 0 else '-')
            else:
                # 性状列：智能匹配
                value = farm_data.get(col_name)
                if value is not None:
                    row.append(value)
                else:
                    row.append('-')

        return row

"""
样式管理器
统一管理Excel报告的所有样式
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class StyleManager:
    """统一管理Excel样式"""

    # 颜色定义
    COLOR_HEADER_BG = "2E5C8A"      # 深蓝色
    COLOR_HEADER_TEXT = "FFFFFF"     # 白色
    COLOR_ROW_EVEN = "F5F5F5"        # 浅灰色
    COLOR_TOTAL_BG = "D6E9F8"        # 浅蓝色
    COLOR_RISK_HIGH = "FFE6E6"       # 红色背景
    COLOR_RISK_MEDIUM = "FFF9E6"     # 黄色背景
    COLOR_RISK_LOW = "E6F9E6"        # 绿色背景

    def __init__(self):
        self._init_styles()

    def _init_styles(self):
        """初始化常用样式"""
        # 标题行样式（12号字体）
        self.header_font = Font(
            name='微软雅黑',
            size=12,
            bold=True,
            color=self.COLOR_HEADER_TEXT
        )
        self.header_fill = PatternFill(
            start_color=self.COLOR_HEADER_BG,
            end_color=self.COLOR_HEADER_BG,
            fill_type='solid'
        )
        self.header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        # 数据行样式（10号字体）
        self.data_font = Font(name='微软雅黑', size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='center')
        self.data_alignment_center = Alignment(horizontal='center', vertical='center')
        self.data_alignment_right = Alignment(horizontal='right', vertical='center')

        # 合计行样式（10号字体）
        self.total_font = Font(name='微软雅黑', size=10, bold=True)
        self.total_fill = PatternFill(
            start_color=self.COLOR_TOTAL_BG,
            end_color=self.COLOR_TOTAL_BG,
            fill_type='solid'
        )

        # 标题样式（一级标题，如"一、基本信息"，14号字体）
        self.title_font = Font(
            name='微软雅黑',
            size=14,
            bold=True,
            color="2E5C8A"  # 深蓝色
        )
        self.title_alignment = Alignment(
            horizontal='left',
            vertical='center'
        )

        # 边框
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )

    def apply_title_style(self, cell):
        """应用一级标题样式（如"一、基本信息"）"""
        cell.font = self.title_font
        cell.alignment = self.title_alignment

    def apply_header_style(self, cell):
        """应用表格标题行样式"""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border

    def apply_data_style(self, cell, alignment='left'):
        """应用数据样式"""
        cell.font = self.data_font
        cell.border = self.border
        if alignment == 'center':
            cell.alignment = self.data_alignment_center
        elif alignment == 'right':
            cell.alignment = self.data_alignment_right
        else:
            cell.alignment = self.data_alignment

    def apply_total_style(self, cell):
        """应用合计行样式"""
        cell.font = self.total_font
        cell.fill = self.total_fill
        cell.border = self.border
        cell.alignment = self.data_alignment_center

    def apply_basic_info_label_style(self, cell):
        """应用基本信息表标签列样式（深蓝色背景，白字，12号字体）"""
        cell.font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border

    def apply_basic_info_value_style(self, cell):
        """应用基本信息表值列样式（浅蓝色背景，10号字体）"""
        cell.font = Font(name='微软雅黑', size=10)
        cell.fill = PatternFill(start_color="D6E9F8", end_color="D6E9F8", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border

    def apply_risk_color(self, cell, risk_level: str):
        """
        应用风险等级颜色

        Args:
            cell: 单元格对象
            risk_level: 风险等级 ('high', 'medium', 'low')
        """
        if risk_level == 'high':
            cell.fill = PatternFill(
                start_color=self.COLOR_RISK_HIGH,
                end_color=self.COLOR_RISK_HIGH,
                fill_type='solid'
            )
        elif risk_level == 'medium':
            cell.fill = PatternFill(
                start_color=self.COLOR_RISK_MEDIUM,
                end_color=self.COLOR_RISK_MEDIUM,
                fill_type='solid'
            )
        elif risk_level == 'low':
            cell.fill = PatternFill(
                start_color=self.COLOR_RISK_LOW,
                end_color=self.COLOR_RISK_LOW,
                fill_type='solid'
            )

    def get_risk_fill(self, risk_level: str) -> PatternFill:
        """
        获取风险等级填充样式

        Args:
            risk_level: 风险等级

        Returns:
            PatternFill对象
        """
        color_map = {
            'high': self.COLOR_RISK_HIGH,
            'medium': self.COLOR_RISK_MEDIUM,
            'low': self.COLOR_RISK_LOW,
        }
        color = color_map.get(risk_level, 'FFFFFF')
        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )

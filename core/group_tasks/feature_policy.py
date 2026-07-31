"""牧场组“按页面参数逐场分析”的独立产物清单。

这里的功能清单与 ``data -> analysis -> child_excel`` 主报告阶段完全
分离。页面内批量计算只提交本功能的结果，不能把整个牧场误标成已经
完成自动报告，也不能直接触发牧场组汇总。
"""

from __future__ import annotations

import hashlib
import fnmatch
import math
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from core.group_tasks.stage_manifest import (
    commit_stage_manifest,
    compute_xlsx_identifier_multiset,
    validate_stage_manifest,
)


FEATURE_POLICY_REVISION = 1
_VERSION_UNSET = object()
FEATURE_MANIFEST_DIRECTORY = (
    Path("group_store") / "stage_manifests" / "features"
)

TRAIT_OPERATIONS = {
    "cow_traits",
    "bull_traits",
    "mated_bull_traits",
}
INDEX_OPERATIONS = {"cow_index", "bull_index"}
INBREEDING_OPERATIONS = {
    "cow_self_inbreeding",
    "mated_inbreeding",
    "candidate_inbreeding",
}
SUPPORTED_FEATURE_OPERATIONS = (
    TRAIT_OPERATIONS | INDEX_OPERATIONS | INBREEDING_OPERATIONS
)

FEATURE_TITLES = {
    "cow_traits": "在群母牛关键育种性状分析",
    "bull_traits": "备选公牛关键育种性状分析",
    "mated_bull_traits": "已配公牛关键育种性状分析",
    "cow_index": "母牛群指数排名",
    "bull_index": "备选公牛指数排名",
    "cow_self_inbreeding": "母牛近交系数及隐性基因分析",
    "mated_inbreeding": "已配公牛近交系数及隐性基因分析",
    "candidate_inbreeding": "备选公牛近交系数及隐性基因分析",
}

_FEATURE_INPUTS = {
    "cow_traits": (
        Path("standardized_data") / "processed_cow_data.xlsx",
    ),
    "bull_traits": (
        Path("standardized_data") / "processed_bull_data.xlsx",
    ),
    "mated_bull_traits": (
        Path("standardized_data") / "processed_breeding_data.xlsx",
    ),
    "cow_index": (
        Path("standardized_data") / "processed_cow_data.xlsx",
    ),
    "bull_index": (
        Path("standardized_data") / "processed_bull_data.xlsx",
    ),
    "cow_self_inbreeding": (
        Path("standardized_data") / "processed_cow_data.xlsx",
    ),
    "mated_inbreeding": (
        Path("standardized_data") / "processed_cow_data.xlsx",
        Path("standardized_data") / "processed_breeding_data.xlsx",
    ),
    "candidate_inbreeding": (
        Path("standardized_data") / "processed_cow_data.xlsx",
        Path("standardized_data") / "processed_bull_data.xlsx",
    ),
}

_FIXED_FEATURE_OUTPUTS = {
    "cow_traits": (
        Path("analysis_results")
        / "processed_cow_data_key_traits_final.xlsx",
        Path("analysis_results") / "关键育种性状分析结果.xlsx",
        Path("analysis_results") / "系谱识别分析结果.xlsx",
        Path("analysis_results")
        / "sire_traits_mean_by_cow_birth_year.xlsx",
    ),
    "bull_traits": (
        Path("analysis_results") / "processed_bull_data_key_traits.xlsx",
    ),
    "mated_bull_traits": (
        Path("analysis_results") / "processed_mated_bull_traits.xlsx",
    ),
    "cow_index": (
        Path("analysis_results") / "processed_index_cow_index_scores.xlsx",
    ),
    "bull_index": (
        Path("analysis_results") / "processed_index_bull_scores.xlsx",
    ),
    "cow_self_inbreeding": (
        Path("analysis_results") / "母牛近交系数分析结果.xlsx",
    ),
}

_LATEST_OUTPUT_PATTERNS = {
    "mated_inbreeding": "已配公牛_近交系数及隐性基因分析结果_*.xlsx",
    "candidate_inbreeding": "备选公牛_近交系数及隐性基因分析结果_*.xlsx",
}

# 完整报告 analysis 阶段允许“有提示完成”；母牛性状的核心证据与
# ``stage_policy.REQUIRED_ANALYSIS_OUTPUTS`` 保持一致，不把按出生年
# 父号均值这类附加明细误当成阻塞条件。
_REPORT_OPERATION_FIXED_OUTPUTS = {
    **_FIXED_FEATURE_OUTPUTS,
    "cow_traits": (
        Path("analysis_results")
        / "processed_cow_data_key_traits_final.xlsx",
        Path("analysis_results") / "关键育种性状分析结果.xlsx",
        Path("analysis_results") / "系谱识别分析结果.xlsx",
    ),
}

_COW_INDEX_SCORE_INPUTS = (
    Path("analysis_results")
    / "processed_cow_data_key_traits_scores_genomic.xlsx",
    Path("analysis_results")
    / "processed_cow_data_key_traits_scores_pedigree.xlsx",
)

_COW_ID_COLUMNS = ("cow_id", "母牛号", "牛号", "耳号")
_BULL_ID_COLUMNS = ("bull_id", "公牛号", "NAAB", "BULL NAAB")
_MISSING_IDENTIFIERS = {"", "nan", "none", "null", "nat", "<na>", "n/a"}
_IDENTIFIER_HASH_MODULUS = 1 << 256


class FeaturePolicyError(RuntimeError):
    """页面参数、直接输入或功能产物不满足提交要求。"""


def feature_manifest_path(operation: str) -> Path:
    if operation not in SUPPORTED_FEATURE_OPERATIONS:
        raise ValueError(f"不支持的牧场组分析功能: {operation}")
    return FEATURE_MANIFEST_DIRECTORY / f"{operation}.json"


def normalize_feature_parameters(
    operation: str,
    parameters: Mapping[str, Any] | None,
) -> dict[str, Any]:
    """严格规范页面参数，拒绝请求中夹带其它字段。"""
    if operation not in SUPPORTED_FEATURE_OPERATIONS:
        raise FeaturePolicyError(f"不支持的分析功能: {operation}")
    if parameters is None:
        parameters = {}
    if not isinstance(parameters, Mapping):
        raise FeaturePolicyError("分析参数必须是对象")
    payload = dict(parameters)

    if operation in TRAIT_OPERATIONS:
        if set(payload) != {"traits"}:
            raise FeaturePolicyError("性状分析参数只能包含 traits")
        traits = payload.get("traits")
        if (
            not isinstance(traits, list)
            or not traits
            or len(traits) > 128
            or not all(isinstance(value, str) for value in traits)
        ):
            raise FeaturePolicyError("traits 必须是非空字符串数组")
        normalized_traits = [value.strip() for value in traits]
        if (
            any(not value or len(value) > 100 for value in normalized_traits)
            or len(normalized_traits) != len(set(normalized_traits))
        ):
            raise FeaturePolicyError("traits 包含空值、重复值或超长值")
        from core.breeding_calc.cow_traits_calc import TRAITS_TRANSLATION

        unknown = [
            value
            for value in normalized_traits
            if value not in TRAITS_TRANSLATION
        ]
        if unknown:
            raise FeaturePolicyError(
                "存在不支持的性状：" + "、".join(unknown[:5])
            )
        return {"traits": normalized_traits}

    if operation in INDEX_OPERATIONS:
        if set(payload) != {"weight_name", "weight_values"}:
            raise FeaturePolicyError(
                "指数参数必须包含 weight_name 和 weight_values 快照"
            )
        weight_name = payload.get("weight_name")
        if (
            not isinstance(weight_name, str)
            or not weight_name.strip()
            or len(weight_name.strip()) > 100
            or any(ord(character) < 32 for character in weight_name)
        ):
            raise FeaturePolicyError("weight_name 格式无效")
        values = payload.get("weight_values")
        if not isinstance(values, Mapping) or not values:
            raise FeaturePolicyError("weight_values 必须是非空对象")
        from core.breeding_calc.index_calculation import TRAIT_SD

        normalized_values: dict[str, float] = {}
        for raw_trait, raw_value in values.items():
            trait = str(raw_trait).strip()
            if trait not in TRAIT_SD:
                raise FeaturePolicyError(f"权重包含不支持的性状: {trait}")
            if isinstance(raw_value, bool):
                raise FeaturePolicyError("权重值不能是布尔值")
            try:
                value = float(raw_value)
            except (TypeError, ValueError, OverflowError) as exc:
                raise FeaturePolicyError("权重值必须是数字") from exc
            if not math.isfinite(value):
                raise FeaturePolicyError("权重值必须是有限数字")
            if value:
                normalized_values[trait] = value
        if not normalized_values:
            raise FeaturePolicyError("权重配置不能全部为 0")
        if abs(sum(abs(value) for value in normalized_values.values()) - 100) > 0.0001:
            raise FeaturePolicyError("权重绝对值之和必须为 100")
        return {
            "weight_name": weight_name.strip(),
            "weight_values": {
                key: normalized_values[key]
                for key in sorted(normalized_values)
            },
        }

    if payload:
        raise FeaturePolicyError("该近交分析不接受额外参数")
    return {}


def feature_prerequisite(
    project_path: Path,
    operation: str,
    *,
    dataset_selection: Mapping[str, Any] | None = None,
) -> tuple[str, str]:
    """返回 ``(状态, 说明)``；状态为 ready、skipped 或 failed。"""
    root = Path(project_path)
    selected = dict(dataset_selection or {})
    herd_selected = bool(selected.get("herd", True))
    breeding_selected = bool(selected.get("breeding", True))
    cow_file = root / "standardized_data" / "processed_cow_data.xlsx"
    bull_file = root / "standardized_data" / "processed_bull_data.xlsx"
    breeding_file = (
        root / "standardized_data" / "processed_breeding_data.xlsx"
    )

    if operation in {
        "cow_traits",
        "cow_index",
        "cow_self_inbreeding",
        "mated_inbreeding",
        "candidate_inbreeding",
    }:
        if not herd_selected:
            return "failed", "创建牧场组时未选择牛群/系谱数据"
        if not cow_file.is_file():
            return "failed", "子项目缺少标准化母牛数据"

    if operation in {"bull_traits", "bull_index", "candidate_inbreeding"}:
        if not bull_file.is_file():
            return "skipped", "该牧场未上传备选公牛文件，已跳过"

    if operation in {"mated_bull_traits", "mated_inbreeding"}:
        if not breeding_selected:
            return "skipped", "创建牧场组时未选择配种记录，已跳过"
        if not breeding_file.is_file():
            return "skipped", "该牧场没有可用配种记录，已跳过"

    return "ready", ""


def _feature_inputs(root: Path, operation: str) -> dict[str, Path]:
    result = {}
    for relative in _FEATURE_INPUTS[operation]:
        path = root / relative
        if not path.is_file():
            raise FeaturePolicyError(f"缺少直接输入: {relative.as_posix()}")
        result[relative.as_posix()] = path

    if operation == "cow_index":
        # 母牛指数实际直接读取性状得分文件；仅登记母牛原始表会令得分
        # 文件被单场分析改写后仍错误复用旧指数。计算逻辑优先 genomic，
        # 不存在时才使用 pedigree，这里必须采用完全相同的选择顺序。
        score_input = next(
            (
                root / relative
                for relative in _COW_INDEX_SCORE_INPUTS
                if (root / relative).is_file()
            ),
            None,
        )
        if score_input is None:
            raise FeaturePolicyError("缺少母牛指数实际使用的性状得分文件")
        relative = score_input.relative_to(root).as_posix()
        result[relative] = score_input

    genomic = root / "standardized_data" / "processed_genomic_data.xlsx"
    if operation in {"cow_traits", "cow_index"} and genomic.is_file():
        relative = genomic.relative_to(root).as_posix()
        result[relative] = genomic
    return result


def _output_file_state(path: Path) -> dict[str, int]:
    stat = path.stat()
    return {
        "size_bytes": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
    }


def capture_feature_output_state(
    root_path: Path,
    operation: str,
) -> dict[str, dict[str, int]]:
    """记录运行前的时间戳结果集合，用于只提交本次新建或更新的文件。"""
    if operation not in SUPPORTED_FEATURE_OPERATIONS:
        raise FeaturePolicyError(f"不支持的分析功能: {operation}")
    pattern = _LATEST_OUTPUT_PATTERNS.get(operation)
    if pattern is None:
        return {}
    root = Path(root_path).resolve()
    return {
        path.relative_to(root).as_posix(): _output_file_state(path)
        for path in (root / "analysis_results").glob(pattern)
        if path.is_file()
    }


def resolve_feature_outputs(
    root_path: Path,
    operation: str,
    *,
    output_baseline: Mapping[str, Mapping[str, Any]] | None = None,
) -> dict[str, Path]:
    root = Path(root_path).resolve()
    fixed = _FIXED_FEATURE_OUTPUTS.get(operation)
    if fixed is not None:
        outputs = [root / relative for relative in fixed]
    else:
        if output_baseline is None:
            raise FeaturePolicyError(
                "时间戳结果提交必须提供本次运行前文件快照"
            )
        pattern = _LATEST_OUTPUT_PATTERNS[operation]
        candidates = [
            path
            for path in (root / "analysis_results").glob(pattern)
            if path.is_file()
        ]
        changed = []
        new = []
        for path in candidates:
            relative = path.relative_to(root).as_posix()
            state = _output_file_state(path)
            previous = output_baseline.get(relative)
            if previous is None:
                new.append(path)
                changed.append(path)
            elif dict(previous) != state:
                changed.append(path)
        # 正常运行会产生一个新时间戳文件。若同一秒重算覆盖同名文件，
        # 则退回本次确实发生变化的候选；未变化的“未来 mtime”旧文件
        # 永远不会参与本次提交。
        fresh = new or changed
        outputs = (
            [
                max(
                    fresh,
                    key=lambda path: (
                        path.stat().st_mtime_ns,
                        path.name,
                    ),
                )
            ]
            if fresh
            else []
        )
    missing = [
        path.relative_to(root).as_posix()
        for path in outputs
        if not path.is_file()
    ]
    if not outputs or missing:
        raise FeaturePolicyError(
            "分析未生成本次可确认的完整结果"
            + (f"：{'、'.join(missing)}" if missing else "")
        )
    return {
        path.relative_to(root).as_posix(): path
        for path in outputs
    }


def _first_sheet_headers(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return {
            str(value).strip()
            for value in row
            if value is not None and str(value).strip()
        }
    finally:
        workbook.close()


def _multiset_signature(value: Mapping[str, Any]) -> tuple[Any, ...]:
    return (
        int(value.get("row_count", -1)),
        int(value.get("identifier_count", -1)),
        int(value.get("blank_count", -1)),
        str(value.get("hash_sum") or ""),
        str(value.get("hash_square_sum") or ""),
        str(value.get("hash_xor") or ""),
    )


def _normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value).strip()
    if text.casefold() in _MISSING_IDENTIFIERS:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _new_identifier_multiset() -> dict[str, int]:
    return {
        "row_count": 0,
        "identifier_count": 0,
        "blank_count": 0,
        "hash_sum": 0,
        "hash_square_sum": 0,
        "hash_xor": 0,
    }


def _add_identifier_to_multiset(
    state: dict[str, int],
    value: Any,
) -> None:
    state["row_count"] += 1
    identifier = _normalize_identifier(value)
    if not identifier:
        state["blank_count"] += 1
        return
    state["identifier_count"] += 1
    number = int.from_bytes(
        hashlib.sha256(identifier.encode("utf-8")).digest(),
        "big",
    )
    state["hash_sum"] = (
        state["hash_sum"] + number
    ) % _IDENTIFIER_HASH_MODULUS
    state["hash_square_sum"] = (
        state["hash_square_sum"] + number * number
    ) % _IDENTIFIER_HASH_MODULUS
    state["hash_xor"] ^= number


def _public_identifier_multiset(
    state: Mapping[str, int],
) -> dict[str, Any]:
    return {
        "row_count": int(state["row_count"]),
        "identifier_count": int(state["identifier_count"]),
        "blank_count": int(state["blank_count"]),
        "hash_sum": f"{int(state['hash_sum']):064x}",
        "hash_square_sum": f"{int(state['hash_square_sum']):064x}",
        "hash_xor": f"{int(state['hash_xor']):064x}",
    }


def _filtered_dairy_cow_multiset(path: Path) -> dict[str, Any]:
    """按性状计算相同的奶牛母牛口径流式计算牛号多重集。"""
    from config.breed_constants import is_dairy_breed

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, ())
        normalized_headers = {
            str(value).strip().casefold(): index
            for index, value in enumerate(headers)
            if value is not None and str(value).strip()
        }

        def column_index(candidates: tuple[str, ...]) -> int | None:
            for candidate in candidates:
                index = normalized_headers.get(candidate.casefold())
                if index is not None:
                    return index
            return None

        cow_index = column_index(_COW_ID_COLUMNS)
        if cow_index is None:
            raise FeaturePolicyError("母牛输入缺少可识别的牛号列")
        sex_index = column_index(("sex", "性别"))
        breed_index = column_index(("breed", "品种"))
        all_female = _new_identifier_multiset()
        dairy_female = _new_identifier_multiset()

        for row in rows:
            # 与通用 XLSX 多重集实现一致：仅设置了样式的模板空尾不是
            # 业务记录，不能把它计入 blank_count 后误判为结果丢行。
            if all(
                value is None
                or (
                    isinstance(value, str)
                    and not value.strip()
                )
                for value in row
            ):
                continue
            sex = (
                row[sex_index]
                if sex_index is not None and sex_index < len(row)
                else None
            )
            if sex_index is not None and str(
                "母" if sex is None else sex
            ).strip() == "公":
                continue
            cow_id = row[cow_index] if cow_index < len(row) else None
            _add_identifier_to_multiset(all_female, cow_id)
            breed = (
                row[breed_index]
                if breed_index is not None and breed_index < len(row)
                else None
            )
            if breed_index is None or is_dairy_breed(breed):
                _add_identifier_to_multiset(dairy_female, cow_id)

        selected = dairy_female
        if (
            breed_index is not None
            and all_female["row_count"] > 0
            and dairy_female["row_count"] == 0
        ):
            # 与 filter_dairy_cows 的“全部被未知品种过滤时保留母牛”
            # 安全兜底保持一致。
            selected = all_female
        return _public_identifier_multiset(selected)
    finally:
        workbook.close()


def _find_artifact(
    artifacts: Mapping[str, Path],
    filename: str,
) -> Path:
    for logical_name, path in artifacts.items():
        if logical_name.endswith(filename):
            return path
    raise FeaturePolicyError(f"功能产物清单缺少文件: {filename}")


def _validate_identifier_flow(
    operation: str,
    inputs: Mapping[str, Path],
    outputs: Mapping[str, Path],
) -> None:
    """首次提交前核验不会丢牛、重复牛或串到其它牛号。

    牛只指数会排序，性状表也可能改变行序，因此使用顺序无关且保留
    重复次数的多重集。近交配对属于一对多扩张，不适用该相等约束。
    """
    if operation not in {
        "cow_traits",
        "bull_traits",
        "mated_bull_traits",
        "cow_index",
        "bull_index",
    }:
        return

    if operation == "cow_traits":
        expected = _filtered_dairy_cow_multiset(
            _find_artifact(inputs, "processed_cow_data.xlsx")
        )
        actual = compute_xlsx_identifier_multiset(
            _find_artifact(
                outputs,
                "processed_cow_data_key_traits_final.xlsx",
            ),
            _COW_ID_COLUMNS,
        )
    elif operation == "bull_traits":
        expected = compute_xlsx_identifier_multiset(
            _find_artifact(inputs, "processed_bull_data.xlsx"),
            _BULL_ID_COLUMNS,
        )
        actual = compute_xlsx_identifier_multiset(
            _find_artifact(
                outputs,
                "processed_bull_data_key_traits.xlsx",
            ),
            _BULL_ID_COLUMNS,
        )
    elif operation == "mated_bull_traits":
        expected = compute_xlsx_identifier_multiset(
            _find_artifact(inputs, "processed_breeding_data.xlsx"),
            ("耳号", "cow_id", "母牛号"),
        )
        actual = compute_xlsx_identifier_multiset(
            _find_artifact(
                outputs,
                "processed_mated_bull_traits.xlsx",
            ),
            ("耳号", "cow_id", "母牛号"),
        )
    elif operation == "cow_index":
        score_input = next(
            (
                path
                for logical_name, path in inputs.items()
                if logical_name.endswith(
                    "processed_cow_data_key_traits_scores_genomic.xlsx"
                )
                or logical_name.endswith(
                    "processed_cow_data_key_traits_scores_pedigree.xlsx"
                )
            ),
            None,
        )
        if score_input is None:
            raise FeaturePolicyError("母牛指数清单缺少实际性状得分输入")
        expected = compute_xlsx_identifier_multiset(
            score_input,
            _COW_ID_COLUMNS,
        )
        actual = compute_xlsx_identifier_multiset(
            _find_artifact(
                outputs,
                "processed_index_cow_index_scores.xlsx",
            ),
            _COW_ID_COLUMNS,
        )
    else:
        expected = compute_xlsx_identifier_multiset(
            _find_artifact(inputs, "processed_bull_data.xlsx"),
            _BULL_ID_COLUMNS,
        )
        actual = compute_xlsx_identifier_multiset(
            _find_artifact(
                outputs,
                "processed_index_bull_scores.xlsx",
            ),
            _BULL_ID_COLUMNS,
        )

    if _multiset_signature(expected) != _multiset_signature(actual):
        raise FeaturePolicyError(
            "结果牛号多重集与本次分析直接输入不一致，拒绝提交"
        )


def _validate_output_columns(
    operation: str,
    parameters: Mapping[str, Any],
    outputs: Mapping[str, Path],
) -> None:
    primary = next(iter(outputs.values()))
    headers = _first_sheet_headers(primary)
    if operation == "cow_traits":
        expected = {
            f"{trait}_score"
            for trait in parameters["traits"]
        }
    elif operation in {"bull_traits", "mated_bull_traits"}:
        expected = set(parameters["traits"])
    elif operation in INDEX_OPERATIONS:
        expected = {
            f"{parameters['weight_name']}_index",
            "ranking",
        }
    elif operation == "cow_self_inbreeding":
        expected = {"母牛号", "近交系数"}
    elif operation == "mated_inbreeding":
        expected = {"母牛号", "配种公牛号", "后代近交系数"}
    else:
        expected = {"母牛号", "备选公牛号", "后代近交系数"}
    missing = sorted(expected - headers)
    if missing:
        raise FeaturePolicyError(
            f"结果文件缺少必要列：{'、'.join(missing[:8])}"
        )


def _cow_id_sources(
    inputs: Mapping[str, Path],
    outputs: Mapping[str, Path],
    operation: str,
) -> dict[str, dict[str, Any]]:
    sources: dict[str, dict[str, Any]] = {}
    for logical_name in inputs:
        if "processed_bull_data" in logical_name or operation in {
            "bull_traits",
            "bull_index",
        }:
            sources[logical_name] = {"columns": _BULL_ID_COLUMNS}
        elif (
            "processed_breeding_data" in logical_name
            or operation == "mated_bull_traits"
        ):
            sources[logical_name] = {"columns": ("耳号", "cow_id", "母牛号")}
        elif operation in INBREEDING_OPERATIONS and logical_name in outputs:
            sources[logical_name] = {
                "columns": ("母牛号", "cow_id", "耳号"),
                "sheet_name": "配对明细表",
            }
        else:
            sources[logical_name] = {"columns": _COW_ID_COLUMNS}
    primary_output_names = {
        "cow_traits": "processed_cow_data_key_traits_final.xlsx",
        "bull_traits": "processed_bull_data_key_traits.xlsx",
        "mated_bull_traits": "processed_mated_bull_traits.xlsx",
        "cow_index": "processed_index_cow_index_scores.xlsx",
        "bull_index": "processed_index_bull_scores.xlsx",
        "cow_self_inbreeding": "母牛近交系数分析结果.xlsx",
        "mated_inbreeding": "已配公牛_近交系数及隐性基因分析结果",
        "candidate_inbreeding": "备选公牛_近交系数及隐性基因分析结果",
    }
    primary_token = primary_output_names[operation]
    for logical_name in outputs:
        if primary_token not in logical_name:
            continue
        if operation in {"bull_traits", "bull_index"}:
            sources[logical_name] = {"columns": _BULL_ID_COLUMNS}
        elif operation == "mated_bull_traits":
            sources[logical_name] = {
                "columns": ("耳号", "cow_id", "母牛号")
            }
        elif operation in INBREEDING_OPERATIONS:
            sources[logical_name] = {
                "columns": ("母牛号", "cow_id", "耳号"),
                "sheet_name": "配对明细表",
            }
        else:
            sources[logical_name] = {"columns": _COW_ID_COLUMNS}
    return sources


def _feature_config(
    operation: str,
    parameters: Mapping[str, Any],
    *,
    bull_library_version: Any = _VERSION_UNSET,
) -> dict[str, Any]:
    if bull_library_version is _VERSION_UNSET:
        from core.data.update_manager import get_local_db_version

        bull_library_version = get_local_db_version()

    return {
        "policy_revision": FEATURE_POLICY_REVISION,
        "operation": operation,
        "parameters": dict(parameters),
        "bull_library_version": bull_library_version,
        "analysis_calendar_year": datetime.now().year,
    }


def commit_feature_manifest(
    root_path: Path,
    operation: str,
    parameters: Mapping[str, Any],
    *,
    expected_task_id: str,
    expected_farm_code: str,
    bull_library_version: Any = _VERSION_UNSET,
    output_baseline: Mapping[str, Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    """校验列、XLSX、输入哈希及牛号多重集后原子提交功能清单。"""
    root = Path(root_path).resolve()
    normalized = normalize_feature_parameters(operation, parameters)
    inputs = _feature_inputs(root, operation)
    outputs = resolve_feature_outputs(
        root,
        operation,
        output_baseline=output_baseline,
    )
    _validate_output_columns(operation, normalized, outputs)
    _validate_identifier_flow(operation, inputs, outputs)
    return commit_stage_manifest(
        root,
        feature_manifest_path(operation),
        task_id=expected_task_id,
        farm_code=expected_farm_code,
        stage=f"feature:{operation}",
        config=_feature_config(
            operation,
            normalized,
            bull_library_version=bull_library_version,
        ),
        inputs=inputs,
        outputs=outputs,
        cow_id_sources=_cow_id_sources(inputs, outputs, operation),
    )


def validate_feature_manifest(
    root_path: Path,
    operation: str,
    parameters: Mapping[str, Any],
    *,
    expected_task_id: str,
    expected_farm_code: str,
    verification: str = "full",
    bull_library_version: Any = _VERSION_UNSET,
) -> dict[str, Any]:
    root = Path(root_path).resolve()
    normalized = normalize_feature_parameters(operation, parameters)
    validation = validate_stage_manifest(
        root,
        feature_manifest_path(operation),
        expected_task_id=expected_task_id,
        expected_farm_code=expected_farm_code,
        expected_stage=f"feature:{operation}",
        expected_config=_feature_config(
            operation,
            normalized,
            bull_library_version=bull_library_version,
        ),
        verification=verification,
    )
    if not validation.get("valid"):
        return validation

    return _validate_current_feature_input_set(
        root,
        operation,
        validation,
    )


def _validate_current_feature_input_set(
    root: Path,
    operation: str,
    validation: Mapping[str, Any],
) -> dict[str, Any]:
    """补充核对通用清单无法发现的新增/切换直接输入。"""
    validation = dict(validation)
    if not validation.get("valid"):
        return validation

    # 通用清单会核验已登记文件的内容，但不会发现一个新的、更高优先级
    # 可选输入刚刚出现。重新解析当前直接输入集合，避免 cow_index 从
    # pedigree 切换到 genomic 后仍复用旧指数。
    try:
        current_inputs = set(_feature_inputs(root, operation))
    except Exception as exc:
        return {
            **validation,
            "valid": False,
            "status": "artifact_missing",
            "issues": [
                *validation.get("issues", []),
                {
                    "code": "artifact_missing",
                    "artifact": "",
                    "message": str(exc),
                },
            ],
        }
    manifest = validation.get("manifest") or {}
    stored_inputs = {
        str(item.get("logical_name") or "")
        for item in manifest.get("inputs", [])
        if item.get("logical_name")
    }
    if stored_inputs != current_inputs:
        return {
            **validation,
            "valid": False,
            "status": "artifact_mismatch",
            "issues": [
                *validation.get("issues", []),
                {
                    "code": "input_selection_changed",
                    "artifact": "",
                    "message": "分析实际输入集合已经变化",
                },
            ],
        }
    return validation


def validate_recorded_feature_manifest(
    root_path: Path,
    operation: str,
    *,
    expected_task_id: str,
    expected_farm_code: str,
    verification: str = "stat",
) -> dict[str, Any]:
    """核验最近一次页面分析，不要求调用方重新提供当时页面参数。

    该入口只用于只读状态展示。它会验证身份、全部登记文件和当前直接
    输入集合，但不会把页面分析提升为完整报告 analysis 阶段。
    """
    root = Path(root_path).resolve()
    validation = validate_stage_manifest(
        root,
        feature_manifest_path(operation),
        expected_task_id=expected_task_id,
        expected_farm_code=expected_farm_code,
        expected_stage=f"feature:{operation}",
        verification=verification,
    )
    return _validate_current_feature_input_set(
        root,
        operation,
        validation,
    )


def manifest_declares_feature_outputs(
    validation: Mapping[str, Any],
    operation: str,
) -> bool:
    """判断一个有效清单是否登记了某项分析的完整正式输出。"""
    if operation not in SUPPORTED_FEATURE_OPERATIONS:
        raise ValueError(f"不支持的牧场组分析功能: {operation}")
    if not validation.get("valid"):
        return False
    manifest = validation.get("manifest") or {}
    output_paths = {
        str(item.get("relative_path") or "")
        for item in manifest.get("outputs", [])
        if isinstance(item, Mapping) and item.get("relative_path")
    }
    fixed = _FIXED_FEATURE_OUTPUTS.get(operation)
    if fixed is not None:
        required = {path.as_posix() for path in fixed}
        return required.issubset(output_paths)
    pattern = _LATEST_OUTPUT_PATTERNS[operation]
    return any(
        fnmatch.fnmatch(Path(relative).name, pattern)
        for relative in output_paths
    )


def manifest_declares_report_analysis_outputs(
    validation: Mapping[str, Any],
    operation: str,
) -> bool:
    """判断完整报告阶段清单是否登记了某项分析的报告口径结果。"""
    if operation not in SUPPORTED_FEATURE_OPERATIONS:
        raise ValueError(f"不支持的牧场组分析功能: {operation}")
    if not validation.get("valid"):
        return False
    manifest = validation.get("manifest") or {}
    output_paths = {
        str(item.get("relative_path") or "")
        for item in manifest.get("outputs", [])
        if isinstance(item, Mapping) and item.get("relative_path")
    }
    fixed = _REPORT_OPERATION_FIXED_OUTPUTS.get(operation)
    if fixed is not None:
        required = {path.as_posix() for path in fixed}
        return required.issubset(output_paths)
    pattern = _LATEST_OUTPUT_PATTERNS[operation]
    return any(
        fnmatch.fnmatch(Path(relative).name, pattern)
        for relative in output_paths
    )


def _archive_manifest(root: Path, relative: Path, reason: str) -> None:
    source = root / relative
    if not source.is_file():
        return
    history = source.parent / "history"
    history.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    target = history / f"{source.stem}_{reason}_{stamp}_{os.getpid()}.json"
    os.replace(source, target)


def discard_feature_manifest(
    root_path: Path,
    operation: str,
    reason: str,
) -> None:
    """归档一项刚提交但不能安全复用的功能清单。"""
    _archive_manifest(
        Path(root_path).resolve(),
        feature_manifest_path(operation),
        str(reason or "discarded"),
    )


def invalidate_before_feature_run(root_path: Path, operation: str) -> None:
    """只归档清单，不移动已有分析结果。

    页面参数重算会令完整分析、单场报告及最终汇总过期，但不能像完整
    ``analysis`` 阶段那样把其它功能的正式结果整体移走。
    """
    root = Path(root_path).resolve()
    _archive_manifest(
        root,
        feature_manifest_path(operation),
        "replaced",
    )
    if operation == "cow_traits":
        _archive_manifest(
            root,
            feature_manifest_path("cow_index"),
            "upstream_changed",
        )
    from core.group_tasks.stage_policy import stage_manifest_path

    _archive_manifest(root, stage_manifest_path("analysis"), "feature_changed")
    _archive_manifest(
        root,
        stage_manifest_path("child_excel"),
        "feature_changed",
    )


def manifest_artifacts(validation: Mapping[str, Any]) -> list[str]:
    manifest = validation.get("manifest") or {}
    return [
        str(item.get("relative_path") or "")
        for item in manifest.get("outputs", [])
        if item.get("relative_path")
    ]

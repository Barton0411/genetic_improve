"""牧场组单阶段产物的原子提交与完整性验证。

阶段清单只保存项目根目录内的相对路径、身份字段和不可逆指纹，不保存
配置原文、绝对路径或凭据。大文件哈希按块计算，XLSX 结构检查直接复用
ZIP/XML 流式检查器；可选的牛号多重集也通过只读流式工作簿计算。
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import posixpath
import re
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, Mapping, Optional, Sequence, Tuple, Union
from xml.etree import ElementTree

from openpyxl import load_workbook


SCHEMA_VERSION = 1
DEFAULT_HASH_CHUNK_SIZE = 1024 * 1024
_MISSING_IDENTIFIERS = {"", "nan", "none", "null", "nat", "<na>", "n/a"}
_MODULUS = 1 << 256
_UNSET = object()
PathLike = Union[str, os.PathLike]
_CELL_REFERENCE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")
_RELATIONSHIP_ID = (
    "{http://schemas.openxmlformats.org/officeDocument/2006/"
    "relationships}id"
)


class StageManifestError(RuntimeError):
    """阶段清单无法安全创建时抛出的错误。"""


def _local_name(tag: str) -> str:
    return str(tag).rsplit("}", 1)[-1]


def _cell_position(reference: str) -> Tuple[int, int]:
    match = _CELL_REFERENCE.fullmatch(str(reference or "").strip())
    if match is None:
        raise ValueError(f"无效单元格引用: {reference!r}")
    column = 0
    for character in match.group(1).upper():
        column = column * 26 + ord(character) - ord("A") + 1
    row = int(match.group(2))
    if row > 1_048_576 or column > 16_384:
        raise ValueError(f"单元格引用超出 Excel 上限: {reference!r}")
    return row, column


def _sheet_extent(
    archive: zipfile.ZipFile,
    member_name: str,
) -> Tuple[int, int]:
    """优先读取 dimension；缺失时流式扫描单元格引用。"""
    max_row = 0
    max_column = 0
    parser = ElementTree.XMLPullParser(events=("start", "end"))
    with archive.open(member_name, "r") as stream:
        while True:
            chunk = stream.read(64 * 1024)
            if not chunk:
                break
            parser.feed(chunk)
            for event, element in parser.read_events():
                local_name = _local_name(element.tag)
                if event == "start" and local_name == "dimension":
                    reference = str(element.attrib.get("ref") or "")
                    if not reference:
                        raise ValueError("工作表 dimension 缺少 ref")
                    positions = [
                        _cell_position(item)
                        for item in reference.split(":")
                    ]
                    return (
                        max(item[0] for item in positions),
                        max(item[1] for item in positions),
                    )
                if event == "start" and local_name == "c":
                    reference = element.attrib.get("r")
                    if not reference:
                        raise ValueError("工作表存在缺少 r 的单元格")
                    row, column = _cell_position(reference)
                    max_row = max(max_row, row)
                    max_column = max(max_column, column)
                elif event == "end":
                    element.clear()
        parser.close()
    return max_row, max_column


def _xlsx_structure(path: Path) -> Dict[str, Any]:
    """独立、低内存地核验 XLSX ZIP 和 Sheet 行列边界。

    这里不依赖 ``core.group_report``，避免任务调度接入阶段清单后触发
    报告包 ``__init__`` 所造成的循环导入。
    """
    sheets = []
    try:
        with zipfile.ZipFile(path, "r") as archive:
            members = archive.namelist()
            if len(members) != len(set(members)):
                raise ValueError("XLSX ZIP 中存在重复成员路径")
            required = {
                "[Content_Types].xml",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
            }
            missing = sorted(required.difference(members))
            if missing:
                raise ValueError(
                    "XLSX 缺少必要结构: " + ", ".join(missing)
                )
            corrupt_member = archive.testzip()
            if corrupt_member is not None:
                raise ValueError(
                    f"XLSX ZIP 成员 CRC 校验失败: {corrupt_member}"
                )

            with archive.open(
                "xl/_rels/workbook.xml.rels", "r"
            ) as stream:
                relationship_root = ElementTree.parse(stream).getroot()
            relationships = {}
            for element in relationship_root.iter():
                if _local_name(element.tag) != "Relationship":
                    continue
                relationship_id = element.attrib.get("Id")
                if relationship_id:
                    relationships[str(relationship_id)] = {
                        "target": str(element.attrib.get("Target") or ""),
                        "mode": str(element.attrib.get("TargetMode") or ""),
                        "type": str(element.attrib.get("Type") or ""),
                    }

            with archive.open("xl/workbook.xml", "r") as stream:
                workbook_root = ElementTree.parse(stream).getroot()
            for element in workbook_root.iter():
                if _local_name(element.tag) != "sheet":
                    continue
                name = str(element.attrib.get("name") or "")
                relationship_id = element.attrib.get(_RELATIONSHIP_ID)
                if relationship_id is None:
                    relationship_id = next(
                        (
                            value
                            for key, value in element.attrib.items()
                            if _local_name(key) == "id"
                        ),
                        None,
                    )
                relationship = relationships.get(
                    str(relationship_id or "")
                )
                if relationship is None:
                    raise ValueError(f"工作表 {name!r} 缺少 relationship")
                if relationship["mode"].casefold() == "external":
                    raise ValueError(f"工作表 {name!r} 使用外部 relationship")
                kind = relationship["type"].rstrip("/").rsplit("/", 1)[-1]
                if kind != "worksheet":
                    # 阶段数据产物只允许普通 worksheet。图表页作为独立
                    # workbook sheet 会令牛号/行数审计含义不明确。
                    raise ValueError(
                        f"工作表 {name!r} 类型不受支持: {kind!r}"
                    )
                target = relationship["target"].replace("\\", "/").strip()
                member_name = (
                    posixpath.normpath(target.lstrip("/"))
                    if target.startswith("/")
                    else posixpath.normpath(posixpath.join("xl", target))
                )
                if (
                    member_name == ".."
                    or member_name.startswith("../")
                    or member_name not in members
                ):
                    raise ValueError(
                        f"工作表 {name!r} XML 路径无效"
                    )
                max_row, max_column = _sheet_extent(
                    archive,
                    member_name,
                )
                sheets.append(
                    {
                        "name": name,
                        "state": str(
                            element.attrib.get("state", "visible")
                        ),
                        "max_row": max_row,
                        "max_column": max_column,
                    }
                )
            if not sheets:
                raise ValueError("XLSX 工作簿不包含普通 worksheet")
    except Exception as exc:
        return {
            "valid": False,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "error": f"{type(exc).__name__}: {str(exc)[:500]}",
        }
    return {
        "valid": True,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "error": "",
    }


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def stream_sha256(
    path: PathLike,
    chunk_size: int = DEFAULT_HASH_CHUNK_SIZE,
) -> str:
    """按块计算文件 SHA-256，不把整个文件读入内存。"""
    if int(chunk_size) <= 0:
        raise ValueError("chunk_size 必须大于 0")
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        while True:
            chunk = stream.read(int(chunk_size))
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return value.as_posix()
    if isinstance(value, (set, frozenset)):
        return sorted(value, key=lambda item: str(item))
    if isinstance(value, bytes):
        return {"bytes_sha256": hashlib.sha256(value).hexdigest()}
    raise TypeError(f"配置中存在不可序列化类型: {type(value).__name__}")


def compute_config_fingerprint(config: Any) -> str:
    """计算稳定配置指纹；配置原文不会写入阶段清单。"""
    serialized = json.dumps(
        config,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
        default=_json_default,
    ).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()


def _normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value).strip()
    if text.casefold() in _MISSING_IDENTIFIERS:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _identifier_multiset_state() -> Dict[str, int]:
    return {
        "row_count": 0,
        "identifier_count": 0,
        "blank_count": 0,
        "hash_sum": 0,
        "hash_square_sum": 0,
        "hash_xor": 0,
    }


def _add_identifier(state: Dict[str, int], value: Any) -> None:
    state["row_count"] += 1
    identifier = _normalize_identifier(value)
    if not identifier:
        state["blank_count"] += 1
        return
    state["identifier_count"] += 1
    number = int.from_bytes(
        hashlib.sha256(identifier.encode("utf-8")).digest(),
        "big",
    )
    state["hash_sum"] = (state["hash_sum"] + number) % _MODULUS
    state["hash_square_sum"] = (
        state["hash_square_sum"] + number * number
    ) % _MODULUS
    state["hash_xor"] ^= number


def _public_multiset(
    state: Dict[str, int],
    *,
    column_name: str,
    column_index: int,
    sheet_name: str,
) -> Dict[str, Any]:
    components = {
        "row_count": int(state["row_count"]),
        "identifier_count": int(state["identifier_count"]),
        "blank_count": int(state["blank_count"]),
        "hash_sum": f"{int(state['hash_sum']):064x}",
        "hash_square_sum": f"{int(state['hash_square_sum']):064x}",
        "hash_xor": f"{int(state['hash_xor']):064x}",
    }
    fingerprint = hashlib.sha256(
        json.dumps(
            components,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    return {
        "algorithm": "sha256-multiset-v1",
        "sheet_name": sheet_name,
        "column_name": column_name,
        "column_index": int(column_index),
        **components,
        "fingerprint": fingerprint,
    }


def compute_xlsx_identifier_multiset(
    path: PathLike,
    columns: Union[str, Sequence[str]],
    *,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """流式计算 XLSX 指定牛号列的顺序无关、多重性敏感指纹。"""
    candidates = [columns] if isinstance(columns, str) else list(columns)
    normalized_candidates = {
        str(candidate).strip().casefold()
        for candidate in candidates
        if str(candidate).strip()
    }
    if not normalized_candidates:
        raise ValueError("至少需要一个牛号列候选名称")

    workbook = load_workbook(
        Path(path),
        read_only=True,
        data_only=True,
    )
    try:
        if sheet_name is None:
            worksheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作簿不存在 Sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ValueError("工作表为空，无法识别牛号列") from exc

        column_index = next(
            (
                index
                for index, header in enumerate(headers)
                if str(header or "").strip().casefold()
                in normalized_candidates
            ),
            None,
        )
        if column_index is None:
            raise ValueError(
                "未找到牛号列，候选列为: "
                + "、".join(str(item) for item in candidates)
            )
        column_name = str(headers[column_index] or "").strip()
        state = _identifier_multiset_state()
        for row in rows:
            # 跳过完全空白的物理行。某些模板只设置了远端行的格式，
            # openpyxl 仍会把这些行包含在 max_row 中；它们不是业务记录。
            if all(
                value is None
                or (
                    isinstance(value, str)
                    and not value.strip()
                )
                for value in row
            ):
                continue
            value = row[column_index] if column_index < len(row) else None
            _add_identifier(state, value)
        return _public_multiset(
            state,
            column_name=column_name,
            column_index=column_index + 1,
            sheet_name=worksheet.title,
        )
    finally:
        workbook.close()


def _atomic_write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary_path = Path(temporary_name)
    descriptor_open = True
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as stream:
            descriptor_open = False
            json.dump(
                payload,
                stream,
                ensure_ascii=False,
                indent=2,
                allow_nan=False,
            )
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary_path, path)
        try:
            directory_descriptor = os.open(
                str(path.parent),
                os.O_RDONLY,
            )
            try:
                os.fsync(directory_descriptor)
            finally:
                os.close(directory_descriptor)
        except OSError:
            # Windows/部分文件系统不支持目录 fsync；文件本身仍已 fsync
            # 并通过 os.replace 原子替换。
            pass
    finally:
        if descriptor_open:
            try:
                os.close(descriptor)
            except OSError:
                pass
        temporary_path.unlink(missing_ok=True)


def _safe_relative_path(root: Path, path: PathLike) -> Tuple[Path, str]:
    root = root.resolve()
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = root / candidate
    resolved = candidate.resolve()
    try:
        relative = resolved.relative_to(root)
    except ValueError as exc:
        raise StageManifestError("阶段文件路径超出项目根目录") from exc
    if resolved.is_symlink() or candidate.is_symlink():
        raise StageManifestError("阶段文件不能是符号链接")
    if not resolved.is_file():
        raise StageManifestError(f"阶段文件不存在: {relative.as_posix()}")
    return resolved, relative.as_posix()


def _manifest_relative_path(root: Path, manifest_path: PathLike) -> Path:
    root = root.resolve()
    candidate = Path(manifest_path)
    if not candidate.is_absolute():
        candidate = root / candidate
    # manifest 可以尚不存在，因此只解析父目录。
    parent = candidate.parent.resolve()
    try:
        parent.relative_to(root)
    except ValueError as exc:
        raise StageManifestError("manifest 路径超出项目根目录") from exc
    return parent / candidate.name


def _normalize_artifacts(
    artifacts: Union[Mapping[str, PathLike], Sequence[PathLike]],
    kind: str,
) -> Sequence[Tuple[str, PathLike]]:
    if isinstance(artifacts, Mapping):
        items = [(str(name), path) for name, path in artifacts.items()]
    else:
        items = [
            (f"{kind}_{index:03d}", path)
            for index, path in enumerate(artifacts, start=1)
        ]
    logical_names = [name for name, _ in items]
    if any(not name.strip() for name in logical_names):
        raise StageManifestError("阶段文件逻辑名称不能为空")
    if len(logical_names) != len(set(logical_names)):
        raise StageManifestError("阶段文件逻辑名称不能重复")
    return items


def _cow_spec(
    cow_id_sources: Optional[Mapping[str, Any]],
    logical_name: str,
) -> Optional[Tuple[Union[str, Sequence[str]], Optional[str]]]:
    if not cow_id_sources or logical_name not in cow_id_sources:
        return None
    value = cow_id_sources[logical_name]
    if isinstance(value, Mapping):
        columns = value.get("columns") or value.get("column")
        sheet_name = value.get("sheet_name") or value.get("sheet")
    else:
        columns = value
        sheet_name = None
    if not columns:
        raise StageManifestError(
            f"牛号多重集配置缺少列名: {logical_name}"
        )
    return columns, str(sheet_name) if sheet_name else None


def _capture_artifact(
    root: Path,
    logical_name: str,
    path: PathLike,
    kind: str,
    cow_id_sources: Optional[Mapping[str, Any]],
) -> Dict[str, Any]:
    absolute_path, relative_path = _safe_relative_path(root, path)
    before = absolute_path.stat()
    entry: Dict[str, Any] = {
        "logical_name": logical_name,
        "kind": kind,
        "relative_path": relative_path,
        "size_bytes": int(before.st_size),
        "mtime_ns": int(before.st_mtime_ns),
        "sha256": stream_sha256(absolute_path),
        "xlsx": None,
        "cow_id_multiset": None,
    }
    if absolute_path.suffix.casefold() == ".xlsx":
        structure = _xlsx_structure(absolute_path)
        if not structure.get("valid"):
            raise StageManifestError(
                f"XLSX 结构校验失败 {relative_path}: "
                f"{structure.get('error', '')}"
            )
        entry["xlsx"] = structure
        spec = _cow_spec(cow_id_sources, logical_name)
        if spec is not None:
            entry["cow_id_multiset"] = compute_xlsx_identifier_multiset(
                absolute_path,
                spec[0],
                sheet_name=spec[1],
            )
    elif _cow_spec(cow_id_sources, logical_name) is not None:
        raise StageManifestError(
            f"牛号多重集当前只支持 XLSX: {relative_path}"
        )

    after = absolute_path.stat()
    if (
        before.st_size != after.st_size
        or before.st_mtime_ns != after.st_mtime_ns
    ):
        raise StageManifestError(
            f"扫描期间阶段文件发生变化: {relative_path}"
        )
    return entry


def commit_stage_manifest(
    root_path: PathLike,
    manifest_path: PathLike,
    *,
    task_id: str,
    farm_code: str,
    stage: str,
    config: Any,
    inputs: Union[Mapping[str, PathLike], Sequence[PathLike]],
    outputs: Union[Mapping[str, PathLike], Sequence[PathLike]],
    cow_id_sources: Optional[Mapping[str, Any]] = None,
) -> Dict[str, Any]:
    """扫描阶段输入/输出并以原子替换方式提交 committed manifest。"""
    identity = {
        "task_id": str(task_id).strip(),
        "farm_code": str(farm_code).strip(),
        "stage": str(stage).strip(),
    }
    if any(not value for value in identity.values()):
        raise StageManifestError("task_id、farm_code、stage 均不能为空")
    root = Path(root_path).resolve()
    if not root.is_dir():
        raise StageManifestError("项目根目录不存在")
    target = _manifest_relative_path(root, manifest_path)

    input_entries = [
        _capture_artifact(
            root,
            logical_name,
            path,
            "input",
            cow_id_sources,
        )
        for logical_name, path in _normalize_artifacts(inputs, "input")
    ]
    output_entries = [
        _capture_artifact(
            root,
            logical_name,
            path,
            "output",
            cow_id_sources,
        )
        for logical_name, path in _normalize_artifacts(outputs, "output")
    ]
    if not output_entries:
        raise StageManifestError("阶段至少需要一个输出文件")
    all_paths = [
        entry["relative_path"]
        for entry in input_entries + output_entries
    ]
    if len(all_paths) != len(set(all_paths)):
        raise StageManifestError("同一文件不能在阶段清单中重复登记")

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "status": "committed",
        "committed_at": _utc_now(),
        **identity,
        "config_fingerprint": compute_config_fingerprint(config),
        "inputs": input_entries,
        "outputs": output_entries,
    }
    _atomic_write_json(target, manifest)
    return manifest


def _resolve_stored_path(root: Path, value: Any) -> Path:
    text = str(value or "")
    if not text or "\\" in text:
        raise ValueError("manifest 中包含无效相对路径")
    pure = PurePosixPath(text)
    if pure.is_absolute() or any(part in {"", ".", ".."} for part in pure.parts):
        raise ValueError("manifest 中包含不安全相对路径")
    candidate = (root / Path(*pure.parts)).resolve()
    try:
        candidate.relative_to(root)
    except ValueError as exc:
        raise ValueError("manifest 文件路径超出项目根目录") from exc
    if candidate.is_symlink():
        raise ValueError("manifest 文件路径指向符号链接")
    return candidate


def _artifact_validation_issues(
    root: Path,
    expected: Dict[str, Any],
    *,
    verification: str,
) -> Tuple[Sequence[Dict[str, str]], Optional[Dict[str, Any]]]:
    logical_name = str(expected.get("logical_name") or "")
    relative_path = str(expected.get("relative_path") or "")
    issues = []
    try:
        path = _resolve_stored_path(root, relative_path)
    except Exception as exc:
        return (
            [
                {
                    "code": "unsafe_artifact_path",
                    "artifact": logical_name,
                    "message": str(exc),
                }
            ],
            None,
        )
    if not path.is_file():
        return (
            [
                {
                    "code": "artifact_missing",
                    "artifact": logical_name,
                    "message": f"文件不存在: {relative_path}",
                }
            ],
            None,
        )

    before = path.stat()
    artifact_state = {
        "logical_name": logical_name,
        "kind": str(expected.get("kind") or ""),
        "relative_path": relative_path,
        "size_bytes": int(before.st_size),
        "mtime_ns": int(before.st_mtime_ns),
    }
    if int(expected.get("size_bytes", -1)) != int(before.st_size):
        issues.append(
            {
                "code": "artifact_size_mismatch",
                "artifact": logical_name,
                "message": f"文件大小已变化: {relative_path}",
            }
        )
    stored_mtime_ns = expected.get("mtime_ns")
    has_stored_mtime = (
        isinstance(stored_mtime_ns, int)
        and not isinstance(stored_mtime_ns, bool)
        and stored_mtime_ns >= 0
    )
    effective_verification = (
        verification
        if verification == "full" or has_stored_mtime
        else "full"
    )
    if (
        has_stored_mtime
        and int(stored_mtime_ns) != int(before.st_mtime_ns)
    ):
        issues.append(
            {
                "code": "artifact_mtime_mismatch",
                "artifact": logical_name,
                "message": f"文件修改时间已变化: {relative_path}",
            }
        )

    if effective_verification == "full":
        current_sha256 = stream_sha256(path)
        if str(expected.get("sha256") or "") != current_sha256:
            issues.append(
                {
                    "code": "artifact_hash_mismatch",
                    "artifact": logical_name,
                    "message": f"文件 SHA-256 已变化: {relative_path}",
                }
            )

        stored_xlsx = expected.get("xlsx")
        if stored_xlsx is not None:
            current_xlsx = _xlsx_structure(path)
            if not current_xlsx.get("valid"):
                issues.append(
                    {
                        "code": "xlsx_invalid",
                        "artifact": logical_name,
                        "message": (
                            f"XLSX 结构无效: {relative_path}: "
                            f"{current_xlsx.get('error', '')}"
                        ),
                    }
                )
            elif current_xlsx != stored_xlsx:
                issues.append(
                    {
                        "code": "xlsx_structure_mismatch",
                        "artifact": logical_name,
                        "message": (
                            f"XLSX Sheet 行列结构已变化: {relative_path}"
                        ),
                    }
                )

        stored_multiset = expected.get("cow_id_multiset")
        if stored_multiset is not None and stored_xlsx is not None:
            try:
                current_multiset = compute_xlsx_identifier_multiset(
                    path,
                    str(stored_multiset.get("column_name") or ""),
                    sheet_name=str(stored_multiset.get("sheet_name") or ""),
                )
                if current_multiset != stored_multiset:
                    issues.append(
                        {
                            "code": "cow_id_multiset_mismatch",
                            "artifact": logical_name,
                            "message": f"牛号多重集已变化: {relative_path}",
                        }
                    )
            except Exception as exc:
                issues.append(
                    {
                        "code": "cow_id_multiset_invalid",
                        "artifact": logical_name,
                        "message": (
                            f"无法复核牛号多重集 {relative_path}: "
                            f"{type(exc).__name__}: {exc}"
                        ),
                    }
                )

    after = path.stat()
    if (
        before.st_size != after.st_size
        or before.st_mtime_ns != after.st_mtime_ns
    ):
        issues.append(
            {
                "code": "artifact_changed_during_validation",
                "artifact": logical_name,
                "message": f"验证期间文件发生变化: {relative_path}",
            }
        )
    return issues, artifact_state


def _validation_status(issues: Sequence[Dict[str, str]]) -> str:
    if not issues:
        return "valid"
    codes = {issue["code"] for issue in issues}
    for status, matching_codes in (
        ("manifest_invalid", {"manifest_invalid", "manifest_not_committed"}),
        ("identity_mismatch", {"identity_mismatch"}),
        ("config_mismatch", {"config_mismatch"}),
        ("artifact_missing", {"artifact_missing"}),
        (
            "artifact_mismatch",
            {
                "unsafe_artifact_path",
                "artifact_size_mismatch",
                "artifact_mtime_mismatch",
                "artifact_hash_mismatch",
                "xlsx_invalid",
                "xlsx_structure_mismatch",
                "cow_id_multiset_mismatch",
                "cow_id_multiset_invalid",
                "artifact_changed_during_validation",
            },
        ),
    ):
        if codes.intersection(matching_codes):
            return status
    return "invalid"


def validate_stage_manifest(
    root_path: PathLike,
    manifest_path: PathLike,
    *,
    expected_task_id: Optional[str] = None,
    expected_farm_code: Optional[str] = None,
    expected_stage: Optional[str] = None,
    expected_config: Any = _UNSET,
    expected_config_fingerprint: Optional[str] = None,
    verification: str = "full",
) -> Dict[str, Any]:
    """重新核验阶段身份、配置和所有输入/输出产物。

    返回 ``{"valid": bool, "status": str, "issues": [...]}``。status
    会明确区分 manifest 缺失/无效、身份不符、配置不符、文件缺失和产物
    内容不符。``full`` 会复核哈希、XLSX 结构和牛号多重集；``stat``
    只复核大小和修改时间。旧清单缺少 ``mtime_ns`` 时，对应产物会自动
    回退到 ``full``。
    """
    if verification not in {"full", "stat"}:
        raise ValueError("verification 只能是 'full' 或 'stat'")
    root = Path(root_path).resolve()
    try:
        target = _manifest_relative_path(root, manifest_path)
    except Exception as exc:
        return {
            "valid": False,
            "status": "manifest_invalid",
            "issues": [
                {
                    "code": "manifest_invalid",
                    "artifact": "",
                    "message": str(exc),
                }
            ],
        }
    if not target.is_file():
        return {
            "valid": False,
            "status": "manifest_missing",
            "issues": [
                {
                    "code": "manifest_missing",
                    "artifact": "",
                    "message": "阶段 manifest 不存在",
                }
            ],
        }

    try:
        manifest = json.loads(target.read_text(encoding="utf-8"))
        if not isinstance(manifest, dict):
            raise ValueError("manifest 根节点不是对象")
    except Exception as exc:
        return {
            "valid": False,
            "status": "manifest_invalid",
            "issues": [
                {
                    "code": "manifest_invalid",
                    "artifact": "",
                    "message": f"{type(exc).__name__}: {exc}",
                }
            ],
        }

    issues = []
    if manifest.get("schema_version") != SCHEMA_VERSION:
        issues.append(
            {
                "code": "manifest_invalid",
                "artifact": "",
                "message": "manifest schema_version 不受支持",
            }
        )
    if manifest.get("status") != "committed":
        issues.append(
            {
                "code": "manifest_not_committed",
                "artifact": "",
                "message": "阶段 manifest 尚未提交",
            }
        )
    for key in ("task_id", "farm_code", "stage", "config_fingerprint"):
        if not str(manifest.get(key) or "").strip():
            issues.append(
                {
                    "code": "manifest_invalid",
                    "artifact": "",
                    "message": f"manifest 缺少字段: {key}",
                }
            )

    expected_identity = {
        "task_id": expected_task_id,
        "farm_code": expected_farm_code,
        "stage": expected_stage,
    }
    for key, expected_value in expected_identity.items():
        if (
            expected_value is not None
            and str(manifest.get(key) or "") != str(expected_value)
        ):
            issues.append(
                {
                    "code": "identity_mismatch",
                    "artifact": "",
                    "message": f"阶段身份不一致: {key}",
                }
            )

    if (
        expected_config_fingerprint is not None
        and expected_config is not _UNSET
    ):
        raise ValueError(
            "expected_config 与 expected_config_fingerprint 只能提供一个"
        )
    if expected_config is not _UNSET:
        expected_config_fingerprint = compute_config_fingerprint(
            expected_config
        )
    if (
        expected_config_fingerprint is not None
        and str(manifest.get("config_fingerprint") or "")
        != str(expected_config_fingerprint)
    ):
        issues.append(
            {
                "code": "config_mismatch",
                "artifact": "",
                "message": "阶段配置指纹不一致",
            }
        )

    artifacts = []
    for kind in ("inputs", "outputs"):
        value = manifest.get(kind)
        if not isinstance(value, list):
            issues.append(
                {
                    "code": "manifest_invalid",
                    "artifact": "",
                    "message": f"manifest {kind} 不是列表",
                }
            )
            continue
        if kind == "outputs" and not value:
            issues.append(
                {
                    "code": "manifest_invalid",
                    "artifact": "",
                    "message": "manifest 没有输出文件",
                }
            )
        valid_items = [
            item for item in value if isinstance(item, dict)
        ]
        artifacts.extend(valid_items)
        if len(valid_items) != len(value):
            issues.append(
                {
                    "code": "manifest_invalid",
                    "artifact": "",
                    "message": "manifest 存在非对象产物记录",
                }
            )
        expected_kind = "input" if kind == "inputs" else "output"
        for item in valid_items:
            logical_name = str(item.get("logical_name") or "").strip()
            relative_path = str(item.get("relative_path") or "").strip()
            sha256 = str(item.get("sha256") or "").strip()
            size_bytes = item.get("size_bytes")
            if not logical_name or not relative_path:
                issues.append(
                    {
                        "code": "manifest_invalid",
                        "artifact": logical_name,
                        "message": "manifest 产物缺少逻辑名称或相对路径",
                    }
                )
            if item.get("kind") != expected_kind:
                issues.append(
                    {
                        "code": "manifest_invalid",
                        "artifact": logical_name,
                        "message": f"manifest 产物类型应为 {expected_kind}",
                    }
                )
            try:
                valid_sha256 = (
                    len(sha256) == 64
                    and int(sha256, 16) >= 0
                )
            except ValueError:
                valid_sha256 = False
            if not valid_sha256:
                issues.append(
                    {
                        "code": "manifest_invalid",
                        "artifact": logical_name,
                        "message": "manifest 产物 SHA-256 无效",
                    }
                )
            if (
                not isinstance(size_bytes, int)
                or isinstance(size_bytes, bool)
                or size_bytes < 0
            ):
                issues.append(
                    {
                        "code": "manifest_invalid",
                        "artifact": logical_name,
                        "message": "manifest 产物文件大小无效",
                    }
                )
            mtime_ns = item.get("mtime_ns")
            if (
                mtime_ns is not None
                and (
                    not isinstance(mtime_ns, int)
                    or isinstance(mtime_ns, bool)
                    or mtime_ns < 0
                )
            ):
                issues.append(
                    {
                        "code": "manifest_invalid",
                        "artifact": logical_name,
                        "message": "manifest 产物修改时间无效",
                    }
                )

    seen_paths = set()
    artifact_stats = []
    for artifact in artifacts:
        relative_path = str(artifact.get("relative_path") or "")
        if relative_path in seen_paths:
            issues.append(
                {
                    "code": "manifest_invalid",
                    "artifact": str(artifact.get("logical_name") or ""),
                    "message": "manifest 存在重复文件路径",
                }
            )
            continue
        seen_paths.add(relative_path)
        artifact_issues, artifact_state = _artifact_validation_issues(
            root,
            artifact,
            verification=verification,
        )
        issues.extend(artifact_issues)
        if artifact_state is not None:
            artifact_stats.append(artifact_state)

    status = _validation_status(issues)
    return {
        "valid": status == "valid",
        "status": status,
        "issues": issues,
        "manifest": manifest,
        "artifact_stats": artifact_stats,
    }

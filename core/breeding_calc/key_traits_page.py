#   core/breeding_calc/key_traits_page.py


"""
关键性状计算界面模块
整合了母牛、公牛和已配公牛的关键性状计算功能
"""

from PyQt6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QPushButton, QLabel,
    QListWidget, QListWidgetItem, QAbstractItemView, QMessageBox,
    QMainWindow
)
from PyQt6.QtCore import Qt
import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
import datetime
from pathlib import Path
from sklearn.linear_model import LinearRegression
import shutil

from core.breeding_calc.traits_calculation import TraitsCalculation
from core.data.update_manager import LOCAL_DB_PATH
from gui.progress import ProgressDialog

# 复用已有的性状翻译字典
TRAITS_TRANSLATION = {
    'TPI': '育种综合指数', 'NM$': '净利润值', 'CM$': '奶酪净值', 'FM$': '液奶净值', 'GM$': '放牧净值',
    'MILK': '产奶量', 'FAT': '乳脂量', 'PROT': '乳蛋白量', 'FAT %': '乳脂率', 'PROT%': '乳蛋白率',
    'SCS': '体细胞指数', 'DPR': '女儿怀孕率', 'HCR': '青年牛受胎率', 'CCR': '成母牛受胎率', 'PL': '生产寿命',
    'SCE': '配种产犊难易度', 'DCE': '女儿产犊难易度', 'SSB': '配种死胎率', 'DSB': '女儿死淘率',
    'PTAT': '体型综合指数', 'UDC': '乳房综合指数', 'FLC': '肢蹄综合指数', 'BDC': '体重综合指数',
    'ST': '体高', 'SG': '强壮度', 'BD': '体深', 'DF': '乳用特征', 'RA': '尻角度', 'RW': '尻宽',
    'LS': '后肢侧视', 'LR': '后肢后视', 'FA': '蹄角度', 'FLS': '肢蹄评分', 'FU': '前方附着',
    'UH': '后房高度', 'UW': '后方宽度', 'UC': '悬韧带', 'UD': '乳房深度', 'FT': '前乳头位置',
    'RT': '后乳头位置', 'TL': '乳头长度', 'FE': '饲料效率指数', 'FI': '繁殖力指数', 'HI': '健康指数',
    'LIV': '存活率', 'GL': '妊娠长度', 'MAST': '乳房炎抗病性', 'MET': '子宫炎抗病性', 'RP': '胎衣不下抗病性',
    'KET': '酮病抗病性', 'DA': '变胃抗病性', 'MFV': '产后瘫抗病性', 'EFC': '首次产犊月龄',
    'HLiv': '后备牛存活率', 'FS': '饲料节约指数', 'RFI': '剩余饲料采食量', 'Milk Speed': '排乳速度',
    'Eval Date': '数据日期'
}

class KeyTraitsPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # 添加按钮属性
        self.cow_calc_btn = None
        self.bull_calc_btn = None
        self.breeding_bull_calc_btn = None
        
        self.default_traits = [
            'NM$', 'TPI', 'MILK', 'FAT', 'FAT %', 'PROT', 'PROT%', 
            'SCS', 'PL', 'DPR', 'PTAT', 'UDC', 'FLC', 'RFI', 'Eval Date'
        ]
        
        self.all_traits = [
            'TPI', 'NM$', 'CM$', 'FM$', 'GM$', 'MILK', 'FAT', 'PROT', 
            'FAT %', 'PROT%', 'SCS', 'DPR', 'HCR', 'CCR', 'PL', 'SCE', 
            'DCE', 'SSB', 'DSB', 'PTAT', 'UDC', 'FLC', 'BDC', 'ST', 'SG', 
            'BD', 'DF', 'RA', 'RW', 'LS', 'LR', 'FA', 'FLS', 'FU', 'UH', 
            'UW', 'UC', 'UD', 'FT', 'RT', 'TL', 'FE', 'FI', 'HI', 'LIV', 
            'GL', 'MAST', 'MET', 'RP', 'KET', 'DA', 'MFV', 'EFC', 'HLiv', 
            'FS', 'RFI', 'Milk Speed', 'Eval Date'
        ]
        
        # 必选性状（用于生成正态分布图，不能被移除）
        self.required_traits = ['NM$', 'TPI']
        self.required_trait = 'NM$'  # 保持向后兼容
        self.setup_ui()

    def setup_ui(self):
        layout = QHBoxLayout(self)
        
        # 左侧：全部性状列表
        left_layout = self.create_left_panel()
        layout.addLayout(left_layout)
        
        # 中间：按钮区域
        button_layout = self.create_middle_panel()
        layout.addLayout(button_layout)
        
        # 右侧：已选择性状列表和计算按钮
        right_layout = self.create_right_panel()
        layout.addLayout(right_layout)
        
        # 设置布局的伸缩因子
        layout.setStretch(0, 2)  # 左侧列表
        layout.setStretch(1, 1)  # 中间按钮区域
        layout.setStretch(2, 2)  # 右侧列表

    def create_left_panel(self):
        layout = QVBoxLayout()
        left_label = QLabel("全部性状")
        self.all_traits_list = QListWidget()
        
        # 添加性状，显示中英文
        for trait in self.all_traits:
            item = QListWidgetItem(f"{trait} - {TRAITS_TRANSLATION[trait]}")
            item.setData(Qt.ItemDataRole.UserRole, trait)
            self.all_traits_list.addItem(item)
            
        self.all_traits_list.itemDoubleClicked.connect(self.add_trait)
        layout.addWidget(left_label)
        layout.addWidget(self.all_traits_list)
        return layout

    def create_middle_panel(self):
        layout = QVBoxLayout()
        
        # 操作按钮
        add_button = QPushButton("添加 >>")
        remove_button = QPushButton("<< 移除")
        select_all_button = QPushButton("全选")
        reset_button = QPushButton("恢复默认")
        
        # 连接信号
        add_button.clicked.connect(self.add_trait)
        remove_button.clicked.connect(self.remove_trait)
        select_all_button.clicked.connect(self.select_all_traits)
        reset_button.clicked.connect(self.reset_traits)
        
        # 设置按钮样式
        button_style = """
            QPushButton {
                padding: 8px 16px;
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """
        for btn in [add_button, remove_button, select_all_button, reset_button]:
            btn.setStyleSheet(button_style)
        
        # 添加到布局
        layout.addStretch()
        for btn in [add_button, remove_button, select_all_button, reset_button]:
            layout.addWidget(btn)
        layout.addStretch()
        
        return layout

    def create_right_panel(self):
        layout = QVBoxLayout()
        
        # 上部分：已选择性状列表
        right_label = QLabel("已选择性状\n（可拖拽调整顺序）")
        self.selected_traits_list = QListWidget()
        
        # 添加默认性状
        for trait in self.default_traits:
            item = QListWidgetItem(f"{trait} - {TRAITS_TRANSLATION[trait]}")
            item.setData(Qt.ItemDataRole.UserRole, trait)
            self.selected_traits_list.addItem(item)
            
        self.selected_traits_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.selected_traits_list.itemDoubleClicked.connect(self.remove_trait)
        
        layout.addWidget(right_label)
        layout.addWidget(self.selected_traits_list)
        
        # 下部分：计算按钮
        calc_button_layout = QVBoxLayout()
        
        # 创建三个计算按钮
        self.cow_calc_btn = QPushButton("在群母牛关键性状计算")
        self.bull_calc_btn = QPushButton("备选公牛关键性状计算")
        self.breeding_bull_calc_btn = QPushButton("已配公牛关键性状计算")
        
        # 设置按钮样式
        button_style = """
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                margin: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
                color: #95a5a6;
            }
        """
        
        for btn in [self.cow_calc_btn, self.bull_calc_btn, self.breeding_bull_calc_btn]:
            btn.setStyleSheet(button_style)
        
        # 连接信号
        self.cow_calc_btn.clicked.connect(self.start_cow_traits_calculation)
        self.bull_calc_btn.clicked.connect(self.start_bull_traits_calculation)
        self.breeding_bull_calc_btn.clicked.connect(self.start_breeding_bull_calculation)
        
        # 默认禁用所有按钮
        for btn in [self.cow_calc_btn, self.bull_calc_btn, self.breeding_bull_calc_btn]:
            btn.setEnabled(False)
        
        # 添加按钮到布局
        calc_button_layout = QVBoxLayout()
        calc_button_layout.addWidget(self.cow_calc_btn)
        calc_button_layout.addWidget(self.bull_calc_btn)
        calc_button_layout.addWidget(self.breeding_bull_calc_btn)
        
        layout.addLayout(calc_button_layout)
        
        return layout

    def add_trait(self, item=None):
        if not item:
            item = self.all_traits_list.currentItem()
        if item:
            trait = item.data(Qt.ItemDataRole.UserRole)
            if not self.is_trait_selected(trait):
                new_item = QListWidgetItem(f"{trait} - {TRAITS_TRANSLATION[trait]}")
                new_item.setData(Qt.ItemDataRole.UserRole, trait)
                self.selected_traits_list.addItem(new_item)

    def remove_trait(self, item=None):
        if not item:
            item = self.selected_traits_list.currentItem()
        if item:
            trait = item.data(Qt.ItemDataRole.UserRole)
            # 检查是否为必选性状（NM$ 和 TPI 用于生成正态分布图）
            if trait not in self.required_traits:
                self.selected_traits_list.takeItem(self.selected_traits_list.row(item))

    def is_trait_selected(self, trait):
        for i in range(self.selected_traits_list.count()):
            item = self.selected_traits_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == trait:
                return True
        return False

    def select_all_traits(self):
        self.selected_traits_list.clear()
        for trait in self.all_traits:
            item = QListWidgetItem(f"{trait} - {TRAITS_TRANSLATION[trait]}")
            item.setData(Qt.ItemDataRole.UserRole, trait)
            self.selected_traits_list.addItem(item)

    def reset_traits(self):
        self.selected_traits_list.clear()
        for trait in self.default_traits:
            item = QListWidgetItem(f"{trait} - {TRAITS_TRANSLATION[trait]}")
            item.setData(Qt.ItemDataRole.UserRole, trait)
            self.selected_traits_list.addItem(item)

    def get_selected_traits(self):
        return [self.selected_traits_list.item(i).data(Qt.ItemDataRole.UserRole) 
                for i in range(self.selected_traits_list.count())]


    def update_calculation_buttons_state(self):
        """更新计算按钮的可用状态"""
        print(f"update_calculation_buttons_state called. selected_project_path: {self.get_main_window().selected_project_path}") # 打印当前选中的项目路径

        main_window = self.get_main_window()
        if not main_window or not main_window.selected_project_path:
            for btn in [self.cow_calc_btn, self.bull_calc_btn, self.breeding_bull_calc_btn]:
                btn.setEnabled(False)
            return

        project_path = main_window.selected_project_path

        # 检查各个数据文件是否存在
        cow_data_exists = (project_path / "standardized_data" / "processed_cow_data.xlsx").exists()
        bull_data_exists = (project_path / "standardized_data" / "processed_bull_data.xlsx").exists()
        breeding_data_exists = (project_path / "standardized_data" / "processed_breeding_data.xlsx").exists()

        print(f"  Cow data exists: {cow_data_exists}")
        print(f"  Bull data exists: {bull_data_exists}")
        print(f"  Breeding data exists: {breeding_data_exists}")

        # 更新按钮状态
        self.cow_calc_btn.setEnabled(cow_data_exists)
        self.bull_calc_btn.setEnabled(bull_data_exists)
        self.breeding_bull_calc_btn.setEnabled(breeding_data_exists)

    def get_main_window(self):
        """获取主窗口实例"""
        parent = self.parent()
        while parent and not isinstance(parent, QMainWindow):
            parent = parent.parent()
        return parent

    def save_with_formatting(self, df, output_path):
        """保存Excel文件并应用格式化

        Args:
            df: DataFrame
            output_path: 输出路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        # 创建一个副本，用于移除source列
        df_to_save = df.copy()

        # 记录source列的信息
        source_info = {}
        for bull_type in ['sire', 'mgs', 'mmgs']:
            source_col = f"{bull_type}_source"
            if source_col in df_to_save.columns:
                source_info[bull_type] = df_to_save[source_col].to_list()
                # 删除source列（用户不需要看到）
                df_to_save = df_to_save.drop(columns=[source_col])

        # 用pandas写入基础数据
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_to_save.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # 查找所有需要格式化的列
            traits_cols = []

            for col_idx, col_name in enumerate(df_to_save.columns, 1):
                # 查找性状列
                for bull_type in ['sire', 'mgs', 'mmgs']:
                    if col_name.startswith(f"{bull_type}_") and not col_name.endswith('_identified'):
                        traits_cols.append((col_idx, col_name, bull_type))

            # 应用格式化
            red_font = Font(color="FF0000")  # 红色字体
            yellow_font = Font(color="FFFF00")  # 亮黄色字体
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # 黑色背景

            # 遍历数据行（从第2行开始，第1行是标题）
            for row_idx in range(2, len(df_to_save) + 2):
                data_idx = row_idx - 2  # DataFrame中的索引

                for col_idx, col_name, bull_type in traits_cols:
                    if bull_type in source_info and data_idx < len(source_info[bull_type]):
                        source_value = source_info[bull_type][data_idx]
                        cell = worksheet.cell(row=row_idx, column=col_idx)

                        if source_value == 2:
                            # 年份预估值 - 红色字体
                            cell.font = red_font
                        elif source_value == 3:
                            # 默认值 - 黑底黄字
                            cell.font = yellow_font
                            cell.fill = black_fill

    def save_results_with_retry(self, df, output_path, apply_formatting=True):
        """尝试保存结果，如果文件被占用提示用户并重试

        Args:
            df: 要保存的DataFrame
            output_path: 输出路径
            apply_formatting: 是否应用格式化（颜色标记）
        """
        while True:
            try:
                if apply_formatting and any(col.endswith('_source') for col in df.columns):
                    # 使用ExcelWriter应用格式
                    self.save_with_formatting(df, output_path)
                else:
                    df.to_excel(output_path, index=False)
                print(f"结果已保存到: {output_path}")
                return True
            except PermissionError:
                reply = QMessageBox.question(
                    self,
                    "文件被占用",
                    f"文件 {output_path.name} 正在被其他程序使用。\n请关闭该文件后点击'重试'继续，或点击'取消'停止操作。",
                    QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel
                )
                if reply == QMessageBox.StandardButton.Cancel:
                    return False
        
    def fill_estimated_values(self, cow_df, yearly_data_dict, selected_traits, engine):
        """
        填充预估值到缺失的育种值字段

        Args:
            cow_df: 母牛数据DataFrame
            yearly_data_dict: 年度平均值字典 {trait: DataFrame}
            selected_traits: 选中的性状列表
            engine: 数据库引擎

        Returns:
            处理后的DataFrame，包含预估值和标记
        """
        print("开始填充预估值...")

        # 1. 转换年度数据格式
        yearly_data = {}
        if isinstance(yearly_data_dict, dict):
            for trait in selected_traits:
                if trait in yearly_data_dict:
                    df_trait = yearly_data_dict[trait]
                    # 创建年份到平均值的映射
                    yearly_data[trait] = df_trait.set_index('birth_year')['mean'].to_dict()
            print(f"成功处理年度数据，包含 {len(yearly_data)} 个性状")
        else:
            print("年度数据格式错误")
            return cow_df

        # 2. 获取默认值（从999HO99999）
        default_values = {}
        with engine.connect() as conn:
            default_bull = conn.execute(
                text("SELECT * FROM bull_library WHERE `BULL NAAB`='999HO99999'")
            ).fetchone()
            if default_bull:
                default_bull_dict = dict(default_bull._mapping)
                for trait in selected_traits:
                    default_values[trait] = default_bull_dict.get(trait, 0)
            else:
                print("警告: 未找到默认公牛999HO99999")
                for trait in selected_traits:
                    default_values[trait] = 0

        # 3. 确保有birth_year, dam_birth_year, mgd_birth_year列
        if 'birth_year' not in cow_df.columns:
            cow_df['birth_year'] = pd.to_datetime(cow_df['birth_date']).dt.year
        if 'dam_birth_year' not in cow_df.columns:
            cow_df['dam_birth_year'] = pd.to_datetime(cow_df['birth_date_dam']).dt.year
        if 'mgd_birth_year' not in cow_df.columns:
            cow_df['mgd_birth_year'] = pd.to_datetime(cow_df['birth_date_mgd']).dt.year

        # 4. 为每个缺失的育种值填充预估值，并创建来源标记
        for bull_type, year_col in [('sire', 'birth_year'),
                                     ('mgs', 'dam_birth_year'),
                                     ('mmgs', 'mgd_birth_year')]:
            # 创建来源标记列（1=真实值，2=年份预估，3=默认值）
            cow_df[f"{bull_type}_source"] = 1  # 默认为真实值

            # 只处理identified=False的记录
            if f"{bull_type}_identified" in cow_df.columns:
                mask_not_identified = cow_df[f"{bull_type}_identified"] == False

                for trait in selected_traits:
                    trait_col = f"{bull_type}_{trait}"
                    if trait_col in cow_df.columns:
                        # 找到需要填充的行（identified=False且值为空）
                        mask_need_fill = mask_not_identified & cow_df[trait_col].isna()

                        # 对需要填充的行进行处理
                        for idx in cow_df[mask_need_fill].index:
                            year = cow_df.loc[idx, year_col]

                            if pd.notna(year) and trait in yearly_data:
                                year = int(year)
                                if year in yearly_data[trait]:
                                    # 使用年份预估值
                                    cow_df.loc[idx, trait_col] = yearly_data[trait][year]
                                    cow_df.loc[idx, f"{bull_type}_source"] = 2
                                else:
                                    # 年份不在数据中，使用默认值
                                    cow_df.loc[idx, trait_col] = default_values[trait]
                                    cow_df.loc[idx, f"{bull_type}_source"] = 3
                            else:
                                # 没有年份信息，使用默认值
                                cow_df.loc[idx, trait_col] = default_values[trait]
                                cow_df.loc[idx, f"{bull_type}_source"] = 3

        print("预估值填充完成")
        return cow_df

    def process_cow_key_traits_by_year(self, data_source, output_path=None, progress_callback=None):
        """处理年度关键性状数据

        Args:
            data_source: DataFrame或文件路径
            output_path: 输出文件路径（可选）
            progress_callback: 进度回调函数
        """
        print("开始处理年度关键性状数据...")

        try:
            # 1. 读取数据
            if isinstance(data_source, pd.DataFrame):
                df = data_source.copy()
            else:
                df = pd.read_excel(data_source)
            
            # 2. 获取出生年份范围
            min_year = df['birth_year'].min()
            max_year = df['birth_year'].max()
            print(f"出生年份范围: {min_year} - {max_year}")

            # 3. 准备所有性状的数据
            results = {}
            
            # 4. 获取默认值（从999HO99999）
            default_values = {}
            engine = create_engine(f'sqlite:///{LOCAL_DB_PATH}')
            with engine.connect() as conn:
                default_bull = conn.execute(
                    text("SELECT * FROM bull_library WHERE `BULL NAAB`='999HO99999'")
                ).fetchone()
                if default_bull:
                    default_bull_dict = dict(default_bull._mapping)
                    for trait in self.get_selected_traits():
                        default_values[trait] = default_bull_dict.get(trait, 0)
                else:
                    print("警告: 未找到默认公牛999HO99999")
                    for trait in self.get_selected_traits():
                        default_values[trait] = 0
                        
            # 5. 处理每个性状
            total_traits = len(self.get_selected_traits())
            for idx, trait in enumerate(self.get_selected_traits(), 1):
                # 更新进度
                if progress_callback:
                    progress = int((idx / total_traits) * 100)
                    progress_callback(progress)
                    
                print(f"处理性状: {trait}")
                trait_col = f'sire_{trait}'
                
                # 5.1 如果性状列不存在，直接使用默认值处理所有年份
                if trait_col not in df.columns:
                    print(f"性状列 {trait_col} 不存在，使用默认值")
                    all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                    all_years['mean'] = default_values[trait]
                    all_years['interpolated'] = True
                else:
                    # 只统计identified=true的记录来计算遗传趋势
                    # 确保sire_identified列存在
                    if 'sire_identified' in df.columns:
                        df_identified = df[df['sire_identified'] == True]
                    else:
                        # 如果没有identified列，使用所有非空数据
                        df_identified = df[df[trait_col].notna()]

                    yearly_means = df_identified.groupby('birth_year').agg({
                        trait_col: ['count', 'mean']
                    }).reset_index()
                    yearly_means.columns = ['birth_year', 'count', 'mean']
                    
                    # 5.2 筛选数据量>=10的年份
                    valid_years = yearly_means[yearly_means['count'] >= 10]
                    
                    if len(valid_years) > 1:
                        # 有多个有效年份，进行正常回归
                        X = valid_years['birth_year'].values.reshape(-1, 1)
                        y = valid_years['mean'].values
                        reg = LinearRegression().fit(X, y)
                        
                        # 生成所有年份的预测值
                        all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                        all_years['mean'] = reg.predict(all_years['birth_year'].values.reshape(-1, 1))
                        all_years['interpolated'] = ~all_years['birth_year'].isin(valid_years['birth_year'])
                        
                    elif len(valid_years) == 1:
                        # 只有一个有效年份
                        valid_year = valid_years['birth_year'].iloc[0]
                        valid_mean = valid_years['mean'].iloc[0]
                        
                        if valid_year == min_year:
                            # 使用最小年份+1的默认值
                            points = {
                                valid_year: valid_mean,
                                valid_year + 1: default_values[trait]
                            }
                            slope = default_values[trait] - valid_mean  # 计算斜率
                        else:
                            # 使用最小年份的默认值
                            points = {
                                min_year: default_values[trait],
                                valid_year: valid_mean
                            }
                            slope = (valid_mean - default_values[trait]) / (valid_year - min_year)  # 计算斜率
                        
                        # 使用这两个点创建线性方程
                        all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                        all_years['mean'] = all_years['birth_year'].apply(
                            lambda x: points.get(x, 
                                # 如果年份大于最大已知年份，使用线性外推
                                valid_mean + slope * (x - valid_year) if x > valid_year else
                                # 否则使用线性插值
                                np.interp(x, list(points.keys()), list(points.values())))
                        )
                        all_years['interpolated'] = ~all_years['birth_year'].isin([valid_year])
                        
                    else:
                        # 没有有效年份，全部使用默认值
                        all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                        all_years['mean'] = default_values[trait]
                        all_years['interpolated'] = True
                
                # 保存处理结果到字典中
                results[trait] = all_years
            
            # 6. 尝试保存所有结果到Excel
            if output_path is None:
                # 如果没有提供输出路径，不保存
                return results

            # 确保output_path是Path对象
            if not isinstance(output_path, Path):
                output_path = Path(output_path)
            while True:
                try:
                    with pd.ExcelWriter(output_path) as writer:
                        for trait, data in results.items():
                            data.to_excel(writer, sheet_name=trait, index=False)
                    print("年度关键性状数据处理完成")
                    break
                except PermissionError:
                    reply = QMessageBox.question(
                        self,
                        "文件被占用",
                        f"文件 {output_path.name} 正在被其他程序使用。\n请关闭该文件后点击'重试'继续，或点击'取消'停止操作。",
                        QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel
                    )
                    if reply == QMessageBox.StandardButton.Cancel:
                        return False
            
            return results  # 返回结果字典，供后续使用

        except Exception as e:
            print(f"处理年度关键性状数据时发生错误: {str(e)}")
            return False

    def perform_cow_traits_calculation(self, main_window, progress_callback=None, task_info_callback=None):
        """执行关键性状计算的核心逻辑"""
        engine = None   # 初始化engine为None
        try:
            # 1. 获取所选性状列表
            selected_traits = self.get_selected_traits()
            print(f"获取到所选性状: {selected_traits}")

            # 2. 读取母牛数据，获取公牛列表
            cow_data_path = main_window.selected_project_path / "standardized_data" / "processed_cow_data.xlsx"
            print(f"母牛数据文件路径: {cow_data_path}")
            
            if not cow_data_path.exists():
                raise FileNotFoundError("请先上传并处理母牛数据")
            
            print("开始读取母牛数据...")
            cow_df = pd.read_excel(cow_data_path)
            
            # 添加birth_year列
            cow_df['birth_year'] = pd.to_datetime(cow_df['birth_date']).dt.year
            
            # 添加sire_identified, mgs_identified, mmgs_identified列
            for col in ['sire', 'mgs', 'mmgs']:
                cow_df[f"{col}_identified"] = False
            
            bull_ids = set()
            for col in ['sire', 'mgs', 'mmgs']:
                if col in cow_df.columns:
                    bull_ids.update(cow_df[col].dropna().unique())
            print(f"提取到的公牛ID数量: {len(bull_ids)}")
            
            # 3. 检查数据库中的公牛,并提取所选性状数据
            print("开始检查数据库中的公牛...")
            engine = create_engine(f'sqlite:///{LOCAL_DB_PATH}')
            bull_traits = {}
            with engine.connect() as conn:
                missing_bulls = []
                for bull_id in bull_ids:
                    if pd.isna(bull_id):  # 跳过空值
                        continue
                    if len(str(bull_id)) > 10:
                        print(f"检查长ID公牛 {bull_id} (BULL REG)")
                        result = conn.execute(
                            text("SELECT * FROM bull_library WHERE `BULL REG`=:bull_id"),
                            {"bull_id": bull_id}
                        ).fetchone()
                    else:
                        print(f"检查短ID公牛 {bull_id} (BULL NAAB)")
                        result = conn.execute(
                            text("SELECT * FROM bull_library WHERE `BULL NAAB`=:bull_id"),
                            {"bull_id": bull_id}
                        ).fetchone()
                        
                    if result:
                        # 如果找到了公牛,提取所选性状数据
                        result_dict = dict(result._mapping)
                        bull_traits[str(bull_id)] = {trait: result_dict.get(trait) for trait in selected_traits}
                    else:
                        missing_bulls.append(bull_id)
                
            print(f"缺失公牛数量: {len(missing_bulls)}")
                        
            # 4. 如果有缺失的公牛，创建缺失公牛的DataFrame并更新到云端数据库
            if missing_bulls:
                print("开始处理缺失公牛...")
                try:
                    # 通过API上传缺失公牛信息
                    from api.api_client import get_api_client
                    api_client = get_api_client()

                    username = main_window.username if hasattr(main_window, 'username') else 'unknown'

                    # 准备上传数据
                    bulls_data = []
                    for bull_id in missing_bulls:
                        bulls_data.append({
                            'bull': bull_id,
                            'source': 'traits_calculation',
                            'time': datetime.datetime.now().isoformat(),
                            'user': username
                        })

                    # 调用API上传
                    print("开始通过API更新缺失公牛到云端数据库...")
                    success = api_client.upload_missing_bulls(bulls_data)

                    if success:
                        print("缺失公牛更新完成")
                    else:
                        print("警告：缺失公牛上传失败")
                except Exception as e:
                    print(f"上传缺失公牛信息时发生错误：{e}")
                
            # 5. 将提取到的公牛性状数据合并到母牛数据中
            print("开始合并性状数据...")
            for bull_type in ['sire', 'mgs', 'mmgs']:
                print(f"处理 {bull_type} 的性状数据")
                for trait in selected_traits:
                    cow_df[f"{bull_type}_{trait}"] = cow_df[bull_type].apply(
                        lambda x: bull_traits.get(str(x), {}).get(trait) if pd.notna(x) else np.nan
                    )
                    print(f"{bull_type}_{trait} 的前5行数据:")
                    print(cow_df[f"{bull_type}_{trait}"].head())
                
                # 更新 identified 列
                cow_df[f"{bull_type}_identified"] = cow_df[bull_type].apply(
                    lambda x: str(x) in bull_traits if pd.notna(x) else False
                )

            # 5.5 先计算并保存年度平均值（供预估使用）
            print("计算年度平均值用于预估...")
            yearly_output_path = main_window.selected_project_path / "analysis_results" / "sire_traits_mean_by_cow_birth_year.xlsx"
            yearly_data_dict = self.process_cow_key_traits_by_year(cow_df, yearly_output_path)
            if yearly_data_dict is False:
                return False, "用户取消了年度数据保存操作"

            # 5.6 使用年度平均值填充缺失的育种值
            print("填充预估值...")
            cow_df = self.fill_estimated_values(cow_df, yearly_data_dict, selected_traits, engine)

            # 6. 保存详细结果
            detail_output_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_detail.xlsx"
            if not self.save_results_with_retry(cow_df, detail_output_path):
                return False, "用户取消了保存操作"

            # 7. 处理年度平均值和回归分析（第二次调用，从文件读取）
            yearly_output_path = main_window.selected_project_path / "analysis_results" / "sire_traits_mean_by_cow_birth_year.xlsx"
            # 注意：这里不需要再次处理，因为已经在5.5步骤处理过了

            # 8. 计算关键性状得分
            pedigree_output_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx"
            if not self.calculate_cow_key_traits_scores(detail_output_path, yearly_output_path):
                return False, "用户取消了得分计算操作"
            
            # 9. 用基因组数据更新关键性状得分
            genomic_data_path = main_window.selected_project_path / "standardized_data" / "processed_genomic_data.xlsx"
            if genomic_data_path.exists():
                if not self.update_with_genomic_data(pedigree_output_path, genomic_data_path):
                    return False, "用户取消了基因组数据更新操作"
            else:
                print("未找到基因组数据文件，跳过基因组数据更新")

            # 10. 生成最终育种值文件（包含所有母牛育种值）
            try:
                print("生成最终育种值文件...")
                genomic_scores_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_scores_genomic.xlsx"
                pedigree_scores_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx"
                final_output_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_final.xlsx"

                # 优先使用genomic文件，如果不存在则使用pedigree文件
                if genomic_scores_path.exists():
                    source_path = genomic_scores_path
                    source_type = "genomic"
                elif pedigree_scores_path.exists():
                    source_path = pedigree_scores_path
                    source_type = "pedigree"
                else:
                    print("未找到scores文件，跳过生成final文件")
                    source_path = None

                if source_path:
                    # 复制为final文件
                    shutil.copy(source_path, final_output_path)
                    print(f"✓ 已生成最终育种值文件: processed_cow_data_key_traits_final.xlsx (来源: {source_type})")
            except Exception as e:
                print(f"生成最终育种值文件失败: {e}")
                # 不影响主流程，继续执行

            # 11. 自动生成系谱识别分析结果（在合并育种值之后）
            try:
                from core.breeding_calc.generate_pedigree_analysis import generate_pedigree_analysis_result
                print("自动生成系谱识别分析结果...")
                generate_pedigree_analysis_result(main_window.selected_project_path)
                print("✓ 系谱识别分析结果已生成")
            except Exception as e:
                print(f"生成系谱识别分析结果失败: {e}")
                # 不影响主流程，继续执行

            return True, "计算完成"

        except Exception as e:
            return False, str(e)
        finally:
            if engine:
                engine.dispose()

    def on_calculation_finished(self):
        """计算完成的处理"""
        self.progress_dialog.close()
        QMessageBox.information(self, "完成", "关键性状计算流程已完成")

    def on_calculation_error(self, error_message):
        """计算错误的处理"""
        self.progress_dialog.close()
        QMessageBox.critical(self, "错误", f"处理过程中发生错误：{error_message}")
    def calculate_cow_key_traits_scores(self, detail_path, yearly_path, progress_callback=None):
        """计算母牛关键性状得分"""
        print("开始计算母牛关键性状得分...")
        
        try:
            # 1. 读取详细数据文件和年度数据文件
            print("读取详细数据文件...")
            df = pd.read_excel(detail_path)
            print(f"已读取 detail_path: {detail_path}")
            
            # 打印所有 sire_, mgs_, mmgs_ 相关列的数据类型
            for col in df.columns:
                if 'sire_' in col or 'mgs_' in col or 'mmgs_' in col:
                    print(f"Data type of column {col}: {df[col].dtype}")
                    # 如果该列是 object 类型，进一步检查唯一值
                    if df[col].dtype == 'object':
                        print(f"Unique values in column {col}: {df[col].unique()}")


            # 2. 确保所有日期列的格式正确
            print("处理日期列...")
            date_columns = ['birth_date', 'birth_date_dam', 'birth_date_mgd']
            for col in date_columns:
                if col in df.columns:
                    # 尝试转换日期，对于无效日期使用NaT
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    # 记录无效日期的数量
                    invalid_dates = df[col].isna().sum()
                    if invalid_dates > 0:
                        print(f"警告: {col} 列中有 {invalid_dates} 个无效日期")
            
            # 3. 从日期提取年份，同时处理无效日期
            print("提取年份...")
            df['birth_year'] = df['birth_date'].dt.year
            df['dam_birth_year'] = df['birth_date_dam'].dt.year
            df['mgd_birth_year'] = df['birth_date_mgd'].dt.year
            
            # 4. 读取所有年度数据到字典中
            print("读取年度数据...")
            yearly_data = {}
            with pd.ExcelFile(yearly_path) as xls:
                for trait in self.get_selected_traits():
                    yearly_data[trait] = pd.read_excel(xls, sheet_name=trait)
                    # 将birth_year设为索引以便快速查找
                    yearly_data[trait].set_index('birth_year', inplace=True)
                    # 记录可用的年份范围
                    print(f"性状 {trait} 的年度数据范围: {yearly_data[trait].index.min()} - {yearly_data[trait].index.max()}")
                    # 打印该性状的 yearly_data[trait]
                    print(f"Yearly data for {trait}:")
                    print(yearly_data[trait])           
            print(f"已读取 yearly_path: {yearly_path}")


            # 5. 获取默认公牛999HO99999的性状值
            print("获取默认公牛数据...")
            engine = create_engine(f'sqlite:///{LOCAL_DB_PATH}')
            default_values = {}
            with engine.connect() as conn:
                default_bull = conn.execute(
                    text("SELECT * FROM bull_library WHERE `BULL NAAB`='999HO99999'")
                ).fetchone()
                if default_bull:
                    default_bull_dict = dict(default_bull._mapping)
                    print(f"default_bull_dict: {default_bull_dict}") # 打印 default_bull_dict
                    print(f"default_bull_dict type: {type(default_bull_dict)}") # 打印 default_bull_dict 的类型
                    
                    for trait in self.get_selected_traits():
                        default_values[trait] = default_bull_dict.get(trait, 0)
                    print(f"获取到默认公牛性状值: {default_values}")
                else:
                    print("警告: 未找到默认公牛999HO99999，使用0作为默认值")
                    for trait in self.get_selected_traits():
                        default_values[trait] = 0

            # 6. 设置权重
            weights = {
                'sire': 0.5,
                'mgs': 0.25,
                'mmgs': 0.125,
                'default': 0.125  # 999HO99999的权重
            }
            
            # 7. 为每个性状计算加权得分
            print("开始计算加权得分...")
            total_traits = len(self.get_selected_traits())
            total_rows = len(df)
            
            for trait_idx, trait in enumerate(self.get_selected_traits(), 1):
                print(f"处理性状: {trait}")
                score_column = f'{trait}_score'
                df[score_column] = 0.0  # 初始化得分列
                
                # 确保性状相关列为数值类型
                df[f'sire_{trait}'] = pd.to_numeric(df[f'sire_{trait}'], errors='coerce')
                df[f'mgs_{trait}'] = pd.to_numeric(df[f'mgs_{trait}'], errors='coerce')
                df[f'mmgs_{trait}'] = pd.to_numeric(df[f'mmgs_{trait}'], errors='coerce')



                for index, row in df.iterrows():
                    # 更新进度
                    if progress_callback:
                        current_position = (trait_idx - 1) * total_rows + index + 1
                        progress = int((current_position / (total_traits * total_rows)) * 100)
                        progress_callback(progress)

                    trait_score = 0.0

                    # 详细的 print 调试
                    print(f"------ 处理行索引: {index}, 性状: {trait} ------")

                    sire_col = f'sire_{trait}'
                    mgs_col = f'mgs_{trait}'
                    mmgs_col = f'mmgs_{trait}'

                    print(f"  sire_col ({sire_col}) value: {row[sire_col]}, type: {type(row[sire_col])}")
                    print(f"  mgs_col ({mgs_col}) value: {row[mgs_col]}, type: {type(row[mgs_col])}")
                    print(f"  mmgs_col ({mmgs_col}) value: {row[mmgs_col]}, type: {type(row[mmgs_col])}")

                    # 处理sire
                    if pd.isna(row[sire_col]):
                        print(f"  sire_{trait} 为空 (NaN)")
                        if pd.notna(row['birth_year']):
                            birth_year = int(row['birth_year'])
                            print(f"    birth_year: {birth_year}")
                            if birth_year in yearly_data[trait].index:
                                yearly_value = yearly_data[trait].loc[birth_year, 'mean']
                                print(f"    yearly_data 中 {birth_year} 年的 {trait} 值为: {yearly_value}, 类型: {type(yearly_value)}")
                                trait_score += weights['sire'] * yearly_value
                            else:
                                print(f"    警告: 未找到 {birth_year} 年的年度数据")
                                trait_score += weights['sire'] * default_values[trait]
                        else:
                            print("    birth_year 为空 (NaN)")
                            trait_score += weights['sire'] * default_values[trait]
                    else:
                        print(f"  sire_{trait} 非空，尝试计算")
                        trait_score += weights['sire'] * row[sire_col]

                    # 处理mgs
                    if pd.isna(row[mgs_col]):
                        print(f"  mgs_{trait} 为空 (NaN)")
                        if pd.notna(row['dam_birth_year']):
                            dam_birth_year = int(row['dam_birth_year'])
                            print(f"    dam_birth_year: {dam_birth_year}")
                            if dam_birth_year in yearly_data[trait].index:
                                yearly_value = yearly_data[trait].loc[dam_birth_year, 'mean']
                                print(f"    yearly_data 中 {dam_birth_year} 年的 {trait} 值为: {yearly_value}, 类型: {type(yearly_value)}")
                                trait_score += weights['mgs'] * yearly_value
                            else:
                                print(f"    警告: 未找到 {dam_birth_year} 年的年度数据")
                                trait_score += weights['mgs'] * default_values[trait]
                        else:
                            print("    dam_birth_year 为空 (NaN)")
                            trait_score += weights['mgs'] * default_values[trait]
                    else:
                        print(f"  mgs_{trait} 非空，尝试计算")
                        trait_score += weights['mgs'] * row[mgs_col]


                    # 处理mmgs
                    if pd.isna(row[mmgs_col]):
                        print(f"  mmgs_{trait} 为空 (NaN)")
                        if pd.notna(row['mgd_birth_year']):
                            mgd_birth_year = int(row['mgd_birth_year'])
                            print(f"    mgd_birth_year: {mgd_birth_year}")
                            if mgd_birth_year in yearly_data[trait].index:
                                yearly_value = yearly_data[trait].loc[mgd_birth_year, 'mean']
                                print(f"    yearly_data 中 {mgd_birth_year} 年的 {trait} 值为: {yearly_value}, 类型: {type(yearly_value)}")
                                trait_score += weights['mmgs'] * yearly_value
                            else:
                                print(f"    警告: 未找到 {mgd_birth_year} 年的年度数据")
                                trait_score += weights['mmgs'] * default_values[trait]
                        else:
                            print("    mgd_birth_year 为空 (NaN)")
                            trait_score += weights['mmgs'] * default_values[trait]
                    else:
                        print(f"  mmgs_{trait} 非空，尝试计算")
                        trait_score += weights['mmgs'] * row[mmgs_col]

                    # 加上默认值
                    print(f"  default_values[{trait}] 的值: {default_values[trait]}, 类型: {type(default_values[trait])}")
                    trait_score += weights['default'] * default_values[trait]

                    print(f"  本行 {trait} 的 trait_score: {trait_score}")
                    df.at[index, score_column] = trait_score

                # 打印每个性状的得分统计信息
                print(f"性状 {trait} 得分统计:")
                print(df[score_column].describe())
                    
            # 8. 保存结果
            print("保存计算结果...")
            output_path = detail_path.parent / "processed_cow_data_key_traits_scores_pedigree.xlsx"
            if not self.save_results_with_retry(df, output_path):
                print("用户取消了得分数据保存操作")
                return False
                
            print("母牛关键性状得分计算完成")
            return True
            
        except Exception as e:
            print(f"计算母牛关键性状得分时发生错误: {str(e)}")
            raise

    def update_with_genomic_data(self, pedigree_path, genomic_path, progress_callback=None):
        """用基因组数据更新关键性状得分"""
        print("开始用基因组数据更新关键性状得分...")
        
        try:
            # 1. 读取系谱计算的得分文件
            df_pedigree = pd.read_excel(pedigree_path)
            
            # 2. 读取基因组数据
            df_genomic = pd.read_excel(genomic_path)
            
            # 3. 为所有性状得分添加来源列，默认为"P"(系谱)
            score_columns = [col for col in df_pedigree.columns if col.endswith('_score')]
            for col in score_columns:
                df_pedigree[f"{col}_source"] = "P"
                
            # 4. 用基因组数据更新得分
            total_traits = len(self.get_selected_traits())
            total_rows = len(df_pedigree)

            for trait_idx, trait in enumerate(self.get_selected_traits(), 1):
                score_col = f"{trait}_score"
                source_col = f"{trait}_score_source"

                # 遍历每一行数据
                for row_idx, (idx, row) in enumerate(df_pedigree.iterrows(), 1):
                    # 更新进度
                    if progress_callback:
                        current_position = (trait_idx - 1) * total_rows + row_idx
                        progress = int((current_position / (total_traits * total_rows)) * 100)
                        progress_callback(progress)
                        
                    cow_id = row['cow_id']
                    # 查找对应的基因组数据
                    genomic_row = df_genomic[df_genomic['cow_id'] == cow_id]
                    
                    if not genomic_row.empty and trait in genomic_row.columns:
                        # 如果找到基因组数据且该性状有值
                        if pd.notna(genomic_row[trait].iloc[0]):
                            # 更新得分和来源标记
                            df_pedigree.at[idx, score_col] = genomic_row[trait].iloc[0]
                            df_pedigree.at[idx, source_col] = "G"
                            print(f"母牛 {cow_id} 的 {trait} 更新为基因组数据")
            
            # 5. 添加一个总结列，显示有多少个性状使用了基因组数据
            genomic_count = df_pedigree[[col for col in df_pedigree.columns if col.endswith('_source')]].apply(
                lambda x: (x == "G").sum(), axis=1
            )
            df_pedigree['genomic_traits_count'] = genomic_count
            
            # 6. 保存结果
            output_path = pedigree_path.parent / "processed_cow_data_key_traits_scores_genomic.xlsx"
            if not self.save_results_with_retry(df_pedigree, output_path):
                print("用户取消了基因组数据更新结果保存")
                return False
            
            print("基因组数据更新完成")
            return True
            
        except Exception as e:
            print(f"更新基因组数据时发生错误: {str(e)}")
            raise


    def start_cow_traits_calculation(self):
            """开始母牛关键性状计算流程"""
            # 获取主窗口实例
            main_window = self.get_main_window()
            if not main_window:
                QMessageBox.warning(self, "警告", "无法获取主窗口")
                return

            if not hasattr(main_window, 'selected_project_path') or not main_window.selected_project_path:
                QMessageBox.warning(self, "警告", "请先选择一个项目")
                return

            try:
                # 每次计算前创建新的计算器实例
                self.traits_calculator = TraitsCalculation()
                # 创建进度对话框
                self.progress_dialog = ProgressDialog(self)
                self.progress_dialog.show()

                # 获取选中的性状列表
                selected_traits = self.get_selected_traits()
                print(f"start_cow_traits_calculation called. selected_traits: {selected_traits}") # 打印选择的性状

                # 定义进度回调函数，能接收进度值和消息
                def progress_callback(progress_value, message=None):
                    print(f"[DEBUG-CALLBACK] 收到回调: progress={progress_value}, message={message}")  # 添加调试输出
                    if progress_value is not None:
                        print(f"[DEBUG-CALLBACK] 更新进度条: {progress_value}")
                        self.progress_dialog.update_progress(progress_value)
                    if message is not None:
                        print(f"[DEBUG-CALLBACK] 更新信息: {message}")
                        self.progress_dialog.update_info(message)

                # 先测试回调函数是否工作
                print("[DEBUG-CALLBACK] 测试回调函数...")
                progress_callback(10, "测试消息：回调函数工作正常")

                # 执行计算
                success, message = self.traits_calculator.process_data(
                    main_window,
                    selected_traits,
                    progress_callback=progress_callback
                )

                if success:
                    self.progress_dialog.close()
                    QMessageBox.information(self, "完成", "关键性状计算完成！")
                else:
                    self.progress_dialog.close()
                    QMessageBox.critical(self, "错误", f"计算过程中发生错误：{message}")

            except Exception as e:
                self.progress_dialog.close()
                print(f"start_cow_traits_calculation: Exception occurred: {e}") # 打印异常信息
                QMessageBox.critical(self, "错误", f"处理过程中发生错误：{str(e)}")
            finally:
                # 清理资源
                if hasattr(self, 'progress_dialog'):
                    self.progress_dialog.close()
                if hasattr(self.traits_calculator, 'engine'):
                    self.traits_calculator.engine.dispose()

    def start_bull_traits_calculation(self):
        """开始公牛关键性状计算流程"""
        # 获取主窗口实例
        main_window = self.get_main_window()
        if not main_window:
            QMessageBox.warning(self, "警告", "无法获取主窗口")
            return
            
        if not main_window.selected_project_path:
            QMessageBox.warning(self, "警告", "请先选择一个项目")
            return

        try:
            # 检查备选公牛数据文件是否存在
            bull_data_path = main_window.selected_project_path / "standardized_data" / "processed_bull_data.xlsx"
            if not bull_data_path.exists():
                QMessageBox.warning(self, "警告", "未找到备选公牛数据文件，请先上传并处理备选公牛数据")
                return

            # 读取备选公牛数据
            try:
                bull_df = pd.read_excel(bull_data_path)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取备选公牛数据文件失败：{str(e)}")
                return

            # 获取选中的性状列表
            selected_traits = self.get_selected_traits()
            if not selected_traits:
                QMessageBox.warning(self, "警告", "请至少选择一个性状")
                return

            # 连接本地数据库
            try:
                engine = create_engine(f'sqlite:///{LOCAL_DB_PATH}')
            except Exception as e:
                QMessageBox.critical(self, "错误", f"连接本地数据库失败：{str(e)}")
                return

            # 处理每个公牛的性状数据
            missing_bulls = []
            trait_data = []

            # 创建进度对话框
            self.progress_dialog = ProgressDialog(self)
            self.progress_dialog.show()

            try:
                total_bulls = len(bull_df)
                for idx, row in bull_df.iterrows():
                    # 更新进度
                    if self.progress_dialog:
                        progress = int((idx + 1) / total_bulls * 100)
                        self.progress_dialog.update_progress(progress)

                    bull_id = str(row['bull_id'])
                    if pd.isna(bull_id) or bull_id.strip() == '':
                        continue

                    with engine.connect() as conn:
                        # 先尝试用 BULL NAAB 查询
                        result = conn.execute(
                            text("SELECT * FROM bull_library WHERE `BULL NAAB`=:bull_id"),
                            {"bull_id": bull_id}
                        ).fetchone()
                        
                        if not result:
                            # 如果找不到，再尝试用 BULL REG 查询
                            result = conn.execute(
                                text("SELECT * FROM bull_library WHERE `BULL REG`=:bull_id"),
                                {"bull_id": bull_id}
                            ).fetchone()

                        if result:
                            # 提取性状数据
                            bull_data = dict(row)  # 保留原始数据
                            result_dict = dict(result._mapping)
                            for trait in selected_traits:
                                bull_data[trait] = result_dict.get(trait)
                            trait_data.append(bull_data)
                        else:
                            missing_bulls.append(bull_id)
                            # 对于缺失的公牛，保留原始数据，性状值设为空
                            bull_data = dict(row)
                            for trait in selected_traits:
                                bull_data[trait] = None
                            trait_data.append(bull_data)

            finally:
                if hasattr(self, 'progress_dialog'):
                    self.progress_dialog.close()

            # 如果有缺失的公牛，上传到云端数据库并提示用户
            if missing_bulls:
                try:
                    # 通过API上传缺失公牛信息
                    from api.api_client import get_api_client
                    api_client = get_api_client()

                    username = main_window.username if hasattr(main_window, 'username') else 'unknown'

                    # 准备上传数据
                    bulls_data = []
                    for bull_id in missing_bulls:
                        bulls_data.append({
                            'bull': bull_id,
                            'source': 'bull_key_traits',
                            'time': datetime.datetime.now().isoformat(),
                            'user': username
                        })

                    # 调用API上传
                    success = api_client.upload_missing_bulls(bulls_data)
                    
                    if success:
                        # 提示用户
                        QMessageBox.warning(
                            self,
                            "警告",
                            f"以下公牛在数据库中未找到：\n{', '.join(missing_bulls)}"
                        )
                    else:
                        raise Exception("API上传失败")

                except Exception as e:
                    QMessageBox.warning(
                        self,
                        "警告",
                        f"上传缺失公牛信息时发生错误：{str(e)}\n"
                        f"缺失的公牛：{', '.join(missing_bulls)}"
                    )

            # 生成结果DataFrame
            result_df = pd.DataFrame(trait_data)

            # 保存结果文件
            output_path = main_window.selected_project_path / "analysis_results" / "processed_bull_data_key_traits.xlsx"
            if not self.save_results_with_retry(result_df, output_path):
                QMessageBox.warning(self, "警告", "用户取消了保存操作")
                return

            QMessageBox.information(self, "成功", "公牛关键性状计算完成！")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"计算过程中发生错误：{str(e)}")
            return
        finally:
            if hasattr(self, 'progress_dialog'):
                self.progress_dialog.close()
            if engine:
                engine.dispose()

    def start_breeding_bull_calculation(self):
        """开始已配公牛关键性状计算"""
        # 获取主窗口实例
        main_window = self.get_main_window()
        if not main_window:
            QMessageBox.warning(self, "警告", "无法获取主窗口")
            return
        
        if not hasattr(main_window, 'selected_project_path') or not main_window.selected_project_path:
            QMessageBox.warning(self, "警告", "请先选择一个项目")
            return

        # 检查配种记录文件是否存在
        breeding_data_path = main_window.selected_project_path / "standardized_data" / "processed_breeding_data.xlsx"
        if not breeding_data_path.exists():
            QMessageBox.warning(self, "警告", "未找到配种记录数据文件，请先上传并处理配种记录数据")
            return

        QMessageBox.information(
            self, 
            "功能开发中", 
            "已配公牛关键性状计算功能正在开发中...\n\n"
            "此功能将根据配种记录中使用的公牛计算关键性状，\n"
            "并按年份区分进行分析。"
        )
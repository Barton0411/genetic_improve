# core/breeding_calc/traits_calculation.py

from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from typing import Tuple, Optional
from sqlalchemy import create_engine, text
from PyQt6.QtWidgets import QMessageBox
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .base_calculation import BaseCowCalculation

class TraitsCalculation(BaseCowCalculation):
    def __init__(self):
        super().__init__()
        self.output_prefix = "processed_cow_data_key_traits"
        self.required_columns = ['cow_id', 'birth_date', 'birth_date_dam', 'birth_date_mgd']
        self.yearly_filename = "sire_traits_mean_by_cow_birth_year.xlsx"  # 新的年度数据文件名

    def process_data(self, main_window, selected_traits: list, progress_callback=None, task_info_callback=None) -> Tuple[bool, str]:
        """执行关键性状计算的核心逻辑"""
        try:
            # 1. 初始化必要的设置
            print("1. 初始化数据库连接...")
            if task_info_callback:
                task_info_callback("初始化数据库连接")
            if progress_callback:
                progress_callback(5, "正在初始化数据库连接...")
            
            if not self.init_db_connection():
                return False, "连接数据库失败"

            project_path = main_window.selected_project_path
            success, error_msg = self.check_project_data(project_path, "processed_cow_data.xlsx")
            if not success:
                return False, error_msg

            # 2. 读取母牛数据，获取公牛列表
            print("2. 读取母牛数据...")
            if task_info_callback:
                task_info_callback("读取母牛数据")
            if progress_callback:
                progress_callback(10, "正在读取母牛数据文件...")
                
            cow_df = self.read_data(project_path, "processed_cow_data.xlsx")
            if cow_df is None:
                return False, "读取母牛数据失败"
            
            print(f"成功读取 {len(cow_df)} 条母牛记录")
            if progress_callback:
                progress_callback(15, f"成功读取 {len(cow_df)} 条母牛记录")
            
            # 添加birth_year列
            cow_df['birth_year'] = pd.to_datetime(cow_df['birth_date']).dt.year
            
            # 添加sire_identified等列
            for col in ['sire', 'mgs', 'mmgs']:
                cow_df[f"{col}_identified"] = False
            
            # 获取所有公牛ID
            bull_ids = set()
            for col in ['sire', 'mgs', 'mmgs']:
                if col in cow_df.columns:
                    bull_ids.update(cow_df[col].dropna().unique())
            
            print(f"共发现 {len(bull_ids)} 个公牛ID")
            if progress_callback:
                progress_callback(20, f"数据预处理完成，共发现 {len(bull_ids)} 个公牛ID")

            # 3. 批量处理公牛性状数据
            print("3. 处理公牛性状数据...")
            if task_info_callback:
                task_info_callback("处理公牛性状数据")
            if progress_callback:
                progress_callback(25, "开始批量查询公牛性状数据...")

            bull_traits = self.get_bull_traits_batch(list(bull_ids), selected_traits)
            print(f"批量查询完成，获取到 {len(bull_traits)} 个公牛的性状数据")

            if progress_callback:
                progress_callback(45, f"公牛性状数据处理完成，获取到 {len(bull_traits)} 个公牛的性状数据")

            # 4. 更新母牛数据中的公牛性状
            print("4. 更新母牛数据中的公牛性状...")
            if task_info_callback:
                task_info_callback("更新母牛数据中的公牛性状")
            if progress_callback:
                progress_callback(50, "开始更新母牛数据中的公牛性状...")
            
            for col in ['sire', 'mgs', 'mmgs']:
                for trait in selected_traits:
                    trait_col = f'{col}_{trait}'
                    # 确保类型匹配，将x转为字符串再查询
                    cow_df[trait_col] = cow_df[col].map(lambda x: bull_traits.get(str(x), {}).get(trait) if pd.notna(x) else None)
                    identified_count = cow_df[trait_col].notna().sum()
                    print(f"{trait_col}: 找到 {identified_count} 条记录")

                # 更新 identified 标记：如果公牛在 bull_traits 中找到，则标记为 True
                cow_df[f"{col}_identified"] = cow_df[col].apply(
                    lambda x: str(x) in bull_traits if pd.notna(x) else False
                )
                identified_bull_count = cow_df[f"{col}_identified"].sum()
                print(f"{col}_identified: 标记了 {identified_bull_count} 头公牛被找到")

            if progress_callback:
                progress_callback(60, "母牛数据中的公牛性状更新完成")

            # 5. 先处理年度数据（需要先计算年度数据，用于填充预估值）
            print("5. 处理年度数据...")
            if task_info_callback:
                task_info_callback("处理年度数据")
            if progress_callback:
                progress_callback(65, "正在处理年度数据...")

            output_dir = project_path / "analysis_results"
            output_dir.mkdir(exist_ok=True)
            yearly_output_path = output_dir / self.yearly_filename

            # 优化：直接传DataFrame，避免临时文件读写
            if not self.process_yearly_data_from_df(cow_df, yearly_output_path, selected_traits):
                return False, "处理年度数据失败"

            # 6. 填充预估值
            print("6. 填充缺失公牛的预估值...")
            if task_info_callback:
                task_info_callback("填充缺失公牛预估值")
            if progress_callback:
                progress_callback(70, "正在填充缺失公牛的预估值...")

            cow_df = self.fill_estimated_values(cow_df, yearly_output_path, selected_traits)

            # 7. 保存详细结果（带格式）
            print("7. 保存详细计算结果...")
            if task_info_callback:
                task_info_callback("保存详细计算结果")
            if progress_callback:
                progress_callback(75, "正在保存详细计算结果...")

            detail_output_path = output_dir / f"{self.output_prefix}_detail.xlsx"
            if not self.save_results_with_retry(cow_df, detail_output_path, apply_formatting=False):
                return False, "保存详细结果失败"

            if progress_callback:
                progress_callback(70, f"详细结果已保存到: {detail_output_path.name}")

            if progress_callback:
                progress_callback(80, f"详细结果已保存到: {detail_output_path.name}")

            # 8. 计算性状得分
            print("8. 计算性状得分...")
            if task_info_callback:
                task_info_callback("计算性状得分")
            if progress_callback:
                progress_callback(85, "正在计算性状得分...")
                
            pedigree_output_path = output_dir / f"{self.output_prefix}_scores_pedigree.xlsx"
            # 优化：直接传DataFrame，避免重复读取Excel文件
            if not self.calculate_trait_scores_from_df(cow_df, yearly_output_path, pedigree_output_path,
                                                       apply_formatting=False, selected_traits=selected_traits):
                return False, "计算性状得分失败"

            if progress_callback:
                progress_callback(90, f"性状得分计算完成，保存到: {pedigree_output_path.name}")

            # 9. 检查并处理基因组数据
            print("9. 检查基因组数据...")
            if task_info_callback:
                task_info_callback("检查并处理基因组数据")
            if progress_callback:
                progress_callback(95, "正在检查基因组数据...")
                
            genomic_data_path = project_path / "standardized_data" / "processed_genomic_data.xlsx"
            genomic_output_path = output_dir / f"{self.output_prefix}_scores_genomic.xlsx"

            if genomic_data_path.exists():
                print("发现基因组数据，开始更新...")
                if progress_callback:
                    progress_callback(97, "发现基因组数据，正在更新...")
                if not self.update_genomic_data(pedigree_output_path, genomic_data_path, genomic_output_path, apply_formatting=False):
                    return False, "更新基因组数据失败"
                print("基因组数据更新完成")
                if progress_callback:
                    progress_callback(99, f"基因组数据更新完成，保存到: {genomic_output_path.name}")
            else:
                print("未找到基因组数据文件，创建基于系谱的基因组文件...")
                if progress_callback:
                    progress_callback(97, "未找到基因组数据，创建占位文件...")

                # 即使没有基因组数据，也创建genomic文件（内容与pedigree相同）
                if not self.create_genomic_placeholder(pedigree_output_path, genomic_output_path, apply_formatting=False):
                    return False, "创建基因组占位文件失败"

                print("基因组占位文件已创建")
                if progress_callback:
                    progress_callback(99, f"基因组文件已创建（仅系谱数据）: {genomic_output_path.name}")

            # 10. 生成最终育种值文件
            import shutil
            try:
                print("生成最终育种值文件...")
                genomic_scores_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_scores_genomic.xlsx"
                pedigree_scores_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx"
                final_output_path = main_window.selected_project_path / "analysis_results" / "processed_cow_data_key_traits_final.xlsx"

                if genomic_scores_path.exists():
                    shutil.copy(genomic_scores_path, final_output_path)
                    print(f"✓ 已生成最终育种值文件 (来源: genomic)")
                elif pedigree_scores_path.exists():
                    shutil.copy(pedigree_scores_path, final_output_path)
                    print(f"✓ 已生成最终育种值文件 (来源: pedigree)")
                else:
                    print("未找到scores文件，跳过生成final文件")
            except Exception as e:
                print(f"生成最终育种值文件失败: {e}")

            # 11. 自动生成系谱识别分析结果
            try:
                from core.breeding_calc.generate_pedigree_analysis import generate_pedigree_analysis_result
                print("自动生成系谱识别分析结果...")
                generate_pedigree_analysis_result(main_window.selected_project_path)
                print("✓ 系谱识别分析结果已生成")
            except Exception as e:
                print(f"生成系谱识别分析结果失败: {e}")

            # 12. 自动生成关键育种性状分析结果
            try:
                from core.breeding_calc.generate_key_traits_analysis import generate_key_traits_analysis_result
                print("自动生成关键育种性状分析结果...")
                generate_key_traits_analysis_result(main_window.selected_project_path)
                print("✓ 关键育种性状分析结果已生成")
            except Exception as e:
                print(f"生成关键育种性状分析结果失败: {e}")

            if progress_callback:
                progress_callback(100, "所有计算步骤已完成！")
            if task_info_callback:
                task_info_callback("计算完成")
            print("所有计算步骤已完成!")
            return True, "计算完成"

        except Exception as e:
            print(f"计算过程出错: {str(e)}")
            if progress_callback:
                progress_callback(None, f"计算过程出错: {str(e)}")
            return False, str(e)

    def process_yearly_data(self, detail_path: Path, output_path: Path, selected_traits: list) -> bool:
        """处理年度关键性状数据"""
        try:
            # 读取详细数据
            df = pd.read_excel(detail_path)
            
            # 检查birth_year列是否存在
            if 'birth_year' not in df.columns:
                print("错误：数据中缺少birth_year列")
                return False
            
            # 清理birth_year数据，移除空值和无效值
            df = df.dropna(subset=['birth_year'])
            df['birth_year'] = pd.to_numeric(df['birth_year'], errors='coerce')
            df = df.dropna(subset=['birth_year'])
            
            # 过滤合理的年份范围（1900-当前年份+10）
            current_year = datetime.datetime.now().year
            df = df[(df['birth_year'] >= 1900) & (df['birth_year'] <= current_year + 10)]
            
            if len(df) == 0:
                print("错误：没有有效的出生年份数据")
                return False
            
            # 获取出生年份范围
            min_year = int(df['birth_year'].min())
            max_year = int(df['birth_year'].max())
            
            # 必需的性状（用于生成正态分布图）
            required_traits = ['NM$', 'TPI']

            # 确保必需性状被包含在处理列表中
            all_traits_to_process = list(selected_traits)
            for req_trait in required_traits:
                if req_trait not in all_traits_to_process:
                    all_traits_to_process.append(req_trait)
                    print(f"信息: 自动添加必需性状 {req_trait} 到处理列表")

            # 获取默认值（包含所有要处理的性状）
            default_values = self.get_default_values(all_traits_to_process)

            # 处理每个性状
            results = {}
            for trait in all_traits_to_process:
                trait_col = f'sire_{trait}'

                if trait_col not in df.columns:
                    # 使用默认值
                    all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                    all_years['mean'] = default_values[trait]
                    all_years['interpolated'] = True
                else:
                    # 仅使用 sire_identified 为 True 的记录来计算年度平均值
                    if 'sire_identified' in df.columns:
                        df_identified = df[df['sire_identified'] == True]
                        print(f"处理性状 {trait}: 总记录 {len(df)} 条，其中 identified {len(df_identified)} 条")
                    else:
                        # 如果没有 identified 列，使用有数据的记录
                        df_identified = df[df[trait_col].notna()]
                        print(f"处理性状 {trait}: 使用有数据的记录 {len(df_identified)} 条")

                    if len(df_identified) == 0:
                        # 如果没有 identified 记录，使用默认值
                        all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                        all_years['mean'] = default_values[trait]
                        all_years['interpolated'] = True
                    else:
                        # 计算年度均值（仅使用 identified 记录）
                        yearly_means = df_identified.groupby('birth_year').agg({
                            trait_col: ['count', 'mean']
                        }).reset_index()
                        yearly_means.columns = ['birth_year', 'count', 'mean']

                        # 处理数据：传递所有年度数据，让process_trait_yearly_data决定哪些需要插值
                        all_years = self.process_trait_yearly_data(
                            yearly_means, min_year, max_year, default_values[trait]
                        )

                results[trait] = all_years
            
            # 保存结果
            return self.save_yearly_results(results, output_path)

        except Exception as e:
            import traceback
            print(f"处理年度数据失败: {e}")
            print(f"详细错误信息: {traceback.format_exc()}")
            return False

    def process_yearly_data_from_df(self, df: pd.DataFrame, output_path: Path, selected_traits: list) -> bool:
        """处理年度关键性状数据 - 优化版本，直接从DataFrame处理，避免临时文件"""
        try:
            # 创建副本避免修改原始数据
            df = df.copy()

            # 检查birth_year列是否存在
            if 'birth_year' not in df.columns:
                print("错误：数据中缺少birth_year列")
                return False

            # 清理birth_year数据，移除空值和无效值
            df = df.dropna(subset=['birth_year'])
            df['birth_year'] = pd.to_numeric(df['birth_year'], errors='coerce')
            df = df.dropna(subset=['birth_year'])

            # 过滤合理的年份范围（1900-当前年份+10）
            current_year = datetime.datetime.now().year
            df = df[(df['birth_year'] >= 1900) & (df['birth_year'] <= current_year + 10)]

            if len(df) == 0:
                print("错误：没有有效的出生年份数据")
                return False

            # 获取出生年份范围
            min_year = int(df['birth_year'].min())
            max_year = int(df['birth_year'].max())

            # 必需的性状（用于生成正态分布图）
            required_traits = ['NM$', 'TPI']

            # 确保必需性状被包含在处理列表中
            all_traits_to_process = list(selected_traits)
            for req_trait in required_traits:
                if req_trait not in all_traits_to_process:
                    all_traits_to_process.append(req_trait)
                    print(f"信息: 自动添加必需性状 {req_trait} 到处理列表")

            # 获取默认值（包含所有要处理的性状）
            default_values = self.get_default_values(all_traits_to_process)

            # 处理每个性状
            results = {}
            for trait in all_traits_to_process:
                trait_col = f'sire_{trait}'

                if trait_col not in df.columns:
                    # 使用默认值
                    all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                    all_years['mean'] = default_values[trait]
                    all_years['interpolated'] = True
                else:
                    # 仅使用 sire_identified 为 True 的记录来计算年度平均值
                    if 'sire_identified' in df.columns:
                        df_identified = df[df['sire_identified'] == True]
                        print(f"处理性状 {trait}: 总记录 {len(df)} 条，其中 identified {len(df_identified)} 条")
                    else:
                        # 如果没有 identified 列，使用有数据的记录
                        df_identified = df[df[trait_col].notna()]
                        print(f"处理性状 {trait}: 使用有数据的记录 {len(df_identified)} 条")

                    if len(df_identified) == 0:
                        # 如果没有 identified 记录，使用默认值
                        all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
                        all_years['mean'] = default_values[trait]
                        all_years['interpolated'] = True
                    else:
                        # 计算年度均值（仅使用 identified 记录）
                        yearly_means = df_identified.groupby('birth_year').agg({
                            trait_col: ['count', 'mean']
                        }).reset_index()
                        yearly_means.columns = ['birth_year', 'count', 'mean']

                        # 处理数据：传递所有年度数据，让process_trait_yearly_data决定哪些需要插值
                        all_years = self.process_trait_yearly_data(
                            yearly_means, min_year, max_year, default_values[trait]
                        )

                results[trait] = all_years

            # 保存结果
            return self.save_yearly_results(results, output_path)

        except Exception as e:
            import traceback
            print(f"处理年度数据失败: {e}")
            print(f"详细错误信息: {traceback.format_exc()}")
            return False

    def process_trait_yearly_data(self, yearly_data: pd.DataFrame, min_year: int,
                                max_year: int, default_value: float) -> pd.DataFrame:
        """处理单个性状的年度数据

        Args:
            yearly_data: 包含birth_year, count, mean列的DataFrame
            min_year: 最小年份
            max_year: 最大年份
            default_value: 默认值

        Returns:
            包含所有年份数据的DataFrame，保留实际值，仅对缺失或样本量<10的年份插值
        """
        # 创建完整年份范围的DataFrame
        all_years = pd.DataFrame({'birth_year': range(min_year, max_year + 1)})
        all_years['count'] = 0
        all_years['mean'] = np.nan
        all_years['interpolated'] = False

        # 填充实际数据 - 使用向量化操作替代循环
        if not yearly_data.empty:
            # 设置索引以便快速查找
            all_years_indexed = all_years.set_index('birth_year')
            yearly_data_indexed = yearly_data.set_index('birth_year')

            # 找出共同的年份
            common_years = all_years_indexed.index.intersection(yearly_data_indexed.index)

            # 批量更新
            all_years_indexed.loc[common_years, 'count'] = yearly_data_indexed.loc[common_years, 'count']
            all_years_indexed.loc[common_years, 'mean'] = yearly_data_indexed.loc[common_years, 'mean']

            # 重置索引
            all_years = all_years_indexed.reset_index()

        # 找出需要插值的年份（没有数据或样本量<10）
        needs_interpolation = (all_years['count'] < 10) | all_years['mean'].isna()

        # 找出有效的年份（样本量>=10）
        valid_years = all_years[all_years['count'] >= 10].copy()

        if len(valid_years) > 1:
            # 有多个有效年份，使用线性回归进行插值
            X = valid_years['birth_year'].values.reshape(-1, 1)
            y = valid_years['mean'].values
            reg = LinearRegression().fit(X, y)

            # 仅对需要插值的年份应用回归模型
            for idx in all_years[needs_interpolation].index:
                year = all_years.loc[idx, 'birth_year']
                all_years.loc[idx, 'mean'] = reg.predict([[year]])[0]
                all_years.loc[idx, 'interpolated'] = True

        elif len(valid_years) == 1:
            # 只有一个有效年份，使用简单插值
            valid_year = valid_years['birth_year'].iloc[0]
            valid_mean = valid_years['mean'].iloc[0]

            # 计算斜率
            if valid_year == min_year:
                slope = (default_value - valid_mean) / max(1, max_year - min_year)
            elif valid_year == max_year:
                slope = (valid_mean - default_value) / max(1, max_year - min_year)
            else:
                # 中间年份，向两边延伸
                slope = (default_value - valid_mean) / max(abs(valid_year - min_year), abs(max_year - valid_year))

            # 仅对需要插值的年份应用
            for idx in all_years[needs_interpolation].index:
                year = all_years.loc[idx, 'birth_year']
                if year < valid_year:
                    all_years.loc[idx, 'mean'] = np.interp(year, [min_year, valid_year], [default_value, valid_mean])
                else:
                    all_years.loc[idx, 'mean'] = valid_mean + slope * (year - valid_year)
                all_years.loc[idx, 'interpolated'] = True

        else:
            # 没有有效年份（所有样本量都<10），所有年份使用默认值
            all_years.loc[needs_interpolation, 'mean'] = default_value
            all_years.loc[needs_interpolation, 'interpolated'] = True

        # 保留count列用于显示数据质量
        return all_years[['birth_year', 'mean', 'count', 'interpolated']]

    def get_default_values(self, selected_traits: list) -> dict:
        """获取默认值（从999HO99999）"""
        default_values = {}
        try:
            with self.db_engine.connect() as conn:
                default_bull = conn.execute(
                    text("SELECT * FROM bull_library WHERE `BULL NAAB`='999HO99999'")
                ).fetchone()
                if default_bull:
                    default_bull_dict = dict(default_bull._mapping)
                    for trait in selected_traits:
                        default_values[trait] = default_bull_dict.get(trait, 0)
                else:
                    for trait in selected_traits:
                        default_values[trait] = 0
        except Exception as e:
            print(f"获取默认值失败: {e}")
            for trait in selected_traits:
                default_values[trait] = 0
        
        return default_values

    def save_yearly_results(self, results: dict, output_path: Path) -> bool:
        """保存年度结果"""
        try:
            with pd.ExcelWriter(output_path) as writer:
                for trait, data in results.items():
                    data.to_excel(writer, sheet_name=trait, index=False)
            return True
        except Exception as e:
            print(f"保存年度结果失败: {e}")
            return False

    def calculate_trait_scores(self, detail_path: Path, yearly_path: Path,
                             output_path: Path, apply_formatting: bool = False,
                             selected_traits: list = None) -> bool:
        """计算性状得分

        Args:
            detail_path: 详细数据文件路径
            yearly_path: 年度数据文件路径
            output_path: 输出文件路径
            apply_formatting: 是否应用格式化
            selected_traits: 选择的性状列表（用于确保生成所有性状的得分）
        """
        try:
            df = pd.read_excel(detail_path)

            # 处理日期列
            date_columns = ['birth_date', 'birth_date_dam', 'birth_date_mgd']
            for col in date_columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')

            df['birth_year'] = df['birth_date'].dt.year
            df['dam_birth_year'] = df['birth_date_dam'].dt.year
            df['mgd_birth_year'] = df['birth_date_mgd'].dt.year

            # 读取年度数据
            yearly_data = {}
            with pd.ExcelFile(yearly_path) as xls:
                for sheet in xls.sheet_names:
                    yearly_data[sheet] = pd.read_excel(xls, sheet_name=sheet)
                    yearly_data[sheet].set_index('birth_year', inplace=True)

            # 必需的性状（用于生成正态分布图）
            required_traits = ['NM$', 'TPI']

            # 确定要处理的性状列表：优先使用 selected_traits 参数
            if selected_traits:
                all_traits_to_process = list(selected_traits)
                # 确保必需性状被包含
                for req_trait in required_traits:
                    if req_trait not in all_traits_to_process:
                        all_traits_to_process.append(req_trait)
            else:
                # 回退到从 yearly_data 获取（保持向后兼容）
                all_traits_to_process = list(yearly_data.keys())

            # 确保必需的性状存在于 yearly_data 中
            min_year = int(df['birth_year'].min()) if df['birth_year'].notna().any() else 2020
            max_year = int(df['birth_year'].max()) if df['birth_year'].notna().any() else 2024

            # 为所有需要处理的性状创建年度数据（如果不存在）
            for trait in all_traits_to_process:
                if trait not in yearly_data:
                    print(f"警告: yearly 数据中缺少 {trait}，使用默认值创建")
                    # 创建默认的年度数据
                    default_yearly = pd.DataFrame({
                        'birth_year': range(min_year, max_year + 1),
                        'mean': [0] * (max_year - min_year + 1)
                    })
                    default_yearly.set_index('birth_year', inplace=True)
                    yearly_data[trait] = default_yearly

            # 获取默认值（包含所有要处理的性状）
            default_values = self.get_default_values(all_traits_to_process)
            
            # 设置权重
            weights = {
                'sire': 0.5,
                'mgs': 0.25,
                'mmgs': 0.125,
                'default': 0.125
            }
            
            # 计算得分并保留source信息（使用 all_traits_to_process 确保处理所有选择的性状）
            for trait in all_traits_to_process:
                score_column = f'{trait}_score'
                df[score_column] = self.calculate_single_trait_score(
                    df, trait, yearly_data[trait], default_values[trait], weights
                )

                # 复制source列（如果存在）- 使用向量化操作替代apply()
                for bull_type in ['sire', 'mgs', 'mmgs']:
                    source_col = f'{bull_type}_{trait}_source'
                    if source_col not in df.columns:
                        # 如果source列不存在，根据是否有值来判断
                        trait_col = f'{bull_type}_{trait}'
                        if trait_col in df.columns:
                            # 确定年份列
                            if bull_type == 'sire':
                                year_col = 'birth_year'
                            elif bull_type == 'mgs':
                                year_col = 'dam_birth_year'
                            else:
                                year_col = 'mgd_birth_year'

                            # 使用向量化操作设置source值
                            has_trait = df[trait_col].notna()
                            has_year = df[year_col].notna() if year_col in df.columns else pd.Series(False, index=df.index)
                            # source=1 如果有trait值，否则 source=2 如果有年份，否则 source=3
                            df[source_col] = np.where(has_trait, 1, np.where(has_year, 2, 3))
            
            return self.save_results_with_retry(df, output_path, apply_formatting=apply_formatting)

        except Exception as e:
            print(f"计算性状得分失败: {e}")
            return False

    def calculate_trait_scores_from_df(self, df: pd.DataFrame, yearly_path: Path,
                                       output_path: Path, apply_formatting: bool = False,
                                       selected_traits: list = None) -> bool:
        """计算性状得分 - 优化版本，直接从DataFrame处理，避免重复读取Excel

        Args:
            df: 详细数据DataFrame（已在内存中）
            yearly_path: 年度数据文件路径
            output_path: 输出文件路径
            apply_formatting: 是否应用格式化
            selected_traits: 选择的性状列表（用于确保生成所有性状的得分）
        """
        try:
            # 创建副本避免修改原始数据
            df = df.copy()

            # 处理日期列（如果尚未处理）
            date_columns = ['birth_date', 'birth_date_dam', 'birth_date_mgd']
            for col in date_columns:
                if col in df.columns and not pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = pd.to_datetime(df[col], errors='coerce')

            if 'birth_year' not in df.columns or df['birth_year'].isna().all():
                df['birth_year'] = df['birth_date'].dt.year if 'birth_date' in df.columns else None
            if 'dam_birth_year' not in df.columns:
                df['dam_birth_year'] = df['birth_date_dam'].dt.year if 'birth_date_dam' in df.columns else None
            if 'mgd_birth_year' not in df.columns:
                df['mgd_birth_year'] = df['birth_date_mgd'].dt.year if 'birth_date_mgd' in df.columns else None

            # 读取年度数据
            yearly_data = {}
            with pd.ExcelFile(yearly_path) as xls:
                for sheet in xls.sheet_names:
                    yearly_data[sheet] = pd.read_excel(xls, sheet_name=sheet)
                    yearly_data[sheet].set_index('birth_year', inplace=True)

            # 必需的性状（用于生成正态分布图）
            required_traits = ['NM$', 'TPI']

            # 确定要处理的性状列表：优先使用 selected_traits 参数
            if selected_traits:
                all_traits_to_process = list(selected_traits)
                # 确保必需性状被包含
                for req_trait in required_traits:
                    if req_trait not in all_traits_to_process:
                        all_traits_to_process.append(req_trait)
            else:
                # 回退到从 yearly_data 获取（保持向后兼容）
                all_traits_to_process = list(yearly_data.keys())

            # 确保必需的性状存在于 yearly_data 中
            min_year = int(df['birth_year'].min()) if df['birth_year'].notna().any() else 2020
            max_year = int(df['birth_year'].max()) if df['birth_year'].notna().any() else 2024

            # 为所有需要处理的性状创建年度数据（如果不存在）
            for trait in all_traits_to_process:
                if trait not in yearly_data:
                    print(f"警告: yearly 数据中缺少 {trait}，使用默认值创建")
                    # 创建默认的年度数据
                    default_yearly = pd.DataFrame({
                        'birth_year': range(min_year, max_year + 1),
                        'mean': [0] * (max_year - min_year + 1)
                    })
                    default_yearly.set_index('birth_year', inplace=True)
                    yearly_data[trait] = default_yearly

            # 获取默认值（包含所有要处理的性状）
            default_values = self.get_default_values(all_traits_to_process)

            # 设置权重
            weights = {
                'sire': 0.5,
                'mgs': 0.25,
                'mmgs': 0.125,
                'default': 0.125
            }

            # 计算得分并保留source信息（使用 all_traits_to_process 确保处理所有选择的性状）
            for trait in all_traits_to_process:
                score_column = f'{trait}_score'
                df[score_column] = self.calculate_single_trait_score(
                    df, trait, yearly_data[trait], default_values[trait], weights
                )

                # 复制source列（如果存在）- 使用向量化操作替代apply()
                for bull_type in ['sire', 'mgs', 'mmgs']:
                    source_col = f'{bull_type}_{trait}_source'
                    if source_col not in df.columns:
                        # 如果source列不存在，根据是否有值来判断
                        trait_col = f'{bull_type}_{trait}'
                        if trait_col in df.columns:
                            # 确定年份列
                            if bull_type == 'sire':
                                year_col = 'birth_year'
                            elif bull_type == 'mgs':
                                year_col = 'dam_birth_year'
                            else:
                                year_col = 'mgd_birth_year'

                            # 使用向量化操作设置source值
                            has_trait = df[trait_col].notna()
                            has_year = df[year_col].notna() if year_col in df.columns else pd.Series(False, index=df.index)
                            # source=1 如果有trait值，否则 source=2 如果有年份，否则 source=3
                            df[source_col] = np.where(has_trait, 1, np.where(has_year, 2, 3))

            return self.save_results_with_retry(df, output_path, apply_formatting=apply_formatting)

        except Exception as e:
            print(f"计算性状得分失败: {e}")
            return False

    def calculate_single_trait_score(self, df: pd.DataFrame, trait: str,
                                   yearly_data: pd.DataFrame, default_value: float,
                                   weights: dict) -> pd.Series:
        """计算单个性状的得分"""
        trait_score = pd.Series(0.0, index=df.index)
        
        # 处理sire
        sire_col = f'sire_{trait}'
        trait_score += self.calculate_bull_contribution(
            df, sire_col, 'birth_year', yearly_data, default_value, weights['sire']
        )
        
        # 处理mgs
        mgs_col = f'mgs_{trait}'
        trait_score += self.calculate_bull_contribution(
            df, mgs_col, 'dam_birth_year', yearly_data, default_value, weights['mgs']
        )
        
        # 处理mmgs
        mmgs_col = f'mmgs_{trait}'
        trait_score += self.calculate_bull_contribution(
            df, mmgs_col, 'mgd_birth_year', yearly_data, default_value, weights['mmgs']
        )
        
        # 添加默认值的贡献
        trait_score += weights['default'] * default_value
        
        return trait_score

    def calculate_bull_contribution(self, df: pd.DataFrame, trait_col: str,
                                  year_col: str, yearly_data: pd.DataFrame,
                                  default_value: float, weight: float) -> pd.Series:
        """计算单个公牛的贡献

        注意：此方法假设数据已经通过fill_estimated_values填充了所有缺失值，
        所以可以直接使用trait_col中的值进行计算。
        """
        contribution = pd.Series(0.0, index=df.index)

        # 检查trait_col是否存在
        if trait_col not in df.columns:
            # 如果列不存在，使用默认值（这种情况应该很少发生）
            contribution[:] = weight * default_value
        else:
            # 直接使用列中的值（已包含真实值和预估值）
            # 使用 pd.to_numeric 确保数值类型，避免 FutureWarning
            values = pd.to_numeric(df[trait_col], errors='coerce').fillna(default_value)
            contribution = weight * values

        return contribution

    def update_genomic_data(self, pedigree_path: Path, genomic_path: Path, output_path: Path, apply_formatting: bool = False) -> bool:
        """用基因组数据更新关键性状得分 - 优化版本，使用向量化操作替代循环"""
        try:
            # 1. 读取系谱数据文件
            df_pedigree = pd.read_excel(pedigree_path)

            # 2. 读取基因组数据
            df_genomic = pd.read_excel(genomic_path)

            # 打印基因组数据的列名以便调试
            print(f"基因组数据列名: {df_genomic.columns.tolist()[:30]}...")  # 打印前30个
            print(f"基因组数据行数: {len(df_genomic)}")

            # 检查cow_id列是否存在
            if 'cow_id' not in df_genomic.columns:
                print("警告：基因组数据中缺少cow_id列")
                return False

            # 打印部分cow_id以验证
            print(f"基因组数据中的前5个cow_id: {df_genomic['cow_id'].head().tolist()}")
            print(f"系谱数据中的前5个cow_id: {df_pedigree['cow_id'].head().tolist()}")

            # 3. 初始化数据来源标记
            score_columns = [col for col in df_pedigree.columns if col.endswith('_score')]
            for col in score_columns:
                df_pedigree[f"{col}_source"] = "P"  # 默认标记为系谱来源

            # 4. 更新得分和数据来源 - 使用向量化操作
            traits = [col[:-6] for col in score_columns]  # 去掉 '_score' 后缀
            print(f"需要匹配的性状: {traits}")

            # 检查基因组数据中存在哪些性状列
            existing_traits = [trait for trait in traits if trait in df_genomic.columns]
            print(f"基因组数据中存在的性状: {existing_traits}")

            # 确保cow_id类型一致（都转为字符串）
            df_pedigree['cow_id_str'] = df_pedigree['cow_id'].astype(str)
            df_genomic['cow_id_str'] = df_genomic['cow_id'].astype(str)

            # 为基因组数据设置索引以便快速查找
            df_genomic_indexed = df_genomic.set_index('cow_id_str')

            # 找出在基因组数据中存在的母牛
            matched_mask = df_pedigree['cow_id_str'].isin(df_genomic_indexed.index)
            matched_cows = matched_mask.sum()
            print(f"匹配到的母牛数: {matched_cows}")

            # 统计更新数量
            update_count = 0

            # 对每个性状进行向量化更新
            for trait in existing_traits:
                score_col = f'{trait}_score'
                source_col = f'{trait}_score_source'

                if score_col not in df_pedigree.columns:
                    continue

                # 创建cow_id到性状值的映射
                trait_map = df_genomic_indexed[trait].to_dict()

                # 获取匹配母牛的性状值
                matched_ids = df_pedigree.loc[matched_mask, 'cow_id_str']
                new_values = matched_ids.map(trait_map)

                # 只更新有效值（非NaN）
                valid_mask = matched_mask & new_values.reindex(df_pedigree.index).notna()
                if valid_mask.any():
                    df_pedigree.loc[valid_mask, score_col] = new_values.reindex(df_pedigree.index)[valid_mask]
                    df_pedigree.loc[valid_mask, source_col] = "G"
                    trait_update_count = valid_mask.sum()
                    update_count += trait_update_count

            print(f"总共更新了 {update_count} 个性状值")

            # 清理临时列
            df_pedigree.drop(columns=['cow_id_str'], inplace=True)

            # 5. 添加基因组性状计数列 - 优化版本
            source_cols = [col for col in df_pedigree.columns if col.endswith('_score_source')]
            if source_cols:
                # 使用向量化操作计算每行的 "G" 数量
                df_pedigree['genomic_traits_count'] = (df_pedigree[source_cols] == "G").sum(axis=1)
            else:
                df_pedigree['genomic_traits_count'] = 0

            # 6. 保存结果
            return self.save_results_with_retry(df_pedigree, output_path, apply_formatting=apply_formatting)

        except Exception as e:
            print(f"更新基因组数据时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False

    def get_bull_traits_batch(self, bull_ids: list, selected_traits: list) -> dict:
        """
        批量从数据库获取多个公牛的性状数据（替代逐个查询，大幅提升性能）

        Args:
            bull_ids: 公牛ID列表
            selected_traits: 选中的性状列表

        Returns:
            dict: {bull_id_str: {trait: value, ...}, ...}
        """
        bull_traits = {}
        if not bull_ids:
            return bull_traits

        # 清理并分类ID：短ID用NAAB查，长ID用REG查
        naab_ids = []
        reg_ids = []
        for bid in bull_ids:
            if pd.isna(bid):
                continue
            bid_str = str(bid)
            if len(bid_str) <= 10:
                naab_ids.append(bid_str)
            else:
                reg_ids.append(bid_str)

        try:
            with self.db_engine.connect() as conn:
                # 分批查询，避免SQL IN子句过长（每批500个）
                batch_size = 500

                for batch_start in range(0, len(naab_ids), batch_size):
                    batch = naab_ids[batch_start:batch_start + batch_size]
                    placeholders = ','.join([f':id_{i}' for i in range(len(batch))])
                    params = {f'id_{i}': bid for i, bid in enumerate(batch)}
                    sql = text(f"SELECT * FROM bull_library WHERE `BULL NAAB` IN ({placeholders})")
                    results = conn.execute(sql, params).fetchall()
                    for row in results:
                        row_dict = dict(row._mapping)
                        bull_id_key = str(row_dict.get('BULL NAAB', ''))
                        bull_traits[bull_id_key] = {trait: row_dict.get(trait) for trait in selected_traits}

                for batch_start in range(0, len(reg_ids), batch_size):
                    batch = reg_ids[batch_start:batch_start + batch_size]
                    placeholders = ','.join([f':id_{i}' for i in range(len(batch))])
                    params = {f'id_{i}': bid for i, bid in enumerate(batch)}
                    sql = text(f"SELECT * FROM bull_library WHERE `BULL REG` IN ({placeholders})")
                    results = conn.execute(sql, params).fetchall()
                    for row in results:
                        row_dict = dict(row._mapping)
                        bull_id_key = str(row_dict.get('BULL REG', ''))
                        bull_traits[bull_id_key] = {trait: row_dict.get(trait) for trait in selected_traits}

        except Exception as e:
            print(f"批量获取公牛性状数据时发生错误: {e}")

        return bull_traits

    def get_bull_traits(self, bull_id: str, selected_traits: list) -> dict:
        """
        从数据库获取公牛的性状数据
        
        Args:
            bull_id: 公牛ID
            selected_traits: 选中的性状列表
            
        Returns:
            dict: 包含性状数据的字典，如果未找到则返回None
        """
        try:
            if pd.isna(bull_id):
                return None
                
            bull_id = str(bull_id)  # 确保ID是字符串
            with self.db_engine.connect() as conn:
                # 先尝试用 BULL NAAB 查询
                if len(bull_id) <= 10:
                    result = conn.execute(
                        text("SELECT * FROM bull_library WHERE `BULL NAAB`=:bull_id"),
                        {"bull_id": bull_id}
                    ).fetchone()
                else:
                    # 对于长ID，使用 BULL REG 查询
                    result = conn.execute(
                        text("SELECT * FROM bull_library WHERE `BULL REG`=:bull_id"),
                        {"bull_id": bull_id}
                    ).fetchone()
                
                if result:
                    # 如果找到了公牛，提取所选性状数据
                    result_dict = dict(result._mapping)
                    return {trait: result_dict.get(trait) for trait in selected_traits}
                    
                return None
                
        except Exception as e:
            print(f"获取公牛 {bull_id} 的性状数据时发生错误: {e}")
            return None

    def create_genomic_placeholder(self, pedigree_path: Path, output_path: Path, apply_formatting: bool = False) -> bool:
        """当没有基因组数据时，创建基因组占位文件

        Args:
            pedigree_path: 系谱数据文件路径
            output_path: 输出文件路径
            apply_formatting: 是否应用格式化

        Returns:
            bool: 是否成功创建
        """
        try:
            # 读取系谱数据
            df_pedigree = pd.read_excel(pedigree_path)

            # 如果没有source列，添加它们
            score_columns = [col for col in df_pedigree.columns if col.endswith('_score')]
            for col in score_columns:
                source_col = f"{col}_source"
                if source_col not in df_pedigree.columns:
                    df_pedigree[source_col] = "P"  # 全部标记为系谱来源

            # 添加或更新基因组性状计数列（没有基因组数据，所以都是0）
            df_pedigree['genomic_traits_count'] = 0

            # 保存结果
            return self.save_results_with_retry(df_pedigree, output_path, apply_formatting=apply_formatting)

        except Exception as e:
            print(f"创建基因组占位文件时发生错误: {e}")
            return False

    def fill_estimated_values(self, cow_df, yearly_data_path, selected_traits):
        """填充缺失公牛的预估值 - 优化版本，避免DataFrame碎片化"""
        try:
            # 首先确保年份列存在
            if 'birth_year' not in cow_df.columns and 'birth_date' in cow_df.columns:
                cow_df['birth_year'] = pd.to_datetime(cow_df['birth_date'], errors='coerce').dt.year

            if 'dam_birth_year' not in cow_df.columns and 'birth_date_dam' in cow_df.columns:
                cow_df['dam_birth_year'] = pd.to_datetime(cow_df['birth_date_dam'], errors='coerce').dt.year

            if 'mgd_birth_year' not in cow_df.columns and 'birth_date_mgd' in cow_df.columns:
                cow_df['mgd_birth_year'] = pd.to_datetime(cow_df['birth_date_mgd'], errors='coerce').dt.year

            # 读取年度数据并创建年份到均值的映射字典
            yearly_mean_maps = {}  # {trait: {year: mean}}
            with pd.ExcelFile(yearly_data_path) as xlsx:
                for trait in selected_traits:
                    if trait in xlsx.sheet_names:
                        yearly_df = pd.read_excel(xlsx, sheet_name=trait, index_col='birth_year')
                        yearly_mean_maps[trait] = yearly_df['mean'].to_dict()

            # 获取默认值（999HO99999的值）
            default_values = self.get_default_values(selected_traits)

            # 公牛类型和对应的年份列映射
            bull_type_year_cols = {
                'sire': 'birth_year',
                'mgs': 'dam_birth_year',
                'mmgs': 'mgd_birth_year'
            }

            # 优化：预先创建所有source列，避免DataFrame碎片化
            source_cols_to_add = {}
            for bull_type in bull_type_year_cols.keys():
                for trait in selected_traits:
                    source_col = f'{bull_type}_{trait}_source'
                    if source_col not in cow_df.columns:
                        source_cols_to_add[source_col] = np.ones(len(cow_df), dtype=np.int8)

            # 一次性添加所有source列
            if source_cols_to_add:
                cow_df = pd.concat([cow_df, pd.DataFrame(source_cols_to_add, index=cow_df.index)], axis=1)

            for bull_type, year_col in bull_type_year_cols.items():
                identified_col = f'{bull_type}_identified'

                # 获取未识别的公牛掩码（一次性计算）
                if identified_col in cow_df.columns:
                    unidentified_mask = cow_df[identified_col] == False
                else:
                    unidentified_mask = pd.Series(False, index=cow_df.index)

                if not unidentified_mask.any():
                    continue  # 没有未识别的公牛，跳过这个bull_type

                # 获取有效年份的掩码
                has_year_mask = cow_df[year_col].notna() if year_col in cow_df.columns else pd.Series(False, index=cow_df.index)

                # 预计算掩码
                mask_with_year = unidentified_mask & has_year_mask
                mask_no_year = unidentified_mask & ~has_year_mask

                for trait in selected_traits:
                    trait_col = f'{bull_type}_{trait}'
                    source_col = f'{bull_type}_{trait}_source'

                    # 获取年份到均值的映射
                    year_to_mean = yearly_mean_maps.get(trait, {})
                    default_val = default_values.get(trait, 0)

                    # 情况1：未识别 + 有年份数据 -> 使用年份预估值
                    if mask_with_year.any():
                        # 获取年份并转为整数
                        years = cow_df.loc[mask_with_year, year_col].astype(int)
                        # 使用map批量获取年份对应的均值
                        estimated_values = years.map(year_to_mean)

                        # 有年份映射的行 -> source=2
                        has_mapping = estimated_values.notna()
                        has_mapping_idx = mask_with_year & has_mapping.reindex(cow_df.index, fill_value=False)
                        if has_mapping_idx.any():
                            cow_df.loc[has_mapping_idx, trait_col] = estimated_values[has_mapping].values
                            cow_df.loc[has_mapping_idx, source_col] = 2

                        # 没有年份映射的行 -> source=3，使用默认值
                        no_mapping_idx = mask_with_year & (~has_mapping).reindex(cow_df.index, fill_value=True)
                        if no_mapping_idx.any():
                            cow_df.loc[no_mapping_idx, trait_col] = default_val
                            cow_df.loc[no_mapping_idx, source_col] = 3

                    # 情况2：未识别 + 无年份数据 -> 使用默认值
                    if mask_no_year.any():
                        cow_df.loc[mask_no_year, trait_col] = default_val
                        cow_df.loc[mask_no_year, source_col] = 3

            return cow_df

        except Exception as e:
            print(f"填充预估值时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return cow_df

    def save_with_formatting(self, df, output_path):
        """保存Excel文件并应用格式化（红色字体和黑底黄字）"""
        try:
            # 先保存基础数据
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 定义格式
                red_font = Font(color="FF0000")  # 红色
                yellow_font = Font(color="FFFF00")  # 亮黄色
                gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # 深灰色背景

                # 获取列索引映射
                col_map = {col: idx + 1 for idx, col in enumerate(df.columns)}

                # 遍历所有数据行应用格式
                for row_idx in range(len(df)):
                    for col_name in df.columns:
                        # 检查是否是需要格式化的列
                        if '_source' in col_name:
                            continue  # 跳过source列本身

                        # 获取对应的source列名
                        source_col = None
                        for prefix in ['sire_', 'mgs_', 'mmgs_']:
                            if col_name.startswith(prefix) and not col_name.endswith('_identified'):
                                source_col = col_name + '_source'
                                break

                        if source_col and source_col in df.columns:
                            source_value = df.iloc[row_idx][source_col]
                            cell = worksheet.cell(row=row_idx + 2, column=col_map[col_name])  # +2因为有标题行

                            if source_value == 2:
                                # 年份预估值 - 红色字体
                                cell.font = red_font
                            elif source_value == 3:
                                # 默认预估值 - 灰底黄字
                                cell.font = yellow_font
                                cell.fill = gray_fill

            print(f"文件已保存并格式化: {output_path}")
            return True

        except Exception as e:
            print(f"保存格式化文件时发生错误: {e}")
            return False

    def save_results_with_retry(self, df: pd.DataFrame, output_path: Path, apply_formatting: bool = False) -> bool:
        """
        保存结果，如果文件被占用则提供重试选项

        Args:
            df: 要保存的数据
            output_path: 保存路径
            apply_formatting: 是否应用格式化（颜色标记）

        Returns:
            bool: 是否保存成功
        """
        from PyQt6.QtWidgets import QApplication
        from PyQt6.QtCore import QThread

        def _is_main_thread():
            app = QApplication.instance()
            return app is not None and QThread.currentThread() == app.thread()

        while True:
            try:
                # 确保 cow_id 列保持为字符串类型（修复格式变化问题）
                if 'cow_id' in df.columns:
                    df = df.copy()  # 创建副本避免修改原始数据
                    df['cow_id'] = df['cow_id'].astype(str)

                if apply_formatting and any(col.endswith('_source') for col in df.columns):
                    # 使用save_with_formatting应用格式
                    return self.save_with_formatting(df, output_path)
                else:
                    df.to_excel(output_path, index=False)
                    return True
            except PermissionError:
                if _is_main_thread():
                    from PyQt6.QtWidgets import QMessageBox
                    reply = QMessageBox.question(
                        None,
                        "文件被占用",
                        f"文件 {output_path.name} 正在被其他程序使用。\n"
                        "请关闭该文件后点击'重试'继续，或点击'取消'停止操作。",
                        QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel,
                        QMessageBox.StandardButton.Retry
                    )
                    if reply == QMessageBox.StandardButton.Cancel:
                        print(f"用户取消了保存操作: {output_path}")
                        return False
                else:
                    import time
                    for attempt in range(3):
                        time.sleep(1)
                        try:
                            if 'cow_id' in df.columns:
                                df = df.copy()
                                df['cow_id'] = df['cow_id'].astype(str)
                            df.to_excel(output_path, index=False)
                            return True
                        except PermissionError:
                            continue
                    print(f"[警告] 文件 {output_path.name} 被占用，保存失败")
                    return False
            except Exception as e:
                print(f"保存文件失败: {e}")
                return False
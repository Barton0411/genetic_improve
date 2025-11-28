# core/breeding_calc/index_calculation.py

from pathlib import Path
import json
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from core.breeding_calc.traits_calculation import TraitsCalculation

from .base_calculation import BaseCowCalculation
from .cow_traits_calc import TRAITS_TRANSLATION
import os
from pathlib import Path

# 标准差数据
TRAIT_SD = {
    'MILK': 567, 'NM$': 100, 'FS': 56, 'FE': 50, 'RFI': 46.2,
    'FAT': 25, 'PROT': 15, 'MAST': 2.6, 'EFC': 2.05, 'PL': 1.7,
    'CCR': 1.6, 'LIV': 1.6, 'DPR': 1.4, 'MET': 1.4, 'HCR': 1.3,
    'TPI': 100, 'CM$': 100, 'FM$': 100, 'ST': 1, 'SG': 1,
    'BD': 1, 'DF': 1, 'RA': 1, 'RW': 1, 'LS': 1,
    'LR': 1, 'FA': 1, 'FLS': 1, 'FU': 1, 'UH': 1,
    'UW': 1, 'UC': 1, 'UD': 1, 'FT': 1, 'RT': 1,
    'TL': 1, 'GM$': 100, 'KET': 1, 'PTAT': 1, 'RP': 0.9,
    'BDC': 0.76, 'DA': 0.7, 'UDC': 0.65, 'FLC': 0.53,
    'HLiv': 0.4, 'MFV': 0.4, 'SCS': 0.14, 'FAT %': 0.1, 'PROT%': 0.04
}

# 系统预设权重
DEFAULT_WEIGHTS = {
    'NM$权重': {'NM$': 100},
    'TPI权重': {'TPI': 100}
}

class IndexCalculation(BaseCowCalculation):
    def __init__(self):
        super().__init__()
        self.output_prefix = "processed_index"
        self.required_columns = ['cow_id']  # 基本必需列
        self.traits_calculator = TraitsCalculation()  # 初始化 TraitsCalculation 实例
        self.db_engine = None  # 初始化为 None，不在构造函数中连接


    @staticmethod
    def get_global_weights_path() -> Path:
        """获取全局权重配置文件路径"""
        import sys
        import os
        
        # 在打包的应用中，使用用户数据目录
        if hasattr(sys, '_MEIPASS'):
            # Windows: C:\Users\<username>\AppData\Local\genetic_improve
            # Mac: ~/Library/Application Support/genetic_improve
            if sys.platform == 'win32':
                app_data = Path(os.environ['LOCALAPPDATA']) / 'genetic_improve'
            else:
                app_data = Path.home() / 'Library' / 'Application Support' / 'genetic_improve'
            weights_path = app_data / "index_weights"
        else:
            # 开发环境，使用原来的路径
            current_dir = Path(__file__).resolve().parent  # core/breeding_calc
            while current_dir.name != 'genetic_improve':
                current_dir = current_dir.parent
            
            # genetic_projects是genetic_improve的同级目录
            weights_path = current_dir.parent / "genetic_projects" / "index_weights"
        
        try:
            weights_path.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            import logging
            logging.warning(f"Failed to create weights directory: {e}")
            # 返回临时目录作为备选
            import tempfile
            weights_path = Path(tempfile.gettempdir()) / "genetic_improve_weights"
            weights_path.mkdir(parents=True, exist_ok=True)
            
        return weights_path
        
    def load_weights(self) -> Dict[str, Dict[str, float]]:
        """加载所有权重配置（包括系统预设和用户自定义）"""
        weights = DEFAULT_WEIGHTS.copy()
        
        try:
            # 使用全局路径
            weights_file = self.get_global_weights_path() / "custom_weights.json"
            if weights_file.exists():
                with open(weights_file, 'r', encoding='utf-8') as f:
                    custom_weights = json.load(f)
                weights.update(custom_weights)
        except Exception as e:
            print(f"加载用户自定义权重失败: {e}")
                
        return weights
        
    def save_custom_weight(self, weight_name: str, weight_values: Dict[str, float]) -> bool:
        """保存用户自定义权重"""
        try:
            # 使用全局路径
            weights_file = self.get_global_weights_path() / "custom_weights.json"
            
            # 读取现有权重
            existing_weights = {}
            if weights_file.exists():
                with open(weights_file, 'r', encoding='utf-8') as f:
                    existing_weights = json.load(f)
            
            # 更新权重
            existing_weights[weight_name] = weight_values
            
            # 保存
            with open(weights_file, 'w', encoding='utf-8') as f:
                json.dump(existing_weights, f, ensure_ascii=False, indent=4)
                
            return True
            
        except Exception as e:
            print(f"保存自定义权重失败: {e}")
            return False
            
    def delete_custom_weight(self, weight_name: str) -> bool:
        """删除用户自定义权重"""
        try:
            weights_file = self.get_global_weights_path() / "custom_weights.json"
            
            if not weights_file.exists():
                return False
                
            with open(weights_file, 'r', encoding='utf-8') as f:
                weights = json.load(f)
                
            if weight_name in weights:
                del weights[weight_name]
                
                with open(weights_file, 'w', encoding='utf-8') as f:
                    json.dump(weights, f, ensure_ascii=False, indent=4)
                    
                return True
            return False
            
        except Exception as e:
            print(f"删除自定义权重失败: {e}")
            return False
            
    def validate_weight_values(self, weight_values: Dict[str, float]) -> bool:
        """验证权重值是否有效"""
        total = sum(abs(v) for v in weight_values.values())
        return abs(total - 100) < 0.0001  # 允许一点点浮点数误差
        
    def calculate_index_score(self, trait_values: dict, weight_values: dict) -> float:
        """计算指数得分"""
        score = 0
        for trait, weight in weight_values.items():
            if trait in trait_values and trait in TRAIT_SD:
                # 每个性状的得分 = 性状值/性状标准差 × 权重值
                score += (trait_values[trait] / TRAIT_SD[trait]) * weight
        return score
        
    def process_cow_index(self, main_window, weight_name: str) -> Tuple[bool, str]:
        """处理母牛群指数计算"""
        try:
            project_path = main_window.selected_project_path

            # 1. 首先检查是否有 processed_cow_data.xlsx
            cow_data_path = project_path / "standardized_data" / "processed_cow_data.xlsx"
            if not cow_data_path.exists():
                return False, "请先上传母牛数据"

            # 2. 加载权重配置并获取性状列表
            weights = self.load_weights()
            if weight_name not in weights:
                return False, f"未找到权重配置：{weight_name}"
            weight_values = weights[weight_name]
            selected_traits = list(weight_values.keys())

            # 3. 检查是否存在基因组评估结果文件
            genomic_scores_path = project_path / "analysis_results" / "processed_cow_data_key_traits_scores_genomic.xlsx"
            if genomic_scores_path.exists():
                # 3.1 基因组评估结果存在，检查是否完整
                genomic_df = pd.read_excel(genomic_scores_path)
                existing_traits = [col[:-6] for col in genomic_df.columns if col.endswith('_score')]
                missing_traits = [trait for trait in selected_traits if trait not in existing_traits]

                if not missing_traits:
                    # 所有性状都存在，直接使用现有基因组评估结果
                    print("使用现有完整的基因组评估结果")
                    df = genomic_df
                else:
                    # 缺少部分性状，需要重新计算
                    print(f"基因组评估结果缺少性状: {missing_traits}")
                    
                    # 检查是否有基因组数据
                    genomic_data_path = project_path / "standardized_data" / "processed_genomic_data.xlsx"
                    if genomic_data_path.exists():
                        # 有基因组数据，重新计算包含基因组数据
                        success, message = self.traits_calculator.process_data(main_window, selected_traits)
                        if not success:
                            return False, message
                        df = pd.read_excel(project_path / "analysis_results" / "processed_cow_data_key_traits_scores_genomic.xlsx")
                    else:
                        # 没有基因组数据，使用系谱计算后更新基因组评估文件
                        success, message = self.traits_calculator.process_data(main_window, selected_traits)
                        if not success:
                            return False, message
                        
                        # 读取新计算的系谱结果
                        pedigree_df = pd.read_excel(project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx")
                        
                        # 更新基因组评估文件中的缺失性状
                        for trait in missing_traits:
                            score_col = f'{trait}_score'
                            source_col = f'{trait}_score_source'
                            genomic_df[score_col] = pedigree_df[score_col]
                            genomic_df[source_col] = 'P'  # 标记为系谱来源
                        
                        # 保存更新后的基因组评估文件
                        genomic_df.to_excel(genomic_scores_path, index=False)
                        df = genomic_df
            else:
                # 3.2 基因组评估结果不存在，检查是否有基因组数据
                genomic_data_path = project_path / "standardized_data" / "processed_genomic_data.xlsx"
                if genomic_data_path.exists():
                    # 有基因组数据，计算包含基因组数据的结果
                    success, message = self.traits_calculator.process_data(main_window, selected_traits)
                    if not success:
                        return False, message
                    df = pd.read_excel(project_path / "analysis_results" / "processed_cow_data_key_traits_scores_genomic.xlsx")
                else:
                    # 没有基因组数据，检查系谱评估结果
                    pedigree_scores_path = project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx"
                    if pedigree_scores_path.exists():
                        # 检查系谱评估结果是否完整
                        pedigree_df = pd.read_excel(pedigree_scores_path)
                        existing_traits = [col[:-6] for col in pedigree_df.columns if col.endswith('_score')]
                        missing_traits = [trait for trait in selected_traits if trait not in existing_traits]

                        if not missing_traits:
                            # 系谱评估结果完整，直接使用
                            print("使用现有完整的系谱评估结果")
                            df = pedigree_df
                        else:
                            # 系谱评估结果不完整，重新计算
                            print(f"系谱评估结果缺少性状: {missing_traits}")
                            success, message = self.traits_calculator.process_data(main_window, selected_traits)
                            if not success:
                                return False, message
                            df = pd.read_excel(project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx")
                    else:
                        # 没有任何评估结果，计算系谱评估结果
                        success, message = self.traits_calculator.process_data(main_window, selected_traits)
                        if not success:
                            return False, message
                        df = pd.read_excel(project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx")

            # 4. 计算指数得分 (向量化优化)
            weight_values = weights[weight_name]
            score = np.zeros(len(df))
            for trait, weight in weight_values.items():
                if trait in TRAIT_SD:
                    score_col = f'{trait}_score'
                    if score_col in df.columns:
                        # 使用向量化操作，NaN 值用 0 填充
                        trait_scores = df[score_col].fillna(0).values
                        score += (trait_scores / TRAIT_SD[trait]) * weight
            df[f'{weight_name}_index'] = score

            # 5. 排序并添加排名
            df = df.sort_values(f'{weight_name}_index', ascending=False)
            df['ranking'] = range(1, len(df) + 1)

            # 5.5 确保 cow_id 列保持为字符串类型（修复格式变化问题）
            if 'cow_id' in df.columns:
                df['cow_id'] = df['cow_id'].astype(str)

            # 6. 保存结果（应用格式化）
            output_path = project_path / "analysis_results" / f"{self.output_prefix}_cow_index_scores.xlsx"
            if not self.save_results_with_retry(df, output_path, apply_formatting=True):
                return False, "保存结果失败"

            return True, "计算完成"

        except Exception as e:
            print(f"计算母牛群指数时发生错误: {str(e)}")
            return False, str(e)

    def process_bull_index(self, main_window, weight_name: str) -> Tuple[bool, str]:
        """处理公牛指数计算"""
        try:
            project_path = main_window.selected_project_path
            
            # 1. 检查并初始化必要的设置
            bull_data_path = project_path / "standardized_data" / "processed_bull_data.xlsx"
            if not bull_data_path.exists():
                return False, "请先上传备选公牛数据"
                
            # 2. 检查并初始化数据库连接
            if not self.init_db_connection():
                print("数据库连接初始化失败")
                return False, "连接数据库失败"
                
            # 3. 加载权重配置
            weights = self.load_weights()
            if weight_name not in weights:
                return False, f"未找到权重配置：{weight_name}"
                    
            weight_values = weights[weight_name]
            if not weight_values:
                return False, "权重配置为空"
                
            selected_traits = list(weight_values.keys())
            if not selected_traits:
                return False, "未找到需要计算的性状"
            
            # 4. 读取备选公牛数据
            try:
                bull_df = pd.read_excel(bull_data_path)
                if bull_df.empty:
                    return False, "备选公牛数据为空"
            except Exception as e:
                return False, f"读取备选公牛数据失败: {str(e)}"
                    
            # 5. 为每个公牛查询性状数据
            missing_bulls = []
            for idx, row in bull_df.iterrows():
                bull_id = str(row['bull_id'])
                trait_data, found = self.query_bull_traits(bull_id, selected_traits)
                if found:
                    for trait, value in trait_data.items():
                        bull_df.at[idx, trait] = value
                else:
                    missing_bulls.append(bull_id)
                    # 如果找不到公牛数据，所有性状值设为None
                    for trait in selected_traits:
                        bull_df.at[idx, trait] = None

            # 6. 处理缺失的公牛信息
            if missing_bulls:
                print(f"\n[检查点-指数] 在指数排序中发现 {len(missing_bulls)} 个缺失公牛")
                print(f"[检查点-指数] 调用 process_missing_bulls 进行上传...")
                self.process_missing_bulls(missing_bulls, 'bull_index', main_window.username)
            else:
                print("\n[检查点-指数] 所有公牛数据完整，无缺失公牛")

            # 7. 计算指数得分 (向量化优化)
            score = np.zeros(len(bull_df))
            for trait, weight in weight_values.items():
                if trait in TRAIT_SD and trait in bull_df.columns:
                    # 使用向量化操作，NaN 值用 0 填充
                    trait_values = bull_df[trait].fillna(0).values
                    score += (trait_values / TRAIT_SD[trait]) * weight
            bull_df[f'{weight_name}_index'] = score
            
            # 8. 排序并添加排名
            bull_df = bull_df.sort_values(f'{weight_name}_index', ascending=False)
            bull_df['ranking'] = range(1, len(bull_df) + 1)
            
            # 9. 保存结果
            output_path = project_path / "analysis_results" / f"{self.output_prefix}_bull_scores.xlsx"
            if not self.save_results_with_retry(bull_df, output_path):
                return False, "保存结果失败"
                
            return True, "计算完成"
                
        except Exception as e:
            print(f"公牛指数计算发生错误: {str(e)}")
            return False, str(e)
            
        finally:
            # 确保关闭数据库连接
            if hasattr(self, 'db_engine') and self.db_engine is not None:
                try:
                    self.db_engine.dispose()
                    self.db_engine = None
                except Exception as e:
                    print(f"关闭数据库连接时发生错误: {str(e)}")
            
    def calculate_cow_traits(self, project_path: Path, selected_traits: List[str]) -> Tuple[bool, Optional[pd.DataFrame]]:
        """计算母牛关键性状"""
        try:
            cow_data_path = project_path / "standardized_data" / "processed_cow_data.xlsx"
            if not self.init_db_connection():
                return False, None
                
            cow_df = pd.read_excel(cow_data_path)
            
            # 对每个母牛计算选中的性状
            for trait in selected_traits:
                trait_col = f'sire_{trait}'
                
                # 从数据库查询公牛数据
                bull_ids = cow_df['sire'].dropna().unique()
                bull_traits = {}
                
                for bull_id in bull_ids:
                    if pd.isna(bull_id):
                        continue
                    traits_data, found = self.query_bull_traits(str(bull_id), [trait])
                    if found:
                        bull_traits[str(bull_id)] = traits_data[trait]
                
                # 设置性状值
                cow_df[trait] = cow_df['sire'].apply(
                    lambda x: bull_traits.get(str(x)) if pd.notna(x) else None
                )
            
            return True, cow_df
            
        except Exception as e:
            print(f"计算母牛性状失败: {e}")
            return False, None
        

    # 检查是否已有关键性状结果,如果有,检查是否包含所有选中性状。
    def check_existing_traits_results(self, project_path: Path, selected_traits: list) -> Tuple[Optional[pd.DataFrame], bool]:
        """
        检查是否已有关键性状结果,如果有,检查是否包含所有选中性状。
        
        Args:
            project_path: 项目路径
            selected_traits: 选中的性状列表
            
        Returns:
            Tuple[Optional[pd.DataFrame], bool]: (现有性状结果的DataFrame, 是否包含所有选中性状)
        """
        genomic_path = project_path / "analysis_results" / "processed_cow_data_key_traits_scores_genomic.xlsx"
        pedigree_path = project_path / "analysis_results" / "processed_cow_data_key_traits_scores_pedigree.xlsx"
        
        if genomic_path.exists():
            df = pd.read_excel(genomic_path)
        elif pedigree_path.exists():
            df = pd.read_excel(pedigree_path)
        else:
            return None, False
        
        existing_traits = [col[:-6] for col in df.columns if col.endswith('_score')]
        missing_traits = [trait for trait in selected_traits if trait not in existing_traits]

        return df, len(missing_traits) == 0

    def save_with_formatting(self, df, output_path):
        """保存Excel文件并应用格式化（红色字体和灰底黄字）"""
        try:
            # 检查是否需要加载detail文件来获取原始source值
            need_detail = False
            for col in df.columns:
                if col.endswith('_score_source'):
                    # 检查source值是否是P/G格式
                    sample_val = df[col].iloc[0] if len(df) > 0 else None
                    if sample_val in ['P', 'G']:
                        need_detail = True
                        break

            # 如果需要，加载detail文件
            original_sources = {}
            if need_detail:
                # 尝试找到detail文件
                detail_path = output_path.parent / "processed_cow_data_key_traits_detail.xlsx"
                if detail_path.exists():
                    detail_df = pd.read_excel(detail_path)
                    # 只保留需要的列
                    if 'cow_id' in detail_df.columns:
                        original_sources['cow_id'] = detail_df['cow_id']
                        for col in detail_df.columns:
                            if col.endswith('_source') and (col.startswith('sire_') or
                                                            col.startswith('mgs_') or
                                                            col.startswith('mmgs_')):
                                original_sources[col] = detail_df[col]

            # 先保存基础数据
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 定义格式
                red_font = Font(color="FF0000")  # 红色
                yellow_font = Font(color="FFFF00")  # 亮黄色
                gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # 深灰色背景

                # 获取列索引映射
                col_map = {col: idx + 1 for idx, col in enumerate(df.columns)}

                # 如果有原始source数据，创建cow_id到索引的映射
                cow_id_map = {}
                if original_sources and 'cow_id' in original_sources:
                    for idx, cow_id in enumerate(original_sources['cow_id']):
                        cow_id_map[cow_id] = idx

                # 批量收集需要格式化的单元格 (性能优化)
                red_cells = []  # source == 2 (年份预估值)
                yellow_cells = []  # source == 3 (默认预估值)

                # 处理 _score 列
                for col_name in df.columns:
                    if '_score' in col_name and not col_name.endswith('_source'):
                        trait = col_name.replace('_score', '')
                        col_idx = col_map[col_name]

                        # 优先使用原始source数据（从detail文件）
                        if original_sources and cow_id_map:
                            # 对于每个 cow_id，计算最大的 source 值
                            max_sources = pd.Series(index=df.index, dtype=float)
                            for idx, row in df.iterrows():
                                cow_id = row['cow_id'] if 'cow_id' in df.columns else None
                                if cow_id and cow_id in cow_id_map:
                                    detail_idx = cow_id_map[cow_id]
                                    source_vals = []
                                    for prefix in ['sire_', 'mgs_', 'mmgs_']:
                                        source_col = f'{prefix}{trait}_source'
                                        if source_col in original_sources:
                                            val = original_sources[source_col].iloc[detail_idx]
                                            if pd.notna(val):
                                                source_vals.append(val)
                                    if source_vals:
                                        max_sources.at[idx] = max(source_vals)

                            # 使用向量化掩码
                            mask_red = max_sources == 2
                            mask_yellow = max_sources == 3
                            red_cells.extend([(r + 2, col_idx) for r in df.index[mask_red]])
                            yellow_cells.extend([(r + 2, col_idx) for r in df.index[mask_yellow]])
                        else:
                            # 使用直接的 source 列
                            direct_source_col = col_name + '_source'
                            if direct_source_col in df.columns:
                                mask_red = df[direct_source_col] == 2
                                mask_yellow = df[direct_source_col] == 3
                                red_cells.extend([(r + 2, col_idx) for r in df.index[mask_red]])
                                yellow_cells.extend([(r + 2, col_idx) for r in df.index[mask_yellow]])

                # 处理 sire_*、mgs_*、mmgs_* 列
                for col_name in df.columns:
                    if (col_name.startswith('sire_') or col_name.startswith('mgs_') or
                        col_name.startswith('mmgs_')) and not col_name.endswith('_source'):
                        col_idx = col_map[col_name]
                        source_col = col_name + '_source'

                        if source_col in df.columns:
                            mask_red = df[source_col] == 2
                            mask_yellow = df[source_col] == 3
                            red_cells.extend([(r + 2, col_idx) for r in df.index[mask_red]])
                            yellow_cells.extend([(r + 2, col_idx) for r in df.index[mask_yellow]])
                        elif original_sources and cow_id_map and source_col in original_sources:
                            # 从 original_sources 获取
                            for idx, row in df.iterrows():
                                cow_id = row['cow_id'] if 'cow_id' in df.columns else None
                                if cow_id and cow_id in cow_id_map:
                                    detail_idx = cow_id_map[cow_id]
                                    source_value = original_sources[source_col].iloc[detail_idx]
                                    if source_value == 2:
                                        red_cells.append((idx + 2, col_idx))
                                    elif source_value == 3:
                                        yellow_cells.append((idx + 2, col_idx))

                # 批量应用格式
                for row, col in red_cells:
                    worksheet.cell(row=row, column=col).font = red_font
                for row, col in yellow_cells:
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = yellow_font
                    cell.fill = gray_fill

            print(f"文件已保存并格式化: {output_path}")
            return True

        except Exception as e:
            print(f"保存格式化文件时发生错误: {e}")
            return False

    def save_results_with_retry(self, df: pd.DataFrame, output_path: Path, apply_formatting: bool = False) -> bool:
        """
        保存结果，如果文件被占用则提供重试选项

        Args:
            df: 要保存的数据
            output_path: 保存路径
            apply_formatting: 是否应用格式化（颜色标记）

        Returns:
            bool: 是否保存成功
        """
        from PyQt6.QtWidgets import QMessageBox

        while True:
            try:
                # 确保 cow_id 列保持为字符串类型（修复格式变化问题）
                if 'cow_id' in df.columns:
                    df = df.copy()  # 创建副本避免修改原始数据
                    df['cow_id'] = df['cow_id'].astype(str)

                if apply_formatting and any('_source' in col for col in df.columns):
                    # 使用save_with_formatting应用格式
                    return self.save_with_formatting(df, output_path)
                else:
                    df.to_excel(output_path, index=False)
                    return True
            except PermissionError:
                reply = QMessageBox.question(
                    None,
                    "文件被占用",
                    f"文件 {output_path.name} 正在被其他程序使用。\n"
                    "请关闭该文件后点击'重试'继续，或点击'取消'停止操作。",
                    QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel,
                    QMessageBox.StandardButton.Retry
                )

                if reply == QMessageBox.StandardButton.Cancel:
                    print(f"用户取消了保存操作: {output_path}")
                    return False
            except Exception as e:
                print(f"保存文件失败: {e}")
                return False    
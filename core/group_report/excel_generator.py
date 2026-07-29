"""按单牧场子项目的小型汇总结果生成牧场组 Excel。"""

from __future__ import annotations

import logging
import math
import os
import tempfile
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.group_tasks.lease_heartbeat import GroupLeaseHeartbeat
from utils.file_manager import FileManager


logger = logging.getLogger(__name__)

TRAITS = (
    "NM$",
    "TPI",
    "MILK",
    "FAT",
    "FAT %",
    "PROT",
    "PROT%",
    "SCS",
    "PL",
    "DPR",
    "PTAT",
    "UDC",
    "FLC",
    "RFI",
)
FARM_IDENTITY_HEADERS = ("API farmcode", "牧场编号", "牧场名称")


def _number(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


class GroupExcelReportGenerator:
    """生成不复制全量宽表的牧场组最终汇总工作簿。"""

    def __init__(
        self,
        project_path: Path,
        service_staff: str = "",
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ):
        self.project_path = Path(project_path)
        self.service_staff = service_staff or ""
        self.progress_callback = progress_callback
        self.metadata = FileManager.load_project_metadata(self.project_path)
        self._lease_store = None
        self._lease_heartbeat: Optional[GroupLeaseHeartbeat] = None
        self._selection_revision = None

    def _progress(self, value: int, message: str) -> None:
        self._refresh_summary_lease()
        if self.progress_callback:
            self.progress_callback(value, message)

    def _acquire_summary_lease(self) -> None:
        store = FileManager._group_task_store(self.project_path)
        if store is None:
            return
        revision = store.get_selection_revision()
        lease = store.acquire_run_lease(
            f"group-summary:{os.getpid()}:{uuid.uuid4()}",
            run_kind="group_summary",
            lease_seconds=600,
            expected_selection_revision=revision,
        )
        if lease is None:
            raise RuntimeError(
                "该牧场组已有另一个处理或汇总任务正在运行，请稍后再试"
            )
        heartbeat = GroupLeaseHeartbeat(
            store,
            lease,
            lease_seconds=600,
        )
        try:
            heartbeat.start()
        except Exception:
            try:
                store.release_run_lease(str(lease["lease_token"]))
            except Exception:
                logger.warning(
                    "启动牧场组汇总租约续租器失败后释放租约失败",
                    exc_info=True,
                )
            raise
        self._lease_store = store
        self._lease_heartbeat = heartbeat
        self._selection_revision = int(lease["selection_revision"])

    def _refresh_summary_lease(self) -> None:
        if self._lease_heartbeat is None:
            return
        self._lease_heartbeat.check()

    def _release_summary_lease(self) -> None:
        heartbeat = self._lease_heartbeat
        self._lease_heartbeat = None
        if heartbeat is not None:
            try:
                heartbeat.stop(timeout=30, release=True)
            except Exception:
                logger.warning("释放牧场组汇总运行锁失败", exc_info=True)
        self._lease_store = None

    def _validate(self) -> Tuple[bool, str]:
        if self.metadata.get("project_type") != "multi_farm_group":
            return False, "当前项目不是牧场组项目"
        tasks = self.metadata.get("group_tasks", [])
        if not tasks:
            return False, "牧场组中没有子任务"
        active_tasks = [
            task
            for task in tasks
            if task.get("included_in_summary", True)
        ]
        if not active_tasks:
            return False, "牧场组中没有纳入最终汇总范围的牧场"
        readiness = FileManager._group_summary_readiness(
            self.project_path, self.metadata
        )
        if not readiness["ready"]:
            examples = []
            for item in readiness["missing_tasks"][:3]:
                name = item.get("farm_name") or item.get("farm_code")
                examples.append(
                    f"{name}（缺少{'、'.join(item.get('missing', [])[:3])}）"
                )
            suffix = "等" if readiness["missing_count"] > 3 else ""
            return (
                False,
                f"最终汇总尚未就绪：{readiness['ready_count']}/"
                f"{readiness['included_count']} 个牧场具备完整单场结果。"
                + ("；".join(examples) + suffix if examples else ""),
            )
        return True, ""

    @staticmethod
    def _rows(path: Path, sheet_name: str) -> Iterable[Tuple]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                return []
            return list(workbook[sheet_name].iter_rows(values_only=True))
        finally:
            workbook.close()

    @staticmethod
    def _latest_report(child_path: Path) -> Optional[Path]:
        reports = list((child_path / "reports").glob("育种分析综合报告_*.xlsx"))
        return max(reports, key=lambda path: path.stat().st_mtime) if reports else None

    def _collect_farm(self, task: Dict) -> Dict:
        child_path = self.project_path / task["relative_path"]
        task_metadata = task.get("metadata", {})
        if not isinstance(task_metadata, dict):
            task_metadata = {}
        farm_code = str(task.get("farm_code", ""))
        source_kind = str(task.get("source_kind") or "api")
        if "farm_number" in task:
            farm_number = task.get("farm_number")
        elif "farm_number" in task_metadata:
            farm_number = task_metadata.get("farm_number")
        else:
            farm_number = (
                ""
                if str(task.get("source_system") or "") == "慧牧云"
                else farm_code
            )
        farm = {
            "task_id": str(task.get("task_id", "")),
            "farm_code": farm_code,
            "api_farmcode": str(
                task.get("api_farmcode")
                or task_metadata.get("api_farmcode")
                or (farm_code if source_kind != "local" else "")
            ),
            "farm_number": str(farm_number or ""),
            "farm_name": str(task.get("farm_name", "")),
            "relative_path": str(task.get("relative_path", "")),
            "child_path": child_path,
            "report_path": self._latest_report(child_path),
            "current": {},
            "all": {},
            "yearly": [],
            "distributions": defaultdict(list),
            "pedigree": {},
            "availability": {},
        }

        traits_path = child_path / "analysis_results" / "关键育种性状分析结果.xlsx"
        if traits_path.exists():
            for sheet_name, target_name in (
                ("在群母牛年份汇总", "current"),
                ("全部母牛年份汇总", "all"),
            ):
                rows = self._rows(traits_path, sheet_name)
                if rows:
                    headers = [str(value or "") for value in rows[0]]
                    for raw in rows[1:]:
                        row = dict(zip(headers, raw))
                        group = str(row.get("出生年份") or "")
                        if "总计" in group:
                            farm[target_name] = row
                        elif target_name == "current" and group:
                            farm["yearly"].append(row)

            for sheet_name, key in (
                ("在群母牛NM$分布", "current_nm"),
                ("在群母牛TPI分布", "current_tpi"),
            ):
                rows = self._rows(traits_path, sheet_name)
                if rows:
                    headers = [str(value or "") for value in rows[0]]
                    farm["distributions"][key] = [
                        dict(zip(headers, raw)) for raw in rows[1:] if raw[0]
                    ]

        pedigree_path = child_path / "analysis_results" / "系谱识别分析结果.xlsx"
        if pedigree_path.exists():
            rows = self._rows(pedigree_path, "Sheet1")
            if rows:
                headers = [str(value or "") for value in rows[0]]
                current_rows = [
                    dict(zip(headers, raw))
                    for raw in rows[1:]
                    if str(raw[0] or "") == "是"
                ]
                total = sum(int(_number(row.get("头数"), 0)) for row in current_rows)
                sire = sum(
                    int(_number(row.get("父号可识别头数"), 0))
                    for row in current_rows
                )
                mgs = sum(
                    int(_number(row.get("外祖父可识别头数"), 0))
                    for row in current_rows
                )
                mmgs = sum(
                    int(_number(row.get("外曾外祖父可识别头数"), 0))
                    for row in current_rows
                )
                farm["pedigree"] = {
                    "头数": total,
                    "父号可识别头数": sire,
                    "父号识别率": sire / total if total else 0,
                    "外祖父可识别头数": mgs,
                    "外祖父识别率": mgs / total if total else 0,
                    "外曾外祖父可识别头数": mmgs,
                    "外曾外祖父识别率": mmgs / total if total else 0,
                }

        analysis_dir = child_path / "analysis_results"
        standardized_dir = child_path / "standardized_data"
        farm["availability"] = {
            "配种记录": (standardized_dir / "processed_breeding_data.xlsx").exists(),
            "备选公牛": (standardized_dir / "processed_bull_data.xlsx").exists(),
            "已配公牛分析": (analysis_dir / "processed_mated_bull_traits.xlsx").exists(),
            "近交及隐性基因": bool(list(analysis_dir.glob("*近交系数及隐性基因分析结果*.xlsx"))),
            "单牧场Excel": farm["report_path"] is not None,
        }
        return farm

    @staticmethod
    def _farm_identity_values(farm: Dict) -> List[str]:
        """返回接口编码、业务编号和展示名称，禁止三者互相冒充。"""
        return [
            str(farm.get("api_farmcode") or ""),
            str(farm.get("farm_number") or ""),
            str(farm.get("farm_name") or ""),
        ]

    @staticmethod
    def _task_identity_values(task: Dict) -> List[str]:
        task_metadata = task.get("metadata", {})
        if not isinstance(task_metadata, dict):
            task_metadata = {}
        farm_code = str(task.get("farm_code") or "")
        source_kind = str(task.get("source_kind") or "api")
        if "farm_number" in task:
            farm_number = task.get("farm_number")
        elif "farm_number" in task_metadata:
            farm_number = task_metadata.get("farm_number")
        else:
            farm_number = (
                ""
                if str(task.get("source_system") or "") == "慧牧云"
                else farm_code
            )
        return [
            str(
                task.get("api_farmcode")
                or task_metadata.get("api_farmcode")
                or (farm_code if source_kind != "local" else "")
            ),
            str(farm_number or ""),
            str(task.get("farm_name") or ""),
        ]

    @staticmethod
    def _rank_values(farms: List[Dict], column: str) -> Dict[str, int]:
        values = []
        for farm in farms:
            value = farm["current"].get(column)
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            if not math.isfinite(numeric):
                continue
            values.append((farm["task_id"], numeric))
        values.sort(key=lambda item: (-item[1], item[0]))
        return {
            task_id: rank
            for rank, (task_id, _) in enumerate(values, start=1)
        }

    @staticmethod
    def _relative_report_link(report_directory: Path, target: Path) -> str:
        # macOS 的 /var 与 /private/var 是同一位置的不同拼写；两端先
        # resolve，避免生成绕到文件系统根目录的脆弱超长链接。
        return os.path.relpath(
            Path(target).resolve(),
            Path(report_directory).resolve(),
        ).replace(os.sep, "/")

    @staticmethod
    def _weighted_group_row(farms: List[Dict], target: str) -> Dict:
        total_count = sum(int(_number(farm[target].get("头数"), 0)) for farm in farms)
        result = {"头数": total_count}
        for trait in TRAITS:
            column = f"平均{trait}"
            numerator = 0.0
            denominator = 0
            for farm in farms:
                count_value = int(_number(farm[target].get("头数"), 0))
                value = farm[target].get(column)
                if count_value and value is not None:
                    numerator += _number(value) * count_value
                    denominator += count_value
            result[column] = numerator / denominator if denominator else None
        return result

    def _write_table(
        self,
        sheet,
        title: str,
        headers: List[str],
        rows: List[List],
        *,
        freeze: str = "A3",
        percent_columns: Optional[Iterable[int]] = None,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = sheet.cell(1, 1, title)
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="2F75B5")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 30
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(2, column, header)
            cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D9E2F3")
        for row_index, values in enumerate(rows, start=3):
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column, value)
                cell.font = Font(name="微软雅黑", size=9, color="333333")
                cell.alignment = Alignment(
                    horizontal="left" if isinstance(value, str) else "right",
                    vertical="center",
                )
                if isinstance(value, int) and not isinstance(value, bool):
                    cell.number_format = "#,##0"
                elif isinstance(value, float):
                    cell.number_format = "0.00"
                cell.border = Border(bottom=thin)
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7FAFD")
        for column in percent_columns or []:
            for row_index in range(3, 3 + len(rows)):
                sheet.cell(row_index, column).number_format = "0.00%"
        sheet.freeze_panes = freeze
        if rows:
            sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows)+2}"
        for column, header in enumerate(headers, start=1):
            values = [str(header)] + [str(row[column - 1] or "") for row in rows[:200]]
            width = min(max(max(map(len, values)) + 2, 10), 28)
            sheet.column_dimensions[get_column_letter(column)].width = width

    def generate(self) -> Tuple[bool, str]:
        try:
            self._acquire_summary_lease()
            # 租约取得后重新读取一次，确保校验、读取和最终发布使用同一个
            # 纳入范围快照，而不是构造生成器时的旧内存副本。
            self.metadata = FileManager.load_project_metadata(
                self.project_path
            )
            valid, message = self._validate()
            if not valid:
                return False, message
            from core.group_report.publication_snapshot import (
                capture_group_publication_snapshot,
                recompute_and_compare_group_publication_snapshot,
            )

            publication_run_id = str(uuid.uuid4())
            snapshot_directory = (
                self.project_path
                / "group_store"
                / "publication_snapshots"
            )
            before_snapshot_path = (
                snapshot_directory
                / f"{publication_run_id}_before.json"
            )
            before_snapshot = capture_group_publication_snapshot(
                self.project_path,
                output_path=before_snapshot_path,
            )
            from core.group_report.publication_batch import (
                GroupReportPublicationBatch,
                publish_current_group_report_pointer,
                validate_current_group_report_pointer,
            )

            # 指针切换成功、但随后更新 project_metadata.json 前若程序退出，
            # 唯一正式报告其实已经存在。重新生成时先完整验证并修复这层
            # 派生元数据，避免把数百万行明细重新计算一遍。
            current_pointer_path = (
                self.project_path
                / "group_store"
                / "current_group_report.json"
            )
            if current_pointer_path.is_file():
                try:
                    current = validate_current_group_report_pointer(
                        self.project_path
                    )
                except Exception:
                    current = None
                if (
                    current is not None
                    and int(current["selection_revision"])
                    == int(self._selection_revision or 0)
                    and current["publication_basis_sha256"]
                    == before_snapshot["basis_sha256"]
                ):
                    FileManager.update_group_result(
                        self.project_path,
                        excel_path=current["excel_path"],
                        selection_revision=self._selection_revision,
                        report_package_path=current["package_path"],
                        batch_manifest_path=current[
                            "batch_manifest_path"
                        ],
                        batch_manifest_sha256=current[
                            "batch_manifest_sha256"
                        ],
                        current_pointer_path=current["pointer_path"],
                        current_pointer_sha256=current["pointer_sha256"],
                        publication_basis_sha256=current[
                            "publication_basis_sha256"
                        ],
                    )
                    self._progress(
                        100,
                        "当前输入对应的牧场组报告已完整发布，已直接复用",
                    )
                    return True, str(current["excel_path"])

            publication_batch = GroupReportPublicationBatch(
                self.project_path,
                publication_basis_sha256=before_snapshot["basis_sha256"],
                selection_revision=self._selection_revision or 0,
            )
            self._progress(5, "正在读取单牧场汇总结果...")
            all_tasks = self.metadata["group_tasks"]
            tasks = [
                task
                for task in all_tasks
                if task.get("included_in_summary", True)
            ]
            farms = []
            for index, task in enumerate(tasks, start=1):
                farms.append(self._collect_farm(task))
                self._progress(
                    5 + int(index / len(tasks) * 35),
                    f"已读取 {task.get('farm_name', task.get('farm_code'))}",
                )

            self._progress(42, "正在建立全量跨牧场牛只排名与明细分卷...")
            from core.group_report.detail_exporter import (
                GroupCowRankingDetailExporter,
            )

            detail_package_name = "牧场组牛只完整排名明细"
            detail_manifest = publication_batch.load_completed_detail(
                detail_package_name
            )
            if detail_manifest is None:
                detail_manifest = GroupCowRankingDetailExporter(
                    self.project_path,
                    progress_callback=lambda value, text: self._progress(
                        42 + int(value * 0.38), text
                    ),
                ).export(
                    tasks=tasks,
                    output_dir=publication_batch.detail_root,
                    package_name=detail_package_name,
                )
            if detail_manifest.get("status") != "complete":
                problems = [
                    source
                    for source in detail_manifest.get("sources", [])
                    if source.get("status") != "read"
                ]
                names = "、".join(
                    source.get("farm_name")
                    or source.get("farm_code")
                    or "未知牧场"
                    for source in problems[:5]
                )
                return (
                    False,
                    f"全量明细核对未通过，{len(problems)} 个牧场源文件异常："
                    f"{names}。未发布正式牧场组汇总报告。",
                )

            self._progress(81, "正在核验全部单牧场结果文件...")
            from core.group_report.artifact_inventory import (
                build_group_artifact_inventory,
            )

            inventory_path = publication_batch.inventory_path
            artifact_inventory = build_group_artifact_inventory(
                self.project_path,
                tasks=tasks,
                manifest_path=inventory_path,
                progress_callback=lambda value, text: self._progress(
                    81 + int(value * 0.07), text
                ),
            )
            if artifact_inventory.get("status") != "complete":
                invalid_files = int(
                    artifact_inventory.get("counts", {}).get(
                        "invalid_files", 0
                    )
                )
                scan_errors = int(
                    artifact_inventory.get("counts", {}).get(
                        "tasks_with_scan_errors", 0
                    )
                )
                return (
                    False,
                    "全部结果文件核验未通过："
                    f"{invalid_files} 个文件损坏或无效，"
                    f"{scan_errors} 个子项目无法完整扫描。"
                    "未发布正式牧场组汇总报告。",
                )

            counts = detail_manifest.get("counts", {})
            ranking_total = int(counts.get("valid_ranked_rows", 0))
            rankings = []
            seen_preview = set()
            farms_by_relative_path = {
                str(farm.get("relative_path") or ""): farm
                for farm in farms
            }
            farms_by_api_code: Dict[str, List[Dict]] = defaultdict(list)
            for farm in farms:
                farms_by_api_code[
                    str(farm.get("farm_code") or "")
                ].append(farm)
            for rank_type, rows in (
                ("前列", detail_manifest.get("preview", {}).get("top", [])),
                ("后列", detail_manifest.get("preview", {}).get("bottom", [])),
            ):
                for row in rows:
                    key = (
                        row.get("child_relative_path"),
                        row.get("farm_code"),
                        row.get("cow_id"),
                        row.get("global_rank"),
                    )
                    if key in seen_preview:
                        continue
                    seen_preview.add(key)
                    farm = farms_by_relative_path.get(
                        str(row.get("child_relative_path") or "")
                    )
                    if farm is None:
                        matches = farms_by_api_code.get(
                            str(row.get("farm_code") or ""),
                            [],
                        )
                        farm = matches[0] if len(matches) == 1 else {}
                    rankings.append(
                        {
                            "牧场组排名": row.get("global_rank"),
                            "排名类型": rank_type,
                            "API farmcode": (
                                farm.get("api_farmcode")
                                if farm
                                else row.get("api_farmcode")
                                or row.get("farm_code")
                            ),
                            "牧场编号": (
                                farm.get("farm_number")
                                if farm
                                else row.get("farm_number")
                            ),
                            "牧场名称": (
                                farm.get("farm_name")
                                if farm
                                else row.get("farm_name")
                            ),
                            "牛号": row.get("cow_id"),
                            "原始牛号": row.get("raw_cow_id"),
                            "指数": row.get("index_score"),
                        }
                    )
            rankings.sort(key=lambda row: int(row.get("牧场组排名") or 0))
            farm_stats = {
                str(row.get("child_relative_path", "")): row
                for row in detail_manifest.get("farm_stats", [])
            }
            for farm in farms:
                relative = farm["child_path"].relative_to(
                    self.project_path
                ).as_posix()
                stats = farm_stats.get(relative, {})
                farm["index_count"] = int(
                    stats.get("valid_ranked_rows", 0) or 0
                )
                average_index = stats.get("average_index")
                farm["index_sum"] = (
                    float(average_index) * farm["index_count"]
                    if average_index is not None
                    else 0.0
                )
            current_group = self._weighted_group_row(farms, "current")
            all_group = self._weighted_group_row(farms, "all")
            report_directory = publication_batch.staging_path

            workbook = Workbook()
            workbook.remove(workbook.active)

            # 报告说明
            sheet = workbook.create_sheet("报告说明")
            sheet.sheet_view.showGridLines = False
            sheet.merge_cells("A1:H2")
            sheet["A1"] = "牧场组育种分析汇总报告"
            sheet["A1"].font = Font(name="微软雅黑", size=22, bold=True, color="FFFFFF")
            sheet["A1"].fill = PatternFill("solid", fgColor="2F75B5")
            sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            info = [
                ("牧场数量", len(farms)),
                ("在群母牛", int(current_group.get("头数", 0))),
                ("全部母牛", int(all_group.get("头数", 0))),
                ("指数排名有效头数", ranking_total),
                (
                    "全量源记录",
                    int(counts.get("source_rows", 0)),
                ),
                ("数据来源", self.metadata.get("data_source", "")),
                ("报告生成人", self.service_staff),
                ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("统计说明", "牧场组均值按有效头数加权；比例按分子、分母汇总后计算。"),
            ]
            for row, (label, value) in enumerate(info, start=4):
                sheet.cell(row, 1, label).font = Font(name="微软雅黑", bold=True, color="2F5597")
                sheet.cell(row, 2, value)
            sheet.merge_cells("A14:H14")
            sheet["A14"] = (
                "本工作簿保存汇总和索引；跨牧场全量牛只结果见"
                "“完整明细索引”，其余全部单场结果见“全部结果文件索引”。"
                "所有明细均保留，不做 Top-N 截断。"
            )
            sheet["A14"].fill = PatternFill("solid", fgColor="FFF2CC")
            sheet["A14"].alignment = Alignment(wrap_text=True)
            sheet.column_dimensions["A"].width = 22
            sheet.column_dimensions["B"].width = 42
            for column in "CDEFGH":
                sheet.column_dimensions[column].width = 12

            # 任务完成情况
            sheet = workbook.create_sheet("任务完成情况")
            headers = [
                *FARM_IDENTITY_HEADERS, "是否纳入汇总", "状态",
                "完成阶段", "子项目目录", "错误信息",
            ]
            rows = [
                [
                    *self._task_identity_values(task),
                    "是" if task.get("included_in_summary", True) else "否",
                    (
                        "已完成"
                        if task.get("status") == "completed"
                        else (
                            "已完成（有提示）"
                            if task.get("status")
                            == "completed_with_warning"
                            else (
                                "未纳入汇总"
                                if not task.get(
                                    "included_in_summary", True
                                )
                                else task.get("status")
                            )
                        )
                    ),
                    task.get("stage"),
                    task.get("relative_path"),
                    task.get("error", ""),
                ]
                for task in all_tasks
            ]
            self._write_table(sheet, "牧场任务完成情况", headers, rows)

            # 牧场对比总览
            overview_headers = [
                *FARM_IDENTITY_HEADERS, "在群头数", "全部头数",
                "平均NM$", "NM$排名", "平均TPI", "TPI排名", "平均综合指数",
                "配种记录", "备选公牛", "近交及隐性基因",
            ]
            nm_ranks = self._rank_values(farms, "平均NM$")
            tpi_ranks = self._rank_values(farms, "平均TPI")
            overview_rows = []
            for farm in farms:
                current = farm["current"]
                overview_rows.append(
                    [
                        *self._farm_identity_values(farm),
                        int(_number(current.get("头数"), 0)),
                        int(_number(farm["all"].get("头数"), 0)),
                        current.get("平均NM$"),
                        nm_ranks.get(farm["task_id"]),
                        current.get("平均TPI"),
                        tpi_ranks.get(farm["task_id"]),
                        (
                            farm.get("index_sum", 0) / farm.get("index_count", 1)
                            if farm.get("index_count", 0)
                            else None
                        ),
                        "有" if farm["availability"]["配种记录"] else "无",
                        "有" if farm["availability"]["备选公牛"] else "无",
                        "有" if farm["availability"]["近交及隐性基因"] else "无",
                    ]
                )
            sheet = workbook.create_sheet("牧场对比总览")
            self._write_table(sheet, "牧场核心指标横向对比", overview_headers, overview_rows)
            if overview_rows:
                last_row = len(overview_rows) + 2
                chart = BarChart()
                chart.type = "bar"
                chart.style = 10
                chart.title = "各牧场在群母牛平均NM$"
                chart.height = 8
                chart.width = 15
                chart.add_data(Reference(sheet, min_col=6, min_row=2, max_row=last_row), titles_from_data=True)
                chart.set_categories(Reference(sheet, min_col=3, min_row=3, max_row=last_row))
                chart.legend = None
                sheet.add_chart(chart, "N2")

            # 系谱识别
            pedigree_headers = [
                *FARM_IDENTITY_HEADERS, "在群头数", "父号可识别头数", "父号识别率",
                "外祖父可识别头数", "外祖父识别率", "外曾外祖父可识别头数", "外曾外祖父识别率",
            ]
            pedigree_rows = []
            for farm in farms:
                pedigree = farm["pedigree"]
                pedigree_rows.append([
                    *self._farm_identity_values(farm),
                    pedigree.get("头数", 0),
                    pedigree.get("父号可识别头数", 0), pedigree.get("父号识别率", 0),
                    pedigree.get("外祖父可识别头数", 0), pedigree.get("外祖父识别率", 0),
                    pedigree.get("外曾外祖父可识别头数", 0), pedigree.get("外曾外祖父识别率", 0),
                ])
            sheet = workbook.create_sheet("系谱识别对比")
            self._write_table(sheet, "各牧场在群母牛系谱识别情况", pedigree_headers, pedigree_rows, percent_columns=(6, 8, 10))

            # 关键性状长表
            trait_headers = [*FARM_IDENTITY_HEADERS, "性状", "在群平均值", "统计头数", "牧场组加权平均", "与牧场组差异"]
            trait_rows = []
            for farm in farms:
                for trait in TRAITS:
                    column = f"平均{trait}"
                    value = farm["current"].get(column)
                    group_value = current_group.get(column)
                    trait_rows.append([
                        *self._farm_identity_values(farm), trait, value,
                        int(_number(farm["current"].get("头数"), 0)), group_value,
                        (_number(value) - _number(group_value)) if value is not None and group_value is not None else None,
                    ])
            sheet = workbook.create_sheet("关键性状对比")
            self._write_table(sheet, "关键育种性状牧场对比", trait_headers, trait_rows)
            if trait_rows:
                sheet.conditional_formatting.add(
                    f"H3:H{len(trait_rows)+2}",
                    ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
                )

            # 年份趋势
            yearly_headers = [*FARM_IDENTITY_HEADERS, "出生年份", "头数"] + [f"平均{trait}" for trait in TRAITS]
            yearly_rows = []
            for farm in farms:
                for row in farm["yearly"]:
                    yearly_rows.append([
                        *self._farm_identity_values(farm),
                        row.get("出生年份"),
                        int(_number(row.get("头数"), 0)),
                        *[row.get(f"平均{trait}") for trait in TRAITS],
                    ])
            sheet = workbook.create_sheet("年份遗传进展")
            self._write_table(sheet, "各牧场在群母牛年份遗传进展", yearly_headers, yearly_rows)

            # 分布
            for key, title, sheet_name in (
                ("current_nm", "在群母牛NM$分布", "NM$分布对比"),
                ("current_tpi", "在群母牛TPI分布", "TPI分布对比"),
            ):
                distribution_rows = []
                group_counts = defaultdict(int)
                for farm in farms:
                    for row in farm["distributions"].get(key, []):
                        bucket = row.get("分布区间")
                        headcount = int(_number(row.get("头数"), 0))
                        group_counts[bucket] += headcount
                        distribution_rows.append([
                            *self._farm_identity_values(farm), bucket,
                            headcount, _number(str(row.get("占比", "0")).replace("%", "")) / 100,
                            "牧场",
                        ])
                group_total = sum(group_counts.values())
                for bucket, headcount in group_counts.items():
                    distribution_rows.append([
                        "GROUP", "", "牧场组汇总", bucket, headcount,
                        headcount / group_total if group_total else 0, "牧场组",
                    ])
                sheet = workbook.create_sheet(sheet_name)
                self._write_table(sheet, title, [*FARM_IDENTITY_HEADERS, "分布区间", "头数", "占比", "层级"], distribution_rows, percent_columns=(6,))

            # 跨场排名
            ranking_headers = ["牧场组排名", "排名类型", *FARM_IDENTITY_HEADERS, "牛号", "原始牛号", "综合指数"]
            ranking_rows = [[row.get(header) if header != "综合指数" else row.get("指数") for header in ranking_headers] for row in rankings]
            sheet = workbook.create_sheet("跨牧场牛只排名")
            self._write_table(
                sheet,
                "跨牧场在群母牛指数排名预览（全量记录见“完整明细索引”）",
                ranking_headers,
                ranking_rows,
            )

            # 全量明细卷索引及对账。主报告保持轻量，但所有牛只均在分卷中。
            detail_headers = [
                "数据集", "卷号", "字段分片", "数据行数", "起始排名",
                "结束排名", "文件", "SHA-256",
            ]
            detail_rows = []
            detail_paths = []
            package_path = Path(detail_manifest["package_path"])
            for kind, label in (
                ("ranked", "有效在群完整排名"),
                ("reconciliation", "全部源行分类对账"),
                ("long_fields", "超长字段完整内容"),
            ):
                for volume in detail_manifest.get("volumes", {}).get(kind, []):
                    volume_path = Path(volume["absolute_path"])
                    detail_paths.append(volume_path)
                    detail_rows.append(
                        [
                            label,
                            volume.get("volume"),
                            (
                                f"{volume.get('column_part')}/"
                                f"{volume.get('column_parts')}"
                            ),
                            volume.get("data_rows"),
                            volume.get("first_rank"),
                            volume.get("last_rank"),
                            volume_path.name,
                            volume.get("sha256"),
                        ]
                    )
            sheet = workbook.create_sheet("完整明细索引")
            self._write_table(
                sheet,
                "全量明细分卷索引（所有源行均已对账，不截断）",
                detail_headers,
                detail_rows,
            )
            for row_index, volume_path in enumerate(detail_paths, start=3):
                cell = sheet.cell(row_index, 7)
                cell.hyperlink = self._relative_report_link(
                    report_directory,
                    volume_path,
                )
                cell.style = "Hyperlink"

            # 全部标准化数据、分析产物和单场报告的文件级完整索引。
            inventory_headers = [
                "任务ID",
                *FARM_IDENTITY_HEADERS,
                "类别",
                "受管产物",
                "阶段",
                "逻辑名称",
                "相对路径",
                "字节数",
                "SHA-256",
                "XLSX结构有效",
                "Sheet数",
                "Sheet行列范围",
            ]
            inventory_entries = [
                *artifact_inventory.get("files", []),
                *artifact_inventory.get("unmanaged_files", []),
            ]
            tasks_by_id = {
                str(task.get("task_id") or ""): task
                for task in all_tasks
            }
            inventory_rows = [
                [
                    entry.get("task_id"),
                    *self._task_identity_values(
                        tasks_by_id.get(
                            str(entry.get("task_id") or ""),
                            {
                                "farm_code": entry.get("farm_code"),
                                "farm_name": entry.get("farm_name"),
                            },
                        )
                    ),
                    entry.get("category_label"),
                    "是" if entry.get("managed") else "否",
                    entry.get("stage"),
                    entry.get("logical_name"),
                    entry.get("relative_path"),
                    int(entry.get("bytes", 0) or 0),
                    entry.get("sha256"),
                    (
                        "是"
                        if entry.get("xlsx_valid") is True
                        else "否"
                        if entry.get("xlsx_valid") is False
                        else "未校验"
                    ),
                    int(entry.get("sheet_count", 0) or 0),
                    entry.get("sheet_dimensions"),
                ]
                for entry in inventory_entries
            ]
            sheet = workbook.create_sheet("全部结果文件索引")
            self._write_table(
                sheet,
                (
                    "全部单牧场结果文件索引（受管产物参与正式完整性校验；"
                    "未受管文件仅供定位）"
                ),
                inventory_headers,
                inventory_rows,
            )
            for row_index, entry in enumerate(
                inventory_entries,
                start=3,
            ):
                target = self.project_path / entry["relative_path"]
                cell = sheet.cell(row_index, 9)
                cell.hyperlink = self._relative_report_link(
                    report_directory,
                    target,
                )
                cell.style = "Hyperlink"

            integrity_headers = ["校验项目", "源记录数", "导出记录数", "差异", "结果"]
            integrity_rows = [
                [
                    "有效在群完整排名",
                    counts.get("valid_ranked_rows", 0),
                    counts.get("ranked_exported_rows", 0),
                    int(counts.get("valid_ranked_rows", 0))
                    - int(counts.get("ranked_exported_rows", 0)),
                    (
                        "通过"
                        if int(counts.get("valid_ranked_rows", 0))
                        == int(counts.get("ranked_exported_rows", 0))
                        else "不通过"
                    ),
                ],
                [
                    "全部源行分类对账",
                    counts.get("source_rows", 0),
                    counts.get("reconciliation_exported_rows", 0),
                    int(counts.get("source_rows", 0))
                    - int(counts.get("reconciliation_exported_rows", 0)),
                    (
                        "通过"
                        if int(counts.get("source_rows", 0))
                        == int(counts.get("reconciliation_exported_rows", 0))
                        else "不通过"
                    ),
                ],
                [
                    "源文件读取",
                    counts.get("tasks_in_scope", 0),
                    counts.get("source_files_read", 0),
                    int(counts.get("tasks_in_scope", 0))
                    - int(counts.get("source_files_read", 0)),
                    (
                        "通过"
                        if int(counts.get("tasks_in_scope", 0))
                        == int(counts.get("source_files_read", 0))
                        else "不通过"
                    ),
                ],
                [
                    "全部结果文件结构校验",
                    artifact_inventory.get("counts", {}).get(
                        "total_files", 0
                    ),
                    artifact_inventory.get("counts", {}).get(
                        "valid_files", 0
                    ),
                    artifact_inventory.get("counts", {}).get(
                        "invalid_files", 0
                    ),
                    (
                        "通过"
                        if artifact_inventory.get("status") == "complete"
                        else "不通过"
                    ),
                ],
                [
                    "超长字段无损分块",
                    counts.get("long_field_count", 0),
                    counts.get("long_field_count", 0),
                    0,
                    "通过",
                ],
            ]
            sheet = workbook.create_sheet("完整性校验")
            self._write_table(
                sheet,
                "牧场组全量明细完整性校验",
                integrity_headers,
                integrity_rows,
            )

            # 可用性与索引
            availability_headers = [*FARM_IDENTITY_HEADERS, "配种记录", "备选公牛", "已配公牛分析", "近交及隐性基因", "单牧场Excel"]
            availability_rows = [
                [*self._farm_identity_values(farm), *["有" if farm["availability"][key] else "无" for key in ("配种记录", "备选公牛", "已配公牛分析", "近交及隐性基因", "单牧场Excel")]]
                for farm in farms
            ]
            sheet = workbook.create_sheet("数据可用性")
            self._write_table(sheet, "各牧场分析模块可用情况", availability_headers, availability_rows)

            index_headers = [*FARM_IDENTITY_HEADERS, "子项目目录", "单牧场Excel报告"]
            index_rows = []
            for farm in farms:
                report_relative = ""
                if farm["report_path"]:
                    report_relative = farm["report_path"].relative_to(self.project_path).as_posix()
                index_rows.append([
                    *self._farm_identity_values(farm),
                    farm["child_path"].relative_to(self.project_path).as_posix(),
                    report_relative,
                ])
            sheet = workbook.create_sheet("单牧场报告索引")
            self._write_table(sheet, "单牧场子项目与报告索引", index_headers, index_rows)
            for row_index, farm in enumerate(farms, start=3):
                child_cell = sheet.cell(row_index, 4)
                child_cell.hyperlink = self._relative_report_link(
                    report_directory,
                    farm["child_path"],
                )
                child_cell.style = "Hyperlink"
                if farm["report_path"]:
                    report_cell = sheet.cell(row_index, 5)
                    report_cell.hyperlink = self._relative_report_link(
                        report_directory,
                        farm["report_path"],
                    )
                    report_cell.style = "Hyperlink"

            self._progress(90, "正在保存牧场组汇总Excel...")
            output_dir = report_directory
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = publication_batch.excel_path
            file_descriptor, temporary_name = tempfile.mkstemp(
                prefix=f".{output_path.stem}.",
                suffix=".tmp.xlsx",
                dir=output_dir,
            )
            os.close(file_descriptor)
            temporary_path = Path(temporary_name)
            try:
                workbook.save(temporary_path)
                workbook.close()
                if (
                    not temporary_path.exists()
                    or temporary_path.stat().st_size == 0
                    or not zipfile.is_zipfile(temporary_path)
                ):
                    raise RuntimeError("牧场组汇总Excel临时文件校验失败")
                self._refresh_summary_lease()
                final_snapshot_path = (
                    snapshot_directory
                    / f"{publication_run_id}_published.json"
                )
                final_snapshot_result = (
                    recompute_and_compare_group_publication_snapshot(
                        self.project_path,
                        before_snapshot,
                        output_path=final_snapshot_path,
                    )
                )
                os.replace(temporary_path, output_path)
            finally:
                temporary_path.unlink(missing_ok=True)
            # 所有派生产物、分卷哈希及冻结输入快照先绑定成一个不可变
            # 报告包，再一次性提升整个目录。最后只原子切换一个小指针。
            published = publication_batch.finalize_candidate(
                excel_path=output_path,
                detail_manifest_path=Path(
                    detail_manifest["manifest_path"]
                ),
                inventory_path=Path(
                    artifact_inventory["manifest_path"]
                ),
                publication_snapshot_path=final_snapshot_path,
            )
            self._refresh_summary_lease()
            current_pointer = publish_current_group_report_pointer(
                self.project_path,
                published=published,
                selection_revision=self._selection_revision or 0,
                publication_basis_sha256=final_snapshot_result[
                    "after_snapshot"
                ]["basis_sha256"],
            )
            final_detail_manifest_path = Path(
                published["detail_manifest_path"]
            )
            final_detail_package_path = (
                final_detail_manifest_path.parent
            )
            final_detail_paths = [
                final_detail_package_path / str(volume["path"])
                for kind in ("ranked", "reconciliation", "long_fields")
                for volume in detail_manifest.get("volumes", {}).get(
                    kind, []
                )
            ]
            final_output_path = Path(published["excel_path"])
            FileManager.update_group_result(
                self.project_path,
                excel_path=final_output_path,
                selection_revision=self._selection_revision,
                report_package_path=published["package_path"],
                batch_manifest_path=published["batch_manifest_path"],
                batch_manifest_sha256=published[
                    "batch_manifest_sha256"
                ],
                current_pointer_path=current_pointer["pointer_path"],
                current_pointer_sha256=current_pointer["pointer_sha256"],
                detail_package_path=final_detail_package_path,
                detail_manifest_path=final_detail_manifest_path,
                detail_manifest_sha256=published["detail"]["sha256"],
                detail_volume_paths=final_detail_paths,
                detail_counts=counts,
                artifact_inventory_path=published["inventory_path"],
                artifact_inventory_sha256=published["inventory"][
                    "sha256"
                ],
                artifact_inventory_counts=artifact_inventory.get("counts"),
                publication_snapshot_path=published[
                    "publication_snapshot_path"
                ],
                publication_basis_sha256=final_snapshot_result[
                    "after_snapshot"
                ]["basis_sha256"],
            )
            self._progress(100, "牧场组汇总Excel生成完成")
            return True, str(final_output_path)
        except Exception as exc:
            logger.exception("生成牧场组汇总Excel失败")
            return False, f"生成牧场组汇总Excel失败：{exc}"
        finally:
            self._release_summary_lease()

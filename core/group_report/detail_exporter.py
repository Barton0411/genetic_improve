"""牧场组全量牛只排名明细的低内存、可分卷导出。

本模块刻意不把多个牧场的宽表同时装入 pandas 或内存。源工作簿逐行写入
临时 SQLite，SQLite 以未舍入的综合指数完成稳定全局排序，随后再以
XlsxWriter ``constant_memory`` 模式分卷落盘。

导出包同时包含两套文件：

* ``有效在群完整排名``：所有符合条件的在群母牛，不做 Top-N 截断；
* ``全部源行分类对账``：每一个可读取的源数据行，以及它是否参与排名、
  未参与排名的具体原因。

单个 xlsx 只受 Excel 自身行列上限约束；超过阈值后自动增加卷，不牺牲
明细完整性。整个导出包先在同盘临时目录生成，全部写完并完成行数核对后
再原子改名为正式目录。
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import math
import numbers
import os
import shutil
import sqlite3
import tempfile
import zipfile
from collections import Counter
from pathlib import Path
from typing import Callable, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import xlsxwriter
from openpyxl import load_workbook

from core.group_report.exact_decimal import (
    ExactDecimal,
    ExactDecimalError,
    parse_exact_decimal,
    sqlite_order_by_clause,
)


EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
EXCEL_MAX_CELL_CHARACTERS = 32_767
DETAIL_EXPORT_SCHEMA_VERSION = 2
LONG_TEXT_CHUNK_CHARACTERS = 30_000
DEFAULT_ROWS_PER_VOLUME = None
DEFAULT_MAX_CELLS_PER_VOLUME = 5_000_000
DEFAULT_MAX_ROWS_PER_VOLUME = 500_000
DEFAULT_MAX_SOURCE_COLUMNS_PER_PART = 256
DEFAULT_MAX_ESTIMATED_XML_BYTES_PER_VOLUME = 512 * 1024 * 1024
DEFAULT_SQLITE_BATCH_ROWS = 1_000
DEFAULT_SQLITE_BATCH_BYTES = 8 * 1024 * 1024
DEFAULT_DISK_RESERVE_BYTES = 2 * 1024 * 1024 * 1024
DEFAULT_DISK_CHECK_INTERVAL_BYTES = 64 * 1024 * 1024

RANKED_CLASSIFICATION = "有效在群排名"
UNRANKED_CLASSIFICATION = "未排名"

_YES_VALUES = frozenset({"是", "在场", "在群", "1", "true", "y", "yes"})
_MISSING_IDENTIFIER_VALUES = frozenset(
    {"", "nan", "none", "null", "nat", "<na>", "n/a"}
)
_TEXT_SOURCE_COLUMNS = frozenset(
    {
        "cow_id",
        "raw_cow_id",
        "母牛号",
        "牛号",
        "耳号",
        "sire",
        "dam",
        "mgs",
        "mgd",
        "mmgs",
        "父号",
        "母号",
        "外祖父号",
        "外祖母号",
        "外曾祖父号",
        "farmcode",
        "farm_code",
        "api farmcode",
        "牧场编号",
    }
)

_FIXED_HEADERS = (
    "牧场组排名",
    "分类结果",
    "未排名原因",
    "API farmcode",
    "牧场名称",
    "子项目相对目录",
    "源文件",
    "源数据行号",
    "指数列",
    "综合指数_未舍入",
    "综合指数_精确文本",
)


class GroupDetailExportPaused(RuntimeError):
    """资源不足时的可恢复暂停，而不是丢弃进度的计算失败。"""

    def __init__(self, message: str, *, phase: str, details: Optional[Dict] = None):
        super().__init__(message)
        self.phase = phase
        self.details = dict(details or {})


def _progress(
    callback: Optional[Callable[[int, str], None]],
    value: int,
    message: str,
) -> None:
    if callback:
        callback(max(0, min(int(value), 100)), message)


def _sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            chunk = stream.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _write_json_atomic(path: Path, value: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.stem}.",
        suffix=".tmp.json",
        dir=path.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(file_descriptor, "w", encoding="utf-8") as stream:
            json.dump(
                value,
                stream,
                ensure_ascii=False,
                indent=2,
                allow_nan=False,
            )
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary_path, path)
    except Exception:
        try:
            os.close(file_descriptor)
        except OSError:
            pass
        temporary_path.unlink(missing_ok=True)
        raise


def _identifier(value) -> str:
    """把 Excel 数值形式的标识符稳定转为文本。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, numbers.Integral):
        return str(int(value))
    if isinstance(value, numbers.Real):
        numeric_value = float(value)
        if math.isnan(numeric_value):
            return ""
        if math.isfinite(numeric_value) and numeric_value.is_integer():
            return str(int(numeric_value))
    text = str(value).strip()
    return "" if text.casefold() in _MISSING_IDENTIFIER_VALUES else text


def _task_farm_identity(task: Dict) -> Tuple[str, str, str]:
    metadata = task.get("metadata", {})
    if not isinstance(metadata, dict):
        metadata = {}
    farm_code = _identifier(task.get("farm_code"))
    source_kind = str(task.get("source_kind") or "api")
    api_farmcode = _identifier(
        task.get("api_farmcode")
        if "api_farmcode" in task
        else metadata.get(
            "api_farmcode",
            farm_code if source_kind != "local" else "",
        )
    )
    if "farm_number" in task:
        farm_number = _identifier(task.get("farm_number"))
    elif "farm_number" in metadata:
        farm_number = _identifier(metadata.get("farm_number"))
    else:
        farm_number = (
            ""
            if str(task.get("source_system") or "") == "慧牧云"
            else farm_code
        )
    return api_farmcode, farm_number, str(task.get("farm_name") or "")


def _identity_state() -> Dict[str, int]:
    return {
        "row_count": 0,
        "blank_id_count": 0,
        "hash_sum": 0,
        "hash_xor": 0,
    }


def _add_identity(state: Dict[str, int], identifier: str) -> None:
    """构建与行序无关、保留重复次数的牛号多重集指纹。"""
    normalized = _identifier(identifier)
    state["row_count"] += 1
    if not normalized:
        state["blank_id_count"] += 1
    digest = int.from_bytes(
        hashlib.sha256(normalized.encode("utf-8")).digest(),
        "big",
    )
    modulus_mask = (1 << 256) - 1
    state["hash_sum"] = (state["hash_sum"] + digest) & modulus_mask
    state["hash_xor"] ^= digest


def _public_identity(state: Dict[str, int]) -> Dict:
    return {
        "row_count": int(state["row_count"]),
        "blank_id_count": int(state["blank_id_count"]),
        "hash_sum": f"{int(state['hash_sum']):064x}",
        "hash_xor": f"{int(state['hash_xor']):064x}",
    }


def _json_safe(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        numeric_value = float(value)
        if not math.isfinite(numeric_value):
            return str(value)
        return numeric_value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return str(value)
    if isinstance(value, str):
        return value
    return str(value)


def _score(value) -> Tuple[Optional[ExactDecimal], str]:
    if value is None:
        return None, "综合指数为空"
    if isinstance(value, bool):
        return None, "综合指数不是数值"
    if isinstance(value, str) and not value.strip():
        return None, "综合指数为空"
    try:
        numeric_value = parse_exact_decimal(value)
    except (ExactDecimalError, TypeError, ValueError, OverflowError):
        return None, "综合指数不是数值"
    return numeric_value, ""


def _approximate_score(value: Optional[ExactDecimal]) -> Optional[float]:
    """仅用于平均值和兼容展示；排名始终使用精确十进制字段。"""
    if value is None:
        return None
    try:
        approximate = float(value.value)
    except (OverflowError, ValueError):
        return None
    return approximate if math.isfinite(approximate) else None


def _deduplicate_headers(values: Sequence) -> List[str]:
    counts: Counter = Counter()
    output = []
    for column_index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else ""
        if not base:
            base = f"未命名列_{column_index}"
        counts[base] += 1
        output.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return output


def _first_column(headers: Sequence[str], candidates: Iterable[str]) -> Optional[str]:
    positions = {str(header).strip().casefold(): header for header in headers}
    for candidate in candidates:
        found = positions.get(candidate.casefold())
        if found is not None:
            return found
    return None


def _detect_score_column(
    headers: Sequence[str],
    requested: Optional[str],
) -> Optional[str]:
    if requested:
        return _first_column(headers, (requested,))
    candidates = [
        header
        for header in headers
        if str(header).strip().casefold().endswith("_index")
    ]
    if not candidates:
        return None
    # 计算模块把当前使用的指数列放在 ranking 前；选择最后一个可兼容文件中
    # 仍保留旧指数列的情况。
    return candidates[-1]


def _is_blank_row(values: Sequence) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _text_source_column(column_name: str) -> bool:
    folded = column_name.strip().casefold()
    return folded in _TEXT_SOURCE_COLUMNS or folded.endswith(
        ("_id", "_code", "编号", "牛号", "耳号")
    )


class _AtomicXlsxVolume:
    """一个恒定内存、原子替换的 xlsx 卷。"""

    def __init__(
        self,
        output_path: Path,
        sheet_name: str,
        headers: Sequence[str],
        text_column_indexes: Iterable[int],
        max_estimated_xml_bytes: int = (
            DEFAULT_MAX_ESTIMATED_XML_BYTES_PER_VOLUME
        ),
    ):
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        file_descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{output_path.stem}.",
            suffix=".tmp.xlsx",
            dir=output_path.parent,
        )
        os.close(file_descriptor)
        self.temporary_path = Path(temporary_name)
        self.workbook = xlsxwriter.Workbook(
            str(self.temporary_path),
            {
                "constant_memory": True,
                "use_zip64": True,
                # XML 临时文件与输出卷放在同一文件系统，磁盘检查与
                # 原子替换才覆盖真实峰值，也避免系统临时盘先被写满。
                "tmpdir": str(self.output_path.parent),
                "strings_to_formulas": False,
                "strings_to_urls": False,
                "remove_timezone": True,
                "default_date_format": "yyyy-mm-dd hh:mm:ss",
            },
        )
        # 兼容不识别构造参数的旧版 XlsxWriter。
        self.workbook.use_zip64()
        self.worksheet = self.workbook.add_worksheet(sheet_name[:31])
        self.header_format = self.workbook.add_format(
            {
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": "#4472C4",
                "border": 1,
                "align": "center",
                "valign": "vcenter",
            }
        )
        self.text_format = self.workbook.add_format({"num_format": "@"})
        self.score_format = self.workbook.add_format(
            {"num_format": "0.000000000000"}
        )
        self.headers = list(headers)
        self.text_column_indexes = set(text_column_indexes)
        self.max_estimated_xml_bytes = max(
            64 * 1024 * 1024,
            int(max_estimated_xml_bytes),
        )
        self.data_rows = 0
        self.first_rank: Optional[int] = None
        self.last_rank: Optional[int] = None
        self.estimated_xml_bytes = self.estimate_row_bytes(self.headers)
        self.closed = False

        self.worksheet.write_row(0, 0, self.headers, self.header_format)
        self.worksheet.freeze_panes(1, 0)
        for column_index in self.text_column_indexes:
            self.worksheet.set_column(
                column_index,
                column_index,
                None,
                self.text_format,
            )
        # 固定审计列给出可读宽度；源字段保持适中，避免打开超宽工作簿时卡顿。
        fixed_widths = (12, 14, 22, 16, 24, 32, 38, 12, 22, 20, 24)
        for column_index, width in enumerate(fixed_widths[: len(self.headers)]):
            self.worksheet.set_column(column_index, column_index, width)
        if len(self.headers) > len(_FIXED_HEADERS):
            self.worksheet.set_column(
                len(_FIXED_HEADERS),
                len(self.headers) - 1,
                14,
            )
        if len(self.headers) >= 10:
            self.worksheet.set_column(9, 9, 20, self.score_format)

    @staticmethod
    def estimate_row_bytes(values: Sequence) -> int:
        """粗估 sheet XML 增量，用于在关闭卷前限制临时文件峰值。"""
        total = 64
        for value in values:
            if value is None:
                total += 24
            else:
                total += len(str(value).encode("utf-8")) * 2 + 64
        return total

    def can_fit(self, values: Sequence) -> bool:
        if self.data_rows == 0:
            return True
        return (
            self.estimated_xml_bytes + self.estimate_row_bytes(values)
            <= self.max_estimated_xml_bytes
        )

    def write(self, values: Sequence, rank: Optional[int]) -> None:
        normalized = []
        for column_index, value in enumerate(values):
            if value is None:
                normalized.append(None)
            elif column_index in self.text_column_indexes:
                normalized.append(_identifier(value))
            else:
                normalized.append(_json_safe(value))
        for column_index, value in enumerate(normalized):
            if (
                isinstance(value, str)
                and len(value) > EXCEL_MAX_CELL_CHARACTERS
            ):
                digest = hashlib.sha256(value.encode("utf-8")).hexdigest()
                normalized[column_index] = (
                    "[超长字段，完整内容见“超长字段完整内容”分卷；"
                    f"字符数={len(value)}；SHA256={digest}]"
                )
        self.worksheet.write_row(self.data_rows + 1, 0, normalized)
        self.estimated_xml_bytes += self.estimate_row_bytes(normalized)
        self.data_rows += 1
        if rank is not None:
            self.first_rank = rank if self.first_rank is None else self.first_rank
            self.last_rank = rank

    def close(self) -> Dict:
        if self.closed:
            raise RuntimeError("Excel 卷已经关闭")
        try:
            if self.headers:
                self.worksheet.autofilter(
                    0,
                    0,
                    max(0, self.data_rows),
                    len(self.headers) - 1,
                )
            self.workbook.close()
            if (
                not self.temporary_path.exists()
                or self.temporary_path.stat().st_size == 0
                or not zipfile.is_zipfile(self.temporary_path)
            ):
                raise RuntimeError("Excel分卷临时文件校验失败")
            with zipfile.ZipFile(self.temporary_path) as archive:
                required = {"[Content_Types].xml", "xl/workbook.xml"}
                if not required.issubset(archive.namelist()):
                    raise RuntimeError("Excel分卷缺少必需结构文件")
                bad_member = archive.testzip()
                if bad_member is not None:
                    raise RuntimeError(
                        f"Excel分卷CRC校验失败：{bad_member}"
                    )
            os.replace(self.temporary_path, self.output_path)
            self.closed = True
            return {
                "data_rows": self.data_rows,
                "first_rank": self.first_rank,
                "last_rank": self.last_rank,
                "bytes": self.output_path.stat().st_size,
                "estimated_xml_bytes": int(self.estimated_xml_bytes),
                "sha256": _sha256(self.output_path),
            }
        except Exception:
            self.abort()
            raise

    def abort(self) -> None:
        if self.closed:
            return
        # ``Workbook.close()`` 会把 sheet XML 再压成 ZIP，不是真正的
        # abort；在磁盘余量触发安全暂停时调用它反而可能耗尽剩余空间。
        # constant_memory 模式尚未打包时只需关闭并删除各 sheet 临时文件。
        for worksheet in self.workbook.worksheets():
            try:
                worksheet._opt_close()
            except Exception:
                pass
            row_data_filename = getattr(
                worksheet,
                "row_data_filename",
                None,
            )
            if row_data_filename:
                Path(row_data_filename).unlink(missing_ok=True)
        self.workbook.fileclosed = True
        self.temporary_path.unlink(missing_ok=True)
        self.closed = True


class GroupCowRankingDetailExporter:
    """流式生成牧场组牛只全量排名与源行分类对账包。"""

    def __init__(
        self,
        project_path: Path,
        *,
        rows_per_volume: Optional[int] = DEFAULT_ROWS_PER_VOLUME,
        score_column: Optional[str] = None,
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ):
        self.project_path = Path(project_path)
        self.rows_per_volume = (
            int(rows_per_volume)
            if rows_per_volume is not None
            else None
        )
        if self.rows_per_volume is not None and not (
            1 <= self.rows_per_volume <= EXCEL_MAX_ROWS - 1
        ):
            raise ValueError(
                f"每卷数据行数必须在 1 到 {EXCEL_MAX_ROWS - 1:,} 之间"
            )
        self.score_column = score_column
        self.progress_callback = progress_callback

    def _load_tasks(self, tasks: Optional[Sequence[Dict]]) -> List[Dict]:
        if tasks is None:
            metadata_path = self.project_path / "project_metadata.json"
            with metadata_path.open("r", encoding="utf-8") as stream:
                metadata = json.load(stream)
            tasks = metadata.get("group_tasks", [])
        active = [
            dict(task)
            for task in tasks
            if task.get("included_in_summary", True)
        ]
        return sorted(
            active,
            key=lambda task: (
                str(task.get("farm_code", "")),
                str(task.get("farm_name", "")),
                str(task.get("relative_path", task.get("child_path", ""))),
            ),
        )

    def _disk_preflight(
        self,
        task_list: Sequence[Dict],
        output_dir: Path,
    ) -> Dict:
        """做启动前粗估；运行中仍按真实增长量持续检查并可恢复暂停。"""
        compressed_input_bytes = 0
        for task in task_list:
            relative_path = str(
                task.get("relative_path")
                or task.get("child_path")
                or ""
            )
            child_path = (
                Path(task["child_path"])
                if task.get("child_path")
                else self.project_path / relative_path
            )
            for filename in (
                "processed_index_cow_index_scores.xlsx",
                "processed_cow_data_key_traits_final.xlsx",
            ):
                path = child_path / "analysis_results" / filename
                if path.is_file():
                    compressed_input_bytes += path.stat().st_size
        # xlsx 是压缩容器，SQLite payload、排序临时空间、排名与对账两套
        # 分卷的峰值不能只按压缩文件的 6 倍估算。这里用更保守的启动
        # 粗估；真正决定是否暂停的是每个安全检查点的动态磁盘检查。
        estimated_work_bytes = max(
            512 * 1024 * 1024,
            compressed_input_bytes * 10,
        )
        reserve_bytes = self._disk_reserve_bytes(output_dir)
        free_bytes = shutil.disk_usage(output_dir).free
        required_free_bytes = estimated_work_bytes + reserve_bytes
        result = {
            "compressed_input_bytes": int(compressed_input_bytes),
            "estimated_work_bytes": int(estimated_work_bytes),
            "reserve_bytes": int(reserve_bytes),
            "required_free_bytes": int(required_free_bytes),
            "free_bytes": int(free_bytes),
            "passed": free_bytes >= required_free_bytes,
        }
        if not result["passed"]:
            gib = 1024 ** 3
            raise GroupDetailExportPaused(
                "牧场组完整明细已安全暂停：磁盘可用空间约 "
                f"{free_bytes / gib:.1f} GiB，启动当前批次预计至少需要 "
                f"{required_free_bytes / gib:.1f} GiB。"
                "单牧场结果均已保留，释放空间后可从断点继续。",
                phase="preflight",
                details=result,
            )
        return result

    @staticmethod
    def _disk_reserve_bytes(path: Path) -> int:
        total = int(shutil.disk_usage(path).total)
        return max(
            512 * 1024 * 1024,
            min(DEFAULT_DISK_RESERVE_BYTES, int(total * 0.02)),
        )

    def _ensure_free_space(
        self,
        path: Path,
        *,
        phase: str,
        transient_bytes: int = 0,
        details: Optional[Dict] = None,
    ) -> Dict:
        """在可提交边界检查磁盘，不让文件系统被写满后才失败。"""
        usage = shutil.disk_usage(path)
        reserve_bytes = self._disk_reserve_bytes(path)
        required_free_bytes = reserve_bytes + max(0, int(transient_bytes))
        state = {
            "phase": phase,
            "total_bytes": int(usage.total),
            "used_bytes": int(usage.used),
            "free_bytes": int(usage.free),
            "reserve_bytes": reserve_bytes,
            "transient_bytes": max(0, int(transient_bytes)),
            "required_free_bytes": required_free_bytes,
            **dict(details or {}),
        }
        if usage.free < required_free_bytes:
            gib = 1024 ** 3
            raise GroupDetailExportPaused(
                "牧场组完整明细已安全暂停："
                f"{phase}阶段可用空间约 {usage.free / gib:.1f} GiB，"
                f"至少需保留 {required_free_bytes / gib:.1f} GiB。"
                "已完成的牧场和分卷不会重算，释放空间后可继续。",
                phase=phase,
                details=state,
            )
        return state

    def _volume_xml_budget(self, path: Path, *, phase: str) -> int:
        """按当前磁盘余量自适应单卷临时 XML 上限。"""
        usage = shutil.disk_usage(path)
        reserve = self._disk_reserve_bytes(path)
        headroom = int(usage.free) - reserve
        minimum = 64 * 1024 * 1024
        if headroom < minimum * 3:
            self._ensure_free_space(
                path,
                phase=phase,
                transient_bytes=minimum * 3,
            )
        return min(
            DEFAULT_MAX_ESTIMATED_XML_BYTES_PER_VOLUME,
            max(minimum, headroom // 3),
        )

    @staticmethod
    def _database_footprint(path: Path) -> int:
        return sum(
            candidate.stat().st_size
            for candidate in (
                path,
                Path(f"{path}-wal"),
                Path(f"{path}-shm"),
            )
            if candidate.exists()
        )

    @staticmethod
    def _create_database(path: Path) -> sqlite3.Connection:
        path.parent.mkdir(parents=True, exist_ok=True)

        def quarantine_database(reason: str) -> None:
            history = path.parent / "corrupt_history"
            history.mkdir(parents=True, exist_ok=True)
            stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            moved = []
            for suffix in ("", "-wal", "-shm"):
                candidate = Path(f"{path}{suffix}")
                if not candidate.exists():
                    continue
                destination = history / f"{stamp}_{candidate.name}"
                os.replace(candidate, destination)
                moved.append(destination.name)
            _write_json_atomic(
                history / f"{stamp}_recovery.json",
                {
                    "schema_version": 1,
                    "detected_at": dt.datetime.now().isoformat(
                        timespec="seconds"
                    ),
                    "reason": reason,
                    "quarantined_files": moved,
                    "action": "已隔离损坏的未发布磁盘索引，将从已完成源重新建立",
                },
            )

        def open_connection() -> sqlite3.Connection:
            connection_to_open = sqlite3.connect(str(path))
            try:
                connection_to_open.execute("PRAGMA journal_mode=WAL")
                connection_to_open.execute("PRAGMA synchronous=NORMAL")
                connection_to_open.execute("PRAGMA temp_store=FILE")
                connection_to_open.execute("PRAGMA cache_size=-32768")
                connection_to_open.execute("PRAGMA busy_timeout=30000")
                # SQLite 排序临时文件固定放在可监控的同一工作目录。部分
                # SQLite 构建可能禁用该 PRAGMA，失败时仍使用系统默认目录。
                escaped = str(path.parent).replace("'", "''")
                try:
                    connection_to_open.execute(
                        f"PRAGMA temp_store_directory='{escaped}'"
                    )
                except sqlite3.DatabaseError:
                    pass
                return connection_to_open
            except Exception:
                connection_to_open.close()
                raise

        try:
            connection = open_connection()
            quick_check = connection.execute("PRAGMA quick_check").fetchone()
            if not quick_check or str(quick_check[0]).casefold() != "ok":
                reason = (
                    str(quick_check[0])
                    if quick_check
                    else "quick_check 未返回结果"
                )
                connection.close()
                quarantine_database(reason)
                connection = open_connection()
        except sqlite3.DatabaseError as exc:
            try:
                connection.close()
            except (UnboundLocalError, sqlite3.Error):
                pass
            quarantine_database(f"{type(exc).__name__}: {exc}")
            connection = open_connection()
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_key TEXT NOT NULL,
                farm_code TEXT NOT NULL,
                farm_name TEXT NOT NULL,
                child_relative_path TEXT NOT NULL,
                source_file TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                cow_id TEXT NOT NULL,
                raw_cow_id TEXT NOT NULL,
                in_herd_raw TEXT NOT NULL,
                score_column TEXT NOT NULL,
                score_value REAL,
                score_text TEXT NOT NULL DEFAULT '',
                score_sign INTEGER NOT NULL DEFAULT 0,
                score_adjusted_exponent INTEGER NOT NULL DEFAULT 0,
                score_digits TEXT NOT NULL DEFAULT '0',
                classification TEXT NOT NULL,
                unranked_reason TEXT NOT NULL,
                has_long_text INTEGER NOT NULL DEFAULT 0,
                payload_json TEXT NOT NULL
            )
            """
        )
        existing_columns = {
            str(row[1])
            for row in connection.execute("PRAGMA table_info(records)")
        }
        migrations = {
            "score_text": "TEXT NOT NULL DEFAULT ''",
            "score_sign": "INTEGER NOT NULL DEFAULT 0",
            "score_adjusted_exponent": (
                "INTEGER NOT NULL DEFAULT 0"
            ),
            "score_digits": "TEXT NOT NULL DEFAULT '0'",
            "has_long_text": "INTEGER NOT NULL DEFAULT 0",
        }
        for column, definition in migrations.items():
            if column not in existing_columns:
                connection.execute(
                    f'ALTER TABLE records ADD COLUMN "{column}" '
                    f"{definition}"
                )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS records_long_text
            ON records(id)
            WHERE has_long_text = 1
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS source_checkpoints (
                source_key TEXT PRIMARY KEY,
                input_fingerprint TEXT NOT NULL,
                source_result_json TEXT NOT NULL,
                source_columns_json TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                completed_at TEXT NOT NULL
            )
            """
        )
        return connection

    @staticmethod
    def _classify(
        row: Sequence,
        positions: Dict[str, int],
        cow_column: Optional[str],
        raw_cow_column: Optional[str],
        in_herd_column: Optional[str],
        score_column: Optional[str],
    ) -> Tuple[str, str, str, str, Optional[ExactDecimal]]:
        cow_id = _identifier(row[positions[cow_column]]) if cow_column else ""
        raw_cow_id = (
            _identifier(row[positions[raw_cow_column]])
            if raw_cow_column
            else cow_id
        )
        score_value = None
        score_error = ""
        if score_column is not None:
            score_value, score_error = _score(row[positions[score_column]])
        if in_herd_column is None:
            return (
                UNRANKED_CLASSIFICATION,
                "缺少是否在场列",
                cow_id,
                raw_cow_id,
                score_value,
            )
        in_herd_raw = _identifier(row[positions[in_herd_column]])
        if in_herd_raw.strip().casefold() not in _YES_VALUES:
            reason = "是否在场为空" if not in_herd_raw else "非在群母牛"
            return (
                UNRANKED_CLASSIFICATION,
                reason,
                cow_id,
                raw_cow_id,
                score_value,
            )
        if not cow_id:
            return (
                UNRANKED_CLASSIFICATION,
                "牛号为空",
                cow_id,
                raw_cow_id,
                score_value,
            )
        if score_column is None:
            return (
                UNRANKED_CLASSIFICATION,
                "缺少综合指数列",
                cow_id,
                raw_cow_id,
                None,
            )
        if score_error:
            return (
                UNRANKED_CLASSIFICATION,
                score_error,
                cow_id,
                raw_cow_id,
                None,
            )
        return RANKED_CLASSIFICATION, "", cow_id, raw_cow_id, score_value

    @staticmethod
    def _read_identity_source(path: Path) -> Dict:
        """流式读取指数的直接性状输入，只计算行数与牛号多重集。"""
        result = {
            "path": str(path),
            "status": "pending",
            "sha256": "",
            "identity": _public_identity(_identity_state()),
            "error": "",
        }
        if not path.is_file():
            result["status"] = "missing"
            result["error"] = "缺少指数直接输入文件"
            return result

        workbook = None
        try:
            result["sha256"] = _sha256(path)
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = _deduplicate_headers(next(rows))
            except StopIteration:
                headers = []
            cow_column = _first_column(
                headers,
                ("cow_id", "母牛号", "牛号", "耳号"),
            )
            if cow_column is None:
                result["status"] = "invalid_schema"
                result["error"] = "指数直接输入缺少牛号列"
                return result
            cow_position = headers.index(cow_column)
            identity = _identity_state()
            for values in rows:
                row = tuple(values)
                if _is_blank_row(row):
                    continue
                value = (
                    row[cow_position]
                    if cow_position < len(row)
                    else None
                )
                _add_identity(identity, _identifier(value))
            result["status"] = "read"
            result["identity"] = _public_identity(identity)
            return result
        except Exception as exc:
            result["status"] = "failed"
            result["error"] = f"{type(exc).__name__}: {exc}"
            return result
        finally:
            if workbook is not None:
                workbook.close()

    def _source_checkpoint_identity(self, task: Dict) -> Dict:
        relative_path = str(
            task.get("relative_path")
            or task.get("child_path")
            or ""
        )
        child_path = (
            Path(task["child_path"])
            if task.get("child_path")
            else self.project_path / relative_path
        )
        source_path = (
            child_path
            / "analysis_results"
            / "processed_index_cow_index_scores.xlsx"
        )
        direct_path = (
            child_path
            / "analysis_results"
            / "processed_cow_data_key_traits_final.xlsx"
        )
        metadata_path = child_path / "project_metadata.json"
        task_id = str(task.get("task_id") or "").strip()
        api_farmcode, farm_number, farm_name = _task_farm_identity(task)
        source_key = task_id or hashlib.sha256(
            f"{task.get('farm_code', '')}\0{relative_path}".encode("utf-8")
        ).hexdigest()

        def file_state(path: Path) -> Dict:
            if not path.is_file():
                return {
                    "exists": False,
                    "bytes": 0,
                    "mtime_ns": 0,
                }
            stat = path.stat()
            return {
                "exists": True,
                "bytes": int(stat.st_size),
                "mtime_ns": int(stat.st_mtime_ns),
            }

        identity = {
            "detail_export_schema_version": DETAIL_EXPORT_SCHEMA_VERSION,
            "task_id": task_id,
            "farm_code": _identifier(task.get("farm_code")),
            "api_farmcode": api_farmcode,
            "farm_number": farm_number,
            "farm_name": farm_name,
            "relative_path": relative_path,
            "task_status": str(task.get("status") or ""),
            "requested_score_column": self.score_column or "",
            "source": file_state(source_path),
            "direct_input": file_state(direct_path),
            "child_metadata": file_state(metadata_path),
        }
        fingerprint = hashlib.sha256(
            json.dumps(
                identity,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()
        return {
            "source_key": source_key,
            "fingerprint": fingerprint,
        }

    def _reuse_source_checkpoint(
        self,
        connection: sqlite3.Connection,
        task: Dict,
        all_source_columns: List[str],
    ) -> Optional[Dict]:
        identity = self._source_checkpoint_identity(task)
        row = connection.execute(
            """
            SELECT input_fingerprint, source_result_json,
                   source_columns_json, row_count
            FROM source_checkpoints
            WHERE source_key = ?
            """,
            (identity["source_key"],),
        ).fetchone()
        if row is None or str(row[0]) != identity["fingerprint"]:
            return None
        actual_count = int(
            connection.execute(
                "SELECT COUNT(*) FROM records WHERE source_key = ?",
                (identity["source_key"],),
            ).fetchone()[0]
        )
        if actual_count != int(row[3]):
            return None
        try:
            source = json.loads(row[1])
            source_columns = json.loads(row[2])
        except (TypeError, json.JSONDecodeError):
            return None
        for column in source_columns:
            if column not in all_source_columns:
                all_source_columns.append(column)
        return source

    def _save_source_checkpoint(
        self,
        connection: sqlite3.Connection,
        task: Dict,
        source: Dict,
    ) -> None:
        identity = self._source_checkpoint_identity(task)
        source_columns = list(source.pop("_source_columns", []))
        public_source = dict(source)
        connection.execute(
            """
            INSERT OR REPLACE INTO source_checkpoints (
                source_key, input_fingerprint, source_result_json,
                source_columns_json, row_count, completed_at
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                identity["source_key"],
                identity["fingerprint"],
                json.dumps(
                    public_source,
                    ensure_ascii=False,
                    separators=(",", ":"),
                    allow_nan=False,
                ),
                json.dumps(
                    source_columns,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                int(source.get("rows_read", 0) or 0),
                dt.datetime.now().isoformat(timespec="seconds"),
            ),
        )
        connection.commit()

    def _ingest_source(
        self,
        connection: sqlite3.Connection,
        task: Dict,
        task_index: int,
        all_source_columns: List[str],
    ) -> Dict:
        relative_path = str(
            task.get("relative_path")
            or task.get("child_path")
            or ""
        )
        child_path = Path(task["child_path"]) if task.get("child_path") else self.project_path / relative_path
        source_path = (
            child_path
            / "analysis_results"
            / "processed_index_cow_index_scores.xlsx"
        )
        direct_input_path = (
            child_path
            / "analysis_results"
            / "processed_cow_data_key_traits_final.xlsx"
        )
        task_id = str(task.get("task_id") or "").strip()
        api_farmcode, farm_number, farm_name = _task_farm_identity(task)
        source_key = task_id or hashlib.sha256(
            f"{task.get('farm_code', '')}\0{relative_path}".encode("utf-8")
        ).hexdigest()
        source = {
            "source_key": source_key,
            "task_id": task_id,
            "farm_code": _identifier(task.get("farm_code")),
            "api_farmcode": api_farmcode,
            "farm_number": farm_number,
            "farm_name": farm_name,
            "child_relative_path": relative_path,
            "absolute_path": str(source_path.resolve()),
            "path": (
                Path(relative_path)
                / "analysis_results"
                / "processed_index_cow_index_scores.xlsx"
            ).as_posix(),
            "status": "pending",
            "rows_read": 0,
            "sha256": "",
            "score_column": "",
            "identity": _public_identity(_identity_state()),
            "direct_input": {},
            "identity_match": False,
            "duplicate_cow_id_count": 0,
            "lineage_column": "",
            "farm_number_column": "",
            "lineage_mismatch_rows": 0,
            "error": "",
        }
        # 未提交检查点的上次尝试可能留下部分行。单个来源重跑时只清理
        # 自己的记录，不影响已经完成并校验过的其他牧场。
        connection.execute(
            "DELETE FROM records WHERE source_key = ?",
            (source_key,),
        )
        connection.execute(
            "DELETE FROM source_checkpoints WHERE source_key = ?",
            (source_key,),
        )
        connection.commit()

        if task.get("status") not in (
            None,
            "",
            "completed",
            "completed_with_warning",
        ):
            source["status"] = "task_not_completed"
            source["error"] = f"子任务状态为 {task.get('status')}"
            return source
        if not source_path.is_file():
            source["status"] = "missing"
            source["error"] = "缺少指数结果文件"
            return source

        workbook = None
        inserted_rows = 0
        source_columns: List[str] = []
        actual_identity = _identity_state()
        try:
            source["sha256"] = _sha256(source_path)
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            try:
                raw_headers = next(iterator)
            except StopIteration:
                raw_headers = ()
            source_columns = _deduplicate_headers(raw_headers)
            positions = {
                header: column_index
                for column_index, header in enumerate(source_columns)
            }
            cow_column = _first_column(
                source_columns,
                ("cow_id", "母牛号", "牛号", "耳号"),
            )
            raw_cow_column = _first_column(
                source_columns,
                ("raw_cow_id", "原始牛号", "原牛号"),
            )
            in_herd_column = _first_column(
                source_columns,
                ("是否在场", "是否在群", "在群状态", "is_current"),
            )
            api_lineage_column = _first_column(
                source_columns,
                ("farm_code", "api farmcode", "farmcode"),
            )
            farm_number_column = _first_column(
                source_columns,
                ("牧场编号",),
            )
            lineage_column = api_lineage_column or farm_number_column
            source["lineage_column"] = api_lineage_column or ""
            source["farm_number_column"] = farm_number_column or ""
            index_candidates = [
                header
                for header in source_columns
                if str(header).strip().casefold().endswith("_index")
            ]
            selected_score_column = _detect_score_column(
                source_columns,
                self.score_column,
            )
            source["score_column"] = selected_score_column or ""
            schema_errors = []
            integrity_errors = []
            child_metadata_path = child_path / "project_metadata.json"
            try:
                child_metadata = json.loads(
                    child_metadata_path.read_text(encoding="utf-8")
                )
                expected_task_id = str(task.get("task_id") or "")
                actual_task_id = str(
                    child_metadata.get("group_task_id") or ""
                )
                if expected_task_id and actual_task_id != expected_task_id:
                    integrity_errors.append(
                        "子项目任务ID与父任务不一致"
                    )
                child_farms = child_metadata.get("farms", [])
                child_farm_code = _identifier(
                    child_farms[0].get("code")
                    if child_farms
                    else child_metadata.get("group_farm_code")
                )
                if child_farm_code != source["farm_code"]:
                    integrity_errors.append(
                        "子项目牧场编号与父任务不一致"
                    )
            except Exception as exc:
                integrity_errors.append(
                    f"子项目元数据不可验证：{type(exc).__name__}"
                )
            if cow_column is None:
                schema_errors.append("缺少牛号列")
            if in_herd_column is None:
                schema_errors.append("缺少是否在场列")
            if selected_score_column is None:
                schema_errors.append("缺少综合指数列")
            if self.score_column is None and len(index_candidates) > 1:
                schema_errors.append(
                    "存在多个指数列，未明确跨牧场统一口径："
                    + "、".join(index_candidates)
                )
            if lineage_column is None:
                schema_errors.append("缺少牧场编号来源列")

            insert_sql = """
                INSERT INTO records (
                    source_key, farm_code, farm_name, child_relative_path,
                    source_file, source_row, cow_id, raw_cow_id, in_herd_raw,
                    score_column, score_value, score_text, score_sign,
                    score_adjusted_exponent, score_digits, classification,
                    unranked_reason, has_long_text, payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                          ?, ?, ?, ?)
            """
            batch = []
            batch_bytes = 0
            bytes_since_disk_check = 0
            database_row = connection.execute(
                "PRAGMA database_list"
            ).fetchone()
            database_path = Path(database_row[2]) if database_row else source_path
            for excel_row, values in enumerate(iterator, start=2):
                row = tuple(values)
                if _is_blank_row(row):
                    continue
                if len(row) < len(source_columns):
                    row = row + (None,) * (len(source_columns) - len(row))
                (
                    classification,
                    reason,
                    cow_id,
                    raw_cow_id,
                    score_exact,
                ) = self._classify(
                    row,
                    positions,
                    cow_column,
                    raw_cow_column,
                    in_herd_column,
                    selected_score_column,
                )
                in_herd_raw = (
                    _identifier(row[positions[in_herd_column]])
                    if in_herd_column
                    else ""
                )
                _add_identity(actual_identity, cow_id)
                # “牧场编号”可能是业务侧七位编号，而 task.farm_code 是
                # 慧牧云 API farmcode，两者不能互相冒充。只有显式 API
                # lineage 列才与任务编码核对；两列均原样保留到交付明细。
                if api_lineage_column is not None:
                    lineage_value = _identifier(
                        row[positions[api_lineage_column]]
                    )
                    if lineage_value != source["farm_code"]:
                        source["lineage_mismatch_rows"] += 1
                payload = {}
                for column_index, header in enumerate(source_columns):
                    value = row[column_index] if column_index < len(row) else None
                    payload[header] = (
                        _identifier(value)
                        if _text_source_column(header)
                        else _json_safe(value)
                    )
                payload_json = json.dumps(
                    payload,
                    ensure_ascii=False,
                    separators=(",", ":"),
                    allow_nan=False,
                )
                has_long_text = any(
                    isinstance(value, str)
                    and len(value) > EXCEL_MAX_CELL_CHARACTERS
                    for value in (
                        source["farm_name"],
                        relative_path,
                        source["path"],
                        reason,
                        selected_score_column or "",
                        *payload.values(),
                    )
                )
                record = (
                    source_key,
                    source["farm_code"],
                    source["farm_name"],
                    relative_path,
                    source["path"],
                    excel_row,
                    cow_id,
                    raw_cow_id,
                    in_herd_raw,
                    selected_score_column or "",
                    _approximate_score(score_exact),
                    score_exact.text if score_exact is not None else "",
                    score_exact.sign if score_exact is not None else 0,
                    (
                        score_exact.adjusted_exponent
                        if score_exact is not None
                        else 0
                    ),
                    score_exact.digits if score_exact is not None else "0",
                    classification,
                    reason,
                    1 if has_long_text else 0,
                    payload_json,
                )
                batch.append(record)
                # 宽表不能只按 1000 行分批；极宽或超长字段会让单批内存
                # 突增。按“行数或序列化字节数先到者”提交。
                record_bytes = len(payload_json.encode("utf-8")) + 512
                batch_bytes += record_bytes
                bytes_since_disk_check += record_bytes
                if (
                    len(batch) >= DEFAULT_SQLITE_BATCH_ROWS
                    or batch_bytes >= DEFAULT_SQLITE_BATCH_BYTES
                ):
                    connection.executemany(insert_sql, batch)
                    inserted_rows += len(batch)
                    batch.clear()
                    batch_bytes = 0
                    connection.commit()
                    if (
                        bytes_since_disk_check
                        >= DEFAULT_DISK_CHECK_INTERVAL_BYTES
                    ):
                        self._ensure_free_space(
                            database_path.parent,
                            phase=f"读取牧场 {source['farm_name'] or source['farm_code']}",
                            transient_bytes=max(
                                DEFAULT_SQLITE_BATCH_BYTES * 4,
                                int(database_path.stat().st_size * 0.05)
                                if database_path.exists()
                                else 0,
                            ),
                            details={
                                "source_key": source_key,
                                "rows_committed_for_source": inserted_rows,
                            },
                        )
                        bytes_since_disk_check = 0
            if batch:
                connection.executemany(insert_sql, batch)
                inserted_rows += len(batch)
                connection.commit()
            for column in source_columns:
                if column not in all_source_columns:
                    all_source_columns.append(column)
            source["identity"] = _public_identity(actual_identity)
            duplicate_row = connection.execute(
                """
                SELECT COUNT(*)
                FROM (
                    SELECT cow_id
                    FROM records
                    WHERE source_key = ? AND cow_id <> ''
                    GROUP BY cow_id
                    HAVING COUNT(*) > 1
                )
                """,
                (source_key,),
            ).fetchone()
            source["duplicate_cow_id_count"] = int(
                duplicate_row[0] if duplicate_row else 0
            )
            if source["duplicate_cow_id_count"]:
                integrity_errors.append(
                    f"同一牧场存在 {source['duplicate_cow_id_count']} 个重复牛号"
                )
            if source["lineage_mismatch_rows"]:
                integrity_errors.append(
                    f"有 {source['lineage_mismatch_rows']} 行牧场编号与任务不一致"
                )

            direct_input = self._read_identity_source(direct_input_path)
            try:
                direct_input["path"] = direct_input_path.relative_to(
                    self.project_path
                ).as_posix()
            except ValueError:
                direct_input["path"] = str(direct_input_path)
            source["direct_input"] = direct_input
            source["identity_match"] = (
                direct_input.get("status") == "read"
                and direct_input.get("identity") == source["identity"]
            )
            if direct_input.get("status") != "read":
                integrity_errors.append(
                    direct_input.get("error") or "指数直接输入不可读取"
                )
            elif not source["identity_match"]:
                integrity_errors.append(
                    "指数结果与直接性状输入的行数/牛号多重集不一致"
                )

            if schema_errors:
                source["status"] = "invalid_schema"
            elif integrity_errors:
                source["status"] = "invalid_integrity"
            else:
                source["status"] = "read"
            source["rows_read"] = inserted_rows
            source["error"] = "；".join(schema_errors + integrity_errors)
            source["_source_columns"] = source_columns
            return source
        except GroupDetailExportPaused:
            # 当前牧场尚未形成完整来源检查点，只撤销它自己的部分记录；
            # 其他已提交牧场和已校验分卷全部保留。
            connection.execute(
                "DELETE FROM records WHERE source_key = ?",
                (source_key,),
            )
            connection.commit()
            raise
        except Exception as exc:
            # 一个源文件失败时清除该文件已写入的行，避免把半个牧场伪装为
            # 完整结果；其他牧场仍可继续，最终 manifest 明确标记 partial。
            connection.execute(
                "DELETE FROM records WHERE source_key = ?",
                (source_key,),
            )
            connection.commit()
            source["status"] = "failed"
            source["rows_read"] = 0
            source["error"] = f"{type(exc).__name__}: {exc}"
            source["_source_columns"] = source_columns
            return source
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def _build_ranks(connection: sqlite3.Connection) -> None:
        connection.execute("DROP TABLE IF EXISTS ranks")
        connection.execute("DROP INDEX IF EXISTS records_reconciliation_order")
        connection.execute(
            """
            CREATE TABLE ranks (
                record_id INTEGER PRIMARY KEY,
                global_rank INTEGER NOT NULL UNIQUE
            )
            """
        )
        exact_order = sqlite_order_by_clause(
            sign_column="score_sign",
            exponent_column="score_adjusted_exponent",
            digits_column="score_digits",
            descending=True,
        )
        connection.execute(
            f"""
            INSERT INTO ranks (record_id, global_rank)
            SELECT
                id,
                ROW_NUMBER() OVER (
                    ORDER BY
                        {exact_order},
                        farm_code COLLATE BINARY ASC,
                        cow_id COLLATE BINARY ASC,
                        child_relative_path COLLATE BINARY ASC,
                        source_row ASC,
                        id ASC
                )
            FROM records
            WHERE classification = ?
            """,
            (RANKED_CLASSIFICATION,),
        )
        connection.execute(
            "CREATE INDEX records_reconciliation_order ON records(id)"
        )
        connection.commit()

    @staticmethod
    def _query_rows(
        connection: sqlite3.Connection,
        kind: str,
    ) -> Iterator[Tuple]:
        if kind == "ranked":
            cursor = connection.execute(
                """
                SELECT
                    ranks.global_rank, records.classification,
                    records.unranked_reason, records.farm_code,
                    records.farm_name, records.child_relative_path,
                    records.source_file, records.source_row,
                    records.score_column, records.score_value,
                    records.score_text,
                    records.payload_json
                FROM ranks
                JOIN records ON records.id = ranks.record_id
                ORDER BY ranks.global_rank ASC
                """
            )
        else:
            cursor = connection.execute(
                """
                SELECT
                    ranks.global_rank, records.classification,
                    records.unranked_reason, records.farm_code,
                    records.farm_name, records.child_relative_path,
                    records.source_file, records.source_row,
                    records.score_column, records.score_value,
                    records.score_text,
                    records.payload_json
                FROM records
                LEFT JOIN ranks ON ranks.record_id = records.id
                ORDER BY records.id ASC
                """
            )
        while True:
            rows = cursor.fetchmany(1000)
            if not rows:
                break
            yield from rows

    def _export_kind(
        self,
        connection: sqlite3.Connection,
        package_dir: Path,
        *,
        kind: str,
        source_columns: Sequence[str],
        checkpoint: Dict,
        checkpoint_path: Path,
    ) -> List[Dict]:
        fixed_count = len(_FIXED_HEADERS)
        # 不把 Excel 的 16,384 列上限当作实际分片目标。极宽单卷会造成
        # 巨型 XML 临时文件和打开卡顿；多分几个字段卷仍完整保留全部列。
        max_source_columns = min(
            EXCEL_MAX_COLUMNS - fixed_count,
            DEFAULT_MAX_SOURCE_COLUMNS_PER_PART,
        )
        column_parts = [
            list(source_columns[index : index + max_source_columns])
            for index in range(0, len(source_columns), max_source_columns)
        ] or [[]]
        stem = "有效在群完整排名" if kind == "ranked" else "全部源行分类对账"
        sheet_name = "完整排名" if kind == "ranked" else "分类对账"
        volume_entries: List[Dict] = []
        checkpoint.setdefault("volumes", [])
        checkpoint.setdefault("completed_parts", [])

        for part_number, part_columns in enumerate(column_parts, start=1):
            used_headers = set(_FIXED_HEADERS)
            displayed_part_columns = []
            for source_column in part_columns:
                displayed = source_column
                if displayed in used_headers:
                    displayed = f"源字段_{displayed}"
                suffix = 2
                base = displayed
                while displayed in used_headers:
                    displayed = f"{base}__{suffix}"
                    suffix += 1
                used_headers.add(displayed)
                displayed_part_columns.append(displayed)
            headers = list(_FIXED_HEADERS) + displayed_part_columns
            rows_per_volume = self.rows_per_volume or min(
                DEFAULT_MAX_ROWS_PER_VOLUME,
                max(
                    1_000,
                    DEFAULT_MAX_CELLS_PER_VOLUME // max(len(headers), 1),
                ),
            )
            text_indexes = {3, 10}
            text_indexes.update(
                fixed_count + index
                for index, column in enumerate(part_columns)
                if _text_source_column(column)
            )
            writer: Optional[_AtomicXlsxVolume] = None
            part_key = f"{kind}:{part_number}"
            expected = int(
                connection.execute(
                    "SELECT COUNT(*) FROM ranks"
                    if kind == "ranked"
                    else "SELECT COUNT(*) FROM records"
                ).fetchone()[0]
            )
            existing_entries = sorted(
                (
                    dict(entry)
                    for entry in checkpoint["volumes"]
                    if entry.get("kind") == kind
                    and int(entry.get("column_part", 0)) == part_number
                ),
                key=lambda entry: int(entry.get("volume", 0)),
            )
            valid_entries = []
            for expected_volume, entry in enumerate(
                existing_entries,
                start=1,
            ):
                path = package_dir / str(entry.get("path") or "")
                try:
                    valid = (
                        int(entry.get("volume", 0)) == expected_volume
                        and int(entry.get("rows_per_volume", 0))
                        == rows_per_volume
                        and int(entry.get("column_parts", 0))
                        == len(column_parts)
                        and path.is_file()
                        and path.stat().st_size == int(entry.get("bytes", -1))
                        and _sha256(path) == entry.get("sha256")
                    )
                except OSError:
                    valid = False
                if not valid:
                    break
                valid_entries.append(entry)

            valid_names = {entry["path"] for entry in valid_entries}
            for entry in existing_entries:
                if entry.get("path") not in valid_names:
                    (package_dir / str(entry.get("path") or "")).unlink(
                        missing_ok=True
                    )
            checkpoint["volumes"] = [
                entry
                for entry in checkpoint["volumes"]
                if not (
                    entry.get("kind") == kind
                    and int(entry.get("column_part", 0)) == part_number
                )
            ] + valid_entries

            exported_in_part = sum(
                int(entry.get("data_rows", 0)) for entry in valid_entries
            )
            if exported_in_part > expected:
                for entry in valid_entries:
                    (package_dir / entry["path"]).unlink(missing_ok=True)
                checkpoint["volumes"] = [
                    entry
                    for entry in checkpoint["volumes"]
                    if not (
                        entry.get("kind") == kind
                        and int(entry.get("column_part", 0)) == part_number
                    )
                ]
                valid_entries = []
                exported_in_part = 0
            volume_entries.extend(valid_entries)
            volume_number = len(valid_entries)

            if (
                part_key in checkpoint["completed_parts"]
                and exported_in_part == expected
            ):
                continue
            checkpoint["completed_parts"] = [
                key
                for key in checkpoint["completed_parts"]
                if key != part_key
            ]

            def open_writer() -> _AtomicXlsxVolume:
                nonlocal volume_number
                volume_number += 1
                part_suffix = (
                    f"_字段第{part_number:02d}部分"
                    if len(column_parts) > 1
                    else ""
                )
                filename = (
                    f"{stem}_第{volume_number:04d}卷{part_suffix}.xlsx"
                )
                xml_budget = self._volume_xml_budget(
                    package_dir,
                    phase=f"创建{stem}分卷",
                )
                return _AtomicXlsxVolume(
                    package_dir / filename,
                    sheet_name,
                    headers,
                    text_indexes,
                    max_estimated_xml_bytes=xml_budget,
                )

            def save_volume(writer_to_close: _AtomicXlsxVolume) -> Dict:
                # 关闭工作簿时需在现有 sheet XML 旁再生成 ZIP，提前确认
                # 有足够同盘空间；通过后才将该卷原子登记为检查点。
                self._ensure_free_space(
                    package_dir,
                    phase=f"封装{stem}第{volume_number:04d}卷",
                    transient_bytes=max(
                        64 * 1024 * 1024,
                        writer_to_close.estimated_xml_bytes,
                    ),
                )
                stats = writer_to_close.close()
                entry = {
                    "kind": kind,
                    "volume": volume_number,
                    "column_part": part_number,
                    "column_parts": len(column_parts),
                    "rows_per_volume": rows_per_volume,
                    "path": writer_to_close.output_path.name,
                    **stats,
                }
                volume_entries.append(entry)
                checkpoint["volumes"].append(entry)
                _write_json_atomic(checkpoint_path, checkpoint)
                return entry

            try:
                if exported_in_part < expected or (
                    expected == 0 and not valid_entries
                ):
                    writer = open_writer()
                for row_number, raw in enumerate(
                    self._query_rows(connection, kind)
                ):
                    if row_number < exported_in_part:
                        continue
                    assert writer is not None
                    payload = json.loads(raw[11])
                    values = list(raw[:11]) + [
                        payload.get(column)
                        for column in part_columns
                    ]
                    rank = int(raw[0]) if raw[0] is not None else None
                    if (
                        writer.data_rows >= rows_per_volume
                        or not writer.can_fit(values)
                    ):
                        save_volume(writer)
                        writer = open_writer()
                    writer.write(values, rank)
                    exported_in_part += 1
                if writer is not None:
                    save_volume(writer)
                    writer = None
            except Exception:
                if writer is not None:
                    writer.abort()
                raise

            if exported_in_part != expected:
                raise RuntimeError(
                    f"{stem}第 {part_number} 个字段分片应导出 "
                    f"{expected:,} 行，实际为 {exported_in_part:,} 行"
                )
            checkpoint["completed_parts"].append(part_key)
            _write_json_atomic(checkpoint_path, checkpoint)
        return volume_entries

    def _export_long_fields(
        self,
        connection: sqlite3.Connection,
        package_dir: Path,
        *,
        checkpoint: Dict,
        checkpoint_path: Path,
    ) -> Tuple[List[Dict], Dict[str, int]]:
        """把超过 Excel 单元格上限的文本无损拆到独立分卷。"""
        headers = [
            "API farmcode",
            "牧场名称",
            "子项目相对目录",
            "源文件",
            "源数据行号",
            "字段名",
            "原始字符数",
            "原始SHA-256",
            "分块序号",
            "分块总数",
            "完整内容分块",
        ]
        rows_per_volume = self.rows_per_volume or DEFAULT_MAX_ROWS_PER_VOLUME
        text_indexes = {0, 1, 2, 3, 5, 7, 10}
        checkpoint.setdefault("volumes", [])
        checkpoint.setdefault("completed_parts", [])
        part_key = "long_fields:1"
        existing_entries = sorted(
            (
                dict(entry)
                for entry in checkpoint["volumes"]
                if entry.get("kind") == "long_fields"
            ),
            key=lambda entry: int(entry.get("volume", 0)),
        )
        volume_entries: List[Dict] = []
        for expected_volume, entry in enumerate(existing_entries, start=1):
            path = package_dir / str(entry.get("path") or "")
            try:
                valid = (
                    int(entry.get("volume", 0)) == expected_volume
                    and path.is_file()
                    and path.stat().st_size == int(entry.get("bytes", -1))
                    and _sha256(path) == entry.get("sha256")
                )
            except OSError:
                valid = False
            if not valid:
                break
            volume_entries.append(entry)
        valid_names = {entry["path"] for entry in volume_entries}
        for entry in existing_entries:
            if entry.get("path") not in valid_names:
                (package_dir / str(entry.get("path") or "")).unlink(
                    missing_ok=True
                )
        checkpoint["volumes"] = [
            entry
            for entry in checkpoint["volumes"]
            if entry.get("kind") != "long_fields"
        ] + volume_entries
        exported_chunks = sum(
            int(entry.get("data_rows", 0)) for entry in volume_entries
        )
        writer: Optional[_AtomicXlsxVolume] = None
        volume_number = len(volume_entries)
        field_count = 0
        chunk_count = 0

        def open_writer() -> _AtomicXlsxVolume:
            nonlocal volume_number
            volume_number += 1
            xml_budget = self._volume_xml_budget(
                package_dir,
                phase="创建超长字段分卷",
            )
            return _AtomicXlsxVolume(
                package_dir
                / f"超长字段完整内容_第{volume_number:04d}卷.xlsx",
                "超长字段完整内容",
                headers,
                text_indexes,
                max_estimated_xml_bytes=xml_budget,
            )

        def close_writer(writer_to_close: _AtomicXlsxVolume) -> None:
            self._ensure_free_space(
                package_dir,
                phase=f"封装超长字段第{volume_number:04d}卷",
                transient_bytes=max(
                    64 * 1024 * 1024,
                    writer_to_close.estimated_xml_bytes,
                ),
            )
            stats = writer_to_close.close()
            entry = {
                "kind": "long_fields",
                "volume": volume_number,
                "column_part": 1,
                "column_parts": 1,
                "rows_per_volume": rows_per_volume,
                "path": writer_to_close.output_path.name,
                **stats,
            }
            volume_entries.append(entry)
            checkpoint["volumes"].append(entry)
            _write_json_atomic(checkpoint_path, checkpoint)

        cursor = connection.execute(
            """
            SELECT farm_code, farm_name, child_relative_path, source_file,
                   source_row, unranked_reason, score_column, payload_json
            FROM records
            WHERE has_long_text = 1
            ORDER BY id
            """
        )
        try:
            while True:
                rows = cursor.fetchmany(1000)
                if not rows:
                    break
                for (
                    farm_code,
                    farm_name,
                    child_relative_path,
                    source_file,
                    source_row,
                    unranked_reason,
                    score_column,
                    payload_json,
                ) in rows:
                    values = {
                        "审计字段:牧场名称": farm_name,
                        "审计字段:子项目相对目录": child_relative_path,
                        "审计字段:源文件": source_file,
                        "审计字段:未排名原因": unranked_reason,
                        "审计字段:指数列": score_column,
                    }
                    payload = json.loads(payload_json)
                    values.update(
                        {
                            f"源字段:{field_name}": value
                            for field_name, value in payload.items()
                        }
                    )
                    for field_name, value in values.items():
                        if not isinstance(value, str) or len(
                            value
                        ) <= EXCEL_MAX_CELL_CHARACTERS:
                            continue
                        field_count += 1
                        digest = hashlib.sha256(
                            value.encode("utf-8")
                        ).hexdigest()
                        total_chunks = math.ceil(
                            len(value) / LONG_TEXT_CHUNK_CHARACTERS
                        )
                        for chunk_number in range(1, total_chunks + 1):
                            chunk_count += 1
                            if chunk_count <= exported_chunks:
                                continue
                            if writer is None:
                                writer = open_writer()
                            start = (
                                chunk_number - 1
                            ) * LONG_TEXT_CHUNK_CHARACTERS
                            content = value[
                                start : start + LONG_TEXT_CHUNK_CHARACTERS
                            ]
                            output_values = [
                                farm_code,
                                farm_name,
                                child_relative_path,
                                source_file,
                                int(source_row),
                                field_name,
                                len(value),
                                digest,
                                chunk_number,
                                total_chunks,
                                content,
                            ]
                            if (
                                writer.data_rows >= rows_per_volume
                                or not writer.can_fit(output_values)
                            ):
                                close_writer(writer)
                                writer = open_writer()
                            writer.write(output_values, None)
        except Exception:
            if writer is not None:
                writer.abort()
            raise
        if writer is not None:
            close_writer(writer)
        if exported_chunks > chunk_count:
            raise RuntimeError(
                "超长字段检查点行数超过当前来源内容，需重新生成该部分"
            )
        if part_key not in checkpoint["completed_parts"]:
            checkpoint["completed_parts"].append(part_key)
        _write_json_atomic(checkpoint_path, checkpoint)
        return volume_entries, {
            "long_field_count": field_count,
            "long_field_chunk_count": chunk_count,
        }

    @staticmethod
    def _next_package_path(
        output_dir: Path,
        requested_name: Optional[str],
    ) -> Path:
        base = requested_name or (
            "牧场组牛只完整排名明细_"
            + dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        )
        candidate = output_dir / base
        suffix = 2
        while candidate.exists():
            candidate = output_dir / f"{base}_{suffix}"
            suffix += 1
        return candidate

    def export(
        self,
        *,
        tasks: Optional[Sequence[Dict]] = None,
        output_dir: Optional[Path] = None,
        package_name: Optional[str] = None,
    ) -> Dict:
        """生成完整排名包并返回可审计 manifest。

        ``status`` 为 ``complete`` 时，所有纳入范围的源文件均已读取且：

        * 排名卷导出行数 = 有效在群排名行数；
        * 对账卷导出行数 = 所有源数据行数；
        * 每个输出卷都带 SHA-256。

        个别子项目缺文件、尚未完成或文件损坏时不会吞掉异常或冒充完整。
        其他可读取牧场仍会导出，manifest 的 ``status`` 为 ``partial``，
        并在 ``sources`` 中保留具体原因。
        """
        task_list = self._load_tasks(tasks)
        output_dir = Path(output_dir or (self.project_path / "reports"))
        output_dir.mkdir(parents=True, exist_ok=True)
        disk_preflight = self._disk_preflight(task_list, output_dir)
        final_package = self._next_package_path(output_dir, package_name)
        staging_package = output_dir / f".{final_package.name}.resume"
        staging_package.mkdir(parents=False, exist_ok=True)
        pause_state_path = staging_package / "pause_state.json"
        pause_state_path.unlink(missing_ok=True)
        database_dir = staging_package / ".work"
        database_dir.mkdir(parents=True, exist_ok=True)
        database_path = database_dir / "ranking.sqlite3"
        connection: Optional[sqlite3.Connection] = None

        try:
            _progress(self.progress_callback, 1, "正在建立全量排名磁盘索引...")
            connection = self._create_database(database_path)
            all_source_columns: List[str] = []
            sources = []
            task_total = max(1, len(task_list))
            for task_index, task in enumerate(task_list, start=1):
                source = self._reuse_source_checkpoint(
                    connection,
                    task,
                    all_source_columns,
                )
                resumed = source is not None
                if source is None:
                    source = self._ingest_source(
                        connection,
                        task,
                        task_index,
                        all_source_columns,
                    )
                    self._save_source_checkpoint(
                        connection,
                        task,
                        source,
                    )
                source["resumed_from_checkpoint"] = resumed
                sources.append(source)
                _progress(
                    self.progress_callback,
                    5 + int(task_index / task_total * 40),
                    f"已读取 {task_index}/{len(task_list)} 个牧场："
                    f"{source.get('farm_name') or source.get('farm_code')}",
                )

            def add_source_integrity_error(source: Dict, message: str) -> None:
                if source.get("status") == "read":
                    source["status"] = "invalid_integrity"
                if source.get("status") in {
                    "invalid_schema",
                    "invalid_integrity",
                }:
                    existing = str(source.get("error") or "")
                    source["error"] = (
                        f"{existing}；{message}" if existing else message
                    )

            score_columns = {
                str(source.get("score_column") or "")
                for source in sources
                if source.get("rows_read", 0)
                and source.get("score_column")
            }
            if len(score_columns) > 1:
                detail = "、".join(sorted(score_columns))
                for source in sources:
                    if source.get("score_column"):
                        add_source_integrity_error(
                            source,
                            f"跨牧场指数列口径不一致：{detail}",
                        )

            source_paths: Dict[str, List[Dict]] = {}
            source_keys: Dict[str, List[Dict]] = {}
            for source in sources:
                source_key = str(source.get("source_key") or "")
                source_keys.setdefault(source_key, []).append(source)
                absolute_source = source.get("absolute_path")
                if not absolute_source:
                    continue
                canonical = str(Path(absolute_source).resolve())
                source_paths.setdefault(canonical, []).append(source)
            for duplicates in source_paths.values():
                if len(duplicates) <= 1:
                    continue
                for source in duplicates:
                    add_source_integrity_error(
                        source,
                        "同一指数源文件被多个牧场任务重复引用",
                    )
            for duplicates in source_keys.values():
                if len(duplicates) <= 1:
                    continue
                for source in duplicates:
                    add_source_integrity_error(
                        source,
                        "多个牧场任务使用了相同来源键/任务ID",
                    )
            for source in sources:
                source.pop("absolute_path", None)

            export_fingerprint_payload = {
                "schema_version": DETAIL_EXPORT_SCHEMA_VERSION,
                "requested_score_column": self.score_column or "",
                "rows_per_volume": self.rows_per_volume,
                "source_columns": all_source_columns,
                "sources": [
                    {
                        key: source.get(key)
                        for key in (
                            "source_key",
                            "task_id",
                            "farm_code",
                            "child_relative_path",
                            "sha256",
                            "score_column",
                            "identity",
                            "direct_input",
                            "status",
                            "rows_read",
                            "error",
                        )
                    }
                    for source in sources
                ],
            }
            export_fingerprint = hashlib.sha256(
                json.dumps(
                    export_fingerprint_payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                ).encode("utf-8")
            ).hexdigest()
            checkpoint_path = database_dir / "export_checkpoint.json"
            checkpoint = {}
            if checkpoint_path.is_file():
                try:
                    checkpoint = json.loads(
                        checkpoint_path.read_text(encoding="utf-8")
                    )
                except (OSError, json.JSONDecodeError):
                    checkpoint = {}
            if checkpoint.get("run_fingerprint") != export_fingerprint:
                # 仅删除未发布的本批次分卷；单牧场结果和正式历史报告不动。
                for path in staging_package.glob("*.xlsx"):
                    path.unlink(missing_ok=True)
                checkpoint = {
                    "schema_version": DETAIL_EXPORT_SCHEMA_VERSION,
                    "run_fingerprint": export_fingerprint,
                    "volumes": [],
                    "completed_parts": [],
                }
                _write_json_atomic(checkpoint_path, checkpoint)

            _progress(self.progress_callback, 48, "正在按未舍入指数建立稳定全局排名...")
            connection.execute("PRAGMA wal_checkpoint(PASSIVE)")
            database_bytes = self._database_footprint(database_path)
            self._ensure_free_space(
                database_dir,
                phase="建立全局精确排名",
                transient_bytes=max(
                    512 * 1024 * 1024,
                    database_bytes,
                ),
                details={"database_footprint_bytes": database_bytes},
            )
            self._build_ranks(connection)
            source_rows = int(
                connection.execute("SELECT COUNT(*) FROM records").fetchone()[0]
            )
            ranked_rows = int(
                connection.execute("SELECT COUNT(*) FROM ranks").fetchone()[0]
            )
            reason_counts = {
                str(reason): int(count)
                for reason, count in connection.execute(
                    """
                    SELECT unranked_reason, COUNT(*)
                    FROM records
                    WHERE classification = ?
                    GROUP BY unranked_reason
                    ORDER BY COUNT(*) DESC, unranked_reason ASC
                    """,
                    (UNRANKED_CLASSIFICATION,),
                )
            }
            farm_aggregates = {
                str(source_key): {
                    "source_rows": int(source_rows_for_farm),
                    "valid_ranked_rows": int(valid_count),
                    "unranked_rows": int(
                        source_rows_for_farm - valid_count
                    ),
                    "average_index": (
                        float(average_index)
                        if average_index is not None
                        else None
                    ),
                }
                for (
                    source_key,
                    source_rows_for_farm,
                    valid_count,
                    average_index,
                ) in connection.execute(
                    """
                    SELECT
                        source_key,
                        COUNT(*),
                        SUM(CASE WHEN classification = ? THEN 1 ELSE 0 END),
                        AVG(CASE WHEN classification = ? THEN score_value END)
                    FROM records
                    GROUP BY source_key
                    """,
                    (RANKED_CLASSIFICATION, RANKED_CLASSIFICATION),
                )
            }
            farm_stats = []
            for source in sources:
                aggregate = farm_aggregates.get(
                    str(source.get("source_key") or ""),
                    {
                        "source_rows": 0,
                        "valid_ranked_rows": 0,
                        "unranked_rows": 0,
                        "average_index": None,
                    },
                )
                source_rows_for_farm = int(
                    aggregate["source_rows"]
                )
                valid_count = int(aggregate["valid_ranked_rows"])
                farm_stats.append(
                    {
                        "task_id": source.get("task_id", ""),
                        "farm_code": source.get("farm_code", ""),
                        "api_farmcode": source.get(
                            "api_farmcode", source.get("farm_code", "")
                        ),
                        "farm_number": source.get("farm_number", ""),
                        "farm_name": source.get("farm_name", ""),
                        "child_relative_path": source.get(
                            "child_relative_path", ""
                        ),
                        **aggregate,
                        "valid_coverage": (
                            valid_count / source_rows_for_farm
                            if source_rows_for_farm
                            else 0
                        ),
                        "duplicate_cow_id_count": int(
                            source.get("duplicate_cow_id_count", 0)
                        ),
                        "source_status": source.get("status", ""),
                        "source_error": source.get("error", ""),
                    }
                )

            source_identity_by_relative = {
                str(source.get("child_relative_path") or ""): source
                for source in sources
            }

            def preview_rows(descending: bool, limit: int = 100) -> List[Dict]:
                direction = "DESC" if descending else "ASC"
                rows = connection.execute(
                    f"""
                    SELECT ranks.global_rank, records.farm_code,
                           records.farm_name, records.cow_id,
                           records.raw_cow_id, records.score_value,
                           records.score_text,
                           records.child_relative_path
                    FROM ranks
                    JOIN records ON records.id = ranks.record_id
                    ORDER BY ranks.global_rank {direction}
                    LIMIT ?
                    """,
                    (limit,),
                )
                preview_result = []
                for (
                    rank,
                    farm_code,
                    farm_name,
                    cow_id,
                    raw_cow_id,
                    index_score,
                    index_score_exact,
                    child_relative_path,
                ) in rows:
                    source_identity = source_identity_by_relative.get(
                        str(child_relative_path),
                        {},
                    )
                    preview_result.append({
                        "global_rank": int(rank),
                        "farm_code": str(farm_code),
                        "api_farmcode": str(
                            source_identity.get(
                                "api_farmcode", farm_code
                            )
                        ),
                        "farm_number": str(
                            source_identity.get("farm_number", "")
                        ),
                        "farm_name": str(farm_name),
                        "cow_id": str(cow_id),
                        "raw_cow_id": str(raw_cow_id),
                        "index_score": (
                            float(index_score)
                            if index_score is not None
                            else None
                        ),
                        "index_score_exact": str(index_score_exact),
                        "child_relative_path": str(child_relative_path),
                    })
                return preview_result

            preview = {
                "top": preview_rows(False),
                "bottom": list(reversed(preview_rows(True))),
            }

            _progress(self.progress_callback, 55, "正在分卷写入全部有效在群排名...")
            ranked_volumes = self._export_kind(
                connection,
                staging_package,
                kind="ranked",
                source_columns=all_source_columns,
                checkpoint=checkpoint,
                checkpoint_path=checkpoint_path,
            )
            _progress(self.progress_callback, 75, "正在分卷写入全部源行分类对账...")
            reconciliation_volumes = self._export_kind(
                connection,
                staging_package,
                kind="reconciliation",
                source_columns=all_source_columns,
                checkpoint=checkpoint,
                checkpoint_path=checkpoint_path,
            )
            _progress(
                self.progress_callback,
                90,
                "正在无损拆分超过Excel单元格上限的文本...",
            )
            long_field_volumes, long_field_counts = (
                self._export_long_fields(
                    connection,
                    staging_package,
                    checkpoint=checkpoint,
                    checkpoint_path=checkpoint_path,
                )
            )

            # 字段分片时每个字段部分都会覆盖全部行；按第一字段分片核对
            # 行数，避免把列分片重复计入。
            ranked_exported = sum(
                volume["data_rows"]
                for volume in ranked_volumes
                if volume["column_part"] == 1
            )
            reconciliation_exported = sum(
                volume["data_rows"]
                for volume in reconciliation_volumes
                if volume["column_part"] == 1
            )
            if ranked_exported != ranked_rows:
                raise RuntimeError(
                    f"完整排名行数核对失败：应为 {ranked_rows:,}，"
                    f"实际导出 {ranked_exported:,}"
                )
            if reconciliation_exported != source_rows:
                raise RuntimeError(
                    f"源行对账行数核对失败：应为 {source_rows:,}，"
                    f"实际导出 {reconciliation_exported:,}"
                )

            source_problem_count = sum(
                source["status"] != "read" for source in sources
            )
            manifest = {
                "schema_version": 1,
                "status": "partial" if source_problem_count else "complete",
                "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
                "package_name": final_package.name,
                "ranking_rule": (
                    "综合指数未舍入值降序；同分时依次按API farmcode、牛号、"
                    "子项目相对目录、源数据行号、磁盘记录号升序"
                ),
                "rows_per_volume": (
                    self.rows_per_volume or "按列数动态计算"
                ),
                "score_column_requested": self.score_column or "",
                "disk_preflight": disk_preflight,
                "counts": {
                    "tasks_in_scope": len(task_list),
                    "source_files_read": sum(
                        source["status"] == "read" for source in sources
                    ),
                    "source_files_with_problem": source_problem_count,
                    "source_rows": source_rows,
                    "valid_ranked_rows": ranked_rows,
                    "unranked_rows": source_rows - ranked_rows,
                    "ranked_exported_rows": ranked_exported,
                    "reconciliation_exported_rows": reconciliation_exported,
                    "unranked_reason_counts": reason_counts,
                    **long_field_counts,
                },
                "source_columns": all_source_columns,
                "sources": sources,
                "farm_stats": farm_stats,
                "preview": preview,
                "volumes": {
                    "ranked": ranked_volumes,
                    "reconciliation": reconciliation_volumes,
                    "long_fields": long_field_volumes,
                },
            }
            manifest_path = staging_package / "manifest.json"
            _write_json_atomic(manifest_path, manifest)

            _progress(self.progress_callback, 96, "正在核对文件校验值并完成原子落盘...")
            quick_check = connection.execute("PRAGMA quick_check").fetchone()
            if not quick_check or str(quick_check[0]).casefold() != "ok":
                raise RuntimeError("全量排名磁盘索引完整性校验失败，未发布")
            connection.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            connection.close()
            connection = None
            shutil.rmtree(database_dir, ignore_errors=True)
            os.replace(staging_package, final_package)

            returned = dict(manifest)
            returned["package_path"] = str(final_package)
            returned["manifest_path"] = str(final_package / "manifest.json")
            returned["manifest_sha256"] = _sha256(final_package / "manifest.json")
            for kind in ("ranked", "reconciliation", "long_fields"):
                for volume in returned["volumes"][kind]:
                    volume["absolute_path"] = str(final_package / volume["path"])
            _progress(self.progress_callback, 100, "牧场组全量排名明细导出完成")
            return returned
        except GroupDetailExportPaused as exc:
            if connection is not None:
                try:
                    connection.commit()
                    connection.execute("PRAGMA wal_checkpoint(PASSIVE)")
                except sqlite3.Error:
                    pass
                connection.close()
                connection = None
            _write_json_atomic(
                pause_state_path,
                {
                    "schema_version": 1,
                    "status": "paused",
                    "phase": exc.phase,
                    "message": str(exc),
                    "details": exc.details,
                    "paused_at": dt.datetime.now().isoformat(
                        timespec="seconds"
                    ),
                    "resume_directory": staging_package.name,
                },
            )
            raise
        except Exception:
            if connection is not None:
                connection.close()
            # 保留磁盘索引、已提交来源和已校验分卷，下一次使用相同
            # package_name 时从检查点继续。临时 .tmp.xlsx 会由卷写入器清理。
            raise


def export_group_cow_ranking_details(
    project_path: Path,
    *,
    tasks: Optional[Sequence[Dict]] = None,
    output_dir: Optional[Path] = None,
    package_name: Optional[str] = None,
    rows_per_volume: Optional[int] = DEFAULT_ROWS_PER_VOLUME,
    score_column: Optional[str] = None,
    progress_callback: Optional[Callable[[int, str], None]] = None,
) -> Dict:
    """函数式便捷入口；返回内容与 :meth:`export` 相同。"""
    return GroupCowRankingDetailExporter(
        project_path,
        rows_per_volume=rows_per_volume,
        score_column=score_column,
        progress_callback=progress_callback,
    ).export(
        tasks=tasks,
        output_dir=output_dir,
        package_name=package_name,
    )

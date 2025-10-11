"""
测试Sheet 6生成
"""

from pathlib import Path
import sys

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.excel_report.data_collectors import collect_breeding_inbreeding_data
from core.excel_report.sheet_builders import Sheet6Builder
from core.excel_report.formatters import StyleManager, ChartBuilder
from openpyxl import Workbook
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

logger = logging.getLogger(__name__)


def test_sheet6_data_collection():
    """测试Sheet6数据收集"""
    logger.info("=" * 60)
    logger.info("测试Sheet6数据收集")
    logger.info("=" * 60)

    # 使用测试项目路径
    test_project = Path("/Users/bozhenwang/GeneticImprove/去青青_2025_10_08_15_59")
    analysis_folder = test_project / "analysis_results"

    # 收集数据
    data = collect_breeding_inbreeding_data(analysis_folder)

    # 检查数据结构
    logger.info(f"\n数据键: {list(data.keys())}")

    if 'all_years_distribution' in data:
        dist = data['all_years_distribution']
        logger.info(f"\n全部年份分布:")
        logger.info(f"  总配次: {dist['total']}")
        logger.info(f"  区间数量: {dist['counts']}")
        logger.info(f"  区间占比: {[f'{r*100:.1f}%' for r in dist['ratios']]}")

    if 'recent_12m_distribution' in data:
        dist = data['recent_12m_distribution']
        logger.info(f"\n近12个月分布:")
        logger.info(f"  总配次: {dist['total']}")

    if 'yearly_trend' in data:
        logger.info(f"\n年份趋势: {len(data['yearly_trend'])}年")

    logger.info("\n✓ 数据收集测试通过")
    return data


def test_sheet6_generation(data):
    """测试Sheet6生成"""
    logger.info("\n" + "=" * 60)
    logger.info("测试Sheet6表格生成")
    logger.info("=" * 60)

    # 创建workbook
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # 创建样式管理器和图表构建器
    style_manager = StyleManager()
    chart_builder = ChartBuilder()

    # 创建Sheet6Builder
    builder = Sheet6Builder(wb, style_manager, chart_builder)

    # 生成Sheet
    builder.build(data)

    # 保存测试文件
    output_path = Path("/tmp/test_sheet6.xlsx")
    wb.save(output_path)

    logger.info(f"\n✓ Sheet6生成成功")
    logger.info(f"✓ 测试文件已保存: {output_path}")

    # 验证sheet是否存在
    if "配种记录-近交系数分析" in wb.sheetnames:
        ws = wb["配种记录-近交系数分析"]
        logger.info(f"✓ Sheet名称正确: {ws.title}")
        logger.info(f"✓ 数据行数: {ws.max_row}")
        logger.info(f"✓ 数据列数: {ws.max_column}")

        # 检查标题
        logger.info("\n检查表格标题:")
        for row in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and ('汇总表' in str(cell_value) or '趋势' in str(cell_value)):
                logger.info(f"  行{row}: {cell_value}")

        # 检查图表数量
        logger.info(f"\n图表数量: {len(ws._charts)}")
        for i, chart in enumerate(ws._charts, 1):
            logger.info(f"  图表{i}: {chart.title}")

    else:
        logger.error("✗ Sheet不存在")
        return False

    return True


if __name__ == "__main__":
    try:
        # 测试数据收集
        data = test_sheet6_data_collection()

        # 测试表格生成
        if data:
            success = test_sheet6_generation(data)
            if success:
                logger.info("\n" + "=" * 60)
                logger.info("所有测试通过! ✓")
                logger.info("=" * 60)
            else:
                logger.error("\n测试失败")
                sys.exit(1)
        else:
            logger.error("数据收集失败，无法继续测试")
            sys.exit(1)

    except Exception as e:
        logger.error(f"测试失败: {e}", exc_info=True)
        sys.exit(1)

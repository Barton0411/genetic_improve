"""大型 Excel 低内存写入器测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from utils.large_excel_writer import (
    ExcelSizeError,
    copy_file_atomic,
    normalize_identifier,
    normalize_identifier_key,
    read_excel_identifier_safe,
    write_dataframe_atomic,
)


class LargeExcelWriterTests(unittest.TestCase):
    def test_round_trip_preserves_rows_columns_ids_decimals_and_dates(self):
        frame = pd.DataFrame(
            {
                "cow_id": ["00123", 456.0, None],
                "牧场编号": ["0101017", "0102026", "0102027"],
                "NM$": [123.456789, -0.0049, np.nan],
                "birth_date": pd.to_datetime(
                    ["2020-01-02", "2021-03-04", None]
                ),
                "note": ["=1+1", "普通文本", None],
            }
        )

        with tempfile.TemporaryDirectory() as temporary_dir:
            output = Path(temporary_dir) / "result.xlsx"
            write_dataframe_atomic(frame, output)

            workbook = load_workbook(output, read_only=True, data_only=False)
            worksheet = workbook.active
            values = list(worksheet.iter_rows(values_only=True))
            note_data_type = worksheet["E2"].data_type
            workbook.close()

        self.assertEqual(
            values[0],
            ("cow_id", "牧场编号", "NM$", "birth_date", "note"),
        )
        self.assertEqual(values[1][0], "00123")
        self.assertEqual(values[2][0], "456")
        self.assertEqual(values[1][1], "0101017")
        self.assertAlmostEqual(values[1][2], 123.456789, places=6)
        self.assertAlmostEqual(values[2][2], -0.0049, places=6)
        self.assertEqual(values[1][3].date().isoformat(), "2020-01-02")
        self.assertEqual(values[1][4], "=1+1")
        self.assertEqual(
            note_data_type,
            "s",
            "外部文本不能被当作 Excel 公式执行",
        )

    def test_failed_write_keeps_previous_file_and_removes_temporary_file(self):
        frame = pd.DataFrame({"cow_id": ["1"], "NM$": [1.25]})

        with tempfile.TemporaryDirectory() as temporary_dir:
            output = Path(temporary_dir) / "result.xlsx"
            output.write_bytes(b"previous-valid-content")

            with patch(
                "utils.large_excel_writer.os.replace",
                side_effect=PermissionError("locked"),
            ):
                with self.assertRaises(PermissionError):
                    write_dataframe_atomic(frame, output)

            self.assertEqual(output.read_bytes(), b"previous-valid-content")
            self.assertEqual(
                list(Path(temporary_dir).glob(".result.*.tmp.xlsx")),
                [],
            )

    def test_rejects_data_beyond_excel_row_limit_before_writing(self):
        frame = pd.DataFrame({"cow_id": pd.RangeIndex(1_048_576)})

        with tempfile.TemporaryDirectory() as temporary_dir:
            output = Path(temporary_dir) / "result.xlsx"
            with self.assertRaisesRegex(ExcelSizeError, "超过 Excel"):
                write_dataframe_atomic(frame, output)
            self.assertFalse(output.exists())

    def test_source_formatting_follows_sorted_row_values_not_dataframe_index(self):
        frame = pd.DataFrame(
            {
                "cow_id": ["cow-yellow", "cow-red"],
                "sire_NM$": [10.25, 20.5],
                "sire_NM$_source": [3, 2],
                "mgs_NM$_source": [1, 1],
                "mmgs_NM$_source": [1, 1],
                "NM$_score": [8.5, 18.75],
            },
            index=[99, 3],
        )

        with tempfile.TemporaryDirectory() as temporary_dir:
            output = Path(temporary_dir) / "result.xlsx"
            write_dataframe_atomic(
                frame,
                output,
                apply_source_formatting=True,
            )

            workbook = load_workbook(output, data_only=False)
            worksheet = workbook.active
            self.assertEqual(worksheet["A2"].value, "cow-yellow")
            self.assertEqual(worksheet["B2"].fill.fgColor.rgb, "FF808080")
            self.assertEqual(worksheet["B2"].font.color.rgb, "FFFFFF00")
            self.assertEqual(worksheet["F2"].fill.fgColor.rgb, "FF808080")
            self.assertEqual(worksheet["B3"].font.color.rgb, "FFFF0000")
            self.assertEqual(worksheet["F3"].font.color.rgb, "FFFF0000")
            workbook.close()

    def test_identifier_normalization_handles_excel_numeric_forms(self):
        self.assertEqual(normalize_identifier(" 00123 "), "00123")
        self.assertEqual(normalize_identifier(123), "123")
        self.assertEqual(normalize_identifier(123.0), "123")
        self.assertEqual(normalize_identifier(np.int64(123)), "123")
        self.assertEqual(normalize_identifier(None), "")
        self.assertEqual(normalize_identifier(np.nan), "")
        self.assertEqual(normalize_identifier_key("123.0"), "123")
        self.assertEqual(normalize_identifier_key("00123.0"), "00123.0")

    def test_identifier_safe_read_preserves_lineage_and_farm_leading_zeroes(
        self,
    ):
        frame = pd.DataFrame(
            {
                "cow_id": ["000123", "000456"],
                "raw_dam_id": ["000001", "000002"],
                "牧场编号": ["0101001", "0101001"],
                "NM$": [123.45, -6.75],
            }
        )

        with tempfile.TemporaryDirectory() as temporary_dir:
            source = Path(temporary_dir) / "source.xlsx"
            output = Path(temporary_dir) / "round-trip.xlsx"
            frame.to_excel(source, index=False)

            loaded = read_excel_identifier_safe(source)
            write_dataframe_atomic(loaded, output)
            reloaded = read_excel_identifier_safe(output)

        self.assertEqual(reloaded["cow_id"].tolist(), ["000123", "000456"])
        self.assertEqual(
            reloaded["raw_dam_id"].tolist(),
            ["000001", "000002"],
        )
        self.assertEqual(
            reloaded["牧场编号"].tolist(),
            ["0101001", "0101001"],
        )
        self.assertTrue(pd.api.types.is_numeric_dtype(reloaded["NM$"]))
        self.assertEqual(reloaded["NM$"].tolist(), [123.45, -6.75])

    def test_atomic_copy_failure_preserves_previous_target(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "source.xlsx"
            output = root / "result.xlsx"
            source.write_bytes(b"new-valid-content")
            output.write_bytes(b"previous-valid-content")

            with patch(
                "utils.large_excel_writer.os.replace",
                side_effect=PermissionError("locked"),
            ):
                with self.assertRaises(PermissionError):
                    copy_file_atomic(source, output)

            self.assertEqual(output.read_bytes(), b"previous-valid-content")
            self.assertEqual(
                list(root.glob(".result.*.tmp.xlsx")),
                [],
            )


if __name__ == "__main__":
    unittest.main()

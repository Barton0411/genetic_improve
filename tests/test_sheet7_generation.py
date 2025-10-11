"""
测试Sheet 7生成
"""

from pathlib import Path
import sys

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.excel_report.data_collectors import collect_breeding_detail_data
from core.excel_report.sheet_builders import Sheet7Builder
from core.excel_report.formatters import StyleManager, ChartBuilder
from openpyxl import Workbook
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

logger = logging.getLogger(__name__)


def test_sheet7_data_collection():
    """测试Sheet7数据收集"""
    logger.info("=" * 60)
    logger.info("测试Sheet7数据收集")
    logger.info("=" * 60)

    # 使用测试项目路径
    test_project = Path("/Users/bozhenwang/GeneticImprove/去青青_2025_10_08_15_59")
    analysis_folder = test_project / "analysis_results"

    # 收集数据
    data = collect_breeding_detail_data(analysis_folder)

    # 检查数据结构
    if data:
        logger.info(f"\n数据键: {list(data.keys())}")

        if 'data' in data:
            df = data['data']
            logger.info(f"\n数据形状: {df.shape}")
            logger.info(f"列数: {len(df.columns)}")
            logger.info(f"行数: {len(df)}")
            logger.info(f"\n列名: {df.columns.tolist()}")

        if 'file_path' in data:
            logger.info(f"\n源文件: {data['file_path']}")

    logger.info("\n✓ 数据收集测试通过")
    return data


def test_sheet7_generation(data):
    """测试Sheet7生成"""
    logger.info("\n" + "=" * 60)
    logger.info("测试Sheet7表格生成")
    logger.info("=" * 60)

    # 创建workbook
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # 创建样式管理器和图表构建器
    style_manager = StyleManager()
    chart_builder = ChartBuilder()

    # 创建Sheet7Builder
    builder = Sheet7Builder(wb, style_manager, chart_builder)

    # 生成Sheet
    builder.build(data)

    # 保存测试文件
    output_path = Path("/tmp/test_sheet7.xlsx")
    wb.save(output_path)

    logger.info(f"\n✓ Sheet7生成成功")
    logger.info(f"✓ 测试文件已保存: {output_path}")

    # 验证sheet是否存在
    if "配种记录-隐性基因及近交系数明细" in wb.sheetnames:
        ws = wb["配种记录-隐性基因及近交系数明细"]
        logger.info(f"✓ Sheet名称正确: {ws.title}")
        logger.info(f"✓ 数据行数: {ws.max_row}")
        logger.info(f"✓ 数据列数: {ws.max_column}")

        # 检查表头
        logger.info("\n前5个表头:")
        for col in range(1, min(6, ws.max_column + 1)):
            cell_value = ws.cell(row=1, column=col).value
            logger.info(f"  列{col}: {cell_value}")
    else:
        logger.error("✗ Sheet不存在")
        return False

    return True


if __name__ == "__main__":
    try:
        # 测试数据收集
        data = test_sheet7_data_collection()

        # 测试表格生成
        if data:
            success = test_sheet7_generation(data)
            if success:
                logger.info("\n" + "=" * 60)
                logger.info("所有测试通过! ✓")
                logger.info("=" * 60)
            else:
                logger.error("\n测试失败")
                sys.exit(1)
        else:
            logger.error("数据收集失败，无法继续测试")
            sys.exit(1)

    except Exception as e:
        logger.error(f"测试失败: {e}", exc_info=True)
        sys.exit(1)

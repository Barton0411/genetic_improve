"""
测试完整报告生成（包含Sheet 5）
"""

from pathlib import Path
import sys

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.excel_report.generator import ExcelReportGenerator
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

logger = logging.getLogger(__name__)


def test_full_report_generation():
    """测试完整报告生成（包含Sheet 5）"""
    logger.info("=" * 60)
    logger.info("测试完整报告生成（包含Sheet 5）")
    logger.info("=" * 60)

    # 使用测试项目路径
    test_project = Path("/Users/bozhenwang/GeneticImprove/去青青_2025_10_08_15_59")

    # 创建报告生成器
    generator = ExcelReportGenerator(
        project_folder=test_project,
        service_staff="测试人员"
    )

    # 生成报告
    success, result = generator.generate()

    if success:
        logger.info("\n" + "=" * 60)
        logger.info("✓ 报告生成成功!")
        logger.info(f"✓ 报告路径: {result}")
        logger.info("=" * 60)

        # 验证Sheet 5
        from openpyxl import load_workbook
        wb = load_workbook(result)

        if "配种记录-隐性基因分析" in wb.sheetnames:
            ws = wb["配种记录-隐性基因分析"]
            logger.info("\n✓ Sheet 5验证:")
            logger.info(f"  - Sheet名称: {ws.title}")
            logger.info(f"  - 数据行数: {ws.max_row}")
            logger.info(f"  - 数据列数: {ws.max_column}")

            # 检查标题
            title1 = ws.cell(row=1, column=1).value
            logger.info(f"  - 表1标题: {title1}")

            # 检查表头
            headers = []
            for col in range(1, min(11, ws.max_column + 1)):
                header = ws.cell(row=2, column=col).value
                headers.append(header)
            logger.info(f"  - 表头列: {headers}")

            # 检查第一行数据
            if ws.max_row >= 3:
                gene_name = ws.cell(row=3, column=1).value
                gene_translation = ws.cell(row=3, column=2).value
                logger.info(f"  - 第一个基因: {gene_name} ({gene_translation})")

            logger.info("\n✓ Sheet 5集成成功!")
        else:
            logger.error("\n✗ Sheet 5不存在!")
            return False

        return True
    else:
        logger.error(f"\n✗ 报告生成失败: {result}")
        return False


if __name__ == "__main__":
    try:
        success = test_full_report_generation()
        sys.exit(0 if success else 1)
    except Exception as e:
        logger.error(f"测试失败: {e}", exc_info=True)
        sys.exit(1)

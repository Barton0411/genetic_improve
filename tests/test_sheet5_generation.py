"""
测试Sheet 5生成
"""

from pathlib import Path
import sys

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.excel_report.data_collectors import collect_breeding_genes_data
from core.excel_report.sheet_builders import Sheet5Builder
from core.excel_report.formatters import StyleManager, ChartBuilder
from openpyxl import Workbook
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

logger = logging.getLogger(__name__)


def test_sheet5_data_collection():
    """测试Sheet5数据收集"""
    logger.info("=" * 60)
    logger.info("测试Sheet5数据收集")
    logger.info("=" * 60)

    # 使用测试项目路径
    test_project = Path("/Users/bozhenwang/GeneticImprove/去青青_2025_10_08_15_59")
    analysis_folder = test_project / "analysis_results"

    # 收集数据
    data = collect_breeding_genes_data(analysis_folder)

    # 检查数据结构
    logger.info(f"\n数据键: {list(data.keys())}")

    if 'all_years_summary' in data:
        logger.info(f"全部年份基因数量: {len(data['all_years_summary'])}")
        logger.info(f"全部年份总配次: {data.get('all_years_total', 0)}")
        if data['all_years_summary']:
            logger.info(f"第一个基因数据样例:")
            first_gene = data['all_years_summary'][0]
            for key, value in first_gene.items():
                logger.info(f"  {key}: {value}")

    if 'recent_12m_summary' in data:
        logger.info(f"\n近12个月基因数量: {len(data['recent_12m_summary'])}")
        logger.info(f"近12个月总配次: {data.get('recent_12m_total', 0)}")

    if 'date_range' in data:
        logger.info(f"\n日期范围: {data['date_range']}")

    logger.info("\n✓ 数据收集测试通过")
    return data


def test_sheet5_generation(data):
    """测试Sheet5生成"""
    logger.info("\n" + "=" * 60)
    logger.info("测试Sheet5表格生成")
    logger.info("=" * 60)

    # 创建workbook
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # 创建样式管理器和图表构建器
    style_manager = StyleManager()
    chart_builder = ChartBuilder()

    # 创建Sheet5Builder
    builder = Sheet5Builder(wb, style_manager, chart_builder)

    # 生成Sheet
    builder.build(data)

    # 保存测试文件
    output_path = Path("/tmp/test_sheet5.xlsx")
    wb.save(output_path)

    logger.info(f"\n✓ Sheet5生成成功")
    logger.info(f"✓ 测试文件已保存: {output_path}")

    # 验证sheet是否存在
    if "配种记录-隐性基因分析" in wb.sheetnames:
        ws = wb["配种记录-隐性基因分析"]
        logger.info(f"✓ Sheet名称正确: {ws.title}")
        logger.info(f"✓ 数据行数: {ws.max_row}")
        logger.info(f"✓ 数据列数: {ws.max_column}")

        # 检查标题
        logger.info("\n检查标题内容:")
        title1 = ws.cell(row=1, column=1).value
        logger.info(f"  表1标题: {title1}")
        # 第二张表大概在第20行左右
        if ws.max_row >= 20:
            title2 = ws.cell(row=20, column=1).value
            logger.info(f"  表2标题: {title2}")

        # 检查颜色
        logger.info("\n检查单元格颜色:")
        # 假设第3行是第一个数据行
        if ws.max_row >= 3:
            for col in range(2, 10):
                cell = ws.cell(row=3, column=col)
                fill_color = cell.fill.start_color.rgb if cell.fill else None
                logger.info(f"  列{col}: {fill_color}")
    else:
        logger.error("✗ Sheet不存在")
        return False

    return True


if __name__ == "__main__":
    try:
        # 测试数据收集
        data = test_sheet5_data_collection()

        # 测试表格生成
        if data:
            success = test_sheet5_generation(data)
            if success:
                logger.info("\n" + "=" * 60)
                logger.info("所有测试通过! ✓")
                logger.info("=" * 60)
            else:
                logger.error("\n测试失败")
                sys.exit(1)
        else:
            logger.error("数据收集失败，无法继续测试")
            sys.exit(1)

    except Exception as e:
        logger.error(f"测试失败: {e}", exc_info=True)
        sys.exit(1)

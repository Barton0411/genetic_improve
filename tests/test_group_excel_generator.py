from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

import xlsxwriter
from openpyxl import load_workbook

from core.group_report.excel_generator import GroupExcelReportGenerator
from core.group_tasks.lease_heartbeat import GroupSelectionFenceError
from core.group_tasks.stage_policy import commit_child_stage
from utils.file_manager import FileManager


def _write_book(path: Path, sheets: dict[str, list[list]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(path))
    for sheet_name, rows in sheets.items():
        worksheet = workbook.add_worksheet(sheet_name[:31])
        for row_index, row in enumerate(rows):
            worksheet.write_row(row_index, 0, row)
    workbook.close()


class GroupExcelGeneratorTests(unittest.TestCase):
    def _prepare_project(self, temporary_dir: str) -> tuple[Path, str]:
        project = FileManager.create_group_project(
            Path(temporary_dir),
            [{"code": "010", "name": "测试牧场"}],
            data_source="伊起牛",
            task_mode="analysis",
        )
        metadata = FileManager.load_project_metadata(project)
        task = metadata["group_tasks"][0]
        task_id = task["task_id"]
        child = project / task["relative_path"]
        analysis = child / "analysis_results"

        _write_book(
            child / "raw_data" / "cow_data.xlsx",
            {"Sheet1": [["cow_id"], ["001"]]},
        )
        _write_book(
            child / "standardized_data" / "processed_cow_data.xlsx",
            {"Sheet1": [["cow_id", "牧场编号"], ["001", "010"]]},
        )
        _write_book(
            analysis / "processed_cow_data_key_traits_final.xlsx",
            {"Sheet1": [["cow_id", "牧场编号"], ["001", "010"]]},
        )
        _write_book(
            analysis / "processed_index_cow_index_scores.xlsx",
            {
                "Sheet1": [
                    [
                        "cow_id",
                        "raw_cow_id",
                        "是否在场",
                        "统一指数_index",
                        "牧场编号",
                    ],
                    ["001", "001", "是", 12.345678901, "010"],
                ]
            },
        )
        total_headers = [
            "出生年份",
            "头数",
            "平均NM$",
            "平均TPI",
        ]
        _write_book(
            analysis / "关键育种性状分析结果.xlsx",
            {
                "在群母牛年份汇总": [
                    total_headers,
                    ["2020年", 1, 100.25, 2000.75],
                    ["在群母牛总计", 1, 100.25, 2000.75],
                ],
                "全部母牛年份汇总": [
                    total_headers,
                    ["全部母牛总计", 1, 100.25, 2000.75],
                ],
                "在群母牛NM$分布": [
                    ["分布区间", "头数", "占比"],
                    ["100及以上", 1, "100%"],
                ],
                "在群母牛TPI分布": [
                    ["分布区间", "头数", "占比"],
                    ["2000及以上", 1, "100%"],
                ],
            },
        )
        _write_book(
            analysis / "系谱识别分析结果.xlsx",
            {
                "Sheet1": [
                    [
                        "是否在群",
                        "头数",
                        "父号可识别头数",
                        "外祖父可识别头数",
                        "外曾外祖父可识别头数",
                    ],
                    ["是", 1, 1, 1, 1],
                ]
            },
        )
        _write_book(
            child / "reports" / "育种分析综合报告_测试牧场.xlsx",
            {"报告": [["完整单场报告"]]},
        )
        for stage in ("data", "analysis", "child_excel"):
            commit_child_stage(
                child,
                stage,
                expected_task_id=task_id,
                expected_farm_code="010",
            )
            FileManager.update_group_stage(
                project,
                task_id,
                stage,
                status="completed",
            )
        # 阶段提交后的历史工作簿仅进入浏览索引，不能阻断正式发布。
        _write_book(
            child / "reports" / "历史结果.xlsx",
            {"历史": [["未受管"]]},
        )
        return project, task_id

    def test_publishes_snapshot_bound_summary_and_complete_indexes(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project, _ = self._prepare_project(temporary_dir)

            success, result = GroupExcelReportGenerator(project).generate()
            self.assertTrue(success, result)
            output = Path(result)
            self.assertTrue(output.is_file())

            workbook = load_workbook(output, data_only=False)
            try:
                self.assertIn("完整明细索引", workbook.sheetnames)
                self.assertIn("全部结果文件索引", workbook.sheetnames)
                self.assertIn("完整性校验", workbook.sheetnames)

                detail_link = workbook["完整明细索引"]["G3"].hyperlink
                self.assertIsNotNone(detail_link)
                self.assertFalse(detail_link.target.startswith("file:"))
                file_link = workbook["全部结果文件索引"]["I3"].hyperlink
                self.assertIsNotNone(file_link)
                self.assertTrue(
                    file_link.target.startswith("../../farm_projects/")
                )
                inventory_sheet = workbook["全部结果文件索引"]
                unmanaged_rows = [
                    row
                    for row in inventory_sheet.iter_rows(
                        min_row=3,
                        values_only=True,
                    )
                    if row[8] and str(row[8]).endswith("历史结果.xlsx")
                ]
                self.assertEqual(len(unmanaged_rows), 1)
                self.assertEqual(unmanaged_rows[0][5], "否")
                self.assertEqual(unmanaged_rows[0][11], "未校验")
            finally:
                workbook.close()

            published = FileManager.load_project_metadata(project)[
                "group_results"
            ]
            self.assertEqual(published["status"], "current")
            self.assertEqual(published["selection_revision"], 0)
            self.assertEqual(
                published["artifact_inventory_counts"]["invalid_files"],
                0,
            )
            pointer = project / "group_store" / "current_group_report.json"
            self.assertTrue(pointer.is_file())
            self.assertTrue(Path(published["report_package_path"]).is_dir())
            self.assertTrue(Path(published["batch_manifest_path"]).is_file())
            published_snapshot = Path(
                published["publication_snapshot_path"]
            )
            self.assertTrue(published_snapshot.is_file())
            self.assertEqual(
                published_snapshot.parent,
                Path(published["report_package_path"]),
            )
            self.assertTrue(published["detail_manifest_sha256"])

    def test_selection_change_during_generation_blocks_publication(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project, task_id = self._prepare_project(temporary_dir)
            changed = False

            def change_selection_once(_value, _message):
                nonlocal changed
                if not changed:
                    changed = True
                    FileManager.set_group_task_excluded(
                        project,
                        task_id,
                        True,
                    )

            with patch(
                "core.group_report.excel_generator.GroupLeaseHeartbeat"
            ) as heartbeat_type:
                heartbeat = heartbeat_type.return_value
                heartbeat.check.side_effect = [
                    None,
                    GroupSelectionFenceError(0, 1),
                ]
                success, result = GroupExcelReportGenerator(
                    project,
                    progress_callback=change_selection_once,
                ).generate()
            self.assertFalse(success)
            self.assertIn("范围发生变化", result)
            metadata = FileManager.load_project_metadata(project)
            self.assertNotEqual(
                metadata.get("group_results", {}).get("status"),
                "current",
            )
            heartbeat.start.assert_called_once_with()
            heartbeat.stop.assert_called_once_with(
                timeout=30,
                release=True,
            )

    def test_valid_pointer_repairs_metadata_without_regenerating_details(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project, _ = self._prepare_project(temporary_dir)
            success, first = GroupExcelReportGenerator(project).generate()
            self.assertTrue(success, first)

            metadata = FileManager.load_project_metadata(project)
            metadata["group_results"] = {"status": "stale"}
            FileManager._write_json_atomic(
                project / "project_metadata.json",
                metadata,
            )

            with patch(
                "core.group_report.publication_batch."
                "GroupReportPublicationBatch",
                side_effect=AssertionError(
                    "有效正式指针不应重新建立报告批次"
                ),
            ):
                success, second = GroupExcelReportGenerator(
                    project
                ).generate()

            self.assertTrue(success, second)
            self.assertEqual(Path(second), Path(first))
            repaired = FileManager.load_project_metadata(project)[
                "group_results"
            ]
            self.assertEqual(repaired["status"], "current")
            self.assertEqual(Path(repaired["excel_path"]), Path(first))

    def test_hmy_summary_keeps_api_number_and_display_name_separate(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project, task_id = self._prepare_project(temporary_dir)
            store = FileManager._group_task_store(project)
            self.assertIsNotNone(store)
            store.update_task(
                task_id,
                farm_name="合肥陈刘牧场",
                metadata={
                    "api_farmcode": "010",
                    "farm_number": "0101001",
                    "display_name": "合肥陈刘牧场",
                    "source_farm_name": "0101001合肥陈刘牧场",
                },
            )
            metadata_path = project / "project_metadata.json"
            metadata = FileManager.load_project_metadata(project)
            metadata["data_source"] = "慧牧云"
            metadata["interface_source"] = "慧牧云"
            FileManager._write_json_atomic(metadata_path, metadata)

            success, result = GroupExcelReportGenerator(project).generate()
            self.assertTrue(success, result)

            workbook = load_workbook(result, read_only=True, data_only=True)
            try:
                identity_columns = {
                    "任务完成情况": 1,
                    "牧场对比总览": 1,
                    "系谱识别对比": 1,
                    "关键性状对比": 1,
                    "年份遗传进展": 1,
                    "跨牧场牛只排名": 3,
                    "全部结果文件索引": 2,
                    "数据可用性": 1,
                    "单牧场报告索引": 1,
                }
                for sheet_name, first_column in identity_columns.items():
                    sheet = workbook[sheet_name]
                    headers = [
                        sheet.cell(2, column).value
                        for column in range(
                            first_column,
                            first_column + 3,
                        )
                    ]
                    self.assertEqual(
                        headers,
                        ["API farmcode", "牧场编号", "牧场名称"],
                        sheet_name,
                    )
                    self.assertEqual(
                        sheet.cell(3, first_column).value,
                        "010",
                        sheet_name,
                    )
                    self.assertEqual(
                        sheet.cell(3, first_column + 1).value,
                        "0101001",
                        sheet_name,
                    )
                    self.assertEqual(
                        sheet.cell(3, first_column + 2).value,
                        "合肥陈刘牧场",
                        sheet_name,
                    )
            finally:
                workbook.close()

    def test_summary_lease_lifecycle_uses_background_heartbeat_once(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project, _ = self._prepare_project(temporary_dir)
            store = MagicMock()
            store.get_selection_revision.return_value = 7
            store.acquire_run_lease.return_value = {
                "lease_token": "summary-token",
                "selection_revision": 7,
                "current_selection_revision": 7,
                "selection_is_current": True,
            }
            with (
                patch.object(
                    FileManager,
                    "_group_task_store",
                    return_value=store,
                ),
                patch(
                    "core.group_report.excel_generator.GroupLeaseHeartbeat"
                ) as heartbeat_type,
            ):
                generator = GroupExcelReportGenerator(project)
                generator._acquire_summary_lease()
                generator._refresh_summary_lease()
                generator._release_summary_lease()
                generator._release_summary_lease()

            heartbeat_type.assert_called_once_with(
                store,
                store.acquire_run_lease.return_value,
                lease_seconds=600,
            )
            heartbeat = heartbeat_type.return_value
            heartbeat.start.assert_called_once_with()
            heartbeat.check.assert_called_once_with()
            heartbeat.stop.assert_called_once_with(
                timeout=30,
                release=True,
            )
            store.release_run_lease.assert_not_called()


if __name__ == "__main__":
    unittest.main()

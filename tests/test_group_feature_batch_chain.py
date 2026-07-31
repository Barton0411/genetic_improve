from __future__ import annotations

import io
import os
import sys
import tempfile
import time
import types
import unittest
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402
from PyQt6.QtWidgets import QApplication  # noqa: E402

from core.group_tasks.child_runner import (  # noqa: E402
    ChildExecutionError,
    ChildRequestError,
)
from core.group_tasks.feature_policy import (  # noqa: E402
    FeaturePolicyError,
    commit_feature_manifest,
    feature_manifest_path,
    normalize_feature_parameters,
    validate_feature_manifest,
)
from core.group_tasks.feature_process import (  # noqa: E402
    build_feature_request,
    write_feature_request,
)
from core.group_tasks.feature_runner import (  # noqa: E402
    execute_feature_request,
    validate_feature_request,
)
from gui.group_feature_analysis_worker import (  # noqa: E402
    GroupFeatureAnalysisWorker,
)
from utils.file_manager import FileManager  # noqa: E402


def _write_xlsx(
    path: Path,
    headers: list[str],
    rows: list[list[object]],
    *,
    sheet_name: str = "Sheet",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


class GroupFeatureRequestValidationTests(unittest.TestCase):
    def setUp(self):
        self.temporary_dir = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temporary_dir.name)
        self.parent = FileManager.create_group_project(
            self.base_path,
            [{"code": "1001", "name": "测试牧场"}],
            data_source="伊起牛",
            task_mode="analysis",
        )
        metadata = FileManager.load_project_metadata(self.parent)
        self.task = metadata["group_tasks"][0]

    def tearDown(self):
        self.temporary_dir.cleanup()

    @staticmethod
    def _valid_parameters() -> dict[str, object]:
        return {
            "weight_name": "稳定权重",
            "weight_values": {"NM$": 100},
        }

    def test_index_request_requires_immutable_weight_snapshot(self):
        with self.assertRaisesRegex(
            FeaturePolicyError,
            "weight_name 和 weight_values",
        ):
            build_feature_request(
                self.parent,
                self.task["task_id"],
                "bull_index",
                {"weight_name": "稳定权重"},
            )

    def test_request_rejects_unknown_top_level_and_parameter_fields(self):
        payload = build_feature_request(
            self.parent,
            self.task["task_id"],
            "bull_index",
            self._valid_parameters(),
        )

        unexpected = dict(payload)
        unexpected["api_token"] = "DO-NOT-ECHO"
        with self.assertRaises(ChildRequestError) as context:
            validate_feature_request(unexpected)
        self.assertIn("api_token", str(context.exception))
        self.assertNotIn("DO-NOT-ECHO", str(context.exception))

        unexpected_parameters = dict(payload)
        unexpected_parameters["parameters"] = {
            **self._valid_parameters(),
            "api_token": "DO-NOT-ECHO",
        }
        with self.assertRaises(ChildRequestError) as context:
            validate_feature_request(unexpected_parameters)
        self.assertNotIn("DO-NOT-ECHO", str(context.exception))

        wrong_farm = dict(payload)
        wrong_farm["farm_code"] = "9999"
        with self.assertRaisesRegex(
            ChildRequestError,
            "farm_code 与父组任务不一致",
        ):
            validate_feature_request(wrong_farm)

    def test_operation_parameters_are_strict_and_canonical(self):
        with self.assertRaisesRegex(FeaturePolicyError, "重复"):
            normalize_feature_parameters(
                "cow_traits",
                {"traits": ["NM$", " NM$ "]},
            )
        with self.assertRaisesRegex(FeaturePolicyError, "不支持的性状"):
            normalize_feature_parameters(
                "bull_traits",
                {"traits": ["NOT_A_TRAIT"]},
            )
        with self.assertRaisesRegex(FeaturePolicyError, "不接受额外参数"):
            normalize_feature_parameters(
                "cow_self_inbreeding",
                {"generations": 3},
            )

        normalized = normalize_feature_parameters(
            "cow_index",
            {
                "weight_values": {"TPI": 40, "NM$": 60, "DPR": 0},
                "weight_name": "  稳定权重  ",
            },
        )
        self.assertEqual(normalized["weight_name"], "稳定权重")
        self.assertEqual(
            list(normalized["weight_values"]),
            ["NM$", "TPI"],
        )


class GroupFeatureManifestIntegrityTests(unittest.TestCase):
    def setUp(self):
        self.temporary_dir = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temporary_dir.name)
        self.parent = FileManager.create_group_project(
            self.base_path,
            [{"code": "1501", "name": "完整性牧场"}],
            data_source="伊起牛",
            task_mode="analysis",
        )
        metadata = FileManager.load_project_metadata(self.parent)
        self.task = metadata["group_tasks"][0]
        self.child = (
            self.parent / self.task["relative_path"]
        ).resolve(strict=True)

    def tearDown(self):
        self.temporary_dir.cleanup()

    def test_rejects_output_with_same_count_but_duplicated_wrong_bull_id(self):
        _write_xlsx(
            self.child
            / "standardized_data"
            / "processed_bull_data.xlsx",
            ["bull_id"],
            [["A"], ["B"]],
        )
        _write_xlsx(
            self.child
            / "analysis_results"
            / "processed_index_bull_scores.xlsx",
            ["bull_id", "稳定权重_index", "ranking"],
            [["A", 1.0, 1], ["A", 0.5, 2]],
        )

        with self.assertRaisesRegex(
            FeaturePolicyError,
            "牛号多重集",
        ):
            commit_feature_manifest(
                self.child,
                "bull_index",
                {
                    "weight_name": "稳定权重",
                    "weight_values": {"NM$": 100},
                },
                expected_task_id=self.task["task_id"],
                expected_farm_code="1501",
                bull_library_version="test-db",
            )

    def test_cow_traits_allows_business_filter_and_ignores_styled_blank_tail(
        self,
    ):
        cow_path = (
            self.child
            / "standardized_data"
            / "processed_cow_data.xlsx"
        )
        _write_xlsx(
            cow_path,
            ["cow_id", "sex", "breed"],
            [
                ["COW-A", "母", "荷斯坦"],
                ["COW-B", "公", "荷斯坦"],
                ["COW-C", "母", "安格斯"],
            ],
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["cow_id", "sex", "breed"])
        worksheet.append(["COW-A", "母", "荷斯坦"])
        worksheet.append(["COW-B", "公", "荷斯坦"])
        worksheet.append(["COW-C", "母", "安格斯"])
        worksheet["A50"].fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00",
        )
        workbook.save(cow_path)

        result_dir = self.child / "analysis_results"
        _write_xlsx(
            result_dir / "processed_cow_data_key_traits_final.xlsx",
            ["cow_id", "NM$_score"],
            [["COW-A", 1.25]],
        )
        for filename in (
            "关键育种性状分析结果.xlsx",
            "系谱识别分析结果.xlsx",
            "sire_traits_mean_by_cow_birth_year.xlsx",
        ):
            _write_xlsx(
                result_dir / filename,
                ["result"],
                [["ok"]],
            )

        manifest = commit_feature_manifest(
            self.child,
            "cow_traits",
            {"traits": ["NM$"]},
            expected_task_id=self.task["task_id"],
            expected_farm_code="1501",
            bull_library_version="test-db",
        )
        self.assertEqual(manifest["status"], "committed")

    def test_cow_index_tracks_actual_score_file_as_direct_input(self):
        _write_xlsx(
            self.child
            / "standardized_data"
            / "processed_cow_data.xlsx",
            ["cow_id"],
            [["COW-A"], ["COW-B"]],
        )
        score_path = (
            self.child
            / "analysis_results"
            / "processed_cow_data_key_traits_scores_genomic.xlsx"
        )
        _write_xlsx(
            score_path,
            ["cow_id", "NM$_score"],
            [["COW-A", 1.0], ["COW-B", 0.5]],
        )
        _write_xlsx(
            self.child
            / "analysis_results"
            / "processed_index_cow_index_scores.xlsx",
            ["cow_id", "稳定权重_index", "ranking"],
            [["COW-A", 1.0, 1], ["COW-B", 0.5, 2]],
        )
        parameters = {
            "weight_name": "稳定权重",
            "weight_values": {"NM$": 100},
        }

        manifest = commit_feature_manifest(
            self.child,
            "cow_index",
            parameters,
            expected_task_id=self.task["task_id"],
            expected_farm_code="1501",
            bull_library_version="test-db",
        )
        self.assertIn(
            score_path.relative_to(self.child).as_posix(),
            {
                item["logical_name"]
                for item in manifest["inputs"]
            },
        )

        _write_xlsx(
            score_path,
            ["cow_id", "NM$_score"],
            [["COW-A", 9.0], ["COW-B", 0.5]],
        )
        validation = validate_feature_manifest(
            self.child,
            "cow_index",
            parameters,
            expected_task_id=self.task["task_id"],
            expected_farm_code="1501",
            bull_library_version="test-db",
        )
        self.assertFalse(validation["valid"])
        self.assertEqual(validation["status"], "artifact_mismatch")


class GroupFeatureResumeTests(unittest.TestCase):
    def setUp(self):
        self.temporary_dir = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temporary_dir.name)
        self.parent = FileManager.create_group_project(
            self.base_path,
            [{"code": "2001", "name": "断点牧场"}],
            data_source="伊起牛",
            task_mode="analysis",
        )
        metadata = FileManager.load_project_metadata(self.parent)
        self.task = metadata["group_tasks"][0]
        self.child = (
            self.parent / self.task["relative_path"]
        ).resolve(strict=True)
        _write_xlsx(
            self.child
            / "standardized_data"
            / "processed_bull_data.xlsx",
            ["bull_id"],
            [["001HO00001"], ["001HO00002"]],
        )

    def tearDown(self):
        self.temporary_dir.cleanup()

    @staticmethod
    def _parameters(
        *,
        nm_weight: float,
        tpi_weight: float,
    ) -> dict[str, object]:
        values = {}
        if nm_weight:
            values["NM$"] = nm_weight
        if tpi_weight:
            values["TPI"] = tpi_weight
        return {
            "weight_name": "稳定权重",
            "weight_values": values,
        }

    def _write_index_result(self) -> None:
        _write_xlsx(
            self.child
            / "analysis_results"
            / "processed_index_bull_scores.xlsx",
            ["bull_id", "稳定权重_index", "ranking"],
            [
                ["001HO00001", 1.25, 1],
                ["001HO00002", 0.75, 2],
            ],
        )

    def test_same_parameter_manifest_resumes_but_changed_snapshot_recomputes(self):
        first_parameters = self._parameters(
            nm_weight=100,
            tpi_weight=0,
        )
        first_request = write_feature_request(
            self.parent,
            self.task["task_id"],
            "bull_index",
            first_parameters,
        )
        operation_calls: list[dict[str, object]] = []

        def execute_once(request, progress_callback):
            operation_calls.append(dict(request.parameters))
            progress_callback(50, "测试计算")
            self._write_index_result()
            return True, "完成"

        with (
            patch(
                "core.group_tasks.feature_runner._execute_operation",
                side_effect=execute_once,
            ),
            patch(
                "core.data.update_manager.get_local_db_version",
                return_value="test-bull-db",
            ),
        ):
            first = execute_feature_request(
                first_request,
                output_stream=io.StringIO(),
            )
            resumed = execute_feature_request(
                first_request,
                output_stream=io.StringIO(),
            )

            second_parameters = self._parameters(
                nm_weight=60,
                tpi_weight=40,
            )
            second_request = write_feature_request(
                self.parent,
                self.task["task_id"],
                "bull_index",
                second_parameters,
            )
            recomputed = execute_feature_request(
                second_request,
                output_stream=io.StringIO(),
            )

            current = validate_feature_manifest(
                self.child,
                "bull_index",
                second_parameters,
                expected_task_id=self.task["task_id"],
                expected_farm_code="2001",
            )

        self.assertFalse(first["resumed"])
        self.assertTrue(resumed["resumed"])
        self.assertFalse(recomputed["resumed"])
        self.assertEqual(len(operation_calls), 2)
        self.assertEqual(
            operation_calls,
            [
                normalize_feature_parameters(
                    "bull_index",
                    first_parameters,
                ),
                normalize_feature_parameters(
                    "bull_index",
                    second_parameters,
                ),
            ],
        )
        self.assertTrue(current["valid"])
        history = (
            self.child
            / "group_store"
            / "stage_manifests"
            / "features"
            / "history"
        )
        self.assertTrue(
            any(history.glob("bull_index_replaced_*.json")),
            "参数改变后应归档旧清单并重新计算",
        )

    def test_bull_library_version_change_during_run_refuses_commit(self):
        parameters = self._parameters(
            nm_weight=100,
            tpi_weight=0,
        )
        request_path = write_feature_request(
            self.parent,
            self.task["task_id"],
            "bull_index",
            parameters,
        )

        def execute_once(_request, _progress_callback):
            self._write_index_result()
            return True, "完成"

        versions = iter(("db-v1", "db-v1", "db-v2"))

        with (
            patch(
                "core.group_tasks.feature_runner._execute_operation",
                side_effect=execute_once,
            ),
            patch(
                "core.data.update_manager.get_local_db_version",
                side_effect=lambda: next(versions, "db-v2"),
            ),
        ):
            with self.assertRaisesRegex(
                ChildExecutionError,
                "公牛库版本",
            ):
                execute_feature_request(
                    request_path,
                    output_stream=io.StringIO(),
                )

        self.assertFalse(
            (
                self.child
                / feature_manifest_path("bull_index")
            ).exists(),
            "版本变化的计算结果不能留下可复用清单",
        )


class GroupFeatureTimestampOutputTests(unittest.TestCase):
    def setUp(self):
        self.temporary_dir = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temporary_dir.name)
        self.parent = FileManager.create_group_project(
            self.base_path,
            [{"code": "2501", "name": "时间戳牧场"}],
            data_source="伊起牛",
            task_mode="analysis",
        )
        metadata = FileManager.load_project_metadata(self.parent)
        self.task = metadata["group_tasks"][0]
        self.child = (
            self.parent / self.task["relative_path"]
        ).resolve(strict=True)
        _write_xlsx(
            self.child
            / "standardized_data"
            / "processed_cow_data.xlsx",
            ["cow_id"],
            [["COW-A"]],
        )
        _write_xlsx(
            self.child
            / "standardized_data"
            / "processed_breeding_data.xlsx",
            ["耳号", "冻精编号"],
            [["COW-A", "BULL-A"]],
        )

    def tearDown(self):
        self.temporary_dir.cleanup()

    @staticmethod
    def _write_mated_result(path: Path) -> None:
        _write_xlsx(
            path,
            ["母牛号", "配种公牛号", "后代近交系数"],
            [["COW-A", "BULL-A", 0.01]],
            sheet_name="配对明细表",
        )

    def test_commits_new_result_not_unchanged_old_file_with_future_mtime(self):
        result_dir = self.child / "analysis_results"
        old_path = (
            result_dir
            / "已配公牛_近交系数及隐性基因分析结果_20990101_000000.xlsx"
        )
        self._write_mated_result(old_path)
        future = time.time_ns() + 365 * 24 * 60 * 60 * 1_000_000_000
        os.utime(old_path, ns=(future, future))

        new_path = (
            result_dir
            / "已配公牛_近交系数及隐性基因分析结果_20260731_140000.xlsx"
        )
        request_path = write_feature_request(
            self.parent,
            self.task["task_id"],
            "mated_inbreeding",
            {},
        )

        def execute_once(_request, _progress_callback):
            self._write_mated_result(new_path)
            return True, "完成"

        with (
            patch(
                "core.group_tasks.feature_runner._execute_operation",
                side_effect=execute_once,
            ),
            patch(
                "core.data.update_manager.get_local_db_version",
                return_value="test-db",
            ),
        ):
            result = execute_feature_request(
                request_path,
                output_stream=io.StringIO(),
            )

        self.assertEqual(
            result["artifacts"],
            [new_path.relative_to(self.child).as_posix()],
        )
        self.assertNotIn(
            old_path.relative_to(self.child).as_posix(),
            result["artifacts"],
        )


class GroupFeatureWorkerTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def setUp(self):
        self.temporary_dir = tempfile.TemporaryDirectory()
        self.base_path = Path(self.temporary_dir.name)
        self.parent = FileManager.create_group_project(
            self.base_path,
            [
                {"code": "3001", "name": "一号牧场"},
                {"code": "3002", "name": "二号牧场"},
                {"code": "3003", "name": "三号牧场"},
            ],
            data_source="伊起牛",
            task_mode="analysis",
        )
        self.metadata = FileManager.load_project_metadata(self.parent)

    def tearDown(self):
        self.temporary_dir.cleanup()

    def test_runs_strictly_serial_continues_after_one_failure_and_never_summarizes(
        self,
    ):
        existing_report = (
            self.parent
            / "reports"
            / "牧场组育种分析汇总报告_既有.xlsx"
        )
        _write_xlsx(existing_report, ["result"], [["existing"]])
        pointer = (
            self.parent
            / "group_store"
            / "current_group_report.json"
        )
        pointer.write_text(
            '{"kind":"existing-test-pointer"}',
            encoding="utf-8",
        )
        report_before = (
            existing_report.read_bytes(),
            existing_report.stat().st_mtime_ns,
        )
        pointer_before = (
            pointer.read_bytes(),
            pointer.stat().st_mtime_ns,
        )

        worker = GroupFeatureAnalysisWorker(
            self.parent,
            "cow_traits",
            {"traits": ["NM$", "TPI"]},
        )
        call_order: list[str] = []
        active_children = 0
        max_active_children = 0
        finished_results: list[dict[str, object]] = []
        errors: list[str] = []
        done_events: list[tuple[str, bool]] = []

        worker.finished.connect(finished_results.append)
        worker.error.connect(errors.append)
        worker.sub_task_done.connect(
            lambda task_id, success: done_events.append(
                (task_id, success)
            )
        )

        def run_child(*, task, **_kwargs):
            nonlocal active_children, max_active_children
            task_id = str(task["task_id"])
            call_order.append(task_id)
            active_children += 1
            max_active_children = max(max_active_children, active_children)
            try:
                if str(task["farm_code"]) == "3002":
                    raise RuntimeError("该牧场测试失败")
                return {
                    "success": True,
                    "skipped": False,
                    "resumed": False,
                    "message": "完成",
                }
            finally:
                active_children -= 1

        forbidden_group_report = types.ModuleType("core.group_report")

        class ForbiddenGroupExcelReportGenerator:
            def __init__(self, *_args, **_kwargs):
                raise AssertionError("单项批量绝不允许生成牧场组汇总")

        forbidden_group_report.GroupExcelReportGenerator = (
            ForbiddenGroupExcelReportGenerator
        )

        with (
            patch.object(worker, "_acquire_lease"),
            patch.object(worker, "_release_lease"),
            patch.object(worker, "_check_lease"),
            patch.object(worker, "_run_child", side_effect=run_child),
            patch("gui.group_feature_analysis_worker.logger.warning"),
            patch(
                "core.data.update_manager.reset_pedigree_db"
            ),
            patch.dict(
                sys.modules,
                {"core.group_report": forbidden_group_report},
            ),
        ):
            worker.run()

        expected_task_ids = [
            str(task["task_id"])
            for task in self.metadata["group_tasks"]
        ]
        self.assertEqual(call_order, expected_task_ids)
        self.assertEqual(max_active_children, 1)
        self.assertEqual(errors, [])
        self.assertEqual(len(finished_results), 1)

        result = finished_results[0]
        self.assertEqual(
            [item["farm_code"] for item in result["completed"]],
            ["3001", "3003"],
        )
        self.assertEqual(
            [item["farm_code"] for item in result["failed"]],
            ["3002"],
        )
        self.assertEqual(result["skipped"], [])
        self.assertIsNone(result["excel_path"])
        self.assertIsNone(result["ppt_path"])
        self.assertEqual(
            done_events,
            [
                (expected_task_ids[0], True),
                (expected_task_ids[1], False),
                (expected_task_ids[2], True),
            ],
        )
        self.assertEqual(
            list(self.parent.glob("reports/牧场组*.xlsx")),
            [existing_report],
        )
        self.assertEqual(
            list(self.parent.glob("reports/牧场组*.pptx")),
            [],
        )
        self.assertEqual(
            (
                existing_report.read_bytes(),
                existing_report.stat().st_mtime_ns,
            ),
            report_before,
        )
        self.assertEqual(
            (
                pointer.read_bytes(),
                pointer.stat().st_mtime_ns,
            ),
            pointer_before,
        )


if __name__ == "__main__":
    unittest.main()

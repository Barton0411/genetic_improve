"""牧场组全量牛只排名分卷导出测试。"""

from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
import uuid
from pathlib import Path
from unittest.mock import patch

import xlsxwriter
from openpyxl import load_workbook

from core.group_report.detail_exporter import (
    EXCEL_MAX_CELL_CHARACTERS,
    LONG_TEXT_CHUNK_CHARACTERS,
    GroupDetailExportPaused,
    GroupCowRankingDetailExporter,
    export_group_cow_ranking_details,
)
import core.group_report.detail_exporter as detail_exporter_module


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _write_index_file(
    child: Path,
    rows,
    *,
    task_id: str,
    farm_code: str,
    score_column: str = "测试权重_index",
) -> None:
    output = (
        child
        / "analysis_results"
        / "processed_index_cow_index_scores.xlsx"
    )
    output.parent.mkdir(parents=True)
    workbook = xlsxwriter.Workbook(
        str(output),
        {"strings_to_formulas": False, "strings_to_urls": False},
    )
    worksheet = workbook.add_worksheet("Sheet1")
    worksheet.write_row(
        0,
        0,
        [
            "cow_id",
            "raw_cow_id",
            "是否在场",
            score_column,
            "备注",
            "牧场编号",
        ],
    )
    for row_index, row in enumerate(rows, start=1):
        worksheet.write_row(row_index, 0, row)
    workbook.close()

    direct_input = (
        child
        / "analysis_results"
        / "processed_cow_data_key_traits_final.xlsx"
    )
    workbook = xlsxwriter.Workbook(str(direct_input))
    worksheet = workbook.add_worksheet("Sheet1")
    worksheet.write_row(0, 0, ["cow_id", "marker"])
    for row_index, row in enumerate(rows, start=1):
        worksheet.write_row(row_index, 0, [row[0], 1])
    workbook.close()

    metadata = {
        "project_type": "group_child",
        "group_task_id": task_id,
        "group_farm_code": farm_code,
        "farms": [{"code": farm_code, "name": f"{farm_code}场"}],
    }
    (child / "project_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False),
        encoding="utf-8",
    )


def _all_rows(volume_entries):
    rows = []
    for volume in sorted(
        (
            entry
            for entry in volume_entries
            if entry["column_part"] == 1
        ),
        key=lambda entry: entry["volume"],
    ):
        workbook = load_workbook(
            volume["absolute_path"],
            read_only=True,
            data_only=False,
        )
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        rows.extend(dict(zip(headers, values)) for values in iterator)
        workbook.close()
    return rows


def _task(task_id: str, farm_code: str, relative_path: str):
    return {
        "task_id": task_id,
        "farm_code": farm_code,
        "farm_name": f"{farm_code}场",
        "relative_path": relative_path,
        "status": "completed",
    }


class GroupCowRankingDetailExporterTests(unittest.TestCase):
    def test_source_columns_are_complete_across_independent_parts(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_id = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "wide",
                [["C1", "RAW-C1", "是", 1.25, "备注值", "010"]],
                task_id=task_id,
                farm_code="010",
            )
            with patch.object(
                detail_exporter_module,
                "DEFAULT_MAX_SOURCE_COLUMNS_PER_PART",
                2,
            ):
                manifest = GroupCowRankingDetailExporter(
                    project,
                    rows_per_volume=10,
                ).export(
                    tasks=[
                        _task(
                            task_id,
                            "010",
                            "farm_projects/wide",
                        )
                    ],
                    output_dir=project / "reports",
                    package_name="column-parts",
                )

            expected_columns = list(manifest["source_columns"])
            expected_part_count = (
                len(expected_columns) + 1
            ) // 2
            for kind in ("ranked", "reconciliation"):
                entries = manifest["volumes"][kind]
                self.assertEqual(
                    {entry["column_parts"] for entry in entries},
                    {expected_part_count},
                )
                self.assertEqual(
                    {entry["column_part"] for entry in entries},
                    set(range(1, expected_part_count + 1)),
                )
                exported_headers = []
                for part_number in range(1, expected_part_count + 1):
                    part_entries = [
                        entry
                        for entry in entries
                        if entry["column_part"] == part_number
                    ]
                    self.assertEqual(
                        sum(entry["data_rows"] for entry in part_entries),
                        1,
                    )
                    workbook = load_workbook(
                        part_entries[0]["absolute_path"],
                        read_only=True,
                        data_only=False,
                    )
                    try:
                        headers = [
                            value
                            for value in next(
                                workbook.active.iter_rows(
                                    values_only=True
                                )
                            )
                        ]
                    finally:
                        workbook.close()
                    exported_headers.extend(
                        headers[
                            len(detail_exporter_module._FIXED_HEADERS) :
                        ]
                    )
                self.assertEqual(exported_headers, expected_columns)

    def test_overlong_text_is_losslessly_chunked_and_reassembled(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_id = str(uuid.uuid4())
            child = project / "farm_projects" / "long"
            _write_index_file(
                child,
                [["C1", "C1", "是", 1, "普通字段", "010"]],
                task_id=task_id,
                farm_code="010",
            )
            original = (
                "开头-"
                + "育种数据🙂" * 12_000
                + "-结尾"
            )
            self.assertGreater(len(original), EXCEL_MAX_CELL_CHARACTERS)
            expected_chunks = (
                len(original) + LONG_TEXT_CHUNK_CHARACTERS - 1
            ) // LONG_TEXT_CHUNK_CHARACTERS
            task = _task(task_id, "010", "farm_projects/long")
            task["farm_name"] = original

            manifest = GroupCowRankingDetailExporter(
                project,
                rows_per_volume=1,
            ).export(
                tasks=[task],
                output_dir=project / "reports",
                package_name="long-text",
            )

            self.assertEqual(manifest["status"], "complete")
            self.assertEqual(
                manifest["counts"]["long_field_count"],
                1,
            )
            self.assertEqual(
                manifest["counts"]["long_field_chunk_count"],
                expected_chunks,
            )
            ranked_rows = _all_rows(manifest["volumes"]["ranked"])
            self.assertIn("超长字段", ranked_rows[0]["牧场名称"])

            chunks = []
            for volume in sorted(
                manifest["volumes"]["long_fields"],
                key=lambda entry: entry["volume"],
            ):
                workbook = load_workbook(
                    volume["absolute_path"],
                    read_only=True,
                    data_only=False,
                )
                try:
                    rows = workbook.active.iter_rows(values_only=True)
                    headers = list(next(rows))
                    for values in rows:
                        row = dict(zip(headers, values))
                        self.assertEqual(
                            row["字段名"],
                            "审计字段:牧场名称",
                        )
                        self.assertEqual(
                            int(row["原始字符数"]),
                            len(original),
                        )
                        self.assertEqual(
                            row["原始SHA-256"],
                            hashlib.sha256(
                                original.encode("utf-8")
                            ).hexdigest(),
                        )
                        chunks.append(
                            (
                                int(row["分块序号"]),
                                int(row["分块总数"]),
                                row["完整内容分块"],
                            )
                        )
                finally:
                    workbook.close()
            self.assertEqual(
                [number for number, _total, _content in chunks],
                list(range(1, expected_chunks + 1)),
            )
            self.assertTrue(
                all(total == expected_chunks for _number, total, _ in chunks)
            )
            rebuilt = "".join(content for _number, _total, content in chunks)
            self.assertEqual(rebuilt, original)
            self.assertEqual(
                hashlib.sha256(rebuilt.encode("utf-8")).hexdigest(),
                hashlib.sha256(original.encode("utf-8")).hexdigest(),
            )

    def test_resource_pause_keeps_committed_farm_and_resumes(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_id = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "a",
                [["A1", "A1", "是", 2.0, "A", "010"]],
                task_id=task_id,
                farm_code="010",
            )
            tasks = [_task(task_id, "010", "farm_projects/a")]

            class PausingExporter(GroupCowRankingDetailExporter):
                def _ensure_free_space(self, path, *, phase, **kwargs):
                    if phase == "建立全局精确排名":
                        raise GroupDetailExportPaused(
                            "模拟磁盘安全暂停",
                            phase=phase,
                            details={"free_bytes": 1},
                        )
                    return super()._ensure_free_space(
                        path,
                        phase=phase,
                        **kwargs,
                    )

            with self.assertRaisesRegex(
                GroupDetailExportPaused,
                "模拟磁盘安全暂停",
            ):
                PausingExporter(project).export(
                    tasks=tasks,
                    output_dir=project / "reports",
                    package_name="disk-resume",
                )

            resume = project / "reports" / ".disk-resume.resume"
            pause_state = json.loads(
                (resume / "pause_state.json").read_text(encoding="utf-8")
            )
            self.assertEqual(pause_state["status"], "paused")
            self.assertEqual(pause_state["phase"], "建立全局精确排名")
            self.assertTrue(
                (resume / ".work" / "ranking.sqlite3").is_file()
            )

            ingested = []

            class CountingExporter(GroupCowRankingDetailExporter):
                def _ingest_source(
                    self,
                    connection,
                    task,
                    task_index,
                    all_source_columns,
                ):
                    ingested.append(task["farm_code"])
                    return super()._ingest_source(
                        connection,
                        task,
                        task_index,
                        all_source_columns,
                    )

            result = CountingExporter(project).export(
                tasks=tasks,
                output_dir=project / "reports",
                package_name="disk-resume",
            )
            self.assertEqual(result["status"], "complete")
            self.assertEqual(ingested, [])
            self.assertFalse(resume.exists())

    def test_duplicate_task_id_cannot_publish_incomplete_source_union(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            duplicated_task_id = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "a",
                [["A1", "A1", "是", 2.0, "A", "010"]],
                task_id=duplicated_task_id,
                farm_code="010",
            )
            _write_index_file(
                project / "farm_projects" / "b",
                [["B1", "B1", "是", 1.0, "B", "020"]],
                task_id=duplicated_task_id,
                farm_code="020",
            )

            manifest = GroupCowRankingDetailExporter(project).export(
                tasks=[
                    _task(
                        duplicated_task_id,
                        "010",
                        "farm_projects/a",
                    ),
                    _task(
                        duplicated_task_id,
                        "020",
                        "farm_projects/b",
                    ),
                ],
                output_dir=project / "reports",
                package_name="duplicate-source-key",
            )

            self.assertEqual(manifest["status"], "partial")
            self.assertTrue(
                all(
                    source["status"] == "invalid_integrity"
                    and "相同来源键" in source["error"]
                    for source in manifest["sources"]
                )
            )

    def test_interrupted_volume_is_verified_and_not_rewritten_on_resume(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_id = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "a",
                [
                    ["A1", "A1", "是", 3.0, "A1", "010"],
                    ["A2", "A2", "是", 2.0, "A2", "010"],
                    ["A3", "A3", "是", 1.0, "A3", "010"],
                ],
                task_id=task_id,
                farm_code="010",
            )
            tasks = [_task(task_id, "010", "farm_projects/a")]
            real_atomic_write = detail_exporter_module._write_json_atomic
            interrupted = False

            def interrupt_after_first_volume(path, payload):
                nonlocal interrupted
                real_atomic_write(path, payload)
                if (
                    not interrupted
                    and Path(path).name == "export_checkpoint.json"
                    and len(payload.get("volumes", [])) == 1
                ):
                    interrupted = True
                    raise RuntimeError("模拟分卷后退出")

            with patch.object(
                detail_exporter_module,
                "_write_json_atomic",
                side_effect=interrupt_after_first_volume,
            ):
                with self.assertRaisesRegex(RuntimeError, "模拟分卷后退出"):
                    GroupCowRankingDetailExporter(
                        project,
                        rows_per_volume=1,
                    ).export(
                        tasks=tasks,
                        output_dir=project / "reports",
                        package_name="volume-resume",
                    )

            staging = project / "reports" / ".volume-resume.resume"
            first_volume = (
                staging / "有效在群完整排名_第0001卷.xlsx"
            )
            self.assertTrue(first_volume.is_file())
            first_sha = _sha256(first_volume)
            first_mtime = first_volume.stat().st_mtime_ns

            manifest = GroupCowRankingDetailExporter(
                project,
                rows_per_volume=1,
            ).export(
                tasks=tasks,
                output_dir=project / "reports",
                package_name="volume-resume",
            )
            published_first = (
                Path(manifest["package_path"])
                / "有效在群完整排名_第0001卷.xlsx"
            )
            self.assertEqual(_sha256(published_first), first_sha)
            self.assertEqual(published_first.stat().st_mtime_ns, first_mtime)
            self.assertEqual(
                manifest["counts"]["ranked_exported_rows"],
                3,
            )

    def test_interrupted_export_reuses_committed_sources_and_volumes(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_a = str(uuid.uuid4())
            task_b = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "a",
                [["A1", "A1", "是", 2.0, "A", "010"]],
                task_id=task_a,
                farm_code="010",
            )
            _write_index_file(
                project / "farm_projects" / "b",
                [["B1", "B1", "是", 1.0, "B", "020"]],
                task_id=task_b,
                farm_code="020",
            )
            tasks = [
                _task(task_a, "010", "farm_projects/a"),
                _task(task_b, "020", "farm_projects/b"),
            ]

            def interrupt_after_first_source(_value, message):
                if "已读取 1/2 个牧场" in message:
                    raise RuntimeError("模拟程序退出")

            with self.assertRaisesRegex(RuntimeError, "模拟程序退出"):
                GroupCowRankingDetailExporter(
                    project,
                    rows_per_volume=1,
                    progress_callback=interrupt_after_first_source,
                ).export(
                    tasks=tasks,
                    output_dir=project / "reports",
                    package_name="resume-test",
                )
            resume_dir = project / "reports" / ".resume-test.resume"
            self.assertTrue(
                (resume_dir / ".work" / "ranking.sqlite3").is_file()
            )

            ingested = []

            class CountingExporter(GroupCowRankingDetailExporter):
                def _ingest_source(
                    self,
                    connection,
                    task,
                    task_index,
                    all_source_columns,
                ):
                    ingested.append(task["farm_code"])
                    return super()._ingest_source(
                        connection,
                        task,
                        task_index,
                        all_source_columns,
                    )

            manifest = CountingExporter(
                project,
                rows_per_volume=1,
            ).export(
                tasks=tasks,
                output_dir=project / "reports",
                package_name="resume-test",
            )
            self.assertEqual(ingested, ["020"])
            self.assertTrue(
                manifest["sources"][0]["resumed_from_checkpoint"]
            )
            self.assertFalse(
                manifest["sources"][1]["resumed_from_checkpoint"]
            )
            self.assertEqual(manifest["counts"]["source_rows"], 2)
            self.assertFalse(resume_dir.exists())

    def test_corrupt_unpublished_sqlite_is_quarantined_and_rebuilt(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_id = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "a",
                [["A1", "A1", "是", 1.0, "A", "010"]],
                task_id=task_id,
                farm_code="010",
            )
            tasks = [_task(task_id, "010", "farm_projects/a")]

            def interrupt(_value, message):
                if "已读取 1/1 个牧场" in message:
                    raise RuntimeError("模拟建立索引后退出")

            with self.assertRaisesRegex(RuntimeError, "模拟建立索引后退出"):
                GroupCowRankingDetailExporter(
                    project,
                    progress_callback=interrupt,
                ).export(
                    tasks=tasks,
                    output_dir=project / "reports",
                    package_name="corrupt-resume",
                )

            work = (
                project
                / "reports"
                / ".corrupt-resume.resume"
                / ".work"
            )
            database = work / "ranking.sqlite3"
            database.write_bytes(b"not-a-sqlite-database")
            Path(f"{database}-wal").unlink(missing_ok=True)
            Path(f"{database}-shm").unlink(missing_ok=True)

            class StopAfterRecovery(GroupCowRankingDetailExporter):
                @staticmethod
                def _build_ranks(connection):
                    raise RuntimeError("检查隔离记录")

            with self.assertRaisesRegex(RuntimeError, "检查隔离记录"):
                StopAfterRecovery(project).export(
                    tasks=tasks,
                    output_dir=project / "reports",
                    package_name="corrupt-resume",
                )
            history = work / "corrupt_history"
            self.assertTrue(list(history.glob("*_recovery.json")))

            result = GroupCowRankingDetailExporter(project).export(
                tasks=tasks,
                output_dir=project / "reports",
                package_name="corrupt-resume",
            )
            self.assertEqual(result["status"], "complete")

    def test_full_rows_are_ranked_and_reconciled_without_top_n_truncation(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            project = root / "group"
            project.mkdir()
            child_a = project / "farm_projects" / "010_A场"
            child_b = project / "farm_projects" / "020_B场"
            task_a = str(uuid.uuid4())
            task_b = str(uuid.uuid4())
            _write_index_file(
                child_a,
                [
                    ["0002", "A-0002", "是", 10.123456789, "同分第二", "010"],
                    ["0001", "A-0001", "是", 10.123456789, "=不执行公式", "010"],
                    ["0003", "A-0003", "否", 999.0, "非在群", "010"],
                    ["", "A-empty", "是", 88.0, "牛号空", "010"],
                    ["0005", "A-0005", "是", "bad", "坏指数", "010"],
                ],
                task_id=task_a,
                farm_code="010",
            )
            _write_index_file(
                child_b,
                [
                    ["0001", "B-0001", "是", 10.123456790, "未舍入第一", "020"],
                    [6.0, "B-0006", "是", 5.0, "整数牛号", "020"],
                    ["0007", "B-0007", "", 4.0, "状态空", "020"],
                ],
                task_id=task_b,
                farm_code="020",
            )
            tasks = [
                {
                    "task_id": task_a,
                    "farm_code": "010",
                    "farm_name": "A场",
                    "relative_path": "farm_projects/010_A场",
                    "status": "completed",
                },
                {
                    "task_id": task_b,
                    "farm_code": "020",
                    "farm_name": "B场",
                    "relative_path": "farm_projects/020_B场",
                    "status": "completed",
                },
            ]

            manifest = export_group_cow_ranking_details(
                project,
                tasks=tasks,
                output_dir=project / "reports",
                package_name="acceptance",
                rows_per_volume=2,
            )

            self.assertEqual(manifest["status"], "complete")
            self.assertEqual(manifest["counts"]["source_rows"], 8)
            self.assertEqual(manifest["counts"]["valid_ranked_rows"], 4)
            self.assertEqual(manifest["counts"]["unranked_rows"], 4)
            self.assertEqual(manifest["counts"]["ranked_exported_rows"], 4)
            self.assertEqual(
                manifest["counts"]["reconciliation_exported_rows"],
                8,
            )
            self.assertEqual(
                manifest["counts"]["unranked_reason_counts"],
                {
                    "牛号为空": 1,
                    "综合指数不是数值": 1,
                    "是否在场为空": 1,
                    "非在群母牛": 1,
                },
            )

            ranked_volumes = manifest["volumes"]["ranked"]
            reconciliation_volumes = manifest["volumes"]["reconciliation"]
            self.assertEqual(
                [entry["data_rows"] for entry in ranked_volumes],
                [2, 2],
            )
            self.assertEqual(
                [entry["data_rows"] for entry in reconciliation_volumes],
                [2, 2, 2, 2],
            )

            ranked_rows = _all_rows(ranked_volumes)
            self.assertEqual(
                [row["牧场组排名"] for row in ranked_rows],
                [1, 2, 3, 4],
            )
            self.assertEqual(
                [(row["牧场编号"], row["cow_id"]) for row in ranked_rows],
                [
                    ("020", "0001"),
                    ("010", "0001"),
                    ("010", "0002"),
                    ("020", "6"),
                ],
            )
            self.assertGreater(
                ranked_rows[0]["综合指数_未舍入"],
                ranked_rows[1]["综合指数_未舍入"],
            )
            self.assertEqual(ranked_rows[1]["备注"], "=不执行公式")

            reconciliation_rows = _all_rows(reconciliation_volumes)
            self.assertEqual(len(reconciliation_rows), 8)
            self.assertEqual(
                sum(
                    row["分类结果"] == "有效在群排名"
                    for row in reconciliation_rows
                ),
                4,
            )
            self.assertEqual(
                {row["未排名原因"] for row in reconciliation_rows if row["未排名原因"]},
                {"牛号为空", "综合指数不是数值", "是否在场为空", "非在群母牛"},
            )
            self.assertIn("备注", manifest["source_columns"])
            self.assertEqual(ranked_rows[0]["API farmcode"], "020")
            self.assertIn("牧场编号", ranked_rows[0])

            # 所有标识符都以文本写入，包含前导零，不发生科学计数法或 .0。
            first_ranked_path = Path(ranked_volumes[0]["absolute_path"])
            workbook = load_workbook(first_ranked_path, data_only=False)
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]
            farm_column = headers.index("牧场编号") + 1
            cow_column = headers.index("cow_id") + 1
            self.assertEqual(
                worksheet.cell(2, farm_column).data_type,
                "s",
            )
            self.assertEqual(
                worksheet.cell(2, cow_column).data_type,
                "s",
            )
            workbook.close()

            for kind in ("ranked", "reconciliation"):
                for volume in manifest["volumes"][kind]:
                    path = Path(volume["absolute_path"])
                    self.assertTrue(path.is_file())
                    self.assertEqual(volume["sha256"], _sha256(path))
                    self.assertGreater(volume["bytes"], 0)
            manifest_path = Path(manifest["manifest_path"])
            self.assertEqual(manifest["manifest_sha256"], _sha256(manifest_path))
            on_disk = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(on_disk["counts"]["source_rows"], 8)

    def test_exact_decimal_ranking_and_hmy_farm_number_are_not_conflated(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            child = project / "farm_projects" / "hmy"
            task_id = str(uuid.uuid4())
            api_farmcode = "1100110073"
            _write_index_file(
                child,
                [
                    [
                        "C1",
                        "C1",
                        "是",
                        "1.0000000000000000001",
                        "较小",
                        "0102026",
                    ],
                    [
                        "C2",
                        "C2",
                        "是",
                        "1.0000000000000000002",
                        "较大",
                        "0102026",
                    ],
                ],
                task_id=task_id,
                farm_code=api_farmcode,
            )
            manifest = GroupCowRankingDetailExporter(
                project,
                rows_per_volume=10,
            ).export(
                tasks=[
                    _task(
                        task_id,
                        api_farmcode,
                        "farm_projects/hmy",
                    )
                ],
                output_dir=project / "reports",
                package_name="exact-hmy",
            )

            self.assertEqual(manifest["status"], "complete")
            rows = _all_rows(manifest["volumes"]["ranked"])
            self.assertEqual([row["cow_id"] for row in rows], ["C2", "C1"])
            self.assertEqual(
                [row["API farmcode"] for row in rows],
                [api_farmcode, api_farmcode],
            )
            self.assertEqual(
                [row["牧场编号"] for row in rows],
                ["0102026", "0102026"],
            )
            self.assertNotEqual(
                rows[0]["综合指数_精确文本"],
                rows[1]["综合指数_精确文本"],
            )

    def test_missing_source_produces_auditable_partial_empty_package(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            tasks = [
                {
                    "farm_code": "0099",
                    "farm_name": "缺文件牧场",
                    "relative_path": "farm_projects/missing",
                    "status": "completed",
                }
            ]
            manifest = GroupCowRankingDetailExporter(
                project,
                rows_per_volume=10,
            ).export(
                tasks=tasks,
                output_dir=project / "reports",
                package_name="partial",
            )

            self.assertEqual(manifest["status"], "partial")
            self.assertEqual(manifest["counts"]["source_rows"], 0)
            self.assertEqual(manifest["counts"]["valid_ranked_rows"], 0)
            self.assertEqual(
                manifest["sources"][0]["status"],
                "missing",
            )
            self.assertEqual(
                manifest["volumes"]["ranked"][0]["data_rows"],
                0,
            )
            self.assertEqual(
                manifest["volumes"]["reconciliation"][0]["data_rows"],
                0,
            )

    def test_invalid_volume_threshold_is_rejected_before_any_write(self):
        with self.assertRaisesRegex(ValueError, "每卷数据行数"):
            GroupCowRankingDetailExporter(Path("."), rows_per_volume=0)
        with self.assertRaisesRegex(ValueError, "每卷数据行数"):
            GroupCowRankingDetailExporter(
                Path("."),
                rows_per_volume=1_048_576,
            )

    def test_duplicate_cow_id_and_upstream_mismatch_block_formal_complete(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            child = project / "farm_projects" / "dup"
            task_id = str(uuid.uuid4())
            _write_index_file(
                child,
                [
                    ["001", "001", "是", 1.0, "第一条", "010"],
                    ["001", "001", "是", 2.0, "重复条", "010"],
                ],
                task_id=task_id,
                farm_code="010",
            )
            manifest = export_group_cow_ranking_details(
                project,
                tasks=[_task(task_id, "010", "farm_projects/dup")],
                output_dir=project / "reports",
                package_name="duplicate",
                rows_per_volume=10,
            )
            self.assertEqual(manifest["status"], "partial")
            self.assertEqual(
                manifest["sources"][0]["status"],
                "invalid_integrity",
            )
            self.assertEqual(
                manifest["sources"][0]["duplicate_cow_id_count"],
                1,
            )

            # 改写直接输入为一个牛号，模拟合法但被截断/陈旧的指数结果。
            direct = (
                child
                / "analysis_results"
                / "processed_cow_data_key_traits_final.xlsx"
            )
            workbook = xlsxwriter.Workbook(str(direct))
            worksheet = workbook.add_worksheet("Sheet1")
            worksheet.write_row(0, 0, ["cow_id", "marker"])
            worksheet.write_row(1, 0, ["001", 1])
            workbook.close()
            second = export_group_cow_ranking_details(
                project,
                tasks=[_task(task_id, "010", "farm_projects/dup")],
                output_dir=project / "reports",
                package_name="mismatch",
                rows_per_volume=10,
            )
            self.assertFalse(second["sources"][0]["identity_match"])
            self.assertIn(
                "多重集不一致",
                second["sources"][0]["error"],
            )

    def test_sentinel_ids_are_missing_and_score_columns_must_match(self):
        with tempfile.TemporaryDirectory() as temporary_dir:
            project = Path(temporary_dir) / "group"
            project.mkdir()
            task_a = str(uuid.uuid4())
            task_b = str(uuid.uuid4())
            _write_index_file(
                project / "farm_projects" / "a",
                [
                    ["nan", "nan", "是", 3.0, "缺号", "010"],
                    ["A1", "A1", "是", 2.0, "有效", "010"],
                ],
                task_id=task_a,
                farm_code="010",
                score_column="口径A_index",
            )
            _write_index_file(
                project / "farm_projects" / "b",
                [["B1", "B1", "是", 1.0, "有效", "020"]],
                task_id=task_b,
                farm_code="020",
                score_column="口径B_index",
            )
            manifest = export_group_cow_ranking_details(
                project,
                tasks=[
                    _task(task_a, "010", "farm_projects/a"),
                    _task(task_b, "020", "farm_projects/b"),
                ],
                output_dir=project / "reports",
                package_name="score-mismatch",
                rows_per_volume=10,
            )
            self.assertEqual(manifest["status"], "partial")
            self.assertEqual(manifest["counts"]["valid_ranked_rows"], 2)
            self.assertEqual(
                manifest["counts"]["unranked_reason_counts"]["牛号为空"],
                1,
            )
            self.assertTrue(
                all(
                    "指数列口径不一致" in source["error"]
                    for source in manifest["sources"]
                )
            )


if __name__ == "__main__":
    unittest.main()

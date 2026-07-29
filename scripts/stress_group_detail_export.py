#!/usr/bin/env python3
"""可选运行的牧场组全量 Excel 分卷压力测试。

默认生成 2 个模拟牧场、每场 525,000 行，合计 1,050,000 行，刚好超过
单个 Excel 工作表可容纳的 1,048,575 条数据行。该脚本不会被普通单元
测试发现或执行。

完整验收::

    python scripts/stress_group_detail_export.py \
        --require-cross-limit \
        --max-peak-rss-mib 2048 \
        --max-workspace-peak-gib 20 \
        --keep

快速冒烟::

    python scripts/stress_group_detail_export.py \
        --rows-per-farm 100 --rows-per-volume 75

完整验收只执行一次百万行主导出，随后另建一个很小的恢复场景，模拟首卷
提交后进程异常，再确认续跑复用了来源检查点和已提交分卷。RSS 与工作目录
磁盘峰值由后台采样器记录；阈值可通过命令行调整，设为 0 可仅记录不判定。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import sys
import tempfile
import threading
import time
import uuid
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple

try:
    import resource
except ImportError:  # pragma: no cover - Windows
    resource = None

import xlsxwriter
from openpyxl import load_workbook


REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
if str(REPOSITORY_ROOT) not in sys.path:
    sys.path.insert(0, str(REPOSITORY_ROOT))

from core.group_report.detail_exporter import (  # noqa: E402
    EXCEL_MAX_ROWS,
    GroupCowRankingDetailExporter,
)


INDEX_FILENAME = "processed_index_cow_index_scores.xlsx"
DIRECT_INPUT_FILENAME = "processed_cow_data_key_traits_final.xlsx"
SCORE_COLUMN = "压力测试权重_index"
DEFAULT_MAX_PEAK_RSS_MIB = 2048.0
DEFAULT_MAX_WORKSPACE_PEAK_GIB = 20.0
DEFAULT_RESOURCE_SAMPLE_INTERVAL = 0.25


def _process_peak_rss_bytes() -> int:
    """返回当前进程生命周期峰值 RSS，兼容 macOS/Linux。"""
    if resource is not None:
        peak = int(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
        # macOS 返回 bytes，Linux/BSD 通常返回 KiB。
        return peak if sys.platform == "darwin" else peak * 1024
    if sys.platform == "win32":  # pragma: no cover - Windows
        try:
            import ctypes
            from ctypes import wintypes

            class ProcessMemoryCounters(ctypes.Structure):
                _fields_ = [
                    ("cb", wintypes.DWORD),
                    ("PageFaultCount", wintypes.DWORD),
                    ("PeakWorkingSetSize", ctypes.c_size_t),
                    ("WorkingSetSize", ctypes.c_size_t),
                    ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                    ("PagefileUsage", ctypes.c_size_t),
                    ("PeakPagefileUsage", ctypes.c_size_t),
                ]

            counters = ProcessMemoryCounters()
            counters.cb = ctypes.sizeof(counters)
            process = ctypes.windll.kernel32.GetCurrentProcess()
            ok = ctypes.windll.psapi.GetProcessMemoryInfo(
                process,
                ctypes.byref(counters),
                counters.cb,
            )
            return int(counters.PeakWorkingSetSize) if ok else 0
        except (AttributeError, OSError):
            return 0
    return 0


def _workspace_usage_bytes(root: Path) -> Tuple[int, int]:
    """返回工作目录逻辑字节与实际分配字节，不跟随符号链接。"""
    logical = 0
    allocated = 0
    pending = [Path(root)]
    while pending:
        directory = pending.pop()
        try:
            entries = list(os.scandir(directory))
        except (FileNotFoundError, NotADirectoryError, PermissionError):
            continue
        for entry in entries:
            try:
                if entry.is_symlink():
                    continue
                if entry.is_dir(follow_symlinks=False):
                    pending.append(Path(entry.path))
                    continue
                if not entry.is_file(follow_symlinks=False):
                    continue
                stat = entry.stat(follow_symlinks=False)
                logical += int(stat.st_size)
                allocated += int(
                    getattr(stat, "st_blocks", 0) * 512
                    or stat.st_size
                )
            except (FileNotFoundError, PermissionError, OSError):
                continue
    return logical, allocated


class ResourceMonitor:
    """低频采样进程 RSS 与工作目录占用，记录整个验收期间峰值。"""

    def __init__(self, root: Path, *, interval_seconds: float = 0.25):
        interval = float(interval_seconds)
        if not math.isfinite(interval) or interval <= 0:
            raise ValueError("资源采样间隔必须是大于 0 的有限数值")
        self.root = Path(root)
        self.interval_seconds = interval
        self.peak_rss_bytes = 0
        self.peak_workspace_logical_bytes = 0
        self.peak_workspace_allocated_bytes = 0
        self.samples = 0
        self._stop = threading.Event()
        self._thread: Optional[threading.Thread] = None

    def sample(self) -> None:
        logical, allocated = _workspace_usage_bytes(self.root)
        self.peak_rss_bytes = max(
            self.peak_rss_bytes,
            _process_peak_rss_bytes(),
        )
        self.peak_workspace_logical_bytes = max(
            self.peak_workspace_logical_bytes,
            logical,
        )
        self.peak_workspace_allocated_bytes = max(
            self.peak_workspace_allocated_bytes,
            allocated,
        )
        self.samples += 1

    def _run(self) -> None:
        while not self._stop.wait(self.interval_seconds):
            self.sample()

    def start(self) -> "ResourceMonitor":
        if self._thread is not None:
            raise RuntimeError("资源监控器不能重复启动")
        self.sample()
        self._thread = threading.Thread(
            target=self._run,
            name="group-detail-stress-resource-monitor",
            daemon=True,
        )
        self._thread.start()
        return self

    def stop(self) -> Dict:
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=max(1.0, self.interval_seconds * 4))
        self.sample()
        return {
            "peak_rss_bytes": int(self.peak_rss_bytes),
            "peak_rss_mib": round(
                self.peak_rss_bytes / (1024 ** 2),
                3,
            ),
            "workspace_peak_logical_bytes": int(
                self.peak_workspace_logical_bytes
            ),
            "workspace_peak_logical_gib": round(
                self.peak_workspace_logical_bytes / (1024 ** 3),
                3,
            ),
            "workspace_peak_allocated_bytes": int(
                self.peak_workspace_allocated_bytes
            ),
            "workspace_peak_allocated_gib": round(
                self.peak_workspace_allocated_bytes / (1024 ** 3),
                3,
            ),
            "sampling_interval_seconds": self.interval_seconds,
            "samples": int(self.samples),
            "disk_peak_is_sampled": True,
        }


def _sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            chunk = stream.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _assert_xlsx_zip(path: Path) -> None:
    if not path.is_file() or path.stat().st_size <= 0:
        raise AssertionError(f"Excel 文件不存在或为空：{path}")
    with zipfile.ZipFile(path, "r") as archive:
        required = {"[Content_Types].xml", "xl/workbook.xml"}
        missing = required - set(archive.namelist())
        if missing:
            raise AssertionError(
                f"Excel ZIP 结构不完整：{path.name} 缺少 {sorted(missing)}"
            )
        bad_member = archive.testzip()
        if bad_member is not None:
            raise AssertionError(
                f"Excel ZIP CRC 错误：{path.name} / {bad_member}"
            )


def _new_workbook(path: Path) -> Tuple[xlsxwriter.Workbook, object]:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(
        str(path),
        {
            "constant_memory": True,
            "use_zip64": True,
            "strings_to_formulas": False,
            "strings_to_urls": False,
        },
    )
    workbook.use_zip64()
    return workbook, workbook.add_worksheet("Sheet1")


def _write_farm_inputs(
    child_path: Path,
    *,
    farm_index: int,
    rows: int,
    task_id: str,
) -> Dict:
    """一次循环同时写最小指数结果和指数直接输入。"""

    farm_code = f"STRESS-{farm_index:03d}"
    farm_name = f"压力测试牧场{farm_index:03d}"
    analysis_dir = child_path / "analysis_results"
    index_path = analysis_dir / INDEX_FILENAME
    direct_path = analysis_dir / DIRECT_INPUT_FILENAME
    index_workbook, index_sheet = _new_workbook(index_path)
    direct_workbook, direct_sheet = _new_workbook(direct_path)
    index_sheet.write_row(
        0,
        0,
        ["cow_id", "是否在场", SCORE_COLUMN, "牧场编号"],
    )
    direct_sheet.write_row(0, 0, ["cow_id"])

    try:
        for offset in range(rows):
            cow_id = f"{farm_index:03d}-{offset + 1:09d}"
            # 不舍入的唯一分值，排名输出应形成 1..N 的完整排列。
            score = farm_index * 1_000_000_000 + rows - offset
            excel_row = offset + 1
            index_sheet.write_row(
                excel_row,
                0,
                [cow_id, "是", score, farm_code],
            )
            direct_sheet.write(excel_row, 0, cow_id)
    finally:
        index_workbook.close()
        direct_workbook.close()

    metadata = {
        "project_type": "group_child",
        "group_task_id": task_id,
        "group_farm_code": farm_code,
        "farms": [{"code": farm_code, "name": farm_name}],
    }
    (child_path / "project_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False),
        encoding="utf-8",
    )
    _assert_xlsx_zip(index_path)
    _assert_xlsx_zip(direct_path)
    return {
        "task_id": task_id,
        "farm_code": farm_code,
        "farm_name": farm_name,
        "relative_path": child_path.relative_to(
            child_path.parents[1]
        ).as_posix(),
        "status": "completed",
        "_index_path": index_path,
        "_direct_path": direct_path,
    }


def _expected_permutation(total_rows: int) -> Tuple[int, int, int]:
    expected_sum = total_rows * (total_rows + 1) // 2
    expected_square_sum = (
        total_rows * (total_rows + 1) * (2 * total_rows + 1) // 6
    )
    remainder = total_rows % 4
    expected_xor = (
        total_rows
        if remainder == 0
        else 1
        if remainder == 1
        else total_rows + 1
        if remainder == 2
        else 0
    )
    return expected_sum, expected_square_sum, expected_xor


def _canonical_decimal_text(value) -> str:
    try:
        decimal = Decimal(str(value or "0"))
    except (InvalidOperation, ValueError) as exc:
        raise AssertionError(f"精确指数不是有效十进制数：{value}") from exc
    return format(decimal, "f")


def _verify_volume_rows(
    volume: Dict,
    *,
    package_path: Path,
    expected_next_rank: Optional[int],
    expected_identity_at: Callable[[int], Dict],
    identity_start_index: int,
) -> Tuple[int, Optional[int], Tuple[int, int, int], int, int]:
    """流式读回一卷，逐行核对身份、顺序、排名和 Excel 上限。"""

    path = package_path / volume["path"]
    _assert_xlsx_zip(path)
    actual_sha = _sha256(path)
    if actual_sha != volume["sha256"]:
        raise AssertionError(f"分卷 SHA-256 不一致：{path.name}")
    if path.stat().st_size != int(volume["bytes"]):
        raise AssertionError(f"分卷文件大小不一致：{path.name}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise AssertionError(f"分卷没有表头：{path.name}") from exc
        rank_position = headers.index("牧场组排名")
        class_position = headers.index("分类结果")
        farm_position = headers.index("API farmcode")
        cow_position = headers.index("cow_id")
        exact_score_position = headers.index("综合指数_精确文本")
        data_rows = 0
        rank_sum = 0
        rank_square_sum = 0
        rank_xor = 0
        next_rank = expected_next_rank
        first_observed_rank: Optional[int] = None
        last_observed_rank: Optional[int] = None
        for values in rows:
            if not any(value is not None for value in values):
                continue
            data_rows += 1
            if values[class_position] != "有效在群排名":
                raise AssertionError(
                    f"{path.name} 第 {data_rows + 1} 行分类异常"
                )
            rank = values[rank_position]
            if isinstance(rank, bool) or not isinstance(rank, (int, float)):
                raise AssertionError(
                    f"{path.name} 第 {data_rows + 1} 行排名不是数值"
                )
            integer_rank = int(rank)
            if integer_rank != rank:
                raise AssertionError(
                    f"{path.name} 第 {data_rows + 1} 行排名不是整数"
                )
            expected_identity = expected_identity_at(
                identity_start_index + data_rows - 1
            )
            if first_observed_rank is None:
                first_observed_rank = integer_rank
            last_observed_rank = integer_rank
            actual_identity = {
                "rank": integer_rank,
                "farm_code": str(values[farm_position] or ""),
                "cow_id": str(values[cow_position] or ""),
                "score_exact": _canonical_decimal_text(
                    values[exact_score_position]
                ),
            }
            if actual_identity != expected_identity:
                raise AssertionError(
                    f"{path.name} 第 {data_rows + 1} 行身份或排序异常："
                    f"期望 {expected_identity}，实际 {actual_identity}"
                )
            if next_rank is not None:
                if integer_rank != next_rank:
                    raise AssertionError(
                        f"{path.name} 排名断裂：期望 {next_rank}，"
                        f"实际 {integer_rank}"
                    )
                next_rank += 1
            rank_sum += integer_rank
            rank_square_sum += integer_rank * integer_rank
            rank_xor ^= integer_rank
    finally:
        workbook.close()

    if data_rows != int(volume["data_rows"]):
        raise AssertionError(
            f"{path.name} 实际 {data_rows:,} 行，manifest 为 "
            f"{int(volume['data_rows']):,} 行"
        )
    if data_rows > int(volume["rows_per_volume"]):
        raise AssertionError(f"{path.name} 超过配置的每卷数据行上限")
    if data_rows <= 0:
        raise AssertionError(f"{path.name} 是没有明细的空分卷")
    if data_rows + 1 > EXCEL_MAX_ROWS:
        raise AssertionError(f"{path.name} 超过 Excel 工作表行上限")
    if int(volume.get("first_rank") or -1) != first_observed_rank:
        raise AssertionError(f"{path.name} manifest 首排名不一致")
    if int(volume.get("last_rank") or -1) != last_observed_rank:
        raise AssertionError(f"{path.name} manifest 末排名不一致")
    assert first_observed_rank is not None
    assert last_observed_rank is not None
    return (
        data_rows,
        next_rank,
        (rank_sum, rank_square_sum, rank_xor),
        first_observed_rank,
        last_observed_rank,
    )


def _verify_volume_group(
    volumes: Iterable[Dict],
    *,
    package_path: Path,
    expected_rows: int,
    rows_per_volume: int,
    require_sorted_ranks: bool,
    expected_identity_at: Callable[[int], Dict],
) -> Dict:
    entries = sorted(
        (
            dict(volume)
            for volume in volumes
            if int(volume.get("column_part", 0)) == 1
        ),
        key=lambda volume: int(volume["volume"]),
    )
    minimum_volume_count = max(1, math.ceil(expected_rows / rows_per_volume))
    if not minimum_volume_count <= len(entries) <= max(1, expected_rows):
        raise AssertionError(
            "分卷数不正确：至少应为 "
            f"{minimum_volume_count}，实际 {len(entries)}"
        )
    total = 0
    next_rank = 1 if require_sorted_ranks else None
    rank_sum = 0
    rank_square_sum = 0
    rank_xor = 0
    for expected_number, entry in enumerate(entries, start=1):
        if int(entry["volume"]) != expected_number:
            raise AssertionError("分卷编号不连续")
        if int(entry["rows_per_volume"]) != rows_per_volume:
            raise AssertionError("manifest 每卷行数配置不一致")
        (
            actual_rows,
            next_rank,
            permutation,
            first_rank,
            last_rank,
        ) = _verify_volume_rows(
            entry,
            package_path=package_path,
            expected_next_rank=next_rank,
            expected_identity_at=expected_identity_at,
            identity_start_index=total,
        )
        if require_sorted_ranks:
            expected_first = total + 1
            expected_last = total + actual_rows
            if (first_rank, last_rank) != (expected_first, expected_last):
                raise AssertionError(
                    f"第 {expected_number} 卷排名边界不连续："
                    f"期望 {expected_first}..{expected_last}，"
                    f"实际 {first_rank}..{last_rank}"
                )
        total += actual_rows
        rank_sum += permutation[0]
        rank_square_sum += permutation[1]
        rank_xor ^= permutation[2]
    if total != expected_rows:
        raise AssertionError(
            f"分卷合计行数不正确：期望 {expected_rows:,}，实际 {total:,}"
        )
    if require_sorted_ranks and next_rank != expected_rows + 1:
        raise AssertionError("完整排名没有连续覆盖 1..N")
    if (rank_sum, rank_square_sum, rank_xor) != _expected_permutation(
        expected_rows
    ):
        raise AssertionError("分卷中的排名不是 1..N 的完整排列")
    return {
        "volumes": len(entries),
        "rows": total,
        "identity_and_order_verified_rows": total,
        "volume_boundaries_verified": True,
    }


def _expected_ranked_identity(
    position: int,
    *,
    farms: int,
    rows_per_farm: int,
) -> Dict:
    """压力数据按分值降序时，第 ``position`` 行的唯一预期身份。"""
    farm_block, offset = divmod(position, rows_per_farm)
    farm_index = farms - farm_block
    rank = position + 1
    score = farm_index * 1_000_000_000 + rows_per_farm - offset
    return {
        "rank": rank,
        "farm_code": f"STRESS-{farm_index:03d}",
        "cow_id": f"{farm_index:03d}-{offset + 1:09d}",
        "score_exact": str(score),
    }


def _expected_reconciliation_identity(
    position: int,
    *,
    farms: int,
    rows_per_farm: int,
) -> Dict:
    """来源对账按牧场、源行写入时，第 ``position`` 行的预期身份。"""
    farm_block, offset = divmod(position, rows_per_farm)
    farm_index = farm_block + 1
    rank = (farms - farm_index) * rows_per_farm + offset + 1
    score = farm_index * 1_000_000_000 + rows_per_farm - offset
    return {
        "rank": rank,
        "farm_code": f"STRESS-{farm_index:03d}",
        "cow_id": f"{farm_index:03d}-{offset + 1:09d}",
        "score_exact": str(score),
    }


def _verify_manifest(
    manifest: Dict,
    *,
    tasks: List[Dict],
    rows_per_farm: int,
    rows_per_volume: int,
) -> Dict:
    total_rows = len(tasks) * rows_per_farm
    counts = manifest["counts"]
    expected_counts = {
        "tasks_in_scope": len(tasks),
        "source_files_read": len(tasks),
        "source_files_with_problem": 0,
        "source_rows": total_rows,
        "valid_ranked_rows": total_rows,
        "unranked_rows": 0,
        "ranked_exported_rows": total_rows,
        "reconciliation_exported_rows": total_rows,
    }
    if manifest.get("status") != "complete":
        raise AssertionError(f"导出状态不是 complete：{manifest.get('status')}")
    for key, expected in expected_counts.items():
        actual = int(counts.get(key, -1))
        if actual != expected:
            raise AssertionError(
                f"{key} 对账失败：期望 {expected:,}，实际 {actual:,}"
            )

    task_by_id = {task["task_id"]: task for task in tasks}
    if len(manifest["sources"]) != len(tasks):
        raise AssertionError("manifest 来源数量不等于牧场数量")
    for source in manifest["sources"]:
        task = task_by_id.get(source["task_id"])
        if task is None:
            raise AssertionError("manifest 出现未知 task_id")
        if source["status"] != "read":
            raise AssertionError(
                f"{source['farm_code']} 来源状态异常：{source['status']}"
            )
        if int(source["rows_read"]) != rows_per_farm:
            raise AssertionError(f"{source['farm_code']} 来源行数不正确")
        if not source["identity_match"]:
            raise AssertionError(f"{source['farm_code']} 输入身份对账失败")
        if source["sha256"] != _sha256(task["_index_path"]):
            raise AssertionError(f"{source['farm_code']} 指数源 SHA 不一致")
        if source["direct_input"]["sha256"] != _sha256(
            task["_direct_path"]
        ):
            raise AssertionError(f"{source['farm_code']} 直接输入 SHA 不一致")

    manifest_path = Path(manifest["manifest_path"])
    if manifest["manifest_sha256"] != _sha256(manifest_path):
        raise AssertionError("manifest SHA-256 不一致")
    on_disk = json.loads(manifest_path.read_text(encoding="utf-8"))
    if on_disk["counts"] != counts:
        raise AssertionError("磁盘 manifest 与返回结果计数不一致")

    package_path = Path(manifest["package_path"])
    farm_count = len(tasks)
    ranked = _verify_volume_group(
        manifest["volumes"]["ranked"],
        package_path=package_path,
        expected_rows=total_rows,
        rows_per_volume=rows_per_volume,
        require_sorted_ranks=True,
        expected_identity_at=lambda position: _expected_ranked_identity(
            position,
            farms=farm_count,
            rows_per_farm=rows_per_farm,
        ),
    )
    reconciliation = _verify_volume_group(
        manifest["volumes"]["reconciliation"],
        package_path=package_path,
        expected_rows=total_rows,
        rows_per_volume=rows_per_volume,
        require_sorted_ranks=False,
        expected_identity_at=lambda position: (
            _expected_reconciliation_identity(
                position,
                farms=farm_count,
                rows_per_farm=rows_per_farm,
            )
        ),
    )
    return {
        "total_rows": total_rows,
        "ranked": ranked,
        "reconciliation": reconciliation,
        "manifest_sha256": manifest["manifest_sha256"],
        "zero_detail_rows_lost": (
            ranked["rows"] == total_rows
            and reconciliation["rows"] == total_rows
            and ranked["identity_and_order_verified_rows"] == total_rows
            and reconciliation["identity_and_order_verified_rows"]
            == total_rows
        ),
    }


def run_stress(
    root: Path,
    *,
    farms: int,
    rows_per_farm: int,
    rows_per_volume: int,
) -> Dict:
    project = root / "group_project"
    project.mkdir(parents=True, exist_ok=False)
    tasks: List[Dict] = []
    started = time.monotonic()
    for farm_index in range(1, farms + 1):
        child = (
            project
            / "farm_projects"
            / f"stress_{farm_index:03d}"
        )
        task = _write_farm_inputs(
            child,
            farm_index=farm_index,
            rows=rows_per_farm,
            task_id=str(uuid.uuid4()),
        )
        # _write_farm_inputs 只知道子目录的上两级；在此固定为父组相对路径。
        task["relative_path"] = child.relative_to(project).as_posix()
        tasks.append(task)
        print(
            f"[输入] {farm_index}/{farms} 个牧场完成，"
            f"{rows_per_farm:,} 行",
            file=sys.stderr,
            flush=True,
        )

    def progress(value: int, message: str) -> None:
        print(
            f"[导出 {value:3d}%] {message}",
            file=sys.stderr,
            flush=True,
        )

    manifest = GroupCowRankingDetailExporter(
        project,
        rows_per_volume=rows_per_volume,
        score_column=SCORE_COLUMN,
        progress_callback=progress,
    ).export(
        tasks=tasks,
        output_dir=project / "reports",
        package_name="pressure-detail",
    )
    verification = _verify_manifest(
        manifest,
        tasks=tasks,
        rows_per_farm=rows_per_farm,
        rows_per_volume=rows_per_volume,
    )
    verification.update(
        {
            "success": True,
            "farms": farms,
            "rows_per_farm": rows_per_farm,
            "rows_per_volume": rows_per_volume,
            "crossed_single_sheet_data_limit": (
                verification["total_rows"] > EXCEL_MAX_ROWS - 1
            ),
            "elapsed_seconds": round(time.monotonic() - started, 3),
            "project_path": str(project),
        }
    )
    return verification


class _SimulatedInterruption(RuntimeError):
    """压力脚本主动制造的首卷提交后退出。"""


def run_resume_acceptance(
    root: Path,
    *,
    farms: int = 2,
    rows_per_farm: int = 12,
    rows_per_volume: int = 7,
) -> Dict:
    """用小数据模拟中断并验证续跑，不重复百万行主测试。"""
    if farms < 2 or rows_per_farm < 1 or rows_per_volume < 1:
        raise ValueError("恢复验收参数必须为正，且牧场数至少为 2")
    project = Path(root) / "resume_project"
    project.mkdir(parents=True, exist_ok=False)
    tasks: List[Dict] = []
    for farm_index in range(1, farms + 1):
        child = project / "farm_projects" / f"resume_{farm_index:03d}"
        task = _write_farm_inputs(
            child,
            farm_index=farm_index,
            rows=rows_per_farm,
            task_id=str(uuid.uuid4()),
        )
        task["relative_path"] = child.relative_to(project).as_posix()
        tasks.append(task)

    package_name = "resume-acceptance"
    output_dir = project / "reports"
    staging = output_dir / f".{package_name}.resume"
    import core.group_report.detail_exporter as detail_exporter_module

    real_atomic_write = detail_exporter_module._write_json_atomic
    interruption = {"raised": False}

    def interrupt_after_first_committed_volume(path: Path, payload: Dict):
        real_atomic_write(path, payload)
        if (
            not interruption["raised"]
            and Path(path).name == "export_checkpoint.json"
            and len(payload.get("volumes", [])) == 1
        ):
            interruption["raised"] = True
            raise _SimulatedInterruption("模拟首卷提交后进程退出")

    detail_exporter_module._write_json_atomic = (
        interrupt_after_first_committed_volume
    )
    try:
        try:
            GroupCowRankingDetailExporter(
                project,
                rows_per_volume=rows_per_volume,
                score_column=SCORE_COLUMN,
            ).export(
                tasks=tasks,
                output_dir=output_dir,
                package_name=package_name,
            )
        except _SimulatedInterruption:
            pass
        else:
            raise AssertionError("恢复验收没有触发模拟中断")
    finally:
        detail_exporter_module._write_json_atomic = real_atomic_write

    checkpoint_path = staging / ".work" / "export_checkpoint.json"
    if not checkpoint_path.is_file():
        raise AssertionError("模拟中断后没有保留导出检查点")
    checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8"))
    committed = list(checkpoint.get("volumes") or [])
    if len(committed) != 1:
        raise AssertionError("模拟中断时应恰好提交一卷")
    first_entry = committed[0]
    first_staged_path = staging / str(first_entry["path"])
    _assert_xlsx_zip(first_staged_path)
    first_sha = _sha256(first_staged_path)
    first_mtime_ns = first_staged_path.stat().st_mtime_ns

    manifest = GroupCowRankingDetailExporter(
        project,
        rows_per_volume=rows_per_volume,
        score_column=SCORE_COLUMN,
    ).export(
        tasks=tasks,
        output_dir=output_dir,
        package_name=package_name,
    )
    verification = _verify_manifest(
        manifest,
        tasks=tasks,
        rows_per_farm=rows_per_farm,
        rows_per_volume=rows_per_volume,
    )
    published_first = (
        Path(manifest["package_path"]) / str(first_entry["path"])
    )
    first_volume_reused = (
        published_first.is_file()
        and _sha256(published_first) == first_sha
        and published_first.stat().st_mtime_ns == first_mtime_ns
    )
    resumed_sources = sum(
        bool(source.get("resumed_from_checkpoint"))
        for source in manifest.get("sources", [])
    )
    passed = (
        interruption["raised"]
        and resumed_sources == farms
        and first_volume_reused
        and verification["zero_detail_rows_lost"]
        and not staging.exists()
    )
    if not passed:
        raise AssertionError("模拟中断后的续跑复用或完整性验收失败")
    return {
        "passed": True,
        "simulated_interruption": True,
        "farms": farms,
        "rows_per_farm": rows_per_farm,
        "rows_per_volume": rows_per_volume,
        "sources_reused": resumed_sources,
        "first_committed_volume_reused": first_volume_reused,
        "resume_directory_removed_after_publish": not staging.exists(),
        "verification": verification,
    }


def _evaluate_resource_limits(
    metrics: Dict,
    *,
    max_peak_rss_mib: float,
    max_workspace_peak_gib: float,
) -> Dict:
    """把采样值和可配置阈值转换成明确的验收结论。"""
    rss_limit = float(max_peak_rss_mib)
    disk_limit = float(max_workspace_peak_gib)
    if (
        not math.isfinite(rss_limit)
        or not math.isfinite(disk_limit)
        or rss_limit < 0
        or disk_limit < 0
    ):
        raise ValueError("资源阈值必须是大于等于 0 的有限数值")
    rss_limit_bytes = int(rss_limit * 1024 ** 2)
    disk_limit_bytes = int(disk_limit * 1024 ** 3)
    peak_rss = int(metrics.get("peak_rss_bytes", 0))
    peak_disk = int(metrics.get("workspace_peak_allocated_bytes", 0))
    rss_passed = not rss_limit_bytes or (
        peak_rss > 0 and peak_rss <= rss_limit_bytes
    )
    disk_passed = not disk_limit_bytes or peak_disk <= disk_limit_bytes
    return {
        "max_peak_rss_mib": rss_limit,
        "max_peak_rss_bytes": rss_limit_bytes,
        "peak_rss_within_limit": rss_passed,
        "max_workspace_peak_gib": disk_limit,
        "max_workspace_peak_bytes": disk_limit_bytes,
        "workspace_peak_within_limit": disk_passed,
        "passed": rss_passed and disk_passed,
    }


def _parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="牧场组全量 Excel 分卷可选压力测试",
    )
    parser.add_argument(
        "--farms",
        type=int,
        default=2,
        help="模拟牧场数，至少 2（默认：2）",
    )
    parser.add_argument(
        "--rows-per-farm",
        type=int,
        default=525_000,
        help="每个牧场的数据行数（默认：525000）",
    )
    parser.add_argument(
        "--rows-per-volume",
        type=int,
        default=EXCEL_MAX_ROWS - 1,
        help="每个输出 Excel 卷的数据行数上限",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        help="指定输出根目录；指定后结果会保留",
    )
    parser.add_argument(
        "--keep",
        action="store_true",
        help="未指定输出目录时，保留到当前目录下的 stress_outputs",
    )
    parser.add_argument(
        "--require-cross-limit",
        action="store_true",
        help="要求合计数据行严格超过单个 Excel 工作表数据行上限",
    )
    parser.add_argument(
        "--max-peak-rss-mib",
        type=float,
        default=DEFAULT_MAX_PEAK_RSS_MIB,
        help=(
            "进程峰值 RSS 上限 MiB；超过则验收失败，0 表示只记录"
            f"（默认：{DEFAULT_MAX_PEAK_RSS_MIB:g}）"
        ),
    )
    parser.add_argument(
        "--max-workspace-peak-gib",
        type=float,
        default=DEFAULT_MAX_WORKSPACE_PEAK_GIB,
        help=(
            "工作目录实际分配空间峰值上限 GiB；超过则验收失败，"
            f"0 表示只记录（默认：{DEFAULT_MAX_WORKSPACE_PEAK_GIB:g}）"
        ),
    )
    parser.add_argument(
        "--resource-sample-interval",
        type=float,
        default=DEFAULT_RESOURCE_SAMPLE_INTERVAL,
        help=(
            "RSS/磁盘峰值采样间隔秒"
            f"（默认：{DEFAULT_RESOURCE_SAMPLE_INTERVAL:g}）"
        ),
    )
    parser.add_argument(
        "--skip-recovery-check",
        action="store_true",
        help="跳过小数据模拟中断续跑验收（默认执行）",
    )
    parser.add_argument(
        "--recovery-rows-per-farm",
        type=int,
        default=12,
        help="恢复验收每场行数（默认：12，不重复百万行主测试）",
    )
    parser.add_argument(
        "--recovery-rows-per-volume",
        type=int,
        default=7,
        help="恢复验收每卷行数（默认：7）",
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = _parse_args(argv)
    if args.farms < 2:
        raise SystemExit("--farms 必须至少为 2")
    if not 1 <= args.rows_per_farm <= EXCEL_MAX_ROWS - 1:
        raise SystemExit(
            f"--rows-per-farm 必须在 1 到 {EXCEL_MAX_ROWS - 1:,} 之间"
        )
    if not 1 <= args.rows_per_volume <= EXCEL_MAX_ROWS - 1:
        raise SystemExit(
            f"--rows-per-volume 必须在 1 到 {EXCEL_MAX_ROWS - 1:,} 之间"
        )
    total_rows = args.farms * args.rows_per_farm
    if args.require_cross_limit and total_rows <= EXCEL_MAX_ROWS - 1:
        raise SystemExit(
            "--require-cross-limit 已启用，但合计数据行没有超过 "
            f"{EXCEL_MAX_ROWS - 1:,}"
        )
    if args.recovery_rows_per_farm < 1:
        raise SystemExit("--recovery-rows-per-farm 必须至少为 1")
    if args.recovery_rows_per_volume < 1:
        raise SystemExit("--recovery-rows-per-volume 必须至少为 1")
    try:
        _evaluate_resource_limits(
            {},
            max_peak_rss_mib=args.max_peak_rss_mib,
            max_workspace_peak_gib=args.max_workspace_peak_gib,
        )
        if (
            not math.isfinite(args.resource_sample_interval)
            or args.resource_sample_interval <= 0
        ):
            raise ValueError("资源采样间隔必须是大于 0 的有限数值")
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc

    temporary: Optional[tempfile.TemporaryDirectory] = None
    if args.output_root is not None:
        output_root = args.output_root.expanduser().resolve()
        output_root.mkdir(parents=True, exist_ok=True)
        root = output_root / f"group-detail-stress-{uuid.uuid4().hex[:10]}"
        root.mkdir()
        keep = True
    elif args.keep:
        output_root = Path.cwd() / "stress_outputs"
        output_root.mkdir(parents=True, exist_ok=True)
        root = output_root / f"group-detail-stress-{uuid.uuid4().hex[:10]}"
        root.mkdir()
        keep = True
    else:
        temporary = tempfile.TemporaryDirectory(
            prefix="group-detail-stress-"
        )
        root = Path(temporary.name)
        keep = False

    monitor = ResourceMonitor(
        root,
        interval_seconds=args.resource_sample_interval,
    ).start()
    run_started = time.monotonic()
    try:
        try:
            result = run_stress(
                root,
                farms=args.farms,
                rows_per_farm=args.rows_per_farm,
                rows_per_volume=args.rows_per_volume,
            )
            if args.skip_recovery_check:
                recovery = {"skipped": True, "passed": None}
            else:
                recovery = run_resume_acceptance(
                    root,
                    rows_per_farm=args.recovery_rows_per_farm,
                    rows_per_volume=args.recovery_rows_per_volume,
                )
        finally:
            resource_metrics = monitor.stop()

        resource_limits = _evaluate_resource_limits(
            resource_metrics,
            max_peak_rss_mib=args.max_peak_rss_mib,
            max_workspace_peak_gib=args.max_workspace_peak_gib,
        )
        crossed_limit = bool(result["crossed_single_sheet_data_limit"])
        ranked_complete = bool(
            result["ranked"]["rows"] == result["total_rows"]
            and result["ranked"]["identity_and_order_verified_rows"]
            == result["total_rows"]
        )
        reconciliation_complete = bool(
            result["reconciliation"]["rows"] == result["total_rows"]
            and result["reconciliation"][
                "identity_and_order_verified_rows"
            ]
            == result["total_rows"]
        )
        recovery_passed = bool(
            args.skip_recovery_check or recovery.get("passed")
        )
        acceptance = {
            "crossed_single_sheet_data_limit": crossed_limit,
            "cross_limit_requirement_passed": (
                crossed_limit or not args.require_cross_limit
            ),
            "ranked_details_zero_loss_and_sorted": ranked_complete,
            "reconciliation_details_zero_loss_and_ordered": (
                reconciliation_complete
            ),
            "volume_numbers_and_boundaries_continuous": bool(
                result["ranked"]["volume_boundaries_verified"]
                and result["reconciliation"][
                    "volume_boundaries_verified"
                ]
            ),
            "resume_reuses_committed_work": (
                None if args.skip_recovery_check else recovery_passed
            ),
            "resource_limits_passed": resource_limits["passed"],
        }
        acceptance["all_required_checks_passed"] = all(
            (
                acceptance["cross_limit_requirement_passed"],
                ranked_complete,
                reconciliation_complete,
                acceptance["volume_numbers_and_boundaries_continuous"],
                recovery_passed,
                resource_limits["passed"],
            )
        )
        result["success"] = acceptance["all_required_checks_passed"]
        result["recovery_acceptance"] = recovery
        result["resources"] = resource_metrics
        result["resource_limits"] = resource_limits
        result["acceptance"] = acceptance
        result["total_elapsed_seconds"] = round(
            time.monotonic() - run_started,
            3,
        )
        result["kept"] = keep
        if not keep:
            result["project_path"] = ""
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if result["success"] else 2
    finally:
        if temporary is not None:
            temporary.cleanup()


if __name__ == "__main__":
    raise SystemExit(main())

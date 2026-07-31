#!/usr/bin/env python3
"""牧场组验收用的逐牛公式复算。

该模块只读分析结果文件，并直接复用生产代码中的
``TraitsCalculation.calculate_single_trait_score`` 与
``IndexCalculation.calculate_index_score``。所有不一致只以聚合计数和
不可逆组合指纹返回，不返回牛号或单行明细。
"""

from __future__ import annotations

import contextlib
import hashlib
import io
import json
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook


PEDIGREE_WEIGHTS = {
    "sire": 0.5,
    "mgs": 0.25,
    "mmgs": 0.125,
    "default": 0.125,
}
SCORE_SUFFIX = "_score"
INDEX_SUFFIX = "_index"
IDENTIFIER_HEADERS = ("cow_id", "母牛号", "牛号", "耳号")


class FormulaValidationError(RuntimeError):
    """公式验收无法可靠执行。"""


def _normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value).strip()
    if text.casefold() in {"", "nan", "none", "null", "nat", "<na>", "n/a"}:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _is_blank_row(row: Sequence[Any]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row
    )


def _header_positions(headers: Sequence[Any]) -> Dict[str, int]:
    positions: Dict[str, int] = {}
    for index, value in enumerate(headers):
        name = str(value or "").strip()
        if not name:
            continue
        if name in positions:
            raise FormulaValidationError("公式输入存在重复列名")
        positions[name] = index
    return positions


def _identifier_position(positions: Mapping[str, int]) -> int:
    for candidate in IDENTIFIER_HEADERS:
        if candidate in positions:
            return positions[candidate]
    raise FormulaValidationError("公式输入缺少牛只标识列")


def _as_finite_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _numbers_match(
    expected: Any,
    actual: Any,
    *,
    relative_tolerance: float,
    absolute_tolerance: float,
) -> bool:
    expected_number = _as_finite_number(expected)
    actual_number = _as_finite_number(actual)
    if expected_number is None or actual_number is None:
        return expected_number is None and actual_number is None
    return math.isclose(
        expected_number,
        actual_number,
        rel_tol=relative_tolerance,
        abs_tol=absolute_tolerance,
    )


def _canonical_number(value: Any) -> str:
    number = _as_finite_number(value)
    if number is None:
        return "<missing>"
    return format(number, ".17g")


def _configuration_fingerprint(value: Mapping[str, Any]) -> str:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


@dataclass
class _MismatchAccumulator:
    """只保留不一致数量和不可逆聚合摘要。"""

    mismatch_cells: int = 0
    mismatch_rows: int = 0
    _digest: Any = field(default_factory=hashlib.sha256, repr=False)

    def add_row(
        self,
        identifier: Any,
        mismatches: Iterable[tuple[str, Any, Any]],
    ) -> None:
        values = list(mismatches)
        if not values:
            return
        self.mismatch_rows += 1
        identifier_digest = hashlib.sha256(
            _normalize_identifier(identifier).encode("utf-8")
        ).hexdigest()
        for label, expected, actual in values:
            self.mismatch_cells += 1
            payload = (
                identifier_digest,
                str(label),
                _canonical_number(expected),
                _canonical_number(actual),
            )
            self._digest.update(
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ).encode("utf-8")
            )

    def public_fingerprint(self) -> Dict[str, Any]:
        return {
            "algorithm": "sha256-ordered-formula-mismatch-v1",
            "mismatch_cells": int(self.mismatch_cells),
            "mismatch_rows": int(self.mismatch_rows),
            "digest": self._digest.hexdigest(),
        }


def _iter_nonblank_rows(
    worksheet: Any,
) -> tuple[List[Any], Iterator[Sequence[Any]]]:
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration as exc:
        raise FormulaValidationError("公式输入工作表为空") from exc

    def rows() -> Iterator[Sequence[Any]]:
        for row in iterator:
            if not _is_blank_row(row):
                yield row

    return headers, rows()


def _load_trait_defaults(
    traits: Sequence[str],
) -> tuple[Any, Dict[str, float]]:
    """通过生产计算类和本地公牛库取得本次公式使用的默认值。"""
    from core.breeding_calc.traits_calculation import TraitsCalculation

    calculator = TraitsCalculation()
    # 生产类仍保留了面向桌面运行的进度打印。验收器不能让这些输出携带
    # 本地路径或数据片段，因此在本地配置加载期间静默标准输出。
    with contextlib.redirect_stdout(io.StringIO()):
        if not calculator.init_db_connection():
            raise FormulaValidationError("无法加载性状公式默认配置")
        try:
            raw_defaults = calculator.get_default_values(list(traits))
        finally:
            if calculator.db_engine is not None:
                calculator.db_engine.dispose()
                calculator.db_engine = None

    defaults: Dict[str, float] = {}
    for trait in traits:
        value = _as_finite_number(raw_defaults.get(trait))
        if value is None:
            raise FormulaValidationError("性状公式默认配置存在非数值")
        defaults[str(trait)] = value
    return calculator, defaults


def _trait_formula_check(
    path: Path,
    *,
    require_pedigree_source_marker: bool,
    batch_size: int,
    relative_tolerance: float,
    absolute_tolerance: float,
) -> Dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers, rows = _iter_nonblank_rows(worksheet)
        positions = _header_positions(headers)
        identifier_index = _identifier_position(positions)
        score_columns = [
            name
            for name in positions
            if name.endswith(SCORE_SUFFIX)
        ]
        if not score_columns:
            raise FormulaValidationError("性状结果缺少得分列")
        traits = [name[: -len(SCORE_SUFFIX)] for name in score_columns]
        calculator, defaults = _load_trait_defaults(traits)
        config_fingerprint = _configuration_fingerprint(
            {
                "formula": "TraitsCalculation.calculate_single_trait_score",
                "weights": PEDIGREE_WEIGHTS,
                "defaults": defaults,
                "traits": traits,
            }
        )

        required_source_columns = sorted(
            {
                f"{ancestor}_{trait}"
                for trait in traits
                for ancestor in ("sire", "mgs", "mmgs")
            }
        )
        source_positions = {
            name: positions[name]
            for name in required_source_columns
            if name in positions
        }
        score_positions = {
            name: positions[name] for name in score_columns
        }
        source_marker_positions: Dict[str, int] = {}
        if require_pedigree_source_marker:
            for score_column in score_columns:
                marker_column = f"{score_column}_source"
                if marker_column not in positions:
                    raise FormulaValidationError(
                        "发布性状结果缺少得分来源列"
                    )
                source_marker_positions[score_column] = positions[
                    marker_column
                ]
        mismatch = _MismatchAccumulator()
        checked_rows = 0
        checked_cells = 0
        skipped_genomic_cells = 0
        batch: List[
            tuple[Any, Dict[str, Any], Dict[str, Any], set[str]]
        ] = []

        def check_batch() -> None:
            nonlocal checked_rows, checked_cells
            if not batch:
                return
            source_frame = pd.DataFrame(
                [
                    source
                    for _identifier, source, _actual, _eligible in batch
                ]
            )
            expected_by_trait = {
                trait: calculator.calculate_single_trait_score(
                    source_frame,
                    trait,
                    pd.DataFrame(),
                    defaults[trait],
                    PEDIGREE_WEIGHTS,
                )
                for trait in traits
            }
            for row_index, (
                identifier,
                _source,
                actual,
                eligible_traits,
            ) in enumerate(batch):
                row_mismatches = []
                for trait in traits:
                    if trait not in eligible_traits:
                        continue
                    expected = expected_by_trait[trait].iloc[row_index]
                    actual_value = actual[f"{trait}{SCORE_SUFFIX}"]
                    if not _numbers_match(
                        expected,
                        actual_value,
                        relative_tolerance=relative_tolerance,
                        absolute_tolerance=absolute_tolerance,
                    ):
                        row_mismatches.append(
                            (trait, expected, actual_value)
                        )
                mismatch.add_row(identifier, row_mismatches)
            checked_rows += len(batch)
            checked_cells += sum(
                len(eligible_traits)
                for _identifier, _source, _actual, eligible_traits in batch
            )
            batch.clear()

        for row in rows:
            source = {
                name: row[index] if index < len(row) else None
                for name, index in source_positions.items()
            }
            actual = {
                name: row[index] if index < len(row) else None
                for name, index in score_positions.items()
            }
            identifier = (
                row[identifier_index]
                if identifier_index < len(row)
                else None
            )
            eligible_traits = set(traits)
            if require_pedigree_source_marker:
                eligible_traits.clear()
                for trait in traits:
                    score_column = f"{trait}{SCORE_SUFFIX}"
                    marker_index = source_marker_positions[score_column]
                    marker = str(
                        row[marker_index]
                        if marker_index < len(row)
                        else ""
                    ).strip().upper()
                    if marker == "P":
                        eligible_traits.add(trait)
                    elif marker == "G":
                        skipped_genomic_cells += 1
                    else:
                        raise FormulaValidationError(
                            "发布性状结果存在未知得分来源"
                        )
            batch.append((identifier, source, actual, eligible_traits))
            if len(batch) >= batch_size:
                check_batch()
        check_batch()

        if checked_rows == 0:
            raise FormulaValidationError("性状结果没有数据行")
        return {
            "checked_rows": checked_rows,
            "checked_cells": checked_cells,
            "checked_traits": len(traits),
            "skipped_genomic_cells": skipped_genomic_cells,
            "mismatch_cells": mismatch.mismatch_cells,
            "mismatch_rows": mismatch.mismatch_rows,
            "mismatch_fingerprint": mismatch.public_fingerprint(),
            "configuration_fingerprint": config_fingerprint,
            "passed": mismatch.mismatch_cells == 0,
        }
    finally:
        workbook.close()


def _index_formula_check(
    path: Path,
    *,
    relative_tolerance: float,
    absolute_tolerance: float,
) -> Dict[str, Any]:
    from core.breeding_calc.index_calculation import (
        IndexCalculation,
        TRAIT_SD,
    )

    calculator = IndexCalculation()
    with contextlib.redirect_stdout(io.StringIO()):
        weights = calculator.load_weights()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers, rows = _iter_nonblank_rows(worksheet)
        positions = _header_positions(headers)
        identifier_index = _identifier_position(positions)
        index_columns = [
            name
            for name in positions
            if name.endswith(INDEX_SUFFIX)
        ]
        if not index_columns:
            raise FormulaValidationError("指数结果缺少指数列")

        configured: Dict[str, Mapping[str, float]] = {}
        for column in index_columns:
            weight_name = column[: -len(INDEX_SUFFIX)]
            weight_values = weights.get(weight_name)
            if not isinstance(weight_values, Mapping) or not weight_values:
                raise FormulaValidationError("指数结果无法匹配权重配置")
            configured[column] = weight_values

        relevant_sd = {
            trait: float(TRAIT_SD[trait])
            for values in configured.values()
            for trait in values
            if trait in TRAIT_SD
        }
        config_fingerprint = _configuration_fingerprint(
            {
                "formula": "IndexCalculation.calculate_index_score",
                "weights": {
                    column: {
                        str(trait): float(weight)
                        for trait, weight in sorted(
                            values.items(),
                            key=lambda item: str(item[0]),
                        )
                    }
                    for column, values in sorted(configured.items())
                },
                "trait_sd": relevant_sd,
            }
        )

        mismatch = _MismatchAccumulator()
        checked_rows = 0
        checked_cells = 0
        for row in rows:
            identifier = (
                row[identifier_index]
                if identifier_index < len(row)
                else None
            )
            row_mismatches = []
            for column, weight_values in configured.items():
                trait_values = {}
                for trait in weight_values:
                    score_column = f"{trait}{SCORE_SUFFIX}"
                    position = positions.get(score_column)
                    if position is None:
                        continue
                    value = _as_finite_number(
                        row[position] if position < len(row) else None
                    )
                    if value is not None:
                        trait_values[trait] = value
                expected = calculator.calculate_index_score(
                    trait_values,
                    dict(weight_values),
                )
                actual_position = positions[column]
                actual = (
                    row[actual_position]
                    if actual_position < len(row)
                    else None
                )
                if not _numbers_match(
                    expected,
                    actual,
                    relative_tolerance=relative_tolerance,
                    absolute_tolerance=absolute_tolerance,
                ):
                    row_mismatches.append((column, expected, actual))
            mismatch.add_row(identifier, row_mismatches)
            checked_rows += 1
            checked_cells += len(configured)

        if checked_rows == 0:
            raise FormulaValidationError("指数结果没有数据行")
        return {
            "checked_rows": checked_rows,
            "checked_cells": checked_cells,
            "checked_indexes": len(configured),
            "mismatch_cells": mismatch.mismatch_cells,
            "mismatch_rows": mismatch.mismatch_rows,
            "mismatch_fingerprint": mismatch.public_fingerprint(),
            "configuration_fingerprint": config_fingerprint,
            "passed": mismatch.mismatch_cells == 0,
        }
    finally:
        workbook.close()


def validate_cow_formulas(
    child_path: Path,
    *,
    batch_size: int = 1_000,
    relative_tolerance: float = 1e-9,
    absolute_tolerance: float = 1e-8,
) -> Dict[str, Any]:
    """逐牛复算性状得分和指数，返回无牛号的聚合结果。"""
    if int(batch_size) <= 0:
        raise ValueError("batch_size 必须大于 0")
    child = Path(child_path)
    analysis = child / "analysis_results"
    trait_path = (
        analysis
        / "processed_cow_data_key_traits_scores_pedigree.xlsx"
    )
    final_trait_path = (
        analysis
        / "processed_cow_data_key_traits_final.xlsx"
    )
    index_path = analysis / "processed_index_cow_index_scores.xlsx"
    for path in (trait_path, final_trait_path, index_path):
        if not path.is_file() or path.is_symlink():
            raise FormulaValidationError("逐牛公式验收所需文件不存在")

    pedigree_trait = _trait_formula_check(
        trait_path,
        require_pedigree_source_marker=False,
        batch_size=int(batch_size),
        relative_tolerance=float(relative_tolerance),
        absolute_tolerance=float(absolute_tolerance),
    )
    final_trait = _trait_formula_check(
        final_trait_path,
        require_pedigree_source_marker=True,
        batch_size=int(batch_size),
        relative_tolerance=float(relative_tolerance),
        absolute_tolerance=float(absolute_tolerance),
    )
    trait_digest = hashlib.sha256(
        json.dumps(
            [
                pedigree_trait["mismatch_fingerprint"]["digest"],
                final_trait["mismatch_fingerprint"]["digest"],
            ],
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    trait = {
        "checked_rows": max(
            int(pedigree_trait["checked_rows"]),
            int(final_trait["checked_rows"]),
        ),
        "checked_cells": (
            int(pedigree_trait["checked_cells"])
            + int(final_trait["checked_cells"])
        ),
        "checked_traits": max(
            int(pedigree_trait["checked_traits"]),
            int(final_trait["checked_traits"]),
        ),
        "skipped_genomic_cells": int(
            final_trait["skipped_genomic_cells"]
        ),
        "mismatch_cells": (
            int(pedigree_trait["mismatch_cells"])
            + int(final_trait["mismatch_cells"])
        ),
        # 同一牛可能在两个正式产物中各出现一次，因此这是产物行次，
        # 不是去重后的牛数。
        "mismatch_rows": (
            int(pedigree_trait["mismatch_rows"])
            + int(final_trait["mismatch_rows"])
        ),
        "mismatch_fingerprint": {
            "algorithm": "sha256-formula-artifact-set-v1",
            "mismatch_cells": (
                int(pedigree_trait["mismatch_cells"])
                + int(final_trait["mismatch_cells"])
            ),
            "mismatch_rows": (
                int(pedigree_trait["mismatch_rows"])
                + int(final_trait["mismatch_rows"])
            ),
            "digest": trait_digest,
        },
        "configuration_fingerprint": _configuration_fingerprint(
            {
                "pedigree": pedigree_trait[
                    "configuration_fingerprint"
                ],
                "final": final_trait["configuration_fingerprint"],
            }
        ),
        "pedigree_artifact": pedigree_trait,
        "final_artifact": final_trait,
        "passed": bool(
            pedigree_trait["passed"] and final_trait["passed"]
        ),
    }
    index = _index_formula_check(
        index_path,
        relative_tolerance=float(relative_tolerance),
        absolute_tolerance=float(absolute_tolerance),
    )
    return {
        "kind": "cow_formula_recalculation",
        "trait": trait,
        "index": index,
        "passed": bool(trait["passed"] and index["passed"]),
    }

#!/usr/bin/env python3
"""牧场组终态后的补充验收计算工具。

本工具只补跑自动分析链尚未覆盖的两个功能，并重新生成依赖这些结果的
单牧场 Excel 和牧场组最终汇总 Excel：

1. 母牛自身近交分析；
2. 强制重跑自动备选公牛/已配公牛近交分析；
3. 个体选配；
4. 单牧场 Excel（不生成 PPT）；
5. analysis / child_excel 阶段清单的 full 校验；
6. 所有牧场均成功后重新生成牧场组汇总 Excel。

安全约束：

* 默认仅 dry-run；只有显式传入 ``--execute`` 才会读取业务 Excel 或写文件；
* 执行前要求所有纳入汇总的父任务均已处于成功终态；
* 不导入或调用慧牧云推送模块，不生成任何 PPT；
* 每个牧场使用独立子进程串行执行，底层详细输出全部静默，外层只报告
  聚合进度，不打印牛号、公牛号、鉴权信息或底层 SQL；
* 新结果先在同一子项目文件系统内隔离生成并校验，再事务式提升；提交
  清单或 full 校验失败时恢复提升前的结果和阶段清单。

预览（默认模式）::

    python scripts/run_group_supplemental_acceptance.py \
        --group /path/to/牧场组_YYYYMMDD

真实执行::

    python scripts/run_group_supplemental_acceptance.py \
        --group /path/to/牧场组_YYYYMMDD \
        --execute

单场 canary（成功后不生成组汇总）::

    python scripts/run_group_supplemental_acceptance.py \
        --group /path/to/牧场组_YYYYMMDD \
        --canary-task-id <task-uuid> \
        --execute
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import inspect
import json
import logging
import os
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, Optional, Sequence


REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
if str(REPOSITORY_ROOT) not in sys.path:
    sys.path.insert(0, str(REPOSITORY_ROOT))


SUCCESS_TASK_STATUSES = {"completed", "completed_with_warning"}
SUPPLEMENTAL_FILENAMES = (
    "母牛近交系数分析结果.xlsx",
    "个体选配推荐矩阵.xlsx",
    "个体选配报告.xlsx",
    "individual_mating_report.xlsx",
)
PAIR_INBREEDING_PREFIXES = (
    "备选公牛_近交系数及隐性基因分析结果",
    "已配公牛_近交系数及隐性基因分析结果",
)
MUTABLE_ANALYSIS_FILENAMES = {
    *SUPPLEMENTAL_FILENAMES,
    "processed_index_cow_index_scores.xlsx",
}
RESULT_FILENAME = "supplemental_acceptance_result.json"
SUPPLEMENTAL_MANIFEST_FILENAME = "supplemental_acceptance.json"
CHILD_RESULT_SCHEMA = 1


class SupplementalAcceptanceError(RuntimeError):
    """补充验收前置条件、计算或校验失败。"""


def _utc_timestamp() -> str:
    return (
        datetime.now(timezone.utc)
        .isoformat(timespec="seconds")
        .replace("+00:00", "Z")
    )


def _atomic_write_json(path: Path, payload: Dict[str, Any]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(
                payload,
                handle,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _read_json(path: Path) -> Dict[str, Any]:
    try:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise SupplementalAcceptanceError("项目描述不可读") from exc
    if not isinstance(payload, dict):
        raise SupplementalAcceptanceError("项目描述格式不正确")
    return payload


def _read_group_tasks_readonly(group_path: Path) -> list[Dict[str, Any]]:
    database_path = group_path / "group_store" / "group_tasks.sqlite3"
    if not database_path.is_file():
        raise SupplementalAcceptanceError("牧场组任务状态库不存在")
    uri = f"{database_path.resolve().as_uri()}?mode=ro"
    try:
        connection = sqlite3.connect(uri, uri=True, timeout=30)
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT task_id, farm_code, farm_name, relative_path,
                   source_kind, source_system, included_in_summary,
                   status, sort_order
            FROM group_tasks
            WHERE included_in_summary = 1
            ORDER BY sort_order, created_at, task_id
            """
        ).fetchall()
        return [dict(row) for row in rows]
    except sqlite3.Error as exc:
        raise SupplementalAcceptanceError("无法只读访问牧场组任务状态") from exc
    finally:
        if "connection" in locals():
            connection.close()


def _has_active_group_lease(group_path: Path) -> bool:
    database_path = group_path / "group_store" / "group_tasks.sqlite3"
    uri = f"{database_path.resolve().as_uri()}?mode=ro"
    try:
        connection = sqlite3.connect(uri, uri=True, timeout=30)
        connection.row_factory = sqlite3.Row
        row = connection.execute(
            """
            SELECT lease_token, lease_expires_at
            FROM group_run_control
            WHERE singleton_id = 1
            """
        ).fetchone()
    except sqlite3.Error as exc:
        raise SupplementalAcceptanceError("无法检查牧场组运行锁") from exc
    finally:
        if "connection" in locals():
            connection.close()
    if row is None or not row["lease_token"] or not row["lease_expires_at"]:
        return False
    now = _utc_timestamp()
    return str(row["lease_expires_at"]) > now


def _resolve_group(path: Path) -> tuple[Path, Dict[str, Any], list[Dict[str, Any]]]:
    try:
        group_path = Path(path).expanduser().resolve(strict=True)
    except (OSError, RuntimeError) as exc:
        raise SupplementalAcceptanceError("牧场组目录不存在") from exc
    if not group_path.is_dir():
        raise SupplementalAcceptanceError("牧场组路径不是目录")
    metadata = _read_json(group_path / "project_metadata.json")
    if metadata.get("project_type") != "multi_farm_group":
        raise SupplementalAcceptanceError("指定目录不是牧场组项目")
    tasks = _read_group_tasks_readonly(group_path)
    if not tasks:
        raise SupplementalAcceptanceError("没有纳入汇总范围的牧场任务")
    return group_path, metadata, tasks


def _validate_terminal_tasks(tasks: Sequence[Dict[str, Any]]) -> None:
    nonterminal = [
        task for task in tasks
        if str(task.get("status") or "") not in SUCCESS_TASK_STATUSES
    ]
    if nonterminal:
        raise SupplementalAcceptanceError(
            f"父任务尚未全部成功终止（未就绪 {len(nonterminal)} 个）"
        )


def _safe_child_path(group_path: Path, task: Dict[str, Any]) -> Path:
    relative = Path(str(task.get("relative_path") or ""))
    if not relative.parts or relative.is_absolute():
        raise SupplementalAcceptanceError("子项目相对路径无效")
    child = (group_path / relative).resolve(strict=True)
    try:
        child.relative_to(group_path.resolve())
    except ValueError as exc:
        raise SupplementalAcceptanceError("子项目路径越出牧场组目录") from exc
    if not child.is_dir():
        raise SupplementalAcceptanceError("子项目目录不存在")
    child_metadata = _read_json(child / "project_metadata.json")
    if child_metadata.get("project_type") != "group_child":
        raise SupplementalAcceptanceError("子项目身份无效")
    if str(child_metadata.get("group_task_id") or "") != str(
        task.get("task_id") or ""
    ):
        raise SupplementalAcceptanceError("子项目 task_id 与父任务不一致")
    if str(child_metadata.get("group_farm_code") or "") != str(
        task.get("farm_code") or ""
    ):
        raise SupplementalAcceptanceError("子项目牧场身份与父任务不一致")
    return child


def _xlsx_summary(
    path: Path,
    *,
    required_sheets: Iterable[str] = (),
    require_data: bool = True,
) -> Dict[str, Any]:
    path = Path(path)
    if not path.is_file() or path.stat().st_size <= 0:
        raise SupplementalAcceptanceError("工作簿不存在或为空")
    if not zipfile.is_zipfile(path):
        raise SupplementalAcceptanceError("工作簿不是有效 xlsx 容器")
    with zipfile.ZipFile(path) as archive:
        if archive.testzip() is not None:
            raise SupplementalAcceptanceError("工作簿 ZIP CRC 校验失败")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook.sheetnames)
        missing = set(required_sheets) - set(sheet_names)
        if missing:
            raise SupplementalAcceptanceError("工作簿缺少必需工作表")
        max_rows = {
            sheet_name: int(workbook[sheet_name].max_row or 0)
            for sheet_name in sheet_names
        }
        if require_data and not any(value >= 2 for value in max_rows.values()):
            raise SupplementalAcceptanceError("工作簿没有明细数据")
        return {
            "size": path.stat().st_size,
            "sheet_count": len(sheet_names),
            "max_rows": max_rows,
        }
    finally:
        workbook.close()


def _sha256(path: Path, block_size: int = 4 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        while True:
            chunk = handle.read(block_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


@contextlib.contextmanager
def _silence_sensitive_output() -> Iterator[None]:
    """静默底层遗留调试输出，避免打印动物号、SQL 或其它明细。"""

    previous_disable = logging.root.manager.disable
    logging.disable(logging.CRITICAL)
    with open(os.devnull, "w", encoding="utf-8") as sink:
        with contextlib.redirect_stdout(sink), contextlib.redirect_stderr(sink):
            try:
                yield
            finally:
                logging.disable(previous_disable)


def _link_or_copy(source: Path, target: Path) -> None:
    if source.is_symlink():
        raise SupplementalAcceptanceError("子项目包含不受支持的符号链接")
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        os.symlink(source.resolve(), target)
    except (OSError, NotImplementedError):
        shutil.copy2(source, target)


def _mirror_tree_readonly(
    source_root: Path,
    target_root: Path,
    *,
    copied_names: Iterable[str] = (),
    omitted_names: Iterable[str] = (),
    omitted_prefixes: Iterable[str] = (),
) -> None:
    copied = set(copied_names)
    omitted = set(omitted_names)
    prefixes = tuple(omitted_prefixes)
    if not source_root.is_dir():
        return
    for source in sorted(source_root.rglob("*")):
        relative = source.relative_to(source_root)
        target = target_root / relative
        if source.is_dir():
            target.mkdir(parents=True, exist_ok=True)
            continue
        if (
            not source.is_file()
            or source.name in omitted
            or source.name.startswith(prefixes)
        ):
            continue
        target.parent.mkdir(parents=True, exist_ok=True)
        if source.name in copied:
            shutil.copy2(source, target)
        else:
            _link_or_copy(source, target)


def _create_shadow_project(child_path: Path) -> tuple[Path, Path]:
    staging_parent = child_path / "group_store" / "supplemental_staging"
    staging_parent.mkdir(parents=True, exist_ok=True)
    run_root = Path(
        tempfile.mkdtemp(prefix="run-", dir=staging_parent)
    )
    shadow = run_root / "project"
    shadow.mkdir()
    shutil.copy2(
        child_path / "project_metadata.json",
        shadow / "project_metadata.json",
    )
    _mirror_tree_readonly(
        child_path / "raw_data",
        shadow / "raw_data",
    )
    _mirror_tree_readonly(
        child_path / "standardized_data",
        shadow / "standardized_data",
    )
    _mirror_tree_readonly(
        child_path / "analysis_results",
        shadow / "analysis_results",
        copied_names={"processed_index_cow_index_scores.xlsx"},
        omitted_names=SUPPLEMENTAL_FILENAMES,
        omitted_prefixes=PAIR_INBREEDING_PREFIXES,
    )
    (shadow / "reports").mkdir(parents=True, exist_ok=True)
    return run_root, shadow


def _json_safe_cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _write_cow_self_workbook(
    page: Any,
    results: list[Dict[str, Any]],
    output_path: Path,
) -> int:
    import pandas as pd

    detail = pd.DataFrame(results)
    if detail.empty:
        raise SupplementalAcceptanceError("母牛自身近交分析没有产生明细")
    if "近交系数" in detail.columns:
        numeric = (
            detail["近交系数"]
            .astype(str)
            .str.rstrip("%")
        )
        detail = (
            detail.assign(
                _sort_value=pd.to_numeric(numeric, errors="coerce").fillna(-1)
            )
            .sort_values("_sort_value", ascending=False)
            .drop(columns=["_sort_value"])
        )
    for column in detail.columns:
        detail[column] = detail[column].map(_json_safe_cell)
    abnormal, stats = page.collect_cow_self_abnormal(results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.name}.",
        suffix=".tmp.xlsx",
        dir=output_path.parent,
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
            detail.to_excel(writer, sheet_name="配对明细表", index=False)
            abnormal.to_excel(writer, sheet_name="异常明细表", index=False)
            stats.to_excel(writer, sheet_name="统计表", index=False)
        _xlsx_summary(
            temporary,
            required_sheets=("配对明细表", "异常明细表", "统计表"),
        )
        os.replace(temporary, output_path)
    finally:
        temporary.unlink(missing_ok=True)
    return len(detail)


def _run_cow_self(shadow: Path) -> int:
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PyQt6.QtWidgets import QApplication
    from core.data.update_manager import get_pedigree_db, reset_pedigree_db
    from core.inbreeding.inbreeding_page import InbreedingPage

    application = QApplication.instance() or QApplication([])
    page = InbreedingPage()
    cow_file = shadow / "standardized_data" / "processed_cow_data.xlsx"
    reset_pedigree_db()
    pedigree_db = get_pedigree_db()
    pedigree_db.build_cow_pedigree(cow_file, lambda *_args: True)
    results = page.analyze_cow_self(shadow)
    count = _write_cow_self_workbook(
        page,
        results,
        shadow / "analysis_results" / "母牛近交系数分析结果.xlsx",
    )
    page.deleteLater()
    application.processEvents()
    reset_pedigree_db()
    return count


def _load_inventory(path: Path) -> tuple[Dict[tuple[str, str], int], int]:
    import pandas as pd

    frame = pd.read_excel(path, dtype={"bull_id": str})
    required = {"bull_id", "semen_type", "支数"}
    if not required.issubset(frame.columns):
        raise SupplementalAcceptanceError("备选公牛库存缺少必需字段")
    inventory: Dict[tuple[str, str], int] = {}
    for row in frame.itertuples(index=False):
        data = row._asdict()
        bull_id = str(data.get("bull_id") or "").strip()
        semen_type = str(data.get("semen_type") or "").strip()
        raw_count = data.get("支数")
        try:
            numeric_count = float(raw_count)
        except (TypeError, ValueError) as exc:
            raise SupplementalAcceptanceError("备选公牛库存不是整数") from exc
        if (
            not bull_id
            or semen_type not in {"常规", "性控"}
            or not numeric_count.is_integer()
            or numeric_count <= 0
        ):
            raise SupplementalAcceptanceError("备选公牛库存存在无效记录")
        key = (bull_id, semen_type)
        if key in inventory:
            raise SupplementalAcceptanceError("备选公牛库存存在重复复合键")
        inventory[key] = int(numeric_count)
    if not inventory:
        raise SupplementalAcceptanceError("备选公牛库存为空")
    return inventory, len(inventory)


def _run_pair_inbreeding(
    shadow: Path,
) -> tuple[Dict[str, Path], Dict[str, int]]:
    """强制重算备选/已配近交；验收模式禁止上报缺失公牛。"""

    from core.auto_analysis_runner import run_inbreeding_analysis
    from core.data.update_manager import reset_pedigree_db

    parameters = inspect.signature(run_inbreeding_analysis).parameters
    if "allow_missing_bull_upload" not in parameters:
        raise SupplementalAcceptanceError(
            "当前自动近交入口不能显式禁用缺失公牛上报"
        )

    analysis_dir = shadow / "analysis_results"
    specifications = (
        (
            "candidate",
            "备选公牛_近交系数及隐性基因分析结果",
        ),
        (
            "mated",
            "已配公牛_近交系数及隐性基因分析结果",
        ),
    )
    generated: Dict[str, Path] = {}
    row_counts: Dict[str, int] = {}

    if not (
        shadow / "standardized_data" / "processed_breeding_data.xlsx"
    ).is_file():
        raise SupplementalAcceptanceError(
            "已配公牛近交重算缺少标准化配种记录"
        )

    for analysis_type, prefix in specifications:
        # shadow 创建时已排除全部历史同前缀文件；这里再做一次断言，
        # 防止旧文件因命名变化被误当成刚生成结果。
        before = set(analysis_dir.glob(f"{prefix}*.xlsx"))
        if before:
            raise SupplementalAcceptanceError("shadow 混入旧近交结果")
        reset_pedigree_db()
        success, _message = run_inbreeding_analysis(
            shadow,
            analysis_type,
            progress_cb=lambda *_args: None,
            allow_missing_bull_upload=False,
        )
        if not success:
            raise SupplementalAcceptanceError("自动近交分析未成功完成")
        after = set(analysis_dir.glob(f"{prefix}*.xlsx"))
        created = sorted(after - before)
        if len(created) != 1:
            raise SupplementalAcceptanceError(
                "自动近交分析没有产生唯一的新工作簿"
            )
        summary = _xlsx_summary(
            created[0],
            required_sheets=("配对明细表",),
        )
        rows = int(summary["max_rows"].get("配对明细表", 0)) - 1
        if rows <= 0:
            raise SupplementalAcceptanceError("自动近交分析没有明细")
        generated[analysis_type] = created[0]
        row_counts[analysis_type] = rows

    reset_pedigree_db()
    return generated, row_counts


def _run_matching(shadow: Path) -> tuple[int, int, int]:
    from core.matching.complete_mating_executor import CompleteMatingExecutor

    inventory, inventory_items = _load_inventory(
        shadow / "standardized_data" / "processed_bull_data.xlsx"
    )
    executor = CompleteMatingExecutor(shadow)
    result = executor.execute(
        bull_inventory=inventory,
        inbreeding_threshold=6.25,
        control_defect_genes=True,
        heifer_age_days=420,
        cycle_days=21,
        # 第一版备选公牛来自用户导入文件；个别库存公牛可能暂时没有
        # 育种指数。按“有多少可用数据就分析多少”的产品口径继续选配，
        # 但只在验收结果中记录跳过数量，不输出具体公牛号。
        skip_missing_bulls=True,
        selected_groups=None,
        grouping_mode="auto",
        progress_callback=lambda *_args: None,
    )
    if not result.get("success"):
        raise SupplementalAcceptanceError("个体选配未成功完成")
    report_path = shadow / "analysis_results" / "个体选配报告.xlsx"
    summary = _xlsx_summary(
        report_path,
        required_sheets=("选配结果",),
    )
    matching_rows = int(summary["max_rows"].get("选配结果", 0)) - 1
    if matching_rows <= 0:
        raise SupplementalAcceptanceError("个体选配报告没有明细")
    skipped_bulls = len(result.get("skipped_bulls") or [])
    return matching_rows, inventory_items, skipped_bulls


def _run_single_excel(shadow: Path, farm_name: str) -> Path:
    from core.auto_analysis_runner import run_excel_report

    success, result = run_excel_report(
        shadow,
        progress_cb=lambda *_args: None,
        service_staff="",
        farm_name=farm_name,
        max_workers=1,
    )
    if not success:
        raise SupplementalAcceptanceError("单牧场 Excel 生成失败")
    report = Path(result)
    try:
        report.resolve().relative_to((shadow / "reports").resolve())
    except ValueError as exc:
        raise SupplementalAcceptanceError("单牧场 Excel 输出目录无效") from exc
    summary = _xlsx_summary(
        report,
        required_sheets=("个体选配推荐结果",),
    )
    if int(summary["max_rows"].get("个体选配推荐结果", 0)) < 2:
        raise SupplementalAcceptanceError("单牧场 Excel 未写入个体选配章节")
    return report


class _PromotionTransaction:
    """同一子项目文件系统内的可回滚文件提升。"""

    def __init__(self, child_path: Path):
        self.child_path = child_path.resolve()
        root = child_path / "group_store" / "supplemental_transactions"
        root.mkdir(parents=True, exist_ok=True)
        self.backup_root = root / uuid.uuid4().hex
        self.backup_root.mkdir()
        self._records: list[tuple[Path, Optional[Path], bool]] = []
        self._committed = False

    def _checked_target(self, target: Path) -> Path:
        resolved_parent = target.parent.resolve(strict=True)
        resolved = resolved_parent / target.name
        try:
            resolved.relative_to(self.child_path)
        except ValueError as exc:
            raise SupplementalAcceptanceError("提升目标越出子项目目录") from exc
        return resolved

    def _backup_path(self, target: Path) -> Path:
        relative = target.relative_to(self.child_path)
        backup = self.backup_root / relative
        backup.parent.mkdir(parents=True, exist_ok=True)
        return backup

    def protect(self, target: Path) -> None:
        target = self._checked_target(Path(target))
        backup = self._backup_path(target)
        existed = target.is_file()
        if existed:
            shutil.copy2(target, backup)
        self._records.append((target, backup if existed else None, True))

    def promote(self, source: Path, target: Path) -> None:
        source = Path(source)
        target = self._checked_target(Path(target))
        if not source.is_file():
            raise SupplementalAcceptanceError("待提升结果不存在")
        target.parent.mkdir(parents=True, exist_ok=True)
        backup = self._backup_path(target)
        existed = target.is_file()
        if existed:
            os.replace(target, backup)
        self._records.append((target, backup if existed else None, False))
        os.replace(source, target)

    def rollback(self) -> None:
        for target, backup, protected_copy in reversed(self._records):
            target.unlink(missing_ok=True)
            if backup is not None and backup.is_file():
                if protected_copy:
                    os.replace(backup, target)
                else:
                    os.replace(backup, target)
        self._records.clear()
        shutil.rmtree(self.backup_root, ignore_errors=True)

    def commit(self) -> None:
        self._committed = True
        self._records.clear()
        shutil.rmtree(self.backup_root, ignore_errors=True)

    def __enter__(self) -> "_PromotionTransaction":
        return self

    def __exit__(self, exc_type, _exc, _traceback) -> bool:
        if exc_type is not None or not self._committed:
            self.rollback()
        return False


def _validate_supplemental_outputs(
    analysis_dir: Path,
    pair_inbreeding_paths: Dict[str, Path],
) -> Dict[str, Any]:
    summaries = {
        "母牛近交系数分析结果.xlsx": _xlsx_summary(
            analysis_dir / "母牛近交系数分析结果.xlsx",
            required_sheets=("配对明细表", "异常明细表", "统计表"),
        ),
        "个体选配推荐矩阵.xlsx": _xlsx_summary(
            analysis_dir / "个体选配推荐矩阵.xlsx",
        ),
        "个体选配报告.xlsx": _xlsx_summary(
            analysis_dir / "个体选配报告.xlsx",
            required_sheets=("选配结果",),
        ),
        "individual_mating_report.xlsx": _xlsx_summary(
            analysis_dir / "individual_mating_report.xlsx",
        ),
        "processed_index_cow_index_scores.xlsx": _xlsx_summary(
            analysis_dir / "processed_index_cow_index_scores.xlsx",
        ),
    }
    for analysis_type in ("candidate", "mated"):
        path = pair_inbreeding_paths.get(analysis_type)
        if path is None:
            raise SupplementalAcceptanceError("缺少新生成的自动近交工作簿")
        summaries[path.name] = _xlsx_summary(
            path,
            required_sheets=("配对明细表",),
        )
    return summaries


def _build_supplemental_manifest(
    child_path: Path,
    workbook_summaries: Dict[str, Any],
    *,
    cow_self_rows: int,
    matching_rows: int,
    inventory_items: int,
    skipped_bulls: int,
    report_path: Path,
    pair_inbreeding_paths: Dict[str, Path],
    pair_inbreeding_rows: Dict[str, int],
) -> Dict[str, Any]:
    artifacts = []
    relative_files = [
        Path("analysis_results") / name
        for name in (
            *SUPPLEMENTAL_FILENAMES,
            "processed_index_cow_index_scores.xlsx",
        )
    ]
    relative_files.extend(
        pair_inbreeding_paths[analysis_type].relative_to(child_path)
        for analysis_type in ("candidate", "mated")
    )
    relative_files.append(report_path.relative_to(child_path))
    for relative in relative_files:
        absolute = child_path / relative
        artifacts.append(
            {
                "relative_path": relative.as_posix(),
                "size": absolute.stat().st_size,
                "sha256": _sha256(absolute),
            }
        )
    return {
        "schema_version": 1,
        "generated_at": _utc_timestamp(),
        "execution": {
            "hmy_push": False,
            "ppt_generated": False,
            "sequential": True,
            "inbreeding_threshold_percent": 6.25,
            "gene_control": True,
        },
        "counts": {
            "cow_self_rows": int(cow_self_rows),
            "matching_rows": int(matching_rows),
            "inventory_items": int(inventory_items),
            "matching_skipped_bulls": int(skipped_bulls),
            "candidate_inbreeding_rows": int(
                pair_inbreeding_rows["candidate"]
            ),
            "mated_inbreeding_rows": int(
                pair_inbreeding_rows["mated"]
            ),
        },
        "workbooks": workbook_summaries,
        "artifacts": artifacts,
    }


def _validated_existing_supplemental_result(
    group_path: Path,
    task: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """返回可安全复用的已完成结果；缺失或不一致时返回 ``None``。

    补跑可能因机器睡眠、程序退出等原因在牧场之间中断。每个子项目只有
    在补充 manifest、自身产物哈希以及三个正式阶段清单均仍一致时才允许
    跳过重算。最终整组验收仍会再次做 full 校验。
    """

    child_path = _safe_child_path(group_path, task)
    manifest_path = (
        child_path
        / "group_store"
        / SUPPLEMENTAL_MANIFEST_FILENAME
    )
    if not manifest_path.is_file() or manifest_path.is_symlink():
        return None
    try:
        payload = _read_json(manifest_path)
        if payload.get("schema_version") != 1:
            return None
        execution = payload.get("execution")
        if (
            not isinstance(execution, dict)
            or execution.get("hmy_push") is not False
            or execution.get("ppt_generated") is not False
            or execution.get("sequential") is not True
        ):
            return None

        counts = payload.get("counts")
        if not isinstance(counts, dict):
            return None
        normalized_counts: Dict[str, int] = {}
        for key in (
            "cow_self_rows",
            "candidate_inbreeding_rows",
            "mated_inbreeding_rows",
            "matching_rows",
            "inventory_items",
            "matching_skipped_bulls",
        ):
            value = counts.get(key)
            if (
                not isinstance(value, int)
                or isinstance(value, bool)
                or value < 0
            ):
                return None
            normalized_counts[key] = int(value)
        if any(
            normalized_counts[key] <= 0
            for key in (
                "cow_self_rows",
                "candidate_inbreeding_rows",
                "mated_inbreeding_rows",
                "matching_rows",
                "inventory_items",
            )
        ):
            return None

        artifacts = payload.get("artifacts")
        if not isinstance(artifacts, list) or len(artifacts) < 8:
            return None
        seen_paths = set()
        artifact_names = set()
        child_root = child_path.resolve()
        for item in artifacts:
            if not isinstance(item, dict):
                return None
            relative_text = str(item.get("relative_path") or "")
            relative = Path(relative_text)
            if (
                not relative_text
                or "\\" in relative_text
                or relative.is_absolute()
                or any(part in {"", ".", ".."} for part in relative.parts)
                or relative_text in seen_paths
            ):
                return None
            seen_paths.add(relative_text)
            target = (child_path / relative).resolve()
            try:
                target.relative_to(child_root)
            except ValueError:
                return None
            if (
                target.is_symlink()
                or not target.is_file()
                or not isinstance(item.get("size"), int)
                or isinstance(item.get("size"), bool)
                or int(item["size"]) != int(target.stat().st_size)
                or str(item.get("sha256") or "") != _sha256(target)
            ):
                return None
            artifact_names.add(target.name)

        required_names = {
            *SUPPLEMENTAL_FILENAMES,
            "processed_index_cow_index_scores.xlsx",
        }
        if not required_names.issubset(artifact_names):
            return None
        if not any(
            name.startswith(PAIR_INBREEDING_PREFIXES[0])
            for name in artifact_names
        ) or not any(
            name.startswith(PAIR_INBREEDING_PREFIXES[1])
            for name in artifact_names
        ):
            return None
        if not any(
            name.startswith("育种分析综合报告_")
            and name.endswith(".xlsx")
            for name in artifact_names
        ):
            return None

        from core.group_tasks.stage_policy import validate_child_stage

        for stage in ("data", "analysis", "child_excel"):
            validation = validate_child_stage(
                child_path,
                stage,
                expected_task_id=str(task["task_id"]),
                expected_farm_code=str(task["farm_code"]),
                verification="stat",
            )
            if not validation.get("valid"):
                return None
        return {
            "schema_version": CHILD_RESULT_SCHEMA,
            "success": True,
            "reused": True,
            "counts": normalized_counts,
        }
    except (OSError, ValueError, SupplementalAcceptanceError):
        return None


def _process_one_child(
    group_path: Path,
    task: Dict[str, Any],
    result_path: Path,
) -> int:
    phase = "preflight"
    run_root: Optional[Path] = None
    try:
        child_path = _safe_child_path(group_path, task)
        run_root, shadow = _create_shadow_project(child_path)

        with _silence_sensitive_output():
            phase = "cow_self"
            cow_self_rows = _run_cow_self(shadow)

            phase = "pair_inbreeding"
            pair_inbreeding_paths, pair_inbreeding_rows = (
                _run_pair_inbreeding(shadow)
            )

            phase = "matching"
            (
                matching_rows,
                inventory_items,
                matching_skipped_bulls,
            ) = _run_matching(shadow)

            phase = "staged_validation"
            staged_summaries = _validate_supplemental_outputs(
                shadow / "analysis_results",
                pair_inbreeding_paths,
            )

            phase = "single_excel"
            staged_report = _run_single_excel(
                shadow,
                str(task.get("farm_name") or "牧场"),
            )
            _xlsx_summary(
                staged_report,
                required_sheets=("个体选配推荐结果",),
            )

            phase = "promotion"
            analysis_manifest = (
                child_path
                / "group_store"
                / "stage_manifests"
                / "analysis.json"
            )
            excel_manifest = (
                child_path
                / "group_store"
                / "stage_manifests"
                / "child_excel.json"
            )
            supplemental_manifest = (
                child_path
                / "group_store"
                / "supplemental_acceptance.json"
            )
            target_report = child_path / "reports" / staged_report.name
            target_pair_inbreeding_paths = {
                analysis_type: (
                    child_path
                    / "analysis_results"
                    / pair_inbreeding_paths[analysis_type].name
                )
                for analysis_type in ("candidate", "mated")
            }

            with _PromotionTransaction(child_path) as transaction:
                transaction.protect(analysis_manifest)
                transaction.protect(excel_manifest)
                transaction.protect(supplemental_manifest)
                for filename in (
                    "processed_index_cow_index_scores.xlsx",
                    *SUPPLEMENTAL_FILENAMES,
                ):
                    transaction.promote(
                        shadow / "analysis_results" / filename,
                        child_path / "analysis_results" / filename,
                    )
                for analysis_type in ("candidate", "mated"):
                    transaction.promote(
                        pair_inbreeding_paths[analysis_type],
                        target_pair_inbreeding_paths[analysis_type],
                    )
                transaction.promote(staged_report, target_report)

                phase = "analysis_manifest"
                from core.group_tasks.stage_policy import (
                    commit_child_stage,
                    validate_child_stage,
                )

                committed_analysis = commit_child_stage(
                    child_path,
                    "analysis",
                    expected_task_id=str(task["task_id"]),
                    expected_farm_code=str(task["farm_code"]),
                )
                committed_outputs = {
                    str(item.get("relative_path") or "")
                    for item in committed_analysis.get("outputs", [])
                    if isinstance(item, dict)
                }
                expected_pair_outputs = {
                    target.relative_to(child_path).as_posix()
                    for target in target_pair_inbreeding_paths.values()
                }
                if not expected_pair_outputs.issubset(committed_outputs):
                    raise SupplementalAcceptanceError(
                        "analysis 清单没有纳入本轮新近交工作簿"
                    )
                analysis_validation = validate_child_stage(
                    child_path,
                    "analysis",
                    expected_task_id=str(task["task_id"]),
                    expected_farm_code=str(task["farm_code"]),
                    verification="full",
                )
                if not analysis_validation.get("valid"):
                    raise SupplementalAcceptanceError(
                        "analysis 阶段 full 校验失败"
                    )

                phase = "child_excel_manifest"
                committed_excel = commit_child_stage(
                    child_path,
                    "child_excel",
                    expected_task_id=str(task["task_id"]),
                    expected_farm_code=str(task["farm_code"]),
                    report_path=target_report,
                )
                committed_excel_inputs = {
                    str(item.get("relative_path") or "")
                    for item in committed_excel.get("inputs", [])
                    if isinstance(item, dict)
                }
                if not expected_pair_outputs.issubset(
                    committed_excel_inputs
                ):
                    raise SupplementalAcceptanceError(
                        "child_excel 清单没有纳入本轮新近交工作簿"
                    )
                for stage in ("data", "analysis", "child_excel"):
                    validation = validate_child_stage(
                        child_path,
                        stage,
                        expected_task_id=str(task["task_id"]),
                        expected_farm_code=str(task["farm_code"]),
                        verification="full",
                    )
                    if not validation.get("valid"):
                        raise SupplementalAcceptanceError(
                            f"{stage} 阶段 full 校验失败"
                        )

                phase = "supplemental_manifest"
                manifest = _build_supplemental_manifest(
                    child_path,
                    staged_summaries,
                    cow_self_rows=cow_self_rows,
                    matching_rows=matching_rows,
                    inventory_items=inventory_items,
                    skipped_bulls=matching_skipped_bulls,
                    report_path=target_report,
                    pair_inbreeding_paths=target_pair_inbreeding_paths,
                    pair_inbreeding_rows=pair_inbreeding_rows,
                )
                _atomic_write_json(supplemental_manifest, manifest)
                transaction.commit()

        _atomic_write_json(
            result_path,
            {
                "schema_version": CHILD_RESULT_SCHEMA,
                "success": True,
                "counts": {
                    "cow_self_rows": int(cow_self_rows),
                    "matching_rows": int(matching_rows),
                    "inventory_items": int(inventory_items),
                    "matching_skipped_bulls": int(
                        matching_skipped_bulls
                    ),
                    "candidate_inbreeding_rows": int(
                        pair_inbreeding_rows["candidate"]
                    ),
                    "mated_inbreeding_rows": int(
                        pair_inbreeding_rows["mated"]
                    ),
                },
            },
        )
        return 0
    except BaseException as exc:
        _atomic_write_json(
            result_path,
            {
                "schema_version": CHILD_RESULT_SCHEMA,
                "success": False,
                "phase": phase,
                "error_type": type(exc).__name__,
            },
        )
        return 1
    finally:
        if run_root is not None:
            shutil.rmtree(run_root, ignore_errors=True)
        try:
            from core.data.update_manager import reset_pedigree_db

            reset_pedigree_db()
        except Exception:
            pass


def _run_child_subprocess(
    group_path: Path,
    task: Dict[str, Any],
) -> Dict[str, Any]:
    result_directory = group_path / "group_store" / "supplemental_child_results"
    result_directory.mkdir(parents=True, exist_ok=True)
    result_path = result_directory / f".{uuid.uuid4().hex}.json"
    command = [
        sys.executable,
        str(Path(__file__).resolve()),
        "--group",
        str(group_path),
        "--_child-task-id",
        str(task["task_id"]),
        "--_result-file",
        str(result_path),
    ]
    environment = os.environ.copy()
    environment.setdefault("QT_QPA_PLATFORM", "offscreen")
    environment["PYTHONUNBUFFERED"] = "1"
    try:
        completed = subprocess.run(
            command,
            cwd=REPOSITORY_ROOT,
            env=environment,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        if not result_path.is_file():
            return {
                "success": False,
                "phase": "child_process",
                "error_type": f"Exit{completed.returncode}",
            }
        payload = _read_json(result_path)
        if payload.get("schema_version") != CHILD_RESULT_SCHEMA:
            return {
                "success": False,
                "phase": "child_protocol",
                "error_type": "SchemaMismatch",
            }
        return payload
    finally:
        result_path.unlink(missing_ok=True)


def _write_parent_result(
    group_path: Path,
    *,
    success_count: int,
    failed_count: int,
    group_report_path: Optional[Path],
    mode: str,
) -> Path:
    result_name = (
        "supplemental_acceptance_canary_result.json"
        if mode == "canary"
        else RESULT_FILENAME
    )
    target = group_path / "group_store" / result_name
    _atomic_write_json(
        target,
        {
            "schema_version": 1,
            "generated_at": _utc_timestamp(),
            "mode": mode,
            "success": (
                failed_count == 0
                and (mode == "canary" or group_report_path is not None)
            ),
            "farm_count": int(success_count + failed_count),
            "success_count": int(success_count),
            "failed_count": int(failed_count),
            "group_excel": (
                str(group_report_path.resolve())
                if group_report_path is not None
                else ""
            ),
            "hmy_push": False,
            "ppt_generated": False,
        },
    )
    return target


def _generate_group_excel(group_path: Path) -> Path:
    from core.group_report.excel_generator import GroupExcelReportGenerator

    last_bucket = {"value": -1}
    aggregate_output = sys.stdout

    def progress(value: int, _message: str) -> None:
        bucket = min(10, max(0, int(value) // 10))
        if bucket > last_bucket["value"]:
            last_bucket["value"] = bucket
            # 底层详细输出处于重定向状态；明确写回进入静默区前保存的
            # 标准输出，只暴露聚合百分比，不转发带牧场名的 message。
            print(
                f"[汇总] 进度 {bucket * 10}%",
                file=aggregate_output,
                flush=True,
            )

    with _silence_sensitive_output():
        generator = GroupExcelReportGenerator(
            group_path,
            service_staff="",
            progress_callback=progress,
        )
        success, result = generator.generate()
    if not success:
        raise SupplementalAcceptanceError("牧场组汇总 Excel 生成失败")
    report_path = Path(result)
    _xlsx_summary(report_path)
    return report_path


def _execute(
    group_path: Path,
    tasks: Sequence[Dict[str, Any]],
    *,
    canary: bool = False,
) -> int:
    if _has_active_group_lease(group_path):
        raise SupplementalAcceptanceError(
            "牧场组仍有运行任务；请等待父任务完全结束后再执行"
        )

    from core.group_tasks.lease_heartbeat import GroupLeaseHeartbeat
    from utils.group_task_store import GroupTaskStore

    store = GroupTaskStore(
        group_path / "group_store" / "group_tasks.sqlite3"
    )
    revision = store.get_selection_revision()
    lease = store.acquire_run_lease(
        f"supplemental-acceptance:{os.getpid()}:{uuid.uuid4()}",
        run_kind=(
            "supplemental_acceptance_canary"
            if canary
            else "supplemental_acceptance"
        ),
        lease_seconds=600,
        expected_selection_revision=revision,
    )
    if lease is None:
        raise SupplementalAcceptanceError("无法取得牧场组排他运行锁")
    heartbeat = GroupLeaseHeartbeat(store, lease, lease_seconds=600)
    heartbeat.start()

    success_count = 0
    failures: list[Dict[str, Any]] = []
    try:
        total = len(tasks)
        print(f"开始串行补跑：共 {total} 个牧场")
        for index, task in enumerate(tasks, start=1):
            heartbeat.check()
            reusable = (
                None
                if canary
                else _validated_existing_supplemental_result(
                    group_path,
                    task,
                )
            )
            if reusable is not None:
                success_count += 1
                counts = reusable["counts"]
                print(
                    f"[{index}/{total}] 已完成且校验一致，复用结果："
                    f"母牛自身近交 {counts['cow_self_rows']:,} 行，"
                    "备选近交 "
                    f"{counts['candidate_inbreeding_rows']:,} 行，"
                    "已配近交 "
                    f"{counts['mated_inbreeding_rows']:,} 行，"
                    f"个体选配 {counts['matching_rows']:,} 行"
                )
                continue
            print(f"[{index}/{total}] 开始")
            result = _run_child_subprocess(group_path, task)
            if result.get("success"):
                success_count += 1
                counts = result.get("counts") or {}
                print(
                    f"[{index}/{total}] 完成："
                    f"母牛自身近交 {int(counts.get('cow_self_rows', 0)):,} 行，"
                    "备选近交 "
                    f"{int(counts.get('candidate_inbreeding_rows', 0)):,} 行，"
                    "已配近交 "
                    f"{int(counts.get('mated_inbreeding_rows', 0)):,} 行，"
                    f"个体选配 {int(counts.get('matching_rows', 0)):,} 行"
                    f"（跳过缺少指数公牛 "
                    f"{int(counts.get('matching_skipped_bulls', 0))} 头）"
                )
            else:
                failures.append(result)
                print(
                    f"[{index}/{total}] 失败："
                    f"阶段 {result.get('phase', 'unknown')}，"
                    f"类型 {result.get('error_type', 'Error')}"
                )
        heartbeat.check()
    finally:
        heartbeat.stop(timeout=30, release=True)

    if failures:
        result_path = _write_parent_result(
            group_path,
            success_count=success_count,
            failed_count=len(failures),
            group_report_path=None,
            mode="canary" if canary else "full",
        )
        print(
            f"补跑结束：成功 {success_count} 个，失败 {len(failures)} 个；"
            "存在失败时未生成新的牧场组汇总 Excel。"
        )
        print(f"聚合结果：{result_path}")
        return 1

    if canary:
        result_path = _write_parent_result(
            group_path,
            success_count=success_count,
            failed_count=0,
            group_report_path=None,
            mode="canary",
        )
        print(
            "单场 canary 的补跑、原子提升及 full 校验通过；"
            "canary 模式不生成牧场组汇总 Excel。"
        )
        print(f"聚合结果：{result_path}")
        return 0

    print("全部单牧场补跑及 full 校验通过，开始重生成牧场组汇总 Excel。")
    group_report = _generate_group_excel(group_path)
    result_path = _write_parent_result(
        group_path,
        success_count=success_count,
        failed_count=0,
        group_report_path=group_report,
        mode="full",
    )
    print(f"补跑完成：成功 {success_count} 个，失败 0 个。")
    print(f"牧场组汇总 Excel：{group_report}")
    print(f"聚合结果：{result_path}")
    return 0


def _dry_run(
    group_path: Path,
    tasks: Sequence[Dict[str, Any]],
    *,
    canary: bool = False,
    all_tasks: Optional[Sequence[Dict[str, Any]]] = None,
) -> int:
    readiness_tasks = list(all_tasks or tasks)
    terminal_count = sum(
        str(task.get("status") or "") in SUCCESS_TASK_STATUSES
        for task in readiness_tasks
    )
    active_lease = _has_active_group_lease(group_path)
    print("补充验收 dry-run（未读取业务 Excel，未写入任何结果）")
    print(f"本次执行范围：{len(tasks)} 个牧场")
    print(
        f"父组成功终态：{terminal_count}/{len(readiness_tasks)}"
    )
    print(f"父任务运行锁：{'仍在运行' if active_lease else '空闲'}")
    print(
        "每个牧场计划：母牛自身近交 → 强制重跑备选/已配近交 → "
        "个体选配 → 单场 Excel → analysis/child_excel full 校验"
    )
    if canary:
        print("canary 模式：单场 full 校验后停止，不生成牧场组汇总 Excel")
    else:
        print("全部成功后计划：重生成牧场组汇总 Excel")
    print("明确禁用：慧牧云结果推送、PPT 生成")
    if terminal_count != len(readiness_tasks) or active_lease:
        print("当前仅可预览；父任务终态且运行锁释放后方可 --execute。")
        return 2
    print("前置状态已就绪；如需真实执行，请显式追加 --execute。")
    return 0


def _find_child_task(
    tasks: Sequence[Dict[str, Any]],
    task_id: str,
) -> Dict[str, Any]:
    matches = [
        task for task in tasks
        if str(task.get("task_id") or "") == str(task_id)
    ]
    if len(matches) != 1:
        raise SupplementalAcceptanceError("内部子任务不存在或不唯一")
    return matches[0]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "牧场组终态后串行补跑母牛自身近交、备选/已配近交、"
            "个体选配和单场 Excel；"
            "默认仅 dry-run，显式 --execute 才会执行。"
        )
    )
    parser.add_argument(
        "--group",
        type=Path,
        required=True,
        help="牧场组项目目录",
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--dry-run",
        action="store_true",
        help="只检查父任务终态并展示计划（默认）",
    )
    mode.add_argument(
        "--execute",
        action="store_true",
        help="真实串行执行补跑、原子提升、full 校验和组汇总",
    )
    parser.add_argument(
        "--canary-task-id",
        help=(
            "仅选择指定 task UUID 做单场 canary；成功后不生成组汇总"
        ),
    )
    parser.add_argument(
        "--_child-task-id",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--_result-file",
        type=Path,
        help=argparse.SUPPRESS,
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        group_path, _metadata, tasks = _resolve_group(args.group)
        if args._child_task_id:
            if args._result_file is None:
                raise SupplementalAcceptanceError("内部结果文件未指定")
            _validate_terminal_tasks(tasks)
            task = _find_child_task(tasks, args._child_task_id)
            return _process_one_child(
                group_path,
                task,
                args._result_file,
            )

        selected_tasks = tasks
        canary = bool(args.canary_task_id)
        if canary:
            selected_tasks = [
                _find_child_task(tasks, args.canary_task_id)
            ]

        if args.execute:
            _validate_terminal_tasks(tasks)
            return _execute(
                group_path,
                selected_tasks,
                canary=canary,
            )
        return _dry_run(
            group_path,
            selected_tasks,
            canary=canary,
            all_tasks=tasks,
        )
    except SupplementalAcceptanceError as exc:
        print(f"无法继续：{exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())

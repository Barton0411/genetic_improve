#!/usr/bin/env python3
"""牧场组全流程结果的只读验收验证器。

这个工具只允许在牧场组任务全部进入成功终态、且没有活动运行租约后读取
XLSX。它不会使用 :class:`utils.group_task_store.GroupTaskStore`，因为该类
构造时会初始化数据库；任务状态库始终通过 SQLite ``mode=ro`` 打开。

验证范围包括：

* 父/子项目身份、任务与阶段终态；
* 三阶段 committed manifest 的官方 ``full`` 校验；
* 原始母牛 -> 标准化母牛 -> 性状结果 -> 指数结果的行数与牛号血缘；
* 配种记录 -> 已配公牛性状结果的逐记录血缘；
* 备选公牛导入 -> 性状 -> 指数的逐记录血缘；
* 正式/补充 XLSX 的 ZIP、公式错误、空表与百分比格式异常；
* 当前牧场组报告指针、不可变报告包和完整明细分卷的无丢失对账；
* 报告包发布快照与当前任务选择、阶段 manifest 的一致性。

输出只包含聚合计数、不可逆摘要和路径，不输出牛号、冻精号、凭据或
接口会话信息。默认要求调用方显式提供项目目录外的 ``--output-dir``。
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import sqlite3
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, Iterator, List, Mapping, Optional, Sequence
from urllib.parse import quote
from xml.etree import ElementTree

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.group_report.artifact_inventory import inspect_xlsx_structure
from core.group_report.publication_batch import (
    GroupReportPublicationError,
    validate_current_group_report_pointer,
)
from core.group_tasks.stage_manifest import (
    stream_sha256,
    validate_stage_manifest,
)
from core.group_tasks.stage_policy import (
    STAGE_ORDER,
    stage_manifest_path,
)
from core.data.processor import format_naab_number
from scripts.acceptance_formula_checks import (
    FormulaValidationError,
    validate_cow_formulas,
)
from scripts.acceptance_scope_checks import validate_child_scope_artifacts


SCHEMA_VERSION = 1
TERMINAL_SUCCESS = {"completed", "completed_with_warning"}
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!")
IDENTIFIER_MISSING = {"", "nan", "none", "null", "nat", "<na>", "n/a"}
COW_STABLE_FIELD_ALIASES = {
    "sire": (
        "sire",
        "父号",
        "父亲号",
        "父亲",
        "公牛号",
        "father",
        "father_num",
        "fathernum",
    ),
    "dam": (
        "dam",
        "母号",
        "母亲号",
        "母亲牛号",
        "mother",
        "mother_num",
        "mothernum",
    ),
    "mgs": (
        "mgs",
        "外祖父",
        "外祖父号",
        "maternal_grandsire",
        "maternal grandsire",
        "grandpa",
    ),
    "mmgs": (
        "mmgs",
        "外曾外祖父",
        "外曾外祖父号",
        "maternal_great_grandsire",
        "maternal great grandsire",
    ),
}
EXCEL_MAX_DATA_ROWS = 1_048_575
BUILTIN_PERCENT_FORMAT_IDS = {9, 10}
TEXT_PERCENT_PATTERN = re.compile(
    r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)\s*%$"
)

CORE_ANALYSIS_FILES = (
    "processed_cow_data_key_traits_detail.xlsx",
    "processed_cow_data_key_traits_final.xlsx",
    "processed_cow_data_key_traits_scores_pedigree.xlsx",
    "sire_traits_mean_by_cow_birth_year.xlsx",
    "processed_index_cow_index_scores.xlsx",
    "关键育种性状分析结果.xlsx",
    "系谱识别分析结果.xlsx",
    "母牛近交系数分析结果.xlsx",
)
OPTIONAL_ANALYSIS_FILES = (
    # 单牧场子项目不一定生成“分牧场年度均值”；存在时仍纳入健康
    # 扫描。跨牧场完整性由最终组报告包的来源与分卷清单负责。
    "sire_traits_mean_by_cow_birth_year_by_farm.xlsx",
)
MATCHING_FILES = (
    "个体选配推荐矩阵.xlsx",
    "个体选配报告.xlsx",
    "individual_mating_report.xlsx",
)
CANDIDATE_FILES = (
    "processed_bull_data_key_traits.xlsx",
    "processed_index_bull_scores.xlsx",
)
MATED_FILES = ("processed_mated_bull_traits.xlsx",)
METRIC_PROFILE_FILES = {
    "processed_cow_data_key_traits_final.xlsx",
    "processed_index_cow_index_scores.xlsx",
    "processed_bull_data_key_traits.xlsx",
    "processed_index_bull_scores.xlsx",
    "processed_mated_bull_traits.xlsx",
    "sire_traits_mean_by_cow_birth_year.xlsx",
    "sire_traits_mean_by_cow_birth_year_by_farm.xlsx",
    "关键育种性状分析结果.xlsx",
}
METRIC_HEADER_TOKENS = (
    "NM$",
    "TPI",
    "MILK",
    "FAT",
    "FAT %",
    "PROT",
    "PROT%",
    "SCS",
    "PL",
    "DPR",
    "PTAT",
    "UDC",
    "FLC",
    "RFI",
    "综合指数",
)


class AcceptanceBlocked(RuntimeError):
    """任务仍在运行或静态门禁不满足，禁止打开项目 XLSX。"""


@dataclass
class Issue:
    severity: str
    scope: str
    code: str
    message: str
    farm_code: str = ""
    relative_path: str = ""


class ResultBuilder:
    def __init__(self, project_path: Path):
        self.project_path = Path(project_path).resolve()
        self.issues: List[Issue] = []
        self.stage_rows: List[Dict[str, Any]] = []
        self.lineage_rows: List[Dict[str, Any]] = []
        self.file_rows: List[Dict[str, Any]] = []
        self.farm_rows: List[Dict[str, Any]] = []
        self.group_report: Dict[str, Any] = {}
        self.gate: Dict[str, Any] = {}

    def add(
        self,
        severity: str,
        scope: str,
        code: str,
        message: str,
        *,
        farm_code: str = "",
        relative_path: str = "",
    ) -> None:
        self.issues.append(
            Issue(
                severity=str(severity),
                scope=str(scope),
                code=str(code),
                message=str(message),
                farm_code=str(farm_code),
                relative_path=str(relative_path),
            )
        )

    def payload(self, *, blocked: bool = False) -> Dict[str, Any]:
        errors = sum(item.severity == "error" for item in self.issues)
        warnings = sum(item.severity == "warning" for item in self.issues)
        status = "blocked" if blocked else ("passed" if errors == 0 else "failed")
        return {
            "schema_version": SCHEMA_VERSION,
            "kind": "multi_farm_acceptance_validation",
            "generated_at": _utc_now(),
            "project_path": str(self.project_path),
            "status": status,
            "counts": {
                "farms": len(self.farm_rows),
                "stage_checks": len(self.stage_rows),
                "lineage_checks": len(self.lineage_rows),
                "xlsx_files": len(self.file_rows),
                "errors": errors,
                "warnings": warnings,
            },
            "gate": self.gate,
            "farms": self.farm_rows,
            "stages": self.stage_rows,
            "lineage": self.lineage_rows,
            "files": self.file_rows,
            "group_report": self.group_report,
            "issues": [asdict(item) for item in self.issues],
        }


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _read_json(path: Path, label: str) -> Dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"{label}无法读取（{type(exc).__name__}）") from exc
    if not isinstance(value, dict):
        raise ValueError(f"{label}根节点不是对象")
    return value


def _normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value).strip()
    if text.casefold() in IDENTIFIER_MISSING:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _is_blank_row(row: Sequence[Any]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row
    )


def _header_key(value: Any) -> str:
    return str(value or "").strip().casefold()


def _first_header(
    headers: Sequence[Any],
    candidates: Sequence[str],
) -> Optional[int]:
    keys = [_header_key(value) for value in headers]
    wanted = {_header_key(value) for value in candidates}
    return next((index for index, key in enumerate(keys) if key in wanted), None)


def _safe_relative(project: Path, value: Any, label: str) -> Path:
    text = str(value or "").strip()
    if not text or "\\" in text:
        raise ValueError(f"{label}不是有效相对路径")
    pure = PurePosixPath(text)
    if pure.is_absolute() or any(part in {"", ".", ".."} for part in pure.parts):
        raise ValueError(f"{label}不是安全相对路径")
    candidate = project.joinpath(*pure.parts)
    resolved = candidate.resolve()
    try:
        resolved.relative_to(project.resolve())
    except ValueError as exc:
        raise ValueError(f"{label}超出牧场组项目") from exc
    current = project
    for part in pure.parts:
        current = current / part
        if current.is_symlink():
            raise ValueError(f"{label}不能经过符号链接")
    return resolved


def _sqlite_ro_connection(path: Path) -> sqlite3.Connection:
    if not path.is_file() or path.is_symlink():
        raise AcceptanceBlocked("牧场组任务状态库不存在或为符号链接")
    uri = f"file:{quote(str(path.resolve()))}?mode=ro"
    connection = sqlite3.connect(uri, uri=True, timeout=30)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA query_only = ON")
    connection.execute("PRAGMA busy_timeout = 30000")
    return connection


def _parse_json_object(value: Any) -> Dict[str, Any]:
    try:
        decoded = json.loads(str(value or "{}"))
    except json.JSONDecodeError:
        return {}
    return decoded if isinstance(decoded, dict) else {}


def _read_task_state(project: Path) -> Dict[str, Any]:
    database = project / "group_store" / "group_tasks.sqlite3"
    with _sqlite_ro_connection(database) as connection:
        task_rows = connection.execute(
            """
            SELECT task_id, farm_code, farm_name, relative_path,
                   source_kind, source_system, included_in_summary,
                   status, current_stage, progress, attempt, error,
                   metadata_json, sort_order, completed_at
            FROM group_tasks
            ORDER BY sort_order, task_id
            """
        ).fetchall()
        stage_rows = connection.execute(
            """
            SELECT task_id, stage, required, status, progress, attempt,
                   output_path, detail_count, completed_at
            FROM group_task_stages
            ORDER BY task_id, CASE stage
                WHEN 'data' THEN 1
                WHEN 'analysis' THEN 2
                WHEN 'child_excel' THEN 3
                ELSE 99 END
            """
        ).fetchall()
        control = connection.execute(
            """
            SELECT selection_revision, lease_token, lease_owner_id,
                   lease_run_kind, lease_expires_at
            FROM group_run_control WHERE singleton_id = 1
            """
        ).fetchone()

    stages: Dict[str, Dict[str, Dict[str, Any]]] = defaultdict(dict)
    for row in stage_rows:
        stages[str(row["task_id"])][str(row["stage"])] = {
            "required": bool(row["required"]),
            "status": str(row["status"] or ""),
            "progress": float(row["progress"] or 0),
            "attempt": int(row["attempt"] or 0),
            "output_path": str(row["output_path"] or ""),
            "detail_count": row["detail_count"],
            "completed_at": str(row["completed_at"] or ""),
        }

    tasks = []
    for row in task_rows:
        metadata = _parse_json_object(row["metadata_json"])
        tasks.append(
            {
                "task_id": str(row["task_id"]),
                "farm_code": str(row["farm_code"] or ""),
                "farm_name": str(row["farm_name"] or ""),
                "relative_path": str(row["relative_path"] or ""),
                "source_kind": str(row["source_kind"] or ""),
                "source_system": str(row["source_system"] or ""),
                "included_in_summary": bool(row["included_in_summary"]),
                "status": str(row["status"] or ""),
                "current_stage": str(row["current_stage"] or ""),
                "progress": float(row["progress"] or 0),
                "attempt": int(row["attempt"] or 0),
                "completed_at": str(row["completed_at"] or ""),
                "identity_metadata": {
                    key: metadata.get(key)
                    for key in (
                        "api_farmcode",
                        "farm_number",
                        "display_name",
                        "source_farm_name",
                    )
                    if key in metadata
                },
                "stages": stages.get(str(row["task_id"]), {}),
            }
        )
    control_value = dict(control) if control is not None else {}
    # 任何租约值都只转成布尔/类型，绝不复制 token 或 owner。
    return {
        "tasks": tasks,
        "selection_revision": int(
            control_value.get("selection_revision", -1)
        ),
        "lease_present": bool(control_value.get("lease_token")),
        "lease_run_kind": str(control_value.get("lease_run_kind") or ""),
        "lease_expires_at": str(control_value.get("lease_expires_at") or ""),
    }


def _lease_is_active(state: Mapping[str, Any]) -> bool:
    if not state.get("lease_present"):
        return False
    expires = str(state.get("lease_expires_at") or "")
    if not expires:
        return True
    try:
        moment = datetime.fromisoformat(expires.replace("Z", "+00:00"))
        if moment.tzinfo is None:
            moment = moment.replace(tzinfo=timezone.utc)
    except ValueError:
        return True
    return moment.astimezone(timezone.utc) > datetime.now(timezone.utc)


def _parent_task_map(metadata: Mapping[str, Any]) -> Dict[str, Dict[str, Any]]:
    result = {}
    tasks = metadata.get("group_tasks")
    if not isinstance(tasks, list):
        return result
    for task in tasks:
        if isinstance(task, dict) and task.get("task_id"):
            result[str(task["task_id"])] = task
    return result


def _expected_identity(
    task: Mapping[str, Any],
    parent_task: Mapping[str, Any],
) -> Dict[str, str]:
    database_metadata = task.get("identity_metadata")
    if not isinstance(database_metadata, Mapping):
        database_metadata = {}
    parent_identity = parent_task.get("metadata")
    if not isinstance(parent_identity, Mapping):
        parent_identity = {}

    def first(*values: Any) -> str:
        for value in values:
            normalized = _normalize_identifier(value)
            if normalized:
                return normalized
        return ""

    source_system = str(
        task.get("source_system") or parent_task.get("source_system") or ""
    )
    source_kind = str(
        task.get("source_kind") or parent_task.get("source_kind") or ""
    )
    farm_code = first(task.get("farm_code"), parent_task.get("farm_code"))
    api_farmcode = first(
        database_metadata.get("api_farmcode"),
        parent_identity.get("api_farmcode"),
        parent_task.get("api_farmcode"),
        farm_code if source_kind != "local" else "",
    )
    farm_number = first(
        database_metadata.get("farm_number"),
        parent_identity.get("farm_number"),
        parent_task.get("farm_number"),
    )
    farm_name = first(
        task.get("farm_name"),
        database_metadata.get("display_name"),
        parent_task.get("farm_name"),
        parent_identity.get("display_name"),
    )
    return {
        "farm_code": farm_code,
        "api_farmcode": api_farmcode,
        "farm_number": farm_number,
        "farm_name": farm_name,
        "source_system": source_system,
        "source_kind": source_kind,
    }


def _static_gate(
    project_path: Path,
    result: ResultBuilder,
) -> tuple[Dict[str, Any], List[Dict[str, Any]]]:
    project = Path(project_path)
    if project.is_symlink() or not project.is_dir():
        raise AcceptanceBlocked("牧场组项目目录不存在或为符号链接")
    project = project.resolve()
    metadata = _read_json(
        project / "project_metadata.json",
        "父项目元数据",
    )
    if metadata.get("project_type") != "multi_farm_group":
        raise AcceptanceBlocked("目标目录不是牧场组父项目")

    state = _read_task_state(project)
    included = [
        task for task in state["tasks"] if task["included_in_summary"]
    ]
    if not included:
        raise AcceptanceBlocked("没有纳入汇总范围的牧场任务")
    if _lease_is_active(state):
        raise AcceptanceBlocked("牧场组仍有活动运行租约，禁止打开 XLSX")

    nonterminal = [
        task
        for task in included
        if task["status"] not in TERMINAL_SUCCESS
    ]
    stage_nonterminal = []
    for task in included:
        required = [
            stage
            for stage in task["stages"].values()
            if stage.get("required")
        ]
        if not required:
            stage_nonterminal.append(task["farm_code"])
            continue
        if any(stage["status"] not in TERMINAL_SUCCESS for stage in required):
            stage_nonterminal.append(task["farm_code"])
    if nonterminal or stage_nonterminal:
        raise AcceptanceBlocked(
            "牧场组任务或必需阶段尚未全部进入成功终态，禁止打开 XLSX"
        )

    pointer_path = project / "group_store" / "current_group_report.json"
    if not pointer_path.is_file() or pointer_path.is_symlink():
        raise AcceptanceBlocked("当前牧场组正式报告指针不存在")

    parent_tasks = _parent_task_map(metadata)
    prepared = []
    for task in included:
        child = _safe_relative(
            project,
            task["relative_path"],
            "牧场子项目目录",
        )
        if not child.is_dir():
            raise AcceptanceBlocked("牧场子项目目录不存在")
        child_metadata = _read_json(
            child / "project_metadata.json",
            "牧场子项目元数据",
        )
        parent_task = parent_tasks.get(task["task_id"], {})
        identity = _expected_identity(task, parent_task)
        if (
            child_metadata.get("project_type") != "group_child"
            or str(child_metadata.get("group_task_id") or "")
            != task["task_id"]
            or _normalize_identifier(
                child_metadata.get("group_farm_code")
            )
            != identity["farm_code"]
        ):
            raise AcceptanceBlocked("牧场子项目身份与任务状态库不一致")
        farms = child_metadata.get("farms")
        if not isinstance(farms, list) or len(farms) != 1:
            raise AcceptanceBlocked("牧场子项目必须且只能包含一个牧场")
        child_farm = farms[0] if isinstance(farms[0], Mapping) else {}
        if identity["api_farmcode"] and _normalize_identifier(
            child_metadata.get("group_api_farmcode")
            or child_farm.get("api_farmcode")
            or child_farm.get("code")
        ) != identity["api_farmcode"]:
            raise AcceptanceBlocked("牧场子项目 API farmcode 身份不一致")
        if identity["farm_number"] and _normalize_identifier(
            child_metadata.get("group_farm_number")
            or child_farm.get("farm_number")
        ) != identity["farm_number"]:
            raise AcceptanceBlocked("牧场子项目业务牧场编号身份不一致")
        child_name = _normalize_identifier(
            child_farm.get("display_name")
            or child_farm.get("name")
        )
        if identity["farm_name"] and child_name != identity["farm_name"]:
            raise AcceptanceBlocked("牧场子项目牧场名称身份不一致")
        prepared.append(
            {
                **task,
                "child_path": child,
                "child_metadata": child_metadata,
                "identity": identity,
            }
        )

    if state.get("lease_present"):
        result.add(
            "warning",
            "gate",
            "expired_lease_record",
            "任务库保留了已过期租约记录，但当前没有活动租约",
        )
    result.gate = {
        "passed": True,
        "project_type": "multi_farm_group",
        "included_tasks": len(prepared),
        "selection_revision": state["selection_revision"],
        "all_tasks_terminal": True,
        "all_required_stages_terminal": True,
        "active_lease": False,
        "current_report_pointer_present": True,
    }
    return state, prepared


def _relative_to_project(project: Path, path: Path) -> str:
    return path.resolve().relative_to(project.resolve()).as_posix()


def _manifest_paths(
    child: Path,
    manifest: Mapping[str, Any],
) -> Iterator[Path]:
    for group in ("inputs", "outputs"):
        entries = manifest.get(group)
        if not isinstance(entries, list):
            continue
        for entry in entries:
            if not isinstance(entry, Mapping):
                continue
            relative = str(entry.get("relative_path") or "")
            if not relative.casefold().endswith(".xlsx"):
                continue
            yield _safe_relative(child, relative, "阶段产物")


def _validate_stage_manifests(
    project: Path,
    tasks: Sequence[Dict[str, Any]],
    result: ResultBuilder,
) -> set[Path]:
    xlsx_paths: set[Path] = set()
    for task in tasks:
        child = task["child_path"]
        for stage_name in STAGE_ORDER:
            stage = task["stages"].get(stage_name, {})
            if not stage.get("required"):
                continue
            validation = validate_stage_manifest(
                child,
                child / stage_manifest_path(stage_name),
                expected_task_id=task["task_id"],
                expected_farm_code=task["farm_code"],
                expected_stage=stage_name,
                verification="full",
            )
            valid = bool(validation.get("valid"))
            result.stage_rows.append(
                {
                    "farm_code": task["farm_code"],
                    "farm_name": task["farm_name"],
                    "stage": stage_name,
                    "database_status": stage.get("status", ""),
                    "manifest_status": str(validation.get("status") or ""),
                    "valid": valid,
                    "issue_count": len(validation.get("issues") or []),
                    "attempt": int(stage.get("attempt", 0) or 0),
                }
            )
            if not valid:
                for item in validation.get("issues") or []:
                    result.add(
                        "error",
                        "stage_manifest",
                        str(item.get("code") or "manifest_invalid"),
                        str(item.get("message") or "阶段 manifest 校验失败"),
                        farm_code=task["farm_code"],
                        relative_path=stage_manifest_path(stage_name).as_posix(),
                    )
                continue
            manifest = validation.get("manifest")
            if isinstance(manifest, Mapping):
                xlsx_paths.update(_manifest_paths(child, manifest))
    return xlsx_paths


def _first_sheet_table(path: Path) -> tuple[List[Any], Iterator[Sequence[Any]], Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration:
        headers = []
    return headers, iterator, workbook


def _first_sheet_headers(path: Path) -> List[Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
        try:
            return list(next(iterator))
        except StopIteration:
            return []
        finally:
            # 让只读 worksheet 的 XML 生成器到达 StopIteration，确保
            # openpyxl 立即释放底层 ZipExtFile，而不是等待垃圾回收。
            for _unused in iterator:
                pass
            close = getattr(iterator, "close", None)
            if callable(close):
                close()
    finally:
        workbook.close()


def _count_table_rows(path: Path) -> int:
    headers, rows, workbook = _first_sheet_table(path)
    del headers
    try:
        return sum(1 for row in rows if not _is_blank_row(row))
    finally:
        workbook.close()


def _identity_expectations(
    headers: Sequence[Any],
    expected: Mapping[str, str],
) -> Dict[str, tuple[int, str]]:
    definitions = (
        (("API farmcode", "farm_code", "farmcode"), "api_farmcode"),
        (("牧场编号",), "farm_number"),
        (("牧场名称",), "farm_name"),
    )
    result = {}
    for candidates, key in definitions:
        index = _first_header(headers, candidates)
        expected_value = _normalize_identifier(expected.get(key))
        if index is not None and expected_value:
            result[key] = (index, expected_value)
    return result


def _required_identity_keys(expected: Mapping[str, str]) -> set[str]:
    source_system = str(expected.get("source_system") or "")
    keys = {"farm_name"} if _normalize_identifier(expected.get("farm_name")) else set()
    if source_system == "慧牧云":
        for key in ("api_farmcode", "farm_number"):
            if _normalize_identifier(expected.get(key)):
                keys.add(key)
    elif _normalize_identifier(expected.get("farm_number")):
        keys.add("farm_number")
    return keys


def _ensure_id_table(connection: sqlite3.Connection, name: str) -> None:
    if not re.fullmatch(r"[a-z][a-z0-9_]*", name):
        raise ValueError("内部血缘表名无效")
    connection.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {name} (
            identifier_hash BLOB PRIMARY KEY,
            source_count INTEGER NOT NULL DEFAULT 0,
            middle_count INTEGER NOT NULL DEFAULT 0,
            target_count INTEGER NOT NULL DEFAULT 0
        ) WITHOUT ROWID
        """
    )


def _profile_identifier_table(
    path: Path,
    *,
    id_candidates: Sequence[str],
    expected_identity: Mapping[str, str],
    connection: sqlite3.Connection,
    table: str,
    slot: str,
) -> Dict[str, Any]:
    if slot not in {"source", "middle", "target"}:
        raise ValueError("内部血缘槽位无效")
    headers, rows, workbook = _first_sheet_table(path)
    try:
        id_index = _first_header(headers, id_candidates)
        if id_index is None:
            raise ValueError("缺少业务标识符列")
        expectations = _identity_expectations(headers, expected_identity)
        required_identity_keys = _required_identity_keys(expected_identity)
        stats = {
            "rows": 0,
            "identifier_count": 0,
            "blank_identifier_count": 0,
            "identity_columns": len(expectations),
            "identity_missing_columns": len(
                required_identity_keys.difference(expectations)
            ),
            "identity_mismatch_rows": 0,
            "identity_blank_rows": 0,
            "identity_unique_counts": defaultdict(set),
        }
        batch: List[tuple[bytes]] = []
        for row in rows:
            if _is_blank_row(row):
                continue
            stats["rows"] += 1
            identifier = _normalize_identifier(
                row[id_index] if id_index < len(row) else None
            )
            if identifier:
                stats["identifier_count"] += 1
                batch.append(
                    (hashlib.sha256(identifier.encode("utf-8")).digest(),)
                )
            else:
                stats["blank_identifier_count"] += 1

            for key, (column_index, expected_value) in expectations.items():
                actual = _normalize_identifier(
                    row[column_index] if column_index < len(row) else None
                )
                if actual:
                    stats["identity_unique_counts"][key].add(
                        hashlib.sha256(actual.encode("utf-8")).digest()
                    )
                else:
                    stats["identity_blank_rows"] += 1
                if actual != expected_value:
                    stats["identity_mismatch_rows"] += 1

            if len(batch) >= 2_000:
                _insert_identifier_batch(connection, table, slot, batch)
                batch.clear()
        if batch:
            _insert_identifier_batch(connection, table, slot, batch)
        connection.commit()
        return {
            key: (
                {
                    name: len(values)
                    for name, values in stats["identity_unique_counts"].items()
                }
                if key == "identity_unique_counts"
                else value
            )
            for key, value in stats.items()
        }
    finally:
        workbook.close()


def _insert_identifier_batch(
    connection: sqlite3.Connection,
    table: str,
    slot: str,
    batch: Sequence[tuple[bytes]],
) -> None:
    connection.executemany(
        f"""
        INSERT INTO {table} (identifier_hash, {slot}_count)
        VALUES (?, 1)
        ON CONFLICT(identifier_hash) DO UPDATE
        SET {slot}_count = {slot}_count + 1
        """,
        batch,
    )


def _multiset_fingerprint_state() -> Dict[str, int]:
    return {
        "row_count": 0,
        "hash_sum": 0,
        "hash_square_sum": 0,
        "hash_xor": 0,
    }


def _add_multiset_digest(
    state: Dict[str, int],
    digest: bytes,
    *,
    count: int = 1,
) -> None:
    occurrences = int(count)
    if occurrences <= 0:
        return
    numeric = int.from_bytes(digest, "big")
    mask = (1 << 256) - 1
    state["row_count"] += occurrences
    state["hash_sum"] = (
        state["hash_sum"] + numeric * occurrences
    ) & mask
    state["hash_square_sum"] = (
        state["hash_square_sum"] + numeric * numeric * occurrences
    ) & mask
    if occurrences % 2:
        state["hash_xor"] ^= numeric


def _public_multiset_fingerprint(state: Mapping[str, int]) -> Dict[str, Any]:
    return {
        "algorithm": "sha256-composite-multiset-v1",
        "row_count": int(state.get("row_count", 0)),
        "hash_sum": f"{int(state.get('hash_sum', 0)):064x}",
        "hash_square_sum": (
            f"{int(state.get('hash_square_sum', 0)):064x}"
        ),
        "hash_xor": f"{int(state.get('hash_xor', 0)):064x}",
    }


def _cow_identity_digest(
    identifier: str,
    identity_values: Mapping[str, str],
    required_identity_keys: Iterable[str],
) -> bytes:
    payload = [
        _normalize_identifier(identifier),
        *[
            _normalize_identifier(identity_values.get(key))
            for key in sorted(required_identity_keys)
        ],
    ]
    return hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    ).digest()


def _cow_stable_field_positions(
    headers: Sequence[Any],
) -> Dict[str, Optional[int]]:
    """返回牛号绑定的稳定系谱字段位置。

    缺失列按全空列参与摘要。这样上游没有提供某一代系谱、下游也保持
    为空时不会误报；若下游凭空出现或串入非空值，仍会被识别。
    """

    return {
        field: _first_header(headers, aliases)
        for field, aliases in COW_STABLE_FIELD_ALIASES.items()
    }


def _cow_stable_fields_digest(
    row: Sequence[Any],
    positions: Mapping[str, Optional[int]],
) -> bytes:
    def canonical_value(field: str, value: Any) -> str:
        normalized = _normalize_identifier(value)
        if field == "dam":
            return normalized
        try:
            formatted, errors = format_naab_number(normalized)
        except Exception:
            return ""
        if errors or not formatted:
            return ""
        return _normalize_identifier(formatted)

    payload = [
        [
            field,
            canonical_value(
                field,
                row[index]
                if index is not None and index < len(row)
                else None,
            ),
        ]
        for field, index in sorted(positions.items())
    ]
    return hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    ).digest()


def _is_male_cow_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return bool(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return math.isfinite(float(value)) and float(value) == 1.0
        except (TypeError, ValueError):
            return False
    return str(value).strip().casefold() in {
        "1",
        "1.0",
        "公",
        "male",
    }


def _insert_cow_audit_batch(
    connection: sqlite3.Connection,
    slot: str,
    batch: Sequence[tuple[bytes, int, bytes, bytes]],
) -> None:
    if slot == "raw":
        connection.executemany(
            """
            INSERT INTO cow_raw_processed_audit (
                row_hash, raw_count, raw_male_count, raw_eligible_count
            )
            VALUES (?, 1, ?, ?)
            ON CONFLICT(row_hash) DO UPDATE SET
                raw_count = raw_count + 1,
                raw_male_count = raw_male_count + excluded.raw_male_count,
                raw_eligible_count = (
                    raw_eligible_count + excluded.raw_eligible_count
                )
            """,
            [
                (
                    identity_digest,
                    int(is_male),
                    int(not is_male),
                )
                for (
                    identity_digest,
                    is_male,
                    _identifier_digest,
                    _stable_digest,
                ) in batch
            ],
        )
        connection.executemany(
            """
            INSERT INTO cow_raw_processed_stable_audit (
                identifier_hash, raw_count, raw_eligible_count,
                raw_stable_hash
            )
            VALUES (?, 1, ?, ?)
            ON CONFLICT(identifier_hash) DO UPDATE SET
                raw_count = raw_count + 1,
                raw_eligible_count = (
                    raw_eligible_count + excluded.raw_eligible_count
                )
            """,
            [
                (
                    identifier_digest,
                    int(not is_male),
                    stable_digest,
                )
                for (
                    _identity_digest,
                    is_male,
                    identifier_digest,
                    stable_digest,
                ) in batch
            ],
        )
        return
    if slot == "processed":
        connection.executemany(
            """
            INSERT INTO cow_raw_processed_audit (
                row_hash, processed_count
            )
            VALUES (?, 1)
            ON CONFLICT(row_hash) DO UPDATE SET
                processed_count = processed_count + 1
            """,
            [
                (identity_digest,)
                for (
                    identity_digest,
                    _unused,
                    _identifier_digest,
                    _stable_digest,
                ) in batch
            ],
        )
        connection.executemany(
            """
            INSERT INTO cow_raw_processed_stable_audit (
                identifier_hash, processed_count, processed_stable_hash
            )
            VALUES (?, 1, ?)
            ON CONFLICT(identifier_hash) DO UPDATE SET
                processed_count = processed_count + 1,
                processed_stable_hash = excluded.processed_stable_hash
            """,
            [
                (identifier_digest, stable_digest)
                for (
                    _identity_digest,
                    _unused,
                    identifier_digest,
                    stable_digest,
                ) in batch
            ],
        )
        return
    raise ValueError("内部母牛拒绝账槽位无效")


def _scan_cow_raw_or_processed(
    path: Path,
    *,
    slot: str,
    id_candidates: Sequence[str],
    expected_identity: Mapping[str, str],
    connection: sqlite3.Connection,
) -> Dict[str, Any]:
    headers, rows, workbook = _first_sheet_table(path)
    try:
        id_index = _first_header(headers, id_candidates)
        if id_index is None:
            raise ValueError("缺少业务标识符列")
        expectations = _identity_expectations(headers, expected_identity)
        required_identity_keys = _required_identity_keys(expected_identity)
        stable_positions = _cow_stable_field_positions(headers)
        sex_index = (
            _first_header(headers, ("性别", "sex"))
            if slot == "raw"
            else None
        )
        stats: Dict[str, Any] = {
            "rows": 0,
            "identifier_count": 0,
            "blank_identifier_count": 0,
            "identity_columns": len(expectations),
            "identity_missing_columns": len(
                required_identity_keys.difference(expectations)
            ),
            "identity_mismatch_rows": 0,
            "identity_blank_rows": 0,
            "stable_field_columns": sum(
                index is not None for index in stable_positions.values()
            ),
            "stable_field_missing_columns": sum(
                index is None for index in stable_positions.values()
            ),
        }
        fingerprint = _multiset_fingerprint_state()
        batch: List[tuple[bytes, int, bytes, bytes]] = []
        for row in rows:
            if _is_blank_row(row):
                continue
            stats["rows"] += 1
            identifier = _normalize_identifier(
                row[id_index] if id_index < len(row) else None
            )
            if identifier:
                stats["identifier_count"] += 1
            else:
                stats["blank_identifier_count"] += 1

            identity_values = {}
            for key in sorted(required_identity_keys):
                expectation = expectations.get(key)
                actual = ""
                if expectation is not None:
                    column_index, expected_value = expectation
                    actual = _normalize_identifier(
                        row[column_index]
                        if column_index < len(row)
                        else None
                    )
                    if not actual:
                        stats["identity_blank_rows"] += 1
                    if actual != expected_value:
                        stats["identity_mismatch_rows"] += 1
                identity_values[key] = actual

            digest = _cow_identity_digest(
                identifier,
                identity_values,
                required_identity_keys,
            )
            _add_multiset_digest(fingerprint, digest)
            is_male = int(
                slot == "raw"
                and sex_index is not None
                and _is_male_cow_value(
                    row[sex_index] if sex_index < len(row) else None
                )
            )
            stable_digest = _cow_stable_fields_digest(
                row,
                stable_positions,
            )
            identifier_digest = hashlib.sha256(
                identifier.encode("utf-8")
            ).digest()
            batch.append(
                (
                    digest,
                    is_male,
                    identifier_digest,
                    stable_digest,
                )
            )
            if len(batch) >= 2_000:
                _insert_cow_audit_batch(connection, slot, batch)
                batch.clear()
        if batch:
            _insert_cow_audit_batch(connection, slot, batch)
        connection.commit()
        stats["identity_fingerprint"] = _public_multiset_fingerprint(
            fingerprint
        )
        return stats
    finally:
        workbook.close()


def _validate_raw_processed_cow_lineage(
    raw: Path,
    processed: Path,
    *,
    expected_identity: Mapping[str, str],
    connection: sqlite3.Connection,
) -> Dict[str, Any]:
    """以磁盘多重集严格核对原始牛群、允许拒绝项和标准化母牛。"""
    connection.execute("DROP TABLE IF EXISTS cow_raw_processed_audit")
    connection.execute(
        "DROP TABLE IF EXISTS cow_raw_processed_stable_audit"
    )
    connection.execute(
        """
        CREATE TABLE cow_raw_processed_audit (
            row_hash BLOB PRIMARY KEY,
            raw_count INTEGER NOT NULL DEFAULT 0,
            raw_male_count INTEGER NOT NULL DEFAULT 0,
            raw_eligible_count INTEGER NOT NULL DEFAULT 0,
            processed_count INTEGER NOT NULL DEFAULT 0
        ) WITHOUT ROWID
        """
    )
    connection.execute(
        """
        CREATE TABLE cow_raw_processed_stable_audit (
            identifier_hash BLOB PRIMARY KEY,
            raw_count INTEGER NOT NULL DEFAULT 0,
            raw_eligible_count INTEGER NOT NULL DEFAULT 0,
            processed_count INTEGER NOT NULL DEFAULT 0,
            raw_stable_hash BLOB,
            processed_stable_hash BLOB
        ) WITHOUT ROWID
        """
    )
    try:
        raw_profile = _scan_cow_raw_or_processed(
            raw,
            slot="raw",
            id_candidates=("cow_id", "母牛号", "牛号", "耳号"),
            expected_identity=expected_identity,
            connection=connection,
        )
        processed_profile = _scan_cow_raw_or_processed(
            processed,
            slot="processed",
            id_candidates=("cow_id", "母牛号", "牛号", "耳号"),
            expected_identity=expected_identity,
            connection=connection,
        )
        aggregate = connection.execute(
            """
            SELECT
              COALESCE(SUM(raw_count), 0),
              COALESCE(SUM(processed_count), 0),
              COALESCE(SUM(raw_male_count), 0),
              COALESCE(SUM(
                CASE WHEN raw_eligible_count > 1
                THEN raw_eligible_count - 1 ELSE 0 END
              ), 0),
              COALESCE(SUM(
                CASE WHEN processed_count > raw_eligible_count
                THEN processed_count - raw_eligible_count ELSE 0 END
              ), 0),
              COALESCE(SUM(
                ABS(
                  processed_count -
                  CASE WHEN raw_eligible_count > 0 THEN 1 ELSE 0 END
                )
              ), 0),
              COALESCE(SUM(
                CASE WHEN processed_count > 1
                THEN processed_count - 1 ELSE 0 END
              ), 0)
            FROM cow_raw_processed_audit
            """
        ).fetchone()
        (
            raw_rows,
            processed_rows,
            rejected_male_rows,
            rejected_duplicate_rows,
            processed_not_eligible_rows,
            retained_multiset_mismatch_rows,
            processed_duplicate_rows,
        ) = (int(value) for value in aggregate)
        stable_aggregate = connection.execute(
            """
            SELECT
              COALESCE(SUM(
                CASE WHEN raw_count > 1 THEN 1 ELSE 0 END
              ), 0),
              COALESCE(SUM(
                CASE WHEN raw_count > 1 THEN raw_count ELSE 0 END
              ), 0),
              COALESCE(SUM(
                CASE
                  WHEN raw_count = 1
                   AND raw_eligible_count = 1
                   AND processed_count = 1
                  THEN 1 ELSE 0
                END
              ), 0),
              COALESCE(SUM(
                CASE
                  WHEN raw_count = 1
                   AND raw_eligible_count = 1
                   AND processed_count = 1
                   AND raw_stable_hash != processed_stable_hash
                  THEN 1 ELSE 0
                END
              ), 0)
            FROM cow_raw_processed_stable_audit
            """
        ).fetchone()
        (
            ambiguous_raw_identifiers,
            ambiguous_raw_rows,
            stable_field_compared_identifiers,
            stable_field_mismatch_identifiers,
        ) = (int(value) for value in stable_aggregate)

        expected_processed = _multiset_fingerprint_state()
        rejected = _multiset_fingerprint_state()
        for row_hash, raw_count, raw_eligible_count in connection.execute(
            """
            SELECT row_hash, raw_count, raw_eligible_count
            FROM cow_raw_processed_audit
            """
        ):
            expected_count = 1 if int(raw_eligible_count) > 0 else 0
            rejected_count = int(raw_count) - expected_count
            _add_multiset_digest(
                expected_processed,
                bytes(row_hash),
                count=expected_count,
            )
            _add_multiset_digest(
                rejected,
                bytes(row_hash),
                count=rejected_count,
            )

        expected_processed_public = _public_multiset_fingerprint(
            expected_processed
        )
        rejected_public = _public_multiset_fingerprint(rejected)
        processed_public = processed_profile["identity_fingerprint"]
        rejected_rows = raw_rows - processed_rows
        audited_rejected_rows = (
            rejected_male_rows + rejected_duplicate_rows
        )
        rejection_balance_delta = rejected_rows - audited_rejected_rows
        passed = (
            raw_rows == int(raw_profile["rows"])
            and processed_rows == int(processed_profile["rows"])
            and raw_rows == processed_rows + audited_rejected_rows
            and rejection_balance_delta == 0
            and processed_not_eligible_rows == 0
            and retained_multiset_mismatch_rows == 0
            and processed_duplicate_rows == 0
            and stable_field_mismatch_identifiers == 0
            and processed_public == expected_processed_public
            and raw_profile["blank_identifier_count"] == 0
            and processed_profile["blank_identifier_count"] == 0
            and raw_profile["identity_missing_columns"] == 0
            and processed_profile["identity_missing_columns"] == 0
            and raw_profile["identity_mismatch_rows"] == 0
            and processed_profile["identity_mismatch_rows"] == 0
        )
        return {
            "raw": raw_profile,
            "processed": processed_profile,
            "rejected_rows": rejected_rows,
            "rejected_male_rows": rejected_male_rows,
            "rejected_duplicate_rows": rejected_duplicate_rows,
            "audited_rejected_rows": audited_rejected_rows,
            "rejection_balance_delta": rejection_balance_delta,
            "processed_not_eligible_rows": processed_not_eligible_rows,
            "retained_multiset_mismatch_rows": (
                retained_multiset_mismatch_rows
            ),
            "processed_duplicate_rows": processed_duplicate_rows,
            "stable_field_compared_identifiers": (
                stable_field_compared_identifiers
            ),
            "stable_field_mismatch_identifiers": (
                stable_field_mismatch_identifiers
            ),
            "stable_field_ambiguous_raw_identifiers": (
                ambiguous_raw_identifiers
            ),
            "stable_field_ambiguous_raw_rows": ambiguous_raw_rows,
            "expected_processed_identity_fingerprint": (
                expected_processed_public
            ),
            "rejected_identity_fingerprint": rejected_public,
            "passed": passed,
        }
    finally:
        connection.execute("DROP TABLE IF EXISTS cow_raw_processed_audit")
        connection.execute(
            "DROP TABLE IF EXISTS cow_raw_processed_stable_audit"
        )


def _identifier_reconciliation(
    connection: sqlite3.Connection,
    table: str,
) -> Dict[str, int]:
    row = connection.execute(
        f"""
        SELECT
          COALESCE(SUM(CASE WHEN source_count > 1
                    THEN source_count - 1 ELSE 0 END), 0),
          COALESCE(SUM(CASE WHEN middle_count > 1
                    THEN middle_count - 1 ELSE 0 END), 0),
          COALESCE(SUM(CASE WHEN target_count > 1
                    THEN target_count - 1 ELSE 0 END), 0),
          COALESCE(SUM(CASE WHEN middle_count > source_count
                    THEN middle_count - source_count ELSE 0 END), 0),
          COALESCE(SUM(CASE WHEN target_count > middle_count
                    THEN target_count - middle_count ELSE 0 END), 0),
          COALESCE(SUM(CASE WHEN middle_count > target_count
                    THEN middle_count - target_count ELSE 0 END), 0),
          COUNT(*)
        FROM {table}
        """
    ).fetchone()
    return {
        "source_duplicate_rows": int(row[0]),
        "middle_duplicate_rows": int(row[1]),
        "target_duplicate_rows": int(row[2]),
        "middle_not_in_source_rows": int(row[3]),
        "target_not_in_middle_rows": int(row[4]),
        "middle_not_in_target_rows": int(row[5]),
        "distinct_identifier_hashes": int(row[6]),
    }


def _canonical_cell(value: Any) -> Any:
    """返回用于跨 Excel 阶段逐行对账的稳定表示。

    pandas/openpyxl 在一次合法的 Excel 往返后，可能把 ``123``、
    ``123.0`` 和文本 ``"123.0"`` 表示成不同 Python 类型；空白值也
    可能在 ``None``、空字符串和 ``NaN`` 之间变化。这些表示差异不应
    被误判为牛号与内容错配。

    带前导零的整数字符串仍按文本处理，因为它通常是牛号或牧场编号，
    ``"00123"`` 变成数值 ``123`` 属于真实的标识符损坏。
    """
    if value is None:
        return ["blank", ""]
    if isinstance(value, bool):
        return ["bool", bool(value)]
    if isinstance(value, datetime):
        return ["datetime", value.isoformat(timespec="microseconds")]
    if isinstance(value, date):
        return ["date", value.isoformat()]
    if isinstance(value, time):
        return ["time", value.isoformat(timespec="microseconds")]
    if isinstance(value, int):
        return ["number", str(value)]
    if isinstance(value, float):
        if math.isnan(value):
            return ["blank", ""]
        if math.isinf(value):
            return ["number", "inf" if value > 0 else "-inf"]
        return ["number", format(Decimal(str(value)).normalize(), "f")]
    if isinstance(value, bytes):
        return ["bytes_sha256", hashlib.sha256(value).hexdigest()]
    text = str(value).strip()
    if text.casefold() in IDENTIFIER_MISSING:
        return ["blank", ""]
    if re.fullmatch(
        r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?",
        text,
    ):
        unsigned = text.lstrip("+-")
        integer_part = unsigned.split(".", 1)[0]
        # ``0.5`` 是普通小数；``00.5``/``00123`` 更可能是标识符。
        has_significant_leading_zero = (
            len(integer_part) > 1 and integer_part.startswith("0")
        )
        if not has_significant_leading_zero:
            try:
                decimal_value = Decimal(text)
                if decimal_value.is_finite():
                    return [
                        "number",
                        format(decimal_value.normalize(), "f"),
                    ]
            except (InvalidOperation, ValueError):
                pass
    return ["text", text]


def _header_positions(headers: Sequence[Any]) -> Dict[str, int]:
    result = {}
    duplicates = set()
    for index, value in enumerate(headers):
        header = str(value or "").strip()
        if not header:
            continue
        if header in result:
            duplicates.add(header)
        else:
            result[header] = index
    if duplicates:
        raise ValueError("工作簿存在重复表头")
    return result


def _projected_row_hash(
    row: Sequence[Any],
    positions: Sequence[int],
) -> bytes:
    values = [
        _canonical_cell(row[index] if index < len(row) else None)
        for index in positions
    ]
    return hashlib.sha256(
        json.dumps(
            values,
            ensure_ascii=False,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    ).digest()


def _insert_content_batch(
    connection: sqlite3.Connection,
    table: str,
    slot: str,
    batch: Sequence[tuple[bytes]],
) -> None:
    if slot not in {"source", "target"}:
        raise ValueError("内部整行血缘槽位无效")
    connection.executemany(
        f"""
        INSERT INTO {table} (row_hash, {slot}_count)
        VALUES (?, 1)
        ON CONFLICT(row_hash) DO UPDATE
        SET {slot}_count = {slot}_count + 1
        """,
        batch,
    )


def _scan_projected_rows(
    path: Path,
    *,
    columns: Sequence[str],
    connection: sqlite3.Connection,
    table: str,
    slot: str,
) -> int:
    headers, rows, workbook = _first_sheet_table(path)
    try:
        positions_by_name = _header_positions(headers)
        positions = [positions_by_name[name] for name in columns]
        batch: List[tuple[bytes]] = []
        count = 0
        for row in rows:
            if _is_blank_row(row):
                continue
            count += 1
            batch.append((_projected_row_hash(row, positions),))
            if len(batch) >= 2_000:
                _insert_content_batch(connection, table, slot, batch)
                batch.clear()
        if batch:
            _insert_content_batch(connection, table, slot, batch)
        connection.commit()
        return count
    finally:
        workbook.close()


def _compare_projected_rows(
    source: Path,
    target: Path,
    *,
    connection: sqlite3.Connection,
    table: str,
    target_may_be_subset: bool,
    excluded_columns: Iterable[str] = (),
) -> Dict[str, int]:
    if not re.fullmatch(r"[a-z][a-z0-9_]*", table):
        raise ValueError("内部整行血缘表名无效")
    source_positions = _header_positions(_first_sheet_headers(source))
    target_positions = _header_positions(_first_sheet_headers(target))
    excluded = {str(name).strip() for name in excluded_columns}
    columns = [
        name
        for name in source_positions
        if name in target_positions and name not in excluded
    ]
    if not columns:
        raise ValueError("上下游工作簿没有可对账的共有列")
    connection.execute(
        f"""
        CREATE TABLE {table} (
            row_hash BLOB PRIMARY KEY,
            source_count INTEGER NOT NULL DEFAULT 0,
            target_count INTEGER NOT NULL DEFAULT 0
        ) WITHOUT ROWID
        """
    )
    try:
        source_rows = _scan_projected_rows(
            source,
            columns=columns,
            connection=connection,
            table=table,
            slot="source",
        )
        target_rows = _scan_projected_rows(
            target,
            columns=columns,
            connection=connection,
            table=table,
            slot="target",
        )
        row = connection.execute(
            f"""
            SELECT
              COALESCE(SUM(CASE WHEN target_count > source_count
                        THEN target_count - source_count ELSE 0 END), 0),
              COALESCE(SUM(CASE WHEN source_count > target_count
                        THEN source_count - target_count ELSE 0 END), 0)
            FROM {table}
            """
        ).fetchone()
        target_not_source = int(row[0])
        source_not_target = int(row[1])
        return {
            "source_rows": source_rows,
            "target_rows": target_rows,
            "shared_columns": len(columns),
            "target_rows_not_in_source": target_not_source,
            "source_rows_not_in_target": source_not_target,
            "passed": int(
                target_not_source == 0
                and (target_may_be_subset or source_not_target == 0)
            ),
        }
    finally:
        connection.execute(f"DROP TABLE IF EXISTS {table}")


def _record_lineage_issue(
    result: ResultBuilder,
    task: Mapping[str, Any],
    check: str,
    passed: bool,
    message: str,
    *,
    warning: bool = False,
) -> None:
    if passed:
        return
    result.add(
        "warning" if warning else "error",
        "lineage",
        check,
        message,
        farm_code=str(task.get("farm_code") or ""),
    )


def _validate_cow_lineage(
    project: Path,
    task: Dict[str, Any],
    connection: sqlite3.Connection,
    result: ResultBuilder,
) -> None:
    child = task["child_path"]
    raw = child / "raw_data" / "cow_data.xlsx"
    processed = child / "standardized_data" / "processed_cow_data.xlsx"
    final = child / "analysis_results" / "processed_cow_data_key_traits_final.xlsx"
    index = child / "analysis_results" / "processed_index_cow_index_scores.xlsx"
    raw_processed = _validate_raw_processed_cow_lineage(
        raw,
        processed,
        expected_identity=task["identity"],
        connection=connection,
    )
    table = "cow_lineage"
    _ensure_id_table(connection, table)
    profiles = {
        "processed": _profile_identifier_table(
            processed,
            id_candidates=("cow_id", "母牛号", "牛号", "耳号"),
            expected_identity=task["identity"],
            connection=connection,
            table=table,
            slot="source",
        ),
        "final": _profile_identifier_table(
            final,
            id_candidates=("cow_id", "母牛号", "牛号", "耳号"),
            expected_identity=task["identity"],
            connection=connection,
            table=table,
            slot="middle",
        ),
        "index": _profile_identifier_table(
            index,
            id_candidates=("cow_id", "母牛号", "牛号", "耳号"),
            expected_identity=task["identity"],
            connection=connection,
            table=table,
            slot="target",
        ),
    }
    reconciliation = _identifier_reconciliation(connection, table)
    connection.execute(f"DROP TABLE {table}")
    processed_final_content = _compare_projected_rows(
        processed,
        final,
        connection=connection,
        table="cow_processed_final_content",
        target_may_be_subset=True,
        # 牧场身份三列已有独立、严格的逐行校验。这里排除它们，
        # 让“牛号绑定内容是否串行”与“编号格式是否损坏”分别报告。
        excluded_columns=("API farmcode", "牧场编号", "牧场名称"),
    )
    final_index_content = _compare_projected_rows(
        final,
        index,
        connection=connection,
        table="cow_final_index_content",
        target_may_be_subset=False,
        # 指数计算后，“牛只分组”会按繁殖状态和选配策略合法回写
        # group；它不是从 final 原样继承的稳定字段。
        excluded_columns=(
            "API farmcode",
            "牧场编号",
            "牧场名称",
            "group",
        ),
    )

    row = {
        "farm_code": task["farm_code"],
        "farm_name": task["farm_name"],
        "lineage": "cow",
        "raw_rows": raw_processed["raw"]["rows"],
        "processed_rows": profiles["processed"]["rows"],
        "final_rows": profiles["final"]["rows"],
        "index_rows": profiles["index"]["rows"],
        "raw_to_processed_excluded": raw_processed["rejected_rows"],
        "raw_to_processed_rejected_male_rows": (
            raw_processed["rejected_male_rows"]
        ),
        "raw_to_processed_rejected_duplicate_rows": (
            raw_processed["rejected_duplicate_rows"]
        ),
        "raw_to_processed_audited_rejected_rows": (
            raw_processed["audited_rejected_rows"]
        ),
        "raw_to_processed_rejection_balance_delta": (
            raw_processed["rejection_balance_delta"]
        ),
        "raw_to_processed_unexplained_rows": (
            raw_processed["retained_multiset_mismatch_rows"]
        ),
        "raw_to_processed_not_eligible_rows": (
            raw_processed["processed_not_eligible_rows"]
        ),
        "raw_to_processed_stable_field_compared_identifiers": (
            raw_processed["stable_field_compared_identifiers"]
        ),
        "raw_to_processed_stable_field_mismatch_identifiers": (
            raw_processed["stable_field_mismatch_identifiers"]
        ),
        "raw_to_processed_stable_field_ambiguous_raw_identifiers": (
            raw_processed["stable_field_ambiguous_raw_identifiers"]
        ),
        "raw_to_processed_stable_field_ambiguous_raw_rows": (
            raw_processed["stable_field_ambiguous_raw_rows"]
        ),
        "raw_stable_field_columns": (
            raw_processed["raw"]["stable_field_columns"]
        ),
        "processed_stable_field_columns": (
            raw_processed["processed"]["stable_field_columns"]
        ),
        "raw_identity_fingerprint": (
            raw_processed["raw"]["identity_fingerprint"]
        ),
        "processed_identity_fingerprint": (
            raw_processed["processed"]["identity_fingerprint"]
        ),
        "expected_processed_identity_fingerprint": (
            raw_processed["expected_processed_identity_fingerprint"]
        ),
        "rejected_identity_fingerprint": (
            raw_processed["rejected_identity_fingerprint"]
        ),
        "processed_to_final_excluded": (
            profiles["processed"]["rows"] - profiles["final"]["rows"]
        ),
        **reconciliation,
        "identity_mismatch_rows": (
            raw_processed["raw"]["identity_mismatch_rows"]
            + raw_processed["processed"]["identity_mismatch_rows"]
            + profiles["final"]["identity_mismatch_rows"]
            + profiles["index"]["identity_mismatch_rows"]
        ),
        "blank_identifier_rows": (
            raw_processed["raw"]["blank_identifier_count"]
            + raw_processed["processed"]["blank_identifier_count"]
            + profiles["final"]["blank_identifier_count"]
            + profiles["index"]["blank_identifier_count"]
        ),
        "processed_final_row_content_mismatch": int(
            processed_final_content["target_rows_not_in_source"]
        ),
        "final_index_row_content_mismatch": int(
            final_index_content["target_rows_not_in_source"]
            + final_index_content["source_rows_not_in_target"]
        ),
        "processed_final_shared_columns": int(
            processed_final_content["shared_columns"]
        ),
        "final_index_shared_columns": int(
            final_index_content["shared_columns"]
        ),
    }
    task["_acceptance_index_rows"] = int(profiles["index"]["rows"])
    result.lineage_rows.append(row)
    passed = (
        bool(raw_processed["passed"])
        and raw_processed["processed"]["rows"]
        == profiles["processed"]["rows"]
        and profiles["processed"]["rows"] >= profiles["final"]["rows"]
        and profiles["final"]["rows"] == profiles["index"]["rows"]
        and reconciliation["middle_not_in_source_rows"] == 0
        and reconciliation["target_not_in_middle_rows"] == 0
        and reconciliation["middle_not_in_target_rows"] == 0
        and reconciliation["source_duplicate_rows"] == 0
        and reconciliation["middle_duplicate_rows"] == 0
        and reconciliation["target_duplicate_rows"] == 0
        and row["identity_mismatch_rows"] == 0
        and row["blank_identifier_rows"] == 0
        and raw_processed["raw"]["identity_missing_columns"] == 0
        and raw_processed["processed"]["identity_missing_columns"] == 0
        and profiles["final"]["identity_missing_columns"] == 0
        and profiles["index"]["identity_missing_columns"] == 0
        and bool(processed_final_content["passed"])
        and bool(final_index_content["passed"])
    )
    _record_lineage_issue(
        result,
        task,
        "cow_lineage_mismatch",
        passed,
        "母牛 raw→拒绝账→processed→final→index 行数、牧场身份、"
        "牛号多重集或牛号绑定系谱字段不一致",
    )
    ambiguous_identifiers = int(
        raw_processed["stable_field_ambiguous_raw_identifiers"]
    )
    _record_lineage_issue(
        result,
        task,
        "cow_duplicate_identity_ambiguous",
        ambiguous_identifiers == 0,
        (
            f"原始牛群存在 {ambiguous_identifiers} 个重复牛号组；"
            "标准化已按去重规则保留一条，稳定系谱绑定未纳入严格逐行比较"
        ),
        warning=True,
    )


def _validate_three_file_lineage(
    task: Dict[str, Any],
    connection: sqlite3.Connection,
    result: ResultBuilder,
    *,
    lineage_name: str,
    paths: Sequence[Path],
    id_candidates: Sequence[str],
    require_equal: bool,
    enforce_identity: bool,
    content_pairs: Sequence[tuple[int, int, bool]],
) -> Optional[Dict[str, Any]]:
    if any(not path.is_file() for path in paths):
        return None
    table = f"{lineage_name}_lineage"
    _ensure_id_table(connection, table)
    profiles = [
        _profile_identifier_table(
            path,
            id_candidates=id_candidates,
            expected_identity=task["identity"],
            connection=connection,
            table=table,
            slot=slot,
        )
        for path, slot in zip(paths, ("source", "middle", "target"))
    ]
    reconciliation = _identifier_reconciliation(connection, table)
    connection.execute(f"DROP TABLE {table}")
    content_results = []
    for pair_index, (source_index, target_index, target_may_be_subset) in enumerate(
        content_pairs,
        start=1,
    ):
        content_results.append(
            _compare_projected_rows(
                paths[source_index],
                paths[target_index],
                connection=connection,
                table=f"{lineage_name}_content_{pair_index}",
                target_may_be_subset=target_may_be_subset,
            )
        )
    row = {
        "farm_code": task["farm_code"],
        "farm_name": task["farm_name"],
        "lineage": lineage_name,
        "raw_rows": profiles[0]["rows"],
        "processed_rows": profiles[1]["rows"],
        "final_rows": profiles[2]["rows"],
        "index_rows": "",
        "raw_to_processed_excluded": (
            profiles[0]["rows"] - profiles[1]["rows"]
        ),
        "processed_to_final_excluded": (
            profiles[1]["rows"] - profiles[2]["rows"]
        ),
        **reconciliation,
        "identity_mismatch_rows": sum(
            profile["identity_mismatch_rows"] for profile in profiles
        ),
        "blank_identifier_rows": sum(
            profile["blank_identifier_count"] for profile in profiles
        ),
        "row_content_mismatch": sum(
            item["target_rows_not_in_source"]
            + (
                0
                if pair[2]
                else item["source_rows_not_in_target"]
            )
            for item, pair in zip(content_results, content_pairs)
        ),
        "row_content_shared_columns": sum(
            item["shared_columns"] for item in content_results
        ),
    }
    result.lineage_rows.append(row)
    passed = (
        (not require_equal or len({profile["rows"] for profile in profiles}) == 1)
        and reconciliation["middle_not_in_source_rows"] == 0
        and reconciliation["target_not_in_middle_rows"] == 0
        and reconciliation["middle_not_in_target_rows"] == 0
        and row["identity_mismatch_rows"] == 0
        and row["blank_identifier_rows"] == 0
        and (
            not enforce_identity
            or all(
                profile["identity_missing_columns"] == 0
                for profile in profiles
            )
        )
        and all(bool(item["passed"]) for item in content_results)
    )
    _record_lineage_issue(
        result,
        task,
        f"{lineage_name}_lineage_mismatch",
        passed,
        f"{lineage_name} 三段行数、身份或标识符血缘不一致",
    )
    return row


def _validate_breeding_lineage(
    task: Dict[str, Any],
    connection: sqlite3.Connection,
    result: ResultBuilder,
) -> None:
    child = task["child_path"]
    raw = child / "raw_data" / "breeding_records.xlsx"
    processed = child / "standardized_data" / "processed_breeding_data.xlsx"
    mated = child / "analysis_results" / "processed_mated_bull_traits.xlsx"
    if not raw.is_file() and not processed.is_file():
        return
    if not (raw.is_file() and processed.is_file() and mated.is_file()):
        _record_lineage_issue(
            result,
            task,
            "breeding_output_missing",
            False,
            "存在配种数据，但 raw、processed 或已配公牛结果不完整",
        )
        return
    _validate_three_file_lineage(
        task,
        connection,
        result,
        lineage_name="breeding",
        paths=(raw, processed, mated),
        id_candidates=("耳号", "cow_id", "母牛号", "牛号"),
        require_equal=True,
        enforce_identity=True,
        content_pairs=((1, 2, False),),
    )


def _validate_bull_lineage(
    task: Dict[str, Any],
    connection: sqlite3.Connection,
    result: ResultBuilder,
) -> None:
    child = task["child_path"]
    source = child / "standardized_data" / "processed_bull_data.xlsx"
    traits = child / "analysis_results" / "processed_bull_data_key_traits.xlsx"
    index = child / "analysis_results" / "processed_index_bull_scores.xlsx"
    if not source.is_file():
        return
    if not traits.is_file() or not index.is_file():
        _record_lineage_issue(
            result,
            task,
            "candidate_bull_output_missing",
            False,
            "存在备选公牛输入，但性状或指数结果不完整",
        )
        return
    _validate_three_file_lineage(
        task,
        connection,
        result,
        lineage_name="bull",
        paths=(source, traits, index),
        id_candidates=("bull_id", "公牛号", "NAAB", "BULL NAAB"),
        require_equal=True,
        enforce_identity=False,
        content_pairs=((0, 1, False), (0, 2, False)),
    )


def _validate_business_scope_lineage(
    task: Dict[str, Any],
    result: ResultBuilder,
) -> None:
    """核对配种业务键及补跑分析范围；不公开任何牛号或公牛号。"""
    issue_definitions = {
        "breeding_business_key": (
            "breeding_business_key_mismatch",
            "配种 raw→processed 业务键多重集不一致",
        ),
        "cow_self_scope": (
            "cow_self_scope_mismatch",
            "母牛自身近交结果范围与奶牛母牛范围不一致",
        ),
        "candidate_cartesian_scope": (
            "candidate_cartesian_scope_mismatch",
            "备选公牛近交结果不是完整母牛×公牛笛卡尔积",
        ),
        "mated_business_key_scope": (
            "mated_business_key_scope_mismatch",
            "已配公牛近交结果与有效配种业务键范围不一致",
        ),
        "matching_scope": (
            "matching_scope_mismatch",
            "个体选配矩阵或报告与在场母牛范围不一致",
        ),
    }
    for check in validate_child_scope_artifacts(task["child_path"]):
        lineage = str(check.get("lineage") or "scope")
        row = {
            "farm_code": task["farm_code"],
            "farm_name": task["farm_name"],
            **check,
        }
        result.lineage_rows.append(row)
        code, message = issue_definitions.get(
            lineage,
            ("business_scope_mismatch", "补跑分析业务范围不一致"),
        )
        _record_lineage_issue(
            result,
            task,
            code,
            bool(check.get("passed")),
            message,
        )


def _validate_lineage(
    project: Path,
    tasks: Sequence[Dict[str, Any]],
    result: ResultBuilder,
) -> None:
    with tempfile.TemporaryDirectory(prefix="multi_farm_acceptance_") as temp:
        database = Path(temp) / "hashed_lineage.sqlite3"
        connection = sqlite3.connect(str(database))
        try:
            connection.execute("PRAGMA journal_mode = OFF")
            connection.execute("PRAGMA synchronous = OFF")
            connection.execute("PRAGMA temp_store = FILE")
            for task in tasks:
                for label, validator in (
                    ("cow", lambda: _validate_cow_lineage(
                        project, task, connection, result
                    )),
                    ("breeding", lambda: _validate_breeding_lineage(
                        task, connection, result
                    )),
                    ("bull", lambda: _validate_bull_lineage(
                        task, connection, result
                    )),
                    ("scope", lambda: _validate_business_scope_lineage(
                        task, result
                    )),
                ):
                    try:
                        validator()
                    except Exception as exc:
                        connection.execute(
                            f"DROP TABLE IF EXISTS {label}_lineage"
                        )
                        result.add(
                            "error",
                            "lineage",
                            f"{label}_lineage_scan_failed",
                            f"{label} 血缘扫描失败（{type(exc).__name__}）",
                            farm_code=task["farm_code"],
                        )
        finally:
            connection.close()


def _validate_formula_integrity(
    tasks: Sequence[Dict[str, Any]],
    result: ResultBuilder,
) -> None:
    """逐牛复算性状和指数；只输出聚合计数与不可逆指纹。"""
    for task in tasks:
        try:
            checks = validate_cow_formulas(task["child_path"])
        except FormulaValidationError as exc:
            result.add(
                "error",
                "formula",
                "cow_formula_scan_failed",
                f"逐牛公式复算无法完成（{type(exc).__name__}）",
                farm_code=task["farm_code"],
            )
            continue
        except Exception as exc:
            result.add(
                "error",
                "formula",
                "cow_formula_scan_failed",
                f"逐牛公式复算异常（{type(exc).__name__}）",
                farm_code=task["farm_code"],
            )
            continue

        trait = checks["trait"]
        index = checks["index"]
        result.lineage_rows.append(
            {
                "farm_code": task["farm_code"],
                "farm_name": task["farm_name"],
                "lineage": "cow_formula",
                "trait_formula_checked_rows": trait["checked_rows"],
                "trait_formula_checked_cells": trait["checked_cells"],
                "trait_formula_checked_traits": trait["checked_traits"],
                "trait_formula_skipped_genomic_cells": int(
                    trait.get("skipped_genomic_cells", 0) or 0
                ),
                "trait_formula_mismatch_rows": trait["mismatch_rows"],
                "trait_formula_mismatch_cells": trait["mismatch_cells"],
                "trait_formula_mismatch_fingerprint": (
                    trait["mismatch_fingerprint"]["digest"]
                ),
                "trait_formula_configuration_fingerprint": (
                    trait["configuration_fingerprint"]
                ),
                "index_formula_checked_rows": index["checked_rows"],
                "index_formula_checked_cells": index["checked_cells"],
                "index_formula_checked_indexes": index["checked_indexes"],
                "index_formula_mismatch_rows": index["mismatch_rows"],
                "index_formula_mismatch_cells": index["mismatch_cells"],
                "index_formula_mismatch_fingerprint": (
                    index["mismatch_fingerprint"]["digest"]
                ),
                "index_formula_configuration_fingerprint": (
                    index["configuration_fingerprint"]
                ),
            }
        )
        if not trait["passed"]:
            result.add(
                "error",
                "formula",
                "cow_trait_formula_mismatch",
                (
                    "逐牛性状公式复算不一致："
                    f"{int(trait['mismatch_rows'])} 行、"
                    f"{int(trait['mismatch_cells'])} 个单元格；"
                    "指纹 "
                    f"{trait['mismatch_fingerprint']['digest']}"
                ),
                farm_code=task["farm_code"],
            )
        if not index["passed"]:
            result.add(
                "error",
                "formula",
                "cow_index_formula_mismatch",
                (
                    "逐牛指数公式复算不一致："
                    f"{int(index['mismatch_rows'])} 行、"
                    f"{int(index['mismatch_cells'])} 个单元格；"
                    "指纹 "
                    f"{index['mismatch_fingerprint']['digest']}"
                ),
                farm_code=task["farm_code"],
            )


def _required_extra_files(
    task: Mapping[str, Any],
) -> tuple[List[Path], List[str]]:
    child = Path(task["child_path"])
    analysis = child / "analysis_results"
    files = [analysis / name for name in CORE_ANALYSIS_FILES]
    files.extend(
        path
        for name in OPTIONAL_ANALYSIS_FILES
        if (path := analysis / name).is_file() and not path.is_symlink()
    )
    missing_patterns = []
    bull_source = child / "standardized_data" / "processed_bull_data.xlsx"
    breeding_source = (
        child / "standardized_data" / "processed_breeding_data.xlsx"
    )
    if bull_source.is_file():
        files.extend(analysis / name for name in CANDIDATE_FILES)
        files.extend(analysis / name for name in MATCHING_FILES)
        candidate = sorted(
            (
                path
                for path in analysis.glob(
                    "备选公牛_近交系数及隐性基因分析结果*.xlsx"
                )
                if path.is_file() and not path.is_symlink()
            ),
            key=lambda path: path.stat().st_mtime_ns,
        )
        if candidate:
            files.append(candidate[-1])
        else:
            missing_patterns.append("备选公牛近交及隐性基因结果")
    if breeding_source.is_file():
        files.extend(analysis / name for name in MATED_FILES)
        mated = sorted(
            (
                path
                for path in analysis.glob(
                    "已配公牛_近交系数及隐性基因分析结果*.xlsx"
                )
                if path.is_file() and not path.is_symlink()
            ),
            key=lambda path: path.stat().st_mtime_ns,
        )
        if mated:
            files.append(mated[-1])
        else:
            missing_patterns.append("已配公牛近交及隐性基因结果")
    return files, missing_patterns


def _collect_required_files(
    project: Path,
    tasks: Sequence[Dict[str, Any]],
    manifest_files: set[Path],
    result: ResultBuilder,
) -> set[Path]:
    paths = set(manifest_files)
    for task in tasks:
        files, missing_patterns = _required_extra_files(task)
        for label in missing_patterns:
            result.add(
                "error",
                "required_output",
                "required_pattern_missing",
                f"缺少{label}",
                farm_code=task["farm_code"],
            )
        for path in files:
            if not path.is_file() or path.is_symlink():
                result.add(
                    "error",
                    "required_output",
                    "required_file_missing",
                    "缺少全流程验收必需结果文件",
                    farm_code=task["farm_code"],
                    relative_path=_relative_to_project(project, path),
                )
            elif path.stat().st_size <= 0:
                result.add(
                    "error",
                    "required_output",
                    "required_file_empty",
                    "全流程验收必需结果文件大小为 0",
                    farm_code=task["farm_code"],
                    relative_path=_relative_to_project(project, path),
                )
            else:
                paths.add(path)
    return paths


def _local_name(tag: str) -> str:
    return str(tag).rsplit("}", 1)[-1]


def _percent_style_indexes(archive: zipfile.ZipFile) -> set[int]:
    if "xl/styles.xml" not in archive.namelist():
        return set()
    with archive.open("xl/styles.xml", "r") as stream:
        root = ElementTree.parse(stream).getroot()
    custom = {}
    cell_xfs = []
    for element in root.iter():
        name = _local_name(element.tag)
        if name == "numFmt":
            try:
                custom[int(element.attrib["numFmtId"])] = str(
                    element.attrib.get("formatCode") or ""
                )
            except (KeyError, ValueError):
                continue
        elif name == "cellXfs":
            cell_xfs = [
                child
                for child in element
                if _local_name(child.tag) == "xf"
            ]
    result = set()
    for index, xf in enumerate(cell_xfs):
        try:
            number_format = int(xf.attrib.get("numFmtId", 0))
        except ValueError:
            continue
        code = custom.get(number_format, "")
        code_without_literals = re.sub(r'"[^"]*"', "", code)
        if number_format in BUILTIN_PERCENT_FORMAT_IDS or "%" in code_without_literals:
            result.add(index)
    return result


def _text_signal(value: str) -> Optional[tuple[str, Optional[float]]]:
    text = str(value or "").strip()
    if text.upper() in ERROR_TOKENS:
        return "error", None
    if TEXT_PERCENT_PATTERN.fullmatch(text):
        try:
            return "percent", float(text.rstrip("%").strip())
        except ValueError:
            return None
    return None


def _shared_string_signals(
    archive: zipfile.ZipFile,
) -> Dict[int, tuple[str, Optional[float]]]:
    member = "xl/sharedStrings.xml"
    if member not in archive.namelist():
        return {}
    signals = {}
    index = -1
    with archive.open(member, "r") as stream:
        for _event, element in ElementTree.iterparse(
            stream,
            events=("end",),
        ):
            if _local_name(element.tag) != "si":
                continue
            index += 1
            text = "".join(
                child.text or ""
                for child in element.iter()
                if _local_name(child.tag) == "t"
            )
            signal = _text_signal(text)
            if signal is not None:
                signals[index] = signal
            element.clear()
    return signals


def _add_text_signal(
    counters: Dict[str, Any],
    signal: Optional[tuple[str, Optional[float]]],
) -> None:
    if signal is None:
        return
    kind, numeric = signal
    if kind == "error":
        counters["literal_error_marker_cells"] += 1
        return
    if kind != "percent" or numeric is None:
        return
    counters["text_percent_cells"] += 1
    counters["text_percent_min"] = (
        numeric
        if counters["text_percent_min"] is None
        else min(counters["text_percent_min"], numeric)
    )
    counters["text_percent_max"] = (
        numeric
        if counters["text_percent_max"] is None
        else max(counters["text_percent_max"], numeric)
    )
    if abs(numeric) > 100.0000001:
        counters["text_percent_abs_gt_100_cells"] += 1


def _scan_xlsx_health(path: Path) -> Dict[str, Any]:
    structure = inspect_xlsx_structure(path)
    if not structure.get("valid"):
        return {
            "valid": False,
            "error": str(structure.get("error") or "XLSX 结构无效"),
            "sheet_count": int(structure.get("sheet_count", 0) or 0),
            "empty_visible_sheets": 0,
            "header_only_visible_sheets": 0,
            "formula_cells": 0,
            "formula_error_cells": 0,
            "error_cells": 0,
            "nonfinite_numeric_cells": 0,
            "percent_numeric_cells": 0,
            "percent_abs_gt_one_cells": 0,
            "numeric_cells": 0,
            "fractional_numeric_cells": 0,
            "tiny_nonzero_numeric_cells": 0,
            "numeric_min": None,
            "numeric_max": None,
            "text_percent_cells": 0,
            "text_percent_abs_gt_100_cells": 0,
            "text_percent_min": None,
            "text_percent_max": None,
            "literal_error_marker_cells": 0,
        }
    counters = {
        "empty_visible_sheets": 0,
        "header_only_visible_sheets": 0,
        "formula_cells": 0,
        "formula_error_cells": 0,
        "error_cells": 0,
        "nonfinite_numeric_cells": 0,
        "percent_numeric_cells": 0,
        "percent_abs_gt_one_cells": 0,
        "numeric_cells": 0,
        "fractional_numeric_cells": 0,
        "tiny_nonzero_numeric_cells": 0,
        "numeric_min": None,
        "numeric_max": None,
        "text_percent_cells": 0,
        "text_percent_abs_gt_100_cells": 0,
        "text_percent_min": None,
        "text_percent_max": None,
        "literal_error_marker_cells": 0,
    }
    try:
        with zipfile.ZipFile(path, "r") as archive:
            percent_styles = _percent_style_indexes(archive)
            shared_signals = _shared_string_signals(archive)
            for sheet in structure["sheets"]:
                nonempty_cells = 0
                max_nonempty_row = 0
                with archive.open(sheet["xml_path"], "r") as stream:
                    for _event, cell in ElementTree.iterparse(
                        stream,
                        events=("end",),
                    ):
                        local = _local_name(cell.tag)
                        if local != "c":
                            if local == "row":
                                cell.clear()
                            continue
                        formula = ""
                        value = None
                        inline_text = ""
                        for child in cell.iter():
                            name = _local_name(child.tag)
                            if name == "f" and child.text:
                                formula += child.text
                            elif name == "v":
                                value = child.text
                            elif name == "t" and child.text:
                                inline_text += child.text
                        has_value = bool(formula or value is not None or inline_text)
                        if has_value:
                            nonempty_cells += 1
                            reference = str(cell.attrib.get("r") or "")
                            match = re.search(r"([1-9][0-9]*)$", reference)
                            if match:
                                max_nonempty_row = max(
                                    max_nonempty_row,
                                    int(match.group(1)),
                                )
                        if formula:
                            counters["formula_cells"] += 1
                            upper = formula.upper()
                            if any(token in upper for token in ERROR_TOKENS):
                                counters["formula_error_cells"] += 1
                        cell_type = str(cell.attrib.get("t") or "")
                        if cell_type == "e":
                            counters["error_cells"] += 1
                        elif cell_type == "s" and value is not None:
                            try:
                                _add_text_signal(
                                    counters,
                                    shared_signals.get(int(value)),
                                )
                            except (TypeError, ValueError):
                                pass
                        elif inline_text:
                            _add_text_signal(
                                counters,
                                _text_signal(inline_text),
                            )
                        if value is not None and cell_type in {"", "n"}:
                            try:
                                numeric = float(value)
                            except (TypeError, ValueError):
                                numeric = None
                            if numeric is not None:
                                counters["numeric_cells"] += 1
                                if not math.isfinite(numeric):
                                    counters["nonfinite_numeric_cells"] += 1
                                else:
                                    counters["numeric_min"] = (
                                        numeric
                                        if counters["numeric_min"] is None
                                        else min(counters["numeric_min"], numeric)
                                    )
                                    counters["numeric_max"] = (
                                        numeric
                                        if counters["numeric_max"] is None
                                        else max(counters["numeric_max"], numeric)
                                    )
                                    if not numeric.is_integer():
                                        counters[
                                            "fractional_numeric_cells"
                                        ] += 1
                                    if 0 < abs(numeric) < 1e-9:
                                        counters[
                                            "tiny_nonzero_numeric_cells"
                                        ] += 1
                                try:
                                    style_index = int(cell.attrib.get("s", 0))
                                except ValueError:
                                    style_index = 0
                                if style_index in percent_styles:
                                    counters["percent_numeric_cells"] += 1
                                    if abs(numeric) > 1.0000001:
                                        counters[
                                            "percent_abs_gt_one_cells"
                                        ] += 1
                        cell.clear()
                if str(sheet.get("state") or "visible") == "visible":
                    if nonempty_cells == 0:
                        counters["empty_visible_sheets"] += 1
                    elif max_nonempty_row <= 1:
                        counters["header_only_visible_sheets"] += 1
    except Exception as exc:
        return {
            "valid": False,
            "error": f"内容扫描失败（{type(exc).__name__}）",
            "sheet_count": int(structure.get("sheet_count", 0) or 0),
            **counters,
        }
    return {
        "valid": True,
        "error": "",
        "sheet_count": int(structure.get("sheet_count", 0) or 0),
        **counters,
    }


def _is_metric_header(value: Any) -> bool:
    header = str(value or "").strip()
    if not header:
        return False
    upper = header.upper()
    parts = set(upper.split("_"))
    return (
        upper == "MEAN"
        or upper.endswith("_SCORE")
        or upper.endswith("_INDEX")
        or any(
            token in parts or upper.startswith(f"平均{token}")
            for token in METRIC_HEADER_TOKENS
        )
    )


def _profile_first_sheet_metrics(path: Path) -> Dict[str, Any]:
    headers, rows, workbook = _first_sheet_table(path)
    try:
        positions = [
            index
            for index, header in enumerate(headers)
            if _is_metric_header(header)
        ]
        numeric_counts = [0] * len(positions)
        nonzero_counts = [0] * len(positions)
        fractional_counts = [0] * len(positions)
        text_percent_counts = [0] * len(positions)
        minimum = None
        maximum = None
        for row in rows:
            if _is_blank_row(row):
                continue
            for offset, position in enumerate(positions):
                value = row[position] if position < len(row) else None
                numeric = None
                if isinstance(value, bool):
                    continue
                if isinstance(value, (int, float)) and math.isfinite(
                    float(value)
                ):
                    numeric = float(value)
                elif isinstance(value, str):
                    signal = _text_signal(value)
                    if signal is not None and signal[0] == "percent":
                        text_percent_counts[offset] += 1
                        numeric = float(signal[1])
                if numeric is None:
                    continue
                numeric_counts[offset] += 1
                if numeric != 0:
                    nonzero_counts[offset] += 1
                if not numeric.is_integer():
                    fractional_counts[offset] += 1
                minimum = numeric if minimum is None else min(minimum, numeric)
                maximum = numeric if maximum is None else max(maximum, numeric)
        all_blank = sum(count == 0 for count in numeric_counts)
        all_zero = sum(
            numeric_counts[index] > 0 and nonzero_counts[index] == 0
            for index in range(len(positions))
        )
        return {
            "metric_columns": len(positions),
            "metric_numeric_cells": sum(numeric_counts),
            "metric_nonzero_cells": sum(nonzero_counts),
            "metric_fractional_cells": sum(fractional_counts),
            "metric_text_percent_cells": sum(text_percent_counts),
            "metric_all_blank_columns": all_blank,
            "metric_all_zero_columns": all_zero,
            "metric_min": minimum,
            "metric_max": maximum,
        }
    finally:
        workbook.close()


def _validate_xlsx_files(
    project: Path,
    paths: Iterable[Path],
    tasks: Sequence[Dict[str, Any]],
    result: ResultBuilder,
) -> None:
    farm_by_child = {
        str(task["child_path"].resolve()): task for task in tasks
    }
    for path in sorted(set(paths), key=lambda value: str(value)):
        try:
            relative = _relative_to_project(project, path)
            owner = next(
                (
                    task
                    for child, task in farm_by_child.items()
                    if path.resolve() == Path(child)
                    or Path(child) in path.resolve().parents
                ),
                None,
            )
            farm_code = str(owner.get("farm_code") or "") if owner else ""
            health = _scan_xlsx_health(path)
            if path.name in METRIC_PROFILE_FILES and health["valid"]:
                health.update(_profile_first_sheet_metrics(path))
            record = {
                "farm_code": farm_code,
                "relative_path": relative,
                "bytes": int(path.stat().st_size) if path.is_file() else 0,
                "sha256": stream_sha256(path) if path.is_file() else "",
                **health,
            }
            result.file_rows.append(record)
        except Exception as exc:
            result.add(
                "error",
                "xlsx",
                "xlsx_scan_failed",
                f"工作簿扫描失败（{type(exc).__name__}）",
                relative_path=str(path),
            )
            continue
        if not health["valid"]:
            result.add(
                "error",
                "xlsx",
                "xlsx_invalid",
                str(health["error"]),
                farm_code=farm_code,
                relative_path=relative,
            )
            continue
        for key, code, message in (
            (
                "empty_visible_sheets",
                "empty_visible_sheet",
                "工作簿存在完全空白的可见 Sheet",
            ),
            (
                "formula_error_cells",
                "formula_reference_error",
                "工作簿公式包含错误引用或错误标记",
            ),
            (
                "error_cells",
                "excel_error_cell",
                "工作簿包含 Excel 错误类型单元格",
            ),
            (
                "nonfinite_numeric_cells",
                "nonfinite_numeric_cell",
                "工作簿包含非有限数值",
            ),
            (
                "literal_error_marker_cells",
                "literal_error_marker",
                "工作簿包含以文本保存的 Excel 错误标记",
            ),
        ):
            if int(health[key]) > 0:
                result.add(
                    "error",
                    "xlsx",
                    code,
                    f"{message}（{int(health[key])} 个）",
                    farm_code=farm_code,
                    relative_path=relative,
                )
        if int(health["header_only_visible_sheets"]) > 0:
            result.add(
                "warning",
                "xlsx",
                "header_only_sheet",
                "工作簿存在只有首行内容的可见 Sheet，需结合缺失数据章节复核",
                farm_code=farm_code,
                relative_path=relative,
            )
        if int(health["percent_abs_gt_one_cells"]) > 0:
            result.add(
                "warning",
                "xlsx",
                "percent_scale_suspicious",
                "百分比格式单元格存在绝对值大于 100% 的数值，需复核小数点口径",
                farm_code=farm_code,
                relative_path=relative,
            )
        if int(health["text_percent_abs_gt_100_cells"]) > 0:
            result.add(
                "warning",
                "xlsx",
                "text_percent_scale_suspicious",
                "文本百分比存在绝对值大于 100% 的结果，需复核小数点口径",
                farm_code=farm_code,
                relative_path=relative,
            )
        metric_columns = int(health.get("metric_columns", 0) or 0)
        if (
            metric_columns > 0
            and int(health.get("metric_all_blank_columns", 0) or 0)
            == metric_columns
        ):
            result.add(
                "warning",
                "data_quality",
                "all_metrics_blank",
                "首个 Sheet 的全部指标列均无可识别数值",
                farm_code=farm_code,
                relative_path=relative,
            )
        elif (
            metric_columns > 0
            and int(health.get("metric_all_zero_columns", 0) or 0)
            == metric_columns
        ):
            result.add(
                "warning",
                "data_quality",
                "all_metrics_zero",
                "首个 Sheet 的全部指标列均为 0，需复核接口数据和公牛识别情况",
                farm_code=farm_code,
                relative_path=relative,
            )


def _detail_audit_digest(
    row: Sequence[Any],
    positions: Mapping[str, int],
) -> bytes:
    fields = (
        "API farmcode",
        "子项目相对目录",
        "源文件",
        "源数据行号",
    )
    return _projected_row_hash(
        row,
        [positions[field] for field in fields],
    )


def _insert_detail_row_audit_batch(
    connection: sqlite3.Connection,
    slot: str,
    batch: Sequence[tuple[bytes]],
) -> None:
    columns = {
        "ranked": "ranked_count",
        "reconciliation_ranked": "reconciliation_ranked_count",
        "reconciliation": "reconciliation_count",
    }
    column = columns.get(slot)
    if column is None:
        raise ValueError("内部完整明细审计槽位无效")
    connection.executemany(
        f"""
        INSERT INTO detail_row_audit (row_hash, {column})
        VALUES (?, 1)
        ON CONFLICT(row_hash) DO UPDATE
        SET {column} = {column} + 1
        """,
        batch,
    )


def _coerce_positive_integer(value: Any) -> Optional[int]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer() and value > 0:
            return int(value)
        return None
    text = str(value).strip()
    if text.isdigit() and int(text) > 0:
        return int(text)
    return None


def _scan_detail_volume(
    path: Path,
    *,
    kind: str,
    include_evidence: bool,
    connection: sqlite3.Connection,
) -> Dict[str, Any]:
    headers, rows, workbook = _first_sheet_table(path)
    try:
        stats: Dict[str, Any] = {
            "actual_rows": 0,
            "classified_ranked_rows": 0,
            "classified_unranked_rows": 0,
            "invalid_fixed_rows": 0,
            "unranked_reason_counts": Counter(),
            "source_row_stats": {},
        }
        if not include_evidence:
            stats["actual_rows"] = sum(
                1 for row in rows if not _is_blank_row(row)
            )
            return stats

        positions = _header_positions(headers)
        required = {
            "牧场组排名",
            "分类结果",
            "未排名原因",
            "API farmcode",
            "子项目相对目录",
            "源文件",
            "源数据行号",
        }
        if not required.issubset(positions):
            raise ValueError("完整明细分卷缺少固定审计列")

        ranked_batch: List[tuple[bytes]] = []
        reconciliation_batch: List[tuple[bytes]] = []
        reconciliation_ranked_batch: List[tuple[bytes]] = []
        rank_batch: List[tuple[int]] = []
        for row in rows:
            if _is_blank_row(row):
                continue
            stats["actual_rows"] += 1
            classification_value = (
                row[positions["分类结果"]]
                if positions["分类结果"] < len(row)
                else None
            )
            reason_value = (
                row[positions["未排名原因"]]
                if positions["未排名原因"] < len(row)
                else None
            )
            source_file_value = (
                row[positions["源文件"]]
                if positions["源文件"] < len(row)
                else None
            )
            classification = (
                "" if classification_value is None
                else str(classification_value).strip()
            )
            reason = (
                "" if reason_value is None
                else str(reason_value).strip()
            )
            source_file = (
                "" if source_file_value is None
                else str(source_file_value).strip()
            )
            source_row = _coerce_positive_integer(
                row[positions["源数据行号"]]
                if positions["源数据行号"] < len(row)
                else None
            )
            rank = _coerce_positive_integer(
                row[positions["牧场组排名"]]
                if positions["牧场组排名"] < len(row)
                else None
            )
            valid = bool(source_file and source_row is not None)
            if kind == "ranked":
                valid = (
                    valid
                    and classification == "有效在群排名"
                    and rank is not None
                    and not reason
                )
                stats["classified_ranked_rows"] += int(
                    classification == "有效在群排名"
                )
            elif kind == "reconciliation":
                if classification == "有效在群排名":
                    stats["classified_ranked_rows"] += 1
                    valid = valid and rank is not None and not reason
                elif classification == "未排名":
                    stats["classified_unranked_rows"] += 1
                    stats["unranked_reason_counts"][reason] += 1
                    valid = valid and rank is None and bool(reason)
                else:
                    valid = False
            else:
                raise ValueError("完整明细分卷类型无效")
            if not valid:
                stats["invalid_fixed_rows"] += 1
                continue

            digest = _detail_audit_digest(row, positions)
            if kind == "ranked":
                ranked_batch.append((digest,))
            else:
                reconciliation_batch.append((digest,))
                if classification == "有效在群排名":
                    reconciliation_ranked_batch.append((digest,))

            if rank is not None:
                rank_batch.append((rank,))
            if kind == "reconciliation":
                source_stats = stats["source_row_stats"].setdefault(
                    source_file,
                    {
                        "count": 0,
                        "minimum": None,
                        "maximum": None,
                        "sum": 0,
                        "square_sum": 0,
                    },
                )
                source_stats["count"] += 1
                source_stats["minimum"] = (
                    source_row
                    if source_stats["minimum"] is None
                    else min(source_stats["minimum"], source_row)
                )
                source_stats["maximum"] = (
                    source_row
                    if source_stats["maximum"] is None
                    else max(source_stats["maximum"], source_row)
                )
                source_stats["sum"] += source_row
                source_stats["square_sum"] += source_row * source_row

            if len(reconciliation_batch) >= 2_000:
                _insert_detail_row_audit_batch(
                    connection,
                    "reconciliation",
                    reconciliation_batch,
                )
                reconciliation_batch.clear()
            if len(reconciliation_ranked_batch) >= 2_000:
                _insert_detail_row_audit_batch(
                    connection,
                    "reconciliation_ranked",
                    reconciliation_ranked_batch,
                )
                reconciliation_ranked_batch.clear()
            if len(ranked_batch) >= 2_000:
                _insert_detail_row_audit_batch(
                    connection,
                    "ranked",
                    ranked_batch,
                )
                ranked_batch.clear()
            if len(rank_batch) >= 2_000:
                connection.executemany(
                    """
                    INSERT INTO detail_rank_audit (rank_value, occurrences)
                    VALUES (?, 1)
                    ON CONFLICT(rank_value) DO UPDATE SET
                        occurrences = occurrences + 1
                    """,
                    rank_batch,
                )
                rank_batch.clear()

        for slot, batch in (
            ("ranked", ranked_batch),
            ("reconciliation", reconciliation_batch),
            ("reconciliation_ranked", reconciliation_ranked_batch),
        ):
            if batch:
                _insert_detail_row_audit_batch(connection, slot, batch)
        if rank_batch:
            connection.executemany(
                """
                INSERT INTO detail_rank_audit (rank_value, occurrences)
                VALUES (?, 1)
                ON CONFLICT(rank_value) DO UPDATE SET
                    occurrences = occurrences + 1
                """,
                rank_batch,
            )
        connection.commit()
        return stats
    finally:
        workbook.close()


def _sum_of_squares(first: int, last: int) -> int:
    if last < first:
        return 0

    def through(value: int) -> int:
        return value * (value + 1) * (2 * value + 1) // 6

    return through(last) - through(first - 1)


def _validate_detail_volume_files(
    manifest: Mapping[str, Any],
    *,
    detail_root: Path,
    expected_rows_by_kind: Mapping[str, int],
    declared_rows_by_kind_part: Mapping[str, Mapping[int, int]],
) -> Dict[str, Any]:
    volumes = manifest.get("volumes")
    sources = manifest.get("sources")
    counts = manifest.get("counts")
    if not isinstance(volumes, Mapping) or not isinstance(sources, list):
        raise ValueError("完整明细缺少来源或分卷清单")
    if not isinstance(counts, Mapping):
        raise ValueError("完整明细缺少 counts")

    expected_sources: Dict[str, int] = {}
    for source in sources:
        if not isinstance(source, Mapping):
            raise ValueError("完整明细来源记录无效")
        source_path = str(source.get("path") or "").strip()
        rows_read = int(source.get("rows_read", -1))
        if (
            not source_path
            or source_path in expected_sources
            or rows_read < 0
        ):
            raise ValueError("完整明细来源路径或行数无效")
        expected_sources[source_path] = rows_read

    actual_rows_by_kind_part: Dict[str, Dict[int, int]] = defaultdict(
        lambda: defaultdict(int)
    )
    actual_volume_count = 0
    classified_ranked_rows = 0
    classified_unranked_rows = 0
    invalid_fixed_rows = 0
    actual_reason_counts: Counter[str] = Counter()
    source_row_stats: Dict[str, Dict[str, int]] = {}

    with tempfile.TemporaryDirectory(
        prefix="multi_farm_detail_audit_"
    ) as temporary_dir:
        connection = sqlite3.connect(
            str(Path(temporary_dir) / "detail_audit.sqlite3")
        )
        try:
            connection.execute("PRAGMA temp_store = FILE")
            connection.execute("PRAGMA cache_size = -16384")
            connection.execute(
                """
                CREATE TABLE detail_row_audit (
                    row_hash BLOB PRIMARY KEY,
                    ranked_count INTEGER NOT NULL DEFAULT 0,
                    reconciliation_ranked_count INTEGER NOT NULL DEFAULT 0,
                    reconciliation_count INTEGER NOT NULL DEFAULT 0
                ) WITHOUT ROWID
                """
            )
            connection.execute(
                """
                CREATE TABLE detail_rank_audit (
                    rank_value INTEGER PRIMARY KEY,
                    occurrences INTEGER NOT NULL DEFAULT 0
                ) WITHOUT ROWID
                """
            )

            for kind in ("ranked", "reconciliation", "long_fields"):
                entries = volumes.get(kind)
                if not isinstance(entries, list):
                    raise ValueError(f"完整明细缺少 {kind} 分卷")
                for entry in entries:
                    if not isinstance(entry, Mapping):
                        raise ValueError("完整明细分卷记录无效")
                    relative = str(entry.get("path") or "")
                    volume_path = _safe_relative(
                        detail_root,
                        relative,
                        "完整明细分卷",
                    )
                    if (
                        not volume_path.is_file()
                        or volume_path.is_symlink()
                    ):
                        raise ValueError("完整明细声明的分卷不存在")
                    declared_bytes = int(entry.get("bytes", -1))
                    declared_sha256 = str(entry.get("sha256") or "")
                    if (
                        declared_bytes <= 0
                        or volume_path.stat().st_size != declared_bytes
                        or not re.fullmatch(
                            r"[0-9a-f]{64}",
                            declared_sha256,
                        )
                        or stream_sha256(volume_path) != declared_sha256
                    ):
                        raise ValueError("完整明细分卷大小或摘要不一致")
                    column_part = int(entry.get("column_part", 1))
                    include_evidence = (
                        column_part == 1
                        and kind in {"ranked", "reconciliation"}
                    )
                    scanned = _scan_detail_volume(
                        volume_path,
                        kind=kind,
                        include_evidence=include_evidence,
                        connection=connection,
                    ) if kind != "long_fields" else {
                        "actual_rows": _count_table_rows(volume_path),
                        "classified_ranked_rows": 0,
                        "classified_unranked_rows": 0,
                        "invalid_fixed_rows": 0,
                        "unranked_reason_counts": Counter(),
                        "source_row_stats": {},
                    }
                    actual_rows = int(scanned["actual_rows"])
                    if actual_rows != int(entry.get("data_rows", -1)):
                        raise ValueError(
                            "完整明细分卷实际数据行数与 manifest 不一致"
                        )
                    actual_rows_by_kind_part[kind][
                        column_part
                    ] += actual_rows
                    actual_volume_count += 1
                    if include_evidence:
                        classified_ranked_rows += int(
                            scanned["classified_ranked_rows"]
                        )
                        classified_unranked_rows += int(
                            scanned["classified_unranked_rows"]
                        )
                        invalid_fixed_rows += int(
                            scanned["invalid_fixed_rows"]
                        )
                        actual_reason_counts.update(
                            scanned["unranked_reason_counts"]
                        )
                        for source_path, stats in scanned[
                            "source_row_stats"
                        ].items():
                            aggregate = source_row_stats.setdefault(
                                source_path,
                                {
                                    "count": 0,
                                    "minimum": None,
                                    "maximum": None,
                                    "sum": 0,
                                    "square_sum": 0,
                                },
                            )
                            aggregate["count"] += int(stats["count"])
                            aggregate["minimum"] = (
                                stats["minimum"]
                                if aggregate["minimum"] is None
                                else min(
                                    aggregate["minimum"],
                                    stats["minimum"],
                                )
                            )
                            aggregate["maximum"] = (
                                stats["maximum"]
                                if aggregate["maximum"] is None
                                else max(
                                    aggregate["maximum"],
                                    stats["maximum"],
                                )
                            )
                            aggregate["sum"] += int(stats["sum"])
                            aggregate["square_sum"] += int(
                                stats["square_sum"]
                            )

            for kind, declared_parts in declared_rows_by_kind_part.items():
                actual_parts = actual_rows_by_kind_part.get(kind, {})
                if dict(actual_parts) != dict(declared_parts):
                    raise ValueError("完整明细实际分卷累计行数不一致")
                expected_rows = int(expected_rows_by_kind[kind])
                if any(
                    int(total) != expected_rows
                    for total in actual_parts.values()
                ):
                    raise ValueError("完整明细实际分片行数未完整覆盖")

            if invalid_fixed_rows:
                raise ValueError("完整明细固定审计列存在无效行")
            expected_ranked = int(expected_rows_by_kind["ranked"])
            expected_reconciliation = int(
                expected_rows_by_kind["reconciliation"]
            )
            expected_unranked = expected_reconciliation - expected_ranked
            if (
                classified_ranked_rows != expected_ranked * 2
                or classified_unranked_rows != expected_unranked
            ):
                raise ValueError("完整明细实际分类行数与总计数不一致")
            declared_reason_counts = counts.get("unranked_reason_counts")
            if not isinstance(declared_reason_counts, Mapping):
                raise ValueError("完整明细未排名原因无效")
            normalized_reasons = {
                str(reason): int(value)
                for reason, value in declared_reason_counts.items()
            }
            if dict(actual_reason_counts) != normalized_reasons:
                raise ValueError("完整明细实际未排名原因与 manifest 不一致")

            audit = connection.execute(
                """
                SELECT
                  COALESCE(SUM(
                    ABS(ranked_count - reconciliation_ranked_count)
                  ), 0),
                  COALESCE(SUM(
                    CASE WHEN ranked_count > 1
                    THEN ranked_count - 1 ELSE 0 END
                  ), 0),
                  COALESCE(SUM(
                    CASE WHEN reconciliation_count > 1
                    THEN reconciliation_count - 1 ELSE 0 END
                  ), 0),
                  COALESCE(SUM(ranked_count), 0),
                  COALESCE(SUM(reconciliation_count), 0)
                FROM detail_row_audit
                """
            ).fetchone()
            if (
                int(audit[0]) != 0
                or int(audit[1]) != 0
                or int(audit[2]) != 0
                or int(audit[3]) != expected_ranked
                or int(audit[4]) != expected_reconciliation
            ):
                raise ValueError("完整排名与全部源行的磁盘血缘不一致")

            rank_stats = connection.execute(
                """
                SELECT
                  COUNT(*),
                  COALESCE(SUM(occurrences), 0),
                  COALESCE(MIN(rank_value), 0),
                  COALESCE(MAX(rank_value), 0),
                  COALESCE(SUM(rank_value * occurrences), 0),
                  COALESCE(SUM(rank_value * rank_value * occurrences), 0),
                  COALESCE(SUM(
                    CASE WHEN occurrences > 2
                    THEN occurrences - 2 ELSE 0 END
                  ), 0)
                FROM detail_rank_audit
                """
            ).fetchone()
            # 每个有效排名应在“排名分卷”和“全部源行分卷”中各出现一次。
            expected_rank_occurrences = expected_ranked * 2
            if expected_ranked:
                rank_valid = (
                    int(rank_stats[0]) == expected_ranked
                    and int(rank_stats[1]) == expected_rank_occurrences
                    and int(rank_stats[2]) == 1
                    and int(rank_stats[3]) == expected_ranked
                    and int(rank_stats[4])
                    == expected_ranked * (expected_ranked + 1)
                    and int(rank_stats[5])
                    == 2 * _sum_of_squares(1, expected_ranked)
                    and int(rank_stats[6]) == 0
                )
            else:
                rank_valid = int(rank_stats[1]) == 0
            if not rank_valid:
                raise ValueError("完整明细实际排名序列不连续或重复")

            if set(source_row_stats) != set(expected_sources):
                raise ValueError("完整明细实际来源文件集合不一致")
            for source_path, expected_rows in expected_sources.items():
                stats = source_row_stats[source_path]
                first = 2
                last = expected_rows + 1
                if (
                    int(stats["count"]) != expected_rows
                    or (
                        expected_rows
                        and (
                            int(stats["minimum"]) != first
                            or int(stats["maximum"]) != last
                            or int(stats["sum"])
                            != (first + last) * expected_rows // 2
                            or int(stats["square_sum"])
                            != _sum_of_squares(first, last)
                        )
                    )
                ):
                    raise ValueError(
                        "完整明细实际来源行号未连续完整覆盖"
                    )
        finally:
            connection.close()

    return {
        "actual_volume_count": actual_volume_count,
        "actual_ranked_rows": int(
            actual_rows_by_kind_part.get("ranked", {}).get(1, 0)
        ),
        "actual_reconciliation_rows": int(
            actual_rows_by_kind_part.get(
                "reconciliation",
                {},
            ).get(1, 0)
        ),
        "actual_long_field_chunk_rows": int(
            actual_rows_by_kind_part.get("long_fields", {}).get(1, 0)
        ),
        "actual_sources_verified": len(source_row_stats),
        "row_lineage_mismatch_rows": 0,
    }


def _validate_detail_counts(
    manifest: Mapping[str, Any],
    *,
    included_task_ids: set[str],
    detail_root: Path,
    expected_source_rows_by_task: Optional[Mapping[str, int]] = None,
) -> Dict[str, Any]:
    if manifest.get("status") != "complete":
        raise ValueError("完整明细 manifest 状态不是 complete")
    counts = manifest.get("counts")
    sources = manifest.get("sources")
    volumes = manifest.get("volumes")
    if not isinstance(counts, Mapping):
        raise ValueError("完整明细缺少 counts")
    if not isinstance(sources, list) or not isinstance(volumes, Mapping):
        raise ValueError("完整明细缺少来源或分卷清单")

    def count(name: str) -> int:
        try:
            value = int(counts[name])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"完整明细计数 {name} 无效") from exc
        if value < 0:
            raise ValueError(f"完整明细计数 {name} 不能为负数")
        return value

    source_rows = count("source_rows")
    ranked_rows = count("valid_ranked_rows")
    unranked_rows = count("unranked_rows")
    ranked_exported = count("ranked_exported_rows")
    reconciliation_exported = count("reconciliation_exported_rows")
    tasks_in_scope = count("tasks_in_scope")
    source_files_read = count("source_files_read")
    source_problems = count("source_files_with_problem")
    long_field_count = count("long_field_count")
    long_chunk_count = count("long_field_chunk_count")
    if (
        source_problems
        or tasks_in_scope != len(included_task_ids)
        or source_files_read != tasks_in_scope
        or len(sources) != tasks_in_scope
        or source_rows != ranked_rows + unranked_rows
        or ranked_rows != ranked_exported
        or source_rows != reconciliation_exported
    ):
        raise ValueError("完整明细总计数未通过无丢失对账")

    seen_source_keys = set()
    source_task_ids = set()
    rows_from_sources = 0
    for source in sources:
        if not isinstance(source, Mapping) or source.get("status") != "read":
            raise ValueError("完整明细存在未成功读取的来源")
        source_key = str(source.get("source_key") or "")
        task_id = str(source.get("task_id") or "")
        if not source_key or source_key in seen_source_keys:
            raise ValueError("完整明细来源键为空或重复")
        seen_source_keys.add(source_key)
        source_task_ids.add(task_id)
        rows_read = int(source.get("rows_read", -1))
        rows_from_sources += rows_read
        if int(source.get("duplicate_cow_id_count", 0) or 0) != 0:
            raise ValueError("完整明细来源存在重复牛号")
        if int(source.get("lineage_mismatch_rows", 0) or 0) != 0:
            raise ValueError("完整明细来源存在牧场归属不一致")
        if source.get("identity_match") is False:
            raise ValueError("完整明细来源指数与直接输入牛号血缘不一致")
        if expected_source_rows_by_task is not None:
            expected_rows = expected_source_rows_by_task.get(task_id)
            if expected_rows is None or rows_read != int(expected_rows):
                raise ValueError("完整明细逐牧场来源行数与当前指数结果不一致")
    if rows_from_sources != source_rows or source_task_ids != included_task_ids:
        raise ValueError("完整明细来源行数或任务范围不一致")

    reason_counts = counts.get("unranked_reason_counts")
    if not isinstance(reason_counts, Mapping) or sum(
        int(value) for value in reason_counts.values()
    ) != unranked_rows:
        raise ValueError("完整明细未排名原因合计不一致")

    volume_totals: Dict[str, Dict[int, int]] = defaultdict(
        lambda: defaultdict(int)
    )
    volume_numbers: Dict[str, Dict[int, List[int]]] = defaultdict(
        lambda: defaultdict(list)
    )
    declared_parts: Dict[str, set[int]] = defaultdict(set)
    volume_count = 0
    seen_paths = set()
    for kind, expected_rows in (
        ("ranked", ranked_rows),
        ("reconciliation", source_rows),
        ("long_fields", long_chunk_count),
    ):
        entries = volumes.get(kind)
        if not isinstance(entries, list):
            raise ValueError(f"完整明细缺少 {kind} 分卷")
        for entry in entries:
            if not isinstance(entry, Mapping):
                raise ValueError("完整明细分卷记录无效")
            relative = str(entry.get("path") or "")
            if not relative or relative in seen_paths:
                raise ValueError("完整明细分卷路径为空或重复")
            seen_paths.add(relative)
            data_rows = int(entry.get("data_rows", -1))
            column_part = int(entry.get("column_part", 1))
            column_parts = int(entry.get("column_parts", 1))
            volume = int(entry.get("volume", -1))
            rows_per_volume = int(entry.get("rows_per_volume", -1))
            if (
                data_rows < 0
                or data_rows > EXCEL_MAX_DATA_ROWS
                or column_part < 1
                or column_parts < column_part
                or volume < 1
                or rows_per_volume < 1
                or data_rows > rows_per_volume
            ):
                raise ValueError("完整明细分卷行数或卷号无效")
            volume_count += 1
            declared_parts[kind].add(column_parts)
            volume_totals[kind][column_part] += data_rows
            volume_numbers[kind][column_part].append(volume)
        if expected_rows and not entries:
            raise ValueError(f"完整明细 {kind} 有数据但没有分卷")
        if entries:
            if len(declared_parts[kind]) != 1:
                raise ValueError(f"完整明细 {kind} 字段分片声明不一致")
            expected_parts = set(range(1, next(iter(declared_parts[kind])) + 1))
            if set(volume_totals[kind]) != expected_parts:
                raise ValueError(f"完整明细 {kind} 字段分片不连续")
            for part, total in volume_totals[kind].items():
                if total != expected_rows:
                    raise ValueError(f"完整明细 {kind} 分片累计行数不一致")
                numbers = sorted(volume_numbers[kind][part])
                if numbers != list(range(1, len(numbers) + 1)):
                    raise ValueError(f"完整明细 {kind} 卷号不连续")
    if (long_field_count == 0) != (long_chunk_count == 0):
        raise ValueError("完整明细超长字段与分块计数不一致")
    actual_evidence = _validate_detail_volume_files(
        manifest,
        detail_root=Path(detail_root),
        expected_rows_by_kind={
            "ranked": ranked_rows,
            "reconciliation": source_rows,
            "long_fields": long_chunk_count,
        },
        declared_rows_by_kind_part={
            kind: dict(parts)
            for kind, parts in volume_totals.items()
        },
    )
    return {
        "tasks_in_scope": tasks_in_scope,
        "source_rows": source_rows,
        "valid_ranked_rows": ranked_rows,
        "unranked_rows": unranked_rows,
        "reconciliation_rows": reconciliation_exported,
        "volume_count": volume_count,
        "long_field_count": long_field_count,
        "long_field_chunk_count": long_chunk_count,
        **actual_evidence,
    }


def _validate_snapshot_is_current(
    project: Path,
    snapshot_path: Path,
    state: Mapping[str, Any],
    tasks: Sequence[Mapping[str, Any]],
) -> None:
    snapshot = _read_json(snapshot_path, "报告包发布快照")
    basis = snapshot.get("basis")
    if not isinstance(basis, Mapping):
        raise ValueError("报告包发布快照缺少 basis")
    if int(basis.get("selection_revision", -1)) != int(
        state["selection_revision"]
    ):
        raise ValueError("报告包发布快照选择版本已过期")
    scope = basis.get("selection_scope")
    if not isinstance(scope, Mapping):
        raise ValueError("报告包发布快照缺少选择范围")
    expected_ids = [str(task["task_id"]) for task in tasks]
    if list(scope.get("included_task_ids") or []) != expected_ids:
        raise ValueError("报告包发布快照牧场范围已过期")
    snapshot_tasks = {
        str(task.get("task_id") or ""): task
        for task in basis.get("tasks", [])
        if isinstance(task, Mapping)
    }
    for task in tasks:
        captured = snapshot_tasks.get(str(task["task_id"]))
        if not isinstance(captured, Mapping):
            raise ValueError("报告包发布快照缺少当前牧场")
        if (
            str(captured.get("status") or "") != str(task["status"])
            or int(captured.get("attempt", 0) or 0)
            != int(task.get("attempt", 0) or 0)
        ):
            raise ValueError("报告包发布快照任务状态已过期")
        captured_stages = {
            str(stage.get("stage") or ""): stage
            for stage in captured.get("stages", [])
            if isinstance(stage, Mapping)
        }
        for stage_name, stage in task["stages"].items():
            if not stage.get("required"):
                continue
            captured_stage = captured_stages.get(stage_name)
            if not isinstance(captured_stage, Mapping):
                raise ValueError("报告包发布快照缺少必需阶段")
            manifest_path = task["child_path"] / stage_manifest_path(stage_name)
            if (
                int(captured_stage.get("attempt", 0) or 0)
                != int(stage.get("attempt", 0) or 0)
                or str(captured_stage.get("manifest_sha256") or "")
                != stream_sha256(manifest_path)
            ):
                raise ValueError("报告包发布快照阶段 manifest 已过期")


def _validate_group_report(
    project: Path,
    state: Mapping[str, Any],
    tasks: Sequence[Dict[str, Any]],
    result: ResultBuilder,
) -> set[Path]:
    paths: set[Path] = set()
    try:
        pointer = validate_current_group_report_pointer(project)
        package = Path(pointer["package_path"])
        batch = _read_json(
            Path(pointer["batch_manifest_path"]),
            "报告包 manifest",
        )
        detail_entry = batch.get("detail")
        if not isinstance(detail_entry, Mapping):
            raise ValueError("报告包 manifest 缺少完整明细")
        detail_manifest_path = package / str(detail_entry.get("relative_path") or "")
        detail_manifest = _read_json(detail_manifest_path, "完整明细 manifest")
        counts = _validate_detail_counts(
            detail_manifest,
            included_task_ids={str(task["task_id"]) for task in tasks},
            detail_root=detail_manifest_path.parent,
            expected_source_rows_by_task={
                str(task["task_id"]): int(task["_acceptance_index_rows"])
                for task in tasks
            },
        )
        snapshot_entry = batch.get("publication_snapshot")
        if not isinstance(snapshot_entry, Mapping):
            raise ValueError("报告包 manifest 缺少发布快照")
        snapshot_path = package / str(snapshot_entry.get("relative_path") or "")
        _validate_snapshot_is_current(project, snapshot_path, state, tasks)
        paths.add(Path(pointer["excel_path"]))
        for entries in detail_manifest.get("volumes", {}).values():
            if not isinstance(entries, list):
                continue
            for entry in entries:
                if isinstance(entry, Mapping):
                    paths.add(
                        detail_manifest_path.parent
                        / str(entry.get("path") or "")
                    )
        result.group_report = {
            "valid": True,
            "package_relative_path": _relative_to_project(project, package),
            "excel_relative_path": _relative_to_project(
                project,
                Path(pointer["excel_path"]),
            ),
            "detail_manifest_relative_path": _relative_to_project(
                project,
                detail_manifest_path,
            ),
            **counts,
        }
    except (GroupReportPublicationError, ValueError, OSError) as exc:
        result.group_report = {"valid": False}
        result.add(
            "error",
            "group_report",
            "group_report_invalid",
            f"牧场组正式报告包或完整明细校验失败（{type(exc).__name__}）",
        )
    return paths


def validate_project(project_path: Path) -> Dict[str, Any]:
    """只读验证一个已完成牧场组，并返回可序列化聚合结果。"""
    project = Path(project_path).resolve()
    result = ResultBuilder(project)
    try:
        state, tasks = _static_gate(project, result)
    except AcceptanceBlocked as exc:
        result.gate = {
            "passed": False,
            "reason": str(exc),
            "xlsx_opened": False,
        }
        result.add("error", "gate", "acceptance_blocked", str(exc))
        return result.payload(blocked=True)

    # 只有上面的静态终态门禁通过后，以下代码才会读取 XLSX。
    manifest_files = _validate_stage_manifests(project, tasks, result)
    required_files = _collect_required_files(
        project,
        tasks,
        manifest_files,
        result,
    )
    _validate_lineage(project, tasks, result)
    _validate_formula_integrity(tasks, result)
    group_files = _validate_group_report(project, state, tasks, result)
    _validate_xlsx_files(
        project,
        required_files | group_files,
        tasks,
        result,
    )

    for task in tasks:
        farm_issues = [
            issue
            for issue in result.issues
            if issue.farm_code == task["farm_code"]
        ]
        result.farm_rows.append(
            {
                "farm_code": task["farm_code"],
                "farm_name": task["farm_name"],
                "task_status": task["status"],
                "required_stages": sum(
                    bool(stage.get("required"))
                    for stage in task["stages"].values()
                ),
                "errors": sum(
                    issue.severity == "error" for issue in farm_issues
                ),
                "warnings": sum(
                    issue.severity == "warning" for issue in farm_issues
                ),
                "passed": not any(
                    issue.severity == "error" for issue in farm_issues
                ),
                "child_relative_path": _relative_to_project(
                    project,
                    task["child_path"],
                ),
            }
        )
    return result.payload()


def _atomic_write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as stream:
            json.dump(
                value,
                stream,
                ensure_ascii=False,
                indent=2,
                allow_nan=False,
            )
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_write_csv(
    path: Path,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(
        dict.fromkeys(
            key for row in rows for key in row.keys()
        )
    )
    if not fieldnames:
        fieldnames = ["empty"]
        rows = [{"empty": ""}]
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(
            descriptor,
            "w",
            encoding="utf-8-sig",
            newline="",
        ) as stream:
            writer = csv.DictWriter(stream, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_outputs(
    payload: Mapping[str, Any],
    output_dir: Path,
) -> Dict[str, str]:
    project = Path(str(payload["project_path"])).resolve()
    output = Path(output_dir).resolve()
    if output == project or project in output.parents:
        raise ValueError("验收输出目录不能位于被验证的牧场组项目内")
    output.mkdir(parents=True, exist_ok=True)
    paths = {
        "json": output / "multi_farm_acceptance.json",
        "farms_csv": output / "multi_farm_acceptance_farms.csv",
        "stages_csv": output / "multi_farm_acceptance_stages.csv",
        "lineage_csv": output / "multi_farm_acceptance_lineage.csv",
        "files_csv": output / "multi_farm_acceptance_files.csv",
        "issues_csv": output / "multi_farm_acceptance_issues.csv",
    }
    _atomic_write_json(paths["json"], payload)
    _atomic_write_csv(paths["farms_csv"], payload.get("farms", []))
    _atomic_write_csv(paths["stages_csv"], payload.get("stages", []))
    _atomic_write_csv(paths["lineage_csv"], payload.get("lineage", []))
    _atomic_write_csv(paths["files_csv"], payload.get("files", []))
    _atomic_write_csv(paths["issues_csv"], payload.get("issues", []))
    return {key: str(path) for key, path in paths.items()}


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="只读验证已完成的多牧场全流程结果",
    )
    parser.add_argument("project_path", type=Path, help="牧场组父项目路径")
    parser.add_argument(
        "--output-dir",
        type=Path,
        required=True,
        help="项目目录外的 JSON/CSV 输出目录",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    arguments = _parser().parse_args(argv)
    payload = validate_project(arguments.project_path)
    try:
        outputs = write_outputs(payload, arguments.output_dir)
    except Exception as exc:
        print(
            json.dumps(
                {
                    "status": "output_failed",
                    "error_type": type(exc).__name__,
                },
                ensure_ascii=False,
            )
        )
        return 3
    print(
        json.dumps(
            {
                "status": payload["status"],
                "counts": payload["counts"],
                "outputs": outputs,
            },
            ensure_ascii=False,
        )
    )
    return 0 if payload["status"] == "passed" else 2


if __name__ == "__main__":
    sys.exit(main())

"""全流程验收中的业务键血缘与补跑范围对账。

本模块只读子项目 XLSX，并以磁盘 SQLite 对不可逆 SHA-256 业务键摘要做
精确多重集比较。公开结果只包含计数与聚合指纹，不包含母牛号、公牛号或
逐行业务键。

口径与生产入口保持一致：

* 配种 raw→processed 使用 ``process_breeding_record_file`` 的冻精号格式化；
* 备选/已配近交使用 ``build_inbreeding_analysis_scope``；
* 母牛自身使用实际分析入口的 ``filter_dairy_cows``；
* 个体选配使用 ``GroupManager.apply_temp_strategy`` 的在场母牛前置范围。
"""

from __future__ import annotations

import hashlib
import json
import math
import sqlite3
import tempfile
from collections.abc import Iterable, Iterator, Mapping, Sequence
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config.breed_constants import filter_dairy_cows
from core.data.processor import format_naab_number
from core.inbreeding.analysis_scope import (
    ORIGINAL_BULL_ID_COLUMN,
    _clean_identifier,
    build_inbreeding_analysis_scope,
)


_MISSING_TOKENS = {"", "nan", "none", "null", "nat", "<na>", "n/a"}
_SEXED_TOKENS = {
    "1",
    "1.0",
    "true",
    "yes",
    "是",
    "性控",
    "性控冻精",
    "超级性控",
}
_REGULAR_TOKENS = {
    "0",
    "0.0",
    "false",
    "no",
    "否",
    "普通",
    "普通冻精",
    "常规",
    "常规冻精",
}
_UNKNOWN_TOKENS = {"未知", "unknown"}
_HASH_MASK = (1 << 256) - 1


class ScopeCheckError(ValueError):
    """补跑范围或业务键文件结构无法验证。"""


def _normalize_identifier(value: Any) -> str:
    return _clean_identifier(value)


def _normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and not math.isfinite(value):
        return ""
    text = str(value).strip()
    if text.casefold() in _MISSING_TOKENS:
        return ""
    try:
        timestamp = pd.Timestamp(value)
    except (TypeError, ValueError, OverflowError):
        return text
    if pd.isna(timestamp):
        return ""
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert("UTC").tz_localize(None)
    return timestamp.isoformat(timespec="seconds")


def _normalize_semen_type(value: Any) -> str:
    text = _normalize_identifier(value)
    token = text.casefold()
    if token in _SEXED_TOKENS:
        return "性控"
    if token in _REGULAR_TOKENS:
        return "常规"
    if token in _UNKNOWN_TOKENS or token in _MISSING_TOKENS:
        return "未知"
    return token


def _normalize_bull_as_processor(value: Any) -> str:
    """复用配种标准化器的 NAAB 规则；失败时按生产逻辑保留原值。"""
    original = _normalize_identifier(value)
    if not original:
        return ""
    formatted, _errors = format_naab_number(original)
    return _normalize_identifier(formatted) if formatted else original


def _key_digest(parts: Sequence[Any]) -> bytes:
    normalized = [_normalize_identifier(value) for value in parts]
    return hashlib.sha256(
        json.dumps(
            normalized,
            ensure_ascii=False,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    ).digest()


def _fingerprint_state() -> dict[str, int]:
    return {
        "rows": 0,
        "hash_sum": 0,
        "hash_square_sum": 0,
        "hash_xor": 0,
    }


def _add_fingerprint(state: dict[str, int], digest: bytes) -> None:
    numeric = int.from_bytes(digest, "big")
    state["rows"] += 1
    state["hash_sum"] = (state["hash_sum"] + numeric) & _HASH_MASK
    state["hash_square_sum"] = (
        state["hash_square_sum"] + numeric * numeric
    ) & _HASH_MASK
    state["hash_xor"] ^= numeric


def _public_fingerprint(state: Mapping[str, int]) -> str:
    """压缩为一个聚合摘要，避免暴露任何单行业务键哈希。"""
    payload = "|".join(
        (
            str(int(state.get("rows", 0))),
            f"{int(state.get('hash_sum', 0)):064x}",
            f"{int(state.get('hash_square_sum', 0)):064x}",
            f"{int(state.get('hash_xor', 0)):064x}",
        )
    )
    return hashlib.sha256(payload.encode("ascii")).hexdigest()


def _header_key(value: Any) -> str:
    return str(value or "").strip().casefold()


def _find_header(
    headers: Sequence[Any],
    candidates: Sequence[str],
    *,
    required: bool = True,
) -> int | None:
    normalized = [_header_key(value) for value in headers]
    wanted = {_header_key(value) for value in candidates}
    match = next(
        (index for index, value in enumerate(normalized) if value in wanted),
        None,
    )
    if match is None and required:
        raise ScopeCheckError("工作簿缺少验收所需业务键列")
    return match


def _cell(row: Sequence[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _is_blank_row(row: Sequence[Any]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row
    )


def _iter_sheet_rows(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> Iterator[tuple[list[Any], Sequence[Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook.active
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            raise ScopeCheckError("工作簿缺少验收所需工作表")
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ScopeCheckError("工作簿没有表头") from exc
        for row in rows:
            if not _is_blank_row(row):
                yield headers, row
    finally:
        workbook.close()


def _coalesce_identifier(*values: Any) -> str:
    for value in values:
        normalized = _normalize_identifier(value)
        if normalized:
            return normalized
    return ""


class _ExactMultisetComparator:
    """在临时 SQLite 中精确比较多重集，同时只公开聚合结果。"""

    def __init__(self) -> None:
        self._temporary = tempfile.TemporaryDirectory(
            prefix="acceptance_scope_"
        )
        self.connection = sqlite3.connect(
            str(Path(self._temporary.name) / "scope.sqlite3")
        )
        self.connection.execute("PRAGMA journal_mode = OFF")
        self.connection.execute("PRAGMA synchronous = OFF")
        self.connection.execute("PRAGMA temp_store = FILE")

    def close(self) -> None:
        self.connection.close()
        self._temporary.cleanup()

    def compare(
        self,
        name: str,
        expected: Iterable[Sequence[Any]],
        actual: Iterable[Sequence[Any]],
    ) -> dict[str, Any]:
        if not name.replace("_", "").isalnum():
            raise ScopeCheckError("内部范围检查名称无效")
        table = f"scope_{name}"
        self.connection.execute(f"DROP TABLE IF EXISTS {table}")
        self.connection.execute(
            f"""
            CREATE TABLE {table} (
                business_key_hash BLOB PRIMARY KEY,
                expected_count INTEGER NOT NULL DEFAULT 0,
                actual_count INTEGER NOT NULL DEFAULT 0
            ) WITHOUT ROWID
            """
        )
        expected_state = _fingerprint_state()
        actual_state = _fingerprint_state()
        self._insert_stream(
            table,
            "expected_count",
            expected,
            expected_state,
        )
        self._insert_stream(
            table,
            "actual_count",
            actual,
            actual_state,
        )
        (
            expected_unique,
            actual_unique,
            missing_rows,
            unexpected_rows,
        ) = (
            int(value)
            for value in self.connection.execute(
                f"""
                SELECT
                    COALESCE(SUM(
                        CASE WHEN expected_count > 0 THEN 1 ELSE 0 END
                    ), 0),
                    COALESCE(SUM(
                        CASE WHEN actual_count > 0 THEN 1 ELSE 0 END
                    ), 0),
                    COALESCE(SUM(
                        CASE WHEN expected_count > actual_count
                        THEN expected_count - actual_count ELSE 0 END
                    ), 0),
                    COALESCE(SUM(
                        CASE WHEN actual_count > expected_count
                        THEN actual_count - expected_count ELSE 0 END
                    ), 0)
                FROM {table}
                """
            ).fetchone()
        )
        self.connection.execute(f"DROP TABLE {table}")
        self.connection.commit()
        expected_rows = int(expected_state["rows"])
        actual_rows = int(actual_state["rows"])
        return {
            "expected_rows": expected_rows,
            "actual_rows": actual_rows,
            "expected_unique_keys": expected_unique,
            "actual_unique_keys": actual_unique,
            "expected_duplicate_rows": expected_rows - expected_unique,
            "actual_duplicate_rows": actual_rows - actual_unique,
            "missing_rows": missing_rows,
            "unexpected_rows": unexpected_rows,
            "expected_fingerprint": _public_fingerprint(expected_state),
            "actual_fingerprint": _public_fingerprint(actual_state),
            "passed": missing_rows == 0 and unexpected_rows == 0,
        }

    def _insert_stream(
        self,
        table: str,
        column: str,
        keys: Iterable[Sequence[Any]],
        state: dict[str, int],
    ) -> None:
        batch: list[tuple[bytes]] = []
        for key in keys:
            digest = _key_digest(key)
            _add_fingerprint(state, digest)
            batch.append((digest,))
            if len(batch) >= 2_000:
                self._insert_batch(table, column, batch)
                batch.clear()
        if batch:
            self._insert_batch(table, column, batch)
        self.connection.commit()

    def _insert_batch(
        self,
        table: str,
        column: str,
        batch: Sequence[tuple[bytes]],
    ) -> None:
        self.connection.executemany(
            f"""
            INSERT INTO {table} (business_key_hash, {column})
            VALUES (?, 1)
            ON CONFLICT(business_key_hash) DO UPDATE SET
                {column} = {column} + 1
            """,
            batch,
        )


def _load_frame(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, dtype=object)


def _iter_raw_breeding_keys(path: Path) -> Iterator[tuple[str, ...]]:
    rows = _iter_sheet_rows(path)
    try:
        headers, first_row = next(rows)
    except StopIteration:
        return
    cow_index = _find_header(
        headers, ("耳号", "牛号", "母牛号", "母牛耳号", "cow_id")
    )
    date_index = _find_header(
        headers,
        ("配种日期", "配种时间", "授精日期", "授精时间", "事件日期", "日期"),
    )
    bull_index = _find_header(
        headers, ("冻精编号", "冻精号", "公牛号", "精液号", "备注")
    )
    type_index = _find_header(
        headers,
        ("冻精类型", "精液类型", "类型", "是否性控"),
    )

    def convert(row: Sequence[Any]) -> tuple[str, ...] | None:
        cow_id = _normalize_identifier(_cell(row, cow_index))
        raw_date = _cell(row, date_index)
        bull_id = _normalize_bull_as_processor(_cell(row, bull_index))
        semen_type = _normalize_semen_type(_cell(row, type_index))
        raw_type = _normalize_identifier(_cell(row, type_index))
        if not cow_id or not bull_id or not _normalize_identifier(raw_date):
            return None
        if not raw_type:
            return None
        return (
            cow_id,
            _normalize_date(raw_date),
            bull_id,
            semen_type,
        )

    first_key = convert(first_row)
    if first_key is not None:
        yield first_key
    for _headers, row in rows:
        key = convert(row)
        if key is not None:
            yield key


def _iter_processed_breeding_keys(
    path: Path,
) -> Iterator[tuple[str, ...]]:
    rows = _iter_sheet_rows(path)
    try:
        headers, first_row = next(rows)
    except StopIteration:
        return
    cow_index = _find_header(headers, ("耳号", "cow_id", "母牛号", "牛号"))
    date_index = _find_header(headers, ("配种日期", "配种时间"))
    bull_index = _find_header(
        headers, ("冻精编号", "冻精号", "配种公牛号", "公牛号")
    )
    type_index = _find_header(headers, ("冻精类型", "精液类型", "类型"))

    def convert(row: Sequence[Any]) -> tuple[str, ...]:
        return (
            _normalize_identifier(_cell(row, cow_index)),
            _normalize_date(_cell(row, date_index)),
            _normalize_bull_as_processor(_cell(row, bull_index)),
            _normalize_semen_type(_cell(row, type_index)),
        )

    yield convert(first_row)
    for _headers, row in rows:
        yield convert(row)


def _iter_cow_self_result_keys(path: Path) -> Iterator[tuple[str]]:
    rows = _iter_sheet_rows(path, sheet_name="配对明细表")
    try:
        headers, first_row = next(rows)
    except StopIteration:
        return
    cow_index = _find_header(headers, ("母牛号", "cow_id", "牛号", "耳号"))
    yield (_normalize_identifier(_cell(first_row, cow_index)),)
    for _headers, row in rows:
        yield (_normalize_identifier(_cell(row, cow_index)),)


def _iter_candidate_result_keys(
    path: Path,
) -> Iterator[tuple[str, str]]:
    rows = _iter_sheet_rows(path, sheet_name="配对明细表")
    try:
        headers, first_row = next(rows)
    except StopIteration:
        return
    cow_index = _find_header(headers, ("母牛号", "cow_id", "牛号", "耳号"))
    original_bull_index = _find_header(
        headers,
        ("原始备选公牛号",),
        required=False,
    )
    bull_index = _find_header(
        headers,
        ("备选公牛号", "公牛号", "bull_id"),
    )

    def convert(row: Sequence[Any]) -> tuple[str, str]:
        return (
            _normalize_identifier(_cell(row, cow_index)),
            _coalesce_identifier(
                _cell(row, original_bull_index),
                _cell(row, bull_index),
            ),
        )

    yield convert(first_row)
    for _headers, row in rows:
        yield convert(row)


def _iter_mated_result_keys(
    path: Path,
) -> Iterator[tuple[str, str, str]]:
    rows = _iter_sheet_rows(path, sheet_name="配对明细表")
    try:
        headers, first_row = next(rows)
    except StopIteration:
        return
    cow_index = _find_header(headers, ("母牛号", "cow_id", "牛号", "耳号"))
    date_index = _find_header(headers, ("配种日期", "配种时间"))
    original_bull_index = _find_header(
        headers,
        ("原始公牛号", "原始配种公牛号"),
        required=False,
    )
    bull_index = _find_header(
        headers,
        ("配种公牛号", "冻精编号", "公牛号", "bull_id"),
    )

    def convert(row: Sequence[Any]) -> tuple[str, str, str]:
        return (
            _normalize_identifier(_cell(row, cow_index)),
            _normalize_date(_cell(row, date_index)),
            _coalesce_identifier(
                _cell(row, original_bull_index),
                _cell(row, bull_index),
            ),
        )

    yield convert(first_row)
    for _headers, row in rows:
        yield convert(row)


def _iter_single_identifier_sheet(
    path: Path,
    *,
    sheet_name: str,
    candidates: Sequence[str],
) -> Iterator[tuple[str]]:
    rows = _iter_sheet_rows(path, sheet_name=sheet_name)
    try:
        headers, first_row = next(rows)
    except StopIteration:
        return
    identifier_index = _find_header(headers, candidates)
    yield (_normalize_identifier(_cell(first_row, identifier_index)),)
    for _headers, row in rows:
        yield (_normalize_identifier(_cell(row, identifier_index)),)


def _latest_workbook(directory: Path, pattern: str) -> Path | None:
    candidates = [
        path
        for path in directory.glob(pattern)
        if path.is_file() and not path.is_symlink()
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime_ns)


def _missing_check(name: str, missing_artifact_count: int) -> dict[str, Any]:
    return {
        "lineage": name,
        "passed": False,
        "missing_artifact_count": int(missing_artifact_count),
        "expected_rows": 0,
        "actual_rows": 0,
        "missing_rows": 0,
        "unexpected_rows": 0,
    }


def _with_name(name: str, comparison: Mapping[str, Any]) -> dict[str, Any]:
    return {"lineage": name, **dict(comparison)}


def _breeding_raw_processed_check(
    comparator: _ExactMultisetComparator,
    raw: Path,
    processed: Path,
) -> dict[str, Any]:
    comparison = comparator.compare(
        "breeding_raw_processed",
        _iter_raw_breeding_keys(raw),
        _iter_processed_breeding_keys(processed),
    )
    raw_rows = sum(1 for _ in _iter_sheet_rows(raw))
    comparison["raw_rows"] = raw_rows
    comparison["eligible_raw_rows"] = comparison["expected_rows"]
    comparison["excluded_raw_rows"] = (
        raw_rows - comparison["expected_rows"]
    )
    return _with_name("breeding_business_key", comparison)


def _cow_self_check(
    comparator: _ExactMultisetComparator,
    cow_file: Path,
    result_file: Path,
) -> dict[str, Any]:
    cow_frame = filter_dairy_cows(
        _load_frame(cow_file),
        log_prefix="验收母牛自身范围：",
    )
    expected = (
        (_normalize_identifier(value),)
        for value in cow_frame["cow_id"]
        if _normalize_identifier(value)
    )
    return _with_name(
        "cow_self_scope",
        comparator.compare(
            "cow_self",
            expected,
            _iter_cow_self_result_keys(result_file),
        ),
    )


def _candidate_check(
    comparator: _ExactMultisetComparator,
    cow_file: Path,
    bull_file: Path,
    result_file: Path,
) -> dict[str, Any]:
    scope = build_inbreeding_analysis_scope(
        "candidate",
        _load_frame(cow_file),
        candidate_bull_df=_load_frame(bull_file),
        standardize_bull_id=_normalize_identifier,
    )
    cow_ids = [
        _normalize_identifier(value)
        for value in scope.cows["cow_id"]
        if _normalize_identifier(value)
    ]
    bull_ids = [
        _normalize_identifier(value)
        for value in scope.candidate_bulls[ORIGINAL_BULL_ID_COLUMN]
        if _normalize_identifier(value)
    ]
    expected = (
        (cow_id, bull_id)
        for cow_id in cow_ids
        for bull_id in bull_ids
    )
    result = _with_name(
        "candidate_cartesian_scope",
        comparator.compare(
            "candidate_cartesian",
            expected,
            _iter_candidate_result_keys(result_file),
        ),
    )
    result["expected_cow_rows"] = len(cow_ids)
    result["expected_bull_rows"] = len(bull_ids)
    result["expected_cartesian_rows"] = len(cow_ids) * len(bull_ids)
    return result


def _mated_check(
    comparator: _ExactMultisetComparator,
    cow_file: Path,
    breeding_file: Path,
    result_file: Path,
) -> dict[str, Any]:
    scope = build_inbreeding_analysis_scope(
        "mated",
        _load_frame(cow_file),
        breeding_df=_load_frame(breeding_file),
    )
    expected = (
        (
            _normalize_identifier(row.get("耳号")),
            _normalize_date(row.get("配种日期")),
            _normalize_identifier(row.get("冻精编号")),
        )
        for row in scope.breeding_records.to_dict("records")
    )
    return _with_name(
        "mated_business_key_scope",
        comparator.compare(
            "mated_business_key",
            expected,
            _iter_mated_result_keys(result_file),
        ),
    )


def _matching_check(
    comparator: _ExactMultisetComparator,
    index_file: Path,
    matrix_file: Path,
    report_file: Path,
) -> dict[str, Any]:
    index_frame = _load_frame(index_file)
    if "cow_id" not in index_frame.columns:
        raise ScopeCheckError("母牛指数结果缺少业务标识符列")
    if "是否在场" not in index_frame.columns or "sex" not in index_frame.columns:
        raise ScopeCheckError("母牛指数结果缺少选配范围列")
    expected_ids = [
        _normalize_identifier(value)
        for value in index_frame.loc[
            index_frame["是否在场"].fillna("").astype(str).str.strip().eq("是")
            & index_frame["sex"].fillna("").astype(str).str.strip().eq("母"),
            "cow_id",
        ]
        if _normalize_identifier(value)
    ]
    matrix = comparator.compare(
        "matching_matrix",
        ((value,) for value in expected_ids),
        _iter_single_identifier_sheet(
            matrix_file,
            sheet_name="推荐汇总",
            candidates=("cow_id", "母牛号", "牛号", "耳号"),
        ),
    )
    report = comparator.compare(
        "matching_report",
        ((value,) for value in expected_ids),
        _iter_single_identifier_sheet(
            report_file,
            sheet_name="选配结果",
            candidates=("母牛号", "cow_id", "牛号", "耳号"),
        ),
    )
    return {
        "lineage": "matching_scope",
        "passed": bool(matrix["passed"] and report["passed"]),
        "expected_rows": len(expected_ids),
        "matrix_rows": matrix["actual_rows"],
        "report_rows": report["actual_rows"],
        "matrix_missing_rows": matrix["missing_rows"],
        "matrix_unexpected_rows": matrix["unexpected_rows"],
        "report_missing_rows": report["missing_rows"],
        "report_unexpected_rows": report["unexpected_rows"],
        "expected_fingerprint": matrix["expected_fingerprint"],
        "matrix_fingerprint": matrix["actual_fingerprint"],
        "report_fingerprint": report["actual_fingerprint"],
        "expected_duplicate_rows": matrix["expected_duplicate_rows"],
        "matrix_duplicate_rows": matrix["actual_duplicate_rows"],
        "report_duplicate_rows": report["actual_duplicate_rows"],
    }


def validate_child_scope_artifacts(child_path: Path) -> list[dict[str, Any]]:
    """验证一个子项目的配种业务键与四类补跑结果范围。

    返回值可直接并入主验收器的 ``lineage`` 数组。所有摘要均为聚合值，
    不包含任何原始业务标识符。
    """
    child = Path(child_path)
    standardized = child / "standardized_data"
    analysis = child / "analysis_results"
    raw_breeding = child / "raw_data" / "breeding_records.xlsx"
    cow_file = standardized / "processed_cow_data.xlsx"
    breeding_file = standardized / "processed_breeding_data.xlsx"
    bull_file = standardized / "processed_bull_data.xlsx"
    cow_self_file = analysis / "母牛近交系数分析结果.xlsx"
    candidate_file = _latest_workbook(
        analysis,
        "备选公牛_近交系数及隐性基因分析结果*.xlsx",
    )
    mated_file = _latest_workbook(
        analysis,
        "已配公牛_近交系数及隐性基因分析结果*.xlsx",
    )
    matrix_file = analysis / "个体选配推荐矩阵.xlsx"
    matching_file = analysis / "个体选配报告.xlsx"
    index_file = analysis / "processed_index_cow_index_scores.xlsx"

    comparator = _ExactMultisetComparator()
    checks: list[dict[str, Any]] = []
    try:
        if raw_breeding.is_file() or breeding_file.is_file():
            if raw_breeding.is_file() and breeding_file.is_file():
                checks.append(
                    _breeding_raw_processed_check(
                        comparator,
                        raw_breeding,
                        breeding_file,
                    )
                )
            else:
                checks.append(_missing_check("breeding_business_key", 1))

        if cow_file.is_file() and cow_self_file.is_file():
            checks.append(
                _cow_self_check(comparator, cow_file, cow_self_file)
            )
        else:
            checks.append(
                _missing_check(
                    "cow_self_scope",
                    int(not cow_file.is_file())
                    + int(not cow_self_file.is_file()),
                )
            )

        if bull_file.is_file():
            if cow_file.is_file() and candidate_file is not None:
                checks.append(
                    _candidate_check(
                        comparator,
                        cow_file,
                        bull_file,
                        candidate_file,
                    )
                )
            else:
                checks.append(
                    _missing_check(
                        "candidate_cartesian_scope",
                        int(not cow_file.is_file())
                        + int(candidate_file is None),
                    )
                )

        if breeding_file.is_file():
            if cow_file.is_file() and mated_file is not None:
                checks.append(
                    _mated_check(
                        comparator,
                        cow_file,
                        breeding_file,
                        mated_file,
                    )
                )
            else:
                checks.append(
                    _missing_check(
                        "mated_business_key_scope",
                        int(not cow_file.is_file())
                        + int(mated_file is None),
                    )
                )

        if bull_file.is_file():
            matching_required = (index_file, matrix_file, matching_file)
            if all(path.is_file() for path in matching_required):
                checks.append(
                    _matching_check(
                        comparator,
                        index_file,
                        matrix_file,
                        matching_file,
                    )
                )
            else:
                checks.append(
                    _missing_check(
                        "matching_scope",
                        sum(
                            not path.is_file()
                            for path in matching_required
                        ),
                    )
                )
        return checks
    finally:
        comparator.close()


__all__ = [
    "ScopeCheckError",
    "validate_child_scope_artifacts",
]
